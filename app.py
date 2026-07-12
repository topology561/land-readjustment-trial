"""
市地重劃前後地價估算系統 (V9.0 完整版)
依據：第四號公報 (113.12.06) / 內政部都市地價指數 / 花蓮縣營造費標準表
土地開發分析法公式：V = [S ÷ (1+R) ÷ (1+i) - (C+M)]
花蓮縣利潤率對應：高雄公會區域利潤率表（第四號公報附表二）
"""

import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, date
import plotly.express as px
import re
import math
import io

class _NamedBytesIO(io.BytesIO):
    """BytesIO 包裝，附帶 .name 屬性，供 pandas read_excel 判斷引擎（xls/xlsx）"""
    def __init__(self, name: str, data: bytes):
        super().__init__(data)
        self.name = name


# ============ 主計總處 SDMX API：營造工程物價總指數 ============
@st.cache_data(ttl=86400, show_spinner=False)
def fetch_construction_price_index(start_year: int = 2024, end_ym: str = None) -> dict:
    """
    從行政院主計總處 SDMX API 取得「營造工程物價總指數」月資料。
    資料源：A030502015（營造工程物價指數）/ fldid=131（總指數）
    🆕 W-D.2 backlog（KL 2026-07-06）：end_ym 預設由寫死 "2026-M3" 改**動態**＝
    當日年月＋2 月緩衝（寫死時間參數＝靜默過期，no-silent-fallback 遠親；
    SDMX endTime 超出既有資料月份僅回傳現有資料，安全）。
    回傳：
      {
        'success': bool,
        'data': [{'ym': '2024-M1', 'roc': '113年1月', 'value': 109.39}, ...],
        'by_roc': {'113年1月': 109.39, ..., '113年11月': 110.68, ...},
        'fetched_at': 'YYYY-MM-DD HH:MM:SS',
        'error': str or None,
      }
    ※ 以 @st.cache_data 快取 24 小時；避免重複呼叫
    """
    import json
    import datetime as _dt
    if end_ym is None:
        _today = _dt.date.today()
        _em = _today.month + 2
        _ey = _today.year + (_em - 1) // 12
        _em = (_em - 1) % 12 + 1
        end_ym = f"{_ey}-M{_em}"
    result = {'success': False, 'data': [], 'by_roc': {}, 'fetched_at': '', 'error': None}
    url = (f"https://nstatdb.dgbas.gov.tw/dgbasAll/webMain.aspx?"
           f"sdmx/A030502015/131...M.&startTime={start_year}&endTime={end_ym}")
    try:
        import requests, urllib3
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        r = requests.get(url, timeout=15, verify=False)
        r.raise_for_status()
        # 政府網站回傳的 Content-Type 宣告 utf-8 但實際可能是 big5，先嘗試 json
        try:
            payload = r.json()
        except Exception:
            # 部分情形 headers 聲明錯誤：手動處理
            payload = json.loads(r.content.decode('utf-8', errors='replace'))

        # observations["index"] = [value]
        series = payload['data']['dataSets'][0]['series']['0']['observations']
        # dimensions.observation[0].values: [{'id':'2024-M1','name':'113年1月'}, ...]
        dim_values = payload['data']['structure']['dimensions']['observation'][0]['values']
        for idx_str, val_list in series.items():
            idx = int(idx_str)
            if idx < 0 or idx >= len(dim_values):
                continue
            dv = dim_values[idx]
            try:
                val = float(val_list[0])
            except (TypeError, ValueError, IndexError):
                continue
            result['data'].append({
                'ym': dv.get('id', ''),
                'roc': dv.get('name', ''),
                'value': val,
            })
            if dv.get('name'):
                result['by_roc'][dv['name']] = val
        result['data'].sort(key=lambda x: x['ym'])
        result['success'] = True
        result['fetched_at'] = _dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    except Exception as ex:
        result['error'] = str(ex)
    return result


# ============ Tab 切換持久化 wrapper ============
# Streamlit 不渲染的 widget 其 key 會被清除；以下 wrapper 另存 _sv_<key> 維持數值
# 先把原始 st.* 函式存成別名，避免批次替換後 wrapper 內部產生無限遞迴
_st_number_input = st.number_input
_st_selectbox    = st.selectbox
_st_checkbox     = st.checkbox
_st_date_input   = st.date_input
_st_slider       = st.slider

def _ni(label, **kwargs):
    """持久化的 st.number_input：切換 Tab 後保留數值"""
    key = kwargs.get('key', '')
    if key and f"_sv_{key}" in st.session_state:
        kwargs['value'] = st.session_state[f"_sv_{key}"]
    val = _st_number_input(label, **kwargs)
    if key:
        st.session_state[f"_sv_{key}"] = val
    return val

def _sb(label, options, **kwargs):
    """持久化的 st.selectbox：切換 Tab 後保留選項"""
    key = kwargs.get('key', '')
    opts = list(options)
    if key and f"_sv_{key}" in st.session_state:
        saved = st.session_state[f"_sv_{key}"]
        if saved in opts:
            kwargs['index'] = opts.index(saved)
    val = _st_selectbox(label, opts, **kwargs)
    if key:
        st.session_state[f"_sv_{key}"] = val
    return val

def _cb(label, value=False, **kwargs):
    """持久化的 st.checkbox：切換 Tab 後保留勾選狀態"""
    key = kwargs.get('key', '')
    # value 可能由位置引數或 keyword 傳入，統一放進 kwargs
    kwargs.setdefault('value', value)
    if key and f"_sv_{key}" in st.session_state:
        kwargs['value'] = st.session_state[f"_sv_{key}"]
    val = _st_checkbox(label, **kwargs)
    if key:
        st.session_state[f"_sv_{key}"] = val
    return val

def _di(label, value=None, **kwargs):
    """持久化的 st.date_input：切換 Tab 後保留日期"""
    key = kwargs.get('key', '')
    # value 可能由位置引數或 keyword 傳入，統一放進 kwargs
    if value is not None:
        kwargs.setdefault('value', value)
    if key and f"_sv_{key}" in st.session_state:
        kwargs['value'] = st.session_state[f"_sv_{key}"]
    val = _st_date_input(label, **kwargs)
    if key:
        st.session_state[f"_sv_{key}"] = val
    return val

def _sl(label, min_value=None, max_value=None, value=None, step=None, **kwargs):
    """持久化的 st.slider：切換 Tab 後保留數值"""
    key = kwargs.get('key', '')
    # 將位置引數合併進 kwargs，避免重複傳遞
    if min_value is not None: kwargs.setdefault('min_value', min_value)
    if max_value is not None: kwargs.setdefault('max_value', max_value)
    if value     is not None: kwargs.setdefault('value', value)
    if step      is not None: kwargs.setdefault('step', step)
    if key and f"_sv_{key}" in st.session_state:
        kwargs['value'] = st.session_state[f"_sv_{key}"]
    val = _st_slider(label, **kwargs)
    if key:
        st.session_state[f"_sv_{key}"] = val
    return val

st.set_page_config(page_title="市地重劃地價估算系統", page_icon="🏗️", layout="wide", initial_sidebar_state="expanded")

# ============ 常數定義 ============
ECONOMIC_LIFE_BY_MATERIAL = {
    "鋼骨造": 50, "鋼骨鋼筋混凝土造": 50, "鋼筋混凝土造": 50, "預鑄混凝土造": 50,
    "鋼筋混凝土加強磚造": 35, "加強磚造": 35, "磚造": 25, "鋼架造": 20, "木造": 10
}
RESIDUAL_RATE_BY_MATERIAL = {
    "鋼骨造": 0.10, "鋼骨鋼筋混凝土造": 0.10, "鋼筋混凝土造": 0.05, "預鑄混凝土造": 0.05,
    "鋼筋混凝土加強磚造": 0.0, "加強磚造": 0.0, "磚造": 0.0, "鋼架造": 0.10, "木造": 0.0
}
FEE_RATES = {"廣告費銷售費": {"min": 3.0, "max": 7.0, "default": 5.0}, "管理費": {"min": 1.5, "max": 3.0, "default": 2.25}, "稅捐": {"min": 0.5, "max": 1.2, "default": 0.85}}
# 花蓮縣對應高雄公會利潤率（依據第四號公報附表二）
HUALIEN_PROFIT_RATES = {
    "1年(含)以下": {"min": 8, "max": 20, "default": 14},
    "超過1年~2年(含)": {"min": 10, "max": 22, "default": 16},
    "超過2年~3年(含)": {"min": 12, "max": 23, "default": 17.5},
    "超過3年~4年(含)": {"min": 13, "max": 25, "default": 19},
    "超過4年~5年(含)": {"min": 14, "max": 26, "default": 20},
    "超過5年": {"min": 15, "max": 27, "default": 21}
}
HUALIEN_CONSTRUCTION_COST = {
    "未達100,000": (60800, 76200), "100,000~150,000": (66800, 83600), "150,000~200,000": (73600, 92200),
    "200,000~250,000": (77400, 101000), "250,000~300,000": (85100, 111000), "300,000~400,000": (90800, 122000),
    "400,000~500,000": (98300, 134000), "500,000以上": (105000, 155000)
}
LAND_PRICE_INDEX_DATES = ["2020/3/31", "2020/9/30", "2021/3/31", "2021/9/30", "2022/3/31", "2022/9/30", "2023/3/31", "2023/9/30", "2024/3/31", "2024/9/30", "2025/3/31"]
LAND_PRICE_INDEX_DATA = {
    "花蓮市": [98.45, 98.6, 98.95, 99.28, 99.75, 99.88, 100, 100.16, 101.05, 101.93, 102.49],
    "鳳林鎮": [98.71, 98.9, 99.09, 99.17, 99.48, 99.91, 100, 100.27, 100.53, 100.81, 100.96],
    "玉里鎮": [98.54, 98.72, 98.96, 99.18, 99.25, 99.74, 100, 100.24, 100.6, 101.12, 101.79],
    "新城鄉": [99.44, 99.48, 99.67, 99.76, 99.83, 99.98, 100, 100.03, 100.14, 100.82, 101.36],
    "吉安鄉": [98.21, 98.26, 98.51, 98.63, 99.05, 99.31, 100, 100.21, 101.52, 102.32, 102.95],
    "壽豐鄉": [98.8, 99, 99.15, 99.33, 99.52, 99.97, 100, 100.35, 101.14, 101.54, 101.95],
    "光復鄉": [97.68, 97.54, 98.39, 99.2, 99.42, 99.76, 100, 100.22, 100.58, 100.81, 100.85],
    "豐濱鄉": [100, 99.83, 99.83, 99.86, 99.93, 99.96, 100, 100.05, 100.12, 100.15, 100.19],
    "瑞穗鄉": [98.99, 99.02, 99.18, 99.32, 99.5, 99.79, 100, 100.34, 100.62, 100.9, 101.44],
    "富里鄉": [97.44, 97.56, 98.19, 98.58, 99.31, 99.79, 100, 100.39, 101.27, 101.97, 102.16],
    "秀林鄉": [98.16, 98.3, 98.88, 99.06, 99.45, 99.98, 100, 100.38, 101.17, 101.78, 102.21]
}

# ============ 工具函數 ============
def get_construction_cost_range(avg_price):
    """根據銷售單價決定營造單價區間，返回區間名稱和(最小值, 最大值)"""
    if avg_price < 100000: rn = "未達100,000"
    elif avg_price < 150000: rn = "100,000~150,000"
    elif avg_price < 200000: rn = "150,000~200,000"
    elif avg_price < 250000: rn = "200,000~250,000"
    elif avg_price < 300000: rn = "250,000~300,000"
    elif avg_price < 400000: rn = "300,000~400,000"
    elif avg_price < 500000: rn = "400,000~500,000"
    else: rn = "500,000以上"
    min_val, max_val = HUALIEN_CONSTRUCTION_COST[rn]
    return rn, min_val, max_val

def get_profit_rate_by_period(total_months):
    years = total_months / 12
    if years <= 1: return "1年(含)以下", HUALIEN_PROFIT_RATES["1年(含)以下"]
    elif years <= 2: return "超過1年~2年(含)", HUALIEN_PROFIT_RATES["超過1年~2年(含)"]
    elif years <= 3: return "超過2年~3年(含)", HUALIEN_PROFIT_RATES["超過2年~3年(含)"]
    elif years <= 4: return "超過3年~4年(含)", HUALIEN_PROFIT_RATES["超過3年~4年(含)"]
    elif years <= 5: return "超過4年~5年(含)", HUALIEN_PROFIT_RATES["超過4年~5年(含)"]
    else: return "超過5年", HUALIEN_PROFIT_RATES["超過5年"]

def get_index_dates():
    return [datetime(int(d.split('/')[0]), int(d.split('/')[1]), int(d.split('/')[2])) for d in LAND_PRICE_INDEX_DATES]

def get_interpolated_index(target_date, district):
    if district not in LAND_PRICE_INDEX_DATA: return None
    index_dates, index_values = get_index_dates(), LAND_PRICE_INDEX_DATA[district]
    if target_date <= index_dates[0]: return index_values[0]
    if target_date >= index_dates[-1]: return index_values[-1]
    for i in range(len(index_dates) - 1):
        if index_dates[i] <= target_date <= index_dates[i + 1]:
            days_total = (index_dates[i + 1] - index_dates[i]).days
            days_from_start = (target_date - index_dates[i]).days
            return index_values[i] + (days_from_start / days_total) * (index_values[i + 1] - index_values[i]) if days_total else index_values[i]
    return index_values[-1]

def calculate_price_adjustment(trans_date, price_date, district):
    if trans_date is None or price_date is None: return {'trans_index': None, 'price_index': None, 'factor': 1.0}
    ti, pi = get_interpolated_index(trans_date, district), get_interpolated_index(price_date, district)
    if ti is None or pi is None or ti == 0: return {'trans_index': ti, 'price_index': pi, 'factor': 1.0}
    return {'trans_index': round(ti, 2), 'price_index': round(pi, 2), 'factor': pi / ti}

def convert_roc_date(roc_date_str):
    try:
        s = re.sub(r'[/\-\.]', '', str(roc_date_str).strip())
        if not s or s == 'nan' or len(s) < 5: return None
        if len(s) <= 6: year, month, day = int(s[:3]) + 1911, int(s[3:5]), 1
        else: year, month, day = int(s[:-4]) + 1911, int(s[-4:-2]), int(s[-2:])
        if month < 1 or month > 12: month = 1
        if day < 1 or day > 31: day = 1
        return datetime(year, month, day)
    except: return None

def parse_area(v):
    try: return float(str(v).replace('㎡', '').replace(',', '').strip()) if pd.notna(v) else 0.0
    except: return 0.0

def parse_price(v):
    try: return float(str(v).replace(',', '').strip()) if pd.notna(v) else 0.0
    except: return 0.0

def get_economic_life(material):
    if pd.isna(material) or not material: return 50, 0.05
    for k, life in ECONOMIC_LIFE_BY_MATERIAL.items():
        if k in str(material): return life, RESIDUAL_RATE_BY_MATERIAL.get(k, 0.05)
    return 50, 0.05

def calc_age(completion_str, trans_date):
    if pd.isna(completion_str) or trans_date is None: return None
    comp = convert_roc_date(completion_str)
    return max(0, round((trans_date - comp).days / 365.25, 1)) if comp else None

# ============ Excel 讀取 ============
def read_land_excel(files):
    """讀取土地交易Excel，只需要案件列表工作表"""
    records = []
    for f in files:
        try:
            # 讀取Excel檔案
            xls = pd.read_excel(f, sheet_name=None, engine='xlrd' if f.name.endswith('.xls') else 'openpyxl', header=None)
            case_df = land_df = None
            
            # 找出案件列表和土地工作表
            for sn, df in xls.items():
                if '案件' in sn: case_df = df
                elif '土地' in sn: land_df = df
            
            if case_df is None: 
                st.warning(f"檔案 {f.name} 無法找到案件列表工作表")
                continue
            
            # 找出欄位標題行（包含「編號」的那一列）
            header_found = False
            for i in range(min(10, len(case_df))):
                if any('編號' in str(v) for v in case_df.iloc[i].values):
                    case_df.columns = case_df.iloc[i]
                    case_df = case_df.iloc[i+1:].reset_index(drop=True)
                    header_found = True
                    break
            
            if not header_found:
                st.warning(f"檔案 {f.name} 案件列表工作表無法找到欄位標題")
                continue
            
            # 處理土地工作表（如果存在）
            if land_df is not None:
                for i in range(min(10, len(land_df))):
                    if any('序號' in str(v) for v in land_df.iloc[i].values):
                        land_df.columns = land_df.iloc[i]
                        land_df = land_df.iloc[i+1:].reset_index(drop=True)
                        break
            
            # 逐筆讀取案件資料
            for _, row in case_df.iterrows():
                cid = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
                if not cid or cid == 'nan' or cid == '': continue
                
                loc, td, tp, up, ar = "", None, 0, 0, 0
                
                # 解析各欄位
                for col in case_df.columns:
                    cs = str(col)
                    val = row[col]
                    
                    # 位置
                    if any(k in cs for k in ['位置', '門牌', '地段']): 
                        loc = str(val) if pd.notna(val) else ""
                    # 交易日期
                    elif '交易日期' in cs: 
                        td = convert_roc_date(val)
                    # 總價 - 改用包含方式匹配
                    elif '總價' in cs and '車位' not in cs: 
                        tp = parse_price(val)
                    # 單價
                    elif '單價' in cs and '元' in cs and '㎡' in cs: 
                        up = parse_price(val)
                    # 面積
                    elif '總面積' in cs or ('面積' in cs and '㎡' in cs and '移轉' not in cs and '車位' not in cs): 
                        ar = parse_area(val)
                
                # 如果沒有面積，從土地工作表取得
                if ar == 0 and land_df is not None and len(land_df) > 0:
                    ac = next((c for c in land_df.columns if '面積' in str(c) and '移轉' in str(c)), None)
                    if ac is None:
                        ac = next((c for c in land_df.columns if '面積' in str(c)), None)
                    if ac:
                        # 匹配編號（格式可能是 "1-1", "1-2" 對應編號 "1"）
                        m = land_df[land_df.iloc[:, 0].astype(str).str.startswith(f"{cid}-")]
                        if not m.empty: 
                            ar = m[ac].apply(parse_area).sum()
                
                # 計算單價（如果未提供）
                if up == 0 and tp > 0 and ar > 0: 
                    up = tp / ar
                
                # 加入有效紀錄
                if tp > 0 and ar > 0:
                    records.append({
                        '檔案': f.name, 
                        '編號': cid, 
                        '位置': loc, 
                        '交易日期': td, 
                        '面積(㎡)': round(ar, 2), 
                        '總價(元)': tp, 
                        '單價(元/㎡)': round(up, 0)
                    })
                    
        except Exception as e: 
            st.error(f"處理 {f.name} 時發生錯誤: {str(e)}")
    
    return pd.DataFrame(records) if records else pd.DataFrame()

def read_building_excel(files):
    """讀取房地交易Excel，需要案件列表、土地、建物三個工作表
    回傳: (有效records的DataFrame, 全部案例土地移轉面積列表)"""
    records = []
    all_land_areas = []  # 收集所有案例的土地移轉面積（含被剔除的）
    for f in files:
        try:
            xls = pd.read_excel(f, sheet_name=None, engine='xlrd' if f.name.endswith('.xls') else 'openpyxl', header=None)
            case_df = land_df = build_df = None
            for sn, df in xls.items():
                if '案件' in sn: case_df = df
                elif '土地' in sn: land_df = df
                elif '建物' in sn: build_df = df
            
            if case_df is None: 
                st.warning(f"檔案 {f.name} 無法找到案件列表工作表")
                continue
                
            def proc(df):
                if df is None: return None
                for i in range(min(10, len(df))):
                    if any(k in ''.join(str(v) for v in df.iloc[i].values) for k in ['編號', '序號', '面積']):
                        df.columns = df.iloc[i]
                        return df.iloc[i+1:].reset_index(drop=True)
                return df
            
            case_df, land_df, build_df = proc(case_df), proc(land_df), proc(build_df)
            
            for _, row in case_df.iterrows():
                cid = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
                if not cid or cid == 'nan' or cid == '': continue
                
                loc, td, tp, ta, age = "", None, 0, 0, None
                for col in case_df.columns:
                    cs, val = str(col), row[col]
                    if any(k in cs for k in ['位置', '門牌']): 
                        loc = str(val) if pd.notna(val) else ""
                    elif '交易日期' in cs: 
                        td = convert_roc_date(val)
                    # 改用包含方式匹配總價
                    elif '總價' in cs and '車位' not in cs: 
                        tp = parse_price(val)
                    elif '總面積' in cs: 
                        ta = parse_area(val)
                    elif cs == '屋齡':
                        try: age = float(val) if pd.notna(val) and str(val).strip() else None
                        except: pass
                
                la = 0
                if land_df is not None and len(land_df) > 0:
                    ac = next((c for c in land_df.columns if '面積' in str(c)), None)
                    if ac:
                        m = land_df[land_df.iloc[:, 0].astype(str).str.startswith(f"{cid}-")]
                        if not m.empty: la = m[ac].apply(parse_area).sum()
                
                # 收集所有案例的土地移轉面積（不管後續是否被剔除）
                if la > 0:
                    all_land_areas.append(round(la, 2))
                
                ba, mat, comp_str, valid_comp = 0, "", None, False
                if build_df is not None and len(build_df) > 0:
                    m = build_df[build_df.iloc[:, 0].astype(str).str.startswith(f"{cid}-")]
                    if not m.empty:
                        fr = m.iloc[0]
                        for col in build_df.columns:
                            cs = str(col)
                            if '面積' in cs and '移轉' in cs: ba = m[col].apply(parse_area).sum()
                            elif '建材' in cs: mat = str(fr[col]) if pd.notna(fr[col]) else ""
                            elif '完成' in cs and '日期' in cs:
                                rv = fr[col]
                                if pd.notna(rv) and str(rv).strip():
                                    comp_str = str(rv)
                                    if convert_roc_date(comp_str): valid_comp = True
                        if (age is None or age == 0) and comp_str and td:
                            ca = calc_age(comp_str, td)
                            if ca and ca > 0: age = ca
                        if age is None or age == 0:
                            for col in build_df.columns:
                                if str(col) == '屋齡':
                                    try: age = float(fr[col]) if pd.notna(fr[col]) and str(fr[col]).strip() else None
                                    except: pass
                                    break
                
                if ba == 0 and ta > 0: ba = ta
                if age is None or age <= 0: continue
                if not valid_comp: continue
                el, rr = get_economic_life(mat)
                if tp > 0 and ba > 0:
                    records.append({
                        '檔案': f.name, 
                        '編號': cid, 
                        '位置': loc, 
                        '交易日期': td, 
                        '屋齡': round(age, 1), 
                        '總價(元)': tp, 
                        '土地面積(㎡)': round(la, 2), 
                        '建物面積(㎡)': round(ba, 2), 
                        '主要建材': mat, 
                        '經濟耐用年數': el, 
                        '殘餘價格率': rr
                    })
        except Exception as e: 
            st.error(f"處理 {f.name} 時發生錯誤: {str(e)}")
    
    df = pd.DataFrame(records) if records else pd.DataFrame()
    return df, all_land_areas

# ============ 計算函數 ============
def calc_building_volume(land_area, bc_pct, far_pct):
    """
    計算建築規劃
    
    附屬建物面積（陽台）計算依據：
    - 每層陽台面積 = max(該層樓地板面積×10%, 5㎡)
    - 總附屬建物面積 = sum(每層陽台面積)
    
    屋頂突出物計算依據：
    - 屋頂突出物 = max(地面層面積×12.5%, 25㎡)
    """
    bc, far = bc_pct / 100, far_pct / 100
    if bc <= 0: return None
    floors = math.ceil(far / bc)
    if floors <= 0: return None
    actual_bc = far / floors
    
    # 一樓面積（建築面積 / 地面層面積）
    gf = land_area * actual_bc
    
    # 二樓以上主建物面積
    uf = land_area * far - gf
    
    # 附屬建物面積（陽台）：每層取 max(該層面積×10%, 5㎡)，共 floors 層
    # 每層樓地板面積 = gf（假設每層相同）
    balcony_per_floor = max(gf * 0.10, 5.0)  # 每層陽台面積（㎡）
    bal = balcony_per_floor * floors  # 總附屬建物面積
    
    # 屋頂突出物：max(地面層面積×12.5%, 25㎡)
    roof = max(gf * 0.125, 25.0)
    
    total = gf + uf + bal + roof
    
    return {
        'floors': floors, 
        'actual_bc_pct': actual_bc * 100, 
        'actual_far_pct': actual_bc * floors * 100, 
        'gf_sqm': gf,  # 一樓面積
        'uf_sqm': uf,  # 二樓以上主建物面積
        'bal_sqm': bal,  # 附屬建物面積
        'bal_per_floor': balcony_per_floor,  # 每層陽台面積
        'roof_sqm': roof,  # 屋頂突出物
        'total_sqm': total, 
        'total_ping': total * 0.3025, 
        'land_ping': land_area * 0.3025
    }

def calc_composite_interest(bm, lm, dr, lr, be, bl, le, ll, bw, lw):
    """
    計算綜合利息資本利率（依據吉安重劃報告書表4-1）
    
    參數：
    bm: 建物興建工期(月) = M1
    lm: 土地持有時期(月) = M2
    dr: 自有資金年利率(%) = A（五大銀行一年期平均存款利率）
    lr: 借貸資金年利率(%) = B（五大銀行平均放款利率）
    be: 建築自有資金比例(%) = C1
    bl: 建築借貸資金比例(%) = D1
    le: 土地自有資金比例(%) = C2
    ll: 土地借貸資金比例(%) = D2
    bw: 建築投資資本利息比例(%) = C3
    lw: 土地投資資本利息比例(%) = D3
    
    公式：
    建築投資資本利息 = (C1×A + D1×B) / 12 × M1 × 0.5
    土地投資資本利息 = (C2×A + D2×B) / 12 × M2
    綜合利息資本利率 = C3×建築投資資本利息 + D3×土地投資資本利息
    
    注意：比例都是百分比形式（如50表示50%）
    """
    # 建築投資加權利率（百分比）
    bld_weighted_rate = (be/100 * dr + bl/100 * lr)  # 結果為百分比
    
    # 土地投資加權利率（百分比）
    land_weighted_rate = (le/100 * dr + ll/100 * lr)  # 結果為百分比
    
    # 建築投資資本利息 = 加權利率 × (工期月數/12) × 0.5
    bi = bld_weighted_rate * (bm / 12) * 0.5
    
    # 土地投資資本利息 = 加權利率 × (持有月數/12)
    li = land_weighted_rate * (lm / 12)
    
    # 綜合利息資本利率 = 建築投資權重×建築利息 + 土地投資權重×土地利息
    composite = (bw/100 * bi) + (lw/100 * li)
    
    return {
        'building_int': bi,      # 建築投資資本利息（%）
        'land_int': li,          # 土地投資資本利息（%）
        'composite': composite   # 綜合利息資本利率（%）
    }

def calc_land_dev(bv, sp, cc, asr, mr, tr, cir, pr, pir=1.0):
    """
    土地開發分析法計算
    公式：V = [S ÷ (1+R) ÷ (1+i) - (C+M)]
    依據：不動產估價技術規則第七十條
    
    S：開發或建築後預期總銷售金額
    R：適當之利潤率
    i：資本利息綜合利率
    C：直接成本（營造費用）
    M：間接成本（廣告銷售費、管理費、稅捐、規劃設計費）
    """
    tfa = bv['total_ping']  # 總樓地板面積(坪)
    la = bv['land_ping']     # 土地面積(坪)
    acc = cc * pir           # 調整後營造單價
    
    # S：總銷售金額
    S = tfa * sp
    
    # C：直接成本（營造總費用）
    C = tfa * acc
    
    # M：間接成本
    M_design = C * 0.02      # 規劃設計費（營造費的2%）
    M_ad = S * asr / 100     # 廣告銷售費
    M_mg = S * mr / 100      # 管理費
    M_tx = S * tr / 100      # 稅捐
    M = M_design + M_ad + M_mg + M_tx
    
    # R：利潤率（轉換為小數）
    R_decimal = pr / 100
    
    # i：資本利息綜合利率（轉換為小數）
    i_decimal = cir / 100
    
    # 土地開發分析法公式：V = [S ÷ (1+R) ÷ (1+i) - (C+M)]
    V = S / (1 + R_decimal) / (1 + i_decimal) - (C + M)
    
    # 回傳計算結果
    return {
        'S': S, 
        'acc': acc, 
        'C': C, 
        'M_design': M_design,
        'M_ad': M_ad, 
        'M_mg': M_mg, 
        'M_tx': M_tx, 
        'M': M, 
        'R_pct': pr,
        'i_pct': cir,
        'V': V,
        'land_sqm': bv['land_ping'] / 0.3025,  # 土地面積(㎡)
        'V_sqm': V / (bv['land_ping'] / 0.3025) if bv['land_ping'] > 0 else 0,  # 每平方公尺單價
        'V_ping': V / bv['land_ping'] if bv['land_ping'] > 0 else 0  # 每坪單價
    }

# ============ DXF 街廓／地籍解析 ============
# Big5 編碼是地籍圖重測系統預設匯出編碼
_DXF_ENCODINGS = ('big5', 'cp950', 'utf-8', 'latin-1')

def _read_dxf_any_encoding(file_bytes: bytes):
    """嘗試用 Big5/cp950/utf-8 等編碼讀取 DXF；回傳 ezdxf Drawing 物件"""
    import ezdxf
    last_err = None
    for enc in _DXF_ENCODINGS:
        try:
            # 使用 TextIOWrapper 處理 CRLF 行尾（ezdxf 要求 \n 行尾）
            stream = io.TextIOWrapper(io.BytesIO(file_bytes), encoding=enc, errors='replace', newline=None)
            doc = ezdxf.read(stream)
            # 若成功讀取但實體全為空，很可能是編碼不對，繼續嘗試下一個
            if len(list(doc.modelspace())) == 0 and enc != _DXF_ENCODINGS[-1]:
                continue
            return doc
        except Exception as e:
            last_err = e
    raise RuntimeError(f"無法解析 DXF：{last_err}")


@st.cache_data(show_spinner=False, max_entries=4)
def parse_block_dxf(file_bytes: bytes) -> dict:
    """
    解析分配街廓 DXF（如 test.dxf）
    ※ 以 @st.cache_data 快取：相同 file_bytes 不再重算，大幅加速 Tab 切換與互動
    回傳：
      {
        'polygons': [ {'id':i, 'vertices':[(x,y),...], 'area_m2':float, 'centroid':(x,y), 'label':'R1', 'is_outer':bool}, ... ],
        'outer_boundary': [(x,y),...] or None,  # 最大多邊形視為重劃範圍外框
        'texts': [{'text':str, 'x':float, 'y':float, 'layer':str}, ...],
        'layer_names': [str,...],
      }
    偵測邏輯：
      - 所有閉合 POLYLINE / LWPOLYLINE 皆視為街廓候選
      - 面積最大者視為重劃範圍外框（排除在街廓清單外，但保留供計算總面積）
      - 街廓內若含 'R<數字>' 格式的 TEXT，取該標籤為街廓編號

    🆕 W-A.3 方案 B（v3.1 圖層重構）：街廓一律維持「BLOCK 圖層的閉合多邊形」直讀，
      **不改為 polygonize(BLOCK 線 ∪ BASELINE 線)**（方案 A 已否決，避免震動 W-A 成果）。
      因此「BLOCK 語意收窄為使用分區界」與「BASELINE 兼任分配線」不在解析層處理，
      改在暫編標記層以 cut_type 區分（見 _annotate_temp_parcel_cut_type，§2）：
      同原地號相鄰暫編 category 不同 → 逕割（使用分區界）；同可建築分區 → 分配（BASELINE 切）。
      使用者繪圖時，同分區分隔處之 BASELINE 與 BLOCK 閉合邊重疊即可。
    """
    from shapely.geometry import Polygon, Point
    import re as _re

    doc = _read_dxf_any_encoding(file_bytes)
    msp = doc.modelspace()

    # 擷取所有閉合多邊形
    # 🆕 Phase 11 Task AA：One-File Workflow 防呆 — 若 DXF 同時包含 BLOCK 與 CADASTRAL
    # 之閉合多邊形，先收集所有閉合多邊形 + 其圖層名，再依圖層名優先過濾
    _block_layer_aliases = ('BLOCK', 'BLOCKS', 'STREET_BLOCK', 'STREETBLOCK',
                             'BLK', 'BLKS', 'R_BLOCK', '街廓')
    _cadastral_layer_aliases = ('CADASTRAL_BOUND', 'CADASTRAL_BOUNDARY', 'CADASTRAL',
                                 'PARCEL_BOUND', 'PARCEL_BOUNDARY', 'PARCEL', 'PARCELS',
                                 '地籍線', '地籍', 'OPAR', 'ROAD')

    def _layer_matches_aliases(lname, aliases):
        if not lname:
            return False
        ln = str(lname).strip().upper()
        return any(ln == a.upper() or a.upper() in ln for a in aliases)

    raw_polys_all = []
    for e in msp.query('POLYLINE'):
        try:
            if e.is_closed:
                pts = [(v.dxf.location.x, v.dxf.location.y) for v in e.vertices]
                if len(pts) >= 3:
                    raw_polys_all.append({'layer': e.dxf.layer, 'vertices': pts})
        except Exception:
            continue
    for e in msp.query('LWPOLYLINE'):
        try:
            if e.closed:
                pts = [(p[0], p[1]) for p in e.get_points()]
                if len(pts) >= 3:
                    raw_polys_all.append({'layer': e.dxf.layer, 'vertices': pts})
        except Exception:
            continue

    # 篩選優先順序：
    # 1. 若有「BLOCK 命名」的圖層 → 只用該圖層的閉合多邊形作為街廓
    # 2. 否則排除「CADASTRAL 命名」的圖層 → 剩餘者均視為街廓候選
    # 3. 都沒有 → 全部閉合多邊形視為街廓候選（原行為，向下相容）
    _has_block_named = any(_layer_matches_aliases(rp['layer'], _block_layer_aliases)
                           for rp in raw_polys_all)
    _has_cadastral_named = any(_layer_matches_aliases(rp['layer'], _cadastral_layer_aliases)
                                for rp in raw_polys_all)
    if _has_block_named:
        raw_polys = [rp for rp in raw_polys_all
                     if _layer_matches_aliases(rp['layer'], _block_layer_aliases)]
    elif _has_cadastral_named:
        raw_polys = [rp for rp in raw_polys_all
                     if not _layer_matches_aliases(rp['layer'], _cadastral_layer_aliases)]
    else:
        raw_polys = raw_polys_all

    # 擷取所有 TEXT / MTEXT 標籤
    texts = []
    for t in msp.query('TEXT'):
        try:
            texts.append({
                'text': (t.dxf.text or '').strip(),
                'x': float(t.dxf.insert.x),
                'y': float(t.dxf.insert.y),
                'layer': t.dxf.layer,
            })
        except Exception:
            continue
    for t in msp.query('MTEXT'):
        try:
            texts.append({
                'text': (t.text or '').strip(),
                'x': float(t.dxf.insert.x),
                'y': float(t.dxf.insert.y),
                'layer': t.dxf.layer,
            })
        except Exception:
            continue

    # 建立 shapely Polygon，計算面積
    poly_objs = []
    for i, rp in enumerate(raw_polys):
        try:
            sp = Polygon(rp['vertices'])
            if not sp.is_valid:
                sp = sp.buffer(0)
            area = sp.area
            cen = sp.centroid
            poly_objs.append({
                'id': i,
                'layer': rp['layer'],
                'vertices': rp['vertices'],
                'shapely': sp,
                'area_m2': area,
                'centroid': (cen.x, cen.y),
            })
        except Exception:
            continue

    if not poly_objs:
        return {'polygons': [], 'outer_boundary': None, 'texts': texts, 'layer_names': sorted(set(rp['layer'] for rp in raw_polys))}

    # 對每個多邊形計算「contain_cnt」＝包含多少「面積比自己小」的其他多邊形中心點
    # 限制「自己必須較大」是為了避免兩個不規則多邊形相互吃到對方中心時誤判
    for p in poly_objs:
        cnt = 0
        for q in poly_objs:
            if q['id'] == p['id']:
                continue
            if p['area_m2'] <= q['area_m2']:
                continue
            try:
                if p['shapely'].contains(Point(q['centroid'][0], q['centroid'][1])):
                    cnt += 1
            except Exception:
                continue
        p['contain_cnt'] = cnt

    # 「葉節點街廓」＝不含任何其他多邊形中心的最細粒度多邊形
    # 這些才是真正的街廓；contain_cnt >= 1 的是複合多邊形（通常是圖面上的群組或外框）
    leaf_polys = [p for p in poly_objs if p['contain_cnt'] == 0]

    # 若沒有葉節點（罕見），則回退使用全部多邊形
    if not leaf_polys:
        leaf_polys = poly_objs

    # 重劃範圍外框：所有葉節點街廓的幾何聯集
    outer_boundary_pts = None
    outer_area = 0.0
    try:
        from shapely.ops import unary_union
        union_geom = unary_union([p['shapely'] for p in leaf_polys])
        outer_area = union_geom.area
        # 擷取外框座標（取外環；若為 MultiPolygon 取最大多邊形）
        if union_geom.geom_type == 'Polygon':
            outer_boundary_pts = list(union_geom.exterior.coords)
        elif union_geom.geom_type == 'MultiPolygon':
            biggest = max(union_geom.geoms, key=lambda g: g.area)
            outer_boundary_pts = list(biggest.exterior.coords)
    except Exception:
        # 退而求其次：用面積最大的複合多邊形
        compound = [p for p in poly_objs if p['contain_cnt'] >= 1]
        if compound:
            biggest_compound = max(compound, key=lambda p: p['area_m2'])
            outer_boundary_pts = biggest_compound['vertices']
            outer_area = biggest_compound['area_m2']

    # 對每個葉節點街廓找標籤 + 分類屬性文字
    # 🆕 Task AB（Phase 11）：放寬至接受任何非空文字 + 解析「R1-住宅」「Rd1_道路」等格式
    # 🐛 Hotfix：必須排除地籍文字層（CADASTRAL_TEXT / OPAR…），否則街廓內之地號（如「628」）
    #            會被誤判為街廓 label。同時對 BLOCK 層之文字加分以提升優先權。
    _BLOCK_TEXT_LAYERS = ('BLOCK', 'BLOCKS', 'STREET_BLOCK', 'STREETBLOCK',
                           'BLK', 'BLKS', 'R_BLOCK', '街廓', 'BLOCK_LABEL', 'BLOCK_TEXT')
    # 引用模組層之地籍文字層別名（已於 parse_ua_cadastral_dxf 區段定義）
    # 為避免 NameError（cache_data 跨檔案環境），使用 globals().get fallback
    _CADASTRAL_TEXT_FILTER = (
        globals().get('_CAD_PARCEL_TEXT_LAYERS')
        or ('OPAR', 'CADASTRAL_TEXT', 'PARCEL_TEXT', 'PARCEL', '地號', 'PARCEL_NO')
    )

    def _layer_in(lname, aliases):
        if not lname:
            return False
        ln = str(lname).strip().upper()
        return any(ln == a.upper() for a in aliases)

    polygons_out = []
    for p in leaf_polys:
        label = ''
        category_raw = ''
        # 收集落在多邊形內的所有候選文字（同時保留圖層資訊）
        _inside_texts = []
        for t in texts:
            try:
                if p['shapely'].contains(Point(t['x'], t['y'])):
                    _inside_texts.append({'text': t['text'], 'layer': t.get('layer', '')})
            except Exception:
                continue
        # 解析每筆文字 → 排序：BLOCK 層 + 含分隔（label+category）優先；其次有 label
        _parsed = []
        for ti in _inside_texts:
            tx = ti['text']
            tly = ti['layer']
            # 🐛 Hotfix：忽略地籍文字（避免地號被誤判為街廓 label）
            if _layer_in(tly, _CADASTRAL_TEXT_FILTER):
                continue
            _lbl, _cat = _parse_block_attribute_text(tx)
            if not _lbl:
                continue
            # 評分：BLOCK 層加 1000 分（最強優先）；含 category 加 100 分；
            #       R 字頭加 10 分；長度 < 10 加 1 分
            _score = 0
            if _layer_in(tly, _BLOCK_TEXT_LAYERS):
                _score += 1000
            if _cat:
                _score += 100
            if str(_lbl).upper().startswith('R'):
                _score += 10
            if len(_lbl) < 10:
                _score += 1
            _parsed.append((_score, _lbl, _cat, tx))
        if _parsed:
            _parsed.sort(key=lambda x: -x[0])
            _, label, category_raw, _ = _parsed[0]
        # ── 任務二：執行幾何還原（道路截角識別 + 真 S 長計算）──────────────
        try:
            geom_restore = restore_block_geometry(p['shapely'])
        except Exception as _gex:
            geom_restore = {
                'classification': {'edges': [], 'front_idx': -1,
                                   'cutoff_idxs': [], 'side_idxs': []},
                'theoretical_corners': [], 'cutoff_total_area': 0.0,
                'true_S_length': 0.0,
                'true_side_length_left': 0.0, 'true_side_length_right': 0.0,
                'mbr_long_m': 0.0, 'mbr_short_m': 0.0, 'mbr_polygon': None,
                's_corrected_by_mbr': False, 'side_corrected_by_mbr': False,
                'notes': [f"幾何還原失敗：{_gex}"],
            }
        polygons_out.append({
            'id': p['id'],
            'layer': p['layer'],
            'vertices': p['vertices'],
            'area_m2': p['area_m2'],
            'centroid': p['centroid'],
            'label': label,
            # 🆕 Task AB：DXF 文字解析出之分類（原始字串，未正規化），供 UI 設定預設值
            'category_raw': category_raw,
            'is_outer': False,
            # 任務二輸出（供 Tab 3 街角地與 G 值演算法使用）
            'geom_restore': geom_restore,
            'cutoff_total_area_m2': geom_restore['cutoff_total_area'],
            'true_S_length':        geom_restore['true_S_length'],
            'true_side_length_left':  geom_restore['true_side_length_left'],
            'true_side_length_right': geom_restore['true_side_length_right'],
            'mbr_long_m':  geom_restore.get('mbr_long_m', 0.0),
            'mbr_short_m': geom_restore.get('mbr_short_m', 0.0),
        })

    return {
        'polygons': polygons_out,
        'outer_boundary': outer_boundary_pts,
        'outer_area_m2': outer_area,
        'texts': texts,
        'layer_names': sorted(set(rp['layer'] for rp in raw_polys)),
    }


# ============ 任務一（深度重構）：DXF 地號 / 面積 Regex 過濾 ============
import re as _re
# 🆕 Phase 11 Hotfix：地號格式擴充
# 支援：
#   1. 純數字（含隱含 -0 子號）：628、713、715、717、718、719
#   2. 母號-子號：628-44、713-12
#   3. 段碼-母號-子號（前綴段碼）：13-628-44
#   4. 母號-子號-段代碼（後綴段碼）：628-0-13、628-44-13
# 正規表示式：1~3 個數字段，以「-」分隔
_PARCEL_NUMBER_RE = _re.compile(r'^\d+(?:-\d+){0,2}$')
_AREA_VALUE_RE = _re.compile(r'^\d+\.?\d*$')              # 面積格式：純數字（含小數）


# 🆕 Phase 11 Task AA：地籍 DXF 圖層名稱別名（One-File Workflow 支援）
# 使用者可能將地籍資料命名為 OPAR / CADASTRAL_TEXT / PARCEL_TEXT / 地號 等不同圖層名
_CAD_PARCEL_TEXT_LAYERS = ('OPAR', 'CADASTRAL_TEXT', 'PARCEL_TEXT', 'PARCEL', '地號', 'PARCEL_NO')
_CAD_AREA_TEXT_LAYERS = ('AREA', 'CADASTRAL_AREA', 'PARCEL_AREA', '面積')
_CAD_BOUNDARY_LAYERS = ('ROAD', 'CADASTRAL_BOUND', 'CADASTRAL_BOUNDARY', 'BOUNDARY',
                         'PARCEL_BOUND', 'PARCEL_BOUNDARY', '地籍線')

# 🆕 Phase 9 Issue 1：地籍解析黑名單 — 凡圖層名稱含此關鍵字者，
# 即使匹配 _CAD_BOUNDARY_LAYERS 也不視為地籍線（避免街廓/輔助線混入）
_CAD_EXCLUDED_LAYERS_FROM_CADASTRAL = (
    'BLOCK', 'BLOCKS', '街廓', 'BLOCK_LINE', 'BLOCK_BOUNDARY',
    'BASELINE', 'BASE_LINE', '基準線',
    'CENTERLINE', 'CENTER_LINE', '中心線',
    'FRONT_LINE', 'FRONTLINE',
    'SIDE_LINE', 'SIDELINE',
)


def _is_cad_layer_match(layer_name: str, candidates: tuple) -> bool:
    """大小寫不敏感比對 DXF 圖層名稱是否屬於候選別名集合。"""
    if not layer_name:
        return False
    ln = str(layer_name).strip().upper()
    return any(ln == c.upper() for c in candidates)


def _is_excluded_from_cadastral(layer_name: str) -> bool:
    """
    🆕 Phase 9 Issue 1：判定圖層是否應排除於地籍線解析

    凡 layer 名稱含 BLOCK / BASELINE / CENTERLINE / FRONT_LINE / SIDE_LINE / 街廓 / 基準線 / 中心線
    即視為「街廓或輔助圖層」，不應參與地籍 polygon 拓樸重建。
    使用「子字串比對」（含 .upper()）以涵蓋變體命名（如 BLOCK_INNER, MY_BASELINE 等）。
    """
    if not layer_name:
        return False
    ln = str(layer_name).strip().upper()
    for ex in _CAD_EXCLUDED_LAYERS_FROM_CADASTRAL:
        if ex.upper() in ln:
            return True
    return False


@st.cache_data(show_spinner=False, max_entries=4)
def parse_ua_cadastral_dxf(file_bytes: bytes) -> dict:
    """
    解析重劃前地籍 DXF（如 UA0013.DXF）
    ※ 以 @st.cache_data 快取：避免 Tab 切換或點選互動觸發重新解析
    🆕 Phase 11 Task AA：支援多種圖層命名（One-File Workflow）
      - 地號標籤：OPAR / CADASTRAL_TEXT / PARCEL_TEXT / PARCEL / 地號
      - 面積標籤：AREA / CADASTRAL_AREA / PARCEL_AREA / 面積
      - 界址線：ROAD / CADASTRAL_BOUND / CADASTRAL_BOUNDARY / BOUNDARY / PARCEL_BOUND / 地籍線
    回傳：
      {
        'parcels': [{'parcel':'628-34-1', 'x':..., 'y':..., 'area_m2':float}, ...],
        'boundary_lines': [((x1,y1),(x2,y2)), ...],  # 地籍線
      }
    """
    doc = _read_dxf_any_encoding(file_bytes)
    msp = doc.modelspace()

    # 地號標籤（中心）— 以 Regex 過濾非地號文字（如「道路」「公園」等注記）
    # 🚨 Phase 9.12 Issue 2：Tab 1 白名單比對 — 不在 Tab 1 之 OPAR 文字（區外文字 / 雜訊）直接捨棄
    _tab1_landno_set = set()
    try:
        _own_map_for_filter = st.session_state.get('t8_ownership_map', {}) or {}
        # 收集 Tab 1 所有地號（含原始 + 正規化）作為白名單
        for k in _own_map_for_filter.keys():
            _ks = str(k).strip()
            if _ks:
                _tab1_landno_set.add(_ks)
                _norm_k = _normalize_landno_module(_ks)
                if _norm_k:
                    _tab1_landno_set.add(_norm_k)
    except Exception:
        _tab1_landno_set = set()

    opar_items = []
    _filtered_opar_texts = []   # 記錄被過濾的文字，供診斷用
    _filtered_by_whitelist = [] # 因不在 Tab 1 白名單被捨棄者
    for t in msp.query('TEXT'):
        try:
            if _is_cad_layer_match(t.dxf.layer, _CAD_PARCEL_TEXT_LAYERS):
                # Issue 2：強制 .strip() 並去除全形空格
                raw_text = (t.dxf.text or '').strip().replace('　', '').replace(' ', '')
                if not raw_text:
                    continue
                if not _PARCEL_NUMBER_RE.match(raw_text):
                    _filtered_opar_texts.append(raw_text)
                    continue
                # 🚨 Issue 2：Tab 1 白名單比對（含正規化雙向比對）
                if _tab1_landno_set:
                    norm_raw = _normalize_landno_module(raw_text)
                    in_whitelist = (
                        raw_text in _tab1_landno_set
                        or norm_raw in _tab1_landno_set
                    )
                    if not in_whitelist:
                        _filtered_by_whitelist.append(raw_text)
                        continue
                opar_items.append({
                    'parcel': raw_text,
                    'x': float(t.dxf.insert.x),
                    'y': float(t.dxf.insert.y),
                })
        except Exception:
            continue
    # 將被過濾的 OPAR 文字存入 session_state 供 UI 診斷
    try:
        st.session_state['f3_dxf_filtered_opar'] = _filtered_opar_texts
        st.session_state['f3_dxf_filtered_by_whitelist'] = _filtered_by_whitelist
    except Exception:
        pass  # 非 Streamlit 環境（如單元測試）靜默略過

    # 面積標籤 — 以 Regex 校驗數值格式
    area_items = []
    for t in msp.query('TEXT'):
        try:
            if _is_cad_layer_match(t.dxf.layer, _CAD_AREA_TEXT_LAYERS):
                raw_text = (t.dxf.text or '').strip()
                if not raw_text:
                    continue
                if _AREA_VALUE_RE.match(raw_text):
                    area_val = float(raw_text)
                else:
                    area_val = 0.0   # 不符數值格式 → 忽略
                area_items.append({
                    'area_m2': area_val,
                    'x': float(t.dxf.insert.x),
                    'y': float(t.dxf.insert.y),
                })
        except Exception:
            continue

    # OPAR 與 AREA 通常落在同一地籍中心（AREA 會略為偏移 1m 左右）
    # 以最近鄰配對
    parcels = []
    used = [False] * len(area_items)
    for o in opar_items:
        best_i, best_d = -1, float('inf')
        for i, a in enumerate(area_items):
            if used[i]:
                continue
            d = (a['x'] - o['x'])**2 + (a['y'] - o['y'])**2
            if d < best_d:
                best_d = d
                best_i = i
        area_m2 = 0.0
        if best_i >= 0 and best_d < 25.0:  # 5m 內視為同一地籍
            area_m2 = area_items[best_i]['area_m2']
            used[best_i] = True
        parcels.append({
            'parcel': o['parcel'],
            'x': o['x'],
            'y': o['y'],
            'area_m2': area_m2,
        })

    # 界址線（地籍線）— 支援多種圖層命名
    # 🆕 Phase 9 Issue 1：黑名單先於白名單，杜絕街廓/基準線/中心線等混入地籍 polygon 拓樸重建
    boundary_lines = []
    _excluded_layer_count = 0
    for e in msp.query('LINE'):
        try:
            # Phase 9：黑名單優先（凡圖層名稱含 BLOCK/BASELINE/CENTERLINE/...，即使在白名單內亦排除）
            if _is_excluded_from_cadastral(e.dxf.layer):
                _excluded_layer_count += 1
                continue
            if _is_cad_layer_match(e.dxf.layer, _CAD_BOUNDARY_LAYERS):
                boundary_lines.append((
                    (float(e.dxf.start.x), float(e.dxf.start.y)),
                    (float(e.dxf.end.x), float(e.dxf.end.y)),
                ))
        except Exception:
            continue
    # 同時掃描 LWPOLYLINE 之線段（部分 CAD 軟體會把界址線存為輕量多邊線而非單獨 LINE）
    for e in msp.query('LWPOLYLINE'):
        try:
            # Phase 9：黑名單優先
            if _is_excluded_from_cadastral(e.dxf.layer):
                _excluded_layer_count += 1
                continue
            if _is_cad_layer_match(e.dxf.layer, _CAD_BOUNDARY_LAYERS):
                pts = [(p[0], p[1]) for p in e.get_points()]
                for i in range(len(pts) - 1):
                    boundary_lines.append((pts[i], pts[i + 1]))
                if e.closed and len(pts) >= 2:
                    boundary_lines.append((pts[-1], pts[0]))
        except Exception:
            continue
    # Phase 9：寫入診斷
    try:
        import streamlit as _st_diag9
        if _excluded_layer_count > 0:
            _st_diag9.session_state['f3_cadastral_excluded_layer_count'] = _excluded_layer_count
    except Exception:
        pass

    # 🚨 Phase 9.4 終極修復：徹底拆除「幾何刪線」(_filter_lines_overlapping_blocks)
    # 法定鐵律：地籍線本來就應該與街廓邊重合（街廓邊上之地號其外緣即為街廓邊）。
    # 用「距離」來刪除地籍線會誤殺合法外緣 → 628-20 等地塊「外牆失蹤」無法封閉成多邊形。
    # 拓樸純淨度 100% 依賴圖層黑名單 (_is_excluded_from_cadastral)，
    # 凡 BLOCK / BASELINE / CENTERLINE / FRONT_LINE / SIDE_LINE 已於上方逐筆 LINE/LWPOLYLINE
    # 讀取階段即跳過。CADASTRAL_BOUND 圖層之線段一律無條件保留並參與 polygonize。

    # 用 ROAD 界址線 + OPAR 地號 → 重建重劃前地籍多邊形
    parcel_polygons = reconstruct_parcel_polygons(boundary_lines, opar_items)

    return {'parcels': parcels, 'boundary_lines': boundary_lines, 'parcel_polygons': parcel_polygons}


def _parse_block_attribute_text(text: str):
    """
    🆕 Phase 7 Module 1：解析街廓內部屬性文字
    格式：'R1-住宅' / 'Rd1_道路' / 'R2 商業區'
    回傳：(label, category) tuple；無分隔字元時 (text, '')
    """
    if not text:
        return None, None
    t = str(text).strip()
    if not t:
        return None, None
    for sep in ('-', '_', ' ', '　'):   # 含全形空格
        if sep in t:
            parts = t.split(sep, 1)
            label = parts[0].strip()
            category = parts[1].strip() if len(parts) > 1 else ''
            return label, category
    return t, ''


def parse_cad_precision_layers(doc, classified_blocks: list) -> dict:
    """
    🆕 Phase 7 Module 1：CAD 精準圖層直讀
    🐛 Phase 11 Hotfix：
      - BASELINE：只綁定至「可建築土地」類別之街廓（避免共用邊界誤綁至道路）；
        若多個可建築街廓皆相鄰 → 取最小距離者
      - SIDE_LINE：分 left/right 儲存（同一街廓兩側皆有 side line 時不再覆寫）

    掃描 DXF 之 BASELINE / CENTERLINE / FRONT_LINE / SIDE_LINE 圖層中之
    LINE / LWPOLYLINE / POLYLINE，自動綁定至對應街廓。

    圖層分流：
      - BASELINE     → f3_manual_baseline
                       （v3.1 正名：BASELINE = 屁股線/街廓內側分配基準線，
                        畫在屁股線本身、不平移 0.5m；存「point + angle_deg」即足夠定義
                        一條線——G 計算只取其方向 angle 當 allocation_dir，
                        point 僅 FRONT_LINE 缺失時當 d_hat 備援錨點。兼任分配線。）
      - CENTERLINE   → f3_manual_road_centerlines（中點落於道路類街廓）
      - FRONT_LINE   → f3_cad_front_lengths（中點落於可建築街廓；推進方向 d_hat 主來源）
      - SIDE_LINE    → f3_cad_side_lengths（左/右獨立；有則該側有街角，見 W-A.3 §4 標記）

    回傳：
      {
        'baselines': {block_label: {'point':(x,y), 'angle_deg':float, 'enabled':True, 'source':'cad'}, ...},
        'centerlines': {road_label: [(x,y), (x,y)], ...},
        'front_lengths': {block_label: float, ...},
        'side_lengths': {block_label: float, ...},  # 向下相容（取 max(left, right)）
        'side_lengths_by_side': {block_label: {'left': float, 'right': float}, ...},  # 🆕
        'diagnostics': {'layers_found':[], 'unbound':[]},
      }
    """
    from shapely.geometry import LineString as _SL, Point as _SP, Polygon as _SP_poly
    import math as _m

    result = {
        'baselines': {}, 'centerlines': {},
        'front_lengths': {}, 'side_lengths': {},
        'side_lengths_by_side': {},
        # 🚨 W-B §1-1：SIDE_LINE 座標含端點（供街角規定範圍構造）
        # 結構：{block_label: {'left': {'p1','p2','mid','length'}, 'right': {...}}}
        'side_lines_by_side': {},
        # 🆕 Phase 11 v2：保存 FRONT_LINE 完整端點資訊，供 G 值迭代使用 d_hat
        # 結構：{block_label: {'p1': (x, y), 'p2': (x, y), 'angle_deg': float, 'length': float}}
        'front_lines': {},
        # 🆕 Hotfix：記錄各圖層在 DXF 中找到對應街廓之原始線條數
        # （不等於 bind 後之 dict 長度；用於 UI 顯示真實 DXF 條數）
        'baselines_matched_count': 0,
        'centerlines_matched_count': 0,
        'front_lines_matched_count': 0,
        'side_lines_matched_count': 0,
        'alloc_dir_by_block': {},          # 🚨 W-B §1-3：宗地分配線方向 {blk: (ux, uy)}
        'side_unmatched_warnings': [],     # 🚨 W-B §1-1b：端點未吻合警示
        'diagnostics': {'layers_found': [], 'unbound': []},
    }
    if doc is None or not classified_blocks:
        return result

    try:
        msp = doc.modelspace()
    except Exception:
        return result

    # 預先建立 block label → (block_dict, shapely polygon, centroid, front_dir)
    blk_polys = {}
    for b in (classified_blocks or []):
        verts = b.get('vertices') or []
        lbl = b.get('label', '')
        if len(verts) >= 3 and lbl:
            try:
                p = _SP_poly(verts)
                if not p.is_valid:
                    p = p.buffer(0)
                cen = p.centroid
                # 取得 front_dir（從 geom_restore.classification.front_idx）
                fdx, fdy = 1.0, 0.0   # 預設 X 軸
                try:
                    _gr = b.get('geom_restore') or {}
                    _cls = _gr.get('classification') or {}
                    _edges = _cls.get('edges') or []
                    _fi = _cls.get('front_idx', -1)
                    if 0 <= _fi < len(_edges):
                        _fe = _edges[_fi]
                        _p1 = _fe.get('p1'); _p2 = _fe.get('p2')
                        if _p1 and _p2:
                            _dx = _p2[0] - _p1[0]; _dy = _p2[1] - _p1[1]
                            _L = (_dx ** 2 + _dy ** 2) ** 0.5
                            if _L > 1e-6:
                                fdx, fdy = _dx / _L, _dy / _L
                except Exception:
                    pass
                blk_polys[lbl] = (b, p, (cen.x, cen.y), (fdx, fdy))
            except Exception:
                continue

    if not blk_polys:
        return result

    _layer_map = {'BASELINE', 'CENTERLINE', 'FRONT_LINE', 'SIDE_LINE', 'ALLOC_LINE'}
    _layers_found = set()

    try:
        entities = list(msp.query('LINE LWPOLYLINE POLYLINE'))
    except Exception:
        return result

    # 🆕 Phase 11 Hotfix：先收集所有 SIDE_LINE，後用「貪婪式 1:1 配對 +
    # 街角地校正」批次處理（避免共用邊界 SIDE_LINE 只綁第一個 block 的問題）
    _side_line_buffer = []   # list of {'pts','length','mid'}

    # 可建築土地類別（依 F3_CATEGORY_BURDEN 反查）
    _buildable_blocks = {
        lbl: (b, p, cen, fd) for lbl, (b, p, cen, fd) in blk_polys.items()
        if F3_CATEGORY_BURDEN.get(b.get('category', ''), '') == '可建築土地'
    }

    for entity in entities:
        try:
            layer = str(entity.dxf.layer).upper()
        except Exception:
            continue
        if layer not in _layer_map:
            continue
        _layers_found.add(layer)

        # 取得線段 polyline 點序列
        try:
            etype = entity.dxftype()
            if etype == 'LINE':
                p1 = (float(entity.dxf.start.x), float(entity.dxf.start.y))
                p2 = (float(entity.dxf.end.x), float(entity.dxf.end.y))
                pts = [p1, p2]
            elif etype == 'LWPOLYLINE':
                pts = [(float(p[0]), float(p[1])) for p in entity.get_points()]
            else:   # POLYLINE
                pts = [(float(v.dxf.location.x), float(v.dxf.location.y))
                       for v in entity.vertices]
            if len(pts) < 2:
                continue
            ls = _SL(pts)
            mid_pt = ls.interpolate(0.5, normalized=True)
        except Exception:
            continue

        if layer == 'BASELINE':
            # 🆕 W-A.3 §5：BASELINE = 屁股線（不平移）。此處將線段起點原封存為 point、
            #   線段角度存為 angle_deg——解析階段全無 0.5m 平移（「取消 0.5m」無需動程式）。
            # 🐛 Hotfix：只綁定至「可建築土地」類別之街廓（避免共用邊界誤綁至道路）
            # 計數所有 DXF 中找得到對應 buildable 街廓之 baseline
            try:
                dx = pts[1][0] - pts[0][0]
                dy = pts[1][1] - pts[0][1]
                angle = _m.degrees(_m.atan2(dy, dx))
            except Exception:
                continue
            _candidates = []   # [(distance, blk_lbl)]
            for blk_lbl, (blk, blk_poly, cen, fd) in _buildable_blocks.items():
                try:
                    d = blk_poly.distance(ls)
                    if d < 2.0:   # 2.0m 容差（CAD 繪製誤差）
                        _candidates.append((d, blk_lbl))
                except Exception:
                    continue
            if _candidates:
                # 🆕 計入 raw matched count
                result['baselines_matched_count'] += 1
                _candidates.sort(key=lambda x: x[0])
                # 貪婪式：嘗試指派給「尚未綁定」之街廓；若全部已綁則指派給最近且線段較長者
                _winner_lbl = None
                for _d, _lbl in _candidates:
                    if _lbl not in result['baselines']:
                        _winner_lbl = _lbl
                        break
                if _winner_lbl is None:
                    _winner_lbl = _candidates[0][1]
                _existing = result['baselines'].get(_winner_lbl)
                if _existing is None or float(ls.length) > _existing.get('_length', 0):
                    result['baselines'][_winner_lbl] = {
                        'point': (float(pts[0][0]), float(pts[0][1])),
                        'angle_deg': float(angle),
                        'enabled': True,
                        'source': 'cad',
                        '_length': float(ls.length),
                    }
            else:
                result['diagnostics']['unbound'].append(('BASELINE', pts[0]))

        elif layer == 'CENTERLINE':
            # 中點落於道路類街廓 → 綁定
            _matched_cl = False
            for blk_lbl, (blk, blk_poly, cen, fd) in blk_polys.items():
                if not _is_road_like_category(blk.get('category', '')):
                    continue
                try:
                    mid_p = _SP(mid_pt.x, mid_pt.y)
                    if blk_poly.contains(mid_p) or blk_poly.distance(mid_p) < 2.0:
                        result['centerlines'][blk_lbl] = [
                            (float(p[0]), float(p[1])) for p in pts
                        ]
                        _matched_cl = True
                        break
                except Exception:
                    continue
            if _matched_cl:
                result['centerlines_matched_count'] += 1

        elif layer == 'FRONT_LINE':
            # 中點所屬街廓 + 線段長度（取最大者，避免被短線覆寫）
            # 🆕 Phase 11 v2：同步保存 FRONT_LINE 端點 + 角度（d_hat 來源）
            length = float(ls.length)
            _matched_fl = False
            for blk_lbl, (blk, blk_poly, cen, fd) in blk_polys.items():
                try:
                    mid_p = _SP(mid_pt.x, mid_pt.y)
                    # FRONT_LINE 容差放寬 5m（CAD 繪製可能略偏外側）
                    if blk_poly.contains(mid_p) or blk_poly.distance(mid_p) < 5.0:
                        _prev = result['front_lengths'].get(blk_lbl, 0.0)
                        if length > _prev:
                            result['front_lengths'][blk_lbl] = length
                            # 同步更新 front_lines（取最長者為主代表）
                            try:
                                _dx_fl = pts[1][0] - pts[0][0]
                                _dy_fl = pts[1][1] - pts[0][1]
                                _ang_fl = _m.degrees(_m.atan2(_dy_fl, _dx_fl))
                            except Exception:
                                _ang_fl = 0.0
                            result['front_lines'][blk_lbl] = {
                                'p1': (float(pts[0][0]), float(pts[0][1])),
                                'p2': (float(pts[-1][0]), float(pts[-1][1])),
                                'angle_deg': float(_ang_fl),
                                'length': length,
                            }
                        _matched_fl = True
                        break
                except Exception:
                    continue
            if _matched_fl:
                result['front_lines_matched_count'] += 1

        elif layer == 'SIDE_LINE':
            # 🆕 Phase 11 Hotfix：先 buffer，後批次處理（greedy 1:1）
            length = float(ls.length)
            _side_line_buffer.append({
                'pts': pts, 'length': length,
                'mid_x': float(mid_pt.x), 'mid_y': float(mid_pt.y),
            })

        elif layer == 'ALLOC_LINE':
            # 🚨 W-B §1-3：宗地分配線 — 取方向單位向量，歸屬至最近可建築街廓
            # spec：位置不重要（使用者隨意位置畫），只取方向
            import math as _m_al
            _best_al_blk = None; _best_al_d = float('inf')
            for blk_lbl, (blk, blk_poly, cen, fd) in _buildable_blocks.items():
                try:
                    mid_p_al = _SP(mid_pt.x, mid_pt.y)
                    d_al = (0.0 if blk_poly.contains(mid_p_al)
                            else float(blk_poly.distance(mid_p_al)))
                    if d_al < _best_al_d:
                        _best_al_d = d_al
                        _best_al_blk = blk_lbl
                except Exception:
                    continue
            if _best_al_blk is not None and _best_al_d < 50.0:
                _dx_al = float(pts[-1][0]) - float(pts[0][0])
                _dy_al = float(pts[-1][1]) - float(pts[0][1])
                _L_al = (_dx_al ** 2 + _dy_al ** 2) ** 0.5
                if _L_al > 1e-6:
                    result['alloc_dir_by_block'][_best_al_blk] = (
                        float(_dx_al / _L_al), float(_dy_al / _L_al)
                    )

    # 🚨 W-B §1-1：SIDE_LINE 配對 — 以 FRONT_LINE 端點容差法定左右側
    # p1=左端、p2=右端（FRONT_LINE 方向約定）；廢除舊 4 階段 cross-product + swap 校正
    _TOL_SL = 0.5   # 端點重合容差（m）
    import math as _math_sl

    def _sl_dist2(a, b):
        return ((a[0] - b[0]) ** 2 + (a[1] - b[1]) ** 2) ** 0.5

    for sl in _side_line_buffer:
        _spts = sl.get('pts') or []
        if len(_spts) < 2:
            continue
        _s_ends = [(float(_spts[0][0]), float(_spts[0][1])),
                   (float(_spts[-1][0]), float(_spts[-1][1]))]
        _mid_sl = (sl['mid_x'], sl['mid_y'])
        _hit_blk = None; _hit_side = None; _dmin_all = float('inf')
        for blk_lbl, _fl in result['front_lines'].items():
            _p1 = _fl['p1']; _p2 = _fl['p2']
            for _se in _s_ends:
                _dmin_all = min(_dmin_all,
                                _sl_dist2(_se, _p1), _sl_dist2(_se, _p2))
            for _se in _s_ends:
                if _sl_dist2(_se, _p1) < _TOL_SL:
                    _hit_blk = blk_lbl; _hit_side = 'left'; break
                if _sl_dist2(_se, _p2) < _TOL_SL:
                    _hit_blk = blk_lbl; _hit_side = 'right'; break
            if _hit_blk:
                break
        if _hit_blk and _hit_side:
            _by_side_wb = result['side_lines_by_side'].setdefault(_hit_blk, {})
            _by_side_wb[_hit_side] = {
                'p1': (float(_spts[0][0]), float(_spts[0][1])),
                'p2': (float(_spts[-1][0]), float(_spts[-1][1])),
                'mid': _mid_sl,
                'length': float(sl['length']),
            }
            # 向下相容：同步更新 side_lengths_by_side / side_lengths
            _legacy_bs = result['side_lengths_by_side'].setdefault(_hit_blk, {})
            _legacy_bs[_hit_side] = float(sl['length'])
            result['side_lengths'][_hit_blk] = max(
                _legacy_bs.get('left', 0.0), _legacy_bs.get('right', 0.0)
            )
            result['side_lines_matched_count'] += 1
        else:
            # §1-1b：端點未吻合任何 FRONT_LINE 端點（容差 0.5m）
            result['side_unmatched_warnings'].append({
                'mid': _mid_sl,
                'dmin': round(_dmin_all, 3),
                'tol': _TOL_SL,
            })

    # 移除 BASELINE 之 _length 內部欄位（不對外輸出）
    for _lbl, _bv in result['baselines'].items():
        _bv.pop('_length', None)

    result['diagnostics']['layers_found'] = sorted(_layers_found)
    return result


def export_allocated_dxf(g_rows: list, classified_blocks: list,
                          manual_road_centerlines: dict = None) -> bytes:
    """
    🆕 Phase 7 Module 4：匯出重劃後地籍 DXF（ezdxf）

    圖層分流：
      - BLOCK_BOUNDARY : 街廓底圖 LWPOLYLINE（灰色）
      - NEW_PARCEL     : 私人分配土地 LWPOLYLINE（綠色 ACI=3）
      - OFFSET_LAND    : 抵費地 LWPOLYLINE（黃色 ACI=2）
      - NEW_CENTERLINE : 道路中心線 LINE（紅色 ACI=1）
      - PARCEL_TEXT    : 暫編地號 + 面積 TEXT 標註

    回傳：DXF bytes（utf-8 編碼字串轉換）
    """
    import ezdxf
    import io as _io_dxf
    from shapely.geometry import Polygon as _SP_poly

    doc = ezdxf.new('R2010', setup=True)
    # 設定圖層顏色（AutoCAD Color Index）
    _layer_specs = [
        ('BLOCK_BOUNDARY', 8),    # 深灰
        ('NEW_PARCEL', 3),        # 綠
        ('OFFSET_LAND', 2),       # 黃
        ('NEW_CENTERLINE', 1),    # 紅
        ('PARCEL_TEXT', 7),       # 白/黑
    ]
    for layer_name, color in _layer_specs:
        if layer_name not in doc.layers:
            doc.layers.add(name=layer_name, color=color)

    msp = doc.modelspace()

    # 1. 街廓底圖
    for b in (classified_blocks or []):
        verts = b.get('vertices') or []
        if len(verts) < 3:
            continue
        try:
            msp.add_lwpolyline(
                list(verts),
                close=True,
                dxfattribs={'layer': 'BLOCK_BOUNDARY'},
            )
        except Exception:
            pass

    # 2. 分配 cut polygons + 文字標註
    for r in (g_rows or []):
        coords = r.get('cut_coords') or []
        if len(coords) < 3:
            continue
        is_offset = (r.get('推進側別') == '抵費地')
        layer = 'OFFSET_LAND' if is_offset else 'NEW_PARCEL'
        try:
            msp.add_lwpolyline(
                list(coords),
                close=True,
                dxfattribs={'layer': layer},
            )
        except Exception:
            continue
        # 質心標註
        try:
            cen = _SP_poly(coords).centroid
            if is_offset:
                _label_text = f"抵費地\\P{r.get('幾何面積(㎡)', 0):.2f}㎡"
            else:
                _label_text = (f"{r.get('暫編地號','')}\\P"
                               f"G={r.get('G(㎡)',0):.2f}㎡\\P"
                               f"S={r.get('S(m)',0):.2f}m")
            msp.add_mtext(
                _label_text,
                dxfattribs={
                    'layer': 'PARCEL_TEXT',
                    'char_height': 1.5,
                    'insert': (float(cen.x), float(cen.y)),
                    'attachment_point': 5,
                },
            )
        except Exception:
            pass

    # 3. 道路中心線
    for road_lbl, pts in (manual_road_centerlines or {}).items():
        if not pts or len(pts) < 2:
            continue
        for i in range(len(pts) - 1):
            try:
                msp.add_line(
                    (float(pts[i][0]), float(pts[i][1])),
                    (float(pts[i+1][0]), float(pts[i+1][1])),
                    dxfattribs={'layer': 'NEW_CENTERLINE'},
                )
            except Exception:
                continue

    # 寫入 BytesIO（ezdxf 預設 UTF-8）
    buf = _io_dxf.StringIO()
    doc.write(buf)
    return buf.getvalue().encode('utf-8')


def _apply_xlsx_page_setup(ws, orientation='landscape', fit_to_width=1):
    """🆕 Phase 7 Module 4：openpyxl 列印排版輔助"""
    try:
        ws.page_setup.orientation = orientation
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = fit_to_width
        ws.page_setup.fitToHeight = 0
        ws.print_options.horizontalCentered = True
    except Exception:
        pass


def _setup_xlsx_header(ws, headers: list, row: int = 1):
    """🆕 Phase 7 Module 4：openpyxl 表頭格式（藍底白字粗體）"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    _border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.fill = PatternFill(
            start_color='4472C4', end_color='4472C4', fill_type='solid'
        )
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = _border


def export_legal_excel(g_rows: list, ownership_map: dict,
                        ownership_groups: dict = None,
                        ownership_full_records: dict = None,
                        B_value: float = 0.0, C_value: float = 0.0,
                        total_area: float = 0.0,
                        pre_avg_price: float = 0.0,
                        post_avg_price: float = 0.0,
                        engineering_cost: float = 0.0,
                        redev_cost: float = 0.0,
                        loan_interest: float = 0.0) -> bytes:
    """
    🆕 Phase 7 Module 4：匯出三大法定報表 Excel

    Sheet 1：重劃前後土地分配對照清冊
    Sheet 2：歸戶負擔計算表
    Sheet 3：公共設施用地負擔統計表

    回傳：xlsx bytes
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import io as _io_xls

    wb = Workbook()

    # ── Sheet 1: 對照清冊 ──
    ws1 = wb.active
    ws1.title = "對照清冊"
    ws1['A1'] = '重劃前後土地分配對照清冊'
    ws1['A1'].font = Font(size=14, bold=True)
    ws1.merge_cells('A1:H1')
    headers1 = ['原地號', '原面積(㎡)', '所屬街廓', '暫編地號',
                 '應分配G(㎡)', '幾何分配(㎡)', '增減(㎡)', '街角地']
    _setup_xlsx_header(ws1, headers1, row=3)
    row = 4
    for r in (g_rows or []):
        if r.get('推進側別') == '抵費地':
            continue
        a = float(r.get('a 面積(㎡)', 0) or 0)
        g = float(r.get('G(㎡)', 0) or 0)
        ws1.cell(row=row, column=1, value=str(r.get('原地號', '')))
        ws1.cell(row=row, column=2, value=round(a, 2))
        ws1.cell(row=row, column=3, value=str(r.get('所屬街廓', '')))
        ws1.cell(row=row, column=4, value=str(r.get('暫編地號', '')))
        ws1.cell(row=row, column=5, value=round(g, 2))
        ws1.cell(row=row, column=6, value=round(float(r.get('幾何面積(㎡)', 0) or 0), 2))
        ws1.cell(row=row, column=7, value=round(g - a, 2))
        ws1.cell(row=row, column=8, value=str(r.get('街角地', '否')))
        row += 1
    # 抵費地另起一段
    _offsets = [r for r in (g_rows or []) if r.get('推進側別') == '抵費地']
    if _offsets:
        row += 1
        ws1.cell(row=row, column=1, value='─── 抵費地 ───').font = Font(bold=True)
        row += 1
        for r in _offsets:
            ws1.cell(row=row, column=1, value=str(r.get('暫編地號', '')))
            ws1.cell(row=row, column=2, value=round(float(r.get('幾何面積(㎡)', 0) or 0), 2))
            ws1.cell(row=row, column=3, value=str(r.get('所屬街廓', '')))
            ws1.cell(row=row, column=8, value='抵費地')
            row += 1
    # 欄寬
    for col_letter, width in zip('ABCDEFGH', [16, 14, 14, 18, 14, 14, 12, 12]):
        ws1.column_dimensions[col_letter].width = width
    _apply_xlsx_page_setup(ws1, orientation='landscape', fit_to_width=1)

    # ── Sheet 2: 歸戶負擔計算表 ──
    # 🆕 V12 模組 4-A：新增 4 欄（重劃前總、應配、實配、分配率）
    # 🔧 微調防護 2：Polygon 面積計算頂點數量檢查（_polygon_area_safe）
    def _polygon_area_safe(cut_coords):
        if not cut_coords or len(cut_coords) < 3:
            return 0.0
        try:
            from shapely.geometry import Polygon as _SP_safe
            p = _SP_safe(cut_coords)
            if not p.is_valid:
                p = p.buffer(0)
            return float(p.area)
        except Exception:
            return 0.0

    ws2 = wb.create_sheet("歸戶負擔計算表")
    ws2['A1'] = '歸戶負擔計算表（依市地重劃實施辦法 §29 附件二；V12 加入分配率）'
    ws2['A1'].font = Font(size=14, bold=True)
    ws2.merge_cells('A1:N1')
    headers2 = ['歸戶ID', '原地號(全部)', '宗地數', '原總面積a(㎡)',
                 '公設負擔 a×B(㎡)', '費用負擔 a×C(㎡)',
                 '應配G(㎡)', '實配(㎡)', '差額(㎡)',
                 # 🆕 V12 新增 4 欄
                 '重劃前土地總面積(㎡)', '重劃後應分配土地面積(㎡)',
                 '重劃後實際分配土地面積(㎡)', '土地分配率(%)',
                 '備註']
    _setup_xlsx_header(ws2, headers2, row=3)
    # 依歸戶群組聚合
    _per_group = {}
    for r in (g_rows or []):
        if r.get('推進側別') == '抵費地':
            continue
        gid = _resolve_ownership(r.get('原地號', ''), ownership_map)
        if not gid:
            gid = '(未歸戶)'
        agg = _per_group.setdefault(gid, {
            'parcels': [], 'a_total': 0.0, 'G_total': 0.0,
            'geom_total': 0.0, 'cut_area_total': 0.0,
        })
        agg['parcels'].append(str(r.get('原地號', '')))
        agg['a_total'] += float(r.get('a 面積(㎡)', 0) or 0)
        agg['G_total'] += float(r.get('G(㎡)', 0) or 0)
        agg['geom_total'] += float(r.get('幾何面積(㎡)', 0) or 0)
        # 🆕 V12 模組 4-A：實際 cut polygon 面積累計（含微調防護 2）
        agg['cut_area_total'] += _polygon_area_safe(r.get('cut_coords'))
    row = 4
    for gid, agg in sorted(_per_group.items()):
        a_t = round(agg['a_total'], 2)
        G_t = round(agg['G_total'], 2)
        cut_t = round(agg['cut_area_total'], 2)
        ratio_pct = round((cut_t / a_t * 100.0), 2) if a_t > 0 else 0.0
        ws2.cell(row=row, column=1, value=gid)
        ws2.cell(row=row, column=2, value='、'.join(sorted(set(agg['parcels']))))
        ws2.cell(row=row, column=3, value=len(agg['parcels']))
        ws2.cell(row=row, column=4, value=a_t)
        ws2.cell(row=row, column=5, value=round(a_t * B_value, 2))
        ws2.cell(row=row, column=6, value=round(a_t * C_value, 2))
        ws2.cell(row=row, column=7, value=G_t)
        ws2.cell(row=row, column=8, value=round(agg['geom_total'], 2))
        ws2.cell(row=row, column=9, value=round(agg['geom_total'] - agg['G_total'], 2))
        # 🆕 V12 4 新欄
        ws2.cell(row=row, column=10, value=a_t)         # 重劃前土地總面積（含已整併公設）
        ws2.cell(row=row, column=11, value=G_t)         # 重劃後應分配土地面積
        ws2.cell(row=row, column=12, value=cut_t)       # 重劃後實際分配土地面積
        ws2.cell(row=row, column=13, value=ratio_pct)   # 土地分配率(%)
        # 備註：欄位 14（指出含整併或孤立地）
        _has_orphan = any(_r.get('推進側別') == '🟠 孤立公設地'
                           for _r in (g_rows or [])
                           if _resolve_ownership(_r.get('原地號', ''),
                                                  ownership_map) == gid)
        ws2.cell(row=row, column=14,
                  value='含孤立公設地（虛擬G）' if _has_orphan else '')
        row += 1
    for col_letter, width in zip('ABCDEFGHIJKLMN',
                                   [12, 30, 8, 14, 14, 14, 14, 14, 12,
                                    18, 18, 18, 14, 18]):
        ws2.column_dimensions[col_letter].width = width
    _apply_xlsx_page_setup(ws2, orientation='landscape', fit_to_width=1)

    # ── Sheet 3: 公設用地負擔統計表 ──
    ws3 = wb.create_sheet("公設負擔統計")
    ws3['A1'] = '計算負擔總計表（市地重劃作業手冊附表 22）'
    ws3['A1'].font = Font(size=14, bold=True)
    ws3.merge_cells('A1:E1')
    _setup_xlsx_header(ws3, ['項目', '面積/數值', '單位', '說明', '備註'], row=3)

    _total_offset_area = sum(
        float(r.get('幾何面積(㎡)', 0) or 0)
        for r in (g_rows or [])
        if r.get('推進側別') == '抵費地'
    )
    _total_g = sum(
        float(r.get('G(㎡)', 0) or 0)
        for r in (g_rows or [])
        if r.get('推進側別') != '抵費地'
    )

    _summary_rows = [
        ('重劃區實測面積', total_area, '㎡',
         '街廓總面積（含可建築 + 公設 + 抵充）', ''),
        ('重劃前平均地價', pre_avg_price, '元/㎡', '預期開發分析法 推算', ''),
        ('重劃後平均地價', post_avg_price, '元/㎡', '土地開發分析法 推算', ''),
        ('重劃前後地價上漲率',
         round(post_avg_price / pre_avg_price, 4) if pre_avg_price > 0 else 0,
         '倍', '= 重劃後 ÷ 重劃前', ''),
        ('B 值（一般負擔係數）', round(B_value, 6), '係數',
         '一般負擔總面積 × 重劃前地價 / [重劃後地價 ×(總面積 − 抵充地)]',
         '附件二 ④；已扣除臨街地特別負擔，非「公設用地平均負擔比率」'),
        ('C 值（費用負擔係數）', round(C_value, 6), '係數',
         '(工程 + 重劃 + 利息) / [重劃後地價 ×(總面積 − 公設總面積)]',
         '附件二 ⑤；分母與「費用負擔比率」分母不同'),
        ('—— 法定分項負擔比率（地主實際感受）——', '', '', '', ''),
        ('公共設施用地平均負擔比率（待輸入）', '需另填', '比率',
         '= (共同負擔公設 − 政府已取得公設 − 抵充地) ÷ (重劃區總面積 − 政府已取得公設 − 抵充地)',
         '需 財務分析 提供「政府已取得公設」+「抵充地」面積'),
        ('費用負擔比率（待輸入）', '需另填', '比率',
         '= (工程 + 重劃 + 利息) ÷ [重劃後地價 ×(總面積 − 政府已取得公設 − 抵充地)]',
         '抵費地（R 街廓內）大致對應此項；實務 10-15%'),
        ('地主實際平均負擔比率', '兩項合計', '比率',
         '= 公設用地平均負擔比率 + 費用負擔比率（**非 B+C**）', ''),
        ('工程費用總額', engineering_cost, '元', '財務分析 工程費', ''),
        ('重劃費用總額', redev_cost, '元', '財務分析 重劃費', ''),
        ('貸款利息總額', loan_interest, '元', '財務分析 利息', ''),
        ('—— 分配結果統計 ——', '', '', '', ''),
        ('應分配 G 值總計', round(_total_g, 2), '㎡', '合計所有私人 G(㎡)', ''),
        ('抵費地總面積', round(_total_offset_area, 2), '㎡',
         '雙向夾擠分配後剩餘空地（含未調配孤立公設地保留額度）', ''),
    ]
    row = 4
    for k, v, u, desc, note in _summary_rows:
        ws3.cell(row=row, column=1, value=k)
        ws3.cell(row=row, column=2, value=v)
        ws3.cell(row=row, column=3, value=u)
        ws3.cell(row=row, column=4, value=desc)
        ws3.cell(row=row, column=5, value=note)
        # 強調 B/C/各負擔比率（紅色粗體）
        if k.startswith(('B 值', 'C 值', '公共設施', '費用負擔', '地主實際')):
            for col in (1, 2, 3, 4, 5):
                ws3.cell(row=row, column=col).font = Font(bold=True, color='C00000')
        row += 1
    for col_letter, width in zip('ABCDE', [22, 16, 8, 50, 14]):
        ws3.column_dimensions[col_letter].width = width
    _apply_xlsx_page_setup(ws3, orientation='landscape', fit_to_width=1)

    out = _io_xls.BytesIO()
    wb.save(out)
    return out.getvalue()


def _has_geopandas() -> bool:
    """偵測 geopandas 是否可用（任務二：GeoJSON/SHP 支援為可選功能）"""
    try:
        import geopandas as _gpd  # noqa: F401
        return True
    except Exception:
        return False


def parse_cadastral_geofile(file_bytes: bytes, file_name: str) -> dict:
    """
    任務二：地籍資料多格式解析（DXF / GeoJSON / Shapefile zip）

    根據檔名副檔名 dispatch：
      - .dxf            → 沿用 parse_ua_cadastral_dxf
      - .geojson / .json → 用 geopandas 讀 + CRS 對齊到 EPSG:3826
      - .zip            → 解壓後讀 .shp + CRS 對齊到 EPSG:3826

    地號欄位偵測順序：'地號' > 'parcel' > 'PARCEL' > 'name' > 'NAME' > 'ID' > 'id'

    回傳格式與 parse_ua_cadastral_dxf 一致：
      {
        'parcels': [{'parcel', 'x', 'y', 'area_m2'}, ...],
        'boundary_lines': [],  # GeoJSON/SHP 模式不重建界址線（無此需求）
        'parcel_polygons': [{'parcel', 'polygon_coords', 'area_m2',
                             'centroid_x', 'centroid_y'}, ...],
      }
    """
    fn_lower = (file_name or '').lower()
    # DXF：沿用既有
    if fn_lower.endswith('.dxf'):
        return parse_ua_cadastral_dxf(file_bytes)

    if not _has_geopandas():
        raise RuntimeError(
            "缺少 geopandas 套件。請執行 `pip install geopandas pyproj` 後再試，"
            "或改用 .dxf 格式上傳。"
        )

    import geopandas as gpd
    from io import BytesIO
    import zipfile, tempfile, os

    # 讀 GeoDataFrame
    gdf = None
    if fn_lower.endswith('.geojson') or fn_lower.endswith('.json'):
        gdf = gpd.read_file(BytesIO(file_bytes))
    elif fn_lower.endswith('.zip'):
        # 解壓到暫存資料夾後找 .shp
        with tempfile.TemporaryDirectory() as tmpdir:
            zpath = os.path.join(tmpdir, 'upload.zip')
            with open(zpath, 'wb') as f:
                f.write(file_bytes)
            with zipfile.ZipFile(zpath) as zf:
                zf.extractall(tmpdir)
            shp_files = []
            for root, _dirs, files in os.walk(tmpdir):
                for fn in files:
                    if fn.lower().endswith('.shp'):
                        shp_files.append(os.path.join(root, fn))
            if not shp_files:
                raise ValueError("ZIP 中找不到 .shp 檔，請確認檔案結構（.shp/.shx/.dbf/.prj）")
            gdf = gpd.read_file(shp_files[0])
    else:
        raise ValueError(f"不支援的檔案格式：{file_name}（僅支援 .dxf / .geojson / .zip[shp]）")

    # CRS 對齊到 EPSG:3826（TWD97 二度分帶）
    if gdf is not None and gdf.crs is not None:
        try:
            if gdf.crs.to_epsg() != 3826:
                gdf = gdf.to_crs(epsg=3826)
        except Exception:
            # 無法判定 EPSG → 假設已是 TWD97（不轉換）
            pass

    # 偵測地號欄位
    parcel_field = None
    for cand in ('地號', 'parcel', 'PARCEL', 'name', 'NAME', 'ID', 'id'):
        if cand in gdf.columns:
            parcel_field = cand
            break

    parcels_out = []
    parcel_polygons = []
    for idx, row in gdf.iterrows():
        try:
            geom = row.geometry
            if geom is None or geom.is_empty:
                continue
            # 取地號（若沒欄位用 idx）
            parcel_no = str(row[parcel_field]) if parcel_field else f"AUTO-{idx}"
            # 處理 Polygon / MultiPolygon
            if geom.geom_type == 'Polygon':
                coords = list(geom.exterior.coords)
                area_m2 = float(geom.area)
                centroid = geom.centroid
                parcel_polygons.append({
                    'parcel': parcel_no,
                    'polygon_coords': coords,
                    'area_m2': area_m2,
                    'centroid_x': float(centroid.x),
                    'centroid_y': float(centroid.y),
                })
                parcels_out.append({
                    'parcel': parcel_no,
                    'x': float(centroid.x), 'y': float(centroid.y),
                    'area_m2': area_m2,
                })
            elif geom.geom_type == 'MultiPolygon':
                for sub_idx, sub in enumerate(geom.geoms):
                    coords = list(sub.exterior.coords)
                    area_m2 = float(sub.area)
                    centroid = sub.centroid
                    sub_no = parcel_no if len(geom.geoms) == 1 else f"{parcel_no}-{sub_idx + 1}"
                    parcel_polygons.append({
                        'parcel': sub_no,
                        'polygon_coords': coords,
                        'area_m2': area_m2,
                        'centroid_x': float(centroid.x),
                        'centroid_y': float(centroid.y),
                    })
                    parcels_out.append({
                        'parcel': sub_no,
                        'x': float(centroid.x), 'y': float(centroid.y),
                        'area_m2': area_m2,
                    })
        except Exception:
            continue

    return {
        'parcels': parcels_out,
        'boundary_lines': [],
        'parcel_polygons': parcel_polygons,
    }


def overlay_parcels_to_blocks(parcels: list, classified_blocks: list) -> list:
    """
    將重劃前地號坐落街廓 — 以地號中心點測試是否在街廓多邊形內
    classified_blocks: [{'id', 'vertices', 'label', 'category', 'is_outer', ...}]
    回傳 parcels 清單（新增 'block_label', 'block_category' 欄位）
    """
    from shapely.geometry import Polygon, Point
    poly_map = []  # (label, category, shapely)
    for b in classified_blocks:
        if b.get('is_outer'):
            continue
        try:
            sp = Polygon(b['vertices'])
            if not sp.is_valid:
                sp = sp.buffer(0)
            poly_map.append((b.get('label', '') or f"區塊#{b['id']}", b.get('category', ''), sp))
        except Exception:
            continue

    results = []
    for p in parcels:
        pt = Point(p['x'], p['y'])
        matched_label, matched_cat = '', '區外'
        for lbl, cat, sp in poly_map:
            try:
                if sp.contains(pt):
                    matched_label, matched_cat = lbl, cat
                    break
            except Exception:
                continue
        results.append({**p, 'block_label': matched_label, 'block_category': matched_cat})
    return results


# ============ Tab 3（街廓互動分析）專用常數與輔助函式 ============
# 15 項分類（依都市計畫分區 + 市地重劃實施辦法 P2, P12, P31）
F3_BLOCK_CATEGORIES = [
    "未分類",
    "住宅區", "商業區",
    "道路", "溝渠", "兒童遊樂場", "鄰里公園", "廣場", "綠地",
    "國民小學", "國民中學", "停車場", "零售市場",
    "社會住宅", "機關用地", "其他非共同負擔之公共設施用地",
]
# 負擔屬性：可建築土地 / 共同負擔 / 非共同負擔 / 未分類
F3_CATEGORY_BURDEN = {
    "未分類": "未分類",
    "住宅區": "可建築土地",
    "商業區": "可建築土地",
    "道路": "共同負擔",
    "溝渠": "共同負擔",
    "兒童遊樂場": "共同負擔",
    "鄰里公園": "共同負擔",
    "廣場": "共同負擔",
    "綠地": "共同負擔",
    "國民小學": "共同負擔",
    "國民中學": "共同負擔",
    "停車場": "共同負擔",
    "零售市場": "共同負擔",
    "社會住宅": "非共同負擔",
    "機關用地": "非共同負擔",
    "其他非共同負擔之公共設施用地": "非共同負擔",
}
F3_CATEGORY_COLORS = {
    "未分類": "#BDBDBD",
    "住宅區": "#2E86C1", "商業區": "#8E44AD",
    "道路": "#F39C12", "溝渠": "#3498DB",
    "兒童遊樂場": "#E74C3C", "鄰里公園": "#27AE60",
    "廣場": "#F1C40F", "綠地": "#16A085",
    "國民小學": "#5D6D7E", "國民中學": "#34495E",
    "停車場": "#7F8C8D", "零售市場": "#D35400",
    "社會住宅": "#C0392B", "機關用地": "#2C3E50",
    "其他非共同負擔之公共設施用地": "#95A5A6",
}


def _normalize_block_category(text: str) -> str:
    """
    🆕 Task AB（Phase 11）：將 DXF 中街廓屬性文字（如「住宅」「住」「Rd-道路」）
    模糊比對至 F3_BLOCK_CATEGORIES 標準分類名稱。
    比對規則（依序）：
      1. 完全相符（如「住宅區」「商業區」「道路」）→ 直接回傳
      2. 縮寫對照表（如「住宅」→「住宅區」、「商業」→「商業區」、「公園」→「鄰里公園」）
      3. 子字串包含（如「丁種建築用地」含「住宅」→ 視為住宅區；含「道」「路」→ 道路）
      4. 皆無 → 回傳空字串（讓上層 fallback 至「未分類」或預設值）
    """
    if not text:
        return ''
    t = str(text).strip()
    if not t:
        return ''
    # 1. 完全相符
    if t in F3_BLOCK_CATEGORIES:
        return t
    # 2. 縮寫 / 別名對照
    _aliases = {
        '住宅':   '住宅區',
        '住':     '住宅區',
        '商業':   '商業區',
        '商':     '商業區',
        '公園':   '鄰里公園',
        '鄰里':   '鄰里公園',
        '兒童':   '兒童遊樂場',
        '遊樂場': '兒童遊樂場',
        '小學':   '國民小學',
        '國小':   '國民小學',
        '中學':   '國民中學',
        '國中':   '國民中學',
        '市場':   '零售市場',
        '零售':   '零售市場',
        '社宅':   '社會住宅',
        '社會':   '社會住宅',
        '機關':   '機關用地',
        '停車':   '停車場',
        '停':     '停車場',
        '綠':     '綠地',
        '溝':     '溝渠',
        '渠':     '溝渠',
        '路':     '道路',
        '道':     '道路',
        'RD':     '道路',
        'ROAD':   '道路',
    }
    if t in _aliases:
        return _aliases[t]
    if t.upper() in _aliases:
        return _aliases[t.upper()]
    # 3. 子字串包含（順序由特定 → 一般，避免誤判）
    _contains_rules = [
        ('社會住宅', '社會住宅'), ('社宅', '社會住宅'),
        ('機關', '機關用地'),
        ('停車', '停車場'),
        ('零售', '零售市場'), ('市場', '零售市場'),
        ('國小', '國民小學'), ('小學', '國民小學'),
        ('國中', '國民中學'), ('中學', '國民中學'),
        ('遊樂', '兒童遊樂場'),
        ('公園', '鄰里公園'),
        ('廣場', '廣場'),
        ('綠地', '綠地'),
        ('溝', '溝渠'), ('渠', '溝渠'),
        ('道路', '道路'), ('道', '道路'), ('路', '道路'),
        ('住宅', '住宅區'), ('住', '住宅區'),
        ('商業', '商業區'), ('商', '商業區'),
    ]
    for kw, cat in _contains_rules:
        if kw in t:
            return cat
    return ''


def _render_f3_overlay_figure(classified_blocks, temp_parcels, zones_map,
                              interactive: bool = False, height: int = 560,
                              highlight_temp_no: str = None,
                              auto_zoom_to_highlight: bool = True,
                              visible_layers: dict = None):
    """
    產生 Tab 3「街廓 × 重劃前地籍」套疊圖（效能優化版）
    - 街廓填色：每街廓一個 trace（必要，以保留 hover）
    - 街廓分配線：合併為一個紅色 trace（避免逐街廓 add_trace）
    - 重劃前地籍暫編地號：依區段色彩分群成少數幾個 trace
      （未標區段 → 深灰黑；已標區段 → 彩色粗線）
    - interactive=True 時每個地號仍保留 customdata 供 plotly_events 點選

    新增參數（任務二：暫編地號圖面定位與高亮）：
      highlight_temp_no       要高亮顯示之暫編地號（可為 '(無地號)' 之特殊值）
      auto_zoom_to_highlight  若 True，將地圖視角置中放大於該地號

    🆕 Task E（深度重構）：visible_layers 圖層可見性控制
      visible_layers          dict[str, bool]，控制各 trace 群組是否渲染：
        - 'block_boundaries' (預設 True): 街廓填色 + 邊框
        - 'pre_cadastral'    (預設 True): 重劃前地籍線
        - 'price_zones'      (預設 True): 地價區段著色
      註：interactive=True 時點擊目標 trace 永遠渲染（不受圖層控制，否則點選失效）
    """
    _vl = visible_layers or {}
    _show_blocks = bool(_vl.get('block_boundaries', True))
    _show_cad    = bool(_vl.get('pre_cadastral', True))
    _show_zones  = bool(_vl.get('price_zones', True))
    import plotly.graph_objects as go
    fig = go.Figure()

    # 1️⃣ 街廓填色（保留 hover）— 受 visible_layers['block_boundaries'] 控制
    if _show_blocks:
        for b in classified_blocks:
            color = F3_CATEGORY_COLORS.get(b.get('category', ''), "#BDBDBD")
            xs = [v[0] for v in b['vertices']] + [b['vertices'][0][0]]
            ys = [v[1] for v in b['vertices']] + [b['vertices'][0][1]]
            fig.add_trace(go.Scatter(
                x=xs, y=ys, mode='lines', fill='toself', fillcolor=color,
                line=dict(color=color, width=0.5), opacity=0.22,
                name=f"街廓 {b.get('label', '')}", hoverinfo='skip', showlegend=False,
            ))

        # 2️⃣ 🟥 街廓分配線（合併一個 trace）
        red_xs, red_ys = [], []
        for b in classified_blocks:
            verts = b.get('vertices', [])
            for x, y in verts:
                red_xs.append(x); red_ys.append(y)
            if verts:
                red_xs.append(verts[0][0]); red_ys.append(verts[0][1])
            red_xs.append(None); red_ys.append(None)
        fig.add_trace(go.Scatter(
            x=red_xs, y=red_ys, mode='lines',
            line=dict(color='#D62728', width=2.2),
            name='街廓分配線（紅）', hoverinfo='skip',
        ))

    # 3️⃣ 重劃前地籍：依區段色彩分群
    palette = ['#1F77B4', '#FF7F0E', '#2CA02C', '#D62728', '#9467BD',
               '#8C564B', '#E377C2', '#17BECF', '#BCBD22']
    uniq_zones = sorted({v for v in zones_map.values() if v})
    zone_color_map = {z: palette[i % len(palette)] for i, z in enumerate(uniq_zones)}
    # 群組：未標 + 各區段 → group_lines
    groups = {'(未標註)': {'xs': [], 'ys': [], 'color': '#2C2C2C', 'width': 0.8}}
    for z in uniq_zones:
        groups[z] = {'xs': [], 'ys': [], 'color': zone_color_map[z], 'width': 2.5}

    for tp in temp_parcels:
        # 🚨 Patch E-0.3：ghost sliver 由 consolidated 渲染處染灰，此處跳過
        if tp.get('_is_ghost_sliver', False):
            continue
        coords = tp.get('polygon_coords') or []
        if not coords:
            continue
        zname = zones_map.get(tp['原地號'], '') or '(未標註)'
        g = groups.setdefault(zname, {'xs': [], 'ys': [], 'color': '#2C2C2C', 'width': 0.8})
        for x, y in coords:
            g['xs'].append(x); g['ys'].append(y)
        g['xs'].append(None); g['ys'].append(None)

    # 統一畫分群線條（interactive 與 preview 均相同，避免重複 trace）
    # _show_cad 控制地籍線是否顯示；_show_zones 控制有色區段是否使用色彩
    if _show_cad:
        for zname, g in groups.items():
            # 若關閉「地價區段著色」，已標區段回退至灰色細線
            if zname != '(未標註)' and not _show_zones:
                _line_color = '#888888'; _line_width = 0.8
            else:
                _line_color = g['color']; _line_width = g['width']
            fig.add_trace(go.Scatter(
                x=g['xs'], y=g['ys'], mode='lines',
                line=dict(color=_line_color, width=_line_width),
                name=f"區段 {zname}", hoverinfo='skip',
                showlegend=(zname != '(未標註)' and _show_zones),
            ))
    # interactive 模式：為每個地籍多邊形加一層「透明但可點擊」的填充層作為點擊靶
    # plotly_events 只能在有資料點且 hoverinfo 非 'skip' 的 trace 上收到點擊
    # 故需以 fill='toself' 讓多邊形內部也成為可點擊區域
    # 點擊後由呼叫端取得 x/y 座標，再用 shapely Point.within(Polygon) 判定
    # 達成「鎖多邊形選取」效果（一次點擊 → 唯一多邊形）
    if interactive:
        for tp in temp_parcels:
            # 🚨 Patch E-0.3：ghost sliver 不應接受點擊（純視覺呈現）
            if tp.get('_is_ghost_sliver', False):
                continue
            coords = tp.get('polygon_coords') or []
            if len(coords) < 3:
                continue
            xs_c = [c[0] for c in coords] + [coords[0][0]]
            ys_c = [c[1] for c in coords] + [coords[0][1]]
            fig.add_trace(go.Scatter(
                x=xs_c, y=ys_c, mode='lines',
                fill='toself',
                fillcolor='rgba(0,0,0,0.001)',  # 幾乎透明但可點擊
                line=dict(color='rgba(0,0,0,0)', width=0),
                hoverinfo='text',
                text=f"原地號 {tp.get('原地號', '')}",
                showlegend=False,
                name=f"_clk_{tp.get('原地號', '')}",
            ))

    # ─── 任務二：高亮指定暫編地號（亮紅 + 粗邊框 + 質心標記 + 自動縮放）───
    highlight_xrange = None
    highlight_yrange = None
    if highlight_temp_no:
        target = None
        for tp in temp_parcels:
            tno = str(tp.get('暫編地號', '') or '')
            ono = str(tp.get('原地號', '') or '')
            # 比對暫編地號或原地號；'(無地號)' 特例：尋找原地號為空之筆
            if highlight_temp_no == '(無地號)' and not ono:
                target = tp
                break
            if tno == highlight_temp_no or ono == highlight_temp_no:
                target = tp
                break
        if target is not None:
            coords = target.get('polygon_coords') or []
            if len(coords) >= 3:
                xs_h = [c[0] for c in coords] + [coords[0][0]]
                ys_h = [c[1] for c in coords] + [coords[0][1]]
                fig.add_trace(go.Scatter(
                    x=xs_h, y=ys_h, mode='lines',
                    fill='toself', fillcolor='rgba(255, 30, 30, 0.45)',
                    line=dict(color='#FF1E1E', width=4),
                    name=f"🔴 高亮：{highlight_temp_no}",
                    hoverinfo='text',
                    text=f"原地號 {target.get('原地號', '(無)')} / 暫編 {target.get('暫編地號', '')}",
                    showlegend=True,
                ))
                # 質心標記
                cx = float(target.get('centroid_x', sum(c[0] for c in coords) / len(coords)))
                cy = float(target.get('centroid_y', sum(c[1] for c in coords) / len(coords)))
                fig.add_trace(go.Scatter(
                    x=[cx], y=[cy], mode='markers+text',
                    marker=dict(symbol='star', size=18, color='#FF1E1E',
                                 line=dict(color='#000', width=1.5)),
                    text=[f"📍 {highlight_temp_no}"],
                    textposition='top center',
                    textfont=dict(size=14, color='#FF1E1E', family='Arial Black'),
                    name='高亮質心', hoverinfo='skip', showlegend=False,
                ))
                # 自動縮放：取地號邊界 + 兩倍範圍 padding
                if auto_zoom_to_highlight:
                    xs_only = [c[0] for c in coords]
                    ys_only = [c[1] for c in coords]
                    minx, maxx = min(xs_only), max(xs_only)
                    miny, maxy = min(ys_only), max(ys_only)
                    pad_x = max((maxx - minx) * 1.5, 10.0)
                    pad_y = max((maxy - miny) * 1.5, 10.0)
                    highlight_xrange = [minx - pad_x, maxx + pad_x]
                    highlight_yrange = [miny - pad_y, maxy + pad_y]

    xaxis_cfg = dict(title="TWD97 X", gridcolor='#E5E5E5', zerolinecolor='#CCCCCC',
                     showline=True, linecolor='#888888')
    yaxis_cfg = dict(title="TWD97 Y", gridcolor='#E5E5E5', zerolinecolor='#CCCCCC',
                     showline=True, linecolor='#888888',
                     scaleanchor="x", scaleratio=1)
    if highlight_xrange:
        xaxis_cfg['range'] = highlight_xrange
    if highlight_yrange:
        yaxis_cfg['range'] = highlight_yrange

    # ═══ 🚨 Patch E-0.3：Ghost Sliver 染灰渲染（一致性保證）═══
    # 跨所有圖層保持一致：ghost sliver 永遠灰色 (碎) 標籤、半透明
    _ghost_xs, _ghost_ys = [], []
    _ghost_hover_pts_x, _ghost_hover_pts_y, _ghost_hover_text = [], [], []
    for tp in (temp_parcels or []):
        if not tp.get('_is_ghost_sliver', False):
            continue
        coords = tp.get('polygon_coords') or []
        if len(coords) < 3:
            continue
        xs_g = [c[0] for c in coords] + [coords[0][0], None]
        ys_g = [c[1] for c in coords] + [coords[0][1], None]
        _ghost_xs.extend(xs_g)
        _ghost_ys.extend(ys_g)
        _ghost_hover_pts_x.append(float(tp.get('centroid_x', 0)))
        _ghost_hover_pts_y.append(float(tp.get('centroid_y', 0)))
        _ghost_hover_text.append(
            f"碎屑：{tp.get('原地號', '')}(碎)<br>"
            f"歸屬於 {tp.get('_attached_to', '')}<br>"
            f"面積已合至主體（{tp.get('_ghost_area_m2', 0):.2f} ㎡）"
        )
    if _ghost_xs:
        fig.add_trace(go.Scatter(
            x=_ghost_xs, y=_ghost_ys, mode='lines',
            fill='toself', fillcolor='rgba(184, 184, 184, 0.6)',  # 中性灰半透明
            line=dict(color='#888888', width=0.6, dash='dot'),
            opacity=0.6,
            name='碎屑 (ghost sliver)',
            hoverinfo='skip',
            showlegend=True,
        ))
        fig.add_trace(go.Scatter(
            x=_ghost_hover_pts_x, y=_ghost_hover_pts_y, mode='markers',
            marker=dict(size=4, color='rgba(0,0,0,0.4)'),
            hoverinfo='text', text=_ghost_hover_text,
            name='_ghost_hover', showlegend=False,
        ))

    fig.update_layout(
        height=height,
        paper_bgcolor='white', plot_bgcolor='white',
        xaxis=xaxis_cfg, yaxis=yaxis_cfg,
        margin=dict(l=40, r=40, t=10, b=40),
        showlegend=True,
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=0,
                    bgcolor='rgba(255,255,255,0.9)', bordercolor='#CCCCCC', borderwidth=1),
    )
    return fig


def _normalize_landno_module(s) -> str:
    """
    🆕 Phase 5 升級 + Phase 11 Hotfix：徹底正規化地號字串（Regex 防呆強化版）

    處理項目：
      1. 段碼前綴：≥3 段且第 1 段較短（1~2 位）→ 視為前綴段碼，取後 2 段
         例：'13-0628-0006' → '0628-0006' → '628-6'
      2. 段碼後綴 🆕：3 段且中間段為「0/00/000…」→ 視為 'XXX-0-段碼' → 取第 1 段
         例：'628-0-13' → '628'
      3. 段碼後綴 🆕：3 段且最後段為短碼（1~2 位）且第 1 段較長 → 取前 2 段
         例：'628-44-13' → '628-44'
      4. 前置零去除：使用 Regex `^0+(?!$)` 精準拔除
         - '00628' → '628'、'0000' → '0'、'0628(建)' → '628(建)'
      5. 子號為 0 視為母地號：'628-0' / '628-00' / '628-0000' → '628'
      6. 非數字段保留原狀（如英文字母段碼）

    保證：'13-0628-0006'、'0628-0006'、'628-6' 經正規化後皆為 '628-6'
          '13-0628-0000'、'0628'、'628'、'628-0-13' 皆為 '628'
          '628-44-13'、'628-44' 皆為 '628-44'
    """
    if not s:
        return ''
    t = str(s).strip()
    if not t:
        return ''
    parts = t.split('-')
    # 1~3. 三段時智慧判定段碼方向（前綴 vs 後綴）
    if len(parts) >= 3:
        _last_is_zero_subnum = (len(parts) == 3
                                  and parts[1].isdigit()
                                  and all(c == '0' for c in parts[1]))
        if _last_is_zero_subnum:
            # 'XXX-0-段碼' → 取第 1 段（如 '628-0-13' → '628'）
            parts = parts[:1]
        elif (len(parts) == 3
              and parts[0].isdigit() and parts[-1].isdigit()
              and len(parts[-1]) <= 2 and len(parts[0]) >= 3):
            # '母-子-段碼' → 取前 2 段（如 '628-44-13' → '628-44'）
            parts = parts[:2]
        else:
            # 預設：前綴段碼 → 取後 2 段（如 '13-628-44' → '628-44'）
            parts = parts[-2:]
    # 4. 前置零去除（Regex `^0+(?!$)` — 安全處理髒字串，不依賴 int 轉換）
    import re as _re_norm
    cleaned = []
    for p in parts:
        ps = p.strip()
        if not ps:
            continue
        # 拔除「開頭的 0」，但若全是 0 則保留一個 '0'
        # 例如：'00628' → '628', '0000' → '0', '0628(建)' → '628(建)'
        ps = _re_norm.sub(r'^0+(?!$)', '', ps)
        cleaned.append(ps)
    if not cleaned:
        return ''
    # 3. 子號為 0 視為母地號
    if len(cleaned) == 2 and cleaned[1] == '0':
        cleaned = cleaned[:1]
    return '-'.join(cleaned)


def _resolve_ownership(parcel_no: str, ownership_map: dict) -> str:
    """
    🆕 Phase 5 強化：四層歸戶查找

    Tier 1: 直接 dict.get
    Tier 2: 正規化後查找
    Tier 3: 反查所有 keys 之正規化值
    Tier 4: 暴力子字串比對 — 結尾匹配 + 至少 3 字元防誤配
    """
    if not parcel_no or not ownership_map:
        return ''
    pn = str(parcel_no).strip()
    if not pn:
        return ''

    # Tier 1: 直接查
    gid = ownership_map.get(pn) or ownership_map.get(parcel_no)
    if gid:
        return str(gid)

    # Tier 2: 正規化後查
    norm = _normalize_landno_module(pn)
    if norm and norm in ownership_map:
        return str(ownership_map[norm])

    # Tier 3: 反查所有 keys 之正規化值
    if norm:
        for k, v in ownership_map.items():
            if _normalize_landno_module(k) == norm:
                return str(v)

    # Tier 4 (Phase 5 NEW): 暴力子字串比對
    # 安全策略：
    #   - 至少 3 字元（防 '1' 誤配 '1-1'、'628-1' 等）
    #   - 必須結尾匹配（'628-6' in '13-628-6' ✓；'628' in '628-1' ✗）
    #   - 雙向匹配：DXF 含於 Tab 1，或 Tab 1 含於 DXF
    if norm and len(norm) >= 3:
        for k, v in ownership_map.items():
            kn = _normalize_landno_module(k)
            if not kn or len(kn) < 3:
                continue
            # A 是 B 的「-結尾子串」或完全相等
            if (kn == norm or norm == kn
                or kn.endswith('-' + norm)
                or norm.endswith('-' + kn)):
                return str(v)
    return ''


def _effective_landno(tp, overrides=None) -> str:
    """
    🆕 Phase 6 Task O：取最終 effective land number（含使用者補正 override）

    若 tp['暫編地號'] 在 overrides 中 → 回傳 override 值
    否則回傳 tp['原地號']
    """
    try:
        import streamlit as _st
        if overrides is None:
            overrides = _st.session_state.get('f3_parcel_no_overrides', {}) or {}
    except Exception:
        overrides = overrides or {}
    k = tp.get('暫編地號', '') if isinstance(tp, dict) else ''
    if k in overrides:
        return overrides[k]
    return tp.get('原地號', '') if isinstance(tp, dict) else ''


def _detect_problem_parcels(temp_parcels: list, ownership_map: dict) -> dict:
    """
    🆕 Phase 6 Task O：偵測兩類問題地塊

    回傳：{
        'no_landno': [(temp_parcel, P_idx), ...],
        'no_owner':  [(temp_parcel, original_landno), ...],
    }
    """
    result = {'no_landno': [], 'no_owner': []}
    try:
        import streamlit as _st
        overrides = _st.session_state.get('f3_parcel_no_overrides', {}) or {}
    except Exception:
        overrides = {}
    p_idx = 1
    for tp in (temp_parcels or []):
        # 取最終地號（含 user override）
        orig_no = (overrides.get(tp.get('暫編地號', ''))
                   or tp.get('原地號', '')
                   or '').strip()
        if not orig_no or orig_no == '(無地號)':
            result['no_landno'].append((tp, p_idx))
            p_idx += 1
            continue
        # 是否歸戶得到？
        gid = _resolve_ownership(orig_no, ownership_map)
        if not gid:
            result['no_owner'].append((tp, orig_no))
    return result


def _ownership_color(gid: str) -> str:
    """🆕 Phase 4 G-4：依歸戶群組 ID hash 產生穩定顏色（同一 gid 永遠同色）"""
    import hashlib
    if not gid:
        return '#BDBDBD'
    h = int(hashlib.md5(str(gid).encode()).hexdigest()[:6], 16)
    palette = ['#1F77B4', '#FF7F0E', '#2CA02C', '#D62728', '#9467BD',
               '#8C564B', '#E377C2', '#17BECF', '#BCBD22', '#5DADE2',
               '#48C9B0', '#F4D03F', '#A569BD', '#52BE80', '#EB984E']
    return palette[h % len(palette)]


def _render_f3_unified_map(classified_blocks, temp_parcels, zones_map=None,
                           ownership_map=None, g_values=None,
                           manual_baselines=None, road_centerlines=None,
                           manual_road_splits=None,
                           manual_road_centerlines=None,
                           params_map=None,
                           visible_layers=None,
                           pick_mode_state=None,
                           road_split_state=None,
                           road_edge_pick=None,
                           action_mode=None,
                           height: int = 620):
    """回傳 (fig, click_target_offset, valid_parcel_list,
              edge_offset, edge_meta, redge_offset, redge_meta)
    - 新增 edge_meta：Task M 街廓邊互動（橙色）— `[(blk_lbl, edge_idx, p1, p2, length), ...]`
    - 新增 redge_meta：Task N 道路邊互動（亮黃色）— `[(road_lbl, edge_idx, p1, p2, length, road_blk), ...]`
    - 新增 manual_road_centerlines：Task N 已生成精準中心線
    - action_mode：用於自動啟用 edge_interactive / road_edges 圖層"""
    """
    🆕 Phase 4 Task G：統一 GIS 互動地圖渲染函式

    單一 Plotly 圖表整合 8 大圖層 + 點擊靶 trace，由 visible_layers dict 控制可見性。

    參數：
      classified_blocks    街廓清單（含 vertices, label, category, area_m2, ...）
      temp_parcels         暫編地號清單（含 polygon_coords, 暫編地號, 原地號, 所屬街廓）
      zones_map            {原地號: 地價區段代碼}，供 price_zones 圖層著色
      ownership_map        {原地號: 歸戶群組 ID}，供 ownership_dist 圖層著色
      g_values             G 值計算結果 list（含 cut_coords, 推進側別, 街角地）→ trial_allocation 圖層
      manual_baselines     {街廓 label: {'point': (x, y), 'angle_deg': float, 'enabled': bool}}
      road_centerlines     {道路 label: list[LineString]}（可選；目前留作未來擴充）
      manual_road_splits   {道路 label: [Polygon, Polygon, ...]}（手動切割道路結果）
      params_map           {暫編地號: {'is_corner': bool, 'side': str, 'W': float}} → corner_status 圖層
      visible_layers       {layer_name: bool} 控制各圖層渲染
      pick_mode_state      {'block': str, 'clicks': [(x,y), ...]} 基準線拾取狀態
      road_split_state     {'road': str, 'clicks': [(x,y), ...]} 道路切割拾取狀態
      height               地圖高度

    回傳：plotly Figure
    """
    import plotly.graph_objects as go
    import math as _math_um

    # 預設圖層可見性
    _vl = dict(visible_layers or {})
    _show_blocks  = _vl.get('block_boundaries', True)
    _show_cad     = _vl.get('pre_cadastral', True)
    _show_zones   = _vl.get('price_zones', True)
    _show_owners  = _vl.get('ownership_dist', False)
    _show_corners = _vl.get('corner_status', True)
    _show_baseline= _vl.get('manual_baseline', True)
    _show_trial   = _vl.get('trial_allocation', False)
    _show_centerl = _vl.get('road_centerlines', False)
    # 🆕 Phase 6 新圖層
    _show_problems = _vl.get('problem_parcels', False)
    # 🆕 V12 模組 4-B：重劃後土地歸戶分佈
    _show_alloc_owners = _vl.get('ownership_after_alloc', False)
    # action_mode 自動啟用 edge_interactive (M) 或 road_edges (N)
    _is_baseline_mode = bool(action_mode and '設定分配基準線' in str(action_mode))
    _is_road_edge_mode = bool(action_mode and '生成精準中心線' in str(action_mode))
    _show_edge_interactive = _is_baseline_mode or _vl.get('edge_interactive', False)
    _show_road_edges = _is_road_edge_mode or _vl.get('road_edges', False)
    # 🆕 Phase 7 Module 3 (M3-1)：圖層獨佔 Hover
    # 進入「📐 設定分配基準線」或「🛣️ 生成精準中心線」模式時，
    # 其他非必要圖層改為 hoverinfo='skip'，僅讓目標邊界 trace 可被 hover
    _exclusive_hover_mode = _is_baseline_mode or _is_road_edge_mode
    # 各圖層 hover 模式（exclusive 時改為 skip）
    _hover_default = 'skip' if _exclusive_hover_mode else 'text'

    fig = go.Figure()
    zones_map     = zones_map or {}
    ownership_map = ownership_map or {}
    params_map    = params_map or {}
    g_values      = g_values or []
    manual_baselines    = manual_baselines or {}
    manual_road_splits  = manual_road_splits or {}
    manual_road_centerlines = manual_road_centerlines or {}
    # 點擊解析用的元資料
    edge_meta = []        # Task M
    redge_meta = []       # Task N
    edge_offset = 0
    redge_offset = 0

    # ═══ 圖層 1：街廓填色 + 邊框 + 紅色街廓分配線 ═══
    if _show_blocks:
        for b in (classified_blocks or []):
            verts = b.get('vertices') or []
            if len(verts) < 3:
                continue
            color = F3_CATEGORY_COLORS.get(b.get('category', ''), "#BDBDBD")
            xs = [v[0] for v in verts] + [verts[0][0]]
            ys = [v[1] for v in verts] + [verts[0][1]]
            fig.add_trace(go.Scatter(
                x=xs, y=ys, mode='lines', fill='toself', fillcolor=color,
                line=dict(color=color, width=0.5), opacity=0.20,
                hoverinfo='skip', showlegend=False,
            ))
        # 街廓 label（合一 trace）
        if classified_blocks:
            _vb = [b for b in classified_blocks if b.get('centroid')]
            if _vb:
                fig.add_trace(go.Scatter(
                    x=[b['centroid'][0] for b in _vb],
                    y=[b['centroid'][1] for b in _vb],
                    mode='text',
                    text=[f"<b>{b.get('label', '')}</b>" for b in _vb],
                    textfont=dict(size=11, color='#444'),
                    hoverinfo='skip', showlegend=False,
                ))
        # 紅色街廓分配線（合一 trace）
        rxs, rys = [], []
        for b in (classified_blocks or []):
            v = b.get('vertices') or []
            if len(v) < 3:
                continue
            for x, y in v:
                rxs.append(x); rys.append(y)
            rxs.append(v[0][0]); rys.append(v[0][1])
            rxs.append(None); rys.append(None)
        if rxs:
            fig.add_trace(go.Scatter(
                x=rxs, y=rys, mode='lines',
                line=dict(color='#D62728', width=2.0),
                name='街廓分配線', hoverinfo='skip',
            ))

    # ═══ 圖層 2：重劃前地籍線（按區段分群著色） ═══
    if _show_cad:
        palette = ['#1F77B4', '#FF7F0E', '#2CA02C', '#D62728', '#9467BD',
                   '#8C564B', '#E377C2', '#17BECF', '#BCBD22']
        uniq_zones = sorted({v for v in zones_map.values() if v})
        zone_color_map = {z: palette[i % len(palette)] for i, z in enumerate(uniq_zones)}
        groups = {'(未標註)': {'xs': [], 'ys': [], 'color': '#666666', 'width': 0.7}}
        for z in uniq_zones:
            groups[z] = {'xs': [], 'ys': [], 'color': zone_color_map[z], 'width': 2.5}
        for tp in (temp_parcels or []):
            # 🚨 Patch E-0.3：ghost sliver 由 consolidated 渲染處染灰
            if tp.get('_is_ghost_sliver', False):
                continue
            coords = tp.get('polygon_coords') or []
            if not coords:
                continue
            zname = zones_map.get(tp.get('原地號', ''), '') or '(未標註)'
            g = groups.setdefault(zname, {'xs': [], 'ys': [], 'color': '#666666', 'width': 0.7})
            for x, y in coords:
                g['xs'].append(x); g['ys'].append(y)
            g['xs'].append(None); g['ys'].append(None)
        for zname, g in groups.items():
            if not g['xs']:
                continue
            # 若關閉「地價區段著色」→ 已標區段回退至灰色細線
            if zname != '(未標註)' and not _show_zones:
                _lc = '#888888'; _lw = 0.7
            else:
                _lc = g['color']; _lw = g['width']
            fig.add_trace(go.Scatter(
                x=g['xs'], y=g['ys'], mode='lines',
                line=dict(color=_lc, width=_lw),
                name=f"區段 {zname}", hoverinfo='skip',
                showlegend=(zname != '(未標註)' and _show_zones),
            ))

    # ═══ 圖層 3：歸戶分佈（fill colored by ownership group）═══
    # 🆕 Phase 4 修正：使用 _resolve_ownership 含正規化查找，解決 Tab1↔Tab4 地號格式差異
    # 🆕 Phase 6 Task O：使用 _effective_landno 含使用者補正 override
    if _show_owners:
        own_groups = {}
        for tp in (temp_parcels or []):
            # 🚨 Patch E-0.3：ghost sliver 不參與歸戶分佈染色
            if tp.get('_is_ghost_sliver', False):
                continue
            gid = _resolve_ownership(_effective_landno(tp), ownership_map)
            if not gid:
                continue
            coords = tp.get('polygon_coords') or []
            if len(coords) < 3:
                continue
            g = own_groups.setdefault(gid, {'xs': [], 'ys': [], 'color': _ownership_color(gid)})
            for x, y in coords:
                g['xs'].append(x); g['ys'].append(y)
            g['xs'].append(coords[0][0]); g['ys'].append(coords[0][1])
            g['xs'].append(None); g['ys'].append(None)
        for gid, g in own_groups.items():
            fig.add_trace(go.Scatter(
                x=g['xs'], y=g['ys'], mode='lines', fill='toself',
                fillcolor=g['color'],
                line=dict(color=g['color'], width=1.2),
                opacity=0.35,
                name=f"歸戶 {gid}",
                # 🆕 Phase 7 M3-1：exclusive hover 模式下不接受 hover
                hoverinfo=_hover_default, text=f"歸戶群組：{gid}",
                showlegend=True,
            ))

    # ═══ 圖層 4：街角狀態（紅色填充已標街角地） ═══
    if _show_corners and params_map:
        for tp in (temp_parcels or []):
            # 🚨 Patch E-0.3：ghost sliver 不會是街角地
            if tp.get('_is_ghost_sliver', False):
                continue
            k = tp.get('暫編地號', '')
            p_meta = params_map.get(k, {})
            if not p_meta.get('is_corner'):
                continue
            coords = tp.get('polygon_coords') or []
            if len(coords) < 3:
                continue
            xs_c = [c[0] for c in coords] + [coords[0][0]]
            ys_c = [c[1] for c in coords] + [coords[0][1]]
            _side = p_meta.get('side', '無')
            # 左：粉紅；右：橙紅
            _fill = ('rgba(244, 114, 182, 0.45)' if _side == '左側'
                     else 'rgba(239, 68, 68, 0.45)' if _side == '右側'
                     else 'rgba(214, 39, 40, 0.35)')
            fig.add_trace(go.Scatter(
                x=xs_c, y=ys_c, mode='lines', fill='toself',
                fillcolor=_fill,
                line=dict(color='#B91C1C', width=1.8),
                # 🆕 Phase 7 M3-1
                hoverinfo=_hover_default, text=f"⭐ {k}（{_side}街角）",
                showlegend=False,
            ))

    # ═══ 圖層 5：手動分配基準線（紅色虛線箭頭） ═══
    if _show_baseline and manual_baselines:
        for blk_lbl, bv in manual_baselines.items():
            if not bv.get('enabled') or not bv.get('point'):
                continue
            try:
                bx, by = bv['point']
                ang = _math_um.radians(float(bv.get('angle_deg', 0.0)))
                _alen = 30.0
                ex = bx + _alen * _math_um.cos(ang)
                ey = by + _alen * _math_um.sin(ang)
                fig.add_trace(go.Scatter(
                    x=[bx, ex], y=[by, ey], mode='lines+markers',
                    line=dict(color='#D62728', width=2, dash='dash'),
                    marker=dict(symbol='circle-open', size=10, color='#D62728'),
                    hoverinfo='text',
                    text=[f"{blk_lbl} 基準起點", f"{blk_lbl} 推進方向 ({bv.get('angle_deg', 0):.1f}°)"],
                    name=f"{blk_lbl} 基準線", showlegend=True,
                ))
            except Exception:
                continue

    # ═══ 圖層 6：試分配地籍圖（cut_coords 三色：綠/橘/黃） ═══
    if _show_trial and g_values:
        for _row in g_values:
            _coords = _row.get('cut_coords') or []
            if len(_coords) < 3:
                continue
            _xs = [c[0] for c in _coords] + [_coords[0][0]]
            _ys = [c[1] for c in _coords] + [_coords[0][1]]
            _is_offset = (_row.get('推進側別') == '抵費地')
            _is_corner = (_row.get('街角地') == '是')
            if _is_offset:
                _fill = 'rgba(255, 200, 0, 0.50)'; _line = '#E8A317'
            elif _is_corner:
                _fill = 'rgba(255,165,0,0.50)';    _line = '#FF8C00'
            else:
                _fill = 'rgba(50,180,80,0.40)';    _line = '#2E7D32'
            fig.add_trace(go.Scatter(
                x=_xs, y=_ys, mode='lines', fill='toself',
                fillcolor=_fill, line=dict(color=_line, width=1.2),
                hoverinfo='text',
                text=(f"抵費地 {_row.get('暫編地號','')}<br>面積 {_row.get('幾何面積(㎡)',0):.2f} ㎡"
                      if _is_offset else
                      f"{_row.get('暫編地號','')}<br>S={_row.get('S(m)',0):.2f}m G={_row.get('G(㎡)',0):.2f}㎡"),
                showlegend=False,
            ))

    # ═══ 圖層 6.5：🆕 V12 模組 4-B 重劃後土地歸戶分佈（含 🚨 補強 D 白邊框）═══
    if _show_alloc_owners and g_values:
        import hashlib as _hl_v12
        from shapely.geometry import Polygon as _SP_alloc

        def _gid_to_color(gid: str, alpha: float = 0.55) -> str:
            """穩定 hash → HSL 色相（同 gid 永遠同色）"""
            if not gid:
                return f"hsla(0, 0%, 60%, {alpha})"
            h = int(_hl_v12.md5(str(gid).encode()).hexdigest()[:6], 16) % 360
            return f"hsla({h}, 70%, 55%, {alpha})"

        def _safe_centroid_x(coords):
            try:
                p = _SP_alloc(coords)
                if not p.is_valid:
                    p = p.buffer(0)
                return float(p.centroid.x)
            except Exception:
                return 0.0

        def _safe_area(coords):
            try:
                p = _SP_alloc(coords)
                if not p.is_valid:
                    p = p.buffer(0)
                return float(p.area)
            except Exception:
                return 0.0

        # 排序：依 cut polygon centroid_x 升序
        rows_with_geom = []
        for r in g_values:
            c = r.get('cut_coords') or []
            if len(c) >= 3 and r.get('推進側別') != '抵費地':
                rows_with_geom.append((_safe_centroid_x(c), r))
        rows_with_geom.sort(key=lambda x: x[0])

        # 統計幾何錯誤（顏色出現於非可建築街廓）
        _BUILDABLE = {'住宅區', '商業區'}
        geometry_error_count = 0
        # 建立 block_label → category 對照
        cat_by_label = {b.get('label', ''): b.get('category', '')
                         for b in (classified_blocks or [])}

        for _, row in rows_with_geom:
            coords = row['cut_coords']
            xs = [c[0] for c in coords] + [coords[0][0]]
            ys = [c[1] for c in coords] + [coords[0][1]]
            orig = row.get('原地號', '')
            gid = (ownership_map.get(orig, '') if ownership_map
                    else '') or '(未歸戶)'
            color = _gid_to_color(gid)
            blk_lbl = row.get('所屬街廓', '')
            blk_cat = cat_by_label.get(blk_lbl, '')
            is_geometry_error = blk_cat not in _BUILDABLE
            if is_geometry_error:
                geometry_error_count += 1

            actual_area = _safe_area(coords)
            hover_text = (
                f"歸戶 <b>{gid}</b><br>"
                f"原地號 {orig}<br>"
                f"街廓 {blk_lbl}（{blk_cat}）<br>"
                f"G = {row.get('G(㎡)', 0):.2f} ㎡<br>"
                f"實配 = {actual_area:.2f} ㎡"
                + ('<br><span style="color:red;">⚠️ 配地溢出非可建築街廓</span>'
                   if is_geometry_error else '')
            )
            fig.add_trace(go.Scatter(
                x=xs, y=ys, mode='lines',
                fill='toself', fillcolor=color,
                # 🚨 補強 D：白色邊框 1.5px 防疊加遮蔽
                line=dict(color='white', width=1.5),
                name=f"歸戶 {gid}",
                hoverinfo='text', text=hover_text,
                showlegend=False,
            ))

        # 幾何錯誤警示：藏入 figure 之 layout.annotations（不影響圖形佈局）
        if geometry_error_count > 0:
            fig.add_annotation(
                xref='paper', yref='paper', x=0.01, y=0.99,
                showarrow=False,
                text=(f"⚠️ <b>{geometry_error_count} 筆配地溢出非可建築街廓</b>"
                      f"<br>請檢視 d_hat / allocation_dir / 公設前置整併"),
                bgcolor='rgba(255, 240, 240, 0.95)',
                bordercolor='#D62728', borderwidth=2,
                font=dict(size=11, color='#D62728'),
                align='left',
            )

    # ═══ 圖層 7：道路中心線（Phase 6 Task N：精準平移生成版 + 舊手動切割相容）═══
    if _show_centerl:
        # 7a. 已生成精準中心線（Task N）— 綠色實線
        for road_lbl, pts in (manual_road_centerlines or {}).items():
            if len(pts) < 2:
                continue
            fig.add_trace(go.Scatter(
                x=[p[0] for p in pts], y=[p[1] for p in pts],
                mode='lines+markers',
                line=dict(color='#27AE60', width=3),
                marker=dict(symbol='circle', size=6, color='#27AE60'),
                hoverinfo='text',
                text=[f"道路 {road_lbl} 中心線（精準平移生成）" for _ in pts],
                name=f"中心線 {road_lbl}", showlegend=True,
            ))
        # 7b. 舊手動切割 pieces（向後相容）
        for road_lbl, pieces in (manual_road_splits or {}).items():
            for i, pp in enumerate(pieces):
                try:
                    coords = list(pp.exterior.coords)
                except Exception:
                    continue
                if len(coords) < 3:
                    continue
                xs_p = [c[0] for c in coords] + [coords[0][0]]
                ys_p = [c[1] for c in coords] + [coords[0][1]]
                fig.add_trace(go.Scatter(
                    x=xs_p, y=ys_p, mode='lines', fill='toself',
                    fillcolor=('rgba(100,200,200,0.20)' if i == 0 else 'rgba(200,100,200,0.20)'),
                    line=dict(color='#0088CC', width=1.5, dash='dot'),
                    hoverinfo='text', text=f"[舊版] 道路 {road_lbl} 手動切割 #{i+1}",
                    showlegend=False,
                ))

    # ═══ 圖層 8：拾取狀態標記（含舊基準線/道路切割相容）═══
    for state, color, label in [
        (pick_mode_state, '#D62728', '基準線'),
        (road_split_state, '#1E90FF', '道路切割'),
    ]:
        if state and state.get('clicks'):
            _picked_pts = state['clicks']
            fig.add_trace(go.Scatter(
                x=[p[0] for p in _picked_pts],
                y=[p[1] for p in _picked_pts],
                mode='markers+text',
                marker=dict(symbol='cross', size=14, color=color,
                            line=dict(color='#FFFFFF', width=2)),
                text=[f"{i+1}" for i in range(len(_picked_pts))],
                textposition='top center',
                textfont=dict(size=14, color=color),
                hoverinfo='skip', showlegend=False,
            ))

    # ═══ 圖層 8b：道路邊鎖定高亮（Task N：road_edge_pick）═══
    if road_edge_pick and road_edge_pick.get('p1') and road_edge_pick.get('p2'):
        _rp1 = road_edge_pick['p1']
        _rp2 = road_edge_pick['p2']
        fig.add_trace(go.Scatter(
            x=[_rp1[0], _rp2[0]], y=[_rp1[1], _rp2[1]],
            mode='lines+markers',
            line=dict(color='#FF1493', width=6),    # 螢光粉色加粗
            marker=dict(symbol='diamond', size=12, color='#FF1493'),
            hoverinfo='text',
            text=f"🔒 已鎖定 {road_edge_pick.get('road_label','')} 第 {road_edge_pick.get('edge_idx',0)+1} 條邊",
            showlegend=False, opacity=0.85,
        ))

    # ═══ 圖層 8c：問題地塊（Phase 6 Task O：紅/黃斜線標記）═══
    if _show_problems:
        try:
            problems = _detect_problem_parcels(temp_parcels, ownership_map)
        except Exception:
            problems = {'no_landno': [], 'no_owner': []}
        # A. 無地號（紅色斜線）
        for tp, p_idx in problems['no_landno']:
            coords = tp.get('polygon_coords') or []
            if len(coords) < 3:
                continue
            xs = [c[0] for c in coords] + [coords[0][0]]
            ys = [c[1] for c in coords] + [coords[0][1]]
            fig.add_trace(go.Scatter(
                x=xs, y=ys, mode='lines', fill='toself',
                fillcolor='rgba(220,38,38,0.30)',
                line=dict(color='#DC2626', width=1.8, dash='dot'),
                hoverinfo='text',
                text=f"⚠️ <b>無地號 P{p_idx}</b><br>暫編 {tp.get('暫編地號','')}<br>面積 {tp.get('分攤登記面積_m2', tp.get('面積_m2', 0)):.2f} ㎡<br><i>(點擊以「✏️ 補正地號」模式補正)</i>",
                showlegend=False,
            ))
            # 標籤 P1, P2...
            try:
                from shapely.geometry import Polygon as _SP_lbl
                _cen = _SP_lbl(coords).centroid
                fig.add_annotation(
                    x=_cen.x, y=_cen.y, text=f"<b>P{p_idx}</b>",
                    showarrow=False, font=dict(size=12, color='#DC2626'),
                    bgcolor='rgba(255,255,255,0.85)', borderpad=2,
                )
            except Exception:
                pass
        # B. 未歸戶（黃色斜線）
        for tp, orig_no in problems['no_owner']:
            coords = tp.get('polygon_coords') or []
            if len(coords) < 3:
                continue
            xs = [c[0] for c in coords] + [coords[0][0]]
            ys = [c[1] for c in coords] + [coords[0][1]]
            fig.add_trace(go.Scatter(
                x=xs, y=ys, mode='lines', fill='toself',
                fillcolor='rgba(245,158,11,0.30)',
                line=dict(color='#F59E0B', width=1.5, dash='dash'),
                hoverinfo='text',
                text=f"⚠️ <b>未歸戶</b><br>原地號 {orig_no}<br>暫編 {tp.get('暫編地號','')}<br>面積 {tp.get('分攤登記面積_m2', tp.get('面積_m2', 0)):.2f} ㎡<br><i>(土地歸戶 查無此地號)</i>",
                showlegend=False,
            ))

    # ═══ 圖層 8d：街廓邊互動（Task M）— 橙色加粗，hover 高亮 ═══
    if _show_edge_interactive:
        edge_offset = len(fig.data)
        for b in (classified_blocks or []):
            verts = b.get('vertices') or []
            if len(verts) < 3:
                continue
            blk_lbl = b.get('label', '')
            for i in range(len(verts)):
                p1 = verts[i]
                p2 = verts[(i + 1) % len(verts)]
                length = ((p2[0]-p1[0])**2 + (p2[1]-p1[1])**2) ** 0.5
                fig.add_trace(go.Scatter(
                    x=[p1[0], p2[0]], y=[p1[1], p2[1]],
                    mode='lines',
                    line=dict(color='#FF8C00', width=4),
                    hoverinfo='text',
                    text=f"📐 街廓 {blk_lbl} - 第 {i+1} 條邊<br>長度 {length:.2f} m<br><b>(點擊鎖定為基準線)</b>",
                    showlegend=False, opacity=0.7,
                    name=f"_edge_{blk_lbl}_{i}",
                ))
                edge_meta.append((blk_lbl, i, p1, p2, length))

    # ═══ 圖層 8e：道路邊互動（Task N）— 亮黃色加粗，限定道路類街廓 ═══
    if _show_road_edges:
        redge_offset = len(fig.data)
        for b in (classified_blocks or []):
            if not _is_road_like_category(b.get('category', '')):
                continue
            verts = b.get('vertices') or []
            if len(verts) < 3:
                continue
            blk_lbl = b.get('label', '')
            for i in range(len(verts)):
                p1 = verts[i]
                p2 = verts[(i + 1) % len(verts)]
                length = ((p2[0]-p1[0])**2 + (p2[1]-p1[1])**2) ** 0.5
                fig.add_trace(go.Scatter(
                    x=[p1[0], p2[0]], y=[p1[1], p2[1]],
                    mode='lines',
                    line=dict(color='#FFD700', width=5),
                    hoverinfo='text',
                    text=f"🛣️ 道路 {blk_lbl} - 第 {i+1} 條邊<br>長度 {length:.2f} m<br><b>(點擊以平移生成中心線)</b>",
                    showlegend=False, opacity=0.85,
                    name=f"_redge_{blk_lbl}_{i}",
                ))
                redge_meta.append((blk_lbl, i, p1, p2, length, b))

    # ═══ 圖層 9：點擊靶 trace（最後渲染 = z-top，被 plotly_events 點擊判定）═══
    # 關鍵：每個 parcel 一個獨立 trace → curveNumber 直接對應 parcel 索引
    # （與 Step D 原版相同的「透明 fill 但可點擊」技術）
    click_target_offset = len(fig.data)
    valid_clk_parcels = []
    for tp in (temp_parcels or []):
        # 🚨 Patch E-0.3：ghost sliver 不參與點擊靶（純視覺呈現）
        if tp.get('_is_ghost_sliver', False):
            continue
        coords = tp.get('polygon_coords') or []
        if len(coords) < 3:
            continue
        xs_t = [c[0] for c in coords] + [coords[0][0]]
        ys_t = [c[1] for c in coords] + [coords[0][1]]
        fig.add_trace(go.Scatter(
            x=xs_t, y=ys_t, mode='lines', fill='toself',
            fillcolor='rgba(0,0,0,0.001)',         # 幾乎透明
            line=dict(color='rgba(0,0,0,0)', width=0),
            hoverinfo='text',
            text=f"暫編 {tp.get('暫編地號','')}<br>原地號 {tp.get('原地號','')}<br>面積 {tp.get('分攤登記面積_m2', tp.get('面積_m2', 0)):.2f} ㎡",
            showlegend=False, name=f"_clk_{tp.get('暫編地號','')}",
        ))
        valid_clk_parcels.append(tp)

    # ═══ 🚨 Patch E-0.3：Ghost Sliver 染灰渲染（一致性保證）═══
    _ghost_xs_u, _ghost_ys_u = [], []
    _ghost_pts_x_u, _ghost_pts_y_u, _ghost_text_u = [], [], []
    for tp in (temp_parcels or []):
        if not tp.get('_is_ghost_sliver', False):
            continue
        coords = tp.get('polygon_coords') or []
        if len(coords) < 3:
            continue
        _ghost_xs_u.extend([c[0] for c in coords] + [coords[0][0], None])
        _ghost_ys_u.extend([c[1] for c in coords] + [coords[0][1], None])
        _ghost_pts_x_u.append(float(tp.get('centroid_x', 0)))
        _ghost_pts_y_u.append(float(tp.get('centroid_y', 0)))
        _ghost_text_u.append(
            f"碎屑：{tp.get('原地號', '')}(碎)<br>"
            f"歸屬於 {tp.get('_attached_to', '')}<br>"
            f"面積已合至主體（{tp.get('_ghost_area_m2', 0):.2f} ㎡）"
        )
    if _ghost_xs_u:
        fig.add_trace(go.Scatter(
            x=_ghost_xs_u, y=_ghost_ys_u, mode='lines',
            fill='toself', fillcolor='rgba(184, 184, 184, 0.6)',
            line=dict(color='#888888', width=0.6, dash='dot'),
            opacity=0.6,
            name='碎屑 (ghost sliver)',
            hoverinfo='skip',
            showlegend=True,
        ))
        fig.add_trace(go.Scatter(
            x=_ghost_pts_x_u, y=_ghost_pts_y_u, mode='markers',
            marker=dict(size=4, color='rgba(0,0,0,0.4)'),
            hoverinfo='text', text=_ghost_text_u,
            name='_ghost_hover_unified', showlegend=False,
        ))

    fig.update_layout(
        height=height,
        paper_bgcolor='white', plot_bgcolor='white',
        xaxis=dict(title="TWD97 X", gridcolor='#E5E5E5'),
        yaxis=dict(title="TWD97 Y", gridcolor='#E5E5E5',
                   scaleanchor='x', scaleratio=1),
        margin=dict(l=40, r=40, t=10, b=40),
        showlegend=True,
        legend=dict(orientation="h", yanchor="bottom", y=1.01, xanchor="left", x=0,
                    bgcolor='rgba(255,255,255,0.85)', bordercolor='#CCCCCC', borderwidth=1),
    )
    return (fig, click_target_offset, valid_clk_parcels,
            edge_offset, edge_meta, redge_offset, redge_meta)


# 🚨 Phase 9.4：_filter_lines_overlapping_blocks 函式已徹底移除
# 法定鐵律：地籍線本就應與街廓邊重合（街廓邊上之地號其外緣即街廓邊）
# 用「距離」刪除地籍線會誤殺合法外緣 → 628-20 等地塊「外牆失蹤」
# 拓樸純淨度 100% 依賴圖層黑名單 _is_excluded_from_cadastral


def reconstruct_parcel_polygons(boundary_lines: list, opar_items: list) -> list:
    """
    由 UA0013.DXF 的 ROAD LINE（界址線）重建重劃前地籍多邊形，
    並以 OPAR TEXT（地號標籤）做配對。

    🆕 Phase 11 Hotfix（質心擴散 + 1:1 貪婪式配對）：
      使用者實務作圖時刻意將地號文字字串中段擺於宗地中心附近，
      故改採「以宗地質心為起點，向外擴散搜尋最近未配對 OPAR」之策略。
      演算法：
        1. 收集所有 (poly_idx, opar_idx, distance) 候選對
        2. 依距離排序遞增
        3. 貪婪式：依序指派 → 若該 poly 與 opar 皆未綁定 → 配對
        4. 此法保證每個 poly 獲得「對它而言最近且尚未被別人占用」的 OPAR
      備援：剩餘未配對之 poly → 標記為空地號（('(無地號)')）
      （不再強制要求 5m 內或 1m buffer；以「對候選最近且 1:1 」為準）

    回傳：[{'parcel','polygon_coords','area_m2','centroid_x','centroid_y'}, ...]
    """
    from shapely.geometry import LineString, Point
    from shapely.ops import polygonize, unary_union

    lines = []
    for s, e in boundary_lines:
        if s == e:
            continue
        try:
            lines.append(LineString([s, e]))
        except Exception:
            continue
    if not lines:
        return []
    try:
        merged = unary_union(lines)
        # 🚨 Phase 9.12 Issue 1：set_precision snap 強制閉合 CAD 浮點縫隙（grid_size=1cm）
        # 解決：CAD 地籍線端點微小不重合 → polygonize 無法閉合 → 出現白色破碎
        # 順序：先 snap → polygonize → 後續 1m² 碎屑過濾（保留 Phase 9.6 機制）
        try:
            from shapely import set_precision as _set_precision
            merged = _set_precision(merged, grid_size=0.01)
        except Exception:
            # shapely < 2.0 fallback：用 buffer-snap
            try:
                merged = merged.buffer(0.005).buffer(-0.005)
            except Exception:
                pass
        polys = list(polygonize(merged))
    except Exception:
        return []

    # 過濾面積 < 1 ㎡ 的碎片
    valid_polys = []
    for poly in polys:
        try:
            if poly.area >= 1.0:
                valid_polys.append(poly)
        except Exception:
            continue
    if not valid_polys:
        return []

    # 🆕 Phase 11 Hotfix / 🔧 Phase 9.3 Hotfix：只保留「葉節點宗地」
    # polygonize() 由界址線拓樸重建時，可能產生「複合多邊形」（內含其他子宗地）
    # 例如重劃外框 / 街廓母界。這些複合多邊形不對應任何地號註記，應排除。
    #
    # 🔧 Phase 9.3：原邏輯過於嚴格 — 只要包含任一較小 poly 的質心即視為非葉
    #    導致相鄰同尺寸 parcel（如 628-20 與 628-21）若質心因浮點誤差落於對方邊內
    #    就被互相誤殺，圖面出現空白破碎。
    #
    # 修正三重條件：
    #   1. 顯著大小差異：other.area < self.area × 0.5（從「< self」放寬至「< 50%」）
    #   2. 忽略雜訊 polygon：other.area < 5㎡ 不參與判定
    #   3. 質心需「深入內部」：距 self.boundary > 0.5m（防邊界 false positive）
    poly_centroids_pre = []
    for poly in valid_polys:
        try:
            c = poly.centroid
            poly_centroids_pre.append((c.x, c.y, poly.area))
        except Exception:
            poly_centroids_pre.append((0.0, 0.0, 0.0))

    leaf_polys = []
    _filtered_as_composite = 0
    for pi, poly in enumerate(valid_polys):
        contains_other = False
        self_area = poly_centroids_pre[pi][2]
        for pj, (cxj, cyj, aj) in enumerate(poly_centroids_pre):
            if pj == pi:
                continue
            # 🔧 條件 1：only consider「顯著小於自己」之 poly（< 50% area）
            if aj >= self_area * 0.5:
                continue
            # 🔧 條件 2：忽略雜訊小 polygon（< 5㎡，可能是浮點碎片）
            if aj < 5.0:
                continue
            try:
                _other_pt = Point(cxj, cyj)
                if not poly.contains(_other_pt):
                    continue
                # 🔧 條件 3：要求質心距 boundary > 0.5m（防止邊界 false positive 誤殺）
                if poly.boundary.distance(_other_pt) <= 0.5:
                    continue
                contains_other = True
                break
            except Exception:
                continue
        if not contains_other:
            leaf_polys.append(poly)
        else:
            _filtered_as_composite += 1

    # 若葉節點過濾把所有都濾掉（罕見），退回使用全部
    if not leaf_polys:
        leaf_polys = valid_polys
    valid_polys = leaf_polys

    # 🔧 Phase 9.3：寫入診斷
    try:
        import streamlit as _st_diag93
        if _filtered_as_composite > 0:
            _st_diag93.session_state['f3_polygons_filtered_as_composite'] = _filtered_as_composite
    except Exception:
        pass

    # 預計算所有 poly 之質心
    poly_centroids = []
    for poly in valid_polys:
        try:
            c = poly.centroid
            poly_centroids.append((c.x, c.y))
        except Exception:
            poly_centroids.append((0.0, 0.0))

    # 🆕 Phase 11 Hotfix：改用「文字點到多邊形邊界距離」而非「文字點到質心距離」
    # 解決小宗地內文字塞不下、被擺在外面 → 質心距離匹配會誤配到鄰近大宗地的問題
    #
    # 🚨 Phase 9.6 強化：碎屑面積門檻 + 最小面積優先排序
    # 當 OPAR 文字點同時落於多個 polygon 內（如「複合大 polygon」+「精準小 polygon」），
    # 應優先選「面積最小且 ≥ 1.0 ㎡ 之合法 polygon」(最精準的外框)
    #
    # 評分規則：
    #   - text 在 poly 內 + area ≥ 1㎡    → cost = -1（最高優先，必中）
    #   - text 在 poly 內 + area < 1㎡    → cost = +999（碎屑，最低優先，應跳過）
    #   - text 在 poly 邊內 0.1m + ≥ 1㎡  → cost = -0.5（次高，貼邊）
    #   - text 在 poly 外                 → cost = poly.distance(pt)²（boundary 距離平方）
    # Tiebreaker（同 cost 時）：
    #   - 同 cost → 較小面積優先（精準外框）
    _SLIVER_AREA_THRESHOLD = 1.0   # 1.0 m² 碎屑門檻
    candidates = []   # list of (cost, area, pi, oi) — 4-tuple，第 2 鍵為面積供 tiebreaker
    for pi, (cx, cy) in enumerate(poly_centroids):
        poly = valid_polys[pi]
        try:
            poly_area = float(poly.area)
        except Exception:
            poly_area = 0.0
        # 🚨 Phase 9.4：預先計算 poly.buffer(0.1)（10cm 容差）供「邊界外微小位移」配對
        try:
            poly_bf = poly.buffer(0.1)
        except Exception:
            poly_bf = poly
        for oi, o in enumerate(opar_items):
            try:
                pt = Point(o['x'], o['y'])
                if poly.contains(pt):
                    if poly_area < _SLIVER_AREA_THRESHOLD:
                        # 🚨 Phase 9.6：碎屑 polygon（< 1㎡），即使包含 OPAR 文字也應跳過
                        cost = 999.0
                    else:
                        cost = -1.0   # Tier 1：文字點在合法 polygon 內 → 必中
                elif poly_bf.contains(pt):
                    if poly_area < _SLIVER_AREA_THRESHOLD:
                        cost = 999.0
                    else:
                        # 🚨 Phase 9.4 Tier 1.5：文字點落於邊界 0.1m 緩衝內 → 視為「貼邊配對」
                        cost = -0.5   # 介於必中與邊界距離之間
                else:
                    # Tier 2：文字點到多邊形邊界之最近距離
                    d_boundary = float(poly.distance(pt))
                    cost = d_boundary * d_boundary   # 距離平方
            except Exception:
                # fallback：傳統質心距離
                cost = (o['x'] - cx) ** 2 + (o['y'] - cy) ** 2
            # 🚨 Phase 9.6：4-tuple，第 2 鍵 area 供 tiebreaker（小面積優先）
            candidates.append((cost, poly_area, pi, oi))

    # 🚨 Phase 9.6：依 (cost asc, area asc) 排序
    # 同 cost 時面積較小者優先 = 「最精準的外框」
    candidates.sort(key=lambda x: (x[0], x[1]))
    used_opar = set()
    poly_match = {}        # pi → oi
    match_tiers = {}       # parcel_no → 1（contained）/2（boundary nearest）

    for cost, area, pi, oi in candidates:
        if pi in poly_match:
            continue
        if oi in used_opar:
            continue
        # 🚨 Phase 9.6：cost = 999（碎屑）絕不配對，留給合法 polygon
        if cost >= 999.0:
            continue
        poly_match[pi] = oi
        used_opar.add(oi)
        match_tiers[opar_items[oi]['parcel']] = 1 if cost < 0 else 2

    # 將配對層次資訊存入 session_state 供 UI 診斷
    try:
        st.session_state['f3_dxf_match_tiers'] = match_tiers
    except Exception:
        pass

    # 組合結果（未配對者 parcel_no 留空）
    results = []
    for pi, poly in enumerate(valid_polys):
        try:
            oi = poly_match.get(pi)
            parcel_no = opar_items[oi]['parcel'] if oi is not None else ''
            cen = poly.centroid
            results.append({
                'parcel': parcel_no,
                'polygon_coords': list(poly.exterior.coords),
                'area_m2': poly.area,
                'centroid_x': cen.x,
                'centroid_y': cen.y,
            })
        except Exception:
            continue
    return results


def validate_parcel_assignments_by_area(parcel_polys: list,
                                          tab1_areas: dict,
                                          area_diff_threshold: float = 0.30,
                                          max_swap_distance_m: float = 30.0) -> dict:
    """
    🆕 Phase 11 Hotfix：以 Tab 1 登記面積交叉驗證地籍配對結果，自動偵測並修正換位錯誤

    背景：當小宗地內地號文字塞不下、被擺在外面時，幾何配對可能誤把該地號配到鄰近大宗地，
    形成「左右互換」的錯配（例如 628-26 ↔ 628-25 互換）。本函式以登記面積對照解決此問題：
      - 若 polygon 實測面積 vs 該 parcel 之 Tab 1 登記面積偏差 > 30%，視為「可疑配對」
      - 對所有可疑配對：嘗試與相鄰（< 30m）之另一可疑配對交換地號
      - 若交換後「兩者面積偏差和」減少 → 採用交換

    參數：
      parcel_polys     reconstruct_parcel_polygons 之輸出
      tab1_areas       {地號: 登記面積㎡}（來自 session_state['t8_parcel_areas']）
      area_diff_threshold  面積偏差閾值（預設 30%）
      max_swap_distance_m  允許交換之最大宗地中心距離（預設 30m）

    回傳：
      {
        'parcel_polys': 修正後之 parcel_polys list（順序維持原樣）,
        'swaps': [{'idx_a', 'idx_b', 'parcel_a_orig', 'parcel_b_orig',
                    'reason': '面積交叉檢核'}, ...],
        'suspects_remaining': [...]   # 修正後仍存疑之 idx
      }
    """
    if not parcel_polys or not tab1_areas:
        return {'parcel_polys': parcel_polys, 'swaps': [], 'suspects_remaining': []}

    def _norm(s):
        try:
            return _normalize_landno_module(s)
        except Exception:
            return str(s or '').strip()

    def _lookup_area(parcel_no):
        if not parcel_no:
            return 0.0
        # 直接查 → 正規化查
        a = tab1_areas.get(parcel_no)
        if a is None:
            a = tab1_areas.get(_norm(parcel_no))
        try:
            return float(a or 0.0)
        except Exception:
            return 0.0

    def _diff_ratio(actual, registered):
        if registered <= 0:
            return None   # 無登記面積 → 無法判定
        return abs(actual - registered) / registered

    # 標記每筆 polygon 之偏差比例
    out = [dict(p) for p in parcel_polys]
    suspects = []     # list of (idx, diff_ratio)
    for i, p in enumerate(out):
        pn = p.get('parcel', '')
        if not pn or pn == '(無地號)':
            continue
        reg = _lookup_area(pn)
        if reg <= 0:
            continue
        actual = float(p.get('area_m2', 0) or 0)
        d = _diff_ratio(actual, reg)
        if d is not None and d > area_diff_threshold:
            suspects.append((i, d))

    if len(suspects) < 2:
        return {'parcel_polys': out, 'swaps': [], 'suspects_remaining': [s[0] for s in suspects]}

    # 嘗試成對交換：對每對 (a, b) ∈ suspects，若中心距離 < max_swap_distance_m
    # 計算交換前後之偏差和，若改善 → 接受交換
    suspect_idxs = [s[0] for s in suspects]
    swapped = set()
    swaps = []
    for i_a in suspect_idxs:
        if i_a in swapped:
            continue
        pa = out[i_a]
        ax, ay = float(pa.get('centroid_x', 0)), float(pa.get('centroid_y', 0))
        pn_a = pa.get('parcel', '')
        actual_a = float(pa.get('area_m2', 0) or 0)
        reg_a = _lookup_area(pn_a)
        for i_b in suspect_idxs:
            if i_b == i_a or i_b in swapped:
                continue
            pb = out[i_b]
            bx, by = float(pb.get('centroid_x', 0)), float(pb.get('centroid_y', 0))
            dist = ((ax - bx) ** 2 + (ay - by) ** 2) ** 0.5
            if dist > max_swap_distance_m:
                continue
            pn_b = pb.get('parcel', '')
            actual_b = float(pb.get('area_m2', 0) or 0)
            reg_b = _lookup_area(pn_b)
            # 交換前偏差和
            err_before = (_diff_ratio(actual_a, reg_a) or 0) + (_diff_ratio(actual_b, reg_b) or 0)
            # 交換後：a 拿 pn_b 的登記面積、b 拿 pn_a 的登記面積
            err_after = (_diff_ratio(actual_a, reg_b) or 0) + (_diff_ratio(actual_b, reg_a) or 0)
            # 改善 ≥ 30%（避免微小波動造成誤交換）
            if err_after < err_before * 0.7:
                # 接受交換
                pa['parcel'] = pn_b
                pb['parcel'] = pn_a
                swapped.add(i_a)
                swapped.add(i_b)
                swaps.append({
                    'idx_a': i_a, 'idx_b': i_b,
                    'parcel_a_orig': pn_a, 'parcel_b_orig': pn_b,
                    'centroid_dist_m': round(dist, 2),
                    'err_before': round(err_before, 4),
                    'err_after': round(err_after, 4),
                })
                break

    # 計算尚未修正之 suspects（交換後仍偏差過大）
    suspects_remaining = []
    for i, _ in suspects:
        if i in swapped:
            continue
        p = out[i]
        pn = p.get('parcel', '')
        reg = _lookup_area(pn)
        if reg > 0:
            actual = float(p.get('area_m2', 0) or 0)
            if (_diff_ratio(actual, reg) or 0) > area_diff_threshold:
                suspects_remaining.append(i)

    return {'parcel_polys': out, 'swaps': swaps, 'suspects_remaining': suspects_remaining}


def _hash_polys_for_overlay(parcel_polys: list, classified_blocks: list) -> str:
    """以 MD5 將兩組多邊形資料壓成單一字串鍵，供 session_state 快取判斷。"""
    import hashlib, json
    def _simplify_parcels(pp):
        return [(p.get('parcel', ''), round(p.get('area_m2', 0), 2),
                 len(p.get('polygon_coords', []) or [])) for p in pp]
    def _simplify_blocks(bb):
        return [(b.get('id'), b.get('label', ''), b.get('category', ''),
                 round(b.get('area_m2', 0), 2), len(b.get('vertices', []) or []))
                for b in bb if not b.get('is_outer')]
    payload = json.dumps([_simplify_parcels(parcel_polys),
                          _simplify_blocks(classified_blocks)],
                         ensure_ascii=False, sort_keys=True)
    return hashlib.md5(payload.encode('utf-8')).hexdigest()


def _is_sliver_polygon(poly, parcel_total_area: float = 0.0,
                        min_area: float = 0.1,
                        sliver_aspect_ratio_strict: float = 15.0,
                        sliver_aspect_ratio_strong: float = 30.0,
                        sliver_threshold_tier1: float = 0.5,
                        sliver_threshold_tier2: float = 5.0) -> bool:
    """
    🚨 Patch E-0.1：複合特徵 sliver 判定（兼容評估期 + 正式期）

    判定邏輯（依面積分四階）：
      1. area < 0.1 ㎡    → 一律 sliver（浮點誤差級別）
      2. 0.1 ≤ area < 0.5 → aspect ratio > 15 才視為 sliver（中等大小細長條）
      3. 0.5 ≤ area < 5.0 → aspect ratio > 30 才視為 sliver（保護合法畸零地）
      4. area ≥ 5.0       → 一律保留（真實宗地切片）

    雙階段相容性：
      - 評估期：宗地切片多在 5-50㎡ → 全部保留
      - 正式期：CAD 浮點誤差多在 < 0.5㎡ → 由 tier1/tier2 配 aspect 過濾
    """
    try:
        if poly is None or poly.is_empty:
            return True
        a = float(poly.area)

        # 階段 1：浮點誤差
        if a < min_area:
            return True

        # 階段 2：0.1-0.5 ㎡ 中等小區塊（用 strict aspect ratio）
        if a < sliver_threshold_tier1:
            try:
                mbr = poly.minimum_rotated_rectangle
                mbr_coords = list(mbr.exterior.coords)
                if mbr_coords and mbr_coords[0] == mbr_coords[-1]:
                    mbr_coords = mbr_coords[:-1]
                if len(mbr_coords) >= 4:
                    edges = []
                    for i in range(4):
                        p1 = mbr_coords[i]; p2 = mbr_coords[(i + 1) % 4]
                        edges.append(((p2[0] - p1[0]) ** 2 + (p2[1] - p1[1]) ** 2) ** 0.5)
                    edges.sort()
                    short_e = edges[0]; long_e = edges[2]
                    if short_e > 0 and (long_e / short_e) > sliver_aspect_ratio_strict:
                        return True
            except Exception:
                pass
            return False  # 0.1-0.5㎡ 但非細長條 → 保留

        # 階段 3：0.5-5 ㎡ 中型區塊（用 strong aspect ratio，保護合法畸零地）
        if a < sliver_threshold_tier2:
            try:
                mbr = poly.minimum_rotated_rectangle
                mbr_coords = list(mbr.exterior.coords)
                if mbr_coords and mbr_coords[0] == mbr_coords[-1]:
                    mbr_coords = mbr_coords[:-1]
                if len(mbr_coords) >= 4:
                    edges = []
                    for i in range(4):
                        p1 = mbr_coords[i]; p2 = mbr_coords[(i + 1) % 4]
                        edges.append(((p2[0] - p1[0]) ** 2 + (p2[1] - p1[1]) ** 2) ** 0.5)
                    edges.sort()
                    short_e = edges[0]; long_e = edges[2]
                    if short_e > 0 and (long_e / short_e) > sliver_aspect_ratio_strong:
                        return True
            except Exception:
                pass
            return False  # 0.5-5㎡ 非極端細長 → 保留

        # 階段 4：≥ 5 ㎡ 一律保留
        return False
    except Exception:
        return False


def _is_perp_dist_below_threshold(inter_poly, block_poly, threshold: float = 0.1) -> bool:
    """
    🚨 W-A 規格三：0.1m 垂距門檻判定

    對 intersection polygon 的界址點（exterior 頂點），排除「與 BLOCK 邊界共用」
    的點（垂距 ≈ 0），對其餘「逕割界址點」算到 BLOCK 邊界的垂距：
      - 若無任何逕割界址點（全部貼在 BLOCK 邊界上）→ 視為沿邊貼合碎片 → True（不切）
      - 若 max(逕割界址點垂距) < threshold → 視為 1/500 圖解誤差 → True（不切）
      - 否則 → False（正常切暫編）

    法理：1/500 圖解區，筆尖 0.2mm ≈ 實地 10cm，同一點。
    """
    try:
        boundary = block_poly.exterior
        if inter_poly.geom_type == 'MultiPolygon':
            inter_poly = max(inter_poly.geoms, key=lambda g: g.area)
        coords = list(inter_poly.exterior.coords)
        if len(coords) < 3:
            return True
        from shapely.geometry import Point as _Pt_wa
        _ON_BOUNDARY_EPS = 0.01   # 垂距 < 1cm 視為與 BLOCK 邊界共用點
        cut_dists = []
        for (x, y) in coords:
            try:
                d = boundary.distance(_Pt_wa(x, y))
            except Exception:
                continue
            if d < _ON_BOUNDARY_EPS:
                continue   # 排除共用點（非有效判定樣本）
            cut_dists.append(d)
        if not cut_dists:
            return True   # 全部頂點貼在 BLOCK 邊界 → 沿邊碎片 → 不切
        return max(cut_dists) < threshold
    except Exception:
        return False


def _assign_four_column_areas(temp_parcels: list, t8_parcel_areas: dict) -> list:
    """
    🚨 W-A 規格一：四欄面積後處理（在 overlay 回傳之後、存 session 之前呼叫）

    overlay 已寫入 `幾何面積_m2`（DXF∩BLOCK 真值）與 `面積_m2`（暫=幾何）。
    本函式補上其餘三欄並重設 `面積_m2`：
      登記面積_m2     = t8_parcel_areas[原地號]（Tab 1 法定面積；查無 → 0）
      分攤登記面積_m2 = 登記面積 × (本暫編幾何 / Σ同原地號各暫編幾何)
                        若登記面積查無或為 0 → 退化用「幾何面積」當原生面積
      面積_m2         = 0.0（純 a' 累加器；G 公式分子 = 分攤登記 + 面積_m2）

    守恆：Σ同原地號各暫編之分攤登記面積 = 該原地號登記面積（圖簿完美對齊）。
    """
    if not temp_parcels:
        return temp_parcels
    t8 = t8_parcel_areas or {}

    # 第一遍：累計同原地號之各暫編幾何面積總和
    geo_sum_by_parent = {}
    for tp in temp_parcels:
        parent = str(tp.get('原地號', '') or '').strip()
        g = float(tp.get('幾何面積_m2', tp.get('面積_m2', 0)) or 0)
        geo_sum_by_parent[parent] = geo_sum_by_parent.get(parent, 0.0) + g

    # 🚨 W-A.2 §3：完整性檢核（缺口率 > 3% → 退化分攤=幾何，不縮放登記面積）
    #   病因：628-12 登記 200 被硬攤在幾何 55.19 的塊上（119.88/80.12 錯）——
    #   §1 救回 G1 塊後缺口率會降到 < 1%；殘留缺口大者多為「圖簿差」或
    #   未救回的真實土地，此時不該縮放分攤、應退化用幾何當原生面積。
    INTEGRITY_GAP_THRESHOLD = 0.03   # 3%
    incomplete_parents = set()
    integrity_warnings = []
    for parent, geo_sum in geo_sum_by_parent.items():
        if not parent or parent.startswith('_UNC') or parent.startswith('_GHOST'):
            continue
        reg = float(t8.get(parent, 0) or 0)
        if reg <= 0:
            try:
                reg = float(t8.get(_normalize_landno_module(parent), 0) or 0)
            except Exception:
                reg = 0.0
        if reg <= 0:
            continue   # 無登記面積本就退化，不計入警示
        gap = reg - geo_sum
        gap_ratio = abs(gap) / reg if reg > 0 else 0
        if gap_ratio > INTEGRITY_GAP_THRESHOLD:
            incomplete_parents.add(parent)
            integrity_warnings.append({
                '原地號': parent,
                '登記面積_m2': round(reg, 2),
                'Σ幾何面積_m2': round(geo_sum, 2),
                '缺口_m2': round(gap, 2),
                '缺口率_%': round(gap_ratio * 100, 2),
            })
    integrity_warnings.sort(key=lambda x: -x['缺口率_%'])

    # 第二遍：賦值四欄
    _largest_tp_by_parent = {}   # 原地號 → (最大幾何之 tp, 該幾何)
    for tp in temp_parcels:
        parent = str(tp.get('原地號', '') or '').strip()
        geo = float(tp.get('幾何面積_m2', tp.get('面積_m2', 0)) or 0)
        reg = float(t8.get(parent, 0) or 0)
        if reg <= 0:
            try:
                reg = float(t8.get(_normalize_landno_module(parent), 0) or 0)
            except Exception:
                reg = 0.0
        geo_sum = geo_sum_by_parent.get(parent, 0.0)
        if parent in incomplete_parents:
            # 🚨 W-A.2 §3：缺口 > 3% → 退化分攤 = 幾何（不縮放登記）
            allotted = round(geo, 2)
        elif reg > 0 and geo_sum > 0:
            allotted = round(reg * geo / geo_sum, 2)
        else:
            allotted = round(geo, 2)   # 退化：無登記面積
        tp['登記面積_m2'] = round(reg, 2)
        tp['分攤登記面積_m2'] = allotted
        tp['面積_m2'] = 0.0   # 純 a' 累加器
        if parent in incomplete_parents:
            tp['_integrity_incomplete'] = True   # 給 GIS 標警示色
        _prev = _largest_tp_by_parent.get(parent)
        if _prev is None or geo > _prev[1]:
            _largest_tp_by_parent[parent] = (tp, geo)

    # 守恆殘差校正：僅對「完整」原地號做（不完整地號分攤=幾何，無需校正）
    _allot_sum_by_parent = {}
    for tp in temp_parcels:
        parent = str(tp.get('原地號', '') or '').strip()
        _allot_sum_by_parent[parent] = (_allot_sum_by_parent.get(parent, 0.0)
                                        + float(tp.get('分攤登記面積_m2', 0) or 0))
    for parent, (tp_max, _g) in _largest_tp_by_parent.items():
        if parent in incomplete_parents:
            continue
        reg = float(tp_max.get('登記面積_m2', 0) or 0)
        if reg <= 0:
            continue
        residual = round(reg - _allot_sum_by_parent.get(parent, 0.0), 2)
        if abs(residual) >= 0.01:
            tp_max['分攤登記面積_m2'] = round(
                float(tp_max.get('分攤登記面積_m2', 0) or 0) + residual, 2)

    # 寫入 session_state 供 UI 顯示警示清單
    try:
        st.session_state['f3_integrity_warnings'] = integrity_warnings
        st.session_state['f3_integrity_incomplete_parents'] = list(incomplete_parents)
    except Exception:
        pass

    return temp_parcels


def _annotate_temp_parcel_cut_type(temp_parcels: list) -> list:
    """
    🆕 W-A.3 §2（方案 B）：標記暫編「切割性質」cut_type，供未來實際作業判定哪些
    線要送地政逕為分割。**只加標記欄位，不改四欄面積、不改 G 計算。**

    判定（依 KL §2 定案）：同一原地號的「相鄰」暫編，兩兩比對所屬街廓 category：
      - 不同 category（使用分區界，如住宅↔道路、住宅↔公園、住宅↔商業）
          → 'cut_逕割'：逕割性質（未來逕為分割成真實子地號，各有登記面積）
      - 同 category 且非道路類（同可建築分區，被 BASELINE 屁股線/分配線切）
          → 'cut_分配'：分配性質（試分配作業用，不逕割，用分攤登記面積）
      - 同 category 且皆道路類（RD↔RD，為切直線道路而畫的 BLOCK 線）
          → 'cut_非分割'：非真實分割（未來 RD 調配以 CENTERLINE 重新處理）
      - 無同原地號之相鄰塊（整筆未跨界）
          → 'cut_整筆'

    相鄰判定：以暫編 polygon 邊界互相接觸（沿用現成 polygon_coords，
    **不做圖層邊比對**）。只比同原地號之相鄰塊，故「住宅A／道路／住宅B」中
    住宅A 與住宅B 因隔著道路不相鄰，不會被誤判為同住宅之分配性質，
    而是各自與中間道路比對、判為逕割。

    單塊 cut_type 取其各相鄰關係之優先序：逕割 > 分配 > 非分割 > 整筆
    （邊界只要有一段是使用分區界，未來該塊即涉逕為分割）。
    """
    if not temp_parcels:
        return temp_parcels
    from shapely.geometry import Polygon as _Poly_ct

    ADJ_TOL = 0.30   # m：暫編 polygon 邊界接觸容差（CAD 浮點/共用邊微縫）

    # 預建 shapely polygon（沿用現成 polygon_coords）
    _polys = {}
    for i, tp in enumerate(temp_parcels):
        coords = tp.get('polygon_coords', []) or []
        if len(coords) >= 3:
            try:
                _p = _Poly_ct(coords)
                if not _p.is_valid:
                    _p = _p.buffer(0)
                _polys[i] = _p if (not _p.is_empty) else None
            except Exception:
                _polys[i] = None
        else:
            _polys[i] = None

    # 依原地號分群（_UNC/_GHOST 匿名殘料不參與跨界判定）
    _by_parent = {}
    for i, tp in enumerate(temp_parcels):
        parent = str(tp.get('原地號', '') or '').strip()
        if not parent or parent.startswith('_UNC') or parent.startswith('_GHOST'):
            continue
        _by_parent.setdefault(parent, []).append(i)

    for i, tp in enumerate(temp_parcels):
        parent = str(tp.get('原地號', '') or '').strip()
        my_cat = str(tp.get('街廓分類', '') or '')
        my_poly = _polys.get(i)
        rels = set()
        if my_poly is not None and parent in _by_parent:
            for j in _by_parent[parent]:
                if j == i:
                    continue
                other_poly = _polys.get(j)
                if other_poly is None:
                    continue
                try:
                    if my_poly.distance(other_poly) <= ADJ_TOL:
                        other_cat = str(temp_parcels[j].get('街廓分類', '') or '')
                        if my_cat != other_cat:
                            rels.add('cut_逕割')
                        elif _is_road_like_category(my_cat):
                            rels.add('cut_非分割')
                        else:
                            rels.add('cut_分配')
                except Exception:
                    continue
        if 'cut_逕割' in rels:
            tp['cut_type'] = 'cut_逕割'
        elif 'cut_分配' in rels:
            tp['cut_type'] = 'cut_分配'
        elif 'cut_非分割' in rels:
            tp['cut_type'] = 'cut_非分割'
        else:
            tp['cut_type'] = 'cut_整筆'
    return temp_parcels


def _annotate_block_corner_flags(classified_blocks: list,
                                 side_lengths_by_side: dict) -> list:
    """
    🆕 W-A.3 §4：依 SIDE_LINE 左右側存在與否，標記街廓有無街角，供 W-B/W-C/W-D 讀取。
    **只加標記欄位，不改任何幾何或 G 計算。**

    判定（依 v3.1 概念 5「街角地非必然」）：
      - side_lengths_by_side[label] = {'left': L, 'right': R}
      - left > 0  → 該街廓左端有街角（corner_sides 含 'left'）
      - right > 0 → 該街廓右端有街角（corner_sides 含 'right'）
      - 皆 0 或查無 → has_corner=False（無 SIDE_LINE 街廓：跳過街角評點與 Rw）

    SIDE_LINE 僅綁定至可建築街廓，故非可建築街廓自然 has_corner=False。
    """
    _sbs = side_lengths_by_side or {}
    for b in classified_blocks or []:
        lbl = b.get('label', '')
        sides = _sbs.get(lbl, {}) or {}
        corner_sides = []
        try:
            if float(sides.get('left', 0) or 0) > 0:
                corner_sides.append('left')
            if float(sides.get('right', 0) or 0) > 0:
                corner_sides.append('right')
        except Exception:
            corner_sides = []
        b['has_corner'] = len(corner_sides) > 0
        b['corner_sides'] = corner_sides
    return classified_blocks


def overlay_polygons_to_blocks(parcel_polys: list, classified_blocks: list) -> list:
    """
    對每筆重劃前宗地多邊形 × 每個街廓多邊形算 intersection。
    🆕 Phase 11 Hotfix：自動過濾 sliver 細長帶狀碎片
      （街廓 DXF 與地籍 DXF 外框不 100% 重疊時，intersection 會產生長條 sliver）
    若 area > 1 ㎡ 且非 sliver 才產生一筆暫編地號。
    回傳：[{'原地號','暫編地號','所屬街廓','街廓分類','面積_m2','polygon_coords',...}]
    ※ 本函式內部以 st.session_state['_f3_overlay_cache'] 做結果快取
      （key = 多邊形資料之 MD5），避免 Tab 切換觸發重算

    🚨 Patch E-1.5：cache 隔離 — cache 與 caller 永遠擁有獨立副本，
       下游 in-place 修改不會污染 cache（解 Tier 1 a' 累加爆量問題）
    """
    import copy as _copy_e15
    try:
        cache_key = _hash_polys_for_overlay(parcel_polys, classified_blocks)
        _c = st.session_state.setdefault('_f3_overlay_cache', {})
        if cache_key in _c:
            # 🚨 Patch E-1.5：cache hit 回傳深拷貝，避免下游 in-place 修改污染 cache
            return _copy_e15.deepcopy(_c[cache_key])
    except Exception:
        cache_key = None
        _c = None
    from shapely.geometry import Polygon

    block_polys = []
    for b in classified_blocks:
        if b.get('is_outer'):
            continue
        try:
            sp = Polygon(b['vertices'])
            if not sp.is_valid:
                sp = sp.buffer(0)
            block_polys.append({
                'label': b.get('label', '') or f"區塊#{b.get('id','?')}",
                'category': b.get('category', ''),
                'sp': sp,
            })
        except Exception:
            continue

    out = []
    seen_cnt = {}
    for pp in parcel_polys:
        parcel_no = pp.get('parcel', '') or '(無地號)'
        parcel_total_area = float(pp.get('area_m2', 0.0) or 0.0)
        try:
            p_sp = Polygon(pp['polygon_coords'])
            if not p_sp.is_valid:
                p_sp = p_sp.buffer(0)
            if parcel_total_area <= 0:
                parcel_total_area = float(p_sp.area or 0.0)
        except Exception:
            continue
        for b in block_polys:
            try:
                inter = p_sp.intersection(b['sp'])
                # 🚨 W-A.2 §1：有原地號歸屬的交集塊一律建為暫編（面積 ≥ 0.05㎡）。
                #   撤回 _is_sliver_polygon / 0.1m 垂距對有主塊的剔除——
                #   形狀過濾只用於 §2 的「無主殘料」分流（_GHOST / _UNC）。
                #   根因：細長條常是跨道路被切出的真實土地（KL 2026-06-15 §0 診斷確認
                #   628-12/35/37/45 等 93-100% 命中真實地號 polygon）。
                #   0.1m 垂距規則原語意為「整筆地號與 BLOCK 線重合」，非「逐片剔除貼邊薄片」——
                #   逐片用法移除以還原語意。
                if inter.is_empty or inter.area < 0.05:
                    continue
                # 🆕 W-C §0.5-F.1：MultiPolygon 逐片拆獨立暫編（修 a 灌大 bug）。
                #   舊碼：面積取 inter.area「總和」、coords 只留「最大片」→ 該暫編「面積≠形狀」，
                #   a（=登記面積×幾何比）被灌進其餘片的幻影面積（如 628-18∩R5 = 381.09+96.82
                #   兩片不相連 → 舊記 a=477.91、只畫 381.09）。
                #   正解：同街廓內多片不相連 → 每片各自成暫編（各自 area+coords，逐片套 sliver 過濾）。
                if inter.geom_type == 'Polygon':
                    _pieces = [inter]
                elif inter.geom_type == 'MultiPolygon':
                    _pieces = list(inter.geoms)
                else:
                    _pieces = []
                for _pc in _pieces:
                    if _pc.is_empty or _pc.area < 0.05:   # 逐片 sliver 過濾（同外層門檻）
                        continue
                    seen_cnt[parcel_no] = seen_cnt.get(parcel_no, 0) + 1
                    temp_no = f"{parcel_no}({seen_cnt[parcel_no]})"
                    cen = _pc.centroid
                    _geo_area = round(_pc.area, 2)
                    out.append({
                        '原地號': parcel_no,
                        '暫編地號': temp_no,
                        '所屬街廓': b['label'],
                        '街廓分類': b['category'],
                        # 🚨 W-A 四欄面積：幾何面積_m2 = DXF∩BLOCK 真值（逐片，唯讀永不污染）
                        '幾何面積_m2': _geo_area,
                        '面積_m2': 0.0,   # 🆕 W-C 裁示②：拆片時 a' 累加器=0（不種幾何，防道路/兄弟片面積注入）
                        '面積_坪': round(_pc.area * 0.3025, 2),
                        '重劃前地價區段': '',
                        'polygon_coords': list(_pc.exterior.coords),
                        'centroid_x': cen.x,
                        'centroid_y': cen.y,
                    })
            except Exception:
                continue

    # W-A.2 §1：撤回有主交集塊形狀剔除後 sliver_dropped 變數已移除
    #   剩餘真．無主殘料的計數請看 f3_unassigned_added_count / f3_unassigned_ghost_count

    # ═══════════════════════════════════════════════════════════
    # 🚨 W-A.2 §2：uncovered 殘料 — 歸屬優先（取代 E-0.2 同 BLOCK 最近邏輯）
    # ───────────────────────────────────────────────────────────
    # 對每個 uncovered sub 子塊：
    #   1. 與每個 parcel_polys 的 polygon 求交集（歸屬修復）：
    #      - inter.area ≥ 0.05㎡ → 該部分歸該地號，補建為正常暫編；
    #        從 sub 扣除該歸屬部分，殘餘繼續比對下一地號。
    #   2. 比對完仍無主之殘餘：
    #      - 距重劃區邊界 < 2m → ghost（邊界帶碎塊，CAD 拓樸誤差）
    #      - 區內 ≥ 50㎡        → _UNC 抵費地候選
    #      - 區內 < 50㎡         → ghost + 計數警示
    #
    # 此設計把 _UNC4 混合塊精確拆回 628-45/628-30/628-31/628-22 各自暫編，
    # 把 _UNC1/2/3 整塊歸還真正地號（628-35/37/12）。
    # ═══════════════════════════════════════════════════════════
    try:
        from shapely.geometry import Polygon as _SP_unc
        from shapely.ops import unary_union as _uu_unc
        UNCOVERED_TOL = 0.05           # 容差 0.05㎡（W-A.2 §1 同步）
        ATTRIB_MIN_AREA = 0.05         # 歸屬交集最小面積
        BOUNDARY_BAND_M = 2.0          # 距重劃區邊界 < 2m → ghost
        UNC_AREA_MIN = 50.0            # ≥ 50㎡ → _UNC 抵費地候選

        # 建重劃區邊界（所有非 outer BLOCK 聯集的 boundary）
        _reform_boundary = None
        try:
            _all_blk_sps = [b['sp'] for b in block_polys
                            if b.get('sp') is not None and not b['sp'].is_empty]
            if _all_blk_sps:
                _reform_union = _uu_unc(_all_blk_sps)
                _reform_boundary = _reform_union.boundary
        except Exception:
            _reform_boundary = None

        # 預建 parcel_polys 的 shapely polygon
        _pp_shapely = []
        for pp in parcel_polys:
            pno = str(pp.get('parcel', '') or '').strip()
            pcoords = pp.get('polygon_coords', []) or []
            if not pno or len(pcoords) < 3:
                continue
            try:
                poly_pp = _SP_unc(pcoords)
                if not poly_pp.is_valid:
                    poly_pp = poly_pp.buffer(0)
                if not poly_pp.is_empty:
                    _pp_shapely.append((pno, poly_pp))
            except Exception:
                continue

        _unc_idx = 0
        _unc_added_count = 0
        _unc_ghost_count = 0
        _attrib_repaired_count = 0   # W-A.2 §2 歸屬修復成功筆數

        for blk in (classified_blocks or []):
            if blk.get('is_outer'):
                continue
            verts = blk.get('vertices', []) or []
            if len(verts) < 3:
                continue
            try:
                blk_poly = _SP_unc(verts)
                if not blk_poly.is_valid:
                    blk_poly = blk_poly.buffer(0)
                if blk_poly.is_empty:
                    continue
            except Exception:
                continue
            blk_label_unc = blk.get('label', '') or f"區塊#{blk.get('id', '?')}"
            blk_cat_unc = blk.get('category', '')

            # 該街廓內所有 temp_parcels 之 polygon 聯集（排除 ghost）
            same_blk_tps = [tp for tp in out
                             if tp.get('所屬街廓', '') == blk_label_unc
                             and not tp.get('_is_ghost_sliver', False)]
            polys = []
            for tp in same_blk_tps:
                cs = tp.get('polygon_coords') or []
                if len(cs) >= 3:
                    try:
                        p = _SP_unc(cs)
                        if not p.is_valid:
                            p = p.buffer(0)
                        if not p.is_empty:
                            polys.append(p)
                    except Exception:
                        pass
            # W-A.2：即使整街廓無 temp（polys 空），整塊 uncovered = blk_poly 仍要處理
            try:
                if polys:
                    covered = _uu_unc(polys)
                    uncovered = blk_poly.difference(covered)
                else:
                    uncovered = blk_poly
            except Exception:
                continue
            if uncovered.is_empty or float(uncovered.area) < UNCOVERED_TOL:
                continue

            sub_polys = (list(uncovered.geoms)
                         if uncovered.geom_type == 'MultiPolygon'
                         else [uncovered])

            for sub in sub_polys:
                sub_area = float(sub.area)
                if sub_area < UNCOVERED_TOL:
                    continue

                # ── §2-1：歸屬修復 ──
                residue = sub
                for pno, pp_poly in _pp_shapely:
                    if residue.is_empty:
                        break
                    try:
                        # 限制在該 BLOCK 內，避免歸屬越界
                        inter = residue.intersection(pp_poly).intersection(blk_poly)
                    except Exception:
                        continue
                    if inter.is_empty:
                        continue
                    ia = float(inter.area)
                    if ia < ATTRIB_MIN_AREA:
                        continue

                    # 🆕 W-C §0.5-F.1：殘料歸屬亦逐片拆暫編（同主切割，修 a 灌大）
                    if inter.geom_type == 'Polygon':
                        _pieces2 = [inter]
                    elif inter.geom_type == 'MultiPolygon':
                        _pieces2 = list(inter.geoms)
                    else:
                        continue
                    for _pc2 in _pieces2:
                        if _pc2.is_empty or _pc2.area < ATTRIB_MIN_AREA:
                            continue
                        inter_coords = list(_pc2.exterior.coords)
                        if len(inter_coords) < 3:
                            continue
                        seen_cnt[pno] = seen_cnt.get(pno, 0) + 1
                        temp_no = f"{pno}({seen_cnt[pno]})"
                        cen_inter = _pc2.centroid
                        _g = round(_pc2.area, 2)
                        out.append({
                            '原地號': pno,
                            '暫編地號': temp_no,
                            '所屬街廓': blk_label_unc,
                            '街廓分類': blk_cat_unc,
                            '幾何面積_m2': _g,
                            '面積_m2': 0.0,   # 🆕 W-C 裁示②：殘料歸屬片 a' 累加器=0
                            '面積_坪': round(_pc2.area * 0.3025, 2),
                            '重劃前地價區段': '',
                            'polygon_coords': inter_coords,
                            'centroid_x': cen_inter.x,
                            'centroid_y': cen_inter.y,
                            '_wa2_attrib_repaired': True,
                        })
                        _attrib_repaired_count += 1

                    # 從 residue 扣除已歸屬部分
                    try:
                        residue = residue.difference(pp_poly)
                        if not residue.is_valid:
                            residue = residue.buffer(0)
                    except Exception:
                        pass

                # ── §2-2：殘餘 → ghost / _UNC 分流 ──
                if residue.is_empty or float(residue.area) < UNCOVERED_TOL:
                    continue

                residue_subs = (list(residue.geoms)
                                if residue.geom_type == 'MultiPolygon'
                                else [residue])

                for res_sub in residue_subs:
                    res_area = float(res_sub.area)
                    if res_area < UNCOVERED_TOL:
                        continue

                    _is_boundary_band = False
                    if _reform_boundary is not None:
                        try:
                            if res_sub.distance(_reform_boundary) < BOUNDARY_BAND_M:
                                _is_boundary_band = True
                        except Exception:
                            pass

                    try:
                        cen = res_sub.representative_point()
                        cx, cy = float(cen.x), float(cen.y)
                    except Exception:
                        cx, cy = 0.0, 0.0
                    res_coords = (list(res_sub.exterior.coords)
                                  if res_sub.geom_type == 'Polygon' else [])

                    if _is_boundary_band or res_area < UNC_AREA_MIN:
                        _reason = ('邊界帶' if _is_boundary_band
                                   else f'<{UNC_AREA_MIN:.0f}㎡ 殘料')
                        out.append({
                            '原地號': '_GHOST',
                            '暫編地號': f'_GHOST_({blk_label_unc})',
                            '所屬街廓': blk_label_unc,
                            '街廓分類': blk_cat_unc,
                            '幾何面積_m2': 0,
                            '面積_m2': 0,
                            '面積_坪': 0,
                            '重劃前地價區段': '',
                            'polygon_coords': res_coords,
                            'centroid_x': cx,
                            'centroid_y': cy,
                            '_is_ghost_sliver': True,
                            '_ghost_area_m2': round(res_area, 2),
                            '_ghost_reason': _reason,
                        })
                        _unc_ghost_count += 1
                    else:
                        _unc_idx += 1
                        out.append({
                            '原地號': f'_UNC{_unc_idx}',
                            '暫編地號': f'_UNC{_unc_idx}_({blk_label_unc})',
                            '所屬街廓': blk_label_unc,
                            '街廓分類': blk_cat_unc,
                            '幾何面積_m2': round(res_area, 2),
                            '面積_m2': 0.0,   # 🆕 W-C 裁示②：無主殘料片 a' 累加器=0（道路片面積不注入兄弟片）
                            '面積_坪': round(res_area * 0.3025, 2),
                            '重劃前地價區段': '',
                            'polygon_coords': res_coords,
                            'centroid_x': cx,
                            'centroid_y': cy,
                            '_unassigned_region': True,
                            '_uncovered_note': 'W-A.2 §2 真．無主殘料（歸屬比對 < 0.05㎡）',
                            '_treat_as_offset': True,
                        })
                        _unc_added_count += 1

        try:
            st.session_state['f3_unassigned_added_count'] = _unc_added_count
            st.session_state['f3_unassigned_ghost_count'] = _unc_ghost_count
            st.session_state['f3_wa2_attrib_repaired_count'] = _attrib_repaired_count
            st.session_state['f3_unassigned_absorbed_count'] = 0
        except Exception:
            pass
    except Exception:
        pass

    # 寫入 session_state 快取
    try:
        if cache_key and _c is not None:
            # 🚨 Patch E-1.5：寫入時存深拷貝，cache 永遠擁有原版（caller 可自由修改 out）
            _c[cache_key] = _copy_e15.deepcopy(out)
            # 限制快取筆數，避免無限成長
            if len(_c) > 6:
                first_key = next(iter(_c))
                _c.pop(first_key, None)
    except Exception:
        pass

    return out


def calc_frontage_burden_scale(road_width_m: float) -> float:
    """
    依市地重劃實施辦法第 29 條附件二 / PDF P18 表計算正街負擔尺度
    - 路寬 ≤ 4 M：不計臨街地特別負擔（回傳 0）
    - 4 < 路寬 < 8 M：(路寬 − 4) ÷ 2
    - 8 ≤ 路寬 < 20 M：路寬 ÷ 4
    - 路寬 ≥ 20 M：固定 5 M
    """
    try:
        w = float(road_width_m)
    except Exception:
        return 0.0
    if w <= 4.0:
        return 0.0
    if w < 8.0:
        return (w - 4.0) / 2.0
    if w < 20.0:
        return w / 4.0
    return 5.0


def calc_side_burden_scale(road_width_m: float) -> float:
    """側街負擔尺度 = 正街尺度 × 1/2（PDF P18）"""
    return calc_frontage_burden_scale(road_width_m) * 0.5


def calc_extension_intersection_length(v_front_a, v_front_b, v_side_a, v_side_b):
    """
    計算臨正街與臨側街道路線段延伸之交點
    輸入：正街線段兩端點 (v_front_a, v_front_b)、側街線段兩端點 (v_side_a, v_side_b)
    回傳：(L_front, L_side, intersect_point)
      L_front = 正街「遠離側街的一端」到延伸交點的距離（不計道路截角）
      L_side  = 側街「遠離正街的一端」到延伸交點的距離
    若兩線平行（無交點）則回傳 (實際長度, 實際長度, None)
    """
    import numpy as _np
    p1 = _np.array(v_front_a, dtype=float)
    p2 = _np.array(v_front_b, dtype=float)
    p3 = _np.array(v_side_a, dtype=float)
    p4 = _np.array(v_side_b, dtype=float)
    d1 = p2 - p1
    d2 = p4 - p3
    denom = d1[0] * d2[1] - d1[1] * d2[0]
    if abs(denom) < 1e-9:
        return (float(_np.linalg.norm(p2 - p1)), float(_np.linalg.norm(p4 - p3)), None)
    t = ((p3[0] - p1[0]) * d2[1] - (p3[1] - p1[1]) * d2[0]) / denom
    ip = p1 + t * d1
    # 取「遠離交點」的端點計算距離
    L_front = float(max(_np.linalg.norm(p1 - ip), _np.linalg.norm(p2 - ip)))
    L_side = float(max(_np.linalg.norm(p3 - ip), _np.linalg.norm(p4 - ip)))
    return (L_front, L_side, (float(ip[0]), float(ip[1])))


def calc_block_extension_lengths(block_vertices: list) -> dict:
    """
    由街廓多邊形自動計算「臨正面道路長度」、「臨左側道路長度」、「臨右側道路長度」
    作法：
      1. 取街廓最小外接旋轉矩形（minimum_rotated_rectangle）→ 4 頂點 4 邊
      2. 長邊配對視為「正面」（北向 + 南向）；短邊視為「側面」（東向 + 西向）
      3. 「延伸交點路長」= 矩形邊長（矩形情形下兩延伸線交點即為矩形頂點）
         → 不計道路截角，與《市地重劃實施辦法》第 29 條附件二意圖一致
      4. 對不規則街廓，以最小外接矩形近似仍符合實務；若精度需求高可人工覆寫
    回傳：
      {
        'front_length': float,       # 正面路長（矩形長邊）
        'left_side_length': float,   # 左側路長（矩形短邊）
        'right_side_length': float,  # 右側路長（矩形短邊；一般同左側）
        'angle_deg': float,          # 矩形主軸偏角
        'rect_coords': [(x,y),...]   # 矩形 4 頂點（可視化用）
      }
    """
    import math
    from shapely.geometry import Polygon
    default = {'front_length': 0.0, 'left_side_length': 0.0, 'right_side_length': 0.0,
               'angle_deg': 0.0, 'rect_coords': []}
    try:
        if not block_vertices or len(block_vertices) < 3:
            return default
        poly = Polygon(block_vertices)
        if not poly.is_valid:
            poly = poly.buffer(0)
        mrr = poly.minimum_rotated_rectangle
        coords = list(mrr.exterior.coords)  # 首尾重複，5 點
        if len(coords) < 5:
            return default
        # 四邊長度
        edges = []
        for i in range(4):
            x1, y1 = coords[i]
            x2, y2 = coords[i + 1]
            L = math.hypot(x2 - x1, y2 - y1)
            edges.append({'len': L, 'p1': (x1, y1), 'p2': (x2, y2)})
        # 相對邊等長：索引 0/2 一組、1/3 一組
        pair_A = (edges[0]['len'] + edges[2]['len']) / 2
        pair_B = (edges[1]['len'] + edges[3]['len']) / 2
        long_edge = max(pair_A, pair_B)
        short_edge = min(pair_A, pair_B)
        # 矩形主軸角（取長邊方向）
        if pair_A >= pair_B:
            dx = edges[0]['p2'][0] - edges[0]['p1'][0]
            dy = edges[0]['p2'][1] - edges[0]['p1'][1]
        else:
            dx = edges[1]['p2'][0] - edges[1]['p1'][0]
            dy = edges[1]['p2'][1] - edges[1]['p1'][1]
        angle = math.degrees(math.atan2(dy, dx))
        return {
            'front_length': round(long_edge, 2),
            'left_side_length': round(short_edge, 2),
            'right_side_length': round(short_edge, 2),
            'angle_deg': round(angle, 2),
            'rect_coords': coords,
        }
    except Exception:
        return default


def calc_special_burden_total(blocks_with_roads: list, C_value: float) -> dict:
    """
    依各可建築街廓之「正面／左側／右側」各自路寬、路長、是否新闢道路
    計算：
      正面道路負擔面積 = 正街尺度 × 正面長度（僅新闢道路計入）
      左側道路負擔面積 = 側街尺度 × 左側長度（僅新闢道路計入）
      右側道路負擔面積 = 側街尺度 × 右側長度（僅新闢道路計入）
      臨街地特別負擔總面積 = (Σ正面 + Σ左側 + Σ右側) × (1 − C)
    ※ 相容舊欄位：若未提供 left_*/right_* 則沿用舊 side_*（視為左側單條）
    """
    front_total = 0.0
    side_total = 0.0
    rows = []
    for b in blocks_with_roads:
        fw = float(b.get('front_width', 0.0) or 0.0)
        fl = float(b.get('front_length', 0.0) or 0.0)
        fn = bool(b.get('front_new', False))
        # 左側：優先讀 left_*，缺欄位則 fallback 到舊 side_*
        lw = float(b.get('left_side_width',  b.get('side_width',  0.0)) or 0.0)
        ll = float(b.get('left_side_length', b.get('side_length', 0.0)) or 0.0)
        ln = bool(b.get('left_side_new',     b.get('side_new',    False)))
        # 右側：無則視為不存在
        rw = float(b.get('right_side_width', 0.0) or 0.0)
        rl = float(b.get('right_side_length', 0.0) or 0.0)
        rn = bool(b.get('right_side_new', False))

        fs  = calc_frontage_burden_scale(fw) if fn else 0.0
        lss = calc_side_burden_scale(lw) if ln else 0.0
        rss = calc_side_burden_scale(rw) if rn else 0.0

        f_area = fs  * fl
        l_area = lss * ll
        r_area = rss * rl
        front_total += f_area
        side_total  += (l_area + r_area)

        rows.append({
            '街廓': b.get('label', ''),
            '正面路寬(m)':  fw, '正街尺度':  round(fs,  4), '正面長度(m)':  fl, '正面面積(㎡)': round(f_area, 2),
            '左側路寬(m)':  lw, '左側尺度':  round(lss, 4), '左側長度(m)':  ll, '左側面積(㎡)': round(l_area, 2),
            '右側路寬(m)':  rw, '右側尺度':  round(rss, 4), '右側長度(m)':  rl, '右側面積(㎡)': round(r_area, 2),
        })
    special_total = (front_total + side_total) * max(0.0, 1.0 - float(C_value or 0.0))
    return {
        'front_total': round(front_total, 2),
        'side_total': round(side_total, 2),
        'special_total': round(special_total, 2),
        'rows': rows,
    }


# ============ 街廓多邊形幾何分析（道路截角還原 + 邊界分類） ============
# 依「升級指南任務二」實作，結合 shapely 計算幾何
# 【2026-04 修復版】改用「拓樸鄰接」演算法，避免拿平行的對立邊去求交點
# 用途：DXF 街廓多邊形常含 3~6m 截角小邊；本模組可：
#   1) 找出拓樸上的截角邊（落於 cutoff_min_m ~ cutoff_max_m）
#   2) 對「每條截角邊」取其「拓樸上相鄰的左右兩條邊」（共享端點）
#   3) 用這兩條相鄰邊延伸求理論街角點（必為相鄰邊，絕不會是平行對立邊）
#   4) 計算截角三角形面積、還原真 S 長與側長

def line_intersection(p1, p2, p3, p4):
    """
    計算「點 p1→p2 連成之直線」與「點 p3→p4 連成之直線」之延伸交點
    參數均為 (x, y)
    回傳：(x, y)；若兩線平行則回傳 None
    """
    x1, y1 = float(p1[0]), float(p1[1])
    x2, y2 = float(p2[0]), float(p2[1])
    x3, y3 = float(p3[0]), float(p3[1])
    x4, y4 = float(p4[0]), float(p4[1])
    denom = (x1 - x2) * (y3 - y4) - (y1 - y2) * (x3 - x4)
    if abs(denom) < 1e-9:
        return None  # 平行
    t = ((x1 - x3) * (y3 - y4) - (y1 - y3) * (x3 - x4)) / denom
    px = x1 + t * (x2 - x1)
    py = y1 + t * (y2 - y1)
    return (round(px, 6), round(py, 6))


def calc_cutoff_triangle_area(cutoff_edge: dict, theoretical_corner) -> float:
    """
    截角三角形面積：以截角邊兩端點 + 理論街角點構成三角形
    使用 Shoelace 公式
    """
    if not cutoff_edge or theoretical_corner is None:
        return 0.0
    a = cutoff_edge['p1']
    b = cutoff_edge['p2']
    c = theoretical_corner
    area = abs((a[0] * (b[1] - c[1]) + b[0] * (c[1] - a[1])
                + c[0] * (a[1] - b[1])) / 2.0)
    return round(area, 4)


def classify_block_edges(polygon, cutoff_min_m: float = 2.0,
                         cutoff_max_m: float = 8.0) -> dict:
    """
    識別街廓多邊形之邊界類型（純資訊，不再用於後續還原）
    回傳 dict：
      {
        'edges': [{'idx': i, 'p1': (x,y), 'p2': (x,y), 'length': L}, ...]
        'front_idx': 最長邊 index（資訊性；正面道路一般是最長邊）
        'cutoff_idxs': 長度落於 cutoff 區間之邊 index 列表
        'side_idxs':   排除正面與截角後之次長邊（最多 2 條）
      }
    ※ 注意：本函式不負責配對截角與相鄰邊，僅做邊長度標籤；
       後續還原請改用 restore_block_geometry，依拓樸鄰接做配對。
    """
    try:
        from shapely.geometry import Polygon as _Poly  # noqa: F401
    except Exception:
        return {'edges': [], 'front_idx': -1, 'cutoff_idxs': [], 'side_idxs': []}

    if polygon is None or polygon.is_empty:
        return {'edges': [], 'front_idx': -1, 'cutoff_idxs': [], 'side_idxs': []}

    coords = list(polygon.exterior.coords)
    if len(coords) < 4:
        return {'edges': [], 'front_idx': -1, 'cutoff_idxs': [], 'side_idxs': []}
    if coords[0] == coords[-1]:
        coords = coords[:-1]

    edges = []
    n = len(coords)
    for i in range(n):
        p1 = coords[i]
        p2 = coords[(i + 1) % n]
        length = ((p2[0] - p1[0]) ** 2 + (p2[1] - p1[1]) ** 2) ** 0.5
        edges.append({'idx': i, 'p1': p1, 'p2': p2, 'length': round(length, 4)})

    if not edges:
        return {'edges': [], 'front_idx': -1, 'cutoff_idxs': [], 'side_idxs': []}

    front_idx = max(range(len(edges)), key=lambda i: edges[i]['length'])  # 資訊性最長邊（front_idx vestigial）
    # 🆕 W-D.1.3-a 切換：廢長度啟發法 cutoff_idxs（截角改 post-pass 拓樸定錨）；保留欄位名為空。
    cutoff_idxs = []
    remaining = [(i, e['length']) for i, e in enumerate(edges)
                 if i != front_idx and i not in cutoff_idxs]
    remaining.sort(key=lambda x: -x[1])
    side_idxs = [i for i, _ in remaining[:2]]

    return {
        'edges': edges, 'front_idx': front_idx,
        'cutoff_idxs': cutoff_idxs, 'side_idxs': side_idxs,
    }


# 🆕 W-D.1.3-a 切換：restore_theoretical_corner / _angle_between_edges_deg（原長度法 restore 迴圈之 helper）
#   隨長度法迴圈整段刪除後全檔零呼叫 → 一併刪（CLAUDE.md：舊函式整個刪、不留 stub）。


def restore_block_geometry(polygon, cutoff_min_m: float = 2.0,
                           cutoff_max_m: float = 8.0,
                           sanity_max_corner_dist_m: float = 50.0,
                           sanity_max_cutoff_area_m2: float = 50.0,
                           simplify_tolerance_m: float = 0.1,
                           min_corner_angle_deg: float = 45.0,
                           max_corner_angle_deg: float = 135.0) -> dict:
    """
    【2026-04 強化版】對街廓多邊形執行幾何還原
    主要改良：
      A. 多邊形簡化：先 polygon.simplify(tolerance) 消除碎共線點，避免 DXF 碎線段干擾
      B. 夾角限制：兩相鄰邊延伸後夾角須落於 [45°, 135°]，避免接近平行的假街角
      C. MBR 基準：以 minimum_rotated_rectangle 之長/短邊作為 S/l 校正基準
      D. 健全性檢查：交點離截角中點 > 50m 或截角面積 > 50㎡ → 跳過

    演算法（嚴格依多邊形頂點順序）：
      1. polygon.simplify(simplify_tolerance_m) 消除微碎點
      2. 遍歷所有邊，找出長度落於 [cutoff_min_m, cutoff_max_m] 之截角候選
      3. 對每條截角邊 c：取拓樸上「前一條邊」與「後一條邊」
      4. 計算前後邊夾角，若不在 [min_corner_angle_deg, max_corner_angle_deg] → 跳過
      5. 否則延伸求理論街角點，做健全性檢查 → 通過則記錄
      6. 真 S 長：兩端理論交點距離；側長同理
      7. MBR 基準：mbr_long_m / mbr_short_m 提供 UI 校正參考

    參數：
      polygon                       shapely Polygon
      cutoff_min_m, cutoff_max_m   截角邊長度判斷範圍（預設 2.0 ~ 8.0 m）
      sanity_max_corner_dist_m     交點離截角中點之合理距離上限（預設 50 m）
      sanity_max_cutoff_area_m2    截角三角形面積合理上限（預設 50 ㎡）
      simplify_tolerance_m          多邊形簡化容差（預設 0.1 m）
      min/max_corner_angle_deg     可接受街角夾角範圍（預設 45° ~ 135°）

    回傳 dict（新增 'mbr_long_m', 'mbr_short_m', 'mbr_polygon' 三鍵）
    """
    notes = []
    if polygon is None or polygon.is_empty:
        return {'classification': {'edges': [], 'front_idx': -1,
                                    'cutoff_idxs': [], 'side_idxs': []},
                'theoretical_corners': [],
                'cutoff_total_area': 0.0,
                'true_S_length': 0.0, 'true_side_length_left': 0.0,
                'true_side_length_right': 0.0,
                'mbr_long_m': 0.0, 'mbr_short_m': 0.0, 'mbr_polygon': None,
                'notes': ['多邊形無效或為空']}

    # 改良 A：多邊形簡化（容差小，僅消除碎共線點，不破壞主要形狀）
    original_area = polygon.area
    try:
        simplified = polygon.simplify(simplify_tolerance_m, preserve_topology=True)
        if (not simplified.is_empty
                and abs(simplified.area - original_area) / max(original_area, 1.0) < 0.05):
            n_before = len(list(polygon.exterior.coords))
            n_after = len(list(simplified.exterior.coords))
            if n_after < n_before:
                notes.append(
                    f"📐 多邊形簡化：頂點 {n_before-1} → {n_after-1}（容差 {simplify_tolerance_m}m，面積誤差 < 5%）"
                )
                polygon = simplified
    except Exception as _ex:
        notes.append(f"⚠️ 多邊形簡化失敗，使用原始多邊形：{_ex}")

    # 改良 C：MBR 基準（最小外接旋轉矩形，提供 S/l 校正參考）
    mbr_long = 0.0
    mbr_short = 0.0
    mbr_poly_geom = None
    try:
        mbr = polygon.minimum_rotated_rectangle
        mbr_poly_geom = mbr
        mbr_coords = list(mbr.exterior.coords)
        if mbr_coords[0] == mbr_coords[-1]:
            mbr_coords = mbr_coords[:-1]
        if len(mbr_coords) >= 4:
            mbr_edge_lens = []
            for i in range(4):
                p1 = mbr_coords[i]
                p2 = mbr_coords[(i + 1) % 4]
                mbr_edge_lens.append(((p2[0] - p1[0]) ** 2
                                      + (p2[1] - p1[1]) ** 2) ** 0.5)
            mbr_edge_lens.sort()
            mbr_short = round(mbr_edge_lens[0], 4)  # 寬（短邊）
            mbr_long = round(mbr_edge_lens[2], 4)   # 長（長邊）
            notes.append(f"📏 MBR 基準：長邊 {mbr_long:.2f} m、短邊 {mbr_short:.2f} m")
    except Exception as _ex:
        notes.append(f"⚠️ MBR 計算失敗：{_ex}")

    cls = classify_block_edges(polygon, cutoff_min_m, cutoff_max_m)
    if not cls['edges']:
        return {'classification': cls, 'theoretical_corners': [],
                'cutoff_total_area': 0.0,
                'true_S_length': 0.0, 'true_side_length_left': 0.0,
                'true_side_length_right': 0.0,
                'mbr_long_m': mbr_long, 'mbr_short_m': mbr_short,
                'mbr_polygon': mbr_poly_geom,
                'notes': notes + ['多邊形無有效邊界']}

    edges = cls['edges']
    n = len(edges)
    front_idx = cls['front_idx']
    front = edges[front_idx]

    # 🆕 W-D.1.3-a 切換：廢長度啟發法（原 cutoff_idxs_topo 長度區間 + 逐截角建 theoretical_corners）。
    #   截角改由 CAD-import 後之 post-pass `_rebuild_corners_topology` 拓樸定錨重建（單一真相源）；
    #   解析期先留空——true_S/true_side 為 vestigial（全檔零消費）、下方尾段吃空集合不炸（結果 0）。
    theoretical_corners = []
    cutoff_total_area = 0.0

    # 6. 真 S 長：若 front 兩端各有截角還原，取兩理論交點距離
    front_corners = []  # 與 front 邊相鄰之理論交點
    for tc in theoretical_corners:
        if tc['prev_idx'] == front_idx or tc['next_idx'] == front_idx:
            front_corners.append(tc['corner'])

    true_S = front['length']
    if len(front_corners) >= 2:
        c1, c2 = front_corners[0], front_corners[1]
        true_S = round(((c2[0] - c1[0]) ** 2 + (c2[1] - c1[1]) ** 2) ** 0.5, 4)
        notes.append(
            f"還原 S 長 = 兩端理論街角點距離 = {true_S:.2f} m（原 front 長 {front['length']:.2f} m）"
        )
    elif len(front_corners) == 1:
        c1 = front_corners[0]
        d1 = ((front['p1'][0] - c1[0]) ** 2 + (front['p1'][1] - c1[1]) ** 2) ** 0.5
        d2 = ((front['p2'][0] - c1[0]) ** 2 + (front['p2'][1] - c1[1]) ** 2) ** 0.5
        far_pt = front['p1'] if d1 > d2 else front['p2']
        true_S = round(((far_pt[0] - c1[0]) ** 2 + (far_pt[1] - c1[1]) ** 2) ** 0.5, 4)
        notes.append(f"還原 S 長 = 理論街角點到 front 對端頂點 = {true_S:.2f} m")

    # 側邊長：取與 front 共享一個截角之兩條邊（即真正的「左/右側」邊）
    # 對每條側邊，找它「兩端」共享頂點的所有理論交點：
    #   兩端皆有 → true_len = 兩交點距離（八邊形完整還原情況）
    #   僅一端有 → true_len = 從邊的「遠端原始頂點」延伸至該交點之距離
    side_pairs = []  # [(side_edge, [corners])]
    for tc in theoretical_corners:
        # 與 front 共享 idx 之截角
        if tc['prev_idx'] == front_idx:
            side_idx = tc['next_idx']
        elif tc['next_idx'] == front_idx:
            side_idx = tc['prev_idx']
        else:
            continue
        s_edge = edges[side_idx]
        # 找該側邊兩端所屬之全部理論交點（含此 tc 自己）
        corners_for_side = []
        for tc2 in theoretical_corners:
            if tc2['prev_idx'] == side_idx or tc2['next_idx'] == side_idx:
                corners_for_side.append(tc2['corner'])
        side_pairs.append((s_edge, corners_for_side))

    side_lengths = []
    for s_edge, corners in side_pairs:
        if len(corners) >= 2:
            c1, c2 = corners[0], corners[1]
            true_len = round(((c2[0] - c1[0]) ** 2
                              + (c2[1] - c1[1]) ** 2) ** 0.5, 4)
        elif len(corners) == 1:
            c = corners[0]
            d1 = ((s_edge['p1'][0] - c[0]) ** 2 + (s_edge['p1'][1] - c[1]) ** 2) ** 0.5
            d2 = ((s_edge['p2'][0] - c[0]) ** 2 + (s_edge['p2'][1] - c[1]) ** 2) ** 0.5
            far = s_edge['p1'] if d1 > d2 else s_edge['p2']
            true_len = round(((far[0] - c[0]) ** 2
                              + (far[1] - c[1]) ** 2) ** 0.5, 4)
        else:
            true_len = s_edge['length']
        side_lengths.append((s_edge, true_len))

    # 排序左右：以兩條邊中點 x 較小者為左側
    side_lengths.sort(key=lambda t: (t[0]['p1'][0] + t[0]['p2'][0]) / 2.0)
    true_side_left = side_lengths[0][1] if len(side_lengths) >= 1 else 0.0
    true_side_right = side_lengths[1][1] if len(side_lengths) >= 2 else 0.0

    # 改良 C：以 MBR 基準對 S 進行合理性校正
    # 若 true_S 與 mbr_long 偏差 > 50%（例如截角還原失誤），改採 MBR 長邊作為 S
    s_corrected_by_mbr = False
    if mbr_long > 0 and true_S > 0:
        ratio = true_S / mbr_long
        if ratio < 0.5 or ratio > 1.5:
            notes.append(
                f"⚠️ 還原 S 長 {true_S:.2f}m 與 MBR 長邊 {mbr_long:.2f}m 偏差 > 50% → 改採 MBR 長邊作為 S"
            )
            true_S = mbr_long
            s_corrected_by_mbr = True

    # 同理：若 true_side 偏離 MBR 短邊太多
    side_corrected_by_mbr = False
    if mbr_short > 0:
        for i, (s_edge, true_len) in enumerate(side_lengths):
            if true_len > 0:
                ratio = true_len / mbr_short
                if ratio < 0.5 or ratio > 1.5:
                    side_lengths[i] = (s_edge, mbr_short)
                    side_corrected_by_mbr = True
        if side_corrected_by_mbr:
            notes.append(f"⚠️ 部分側長偏離 MBR 短邊 {mbr_short:.2f}m > 50% → 已改採 MBR 短邊")
        true_side_left = side_lengths[0][1] if len(side_lengths) >= 1 else mbr_short
        true_side_right = side_lengths[1][1] if len(side_lengths) >= 2 else mbr_short

    return {
        'classification': cls,
        'theoretical_corners': theoretical_corners,
        'cutoff_total_area': round(cutoff_total_area, 4),
        'true_S_length': true_S,
        'true_side_length_left': true_side_left,
        'true_side_length_right': true_side_right,
        'mbr_long_m': mbr_long,
        'mbr_short_m': mbr_short,
        'mbr_polygon': mbr_poly_geom,
        's_corrected_by_mbr': s_corrected_by_mbr,
        'side_corrected_by_mbr': side_corrected_by_mbr,
        'notes': notes,
    }


def _anchor_chamfers_topology(edges, front_line, side_lines,
                              ang_tol_deg=2.0, perp_tol_m=1.0):
    """W-D.1.3-a §4/§4.1：以「共線＋投影重疊」配 FRONT/SIDE→BLOCK 邊，截角＝拓樸夾在
    FRONT-配對邊 與 SIDE-配對邊 之間的 BLOCK 邊（取代長度啟發法）。純函式、可單測。

    §4.1（實座標證）：CAD 的 FRONT/SIDE 線畫到「未截角尖角」、比 BLOCK 邊長約一截角腿
      → 端點重合法一端合一端差、六塊全滅。改共線：方位角 mod180 差 < ang_tol_deg、
      BLOCK 邊兩端到線垂距 < perp_tol_m、邊中點投影落線 span 內（2°/1m 僅捨入緩衝、非擬合參數）。
    line 支援 {'p1','p2'} 或 {'pts':[...]}（聚合線逐段套用；本案為直線、勿寫死單段）。

    edges       [{'idx','p1','p2',...}]（取自 geom_restore.classification.edges）
    front_line  {'p1','p2'} 或 {'pts':[...]} 或 None（f3_cad_front_lines[blk]）
    side_lines  {'left'/'right': {'p1','p2'|'pts'}}（f3_cad_side_lines_by_side[blk]；側別可缺）
    回傳 {'front_idx': int|None,
          'sides': {'left'/'right': {'side_idx','chamfer_idxs','status'} 或 None}}
      status：'截角斜邊'(1)/'無截角(共頂點)'(0)/'多段截角(n)'(≥2)/'配對失敗'。
    """
    import math as _m_anc

    def _segs(line):
        if not line:
            return []
        _pts = line.get('pts')
        if _pts and len(_pts) >= 2:
            return [(_pts[i], _pts[i + 1]) for i in range(len(_pts) - 1)]
        _p1 = line.get('p1'); _p2 = line.get('p2')
        return [(_p1, _p2)] if (_p1 and _p2) else []

    def _az(a, b):
        return _m_anc.degrees(_m_anc.atan2(float(b[1]) - float(a[1]),
                                           float(b[0]) - float(a[0]))) % 180.0

    def _perp(pt, a, b):
        dx = float(b[0]) - float(a[0]); dy = float(b[1]) - float(a[1])
        _L = _m_anc.hypot(dx, dy)
        if _L < 1e-9:
            return _m_anc.hypot(float(pt[0]) - float(a[0]), float(pt[1]) - float(a[1]))
        return abs(dx * (float(pt[1]) - float(a[1]))
                   - dy * (float(pt[0]) - float(a[0]))) / _L

    def _tproj(pt, a, b):
        dx = float(b[0]) - float(a[0]); dy = float(b[1]) - float(a[1])
        _L2 = dx * dx + dy * dy
        if _L2 < 1e-12:
            return 0.0
        return ((float(pt[0]) - float(a[0])) * dx
                + (float(pt[1]) - float(a[1])) * dy) / _L2

    def _edge_on_line(e, segs):
        """BLOCK 邊 e 共線＋投影重疊於 line 任一段 → (matched, edge_len)。"""
        _e1 = e.get('p1'); _e2 = e.get('p2')
        if not _e1 or not _e2:
            return (False, 0.0)
        _eaz = _az(_e1, _e2)
        _emid = ((float(_e1[0]) + float(_e2[0])) / 2.0,
                 (float(_e1[1]) + float(_e2[1])) / 2.0)
        _elen = _m_anc.hypot(float(_e2[0]) - float(_e1[0]),
                             float(_e2[1]) - float(_e1[1]))
        for (a, b) in segs:
            if not a or not b:
                continue
            _da = abs(_eaz - _az(a, b)); _da = min(_da, 180.0 - _da)
            if _da > ang_tol_deg:
                continue
            if _perp(_e1, a, b) > perp_tol_m or _perp(_e2, a, b) > perp_tol_m:
                continue
            if -0.05 <= _tproj(_emid, a, b) <= 1.05:
                return (True, _elen)
        return (False, 0.0)

    def _match_all(line):
        _segs_l = _segs(line)
        if not _segs_l or not edges:
            return []
        _out = []
        for e in edges:
            _ok, _ov = _edge_on_line(e, _segs_l)
            if _ok:
                _out.append((e.get('idx'), _ov))
        return _out

    n = len(edges or [])

    def _between(fi, si):
        """fi 與 si 之間（沿較短弧、不含兩端）之 BLOCK 邊 idx；無法定錨→None。"""
        if fi is None or si is None or fi == si or n == 0:
            return None
        fwd = []
        i = (fi + 1) % n
        _g = 0
        while i != si and _g < n:
            fwd.append(i); i = (i + 1) % n; _g += 1
        bwd = []
        i = (fi - 1) % n
        _g = 0
        while i != si and _g < n:
            bwd.append(i); i = (i - 1) % n; _g += 1
        return fwd if len(fwd) <= len(bwd) else bwd

    # FRONT：共線配對，取重疊最長之 BLOCK 邊為 front_idx（true_S/地基鏈用單一值）
    _fm = _match_all(front_line)
    front_idx = max(_fm, key=lambda x: x[1])[0] if _fm else None
    _fset = [i for i, _ in _fm]

    sides_out = {}
    for _sd in ('left', 'right'):
        _sl = (side_lines or {}).get(_sd)
        if not _sl:
            sides_out[_sd] = None
            continue
        _sm = _match_all(_sl)
        if not _sm or front_idx is None:
            sides_out[_sd] = {'side_idx': None, 'chamfer_idxs': [], 'status': '配對失敗'}
            continue
        # 截角＝(front_set × side_set) 中 _between 最短之弧（自然排除 front/side 邊本身）
        _best = None   # (弧長, chamfer_idxs, side_idx)
        for _fi in (_fset or [front_idx]):
            for _si, _ in _sm:
                _bw = _between(_fi, _si)
                if _bw is None:
                    continue
                if _best is None or len(_bw) < _best[0]:
                    _best = (len(_bw), _bw, _si)
        if _best is None:
            sides_out[_sd] = {'side_idx': None, 'chamfer_idxs': [], 'status': '配對失敗'}
            continue
        _ch = _best[1]
        if len(_ch) == 0:
            status = '無截角(共頂點)'
        elif len(_ch) == 1:
            status = '截角斜邊'
        else:
            status = f'多段截角({len(_ch)})'
        sides_out[_sd] = {'side_idx': _best[2], 'chamfer_idxs': _ch, 'status': status}
    return {'front_idx': front_idx, 'sides': sides_out}


def _rebuild_corners_topology(edges, front_line, side_lines):
    """W-D.1.3-a 切換：以拓樸定錨（`_anchor_chamfers_topology`）重建 theoretical_corners，取代長度啟發法。
    corner = FRONT_LINE ∩ SIDE_LINE(該側) 交點（線畫到未截角尖角；解 reviewer WARNING-2），
    平行/不相交退回 prev/next BLOCK 邊交點。回傳 [{cutoff_idx,prev_idx,next_idx,corner,cutoff_area_m2,
    corner_angle_deg}]（同 schema）；缺 edges/FRONT → None（呼叫端對街角街廓停機）。純函式、可單測（7/7）。
    """
    if not edges or not front_line:
        return None
    _topo = _anchor_chamfers_topology(edges, front_line, side_lines or {})
    if _topo.get('front_idx') is None:
        return None
    n = len(edges)
    _tcs = []
    for _sd in ('left', 'right'):
        _si = _topo['sides'].get(_sd)
        if not _si:
            continue
        _sl = (side_lines or {}).get(_sd) or {}
        for _c in (_si.get('chamfer_idxs') or []):
            if _c is None or _c >= n:
                continue
            _ce = edges[_c]
            _pi = (_c - 1) % n
            _ni = (_c + 1) % n
            # corner = FRONT ∩ SIDE(該側)；平行/不相交 → 退 prev/next BLOCK 邊交點
            _corner = line_intersection(front_line.get('p1'), front_line.get('p2'),
                                        _sl.get('p1'), _sl.get('p2'))
            if _corner is None:
                _corner = line_intersection(
                    edges[_pi].get('p1'), edges[_pi].get('p2'),
                    edges[_ni].get('p1'), edges[_ni].get('p2'))
            if _corner is None:
                continue
            _tcs.append({
                'cutoff_idx': _c, 'prev_idx': _pi, 'next_idx': _ni,
                'corner': _corner,
                'cutoff_area_m2': calc_cutoff_triangle_area(_ce, _corner),
                'corner_angle_deg': 0.0,
                'side': _sd,   # 🆕 W-D.1.3-c.1 側別出生入籍（正典：side_lines_by_side 該側；廢距-FRONT 重猜）
            })
    return _tcs


def _boundary_len_on_line(geom, la, lb, perp_tol=0.05, ang_tol_deg=2.0):
    """W-D.1.3-c：回傳 geom.boundary 落在直線 la–lb（線段 span 內）之總長。
    逐段共線判定（方位 mod180 差<ang_tol、兩端垂距<perp_tol）+ 投影重疊取長。
    自帶 helper（不抽 -a 凍結函式）；容差 perp_tol≥1cm（KL 備忘）。純函式、standalone 10/10。"""
    import math as _m_bl
    if geom is None or la is None or lb is None:
        return 0.0
    try:
        if geom.is_empty:
            return 0.0
    except Exception:
        return 0.0
    lax, lay = float(la[0]), float(la[1])
    lbx, lby = float(lb[0]), float(lb[1])
    ldx, ldy = lbx - lax, lby - lay
    _L2 = ldx * ldx + ldy * ldy
    if _L2 < 1e-12:
        return 0.0
    _L = _m_bl.sqrt(_L2)
    _line_az = _m_bl.degrees(_m_bl.atan2(ldy, ldx)) % 180.0

    def _perp_bl(px, py):
        return abs(ldx * (py - lay) - ldy * (px - lax)) / _L

    def _t_bl(px, py):
        return ((px - lax) * ldx + (py - lay) * ldy) / _L2

    def _rings_bl(g):
        _gt = getattr(g, 'geom_type', '')
        _out = []
        if _gt == 'Polygon':
            _out.append(list(g.exterior.coords))
            _out.extend(list(r.coords) for r in g.interiors)
        elif _gt in ('MultiPolygon', 'GeometryCollection', 'MultiLineString'):
            for _sub in g.geoms:
                _out.extend(_rings_bl(_sub))
        elif _gt == 'LineString':
            _out.append(list(g.coords))
        return _out

    _total = 0.0
    try:
        for _ring in _rings_bl(geom):
            for _i in range(len(_ring) - 1):
                _ax, _ay = _ring[_i][0], _ring[_i][1]
                _bx, _by = _ring[_i + 1][0], _ring[_i + 1][1]
                if _m_bl.hypot(_bx - _ax, _by - _ay) < 1e-9:
                    continue
                _sa = _m_bl.degrees(_m_bl.atan2(_by - _ay, _bx - _ax)) % 180.0
                _da = abs(_sa - _line_az); _da = min(_da, 180.0 - _da)
                if _da > ang_tol_deg:
                    continue
                if _perp_bl(_ax, _ay) > perp_tol or _perp_bl(_bx, _by) > perp_tol:
                    continue
                _ta, _tb = _t_bl(_ax, _ay), _t_bl(_bx, _by)
                _lo, _hi = max(min(_ta, _tb), 0.0), min(max(_ta, _tb), 1.0)
                if _hi > _lo:
                    _total += (_hi - _lo) * _L
    except Exception:
        return 0.0
    return _total


def _chamfer_line_for_side(blk_meta, which_side):
    """W-D.1.3-c.1：取該側截角斜邊 BLOCK 邊端點 (p1,p2)。
    側別直讀 tc['side']（_rebuild_corners_topology 出生入籍、來自 side_lines_by_side 正典）；
    廢除「距 FRONT 邊 p1/p2 重猜」（BLOCK 邊繞向≠語意左右→擲硬幣、五塊翻）。
    tc 無 side（舊快取）→ 不猜、None（呼叫端偵測 None+有 range 警示）。該側無截角 → None。"""
    try:
        _gr = (blk_meta or {}).get('geom_restore') or {}
        _cls = _gr.get('classification') or {}
        _edges = _cls.get('edges') or []
        _tcs = _gr.get('theoretical_corners') or []
        if not _tcs or not _edges:
            return None
        for _tc in _tcs:
            if _tc.get('side') != which_side:
                continue
            _ci = _tc.get('cutoff_idx', -1)
            if _ci < 0 or _ci >= len(_edges):
                continue
            _ep1 = _edges[_ci].get('p1'); _ep2 = _edges[_ci].get('p2')
            if not _ep1 or not _ep2:
                continue
            return (_ep1, _ep2)
        return None
    except (KeyError, IndexError, TypeError, ValueError, AttributeError):
        return None


def calc_B_value(general_burden_area: float, price_before: float, price_after: float,
                 total_area: float, preexisting_public_area: float) -> float:
    """B 值（一般負擔係數）— PDF P18, P21"""
    denom = float(price_after or 0.0) * (float(total_area or 0.0) - float(preexisting_public_area or 0.0))
    if denom <= 0:
        return 0.0
    return (float(general_burden_area or 0.0) * float(price_before or 0.0)) / denom


def calc_C_value(engineering_cost: float, redev_cost: float, loan_interest: float,
                 price_after: float, total_area: float, public_burden_total: float) -> float:
    """C 值（費用負擔係數）— PDF P20, P21"""
    denom = float(price_after or 0.0) * (float(total_area or 0.0) - float(public_burden_total or 0.0))
    if denom <= 0:
        return 0.0
    return (float(engineering_cost or 0.0) + float(redev_cost or 0.0) + float(loan_interest or 0.0)) / denom


def rw_from_width(W: float) -> float:
    """
    街角地側面道路負擔百分率（Rw %）— 依《市地重劃實施辦法》第 29 條附件二
    【104 年版市地重劃作業手冊 P.103 Rw 累積表】
      W (公尺)：0  1    2    3    4    5    6    7    8    9
      Rw (%)  ：0  17.4 24.4 31.3 37.8 44.0 50.0 55.7 61.1 66.3
      W (公尺)：10   11   12   13   14   15   16   17   18+
      Rw (%)  ：71.1 75.7 80.0 84.0 87.8 91.3 94.4 97.4 100
    中間值以線性內插。W ≥ 18m 視為 100%。
    """
    W = float(W or 0.0)
    if W <= 0:
        return 0.0
    table = [0.0, 17.4, 24.4, 31.3, 37.8, 44.0, 50.0, 55.7, 61.1, 66.3,
             71.1, 75.7, 80.0, 84.0, 87.8, 91.3, 94.4, 97.4, 100.0]
    if W >= 18.0:
        return 100.0
    i = int(W)
    frac = W - i
    return table[i] + (table[i + 1] - table[i]) * frac


# ─────────────────────────────────────────────────────────────────────
# W-C §3 / §0.5-B：Rw 累積差額制 + 軸向鎖定（法源：實施辦法第29條附件二）
# ─────────────────────────────────────────────────────────────────────
def rw_increment(W_prev: float, W_cur: float) -> float:
    """Rw 累積差額制（附件二 + v3.1 §6-3c）。
    本筆側街負擔比率 = (R(W_cur) − R(W_prev)) / 100，R = rw_from_width（累積表，W≥18→100）。
    clamp 內建於 rw_from_width → W_prev≥18 之後各筆自動為 0。
    側街負擔總量 = Σ 各筆 = (R(W_last)−R(0))/100 = 100%（telescoping，§3 主驗收）。
    負值（W 倒退，理論不應發生）夾為 0。
    """
    return max(0.0, (rw_from_width(W_cur) - rw_from_width(W_prev)) / 100.0)


def alloc_normal_axis(alloc_dir):
    """§0.5-B 軸向鎖定：由 f3_cad_alloc_dir（地界線/宗地分配線方向 (ux,uy)）
    導出『臨街向』單位向量 = rot90 = (-uy, ux)。

    此單一向量同時是：
      ① 餵 solve_G_binary/_block_strip 的 allocation_dir
         （_block_strip 內 n_hat = rot90(allocation_dir) = -(ux,uy) ∥ f3_cad_alloc_dir
          → 切出的地界線 ∥ 宗地分配線 ✓）。
      ② 附件二 W 量測軸（「向宗地分配線作垂直線」= 沿此臨街向量距）。
    缺值（無 ALLOC_LINE）回傳 None；呼叫端須警示、不得 fallback MBR（鐵律 G/H）。
    """
    import numpy as np
    if alloc_dir is None:
        return None
    try:
        ux, uy = float(alloc_dir[0]), float(alloc_dir[1])
    except (TypeError, ValueError, IndexError):
        return None
    n = (ux * ux + uy * uy) ** 0.5
    if n < 1e-9:
        return None
    return np.array([-uy / n, ux / n])


def _compute_block_depth_alloc(block_vertices, block_area, alloc_dir,
                               front_pts=None, baseline_pts=None):
    """W-C §5a：沿 ALLOC_LINE 量街廓代表深度 D_avg（廢 area/front_len 與 MBR）。
    座標：ux,uy = alloc_dir（深度向）；nx,ny = (-uy,ux)（臨街向 n_alloc）。
      to_wd(p) = (w = p·n_alloc 臨街位置, d = p·alloc 深度位置)。
    方法 A（剖面差值，主用）：FRONT/BASELINE 各頂點投影為 (w,d)，於重疊 w 範圍取樣
      depth(w)=|d_base(w)−d_front(w)|；D_avg=均值、D_min/D_max=極值（診斷）。
    方法 B（面積/臨街跨幅，交叉驗證 / A 缺資料時 fallback）：
      W_block = 街廓頂點 n_alloc 投影跨幅；D_avg_B = area / W_block。
    回傳 dict 或 None（缺 alloc_dir/面積/頂點）。
    """
    if alloc_dir is None or not block_area or block_area <= 0 or not block_vertices:
        return None
    ux, uy = float(alloc_dir[0]), float(alloc_dir[1])
    nrm = (ux * ux + uy * uy) ** 0.5
    if nrm < 1e-9:
        return None
    ux, uy = ux / nrm, uy / nrm
    nx, ny = -uy, ux
    # 方法 B（面積 / 臨街跨幅）
    proj_n = [p[0] * nx + p[1] * ny for p in block_vertices]
    W_block = (max(proj_n) - min(proj_n)) if proj_n else 0.0
    D_avg_B = (block_area / W_block) if W_block > 1e-6 else 0.0

    # 方法 A（剖面差值）
    D_avg_A = D_min_A = D_max_A = None
    def _wd(pts):
        return [(p[0] * nx + p[1] * ny, p[0] * ux + p[1] * uy) for p in pts]
    def _interp_d(seg, w):
        s = sorted(seg, key=lambda q: q[0])
        if w <= s[0][0]:
            return s[0][1]
        if w >= s[-1][0]:
            return s[-1][1]
        for i in range(len(s) - 1):
            w0, d0 = s[i]; w1, d1 = s[i + 1]
            if w0 <= w <= w1 and (w1 - w0) > 1e-9:
                return d0 + (d1 - d0) * (w - w0) / (w1 - w0)
        return s[-1][1]
    if front_pts and baseline_pts and len(front_pts) >= 2 and len(baseline_pts) >= 2:
        fw = _wd(front_pts); bw = _wd(baseline_pts)
        w_lo = max(min(p[0] for p in fw), min(p[0] for p in bw))
        w_hi = min(max(p[0] for p in fw), max(p[0] for p in bw))
        if w_hi - w_lo > 1e-6:
            depths = []
            N = 24
            for i in range(N + 1):
                w = w_lo + (w_hi - w_lo) * i / N
                depths.append(abs(_interp_d(bw, w) - _interp_d(fw, w)))
            if depths:
                D_avg_A = sum(depths) / len(depths)
                D_min_A = min(depths); D_max_A = max(depths)

    # 🆕 W-C §5a-1（實測修正）：深度「值」一律取方法 B（area÷n_alloc 跨幅，不碰 BASELINE）；
    #   方法 A（剖面差值）降為診斷（D_min/D_max + A/B 偏差），不參與值——A 主用會被合成
    #   BASELINE 角度污染（R2 曾算出 D_avg=500）。
    D_avg = D_avg_B
    if D_avg_A is not None and D_avg_A > 1e-6:
        method = 'B(值)/A(診斷)'
        _div = max(D_avg_B, 1e-6)
        note = ('A/B 接近' if abs(D_avg_A - D_avg_B) <= 0.15 * _div
                else f'⚠️ A({D_avg_A:.2f}) vs B({D_avg_B:.2f}) 差異大，街廓恐畫歪或 BASELINE 異常（值採 B）')
        D_min, D_max = D_min_A, D_max_A
    else:
        method = 'B(面積/臨街跨幅)'
        note = ('缺 FRONT/BASELINE 資料 → 僅方法 B' if not (front_pts and baseline_pts)
                else '方法 A 無重疊 w 範圍 → 僅方法 B')
        D_min = D_max = D_avg_B
    return {'D_avg': round(D_avg, 2), 'D_min': round(D_min, 2), 'D_max': round(D_max, 2),
            'D_avg_B': round(D_avg_B, 2), 'method': method, 'note': note}


def iterate_G_S(a: float, A: float, B: float, C: float,
                l_front: float, l_side: float,
                F: float, W: float, avg_depth: float,
                is_corner: bool = False,
                tab6_total_burden: float = None,
                W_prev: float = 0.0,
                # ─ 向後相容保留（W-C 後 W 改累積差額、不再幾何反推） ─
                block_poly=None, parcel_poly=None,
                allocation_dir=None, side_label: str = '左側',
                max_iter: int = 100, tol: float = 0.005) -> dict:
    """
    G 值 / S 值迭代（代數 fallback；solve_G_binary 失敗時用）。
    依《市地重劃實施辦法》第 29 條附件二公式。

    公式：G = [a·(1 − A·B) − Rw·F·l₁ − S·l₂] × (1 − C)
      其中 l₁ = 側面道路負擔尺度（配 Rw·F·l₁）
           l₂ = 正面道路負擔尺度（配 S·l₂）

    🆕 W-C §4 遞進化（取代舊「逐筆獨立、僅街角筆查絕對 Rw」）：
      - 攜帶 `W_prev`（前一筆累積 W）；本筆累積 W_cur = W_prev + S（長方形近似，臨街寬≈S）。
      - Rw 改累積差額：Rw = (R(W_cur) − R(W_prev))/100（rw_increment）。
      - 側街負擔對「角側全筆」皆套（gate 改 F>0 且 l_side>0，廢 is_corner 閘）。
      - 幾何 W 反推（_compute_W_from_S_geom）廢除（W 軸正確性由 solve_G_binary 負責；
        本函式僅長方形近似 fallback）。

    迭代規則：
      初始 S₀ = a·(1 − tb)/avg_depth
      第 k 輪：W_cur = W_prev + S_k；Rw = rw_increment(W_prev, W_cur)
              G_k = [a(1−A·B) − Rw·F·l_side − S_k·l_front]·(1 − C)
              S_(k+1) = G_k / avg_depth
      收斂：|ΔG| < tol 且 |ΔS| < tol
    """
    a = float(a or 0.0); A = float(A or 0.0); B = float(B or 0.0); C = float(C or 0.0)
    l_front = float(l_front or 0.0); l_side = float(l_side or 0.0)
    F = float(F or 0.0); avg_depth = float(avg_depth or 0.0)
    W_prev = float(W_prev or 0.0)

    if avg_depth <= 0 or a <= 0:
        return {'G': 0.0, 'S': 0.0, 'iterations': 0, 'converged': False,
                'trace': [], 'W': round(W_prev, 2), 'Rw_pct': 0.0}

    tb = float(tab6_total_burden) if tab6_total_burden is not None else float(B)
    tb = max(0.0, min(0.9, tb))

    _apply_rw = (F > 0.0 and l_side > 0.0)   # 角側筆才有側街負擔（廢 is_corner 閘）
    S = round(a * (1.0 - tb) / avg_depth, 2)
    W_cur = W_prev
    Rw_pct = 0.0

    trace = []
    last_G = None
    converged = False
    G = 0.0
    it = 0
    for it in range(1, max_iter + 1):
        # 本輪累積 W 與 Rw 差額（長方形近似：本筆臨街寬 ≈ S）
        W_cur = W_prev + S
        Rw = rw_increment(W_prev, W_cur) if _apply_rw else 0.0
        Rw_pct = Rw * 100.0

        # 計算本輪 G、S_new
        G_raw = (a * (1.0 - A * B)
                 - Rw * F * l_side
                 - S * l_front) * (1.0 - C)
        G = round(max(0.0, G_raw), 2)
        S_new = round(G / avg_depth, 2) if avg_depth > 0 else S

        trace.append({
            'step': it,
            'S_before': S, 'W_prev': round(W_prev, 2), 'W_after': round(W_cur, 2),
            'Rw_pct': round(Rw_pct, 2), 'G': G, 'S_after': S_new,
        })

        # 收斂檢查（G、S 穩定）
        if (last_G is not None
                and abs(round(G - last_G, 2)) < tol
                and abs(round(S_new - S, 2)) < tol):
            converged = True
            S = S_new
            break
        last_G = G
        S = S_new

    # 定案：累積 W = W_prev + 最終 S；Rw 用定案 S 重算
    W_cur = W_prev + S
    Rw_pct = (rw_increment(W_prev, W_cur) * 100.0) if _apply_rw else 0.0

    return {
        'G': round(G, 2), 'S': round(S, 2),
        'iterations': it, 'converged': converged, 'trace': trace,
        'W': round(W_cur, 2), 'W_far': round(W_cur, 2), 'Rw_pct': round(Rw_pct, 2),
    }


# ─────────────────────────────────────────────────────────────────────
# 幾何二分法（Binary Search）解 G／S — 升級版步驟 G 演算法
# 適用於非長方形、多邊形街廓；保留原法規公式參數，重寫尋找 S 的方式
# ─────────────────────────────────────────────────────────────────────
def _get_block_d_hat(block_poly, allocation_dir=None):
    """街廓分配線方向 d̂：若有指定 allocation_dir 用之，否則取 MBR 較長邊方向。"""
    import numpy as np
    if allocation_dir is not None:
        d = np.asarray(allocation_dir, dtype=float)
        return d / max(float(np.linalg.norm(d)), 1e-9)
    mbr = block_poly.minimum_rotated_rectangle
    mc = list(mbr.exterior.coords)
    edges = []
    for i in range(4):
        p0 = np.array(mc[i], dtype=float); p1 = np.array(mc[i + 1], dtype=float)
        edges.append((p0, p1, float(np.linalg.norm(p1 - p0))))
    edges.sort(key=lambda e: -e[2])
    dv = edges[0][1] - edges[0][0]
    return dv / max(float(np.linalg.norm(dv)), 1e-9)


def _compute_strip_width(cut_coords: list, d_hat=None) -> float:
    """
    🚨 Patch B-2：從分配條 polygon 之 cut_coords 計算實際寬度。

    若提供 d_hat：寬度 = polygon 在 d_hat「推進方向」上的 1D 投影長度
                 = max(proj) - min(proj)
    若未提供 d_hat：fallback 至 MBR 短邊（最小外接矩形之較短邊長度）

    參數：
      cut_coords  分配條 polygon 之頂點座標 list[[x, y]] 或 list[(x, y)]
      d_hat       推進方向單位向量（np.array shape=(2,) 或 list[float]）

    回傳：strip width (m)；若無法計算回傳 0.0
    """
    import numpy as _np_pw
    if not cut_coords or len(cut_coords) < 3:
        return 0.0
    try:
        if d_hat is not None:
            d_vec = _np_pw.asarray(d_hat, dtype=float)
            d_norm = float(_np_pw.linalg.norm(d_vec))
            if d_norm < 1e-9:
                return 0.0
            d_unit = d_vec / d_norm
            projs = []
            for pt in cut_coords:
                x = float(pt[0]); y = float(pt[1])
                projs.append(float(_np_pw.dot([x, y], d_unit)))
            return float(max(projs) - min(projs))
        # Fallback：MBR 短邊
        from shapely.geometry import Polygon as _SP_pw
        try:
            poly = _SP_pw(cut_coords)
            if not poly.is_valid:
                poly = poly.buffer(0)
            if poly.is_empty:
                return 0.0
            mbr = poly.minimum_rotated_rectangle
            mc = list(mbr.exterior.coords)
            edges = []
            for i in range(len(mc) - 1):
                p0 = mc[i]; p1 = mc[i + 1]
                edges.append(((p1[0] - p0[0]) ** 2 + (p1[1] - p0[1]) ** 2) ** 0.5)
            return float(min(edges)) if edges else 0.0
        except Exception:
            return 0.0
    except Exception:
        return 0.0


def _safe_num_e1(v) -> float:
    """
    🚨 Patch E-1 §6 helper：數值欄位防呆轉 float

    sb_row 內某些欄位（例如「【左】截角(㎡)」、「【右】截角(㎡)」）
    在無臨路時會被寫成字串 '—'（非數字）。直接呼叫 float('—') 會拋 ValueError。
    本 helper 統一把 None/空字串/'—'/不可轉型 一律當 0.0 處理。
    """
    try:
        if v is None or v == '' or v == '—':
            return 0.0
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _find_block_corner(block_poly, d_hat, side_label='左側'):
    """
    找街廓「正面街 ∩ 側街」的角點，並校正 d_hat 方向（從角點指向街廓內部）。
    回傳 (corner_pt, d_hat_corrected) 或 None。
    """
    import numpy as np
    Cb = np.array([block_poly.centroid.x, block_poly.centroid.y])
    coords = list(block_poly.exterior.coords)
    side_edges, front_edges = [], []
    for i in range(len(coords) - 1):
        e0 = np.array(coords[i], dtype=float)
        e1 = np.array(coords[i + 1], dtype=float)
        ev = e1 - e0
        elen = float(np.linalg.norm(ev))
        if elen < 0.05:
            continue
        e_hat = ev / elen
        par_score = abs(float(np.dot(e_hat, d_hat)))
        if par_score < 0.3:
            side_edges.append((e0, e1))
        elif par_score > 0.7:
            front_edges.append((e0, e1))
    if not side_edges or not front_edges:
        return None
    # 依 side_label 篩選對應側邊
    scored = [(e0, e1, float(np.dot((e0 + e1) / 2 - Cb, d_hat))) for (e0, e1) in side_edges]
    sel = (max if side_label == '右側' else min)(scored, key=lambda r: r[2])
    side_pts = [sel[0], sel[1]]
    # 找最近正面街邊端點 → 角點
    best_corner = None; best_d = float('inf')
    for fe0, fe1 in front_edges:
        for fp in [fe0, fe1]:
            for sp in side_pts:
                d = float(np.linalg.norm(fp - sp))
                if d < best_d:
                    best_d = d
                    best_corner = (fp + sp) / 2.0
    if best_corner is None:
        return None
    if best_d > 1.0:
        # 取側邊中靠正面那端
        best_corner = side_pts[0]
    # 校正 d_hat：從角點指向街廓內部（與最遠正面街端點同向）
    far_fp = None; far_d = -1.0
    for fe0, fe1 in front_edges:
        for fp in [fe0, fe1]:
            d = float(np.linalg.norm(fp - best_corner))
            if d > far_d:
                far_d = d; far_fp = fp
    if far_fp is not None and float(np.dot(far_fp - best_corner, d_hat)) < 0:
        d_hat = -d_hat
    return (best_corner, d_hat)


def _block_max_S(block_poly, corner_pt, d_hat):
    """街廓沿 d_hat 從角點起算之最大可能投影距離（= 街廓在 d_hat 方向之長度上限）。"""
    import numpy as np
    coords = list(block_poly.exterior.coords)
    projs = [float(np.dot(np.array(c, dtype=float) - corner_pt, d_hat)) for c in coords]
    return max(0.1, max(projs))


def _split_cutoffs_by_side(blk_meta) -> tuple:
    """
    🆕 W-D.1.3-d：將街廓總截角面積依 tc['side'] 拆分為左右兩側。

    側別直讀 tc['side']（_rebuild_corners_topology 出生入籍、正典＝side_lines_by_side）；
    廢除舊「corner 相對質心 × d_hat 法向點積猜側」（繞向近似＝擲硬幣、診斷會說謊、
    與 -c.1 修掉之猜側同族）。tc 無 side（舊快取）→ 該側不計（不猜）。

    回傳：(cutoff_left_m2, cutoff_right_m2) — 浮點數 tuple
    """
    geom_restore = blk_meta.get('geom_restore') or {}
    theoretical_corners = geom_restore.get('theoretical_corners') or []
    if not theoretical_corners:
        # fallback: 對半切（向後相容；當無 per-corner data 時）
        total = float(blk_meta.get('cutoff_total_area_m2', 0.0) or 0.0)
        return total / 2.0, total / 2.0

    verts = blk_meta.get('vertices') or []
    if len(verts) < 3:
        return 0.0, 0.0

    cl = 0.0; cr = 0.0
    for tc in theoretical_corners:
        area = float(tc.get('cutoff_area_m2', 0.0) or 0.0)
        if area <= 0:
            continue
        _sd = tc.get('side')
        if _sd == 'left':
            cl += area
        elif _sd == 'right':
            cr += area
    return cl, cr


def _resolve_clicked_parcel(cx, cy, temp_parcels, tolerance: float = 0.5):
    """
    🆕 Phase 4 補充 3：點擊解析含 0.5m 容差，避免邊界點選漏接。

    回傳被點擊到的 parcel dict；同樣容差內取面積最大者（避免小碎片優先）。
    """
    from shapely.geometry import Point as _SP, Polygon as _SPoly
    if cx is None or cy is None:
        return None
    try:
        pt = _SP(float(cx), float(cy))
    except Exception:
        return None
    pt_buf = pt.buffer(tolerance)
    candidates = []
    for tp in (temp_parcels or []):
        coords = tp.get('polygon_coords') or []
        if len(coords) < 3:
            continue
        try:
            poly = _SPoly(coords)
            if not poly.is_valid:
                poly = poly.buffer(0)
            if poly.contains(pt) or pt_buf.intersects(poly) or poly.distance(pt) < tolerance:
                candidates.append((tp, poly.area))
        except Exception:
            continue
    if not candidates:
        return None
    candidates.sort(key=lambda x: -x[1])
    return candidates[0][0]


def _is_road_like_category(category: str) -> bool:
    """
    🆕 Phase 5 Task K：判斷 category 是否屬於「道路類」街廓（寬鬆比對）

    解決硬寫死 ('道路', '溝渠') 過於嚴格的問題；
    支援「市區道路」「縣道」「都市計畫道路」「人行道」「水溝」等變體名稱。
    """
    if not category:
        return False
    cat = str(category)
    return ('道' in cat) or ('路' in cat) or ('溝' in cat) or ('渠' in cat)


def _resolve_clicked_block(cx, cy, classified_blocks, tolerance: float = 2.0,
                            category_filter: tuple = None,
                            category_predicate=None):
    """
    🆕 Phase 4 補充 4 + Phase 5 Task K：點擊位置自動鎖定街廓。

    參數：
      category_filter      只考慮指定 category 的 block（精確比對；如 ('道路', '溝渠')）
      category_predicate   🆕 Phase 5：自訂 callable(category) → bool（彈性，優於 filter）
                           若兩者皆提供，優先採 predicate
      tolerance            街廓邊界容差（預設 2m，較寬鬆）

    回傳被點擊到的街廓 dict；面積大者優先。
    """
    from shapely.geometry import Point as _SP, Polygon as _SPoly
    if cx is None or cy is None:
        return None
    try:
        pt = _SP(float(cx), float(cy))
    except Exception:
        return None
    candidates = []
    for b in (classified_blocks or []):
        cat = b.get('category', '')
        # Phase 5：predicate 優先，不行再 fallback 至 filter tuple
        if category_predicate is not None:
            if not category_predicate(cat):
                continue
        elif category_filter and cat not in category_filter:
            continue
        verts = b.get('vertices') or []
        if len(verts) < 3:
            continue
        try:
            poly = _SPoly(verts)
            if not poly.is_valid:
                poly = poly.buffer(0)
            if poly.contains(pt) or poly.distance(pt) < tolerance:
                candidates.append((b, poly.area))
        except Exception:
            continue
    if not candidates:
        return None
    candidates.sort(key=lambda x: -x[1])
    return candidates[0][0]


def _offset_edge_to_centerline(road_label: str, edge_p1: tuple, edge_p2: tuple,
                                width_m: float, classified_blocks: list) -> dict:
    """
    🆕 Phase 6 Task N：CAD 級邊界平行偏移（Parallel Offset）生成精準中心線。

    演算法：
      1. 構造邊界線段 LineString(p1, p2)
      2. 找到對應道路 polygon
      3. 以 W/2 距離平移該線段，自動判斷「道路內側」方向：
         - 取邊中點 + 兩側法向量試探點 → 含於道路 polygon 者為內側
         - 三層 fallback：兩側試探 → 兩側皆內側用 centroid → 兩側皆外側用 centroid 方向
      4. 平移後線段 = 道路中心線（數學精確）
      5. 將中心線兩端適度延伸至道路 polygon 邊界（避免短於道路）

    參數：
      road_label              道路 label
      edge_p1, edge_p2        被點擊邊界之兩端點
      width_m                 計畫道路全寬 W
      classified_blocks       街廓清單

    回傳：
      {'centerline_pts': [(x, y), ...], 'offset_distance': W/2}
      或 None（失敗）
    """
    from shapely.geometry import LineString as _SL, Point as _SP, Polygon as _SP_poly
    import numpy as _np
    if width_m <= 0 or not classified_blocks:
        return None
    # 1. 找到道路 polygon
    road_blk = next((b for b in classified_blocks
                     if b.get('label') == road_label), None)
    if road_blk is None:
        return None
    verts = road_blk.get('vertices') or []
    if len(verts) < 3:
        return None
    try:
        road_poly = _SP_poly(verts)
        if not road_poly.is_valid:
            road_poly = road_poly.buffer(0)
    except Exception:
        return None
    # 2. 構造邊界 + 法向量
    p1 = _np.array([edge_p1[0], edge_p1[1]], dtype=float)
    p2 = _np.array([edge_p2[0], edge_p2[1]], dtype=float)
    edge_vec = p2 - p1
    edge_len = float(_np.linalg.norm(edge_vec))
    if edge_len < 1e-6:
        return None
    edge_dir = edge_vec / edge_len
    n_left = _np.array([-edge_dir[1], edge_dir[0]])
    n_right = -n_left
    half_w = width_m / 2.0
    # 3. 試探：邊中點 + 法向量 0.5m 兩側測試哪側在道路內
    mid = (p1 + p2) / 2.0
    test_left = mid + 0.5 * n_left
    test_right = mid + 0.5 * n_right
    pt_l = _SP(float(test_left[0]), float(test_left[1]))
    pt_r = _SP(float(test_right[0]), float(test_right[1]))
    try:
        inside_left = bool(road_poly.contains(pt_l)) or float(road_poly.distance(pt_l)) < 0.1
        inside_right = bool(road_poly.contains(pt_r)) or float(road_poly.distance(pt_r)) < 0.1
    except Exception:
        inside_left = False
        inside_right = False
    if inside_left and not inside_right:
        n_in = n_left
    elif inside_right and not inside_left:
        n_in = n_right
    elif inside_left and inside_right:
        # 兩側皆視為內側（罕見） → 採離道路 centroid 較近的方向
        try:
            cen = road_poly.centroid
            d_l = ((test_left[0]-cen.x)**2 + (test_left[1]-cen.y)**2) ** 0.5
            d_r = ((test_right[0]-cen.x)**2 + (test_right[1]-cen.y)**2) ** 0.5
            n_in = n_left if d_l < d_r else n_right
        except Exception:
            n_in = n_left
    else:
        # 兩側皆不在道路內 → fallback 至 centroid 方向
        try:
            cen = road_poly.centroid
            v_to_cen = _np.array([cen.x - mid[0], cen.y - mid[1]])
            n_in = n_left if float(_np.dot(v_to_cen, n_left)) > 0 else n_right
        except Exception:
            n_in = n_left
    # 4. 平移生成中心線（W/2 距離）
    cp1 = p1 + half_w * n_in
    cp2 = p2 + half_w * n_in
    # 5. 兩端延伸至道路 polygon 邊界（讓中心線完整覆蓋道路長度）
    try:
        bnd = road_poly.bounds
        diag = ((bnd[2]-bnd[0])**2 + (bnd[3]-bnd[1])**2) ** 0.5
    except Exception:
        diag = max(edge_len * 2.0, 100.0)
    cp1_ext = cp1 - diag * edge_dir
    cp2_ext = cp2 + diag * edge_dir
    try:
        long_line = _SL([(cp1_ext[0], cp1_ext[1]), (cp2_ext[0], cp2_ext[1])])
        # 🆕 Phase 7 Module 2 (M2-3)：buffer(0.05) 微膨脹防呆，避免 TopologyException
        clipped = long_line.intersection(road_poly.buffer(0.05))
        if clipped.is_empty:
            return {
                'centerline_pts': [(float(cp1[0]), float(cp1[1])),
                                    (float(cp2[0]), float(cp2[1]))],
                'offset_distance': half_w,
            }
        if clipped.geom_type == 'LineString':
            pts = list(clipped.coords)
        elif clipped.geom_type == 'MultiLineString':
            longest = max(clipped.geoms, key=lambda ls: ls.length)
            pts = list(longest.coords)
        else:
            return None
        return {
            'centerline_pts': [(float(p[0]), float(p[1])) for p in pts],
            'offset_distance': half_w,
        }
    except Exception:
        return {
            'centerline_pts': [(float(cp1[0]), float(cp1[1])),
                                (float(cp2[0]), float(cp2[1]))],
            'offset_distance': half_w,
        }


def _snap_baseline_to_block_edge(cx, cy, blk_meta, max_dist: float = 15.0):
    """
    🆕 Phase 5 Task J：將點擊位置 (cx, cy) 捕捉到街廓最近邊界，回傳基準線設定 dict。

    取代舊「兩點空中拾取」UX — 使用者僅需點擊一次靠近某邊，
    系統自動以該邊為分配基準線（point + d_hat）。

    回傳：
      {'point': (x, y), 'angle_deg': float, 'edge_idx': int, 'edge_dist_m': float}
      或 None（若距所有邊都 > max_dist 或街廓無效）

    演算法：
      1. 取街廓 exterior 所有相鄰頂點對 (vi, vi+1) 作為線段
      2. 計算 Point(cx, cy) 到每條線段的最短距離
      3. 取距離最小者
      4. baseline.point = vi（第一個端點，依 polygon 頂點順序）
      5. d_hat = (vi+1 - vi) 之單位向量
      6. angle_deg = math.degrees(math.atan2(dy, dx))
    """
    from shapely.geometry import LineString as _SL, Point as _SP
    import math as _math
    verts = blk_meta.get('vertices') or []
    if len(verts) < 3 or cx is None or cy is None:
        return None
    try:
        pt = _SP(float(cx), float(cy))
    except Exception:
        return None
    best = None
    for i in range(len(verts)):
        p1 = verts[i]
        p2 = verts[(i + 1) % len(verts)]
        try:
            seg = _SL([tuple(p1), tuple(p2)])
            d = pt.distance(seg)
            if best is None or d < best['dist']:
                best = {'i': i, 'p1': p1, 'p2': p2, 'dist': d}
        except Exception:
            continue
    if best is None or best['dist'] > max_dist:
        return None
    p1 = best['p1']; p2 = best['p2']
    dx = p2[0] - p1[0]
    dy = p2[1] - p1[1]
    seg_len = _math.hypot(dx, dy)
    if seg_len < 1e-6:
        return None
    angle = _math.degrees(_math.atan2(dy, dx))
    return {
        'point': (float(p1[0]), float(p1[1])),
        'angle_deg': float(angle),
        'edge_idx': int(best['i']),
        'edge_dist_m': float(best['dist']),
    }


def _build_block_front_line(blk_meta):
    """以街廓 MBR 最長邊作為「正面道路 LineString」近似"""
    from shapely.geometry import LineString as _SL
    verts = blk_meta.get('vertices') or []
    if len(verts) < 3:
        return None
    try:
        longest = None; max_L = 0.0
        for i in range(len(verts)):
            p1 = verts[i]
            p2 = verts[(i + 1) % len(verts)]
            L = ((p2[0] - p1[0]) ** 2 + (p2[1] - p1[1]) ** 2) ** 0.5
            if L > max_L:
                max_L = L
                longest = _SL([tuple(p1), tuple(p2)])
        return longest
    except Exception:
        return None


def _auto_detect_corner_side(parcel, blk_meta):
    """
    🆕 Phase 4 H-1：自動偵測街角地左/右側別

    依 parcel centroid 在街廓正面道路 LineString 上的投影距離判斷：
      投影距離 < 街廓正面長度 / 2 → '左側'
      否則                       → '右側'

    回傳 '左側' / '右側' / '' （失敗時）
    """
    coords = parcel.get('polygon_coords') or []
    if not coords:
        return ''
    try:
        cx = sum(c[0] for c in coords) / len(coords)
        cy = sum(c[1] for c in coords) / len(coords)
    except Exception:
        return ''
    front_line = _build_block_front_line(blk_meta)
    if front_line is None:
        return ''
    try:
        return assign_corner_side_by_projection((cx, cy), front_line)
    except Exception:
        return ''



def _projection_order(parcels, front_line_p1, front_line_p2) -> list:
    """
    🆕 W-D.2（§2 正典單一真相源）：重劃前原位次＝宗地沿 FRONT_LINE 之投影順序
    （representative_point 投影、由 p1 端 → p2 端；穩定排序）。

    自 _spatial_order_parcels_v2 抽出成具名純函式，供 v2（位次排序）與 W-D.2 §3
    滑池槽同源消費。PK tiebreaker 之「距角序·暫行」換吃本函式＝baselines v2 轉正
    時刻一併做（v1 凍結期不動，見 TODO(W-D §2) 標記）。
    FRONT_LINE 缺或單筆投影失敗 → 該筆以 0.0 計（維持輸入相對序，穩定排序保證）。
    """
    from shapely.geometry import LineString as _SL_po, Point as _SP_po, Polygon as _P_po
    try:
        line = _SL_po([tuple(front_line_p1), tuple(front_line_p2)])
    except Exception:
        line = None

    def _proj_of(tp):
        cs = tp.get('polygon_coords') or []
        if len(cs) < 3 or line is None:
            return 0.0
        try:
            cen = _P_po(cs).representative_point()
            return float(line.project(_SP_po(cen.x, cen.y)))
        except Exception:
            return 0.0

    return sorted(parcels, key=_proj_of)


def _select_pool_slot(widths, left_side, right_side, rw_func=None) -> dict:
    """
    🆕 W-D.2 §3：滑池槽選位（W-D_細部plan §3.1 STEP1-4；純函式、無 st/session）。

    目標：選池槽 k（左群＝前 k 筆、右群＝其餘、池插中間），使兩側私有地實扛之
    側街負擔加權總和最大：J(k) = ΣRw_L·F_L·l1_L ＋ ΣRw_R·F_R·l1_R（負擔「面積」加權）。
    ΣRw_側 = R(b＋Σw_群) − R(b)（telescoping；R=Rw 累積表、W≥18 飽和 100%）。
    J 平手（ε=1e-6·J* 內）→ 取最靠中央（dev=|Σw_左−Σw_右| 最小；KL 裁示 tie-break）、
    再平手取小 k（序列穩定可重現）。pinned 街角地不可被池頂掉（K 域內建）。

    參數：
      widths       [w_1..w_n] 原位次序（左→右）單筆宗地寬度 w_i（⊥ALLOC、W 軸單位）
      left_side / right_side：{'has': bool, 'F': float, 'l1': float, 'b': float}
        has＝該端臨側街；F＝該側 SIDE_LINE 長；l1＝側街負擔尺度；b＝forced 起始 W（無則 0）
      rw_func      R(W) 累積函數（%）；預設 rw_from_width

    回傳 {'k', 'J_star', 'table': [{'k','J','dev','ΣRw_L','ΣRw_R'}...], 'note'}
    （table＝STEP4 診斷輸出，供 reviewer/KL 逐槽複核。）
    """
    R = rw_func if rw_func is not None else rw_from_width
    w = [float(x or 0.0) for x in (widths or [])]
    n = len(w)
    _L = left_side or {}
    _Rt = right_side or {}
    has_L = bool(_L.get('has'))
    has_R = bool(_Rt.get('has'))
    F_L = float(_L.get('F', 0.0) or 0.0)
    l1_L = float(_L.get('l1', 0.0) or 0.0)
    F_R = float(_Rt.get('F', 0.0) or 0.0)
    l1_R = float(_Rt.get('l1', 0.0) or 0.0)
    b_L = float(_L.get('b', 0.0) or 0.0)
    b_R = float(_Rt.get('b', 0.0) or 0.0)
    note = ''
    # STEP 1：候選槽 K（有左街角 → 左群至少含 p_1；有右街角 → 右群至少含 p_n）
    k_min = 1 if has_L else 0
    k_max = (n - 1) if has_R else n
    if k_min > k_max:
        # 退化（如 n=1 且雙側 pin 衝突）：無合法槽 → 全域比較、具名註記（不靜默）
        K = list(range(0, n + 1))
        note = f'degenerate：n={n} 雙側 pin 衝突無合法槽，取全域 argmax'
    else:
        K = list(range(k_min, k_max + 1))
    # STEP 2：逐槽 J(k) 與偏離 dev(k)
    total = sum(w)
    table = []
    for k in K:
        sL = sum(w[:k])
        sR = total - sL
        rw_L = (R(b_L + sL) - R(b_L)) if has_L else 0.0
        rw_R = (R(b_R + sR) - R(b_R)) if has_R else 0.0
        table.append({
            'k': k,
            # J 單位＝負擔「面積」㎡（Rw 化比率 ×F(m)×l₁；KL 量級裁示 2026-07-05：
            #   如 R2 全飽和 J≈1.00×44.94×1≈45）。等權 F=l1=1 時 ×100 即手冊 %表述（180/150）。
            'J': (rw_L / 100.0) * F_L * l1_L + (rw_R / 100.0) * F_R * l1_R,
            'dev': abs(sL - sR),
            'ΣRw_L': rw_L,
            'ΣRw_R': rw_R,
        })
    # STEP 3：J 最大為主；ε 內平手 → dev 最小（最靠中央）→ 再取小 k
    J_star = max(t['J'] for t in table)
    eps = (1e-6 * J_star) if J_star > 0 else 0.0
    cand = [t for t in table if t['J'] >= J_star - eps]
    pick = min(cand, key=lambda t: (t['dev'], t['k']))
    return {'k': pick['k'], 'J_star': J_star, 'table': table, 'note': note}


def _spatial_order_parcels_v2(parcels_in_block,
                                d_hat,
                                front_line_p1,
                                front_line_p2,
                                pk_winners: dict = None,
                                forced_offset: dict = None) -> dict:
    """
    🚨 Patch D-2：Minimal-Disruption 原位次排序（最小擾動 swap）

    與 v1 的關鍵差異：
      v1：把街角地永遠塞兩端，其他依 proj 排序 → 整列順序與「原投影序列」可能差異很大
      v2：保留原投影序列；PK winner 取街角，被擠下的「原街角實體」退一格，其餘全部位次不動

    例：R6 投影序列 = [628-38(1), 628-18(1), 628-7(1), 628-20(1), 628-4(1), 628-21(1), 628-1(2)]
        PK winner = 628-18(1)
        v2 結果：[628-18(1), 628-38(1), 628-7(1), 628-20(1), 628-4(1), 628-21(1), 628-1(2)]
                  ↑ winner 進位次 1，其他保持原序

    參數：
      parcels_in_block      buildable parcels in this Ri
      d_hat                 推進方向向量
      front_line_p1, p2     FRONT_LINE 兩端點（p1 = 右端, p2 = 左端 by convention）
      pk_winners            {'p1_end': 暫編_R, 'p2_end': 暫編_L}（None 表示無 winner）
      forced_offset         {'left_forced_offset', 'right_forced_offset',
                             'left_corner_min_area', 'right_corner_min_area'}
    回傳：
      {
        'ordered': [{'tp': ..., 'pre_position': i, 'is_corner_winner': bool,
                     'side': '右街角'/'左街角'/'中段'}, ...]
                    從位次 1（右端）→ 位次 N（左端）
        'right_corner_offset_area': float,
        'left_corner_offset_area':  float,
      }
    """
    if not parcels_in_block:
        return {'ordered': [], 'right_corner_offset_area': 0.0,
                'left_corner_offset_area': 0.0}

    pk_winners = pk_winners or {}
    forced_offset = forced_offset or {}

    # ── 投影序列：正典原位次（🆕 W-D.2 §2 單一真相源 _projection_order；由 p1 端 → p2 端）
    pre_seq = _projection_order(parcels_in_block, front_line_p1, front_line_p2)
    pre_seq_meta = [
        {'tp': tp, 'pre_position': i + 1,
         'is_corner_winner': False, 'side': '中段'}
        for i, tp in enumerate(pre_seq)
    ]

    def _is_forced_offset_winner(winner):
        if winner is None:
            return False
        if isinstance(winner, str) and winner.startswith('FORCED_OFFSET'):
            return True
        return False

    # ── 處理右側 winner（p1_end = 位次 1）
    pk_r = pk_winners.get('p1_end')
    if pk_r and not _is_forced_offset_winner(pk_r):
        idx_r = next((i for i, e in enumerate(pre_seq_meta)
                      if e['tp'].get('暫編地號') == pk_r), None)
        if idx_r is not None:
            if idx_r != 0:
                entry = pre_seq_meta.pop(idx_r)
                entry['is_corner_winner'] = True
                entry['side'] = '右街角'
                pre_seq_meta.insert(0, entry)
            else:
                pre_seq_meta[0]['is_corner_winner'] = True
                pre_seq_meta[0]['side'] = '右街角'

    # ── 處理左側 winner（p2_end = 位次 N）
    pk_l = pk_winners.get('p2_end')
    if pk_l and not _is_forced_offset_winner(pk_l):
        idx_l = next((i for i, e in enumerate(pre_seq_meta)
                      if e['tp'].get('暫編地號') == pk_l), None)
        if idx_l is not None:
            last_i = len(pre_seq_meta) - 1
            if idx_l != last_i:
                entry = pre_seq_meta.pop(idx_l)
                entry['is_corner_winner'] = True
                entry['side'] = '左街角'
                pre_seq_meta.append(entry)
            else:
                pre_seq_meta[-1]['is_corner_winner'] = True
                pre_seq_meta[-1]['side'] = '左街角'

    # 🚨 Patch E-2.1：標記首尾 corner_winner 讓下游 first_corner 判定簡化
    if pre_seq_meta:
        pre_seq_meta[0]['is_first_corner_marker'] = bool(
            pre_seq_meta[0].get('is_corner_winner', False))
        pre_seq_meta[-1]['is_first_corner_marker'] = bool(
            pre_seq_meta[-1].get('is_corner_winner', False))

    return {
        'ordered': pre_seq_meta,
        'right_corner_offset_area':
            float(forced_offset.get('right_corner_min_area', 0) or 0)
            if forced_offset.get('right_forced_offset') else 0.0,
        'left_corner_offset_area':
            float(forced_offset.get('left_corner_min_area', 0) or 0)
            if forced_offset.get('left_forced_offset') else 0.0,
    }


def _extend_boundary_through_road(road_poly, adjacent_blocks: list,
                                   min_piece_area: float = 5.0,
                                   min_compactness: float = 0.15) -> list:
    """
    🆕 Task B（深度重構）：透過相鄰街廓邊界線延伸穿越道路，產生精確中心線。

    適用情境：T 字路口、十字路口、多街廓交會處 — 既有「凸殼比 + 反射頂點」演算法
    對複雜路口可能產生不精確中心線；本函式利用相鄰街廓的實際邊界資訊改善。

    演算法：
      1. 對每個相鄰街廓，找與道路邊界 1m 內共享的邊
      2. 將共享邊延伸穿越道路（長度 = 道路 bounds 對角線 × 1.5）
      3. 延伸線 ∩ road_poly 取得切割線
      4. polygonize() 將道路分割為子多邊形
      5. 形狀檢核：丟棄 < 5㎡ 或緊湊度 < 0.15 的畸形碎片
      6. 各合格子矩形呼叫 _mbr_centerline_of() 取中心線

    參數：
      road_poly         shapely Polygon（道路多邊形）
      adjacent_blocks   list[{'label': str, 'shapely': Polygon}]（相鄰街廓）
      min_piece_area    過濾門檻：< 此面積 (㎡) 視為碎片 → 丟棄
      min_compactness   過濾門檻：4π·area/perimeter² < 此值 → 視為狹長 → 丟棄

    回傳：list[LineString]；失敗時回傳空 list（呼叫端應 fallback 至既有演算法）
    """
    from shapely.geometry import LineString, MultiLineString
    from shapely.ops import polygonize, unary_union
    import math as _math

    if road_poly is None or road_poly.is_empty or not adjacent_blocks:
        return []

    try:
        # 1. 道路 bounds 與延伸長度
        bnd = road_poly.bounds
        diag = _math.hypot(bnd[2] - bnd[0], bnd[3] - bnd[1])
        ext_len = diag * 1.5
        road_boundary = road_poly.boundary

        # 2. 提取共享邊並延伸
        cut_lines = []
        for blk in adjacent_blocks:
            blk_poly = blk.get('shapely')
            if blk_poly is None or blk_poly.is_empty:
                continue
            try:
                # 街廓必須與道路相鄰（邊界距離 < 2m）
                if blk_poly.boundary.distance(road_boundary) > 2.0:
                    continue
                blk_coords = list(blk_poly.exterior.coords)
                for i in range(len(blk_coords) - 1):
                    p1 = blk_coords[i]
                    p2 = blk_coords[i + 1]
                    # 雙端點皆在道路邊界 1m 內 → 視為共享邊
                    from shapely.geometry import Point as _Pt
                    if (road_boundary.distance(_Pt(p1)) < 1.0 and
                        road_boundary.distance(_Pt(p2)) < 1.0):
                        # 延伸此邊穿越道路
                        dx = p2[0] - p1[0]
                        dy = p2[1] - p1[1]
                        seg_len = _math.hypot(dx, dy)
                        if seg_len < 1e-6:
                            continue
                        ux, uy = dx / seg_len, dy / seg_len
                        # 從邊中點向兩側各延伸 ext_len
                        mid = ((p1[0] + p2[0]) / 2.0, (p1[1] + p2[1]) / 2.0)
                        ext_line = LineString([
                            (mid[0] - ext_len * ux, mid[1] - ext_len * uy),
                            (mid[0] + ext_len * ux, mid[1] + ext_len * uy),
                        ])
                        # 裁切至道路範圍
                        try:
                            clipped = ext_line.intersection(road_poly)
                            if clipped.is_empty:
                                continue
                            if clipped.geom_type == 'LineString':
                                cut_lines.append(clipped)
                            elif clipped.geom_type == 'MultiLineString':
                                for ls in clipped.geoms:
                                    cut_lines.append(ls)
                        except Exception:
                            continue
            except Exception:
                continue

        if not cut_lines:
            return []

        # 3. polygonize：道路邊界 + 延伸切割線 → 子多邊形
        try:
            all_lines = [road_boundary] + cut_lines
            merged = unary_union(all_lines)
            sub_polys = list(polygonize(merged))
        except Exception:
            return []

        if not sub_polys:
            return []

        # 4. 形狀檢核：過濾畸形碎片
        good_polys = []
        for sp in sub_polys:
            try:
                if not sp.is_valid:
                    sp = sp.buffer(0)
                if sp.is_empty or sp.area < min_piece_area:
                    continue
                # 緊湊度 = 4π·area / perimeter²；圓 = 1，狹長 ≈ 0
                perim = sp.length
                if perim < 1e-6:
                    continue
                compactness = 4.0 * _math.pi * sp.area / (perim * perim)
                if compactness < min_compactness:
                    continue
                good_polys.append(sp)
            except Exception:
                continue

        if not good_polys:
            return []

        # 5. 各合格子矩形取 MBR 中心線
        centerlines = []
        for gp in good_polys:
            try:
                lines = _mbr_centerline_of(gp)
                centerlines.extend(lines)
            except Exception:
                continue

        return centerlines
    except Exception:
        return []


def _extract_road_centerlines(road_poly, max_branches: int = 8,
                               adjacent_blocks: list = None) -> list:
    """
    🆕 任務一：複雜形狀道路多邊形的中心線骨架化（每分支一條 LineString）

    🆕 Task B（深度重構）：若提供 adjacent_blocks，優先嘗試「相鄰街廓邊界線延伸切割」
    （`_extend_boundary_through_road`），失敗才退回原「反射頂點 + 接合區」演算法。

    演算法（純 shapely，不引入新依賴）：
      0. 若 adjacent_blocks 不為空 → 先試 _extend_boundary_through_road；成功即回傳
      1. polygon.simplify(0.5) 消除碎共線點
      2. 凸殼比 = poly.area / convex_hull.area；> 0.85 視為單純路型 → 單條 MBR 中線
      3. 否則進入「反射頂點 + 接合區切除」分支偵測：
         a. 偵測所有「反射（凹）頂點」(reflex vertices)：在多邊形邊界繞行時，
            該點轉向方向與多邊形整體繞行方向相反 → 為凹角（接合處）
         b. 將反射頂點之 bounding box（含小量 padding）作為「接合區 (junction box)」
         c. road.difference(junction_box) → 自動分裂為 N 個分支多邊形
         d. 每個分支跑 _mbr_centerline_of 取一條中心線
         e. 將各中心線最靠近接合中心之端點延長至接合中心，使所有分支匯於一點
      4. 若分支數 > max_branches → 取最大 max_branches 條

    參數：
      road_poly         shapely Polygon（道路用地多邊形）
      max_branches      最多分支數（避免極端碎裂）
      adjacent_blocks   list[{'label': str, 'shapely': Polygon}]（相鄰街廓；可選）

    回傳：list[LineString]（每分支一條中心線；空 list 表示計算失敗）
    """
    from shapely.geometry import Polygon, LineString, MultiPolygon

    if road_poly is None or road_poly.is_empty:
        return []

    # 🆕 Task B：優先嘗試「相鄰街廓邊界線延伸切割」
    if adjacent_blocks:
        try:
            ext_lines = _extend_boundary_through_road(road_poly, adjacent_blocks)
            if ext_lines:
                return ext_lines
        except Exception:
            pass   # fallback 至既有演算法

    try:
        poly = road_poly.simplify(0.5, preserve_topology=True)
        if poly.is_empty or poly.area < 1.0:
            return []

        # 步驟 2：凸殼比判斷路型
        try:
            ch_area = poly.convex_hull.area
            convex_ratio = poly.area / ch_area if ch_area > 0 else 1.0
        except Exception:
            convex_ratio = 1.0

        if convex_ratio >= 0.85:
            # 路型接近凸 → 單條 MBR 中線
            return _mbr_centerline_of(poly)

        # 步驟 3：找反射（凹）頂點
        coords = list(poly.exterior.coords)
        if len(coords) > 0 and coords[0] == coords[-1]:
            coords = coords[:-1]
        n = len(coords)
        if n < 4:
            return _mbr_centerline_of(poly)

        # shapely exterior 之繞向（CCW = 正向）
        is_ccw = bool(poly.exterior.is_ccw)
        reflex_pts = []
        for i in range(n):
            prev_p = coords[(i - 1) % n]
            curr = coords[i]
            next_p = coords[(i + 1) % n]
            v1x = curr[0] - prev_p[0]; v1y = curr[1] - prev_p[1]
            v2x = next_p[0] - curr[0]; v2y = next_p[1] - curr[1]
            cross = v1x * v2y - v1y * v2x
            # CCW 多邊形：凹頂點 cross < 0；CW：凹頂點 cross > 0
            if (is_ccw and cross < -0.01) or ((not is_ccw) and cross > 0.01):
                reflex_pts.append(curr)

        if not reflex_pts:
            return _mbr_centerline_of(poly)

        # 步驟 3-b：接合區 bounding box（含 padding）
        rxs = [p[0] for p in reflex_pts]
        rys = [p[1] for p in reflex_pts]
        # padding 取 MBR 短邊 / 4 與 1.0m 較大者
        try:
            mbr = poly.minimum_rotated_rectangle
            mbr_coords = list(mbr.exterior.coords)
            mbr_edges = []
            for i in range(min(4, len(mbr_coords) - 1)):
                p1 = mbr_coords[i]; p2 = mbr_coords[i + 1]
                mbr_edges.append(((p2[0] - p1[0]) ** 2 + (p2[1] - p1[1]) ** 2) ** 0.5)
            mbr_edges.sort()
            mbr_short = mbr_edges[0] if mbr_edges else 4.0
        except Exception:
            mbr_short = 4.0
        pad = max(mbr_short * 0.25, 1.0)
        jx_min = min(rxs) - pad; jx_max = max(rxs) + pad
        jy_min = min(rys) - pad; jy_max = max(rys) + pad
        junction_box = Polygon([
            (jx_min, jy_min), (jx_max, jy_min),
            (jx_max, jy_max), (jx_min, jy_max),
        ])
        junction_center = ((jx_min + jx_max) / 2.0, (jy_min + jy_max) / 2.0)

        # 步驟 3-c：切除接合區 → 自動分裂為各分支
        try:
            arms_geom = poly.difference(junction_box)
        except Exception:
            arms_geom = None

        if arms_geom is None or arms_geom.is_empty:
            return _mbr_centerline_of(poly)

        if isinstance(arms_geom, MultiPolygon):
            arms = list(arms_geom.geoms)
        else:
            arms = [arms_geom]

        # 過濾極小 arm（< 1㎡）並依面積排序
        arms = [a for a in arms if hasattr(a, 'area') and a.area >= 1.0]
        if not arms:
            return _mbr_centerline_of(poly)
        arms.sort(key=lambda a: -a.area)
        arms = arms[:max_branches]

        # 步驟 3-d：每個 arm 取 MBR 中線
        centerlines = []
        for arm in arms:
            try:
                arm_lines = _mbr_centerline_of(arm)
                centerlines.extend(arm_lines)
            except Exception:
                continue

        if not centerlines:
            return _mbr_centerline_of(poly)

        # 步驟 3-e：將每條中心線靠近接合中心之端點延長至接合中心
        extended = []
        for ln in centerlines:
            try:
                ln_coords = list(ln.coords)
                if len(ln_coords) < 2:
                    continue
                p1 = ln_coords[0]; p2 = ln_coords[-1]
                d1 = ((p1[0] - junction_center[0]) ** 2 + (p1[1] - junction_center[1]) ** 2) ** 0.5
                d2 = ((p2[0] - junction_center[0]) ** 2 + (p2[1] - junction_center[1]) ** 2) ** 0.5
                if d1 < d2:
                    extended.append(LineString([junction_center, p1, p2]))
                else:
                    extended.append(LineString([p1, p2, junction_center]))
            except Exception:
                extended.append(ln)
        return extended
    except Exception:
        return []


def _mbr_centerline_of(poly) -> list:
    """
    以 MBR 兩短邊中點連線取得單條中心線（沿長軸方向）
    例：100×8 矩形 → 中心線為 (0,4)→(100,4)（沿 100m 長邊方向）
    """
    from shapely.geometry import LineString
    if poly is None or poly.is_empty:
        return []
    try:
        mbr = poly.minimum_rotated_rectangle
        coords = list(mbr.exterior.coords)
        if coords and coords[0] == coords[-1]:
            coords = coords[:-1]
        if len(coords) < 4:
            return []
        edges = []
        for i in range(4):
            p1 = coords[i]; p2 = coords[(i + 1) % 4]
            L = ((p2[0] - p1[0]) ** 2 + (p2[1] - p1[1]) ** 2) ** 0.5
            mid = ((p1[0] + p2[0]) / 2.0, (p1[1] + p2[1]) / 2.0)
            edges.append({'p1': p1, 'p2': p2, 'L': L, 'mid': mid})
        # 依長度排序：edges[0,1] 為長邊（平行於長軸）；edges[2,3] 為短邊（垂直於長軸）
        edges.sort(key=lambda e: -e['L'])
        # 中心線 = 兩短邊中點連線（沿長軸方向）
        m1 = edges[2]['mid']; m2 = edges[3]['mid']
        if ((m2[0] - m1[0]) ** 2 + (m2[1] - m1[1]) ** 2) < 1e-6:
            return []
        return [LineString([m1, m2])]
    except Exception:
        return []


def cut_road_by_centerline(road_poly, centerlines: list) -> dict:
    """
    用中心線（任務一輸出）切割道路多邊形 → 左右兩半 polygon
    若有多條中心線（T／十字 等多分支），全部一起 split 後依「靠第一條中心線左/右法向量」歸類。

    回傳：
      {
        'left':  shapely Polygon | MultiPolygon | None,
        'right': shapely Polygon | MultiPolygon | None,
        'method': 'split' | 'split_failed_arithmetic',
      }
    若 split 失敗（含中心線完全在 polygon 外），回傳 left/right 為 None，呼叫端可退回算術平均。
    """
    from shapely.geometry import LineString, Polygon, MultiPolygon
    from shapely.ops import split as shp_split, unary_union
    if road_poly is None or not centerlines:
        return {'left': None, 'right': None, 'method': 'split_failed_arithmetic'}
    try:
        # 1. 合成中心線（多條 → MultiLineString）
        cl_list = [ln for ln in centerlines if hasattr(ln, 'coords') and len(list(ln.coords)) >= 2]
        if not cl_list:
            return {'left': None, 'right': None, 'method': 'split_failed_arithmetic'}
        cl_union = unary_union(cl_list)

        # 2. 🆕 Phase 7 Module 2 (M2-3)：buffer(0.05) 微膨脹確保中心線端點貫穿邊界
        try:
            road_for_split = road_poly.buffer(0.05) if road_poly.is_valid else road_poly.buffer(0).buffer(0.05)
        except Exception:
            road_for_split = road_poly
        try:
            result = shp_split(road_for_split, cl_union)
        except Exception:
            try:
                # fallback：對原 polygon 試一次（不加 buffer）
                result = shp_split(road_poly, cl_union)
            except Exception:
                return {'left': None, 'right': None, 'method': 'split_failed_arithmetic'}
        pieces = list(result.geoms) if hasattr(result, 'geoms') else [result]
        if not pieces:
            return {'left': None, 'right': None, 'method': 'split_failed_arithmetic'}

        # 3. 用第一條中心線的方向計算左右法向量
        import numpy as np
        cl0 = cl_list[0]
        c0 = list(cl0.coords)
        p0 = np.array(c0[0], dtype=float)
        p1 = np.array(c0[-1], dtype=float)
        d = p1 - p0
        d_norm = float(np.linalg.norm(d))
        if d_norm < 1e-9:
            return {'left': None, 'right': None, 'method': 'split_failed_arithmetic'}
        d_hat = d / d_norm
        n_left = np.array([-d_hat[1], d_hat[0]])  # 逆時針 90°（左法向量）

        # 4. 逐塊 piece 歸類
        left_pieces = []
        right_pieces = []
        for piece in pieces:
            if not isinstance(piece, Polygon):
                continue
            c = piece.centroid
            v = np.array([c.x - p0[0], c.y - p0[1]])
            side_dot = float(np.dot(v, n_left))
            if side_dot >= 0:
                left_pieces.append(piece)
            else:
                right_pieces.append(piece)

        left_geom = unary_union(left_pieces) if left_pieces else None
        right_geom = unary_union(right_pieces) if right_pieces else None
        return {'left': left_geom, 'right': right_geom, 'method': 'split'}
    except Exception:
        return {'left': None, 'right': None, 'method': 'split_failed_arithmetic'}


def classify_blocks_by_centerline(blocks: list, centerlines: list) -> dict:
    """
    依中心線把街廓分為左右兩側
    blocks: list[dict]，每個含 'centroid' (x,y) 或 'shapely' polygon
    centerlines: 任務一輸出（list[LineString]）
    回傳：{'left': [block_id1, ...], 'right': [...], 'on_line': [...]}
    """
    import numpy as np
    if not blocks or not centerlines:
        return {'left': [], 'right': [], 'on_line': []}
    cl_list = [ln for ln in centerlines if hasattr(ln, 'coords') and len(list(ln.coords)) >= 2]
    if not cl_list:
        return {'left': [], 'right': [], 'on_line': []}
    # 用第一條中心線決定左右法向量
    c0 = list(cl_list[0].coords)
    p0 = np.array(c0[0], dtype=float)
    p1 = np.array(c0[-1], dtype=float)
    d = p1 - p0
    dn = float(np.linalg.norm(d))
    if dn < 1e-9:
        return {'left': [], 'right': [], 'on_line': []}
    d_hat = d / dn
    n_left = np.array([-d_hat[1], d_hat[0]])

    left, right, on_line = [], [], []
    for b in blocks:
        try:
            if 'centroid' in b and b['centroid']:
                cx, cy = b['centroid']
            elif b.get('shapely') is not None:
                _c = b['shapely'].centroid
                cx, cy = _c.x, _c.y
            else:
                continue
            v = np.array([cx - p0[0], cy - p0[1]])
            s = float(np.dot(v, n_left))
            bid = b.get('id', b.get('label', ''))
            if abs(s) < 0.5:  # 容許 0.5 m 內視為線上
                on_line.append(bid)
            elif s > 0:
                left.append(bid)
            else:
                right.append(bid)
        except Exception:
            continue
    return {'left': left, 'right': right, 'on_line': on_line}


def _find_merge_candidates_v2(tp_pub, all_buildable_with_poly: list,
                                pub_poly,
                                pub_parent_normalized: str,
                                pub_gid: str,
                                ownership_map: dict,
                                adj_dist_m: float = 10.0) -> tuple:
    """
    🆕 Phase 8 Issue 2 / 🔧 Phase 9.1：四層 fallback 候選地搜尋

    Tier 1: 同 parent 地號 + 「相鄰建地」buildable（最理想，地號跨街廓延伸案例）
            例：628-38(2) in RD4 → 找到 628-38(1) in R6（共邊相鄰）
    Tier 2: 同 parent 地號 + 任意位置 buildable（不限相鄰）
            例：628-7(3) in G1 → 找到 628-7(1) in R6 / 628-7(2) in R5
    Tier 3: 同歸戶群組 + 「相鄰建地」buildable（公園地號異於建地之合作案例）
            例：G032 在 G1 公園有 1234-5 → 找到 G032 在 R6 的不同 parent 地號 buildable
    Tier 4: 真孤立 → 空清單（呼叫端走虛擬 G 路徑）

    🔧 Phase 9.1：adj_dist_m 預設由 5.0m 放寬至 10.0m
       原因：實務上公園 polygon 與相鄰建地常隔一條 5~8m 道路；
            若以 5m 為閾值，會錯失隔路相鄰之合併機會 → 誤判孤立 → 抵費地暴增

    參數：
      tp_pub                       公設地的 temp_parcel
      all_buildable_with_poly      預先收集的 [(tp, polygon), ...]
      pub_poly                     公設地之 shapely polygon
      pub_parent_normalized        公設地 parent 地號（已正規化）
      pub_gid                      公設地之歸戶群組 ID（可能空）
      ownership_map                Tab 1 歸戶映射
      adj_dist_m                   地理相鄰之距離閾值（預設 10m，從 5m 放寬）

    回傳：(candidates_list, tier_label)
      candidates_list: [(tp, tp_polygon), ...]
      tier_label: 'tier1' / 'tier2' / 'tier3' / 'tier4'
    """
    if not all_buildable_with_poly or pub_poly is None:
        return [], 'tier4'

    # Tier 1: 同 parent 地號 + 相鄰
    t1 = []
    if pub_parent_normalized:
        for tp_b, b_poly in all_buildable_with_poly:
            if tp_b is tp_pub:
                continue
            b_orig_norm = _normalize_landno_module(tp_b.get('原地號', ''))
            if b_orig_norm != pub_parent_normalized:
                continue
            try:
                if pub_poly.distance(b_poly) < adj_dist_m:
                    t1.append((tp_b, b_poly))
            except Exception:
                continue
    if t1:
        return t1, 'tier1'

    # Tier 2: 同 parent 地號 + 任意位置（NEW 相對 V12）
    t2 = []
    if pub_parent_normalized:
        for tp_b, b_poly in all_buildable_with_poly:
            if tp_b is tp_pub:
                continue
            b_orig_norm = _normalize_landno_module(tp_b.get('原地號', ''))
            if b_orig_norm == pub_parent_normalized:
                t2.append((tp_b, b_poly))
    if t2:
        return t2, 'tier2'

    # Tier 3: 同歸戶 + 相鄰
    t3 = []
    if pub_gid:
        for tp_b, b_poly in all_buildable_with_poly:
            if tp_b is tp_pub:
                continue
            b_gid = ownership_map.get(str(tp_b.get('原地號', '') or '').strip(), '')
            if not b_gid:
                # fallback 透過正規化查找
                try:
                    b_gid = _resolve_ownership(tp_b.get('原地號', ''), ownership_map)
                except Exception:
                    b_gid = ''
            if b_gid != pub_gid:
                continue
            try:
                if pub_poly.distance(b_poly) < adj_dist_m:
                    t3.append((tp_b, b_poly))
            except Exception:
                continue
    if t3:
        return t3, 'tier3'

    # Tier 4: 真孤立
    return [], 'tier4'


# ═══════════════════════════════════════════════════════════════════════
# 🆕 Phase 9.8 Helper Suite：防擠壓模擬引擎之底層 helper
# ═══════════════════════════════════════════════════════════════════════
# 設計目的：將 Step L PK / 公設整併 / G 迭代 等邏輯封裝為可重複呼叫之 helpers，
# 供 Phase 9.9 之 _master_allocation_loop（4-Pass 防擠壓演算法）使用。
# Phase 9.8 階段：僅抽取 helper，不改變既有流程行為。



def _find_orphans_within_distance(blk_meta: dict, all_orphans: list,
                                    ownership_map: dict, dist_m: float = 15.0,
                                    require_same_owner: bool = False,
                                    blk_owners: set = None) -> list:
    """
    🆕 Phase 9.8：找出該街廓 N 公尺內之孤立公設地候選

    參數：
      blk_meta          目標街廓 dict
      all_orphans       全部孤立公設地清單（從 f3_orphan_parcels 取）
      ownership_map     歸戶映射
      dist_m            幾何相鄰距離閾值（預設 15m）
      require_same_owner 是否限定「該街廓內已有同歸戶建地」
      blk_owners        該街廓內既有建地之歸戶 set（搭配 require_same_owner）

    回傳：候選 orphan list（依距離由近至遠排序）
    """
    from shapely.geometry import Polygon as _SP_orph
    if not all_orphans or not blk_meta:
        return []
    verts = blk_meta.get('vertices') or []
    if len(verts) < 3:
        return []
    try:
        blk_poly = _SP_orph(verts)
        if not blk_poly.is_valid:
            blk_poly = blk_poly.buffer(0)
    except Exception:
        return []

    candidates = []
    for orph in all_orphans:
        coords = orph.get('polygon_coords') or []
        if len(coords) < 3:
            continue
        try:
            orph_poly = _SP_orph(coords)
            if not orph_poly.is_valid:
                orph_poly = orph_poly.buffer(0)
            d = float(blk_poly.distance(orph_poly))
            if d > dist_m:
                continue
        except Exception:
            continue
        # 同歸戶過濾（若啟用）
        if require_same_owner and blk_owners:
            try:
                orph_gid = _resolve_ownership(orph.get('原地號', ''), ownership_map)
            except Exception:
                orph_gid = ownership_map.get(orph.get('原地號', ''), '')
            if orph_gid not in blk_owners:
                continue
        candidates.append((d, orph))
    candidates.sort(key=lambda x: x[0])
    return [c[1] for c in candidates]


def _detect_block_allocation_dir(block_poly, front_dir):
    """
    自動推算「街廓分配線」之方向向量（任務二）
    實務定義：街廓分配線通常為街廓背側（屁股）的基準線，與正面道路近似平行。
    本函式利用最小外接矩形 (MBR) 自動判定：
      1. 取街廓 MBR（minimum_rotated_rectangle）的 4 條邊
      2. MBR 的相鄰邊互相垂直 → 共 2 組對邊（idx 0&2, 1&3）
      3. 與 front_dir 較平行（|cos| 較大）的那組對邊 = MBR 之長邊（與正面道路近平行）
      4. 取其方向向量；最終方向取「與 front_dir 同向」者，得 allocation_dir
      5. 切割法向量 = 將 allocation_dir 旋轉 90° → 用於切出宗地兩側界線

    參數：
      block_poly  shapely Polygon（街廓多邊形）
      front_dir   正面道路方向單位向量（np.array shape=(2,)）

    回傳：(allocation_dir, normal_dir)
      allocation_dir  街廓分配線方向（與 front 近平行；單位向量）
      normal_dir      切割法向量（垂直於 allocation_dir；單位向量）
      若計算失敗：回傳 (front_dir, perpendicular_to_front_dir)
    """
    import numpy as np
    f_dir = np.asarray(front_dir, dtype=float)
    f_n = np.linalg.norm(f_dir)
    if f_n < 1e-9:
        return f_dir, np.array([0.0, 0.0])
    f_dir = f_dir / f_n
    fallback_n = np.array([-f_dir[1], f_dir[0]])

    try:
        if block_poly is None or block_poly.is_empty:
            return f_dir, fallback_n
        mbr = block_poly.minimum_rotated_rectangle
        coords = list(mbr.exterior.coords)
        if coords and coords[0] == coords[-1]:
            coords = coords[:-1]
        if len(coords) < 4:
            return f_dir, fallback_n

        edges = []
        for i in range(4):
            p1 = np.asarray(coords[i], dtype=float)
            p2 = np.asarray(coords[(i + 1) % 4], dtype=float)
            v = p2 - p1
            L = float(np.linalg.norm(v))
            if L < 1e-6:
                continue
            edges.append({'idx': i, 'dir': v / L, 'length': L})
        if len(edges) < 4:
            return f_dir, fallback_n

        # MBR 對邊：0&2 一組、1&3 一組；找與 front_dir 較平行者
        cos_02 = abs(float(np.dot(edges[0]['dir'], f_dir)))
        cos_13 = abs(float(np.dot(edges[1]['dir'], f_dir)))
        if cos_02 >= cos_13:
            ad = edges[0]['dir']
        else:
            ad = edges[1]['dir']

        # 反向則翻轉，使 allocation_dir 與 front_dir 同向
        if float(np.dot(ad, f_dir)) < 0:
            ad = -ad

        n_hat = np.array([-ad[1], ad[0]])  # 旋轉 90°
        return ad, n_hat
    except Exception:
        return f_dir, fallback_n


def _block_strip(block_poly, d_hat, baseline_pt, S, allocation_dir=None):
    """
    從 baseline_pt 沿 d_hat 推進 S，切出 block ∩ 切割帶。回傳 (cut_geom, area)。

    切割帶兩側方向 n_hat = rot90(allocation_dir)。allocation_dir 應為
    rot90(f3_cad_alloc_dir)（§0.5-B），使 n_hat ∥ f3_cad_alloc_dir（地界線 ∥ 宗地分配線）。
    🆕 W-C §0.5-C：缺 allocation_dir → 退「⊥d_hat」矩形近似，**不再從 MBR 推算**
    （廢 _detect_block_allocation_dir 靜默退 MBR 長邊；缺 ALLOC 之街廓由上游 st.warning）。
    """
    import numpy as np
    from shapely.geometry import Polygon
    if S <= 0:
        return None, 0.0

    # 🆕 §0.5-C：切割法向量 n_hat。allocation_dir（=rot90(f3_cad_alloc_dir)）由呼叫端必填；
    #   缺/零向量 → 退「⊥d_hat」（界線⊥FRONT 矩形近似），**不再 _detect_block_allocation_dir
    #   靜默退 MBR 長邊**（缺 ALLOC 之街廓上游已 st.warning，避免雙重靜默錯誤）。
    if allocation_dir is not None:
        ad = np.asarray(allocation_dir, dtype=float)
        ad_n = float(np.linalg.norm(ad))
        if ad_n < 1e-9:
            n_hat = np.array([-d_hat[1], d_hat[0]])
        else:
            ad = ad / ad_n
            n_hat = np.array([-ad[1], ad[0]])
    else:
        n_hat = np.array([-d_hat[1], d_hat[0]])

    # 數值穩健：若計算失敗或為零向量，退回舊行為
    if not np.any(np.isfinite(n_hat)) or float(np.linalg.norm(n_hat)) < 1e-9:
        n_hat = np.array([-d_hat[1], d_hat[0]])

    bnd = block_poly.bounds
    big = max(bnd[2] - bnd[0], bnd[3] - bnd[1]) * 4.0 + 100.0
    bp = np.asarray(baseline_pt, dtype=float)
    strip = Polygon([
        tuple(bp - big * n_hat),
        tuple(bp + S * d_hat - big * n_hat),
        tuple(bp + S * d_hat + big * n_hat),
        tuple(bp + big * n_hat),
    ])
    try:
        cut = block_poly.intersection(strip)
        return cut, float(cut.area) if hasattr(cut, 'area') else 0.0
    except Exception:
        return None, 0.0


def solve_G_binary(a: float, A: float, B: float, C: float,
                   l_front: float, l_side: float, F: float,
                   block_poly, d_hat, baseline_pt,
                   S_max_limit: float,
                   is_corner: bool = False,
                   side_label: str = '左側',
                   tol: float = 0.01, max_iter: int = 80,
                   allocation_dir=None,
                   side_mid=None, W_prev: float = 0.0) -> dict:
    """
    幾何驅動的二分法解 S：
      每輪猜 S_guess → 從 baseline_pt 沿 d_hat 切出 block strip 多邊形 → 計算 area_geom；
      比對法規公式 G_target = [a(1-A·B) − Rw·F·l₁ − S·l₂](1-C)；
      |area_geom − G_target| ≤ tol 即收斂。
    搜尋區間：S ∈ [0, S_max_limit]（剩餘正面道路可分配長）

    🆕 任務二（2026-04）：新增 allocation_dir 參數
      若提供，則切割帶之兩側界線方向「垂直於 allocation_dir（街廓分配線）」；
      未提供則由 _block_strip 自動以 MBR 推算（再 fallback 至垂直於 d_hat）。

    回傳：{'G', 'S', 'W', 'Rw_pct', 'area_geom', 'iterations', 'converged', 'trace',
          'allocation_dir'}
    """
    import numpy as np
    a = float(a or 0.0); A = float(A or 0.0); B = float(B or 0.0); C = float(C or 0.0)
    l_front = float(l_front or 0.0); l_side = float(l_side or 0.0); F = float(F or 0.0)
    S_max_limit = max(0.1, float(S_max_limit or 0.1))
    W_prev = float(W_prev or 0.0)

    # §0.5-C：allocation_dir 一律由呼叫端傳入 = rot90(f3_cad_alloc_dir)（臨街向，
    #   = W 量測軸 n_alloc）。廢除 MBR 自動推算 fallback（鐵律 G/H）。
    #   缺值（None）→ W/Rw 不計算（側街負擔=0）；_block_strip 仍可切帶（last-resort）。
    _n_alloc = None
    if allocation_dir is not None:
        try:
            _na = np.asarray(allocation_dir, dtype=float)
            _nn = float(np.linalg.norm(_na))
            if _nn > 1e-9:
                _n_alloc = _na / _nn
        except Exception:
            _n_alloc = None

    S_min, S_max = 0.0, S_max_limit
    trace = []
    converged = False
    S_conv = 0.0; G_conv = 0.0; W_conv = 0.0; Rw_conv = 0.0; area_conv = 0.0
    last_W = 0.0; last_Rw = 0.0; last_G_target = 0.0; last_area = 0.0
    it = 0
    for it in range(1, max_iter + 1):
        S_guess = (S_min + S_max) / 2.0
        cut, area_geom = _block_strip(block_poly, d_hat, baseline_pt, S_guess,
                                       allocation_dir=allocation_dir)

        # 🆕 W-C §1/§3：W 累積 = W_prev + 本筆臨街增量（附件二「向宗地分配線作垂線」之垂距）。
        #   增量 = 本筆 S 在 n_alloc(臨街向)上的投影 = S_guess·|d_hat·n_alloc|
        #   （= 相鄰兩地界線之垂距；首筆 W_prev=0 → 從側街起算，telescoping ΣRw=100%）。
        #   注意：不可從 SIDE_LINE 中點絕對量（中點在半深處，會使首筆 W_prev≠0、ΣRw<100%）。
        #   side_mid 之有無僅作「角側 gate」（None=非角側→W/Rw=0）；W_prev 由外層 thread。
        if _n_alloc is not None and side_mid is not None:
            _dh = np.asarray(d_hat, dtype=float)
            _dhn = float(np.linalg.norm(_dh))
            _cos_dn = abs(float(np.dot(_dh / _dhn, _n_alloc))) if _dhn > 1e-9 else 1.0
            W = W_prev + S_guess * _cos_dn
            Rw = rw_increment(W_prev, W)
            Rw_pct = Rw * 100.0
        else:
            W = 0.0; Rw_pct = 0.0; Rw = 0.0

        G_target = max(0.0, (a * (1.0 - A * B)
                             - Rw * F * l_side
                             - S_guess * l_front) * (1.0 - C))
        diff = area_geom - G_target

        last_W = W; last_Rw = Rw_pct
        last_G_target = G_target; last_area = area_geom

        trace.append({
            'step': it, 'S_guess': round(S_guess, 3),
            'S_min': round(S_min, 3), 'S_max': round(S_max, 3),
            'W': round(W, 2), 'Rw_pct': round(Rw_pct, 2),
            'area_geom': round(area_geom, 2),
            'G_target': round(G_target, 2), 'diff': round(diff, 3),
        })

        if abs(diff) <= tol:
            converged = True
            S_conv = S_guess; G_conv = G_target
            W_conv = W; Rw_conv = Rw_pct; area_conv = area_geom
            break
        if diff > 0:
            S_max = S_guess
        else:
            S_min = S_guess
        if abs(S_max - S_min) < tol / 1000.0:
            break

    if not converged:
        S_conv = (S_min + S_max) / 2.0
        G_conv = last_G_target
        W_conv = last_W
        Rw_conv = last_Rw
        area_conv = last_area

    # 把 allocation_dir 序列化為可儲存的 list（np.ndarray 不易存於 session_state）
    ad_out = None
    if allocation_dir is not None:
        try:
            ad_out = [float(allocation_dir[0]), float(allocation_dir[1])]
        except Exception:
            ad_out = None

    # 🆕 任務七 A：以收斂後 S_conv 再切一次，取得最終 cut polygon 之頂點座標
    #     用以下游「重劃後試分配地籍圖」之 Plotly fill='toself' 渲染
    cut_coords = []
    try:
        final_cut, _final_area = _block_strip(block_poly, d_hat, baseline_pt, S_conv,
                                               allocation_dir=allocation_dir)
        if final_cut is not None and not final_cut.is_empty:
            if hasattr(final_cut, 'geoms'):  # MultiPolygon → 取面積最大者
                _biggest = max(list(final_cut.geoms), key=lambda g: g.area)
                if hasattr(_biggest, 'exterior'):
                    cut_coords = [[float(p[0]), float(p[1])]
                                  for p in list(_biggest.exterior.coords)]
            elif hasattr(final_cut, 'exterior'):  # 單一 Polygon
                cut_coords = [[float(p[0]), float(p[1])]
                              for p in list(final_cut.exterior.coords)]
    except Exception:
        cut_coords = []

    return {
        'G': round(G_conv, 2), 'S': round(S_conv, 2),
        'W': round(W_conv, 2), 'W_far': round(W_conv, 2),  # W_far：本筆遠側累積 W（供外層 thread）
        'Rw_pct': round(Rw_conv, 2),
        'area_geom': round(area_conv, 2),
        'iterations': it, 'converged': converged, 'trace': trace,
        'allocation_dir': ad_out,  # §0.5-B：實際使用之臨街向（rot90(f3_cad_alloc_dir)）
        'cut_coords': cut_coords,  # 任務七：最終配地之多邊形頂點（list[[x,y]]）
    }


# ============ 最小分配面積 / 畸零地附表（花蓮縣畸零地使用規則 §3 第 1 款 一般建築基地） ============
# 依正面路寬分 4 段，對應 5 類使用分區（寬 × 深 公尺）
# 來源：references/1091021花蓮縣畸零地使用規則-附表.pdf
HUALIEN_MIN_LOT_TABLE = {
    # 路寬區間上限 (含)；以 <= 比較，> 25 m 走 999.0
    '住宅區':  [(7.0, 3.00, 12.00), (15.0, 3.50, 14.00), (25.0, 4.00, 16.00), (999.0, 4.00, 16.00)],
    '甲種建築用地': [(7.0, 3.00, 12.00), (15.0, 3.50, 14.00), (25.0, 4.00, 16.00), (999.0, 4.00, 16.00)],
    '乙種建築用地': [(7.0, 3.00, 12.00), (15.0, 3.50, 14.00), (25.0, 4.00, 16.00), (999.0, 4.00, 16.00)],
    '商業區':  [(7.0, 3.50, 11.00), (15.0, 4.00, 15.00), (25.0, 4.50, 15.00), (999.0, 4.30, 18.00)],
    '丙種建築用地': [(7.0, 6.00, 20.00), (15.0, 6.00, 20.00), (25.0, 6.00, 20.00), (999.0, 6.00, 20.00)],
    '風景區':  [(7.0, 6.00, 20.00), (15.0, 6.00, 20.00), (25.0, 6.00, 20.00), (999.0, 6.00, 20.00)],
    '丁種建築用地': [(7.0, 7.00, 16.00), (15.0, 7.00, 16.00), (25.0, 7.00, 16.00), (999.0, 7.00, 16.00)],
    '工業區':  [(7.0, 7.00, 16.00), (15.0, 7.00, 16.00), (25.0, 7.00, 16.00), (999.0, 7.00, 16.00)],
    # 其他分區（預設退回此表）
    '其他':    [(7.0, 3.50, 12.00), (15.0, 4.00, 16.00), (25.0, 4.30, 17.00), (999.0, 4.30, 18.00)],
}


def get_min_lot_size(category: str, front_road_width_m: float) -> dict:
    """
    依街廓分類與正面路寬查詢最小分配寬度 / 深度 / 面積
    （花蓮縣畸零地使用規則 §3 第 1 款 一般建築基地）

    參數：
      category           街廓分類（F3_BLOCK_CATEGORIES 其中之一）
      front_road_width_m 該街廓正面路寬 (m)

    回傳：{'min_width': m, 'min_depth': m, 'min_area': ㎡, 'table_key': 實際查表之分類}
         若非可建築土地（道路、溝渠、公設等），回傳 0/0/0
    """
    cat = (category or '').strip()
    w = float(front_road_width_m or 0.0)

    # 非可建築土地 → 不需最小分配面積檢查
    non_buildable = ('住宅區', '商業區', '甲種建築用地', '乙種建築用地', '丙種建築用地',
                     '丁種建築用地', '風景區', '工業區')
    if cat not in non_buildable and cat not in HUALIEN_MIN_LOT_TABLE:
        return {'min_width': 0.0, 'min_depth': 0.0, 'min_area': 0.0, 'table_key': ''}

    # 取表；若查無，退回「其他」
    table = HUALIEN_MIN_LOT_TABLE.get(cat) or HUALIEN_MIN_LOT_TABLE['其他']
    used_key = cat if cat in HUALIEN_MIN_LOT_TABLE else '其他'

    for upper, mw, md in table:
        if w <= upper:
            return {'min_width': mw, 'min_depth': md, 'min_area': round(mw * md, 2),
                    'table_key': used_key}
    mw, md = table[-1][1], table[-1][2]
    return {'min_width': mw, 'min_depth': md, 'min_area': round(mw * md, 2),
            'table_key': used_key}


def merge_subparcels_by_parent(g_rows: list, min_area_by_block: dict) -> dict:
    """
    同原地號合併：將未達所屬街廓最小分配面積的暫編地號，併入該原地號之最大暫編地號
    （依《平均地權條例》§60-1 + 市地重劃實施辦法 §30）

    參數：
      g_rows             Tab 3 步驟 G 之 G 值清單；每筆含：
                          暫編地號 / 原地號 / 所屬街廓 / G(㎡) / a 面積(㎡) / 重劃前區段
      min_area_by_block  {街廓編號: 該街廓最小分配面積(㎡)}

    規則：
      1. 以「原地號」分組
      2. 每筆暫編地號 G 值若 < 所屬街廓最小分配面積，標為「待合併」
      3. 同原地號內：將所有「待合併」G 值加總後，併入「面積最大且 G ≥ 最小分配面積」的暫編地號
      4. 若同原地號合計 G 值仍未達「主要暫編地號所屬街廓最小分配面積」，則該原地號整體標記
         為「需跨街廓調配」（→ Tab 3 步驟 K 處理）

    回傳：
      {
        'merged_rows': 合併後之各暫編地號（含 '合併來源' 欄位）,
        'cash_compensation': 未達最小分配面積 1/2 的暫編地號（建議現金補償）,
        'cross_block_needed': 需跨街廓調配之原地號列表,
      }
    """
    merged = []
    cash_comp = []
    cross_block = []

    # 🆕 V13 修正 #5：先過濾「抵費地」與「孤立公設地」與「無原地號」之列
    # 這些不應參與「最小分配面積 / 現金補償」檢查
    _SKIP_ROWS = {'抵費地', '🟠 孤立公設地'}
    _filtered_rows = []
    for r in (g_rows or []):
        # 推進側別 落在過濾集中 → 跳過
        if r.get('推進側別', '') in _SKIP_ROWS:
            continue
        # 原地號 為空 / '—' / '(無地號)' → 跳過
        if str(r.get('原地號', '') or '').strip() in ('', '—', '(無地號)'):
            continue
        _filtered_rows.append(r)

    # 依原地號分組
    by_parent = {}
    for r in _filtered_rows:
        p = r.get('原地號', '')
        by_parent.setdefault(p, []).append(dict(r))

    for parent, rows in by_parent.items():
        # 標記每筆狀態
        for r in rows:
            blk = r.get('所屬街廓', '')
            min_area = float(min_area_by_block.get(blk, 0.0) or 0.0)
            g = float(r.get('G(㎡)', 0.0) or 0.0)
            r['_min_area'] = round(min_area, 2)
            # 🚨 Patch D-4：合併觸發條件擴充 — 面積 OR 寬度任一不足即需合併
            _below_width = bool(r.get('_below_min_width', False))
            r['_need_merge'] = (min_area > 0 and g < min_area) or _below_width
            r['_half_min'] = (min_area > 0 and g < min_area / 2.0)
            # 紀錄合併原因（供 UI 顯示）
            if r['_need_merge']:
                if _below_width:
                    r['_merge_reason'] = (
                        f"寬度 {r.get('實際寬度(m)', 0)}m < 法定最小寬"
                    )
                else:
                    r['_merge_reason'] = f"面積 {g:.1f}㎡ < {min_area:.0f}㎡"

        # 1/2 以下 → 建議現金補償（仍先參與合併，但額外標記）
        for r in rows:
            if r['_half_min']:
                cash_comp.append({
                    '原地號': parent, '暫編地號': r['暫編地號'],
                    '所屬街廓': r['所屬街廓'],
                    'G(㎡)': r['G(㎡)'], '最小分配面積(㎡)': r['_min_area'],
                    '狀態': '未達最小分配面積 1/2（建議現金補償）',
                })

        # 若僅一筆 → 直接輸出
        if len(rows) == 1:
            r = rows[0]
            r['合併來源'] = ''
            r['最小分配面積(㎡)'] = r['_min_area']
            r['合併後G(㎡)'] = r['G(㎡)']
            r['是否達最小'] = '✅' if not r['_need_merge'] else '⚠️'
            if r['_need_merge']:
                cross_block.append({
                    '原地號': parent, '主要暫編地號': r['暫編地號'],
                    '目前G(㎡)': r['G(㎡)'], '最小分配面積(㎡)': r['_min_area'],
                    '所屬街廓': r['所屬街廓'],
                })
            # 清除臨時欄位
            for k in ('_min_area', '_need_merge', '_half_min'):
                r.pop(k, None)
            merged.append(r)
            continue

        # 多筆：找主要暫編地號（面積最大者，優先 G ≥ 最小分配面積）
        primary = None
        # 優先：G 已達最小分配面積者中最大
        cand_ok = [r for r in rows if not r['_need_merge']]
        if cand_ok:
            primary = max(cand_ok, key=lambda x: float(x.get('a 面積(㎡)', 0.0) or 0.0))
        else:
            # 全數未達 → 取 a 面積最大者為主
            primary = max(rows, key=lambda x: float(x.get('a 面積(㎡)', 0.0) or 0.0))

        # 合併其他 < 最小分配面積者至 primary
        merged_sources = []
        sum_g_add = 0.0
        for r in rows:
            if r['暫編地號'] == primary['暫編地號']:
                continue
            if r['_need_merge']:
                merged_sources.append(r['暫編地號'])
                sum_g_add += float(r.get('G(㎡)', 0.0) or 0.0)
                # 該筆合併後 G = 0（已併入 primary）
                r_out = dict(r)
                r_out['合併來源'] = ''
                r_out['合併至'] = primary['暫編地號']
                r_out['最小分配面積(㎡)'] = r['_min_area']
                r_out['合併後G(㎡)'] = 0.0
                r_out['是否達最小'] = '🔀 已併入 ' + primary['暫編地號']
                for k in ('_min_area', '_need_merge', '_half_min'):
                    r_out.pop(k, None)
                merged.append(r_out)
            else:
                # 該筆保留原樣
                r_out = dict(r)
                r_out['合併來源'] = ''
                r_out['最小分配面積(㎡)'] = r['_min_area']
                r_out['合併後G(㎡)'] = r['G(㎡)']
                r_out['是否達最小'] = '✅'
                for k in ('_min_area', '_need_merge', '_half_min'):
                    r_out.pop(k, None)
                merged.append(r_out)

        # primary 合併結果
        p_out = dict(primary)
        new_G = float(primary.get('G(㎡)', 0.0) or 0.0) + sum_g_add
        p_out['合併來源'] = '、'.join(merged_sources) if merged_sources else ''
        p_out['最小分配面積(㎡)'] = primary['_min_area']
        p_out['合併後G(㎡)'] = round(new_G, 2)
        # 判斷合併後是否達標
        if primary['_min_area'] > 0 and new_G < primary['_min_area']:
            p_out['是否達最小'] = '⚠️ 仍未達（需跨街廓調配）'
            cross_block.append({
                '原地號': parent, '主要暫編地號': primary['暫編地號'],
                '目前G(㎡)': round(new_G, 2),
                '最小分配面積(㎡)': primary['_min_area'],
                '所屬街廓': primary['所屬街廓'],
            })
        else:
            p_out['是否達最小'] = '✅' if not merged_sources else f'✅ 已合併 {len(merged_sources)} 筆'
        for k in ('_min_area', '_need_merge', '_half_min'):
            p_out.pop(k, None)
        merged.append(p_out)

    return {
        'merged_rows': merged,
        'cash_compensation': cash_comp,
        'cross_block_needed': cross_block,
    }


# ============ 任務三：街角地配地演算法（最小分配面積 + 優先權指數） ============
# 參考《市地重劃作業手冊》P.182-184 + 《花蓮縣畸零地使用規則》
# 與 Tab 1（土地歸戶）資料連動，實作街角地優先分配規則。

def _build_corner_range_v2(side_mid, block_vertices, block_centroid,
                            alloc_dir, shift_distance, chamfer_tri=None):
    """街角規定範圍 = SIDE_LINE 中點 + 宗地分配線法向平移 shift_distance → 切 BLOCK → 扣截角。

    alloc_dir       宗地分配線方向單位向量 (ux, uy)
    shift_distance  退縮 + 畸零地最小寬
    chamfer_tri     道路截角三角形 shapely Polygon（可 None）
    回傳 shapely Polygon 或 None
    """
    try:
        from shapely.geometry import Polygon as _SPv2, LineString as _LSv2
        from shapely.ops import split as _split_v2
        block_poly = _SPv2(block_vertices)
        if not block_poly.is_valid:
            block_poly = block_poly.buffer(0)
        if block_poly.is_empty:
            return None
        ux, uy = alloc_dir
        nx, ny = -uy, ux
        _cen = block_centroid
        _tp = (side_mid[0] + nx, side_mid[1] + ny)
        _tm = (side_mid[0] - nx, side_mid[1] - ny)
        if ((_tm[0]-_cen[0])**2+(_tm[1]-_cen[1])**2) < ((_tp[0]-_cen[0])**2+(_tp[1]-_cen[1])**2):
            nx, ny = -nx, -ny
        shifted_pt = (side_mid[0] + nx * shift_distance,
                      side_mid[1] + ny * shift_distance)
        _ext = 500.0
        cut_line = _LSv2([
            (shifted_pt[0] - ux * _ext, shifted_pt[1] - uy * _ext),
            (shifted_pt[0] + ux * _ext, shifted_pt[1] + uy * _ext),
        ])
        try:
            from shapely.geometry import Point as _Pv2
            pieces = _split_v2(block_poly, cut_line)
            target = min(pieces.geoms,
                         key=lambda g: g.centroid.distance(_Pv2(side_mid)))
        except Exception:
            target = block_poly
        if target.is_empty:
            return None
        if chamfer_tri is not None:
            try:
                target = target.difference(chamfer_tri)
            except Exception:
                pass
        if not target.is_valid:
            target = target.buffer(0)
        return target if not target.is_empty else None
    except Exception:
        return None


def _make_chamfer_tri_wb(blk_meta, which_side):
    """從 blk_meta.geom_restore.theoretical_corners 建截角三角形（left 或 right）。
    W-D.1.3-c.1：側別直讀 tc['side']（正典、來自 side_lines_by_side）；廢除「距 FRONT 端點重猜」
    （BLOCK 邊繞向≠語意左右→五塊翻）。tc 無 side（舊快取）→ 不猜、None。該側無截角 → None。"""
    try:
        from shapely.geometry import Polygon as _SPct
        _gr = (blk_meta or {}).get('geom_restore') or {}
        _cls = _gr.get('classification') or {}
        _edges = _cls.get('edges') or []
        _tcs = _gr.get('theoretical_corners') or []
        if not _tcs or not _edges:
            return None
        for tc in _tcs:
            if tc.get('side') != which_side:
                continue
            _c_idx = tc.get('cutoff_idx', -1)
            if _c_idx < 0 or _c_idx >= len(_edges):
                continue
            _c_edge = _edges[_c_idx]
            _ep1 = _c_edge.get('p1'); _ep2 = _c_edge.get('p2')
            _cpt = tc.get('corner')
            if not all([_ep1, _ep2, _cpt]):
                continue
            return _SPct([_ep1, _ep2, _cpt])
        return None
    except (KeyError, IndexError, TypeError, ValueError, AttributeError):
        return None






# 🆕 W-D.1.3-b：_compute_corner_overlap_area（15m 方框評分）整段刪——項三改真交集/range 面積、
#   全檔已無呼叫（CLAUDE.md：舊函式整個刪、不留 stub）。


def select_corner_lots_both_sides_v12(
    candidates: list,
    front_line_p1: tuple,
    front_line_p2: tuple,
    cutoff_p1_end: float,
    cutoff_p2_end: float,
    base_front_len_m: float,
    base_side_len_m_p1: float,
    base_side_len_m_p2: float,
    min_corner_area_p1: float,
    min_corner_area_p2: float,
    g_values_map: dict = None,
) -> dict:
    """
    🆕 V12 模組 3 + 🚨 補強 C：街角地空間距離綁定 PK

    核心改動：放棄字串 '左側'/'右側' 標籤，改用「質心 → FRONT_LINE p1/p2 之歐式距離」
    自動分流候選人至 p1_end / p2_end 兩組，每組獨立 PK。

    🚨 強化點 2：第一關用 G 值（非物理跨占面積）做最小分配面積門檻判斷

    參數：
      candidates           候選 dict list；需含 'centroid', '暫編地號' 等欄位
                            （可選：'G_value', 'physical_overlap_area', 'front_length',
                              'side_length', '歸戶'）
      front_line_p1, p2    FRONT_LINE 兩端座標 (x, y)
      cutoff_p1_end        p1 端虛擬截角面積（依使用者規範要扣的）
      cutoff_p2_end        p2 端
      base_front_len_m     FRONT_LINE 長度（基準）
      base_side_len_m_p1   p1 端側面長度（用於優先權指數）
      base_side_len_m_p2   p2 端
      min_corner_area_p1   p1 端街角地最小分配面積
      min_corner_area_p2   p2 端
      g_values_map         {暫編地號: 已計算之 G 值}（用於第一關門檻）

    回傳：
      {
        'p1_end': {'winner', 'qualified', 'eliminated', 'note', 'group_size'},
        'p2_end': {'winner', 'qualified', 'eliminated', 'note', 'group_size'},
      }
    """
    import numpy as _np_v12
    g_values_map = g_values_map or {}

    if not candidates:
        return {
            'p1_end': {'winner': None, 'qualified': [], 'eliminated': [],
                       'note': '無候選', 'group_size': 0},
            'p2_end': {'winner': None, 'qualified': [], 'eliminated': [],
                       'note': '無候選', 'group_size': 0},
        }

    p1 = _np_v12.array(front_line_p1, dtype=float)
    p2 = _np_v12.array(front_line_p2, dtype=float)

    # 🆕 Phase 8 Issue 3：FRONT_LINE 方向向量（供 corner overlap 計算用）
    _fl_vec = p2 - p1
    _fl_len = float(_np_v12.linalg.norm(_fl_vec))
    if _fl_len > 1e-9:
        _fl_dir = (_fl_vec / _fl_len).tolist()
    else:
        _fl_dir = [1.0, 0.0]

    # 🚨 Patch B-4：街角地實體交集前置篩選（Phase 9.16 Module 1）
    # 用 setback + legal_min_depth 在 front_line_p1 / p2 端建構真實 corner_polygon，
    # 只保留 polygon 與該 corner_polygon 有「實體交集」之候選。
    # 這彌補 V12 僅用「質心投影距離」分流之缺陷（候選與實際街角範圍未對齊）。
    from shapely.geometry import Polygon as _SP_B4
    try:
        import streamlit as _st_B4
        # 🚨 W-B 修正：退縮 0 陷阱——不可用 `or 3.5`（0.0 是 falsy 會被竄改成 3.5）
        _setback_raw = _st_B4.session_state.get('f3L_setback_default', 3.5)
        _setback_B4 = float(_setback_raw) if _setback_raw is not None else 3.5
    except Exception:
        _setback_B4 = 3.5
    # 🚨 Patch E-1.8：corner_polygon 深度改用「街廓分配深度」（不是法定最小深度 14m）
    try:
        import streamlit as _st_B4_d
        _sb_rows_B4 = _st_B4_d.session_state.get('f3L_sb_rows_by_label', {}) or {}
        _this_blk_B4 = (_st_B4_d.session_state.get('f3_current_pk_block', '') or '')
        _blk_param_B4 = _sb_rows_B4.get(_this_blk_B4, {})
        _legal_depth_B4 = float(_blk_param_B4.get('街廓分配深度(m)', 0.0) or 0.0)
        if _legal_depth_B4 <= 0 and candidates:
            _legal_depth_B4 = float(
                candidates[0].get('平均深度(m)', 0)
                or candidates[0].get('街廓分配深度(m)', 0)
                or 14.0
            )
        if _legal_depth_B4 <= 0:
            _legal_depth_B4 = 14.0
    except Exception:
        _legal_depth_B4 = 14.0
    try:
        # 🚨 W-B 修正：同 setback 防呆寫法（避免 0 被 `or` 竄改）
        _lmw_raw = _blk_param_B4.get('法定最小寬(m)', 3.5)
        _legal_min_width_B4 = float(_lmw_raw) if _lmw_raw is not None else 3.5
    except Exception:
        _legal_min_width_B4 = 3.5

    # 🚨 W-B §5：讀 session 資料，以模組級 _build_corner_range_v2 建街角規定範圍
    try:
        import streamlit as _st_wb5
        _slbs_wb = _st_wb5.session_state.get('f3_cad_side_lines_by_side', {}) or {}
        _adir_wb = _st_wb5.session_state.get('f3_cad_alloc_dir', {}) or {}
        _side_wb = _slbs_wb.get(_this_blk_B4, {})
        _alloc_wb = _adir_wb.get(_this_blk_B4)
        _shift_wb = _setback_B4 + _legal_min_width_B4
        _cls_blks_wb = _st_wb5.session_state.get('f3_classified_blocks', []) or []
        _blk_meta_wb = next(
            (b for b in _cls_blks_wb if b.get('label') == _this_blk_B4), None
        )
        _blk_verts_wb = (_blk_meta_wb.get('vertices') or []) if _blk_meta_wb else []
        _blk_cen_wb = (_blk_meta_wb.get('centroid') or (0.0, 0.0)) if _blk_meta_wb else (0.0, 0.0)
        _chamfer_left_wb = _make_chamfer_tri_wb(_blk_meta_wb, 'left') if _blk_meta_wb else None
        _chamfer_right_wb = _make_chamfer_tri_wb(_blk_meta_wb, 'right') if _blk_meta_wb else None
    except Exception:
        _slbs_wb = {}; _adir_wb = {}; _side_wb = {}; _alloc_wb = None
        _blk_verts_wb = []; _blk_cen_wb = (0.0, 0.0)
        _shift_wb = _setback_B4 + _legal_min_width_B4
        _chamfer_left_wb = None; _chamfer_right_wb = None

    _corner_range_left = None; _corner_range_right = None
    if _alloc_wb and len(_blk_verts_wb) >= 3:
        if 'left' in _side_wb:
            _corner_range_left = _build_corner_range_v2(
                _side_wb['left']['mid'], _blk_verts_wb, _blk_cen_wb,
                _alloc_wb, _shift_wb, _chamfer_left_wb)
        if 'right' in _side_wb:
            _corner_range_right = _build_corner_range_v2(
                _side_wb['right']['mid'], _blk_verts_wb, _blk_cen_wb,
                _alloc_wb, _shift_wb, _chamfer_right_wb)
    if _alloc_wb is None and _this_blk_B4:
        try:
            import streamlit as _st_warn_wb
            _st_warn_wb.warning(
                f"街廓 {_this_blk_B4}：DXF 無 ALLOC_LINE，街角規定範圍以矩形近似"
            )
        except Exception:
            pass

    # 🚨 W-B §6：儲存面積供 UI 驗收
    try:
        import streamlit as _st_cr6
        _cr_areas_6 = _st_cr6.session_state.setdefault('f3_corner_range_areas', {})
        _cr_areas_6[_this_blk_B4] = {
            'left': (round(float(_corner_range_left.area), 2)
                     if _corner_range_left is not None else None),
            'right': (round(float(_corner_range_right.area), 2)
                      if _corner_range_right is not None else None),
        }
        # 🔬 W-C 唯讀診斷：存街角規定範圍多邊形座標（供 forced buffer 落點判定；
        #    純讀取存檔、不參與任何分配決策、不碰守恆）
        _cr_polys_6 = _st_cr6.session_state.setdefault('f3_corner_range_polys', {})
        _cr_polys_6[_this_blk_B4] = {
            'left': ([[float(c[0]), float(c[1])] for c in _corner_range_left.exterior.coords]
                     if _corner_range_left is not None and hasattr(_corner_range_left, 'exterior')
                     else None),
            'right': ([[float(c[0]), float(c[1])] for c in _corner_range_right.exterior.coords]
                      if _corner_range_right is not None and hasattr(_corner_range_right, 'exterior')
                      else None),
        }
    except Exception:
        pass

    # 向下相容：p1=左端、p2=右端（FRONT_LINE 方向約定）
    _corner_poly_p1_B4 = _corner_range_left
    _corner_poly_p2_B4 = _corner_range_right

    # 🆕 W-D.1.3-c：per 端 截角斜邊線 + SIDE_LINE（項一/項二 逐筆化用；分母=range 自身邊）
    _cham_p1 = _chamfer_line_for_side(_blk_meta_wb, 'left')
    _cham_p2 = _chamfer_line_for_side(_blk_meta_wb, 'right')
    _side_p1d = _side_wb.get('left') or {}
    _side_p2d = _side_wb.get('right') or {}
    _sl_p1 = (_side_p1d.get('p1'), _side_p1d.get('p2'))
    _sl_p2 = (_side_p2d.get('p1'), _side_p2d.get('p2'))
    # 分母＝range 多邊形自身之截角邊/側街邊（淺 T 已被 ALLOC@T 裁切；同源、算一次）
    _cut_den_p1 = (_boundary_len_on_line(_corner_poly_p1_B4, _cham_p1[0], _cham_p1[1])
                   if (_cham_p1 and _corner_poly_p1_B4 is not None) else 0.0)
    _cut_den_p2 = (_boundary_len_on_line(_corner_poly_p2_B4, _cham_p2[0], _cham_p2[1])
                   if (_cham_p2 and _corner_poly_p2_B4 is not None) else 0.0)
    _side_den_p1 = (_boundary_len_on_line(_corner_poly_p1_B4, _sl_p1[0], _sl_p1[1])
                    if (_sl_p1[0] and _corner_poly_p1_B4 is not None) else 0.0)
    _side_den_p2 = (_boundary_len_on_line(_corner_poly_p2_B4, _sl_p2[0], _sl_p2[1])
                    if (_sl_p2[0] and _corner_poly_p2_B4 is not None) else 0.0)

    # 🆕 W-D.1.3-c.1：街角端有 range 但截角斜邊線缺失/range 截角邊長≈0 → 具名警示
    #   （不靜默歸零、-a W1 同課；多為舊快取未入籍 tc['side'] 或 FRONT/SIDE 截角拓樸失據）
    try:
        import streamlit as _st_cw1
        for _end_lbl_cw, _rng_cw, _cham_cw, _cutden_cw in (
            ('左', _corner_poly_p1_B4, _cham_p1, _cut_den_p1),
            ('右', _corner_poly_p2_B4, _cham_p2, _cut_den_p2)):
            if _rng_cw is not None and (_cham_cw is None or _cutden_cw < 1e-6):
                _st_cw1.warning(
                    f"⚠️ 街廓 {_this_blk_B4} {_end_lbl_cw}街角端：截角斜邊線缺失或 range 截角邊長≈0"
                    f"（_cham={'None' if _cham_cw is None else 'OK'}、den={_cutden_cw:.4f}）→ 項一分母失據。"
                    f"請清 session cache 重跑（tc 側別入籍）或檢查 FRONT/SIDE 截角拓樸。"
                )
    except Exception:
        pass

    # 預先計算每個候選的 polygon 交集面積（依端別獨立計算）
    _corner_inter_area_p1 = {}   # id(cand) -> 與 p1 corner_polygon 之 inter area
    _corner_inter_area_p2 = {}
    _corner_cut_len_p1 = {}; _corner_cut_len_p2 = {}   # 🆕 W-D.1.3-c 臨截角長（逐候選）
    _side_line_len_p1 = {}; _side_line_len_p2 = {}     # 🆕 W-D.1.3-c 臨側街長（逐候選）
    for cand in candidates:
        coords = cand.get('polygon_coords') or []
        if len(coords) < 3:
            continue
        try:
            cp = _SP_B4(coords)
            if not cp.is_valid:
                cp = cp.buffer(0)
            if cp.is_empty:
                continue
        except Exception:
            continue
        if _corner_poly_p1_B4 is not None:
            try:
                inter1 = cp.intersection(_corner_poly_p1_B4)
                if not inter1.is_empty:
                    _corner_inter_area_p1[id(cand)] = float(inter1.area)
                    # 🆕 W-D.1.3-c 臨截角/臨側街長（additive、不改上兩行）
                    if _cham_p1:
                        _corner_cut_len_p1[id(cand)] = _boundary_len_on_line(inter1, _cham_p1[0], _cham_p1[1])
                    if _sl_p1[0]:
                        _side_line_len_p1[id(cand)] = _boundary_len_on_line(inter1, _sl_p1[0], _sl_p1[1])
            except Exception:
                pass
        if _corner_poly_p2_B4 is not None:
            try:
                inter2 = cp.intersection(_corner_poly_p2_B4)
                if not inter2.is_empty:
                    _corner_inter_area_p2[id(cand)] = float(inter2.area)
                    if _cham_p2:
                        _corner_cut_len_p2[id(cand)] = _boundary_len_on_line(inter2, _cham_p2[0], _cham_p2[1])
                    if _sl_p2[0]:
                        _side_line_len_p2[id(cand)] = _boundary_len_on_line(inter2, _sl_p2[0], _sl_p2[1])
            except Exception:
                pass

    # Step 1：依空間距離分流
    p1_group = []
    p2_group = []
    for cand in candidates:
        cen_raw = cand.get('centroid', (0.0, 0.0))
        try:
            cen = _np_v12.array(cen_raw, dtype=float)
        except Exception:
            continue
        d_to_p1 = float(_np_v12.linalg.norm(cen - p1))
        d_to_p2 = float(_np_v12.linalg.norm(cen - p2))
        cand_copy = dict(cand)
        cand_copy['_dist_to_p1'] = round(d_to_p1, 3)
        cand_copy['_dist_to_p2'] = round(d_to_p2, 3)

        # 🚨 Patch B-4 / E-1.7：實體交集分流（取代「質心投影距離」之單一判定）
        #   - 若候選 polygon 與 p1/p2 corner_polygon 有實體交集 → 加入該端組
        #   - 同時與兩端皆有交集 → 兩組都加入（PK 自然會在每端獨立判定）
        #   - 與兩端皆無交集 → E-1.7 排除出 PK 候選池（手冊圖 7 鐵律）
        _inter_p1_area = _corner_inter_area_p1.get(id(cand), 0.0)
        _inter_p2_area = _corner_inter_area_p2.get(id(cand), 0.0)
        _has_p1 = _inter_p1_area > 1.0   # 🆕 W-D.1.3-b E-1.7 門檻 0.01→1.0㎡（KL 絕對地板；剔 sliver 如 628-40 之 0.33）
        _has_p2 = _inter_p2_area > 1.0

        if _has_p1 or _has_p2:
            if _has_p1:
                cand_p1 = dict(cand_copy)
                cand_p1['cutoff_to_apply'] = float(cutoff_p1_end)
                cand_p1['min_area_to_apply'] = float(min_corner_area_p1)
                cand_p1['side_len_to_apply'] = float(base_side_len_m_p1)
                cand_p1['_assigned_end'] = 'p1'
                cand_p1['_dist_to_side_line'] = d_to_p1
                cand_p1['_dist_to_corner_point'] = d_to_p1
                cand_p1['_corner_pt'] = tuple(p1.tolist())
                cand_p1['_corner_intersection_area'] = round(_inter_p1_area, 2)
                cand_p1['_corner_range_area'] = (round(float(_corner_poly_p1_B4.area), 2)  # 🆕 W-D.1.3-b 項三分母（同 range 多邊形）
                                                 if _corner_poly_p1_B4 is not None else 0.0)
                # 🆕 W-D.1.3-c 項一/項二逐筆量（臨截角/臨側街長 + range 自身邊分母）
                cand_p1['_corner_cut_len'] = round(_corner_cut_len_p1.get(id(cand), 0.0), 3)
                cand_p1['_corner_cut_den'] = round(_cut_den_p1, 3)
                cand_p1['_side_line_len'] = round(_side_line_len_p1.get(id(cand), 0.0), 3)
                cand_p1['_side_line_den'] = round(_side_den_p1, 3)
                p1_group.append(cand_p1)
            if _has_p2:
                cand_p2 = dict(cand_copy)
                cand_p2['cutoff_to_apply'] = float(cutoff_p2_end)
                cand_p2['min_area_to_apply'] = float(min_corner_area_p2)
                cand_p2['side_len_to_apply'] = float(base_side_len_m_p2)
                cand_p2['_assigned_end'] = 'p2'
                cand_p2['_dist_to_side_line'] = d_to_p2
                cand_p2['_dist_to_corner_point'] = d_to_p2
                cand_p2['_corner_pt'] = tuple(p2.tolist())
                cand_p2['_corner_intersection_area'] = round(_inter_p2_area, 2)
                cand_p2['_corner_range_area'] = (round(float(_corner_poly_p2_B4.area), 2)  # 🆕 W-D.1.3-b 項三分母（同 range 多邊形）
                                                 if _corner_poly_p2_B4 is not None else 0.0)
                # 🆕 W-D.1.3-c 項一/項二逐筆量（臨截角/臨側街長 + range 自身邊分母）
                cand_p2['_corner_cut_len'] = round(_corner_cut_len_p2.get(id(cand), 0.0), 3)
                cand_p2['_corner_cut_den'] = round(_cut_den_p2, 3)
                cand_p2['_side_line_len'] = round(_side_line_len_p2.get(id(cand), 0.0), 3)
                cand_p2['_side_line_den'] = round(_side_den_p2, 3)
                p2_group.append(cand_p2)
            # 已用實體交集分流 → skip 後續質心距離 fallback
            cand_copy = None
        else:
            # 🚨 Patch E-1.7：無實體交集 → 排除出 PK 候選池
            # 手冊圖 7 鐵律：「重劃前土地座落於街角最小分配面積範圍內」才能當街角候選
            # 舊 fallback「依距離分流」會把不該是街角的候選強塞進 PK
            cand_copy = None

    # 🆕 W-D.1.3-b：刪 15m 框（_compute_corner_overlap_area）。項三改「真交集 / range 面積」（見 _pk_one_side_v12）；
    #   physical_overlap_area 不再被覆寫、不再作評分來源；候選 _corner_intersection_area / _corner_range_area 為項三真量。

    # Step 2：每組獨立 PK
    return {
        'p1_end': _pk_one_side_v12(p1_group, g_values_map, base_front_len_m),
        'p2_end': _pk_one_side_v12(p2_group, g_values_map, base_front_len_m),
    }


def _pk_one_side_v12(group: list, g_values_map: dict,
                      base_front_len_m: float) -> dict:
    """
    🆕 V12 單側 PK：第一關 G 值門檻 → 第二關優先權指數
    🚨 強化點 2：第一關用 G 值（非物理跨占面積）
    """
    if not group:
        return {'winner': None, 'qualified': [], 'eliminated': [],
                'note': '該端無候選', 'group_size': 0}

    qualified = []
    eliminated = []
    # 第一關：G 值門檻
    for cand in group:
        # 優先讀 g_values_map；fallback 至 cand 內 G_value 或 G_estimated
        cand_G = (g_values_map.get(cand.get('暫編地號', ''))
                   or cand.get('G_value')
                   or cand.get('G_estimated', 0))
        try:
            cand_G = float(cand_G or 0)
        except Exception:
            cand_G = 0.0
        cand['G_for_threshold'] = round(cand_G, 2)
        if cand_G < cand.get('min_area_to_apply', 0):
            cand['eliminated_reason'] = (
                f"G 值 {cand_G:.2f} < 最小分配面積 {cand.get('min_area_to_apply', 0):.2f}"
            )
            eliminated.append(cand)
        else:
            qualified.append(cand)

    if not qualified:
        return {'winner': None, 'qualified': [], 'eliminated': eliminated,
                'note': '🚫 無人達 G 值門檻 → 該端保留為抵費地',
                'group_size': len(group)}

    # 🆕 V13 修正 #3 + #7：第二關優先權指數採用三項法定加權
    # 項一（0.4）：候選宗地與「道路截角線 corner_cut_line」之重疊長度比
    # 項二（0.2）：候選宗地與「SIDE_LINE 評比區內」之重疊長度比
    # 項三（0.4）：候選宗地與「虛擬評比框」之物理交集面積比
    # ※ 對個別 cand 之原始幾何指標（front_length / side_length / physical_overlap_area）
    #   呼叫端在分流時應已計算並寫入 cand 各欄位（fallback 至既有欄位）
    for cand in qualified:
        # 🆕 W-D.1.3-c 手冊三指數逐筆（單一路徑；已刪 V13 死欄位 + front_length/side_length fallback）：
        #   項一=0.4×(臨截角/range 截角邊)、項二=0.2×(臨側街/range 側街邊)、項三=0.4×(真交集/range 面積)；
        #   分子⊆分母（皆取自 parcel∩range vs range 自身邊/面積）→ 值域 ∈[0,權重]。
        _cc_len = float(cand.get('_corner_cut_len', 0) or 0)
        _cc_den = max(float(cand.get('_corner_cut_den', 0) or 0), 1e-9)
        _sl_len = float(cand.get('_side_line_len', 0) or 0)
        _sl_den = max(float(cand.get('_side_line_den', 0) or 0), 1e-9)
        _tj_inter = float(cand.get('_corner_intersection_area', 0) or 0)
        _tj_range = max(float(cand.get('_corner_range_area', 0) or 0), 1e-9)

        score_corner_cut = 0.4 * min(_cc_len / _cc_den, 1.0)   # 項一
        score_side = 0.2 * min(_sl_len / _sl_den, 1.0)         # 項二
        score_overlap = 0.4 * min(_tj_inter / _tj_range, 1.0)  # 項三（-b）

        cand['_score_corner_cut'] = round(score_corner_cut, 6)
        cand['_score_side'] = round(score_side, 6)
        cand['_score_overlap'] = round(score_overlap, 6)
        cand['priority_index'] = round(
            score_corner_cut + score_side + score_overlap, 6
        )

    # 🆕 Phase 8 Issue 3 + 防護一：法定排序主鍵不能用「距角點」為第一鍵
    # （否則大塊合法街角地會被微小畸零地淘汰）
    # 🆕 W-D.2 v2 轉正（§2 tiebreaker 換源，KL 放行 2026-07-05）：pi 同分(4dp) →
    #   重劃前**正典原位次**小者勝（_pre_position_rank＝_projection_order 投影序，
    #   呼叫端寫入；單一真相源）。廢距角序暫行近似（-c 引入、TODO(W-D §2) 就此消滅）。
    qualified.sort(key=lambda c: (
        -round(float(c.get('priority_index', 0)), 4),                 # 1: 手冊優先權指數（4dp 判平手）
        float(c.get('_pre_position_rank', float('inf'))),             # 2: §2 正典原位次（投影序 rank）
    ))
    return {'winner': qualified[0], 'qualified': qualified,
            'eliminated': eliminated, 'note': '✅ PK 完成',
            'group_size': len(group)}


def _compute_per_end_cutoff_areas(block_meta: dict,
                                   front_line_p1: tuple,
                                   front_line_p2: tuple) -> dict:
    """
    🆕 W-D.1.3-d：從街廓 geom_restore 之 theoretical_corners，依 tc['side'] 分配
    p1（左）/ p2（右）兩端各自之截角面積（Σ 該側 cutoff_area_m2；同側多截角則加總）。

    側別直讀 tc['side']（_rebuild_corners_topology 出生入籍、正典＝side_lines_by_side）；
    廢除舊「corner→FRONT p1/p2 最近距離猜側」——單截角塊會把同一 tc 同時指給兩端（雙端同值），
    且繞向近似會翻 side、診斷會說謊（與 -c.1 修掉之猜側同族）。
    front_line_p1/p2 保留於簽名（呼叫端相容）但不再參與側別分配。tc 無 side（舊快取）→ 不計。

    回傳 {
      'p1_end': {'cutoff_area': float, 'theoretical_corner': (x, y) or None},
      'p2_end': {'cutoff_area': float, 'theoretical_corner': (x, y) or None},
    }
    """
    res = {
        'p1_end': {'cutoff_area': 0.0, 'theoretical_corner': None},
        'p2_end': {'cutoff_area': 0.0, 'theoretical_corner': None},
    }
    geom_restore = (block_meta or {}).get('geom_restore') or {}
    tcs = geom_restore.get('theoretical_corners') or []
    if not tcs or front_line_p1 is None or front_line_p2 is None:
        return res
    # p1_end＝左（side=='left'）、p2_end＝右（side=='right'）；corner 欄取該側首個 tc.corner
    _end_of_side = {'left': 'p1_end', 'right': 'p2_end'}
    for tc in tcs:
        _end = _end_of_side.get(tc.get('side'))
        if _end is None:
            continue   # tc 無 side（舊快取）→ 不猜、該截角不計
        res[_end]['cutoff_area'] += float(tc.get('cutoff_area_m2', 0) or 0)
        if res[_end]['theoretical_corner'] is None:
            res[_end]['theoretical_corner'] = tc.get('corner')
    return res


def _extract_corner_cut_line(block_poly,
                              virtual_corner_pt,
                              front_line_pts,
                              side_line_pts,
                              parallel_tol_deg: float = 8.0,
                              min_corner_cut_len: float = 1.5,
                              max_corner_cut_len: float = 15.0) -> dict:
    """
    🆕 V13 Task AF：自動萃取「道路截角線」(Corner Cut Line)

    定義：實體街廓 polygon 邊界上，距離虛擬尖角 V_c 最近、
          且 **不與 FRONT_LINE 或 SIDE_LINE 平行**、長度在合理範圍之線段。

    🛡️ 防護 1（V13 + 後續強化）：
      - 平行容差由 5° 放寬至 **8°**（CAD 繪製誤差容忍）
      - 新增**長度過濾**：截角線長度需介於 [1.5m, 15m]；> 15m 極可能是主街線

    參數：
      block_poly           實體街廓 shapely Polygon
      virtual_corner_pt    虛擬尖角點 V_c (x, y)
      front_line_pts       FRONT_LINE 端點 [(x1,y1), (x2,y2)] 用於平行度判斷
      side_line_pts        SIDE_LINE 端點（同上）
      parallel_tol_deg     平行容差（< 此度數視為平行 → 排除）；預設 8.0
      min_corner_cut_len   截角線最小可接受長度（m）；預設 1.5
      max_corner_cut_len   截角線最大可接受長度（m）；預設 15.0

    回傳：
      {
        'line':       shapely LineString | None,
        'length':     float（截角線長度）,
        'p1', 'p2':   端點座標,
        'angle_deg':  與 X 軸夾角,
      }
    """
    import math as _m_ccl, numpy as _np_ccl
    from shapely.geometry import LineString as _LS_ccl

    if block_poly is None or virtual_corner_pt is None:
        return {'line': None, 'length': 0.0,
                'p1': None, 'p2': None, 'angle_deg': 0.0}

    Vc = _np_ccl.array(virtual_corner_pt, dtype=float)

    # 1. 計算 FRONT / SIDE 之方向向量（單位）
    def _unit_dir_ccl(pts):
        if not pts or len(pts) < 2:
            return None
        dx = pts[-1][0] - pts[0][0]; dy = pts[-1][1] - pts[0][1]
        L = (dx ** 2 + dy ** 2) ** 0.5
        if L < 1e-6:
            return None
        return _np_ccl.array([dx, dy]) / L
    fd = _unit_dir_ccl(front_line_pts)
    sd = _unit_dir_ccl(side_line_pts)

    # 2. 解析 block polygon 邊界線段
    coords = list(block_poly.exterior.coords)
    if coords and coords[0] == coords[-1]:
        coords = coords[:-1]
    if len(coords) < 3:
        return {'line': None, 'length': 0.0,
                'p1': None, 'p2': None, 'angle_deg': 0.0}

    # 3. 評分每邊
    candidates = []
    cos_tol = _m_ccl.cos(_m_ccl.radians(parallel_tol_deg))
    for i in range(len(coords)):
        p1 = _np_ccl.array(coords[i], dtype=float)
        p2 = _np_ccl.array(coords[(i + 1) % len(coords)], dtype=float)
        edge_vec = p2 - p1
        edge_len = float(_np_ccl.linalg.norm(edge_vec))
        if edge_len < 0.01:
            continue
        # 🛡️ 防護 1：長度過濾（[1.5m, 15m]）
        if edge_len < min_corner_cut_len or edge_len > max_corner_cut_len:
            continue
        edge_dir = edge_vec / edge_len

        # 平行度檢查
        is_parallel_to_front = (
            fd is not None
            and abs(float(_np_ccl.dot(edge_dir, fd))) > cos_tol
        )
        is_parallel_to_side = (
            sd is not None
            and abs(float(_np_ccl.dot(edge_dir, sd))) > cos_tol
        )
        if is_parallel_to_front or is_parallel_to_side:
            continue   # 排除平行邊

        mid = (p1 + p2) / 2.0
        d_to_vc = float(_np_ccl.linalg.norm(mid - Vc))
        candidates.append({
            'p1': tuple(p1.tolist()), 'p2': tuple(p2.tolist()),
            'length': edge_len, 'd_to_vc': d_to_vc,
            'angle_deg': _m_ccl.degrees(_m_ccl.atan2(edge_vec[1], edge_vec[0])),
        })

    if not candidates:
        return {'line': None, 'length': 0.0,
                'p1': None, 'p2': None, 'angle_deg': 0.0}

    # 距離 V_c 最近者 = corner_cut_line
    best = min(candidates, key=lambda c: c['d_to_vc'])
    return {
        'line': _LS_ccl([best['p1'], best['p2']]),
        'length': best['length'],
        'p1': best['p1'], 'p2': best['p2'],
        'angle_deg': best['angle_deg'],
    }


def _estimate_G_for_qualification(a_m2: float,
                                    tolerance_m2: float = 0.5) -> float:
    """
    🆕 V12 修正 #4 + 🛡️ V13 防護 2：街角地第一關門檻用之估算 G 值

    🛡️ 防護 2 (a)：對接 Tab 7 (財務分析) 之「重劃總負擔比率」
      公式：G_est = a × (1 − 重劃總負擔比率)
      不需手動傳入 B / C — 直接讀 session_state['f3_total_burden_rate_from_finance']
      （由 Tab 7 之 B + C 寫入；fallback 0.40 當 Tab 7 尚未填）

    🛡️ 防護 2 (b)：加 +0.5 ㎡ 容差（防浮點誤殺地主）

    參數：
      a_m2          重劃前宗地面積（含已整併之公設）
      tolerance_m2  防誤殺容差（預設 +0.5 ㎡）

    回傳：估算 G 值 + 容差
    """
    burden = float(st.session_state.get(
        'f3_total_burden_rate_from_finance', 0.40) or 0.40)
    burden = max(0.0, min(0.95, burden))
    G_est = float(a_m2) * (1.0 - burden)
    return G_est + float(tolerance_m2)


def assign_corner_side_by_projection(parcel_centroid: tuple,
                                      block_front_line) -> str:
    """
    依 LineString.project() 投影距離判定宗地落於街廓正面道路的左／右半段
      - 投影距離 < 正面長度 / 2 → '左側'
      - 否則 → '右側'
    回傳 '左側' / '右側' / '' （失敗時）
    """
    try:
        from shapely.geometry import Point
        if block_front_line is None or block_front_line.length <= 0:
            return ''
        p = Point(parcel_centroid[0], parcel_centroid[1])
        proj = float(block_front_line.project(p))
        return '左側' if proj < block_front_line.length / 2.0 else '右側'
    except Exception:
        return ''


# ============ 任務四：共同／非共同負擔公共設施用地分配 ============
# 與 Tab 1 歸戶連動之公設用地分配模組
# 涵蓋三類：道路、廣場/公園/市場、非共同負擔（機關/停車場）

def allocate_road_to_adjacent_owners(road_area_m2: float,
                                     ownership_residential_holdings: dict,
                                     min_area_m2: float = 0.0,
                                     road_poly=None,
                                     centerlines: list = None) -> dict:
    """
    道路用地共同負擔分配（任務六：幾何驅動版）

    若提供 road_poly + centerlines → 用 shapely.ops.split 沿中心線切道路 →
    依切出的左右兩半 polygon 真實面積比例分配；
    若未提供或 split 失敗 → 退回算術等分（半半）。

    分配邏輯：
      1) 切道路 → left_area / right_area
      2) 各歸戶按其在「左/右側街廓」之持分比例分得對應半邊面積
      3) 一側無持分 → 該側不分；如另一側有持分，將該歸戶整份併入另一側
      4) 若分得面積 < min_area_m2 → 集中向另一側（防呆）
      5) 兩側均無持分 → 標記為跨街廓調配

    參數：
      road_area_m2                      此段道路面積（㎡）
      ownership_residential_holdings    {歸戶群組: {'left': 左側面積, 'right': 右側面積}}
      min_area_m2                       最小分配面積閾值
      road_poly                         (可選) shapely Polygon — 道路範圍
      centerlines                       (可選) list[LineString] — 任務一中心線輸出

    回傳：
      {
        'allocations': [{'歸戶': gid, '左側併入面積': , '右側併入面積': , '備註': ''}],
        'cross_block_needed': [gid, ...],
        'method': '幾何切割' | '算術等分',
        'left_area': float,  # 實際左側面積（㎡）
        'right_area': float,
      }
    """
    # 1. 嘗試幾何切割
    left_area = right_area = road_area_m2 / 2.0
    method = '算術等分'
    if road_poly is not None and centerlines:
        try:
            cut = cut_road_by_centerline(road_poly, centerlines)
            if cut.get('method') == 'split':
                la = float(cut['left'].area) if cut.get('left') is not None else 0.0
                ra = float(cut['right'].area) if cut.get('right') is not None else 0.0
                # 校驗：兩半合計需接近原面積（避免誤切）
                if la + ra > 0 and abs((la + ra) - road_area_m2) / max(road_area_m2, 1e-6) < 0.20:
                    left_area, right_area = la, ra
                    method = '幾何切割'
        except Exception:
            pass

    # 2. 計算各側總持分量（用於比例分配）
    holdings = ownership_residential_holdings or {}
    total_left = sum(max(float(s.get('left', 0.0) or 0.0), 0.0) for s in holdings.values())
    total_right = sum(max(float(s.get('right', 0.0) or 0.0), 0.0) for s in holdings.values())

    allocations = []
    cross_block = []
    for gid, sides in holdings.items():
        left = float(sides.get('left', 0.0) or 0.0)
        right = float(sides.get('right', 0.0) or 0.0)

        # 比例分配
        l_share = (left / total_left) * left_area if (left > 0 and total_left > 0) else 0.0
        r_share = (right / total_right) * right_area if (right > 0 and total_right > 0) else 0.0

        if left > 0 and right > 0:
            row = {
                '歸戶': gid,
                '左側併入面積': round(l_share, 2),
                '右側併入面積': round(r_share, 2),
                '備註': f'{method}：兩側皆有土地，按持分比例分配'
            }
        elif left > 0:
            row = {
                '歸戶': gid,
                '左側併入面積': round(l_share, 2),
                '右側併入面積': 0.0,
                '備註': f'{method}：右側無土地 → 全部併入左側'
            }
        elif right > 0:
            row = {
                '歸戶': gid,
                '左側併入面積': 0.0,
                '右側併入面積': round(r_share, 2),
                '備註': f'{method}：左側無土地 → 全部併入右側'
            }
        else:
            row = {
                '歸戶': gid,
                '左側併入面積': 0.0,
                '右側併入面積': 0.0,
                '備註': f'{method}：兩側皆無土地 → 標記為調配至他街廓'
            }
            cross_block.append(gid)
        allocations.append(row)

        # 防呆：分得面積 < min → 集中向另一側
        if min_area_m2 > 0:
            la = row['左側併入面積']; ra = row['右側併入面積']
            if 0 < la < min_area_m2 and ra > 0:
                row['右側併入面積'] = round(ra + la, 2)
                row['左側併入面積'] = 0.0
                row['備註'] += '；左側未達最小 → 集中向右'
            elif 0 < ra < min_area_m2 and la > 0:
                row['左側併入面積'] = round(la + ra, 2)
                row['右側併入面積'] = 0.0
                row['備註'] += '；右側未達最小 → 集中向左'

    return {
        'allocations': allocations,
        'cross_block_needed': cross_block,
        'method': method,
        'left_area': round(left_area, 2),
        'right_area': round(right_area, 2),
    }


def allocate_plaza_to_adjacent_residential(plaza_area_m2: float,
                                            ownership_residential_holdings: dict,
                                            min_area_m2: float = 0.0) -> dict:
    """
    廣場 / 鄰里公園 / 零售市場 共同負擔分配
    若地主在「相鄰住宅區」內有其他土地 → 將公設用地上的面積合併於該住宅區分配。
    若無 → 標記「調配至他街廓」。

    參數：
      plaza_area_m2                  此公設面積（㎡）
      ownership_residential_holdings {歸戶: 該歸戶於相鄰住宅區之土地面積總和}
      min_area_m2                    最小分配面積閾值
    """
    allocations = []
    cross_block = []
    valid = {gid: a for gid, a in (ownership_residential_holdings or {}).items()
             if float(a or 0.0) > 0}
    if not valid:
        return {'allocations': [],
                'cross_block_needed': list((ownership_residential_holdings or {}).keys()),
                'notes': '無相鄰住宅區歸戶 → 全數標記調配至他街廓'}

    total_resi = sum(valid.values())
    for gid, resi_area in valid.items():
        share_ratio = resi_area / total_resi if total_resi > 0 else 0
        share = round(plaza_area_m2 * share_ratio, 2)
        note = f'依相鄰住宅區面積比 {share_ratio:.2%} 合併於住宅區分配'
        if min_area_m2 > 0 and share < min_area_m2:
            note += f'；併入後 {share:.2f} ㎡ < 最小 {min_area_m2:.2f} ㎡ → 建議現金補償'
        allocations.append({
            '歸戶': gid, '相鄰住宅區面積': round(resi_area, 2),
            '分配面積': share, '備註': note,
        })

    # 找出沒列入 valid 的歸戶（即無相鄰住宅區）
    for gid in (ownership_residential_holdings or {}):
        if gid not in valid:
            cross_block.append(gid)

    return {'allocations': allocations, 'cross_block_needed': cross_block,
            'notes': f'依相鄰住宅區面積比例分配，共 {len(valid)} 個歸戶受配'}


def allocate_non_common_public_land(public_land_area_m2: float,
                                     public_owned_supply_m2: float,
                                     offset_land_supply_m2: float,
                                     private_holdings: dict,
                                     min_area_m2: float = 0.0) -> dict:
    """
    非共同負擔公共設施用地分配（機關用地、停車場）
    優先指配公有地（縣市有 → 鄉鎮有 → 國有，本函式僅以「公有地總供給量」處理）。
    若公有地或抵費地不足指配 → 依該公設範圍內「原私有土地所有權人歸戶面積比例」按比例發還，
    此發還面積不受最小分配面積限制。
    剩餘被排擠的私有地，依歸戶調配至其他街廓。

    參數：
      public_land_area_m2     公共設施用地需指配總量
      public_owned_supply_m2  公有地可供量
      offset_land_supply_m2   抵費地可供量
      private_holdings        {歸戶: 該歸戶於此公設範圍內之原私有面積}
      min_area_m2             最小分配面積閾值（此處供參考；按比例發還不受限制）

    回傳：
      {
        'public_used': 指配公有地量,
        'offset_used': 指配抵費地量,
        'pro_rata_returns': [{'歸戶':, '原私有面積':, '按比例發還':}],
        'shortfall': 不足量,
        'reallocations': [被排擠歸戶清單],
      }
    """
    public_supply = max(0.0, float(public_owned_supply_m2 or 0.0))
    offset_supply = max(0.0, float(offset_land_supply_m2 or 0.0))
    needed = max(0.0, float(public_land_area_m2 or 0.0))

    # 1. 先用公有地
    public_used = min(needed, public_supply)
    needed -= public_used
    # 2. 再用抵費地
    offset_used = min(needed, offset_supply)
    needed -= offset_used
    shortfall = needed  # 剩下需按比例發還之量

    pro_rata_returns = []
    reallocations = []
    if shortfall > 0:
        total_private = sum(float(a or 0.0) for a in (private_holdings or {}).values())
        if total_private > 0:
            for gid, area in private_holdings.items():
                a = float(area or 0.0)
                if a <= 0:
                    continue
                ratio = a / total_private
                rtn = round(shortfall * ratio, 2)
                pro_rata_returns.append({
                    '歸戶': gid, '原私有面積': round(a, 2),
                    '按比例發還': rtn,
                    '備註': f'按比例發還 {ratio:.2%}（不受最小分配面積限制）'
                })
        else:
            # 無原私有 → 全數標為需調配
            reallocations = list((private_holdings or {}).keys())

    return {
        'public_used': round(public_used, 2),
        'offset_used': round(offset_used, 2),
        'pro_rata_returns': pro_rata_returns,
        'shortfall': round(shortfall, 2),
        'reallocations': reallocations,
        'notes': (f'指配公有地 {public_used:.2f} ㎡、抵費地 {offset_used:.2f} ㎡，'
                  f'剩餘 {shortfall:.2f} ㎡按私有歸戶比例發還')
    }


# ============ W-G G.1：§7 七級調配引擎接線層（單一真相源＝verify/wf_f0~f4；禁 fork）============
# 裁定（KL 2026-07-12）：①範圍＝UC9898 複現（非本案旗標告知不硬跑）；
# ②財務 ctx＝β 混源（live 幾何/宗地/Step-G ＋ 快照 財務接線_v3），加誠實圍欄。
# ns＝app 真符號（禁 harvest()，會 re-exec app.py 再入）；缺鍵/缺符號一律 loud（no-silent-fallback）。
_WF_NS_NAMES = [
    "F3_CATEGORY_BURDEN", "calc_B_value", "calc_C_value", "calc_special_burden_total",
    "solve_G_binary", "iterate_G_S", "rw_from_width", "get_min_lot_size",
    "_select_pool_slot", "_projection_order", "_spatial_order_parcels_v2",
    "alloc_normal_axis", "_block_strip",
]


def _wf_ns():
    """引擎經 ns 呼叫之 13 app 真符號 dict（直取本模組 globals，非 harvest）。"""
    g = globals()
    ns = {}
    for n in _WF_NS_NAMES:
        if n not in g:
            raise RuntimeError(f"🔴 七級調配接線：app 缺真符號 {n}（引擎單一真相源不完整）")
        ns[n] = g[n]
    return ns


class _WFSessionShim:
    """曝 .session_state（引擎唯一觸及之 fake_st 介面）；淺拷貝隔離、不污染 live st.session_state。
    未知屬性回 no-op（鏡射 harness _FakeStreamlit，防引擎他路觸及）。"""
    def __init__(self, ss_like):
        object.__setattr__(self, "session_state", dict(ss_like))

    def __getattr__(self, _name):
        def _noop(*a, **k):
            return None
        return _noop


def _wf_load_snapshot(app_file):
    """β：財務 ctx 之凍結來源＝verify/case_params_UC9898.json（裁定②）。回傳完整 snapshot dict。"""
    import os
    import json
    p = os.path.join(os.path.dirname(os.path.abspath(app_file)), "verify", "case_params_UC9898.json")
    with open(p, encoding="utf-8") as f:
        return json.load(f)


def _build_wf_ctx(ss, tag, app_file=__file__):
    """由 live session_state 組單一情境 wf 引擎 ctx（β：live 幾何/宗地/Step-G ＋ 快照財務）。
    ss＝session_state-like dict（app 傳 st.session_state；harness +1 gate 傳 seed dict）。
    tag∈{'0m','3.5m'}。缺鍵 loud。回傳引擎所需 14 欄 ctx。"""
    def _need(k):
        if k not in ss:
            raise RuntimeError(f"🔴 七級調配接線：session 缺 {k}（Step G 未完成或未曝值？）")
        return ss[k]

    # cad：7 鍵 → 引擎單一 cad dict（**含 centerlines**，否則 wf_f3 STRADDLE 宗 KeyError）
    cad = {
        "front_lengths":        _need("f3_cad_front_lengths"),
        "side_lengths_by_side": _need("f3_cad_side_lengths_by_side"),
        "front_lines":          _need("f3_cad_front_lines"),
        "side_lines_by_side":   ss.get("f3_cad_side_lines_by_side", {}) or {},
        "alloc_dir_by_block":   ss.get("f3_cad_alloc_dir", {}) or {},
        "centerlines":          ss.get("f3_manual_road_centerlines", {}) or {},
    }

    # cb_by：list → {label: blk}
    cb_by = {b["label"]: b for b in _need("f3_classified_blocks")}

    # snap：β 混源＝快照 財務接線_v3/global（凍結）＋ live blocks 幾何
    snap_full = _wf_load_snapshot(app_file)
    sb_rows = {r["街廓"]: r for r in _need("f3_sb_rows")}
    depth_by = _need("f3_alloc_depth_by_label")
    mw_by = _need("f3_min_width_by_label")
    blocks = {}
    for lbl in cb_by:
        r = sb_rows.get(lbl)
        if r is None:
            continue  # 非可建築街廓（無臨街負擔列）：引擎僅讀可建築塊
        blocks[lbl] = {
            "正面": {"路寬_m": float(r["正面路寬(m)"]), "負擔尺度_輸入": float(r["正街尺度"]),
                     "期望長度_m": float(r["正面長度(m)"]), "期望負擔面積_m2": float(r["正面面積(㎡)"])},
            "左側": {"路寬_m": float(r["左側路寬(m)"]), "負擔尺度_輸入": float(r["左側尺度"]),
                     "期望長度_m": float(r["左側長度(m)"]), "期望負擔面積_m2": float(r["左側面積(㎡)"])},
            "右側": {"路寬_m": float(r["右側路寬(m)"]), "負擔尺度_輸入": float(r["右側尺度"]),
                     "期望長度_m": float(r["右側長度(m)"]), "期望負擔面積_m2": float(r["右側面積(㎡)"])},
            "街廓分配深度_m": float(depth_by.get(lbl, 0.0)),
            "法定最小寬_m": float(mw_by.get(lbl, 0.0)),
        }
    snap = {"財務接線_v3": snap_full["財務接線_v3"], "global": snap_full["global"], "blocks": blocks}

    return {
        "ns": _wf_ns(),
        "fake_st": _WFSessionShim(ss),
        "cb_by": cb_by,
        "cad": cad,
        "snap": snap,
        "omap": _need("t8_ownership_map"),
        "build": _need("f3_build_parcels"),
        "temp": _need("f3_temp_parcels"),
        "params": _need("f3L_corner_min_table"),
        "winners": _need("f3_corner_winners"),
        "forced": _need("f3L_forced_offset"),
        "setback": float(_need("f3L_setback_default")),
        "gA": _need("f3_G_values"),
        "poolA": _need("f3_wd2_pool_diag"),
    }


def _wf_tag_of(setback):
    """退縮 float → 引擎情境 tag。硬對應 UC9898 雙情境；非 {0,3.5} 時 loud。"""
    if abs(float(setback) - 0.0) < 1e-9:
        return "0m"
    if abs(float(setback) - 3.5) < 1e-9:
        return "3.5m"
    raise RuntimeError(f"🔴 七級調配：退縮 {setback} 非 UC9898 雙情境（0/3.5），引擎凍結錨不適用")


def _is_uc9898(ss):
    """裁定①：§7 引擎為 UC9898 凍結真相。以穩定指紋判本案（R1-R6 街廓 ＋ 33 歸戶群）。"""
    try:
        blocks = {b.get("label") for b in (ss.get("f3_classified_blocks") or [])}
        groups = ss.get("t8_ownership_groups") or {}
        temp = ss.get("f3_temp_parcels") or []
        return ({"R1", "R2", "R3", "R4", "R5", "R6"} <= blocks
                and len(groups) == 33 and len(temp) > 0)
    except Exception:
        return False


# ============ 主程式 ============
def main():
    st.title("🏗️ 市地重劃地價估算系統")
    st.caption("依據：第四號公報 (113.12.06) / 花蓮縣對應高雄公會利潤率 / 土地開發分析法公式：V = S÷(1+R)÷(1+i)-(C+M)")
    
    # Tab 名稱常數（dispatch 改用字串常數，重排順序時不會打破路由）
    TAB_LAND_AGGREGATION = "🏘️ 土地歸戶"
    TAB_LAND_TRADE       = "🏠 土地交易分析"
    TAB_BUILDING_TRADE   = "🏢 房地交易分析"
    TAB_BLOCK_INTERACT   = "🗂️ 街廓互動分析"
    TAB_LAND_DEV         = "📐 土地開發分析法"
    TAB_EXPECTED_DEV     = "📋 預期開發分析法"
    TAB_PRICE_ZONE       = "🗺️ 地價區段分析"
    TAB_FINANCE          = "💰 財務分析"
    # 顯示順序（土地歸戶為首，依任務一規範）
    TAB_NAMES = [
        TAB_LAND_AGGREGATION,
        TAB_LAND_TRADE,
        TAB_BUILDING_TRADE,
        TAB_BLOCK_INTERACT,
        TAB_LAND_DEV,
        TAB_EXPECTED_DEV,
        TAB_PRICE_ZONE,
        TAB_FINANCE,
    ]
    with st.sidebar:
        st.header("📋 功能選單")
        selected_tab = st.radio("選擇功能", TAB_NAMES, key="sel_tab", label_visibility="collapsed")
        st.markdown("---")
        st.caption("市地重劃地價估算系統 v2.0")

    # 🆕 Phase 11：Tab 切換保留各分頁瀏覽位置
    # 透過 JavaScript 將每個 tab 的 scrollY 寫入 sessionStorage，
    # 切換到該 tab 時自動還原至原瀏覽位置
    import streamlit.components.v1 as _components_scroll
    _scroll_key = f'tab_scroll_{selected_tab}'
    _scroll_html = f"""
    <script>
    (function() {{
        const TAB_KEY = {repr(_scroll_key)};
        const ALL_TABS_KEY = 'tab_scroll_all';
        // 在 streamlit iframe 環境中要操作的是 parent 視窗
        const w = window.parent || window;
        const d = w.document;

        // 取得目前 scrollY → 儲存至 sessionStorage（離開 tab 時呼叫）
        function _save() {{
            try {{
                w.sessionStorage.setItem(TAB_KEY, String(w.scrollY || 0));
            }} catch (e) {{}}
        }}

        // 載入時：嘗試還原本 tab 之上次 scrollY（延遲 200ms 等 Streamlit 渲染完成）
        function _restore() {{
            try {{
                const v = w.sessionStorage.getItem(TAB_KEY);
                if (v !== null) {{
                    w.scrollTo({{ top: parseInt(v, 10) || 0, left: 0, behavior: 'instant' }});
                }}
            }} catch (e) {{}}
        }}

        // 滾動時持續更新（debounced 200ms）
        let _t = null;
        w.addEventListener('scroll', function() {{
            if (_t) clearTimeout(_t);
            _t = setTimeout(_save, 200);
        }}, {{ passive: true }});

        // Streamlit rerun 後 DOM 重建 → 多階段嘗試還原
        setTimeout(_restore, 100);
        setTimeout(_restore, 400);
        setTimeout(_restore, 800);
    }})();
    </script>
    """
    try:
        _components_scroll.html(_scroll_html, height=0, width=0)
    except Exception:
        pass

    # ─────────────────────────────────────────────────────────────────
    # 跨 Tab session_state 持久化（影子鍵 _sv_*）
    # 解決：Streamlit 在分頁切換時，未渲染之 widget（如 Tab 5 的 b_cost/k_cost/c_cost）
    #      在多次 rerun 後可能被自動清除 → 切回 Tab 7 顯示時就會回到 fallback 預設。
    # 解法：每次 rerun 起始即把現存 widget 值複製到影子鍵；若 widget 鍵已被清除，
    #      則用影子鍵還原回去，使下次 widget 渲染時延用使用者輸入。
    # ─────────────────────────────────────────────────────────────────
    _persist_keys = (
        # Tab 5：開發成本三項（B 工程、K 拆遷、C 重劃費）
        'b_cost', 'k_cost', 'c_cost',
        # Tab 5：利息計算 / 成熟度修正
        'n_months', 'n_prime', 'maturity_r_ed', 'maturity_n_ed',
        # Tab 5：四項面積（已有 _sv_*，此處保險再做一次）
        'area_offset', 'area_buildable', 'area_public_shared', 'area_public_non_shared',
        # Tab 7：開發年期、貸款利率、實際抵費地面積
        'total_years', 'loan_rate', 'disp_area',
    )
    for _pk in _persist_keys:
        _sv_key = f'_sv_{_pk}'
        if _pk in st.session_state:
            # widget 鍵還在 → 同步至影子鍵
            st.session_state[_sv_key] = st.session_state[_pk]
        elif _sv_key in st.session_state:
            # widget 鍵已被 Streamlit 清除 → 從影子鍵還原
            st.session_state[_pk] = st.session_state[_sv_key]

    # ===== Tab: 土地交易 =====
    if selected_tab == TAB_LAND_TRADE:
        st.markdown("### 功能：上傳土地交易Excel → 地價指數調整 → 土地單價")
        st.caption("📌 僅上傳「土地」交易案件（交易標的為「土地」），程式僅讀取「案件列表」工作表")
        c1, c2 = st.columns([1, 2])
        with c1:
            district1 = _sb("地區", list(LAND_PRICE_INDEX_DATA.keys()), key="d1")
            pd1 = _di("價格日期", date.today(), key="pd1")
            pd1_dt = datetime.combine(pd1, datetime.min.time())
            pi1 = get_interpolated_index(pd1_dt, district1)
            st.info(f"價格日期指數：**{pi1:.2f}**")
        with c2:
            st.markdown("**上傳土地交易Excel（非房地交易）**")
            files1 = st.file_uploader("選擇檔案", type=['xls', 'xlsx'], accept_multiple_files=True, key="f1")
            # 上傳時存入 session_state；切回此頁時從快取還原
            if files1:
                st.session_state['f1_saved'] = [(f.name, f.getvalue()) for f in files1]
            elif st.session_state.get('f1_saved'):
                files1 = [_NamedBytesIO(n, d) for n, d in st.session_state['f1_saved']]
                st.caption("📂 已從快取載入上次上傳的檔案")

        if files1:
            df1 = read_land_excel(files1)
            if df1.empty: 
                st.error("無有效資料")
                st.info("💡 提示：請確認上傳的是「土地」交易案件，不是「房地(土地+建物)」交易案件")
            else:
                st.success(f"✅ {len(df1)} 筆")
                res1 = []
                for _, r in df1.iterrows():
                    adj = calculate_price_adjustment(r['交易日期'], pd1_dt, district1)
                    aup = r['單價(元/㎡)'] * adj['factor']
                    res1.append({'交易日期指數': adj['trans_index'], '價格日期指數': adj['price_index'], '調整率': round(adj['factor'], 4), '調整後單價(元/㎡)': round(aup, 0), '調整後單價(元/坪)': round(aup / 0.3025, 0)})
                df1_res = pd.concat([df1, pd.DataFrame(res1)], axis=1)
                st.dataframe(df1_res, use_container_width=True)
                tot_ar = df1_res['面積(㎡)'].sum()
                wp_sqm = (df1_res['調整後單價(元/㎡)'] * df1_res['面積(㎡)']).sum() / tot_ar if tot_ar > 0 else df1_res['調整後單價(元/㎡)'].mean()
                wp_ping = wp_sqm / 0.3025
                st.metric("加權平均單價", f"{wp_sqm:,.0f} 元/㎡ = {wp_ping:,.0f} 元/坪")
                st.session_state['res_land_sqm'] = wp_sqm
                st.session_state['res_land_ping'] = wp_ping
    
    # ===== Tab: 房地交易 =====
    elif selected_tab == TAB_BUILDING_TRADE:
        st.markdown("### 功能：上傳房地交易Excel → 價格調整+折舊反推 → 新建成屋單價")
        st.caption("📌 上傳「房地(土地+建物)」交易案件，程式會同時讀取「案件列表」、「土地」、「建物」三個工作表")
        st.caption("⚠️ 屋齡為空或無有效建築完成日期的案件將被剔除")
        
        # ===== 房地分析參數（直接顯示於頁面內，不再放於側邊欄） =====
        st.markdown("#### ⚙️ 房地分析參數")
        param_col1, param_col2, param_col3 = st.columns(3)
        with param_col1:
            district2 = _sb("地區", list(LAND_PRICE_INDEX_DATA.keys()), key="d2")
            pd2 = _di("價格日期", date.today(), key="pd2")
            pd2_dt = datetime.combine(pd2, datetime.min.time())
            pi2 = get_interpolated_index(pd2_dt, district2)
            st.info(f"價格日期指數：**{pi2:.2f}**")
        with param_col2:
            lp_from1 = st.session_state.get('res_land_sqm', 0)
            use_lp1 = _cb("使用土地分析結果", value=lp_from1 > 0, key="ulp")
            lp_input = lp_from1 if use_lp1 and lp_from1 > 0 else _ni("土地單價(元/㎡)", min_value=1000, max_value=5000000, value=50000, step=1000, key="lpi")
            if use_lp1 and lp_from1 > 0:
                st.success(f"土地單價：{lp_from1:,.0f} 元/㎡")
        with param_col3:
            age_filter = _cb("啟用屋齡篩選", key="af")
            max_age = _ni("屋齡上限(年)", min_value=1, max_value=100, value=10, step=1, key="ma") if age_filter else None
        
        st.markdown("---")
        st.markdown("**上傳房地交易Excel（土地+建物）**")
        files2 = st.file_uploader("選擇檔案", type=['xls', 'xlsx'], accept_multiple_files=True, key="f2")
        # 上傳時存入 session_state；切回此頁時從快取還原
        if files2:
            st.session_state['f2_saved'] = [(f.name, f.getvalue()) for f in files2]
        elif st.session_state.get('f2_saved'):
            files2 = [_NamedBytesIO(n, d) for n, d in st.session_state['f2_saved']]
            st.caption("📂 已從快取載入上次上傳的檔案")
        if files2:
            df2, all_land_areas = read_building_excel(files2)
            if df2.empty: st.error("無有效資料（或全被剔除）")
            else:
                st.success(f"✅ {len(df2)} 筆有效資料")
                
                # === 全部上傳案例的土地移轉面積統計 ===
                if all_land_areas:
                    all_la_arr = np.array(all_land_areas)
                    st.markdown("#### 📊 土地移轉面積統計")
                    col_s1, col_s2 = st.columns(2)
                    with col_s1:
                        st.markdown(f"**全部上傳案例（{len(all_la_arr)} 筆）**")
                        st.markdown(f"- 平均值：**{all_la_arr.mean():,.2f} ㎡**（{all_la_arr.mean() * 0.3025:,.2f} 坪）")
                        st.markdown(f"- 中位數：**{np.median(all_la_arr):,.2f} ㎡**（{np.median(all_la_arr) * 0.3025:,.2f} 坪）")
                
                # 保存篩選前的有效資料土地面積統計
                df2_all_valid = df2.copy()
                
                if age_filter and max_age:
                    df2 = df2[df2['屋齡'] <= max_age].copy()
                    st.info(f"篩選後 {len(df2)} 筆")
                
                # === 篩選後案例的土地移轉面積統計 ===
                if age_filter and max_age and not df2.empty and all_land_areas:
                    filtered_la = df2['土地面積(㎡)'].values
                    filtered_la = filtered_la[filtered_la > 0]
                    if len(filtered_la) > 0:
                        with col_s2:
                            st.markdown(f"**屋齡篩選後（{len(filtered_la)} 筆）**")
                            st.markdown(f"- 平均值：**{filtered_la.mean():,.2f} ㎡**（{filtered_la.mean() * 0.3025:,.2f} 坪）")
                            st.markdown(f"- 中位數：**{np.median(filtered_la):,.2f} ㎡**（{np.median(filtered_la) * 0.3025:,.2f} 坪）")
                
                if df2.empty: st.warning("篩選後無資料")
                else:
                    res2 = []
                    for _, r in df2.iterrows():
                        ba, la, tp, age, el, rr = r['建物面積(㎡)'], r['土地面積(㎡)'], r['總價(元)'], r['屋齡'], r['經濟耐用年數'], r['殘餘價格率']
                        adj = calculate_price_adjustment(r['交易日期'], pd2_dt, district2)
                        atp = tp * adj['factor']
                        lv = la * lp_input
                        bpv = max(0, atp - lv)
                        ea = min(age, el) if age > 0 else 0
                        dr = (1 - rr) * (ea / el) if el > 0 else 0
                        brc = bpv / (1 - dr) if dr < 1 else bpv
                        ntp = lv + brc
                        nup = ntp / ba / 0.3025 if ba > 0 else 0
                        nup_sqm = ntp / ba if ba > 0 else 0
                        res2.append({
                            '調整率': round(adj['factor'], 4), 
                            '調整後總價': round(atp, 0), 
                            '土地價值': round(lv, 0), 
                            '折舊率(%)': round(dr * 100, 2), 
                            '新建單價(元/坪)': round(nup, 0),
                            '新建單價(元/㎡)': round(nup_sqm, 0)
                        })
                    df2_res = pd.concat([df2.reset_index(drop=True), pd.DataFrame(res2)], axis=1)
                    st.dataframe(df2_res[['編號', '位置', '屋齡', '主要建材', '總價(元)', '土地面積(㎡)', '調整率', '折舊率(%)', '新建單價(元/坪)', '新建單價(元/㎡)']], use_container_width=True)
                    tot_ba = df2_res['建物面積(㎡)'].sum()
                    wnp = (df2_res['新建單價(元/坪)'] * df2_res['建物面積(㎡)']).sum() / tot_ba if tot_ba > 0 else df2_res['新建單價(元/坪)'].mean()
                    wnp_sqm = (df2_res['新建單價(元/㎡)'] * df2_res['建物面積(㎡)']).sum() / tot_ba if tot_ba > 0 else df2_res['新建單價(元/㎡)'].mean()
                    st.metric("加權平均新建單價", f"{wnp:,.0f} 元/坪 = {wnp_sqm:,.0f} 元/㎡")
                    rn, cc_min, cc_max = get_construction_cost_range(wnp)
                    st.info(f"對應營造費區間：**{rn}** → {cc_min:,.0f}~{cc_max:,.0f} 元/坪")
                    st.session_state['res_np'] = wnp
                    st.session_state['res_cc'] = (cc_min + cc_max) / 2  # 存中間值
    
    # ===== Tab: 土地開發分析法 =====
    elif selected_tab == TAB_LAND_DEV:
        st.markdown("### 功能：依第4號公報土地開發分析法計算土地價格")
        st.markdown("**公式：V = S ÷ (1+R) ÷ (1+i) - (C+M)**")
        st.caption("V:土地價格 | S:總銷售金額 | R:利潤率 | i:資本利息綜合利率 | C:直接成本 | M:間接成本")
        c_l, c_r = st.columns([1, 2])
        with c_l:
            st.subheader("📐 基本資料")
            land_ar = _ni("基地面積(㎡)", min_value=1.0, max_value=100000.0, value=500.0, step=10.0, key="la")
            bc_pct = _ni("法定建蔽率(%)", min_value=1.0, max_value=100.0, value=60.0, step=1.0, key="bc")
            far_pct = _ni("法定容積率(%)", min_value=1.0, max_value=1500.0, value=200.0, step=10.0, key="far")
            st.markdown("---")
            st.subheader("💰 銷售單價")
            def_sp = st.session_state.get('res_np', 0)
            if def_sp > 0:
                st.success(f"來自房地分析：{def_sp:,.0f} 元/坪")
                use_sp = _cb("使用", True, key="usp")
                sp = def_sp if use_sp else _ni("手動輸入", min_value=10000, max_value=5000000, value=int(def_sp), step=10000, key="msp")
            else:
                sp = _ni("銷售單價(元/坪)", min_value=10000, max_value=5000000, value=200000, step=10000, key="sp")
            st.markdown("---")
            st.subheader("🔧 營造單價")
            rn, cc_min, cc_max = get_construction_cost_range(sp)
            cc_mid = (cc_min + cc_max) / 2
            st.info(f"區間：{rn} → {cc_min:,.0f}~{cc_max:,.0f} 元/坪")
            base_cc = _sl("營造單價(元/坪)", int(cc_min), int(cc_max), int(cc_mid), 1000, key="bcc")

            # ===== 營造工程物價指數 API 帶入 =====
            st.markdown("**📊 營造工程物價總指數（主計總處 SDMX API）**")
            st.caption("資料源：政府資料開放平臺 A030502015 項目 131（建築工程總指數），113年11月為第11號公報建議基期")
            api_col1, api_col2 = st.columns([1, 2])
            if api_col1.button("🔄 從 API 帶入最新指數", key="btn_fetch_cci", use_container_width=True):
                with st.spinner("呼叫主計總處 API 中…"):
                    cci = fetch_construction_price_index()
                st.session_state['_cci_cache'] = cci
                if cci.get('success'):
                    b_val = cci['by_roc'].get('113年11月')
                    latest = cci['data'][-1] if cci['data'] else None
                    # ⚠️ 關鍵：必須同時修改「widget key（bi/ci）」本身，
                    # 否則 Streamlit 會忽略 value 參數，沿用舊 widget state。
                    if b_val:
                        st.session_state['bi'] = float(b_val)
                        st.session_state['_sv_bi'] = float(b_val)
                    if latest:
                        st.session_state['ci'] = float(latest['value'])
                        st.session_state['_sv_ci'] = float(latest['value'])
                        st.session_state['_cci_latest_label'] = latest['roc']
                    st.success(
                        f"✅ 已取得 {len(cci['data'])} 個月份指數；"
                        f"基期 113/11 = {b_val}、最新 {latest['roc'] if latest else '-'} = {latest['value'] if latest else '-'}"
                    )
                    st.rerun()
                else:
                    st.error(f"API 呼叫失敗：{cci.get('error')}")
            with api_col2:
                _cci = st.session_state.get('_cci_cache', {})
                if _cci.get('success') and _cci.get('data'):
                    rocs = [d['roc'] for d in _cci['data']]
                    options = ['(手動輸入)'] + rocs
                    # 若已有選定月份則記住；否則預設最新月
                    last_pick = st.session_state.get('_cci_month_pick', rocs[-1])
                    try:
                        default_i = options.index(last_pick)
                    except ValueError:
                        default_i = len(options) - 1
                    sel_ci = st.selectbox(
                        "選擇價格日期月份（自動帶入右側欄位）",
                        options, index=default_i, key="cci_month_sel"
                    )
                    if sel_ci != '(手動輸入)':
                        _val = _cci['by_roc'].get(sel_ci)
                        cur = st.session_state.get('ci', None)
                        if _val is not None and (cur is None or abs(float(cur) - float(_val)) > 1e-6):
                            st.session_state['ci'] = float(_val)
                            st.session_state['_sv_ci'] = float(_val)
                            st.session_state['_cci_month_pick'] = sel_ci
                            st.rerun()

            base_idx = _ni("基期物價指數(113.11)", min_value=50.0, max_value=200.0, value=100.0, step=0.1, key="bi")
            curr_idx = _ni("價格日期物價指數", min_value=50.0, max_value=200.0, value=100.0, step=0.1, key="ci")
            pir = curr_idx / base_idx if base_idx > 0 else 1.0
            st.success(f"調整後營造單價：{base_cc * pir:,.0f} 元/坪")

            # API 取得後以表格呈現全部月份
            _cci_disp = st.session_state.get('_cci_cache', {})
            if _cci_disp.get('success') and _cci_disp.get('data'):
                with st.expander("📋 檢視 113/1 ~ 115/3 營造工程總指數全月明細", expanded=False):
                    _df_cci = pd.DataFrame(_cci_disp['data'])[['roc', 'ym', 'value']]
                    _df_cci.columns = ['民國', '西元月份', '指數']
                    st.dataframe(_df_cci, use_container_width=True, hide_index=True, height=280)
                    st.caption(f"資料擷取時間：{_cci_disp.get('fetched_at', '')}")
            st.markdown("---")
            st.subheader("📊 管銷費率")
            asr = _sl("廣告+銷售費率(%)", FEE_RATES["廣告費銷售費"]["min"], FEE_RATES["廣告費銷售費"]["max"], FEE_RATES["廣告費銷售費"]["default"], 0.5, key="asr")
            mr = _sl("管理費率(%)", FEE_RATES["管理費"]["min"], FEE_RATES["管理費"]["max"], FEE_RATES["管理費"]["default"], 0.25, key="mr")
            tr = _sl("稅捐費率(%)", FEE_RATES["稅捐"]["min"], FEE_RATES["稅捐"]["max"], FEE_RATES["稅捐"]["default"], 0.1, key="tr")
            st.markdown("---")
            st.subheader("💵 綜合利息資本利率")
            st.caption("參考吉安重劃報告書表4-1推算方式")
            bm = _ni("建物興建工期 M1 (月)", min_value=1, max_value=120, value=18, step=1, key="bm")
            lm = _ni("土地持有時期 M2 (月)", min_value=1, max_value=120, value=24, step=1, key="lm")
            dr_i = _ni("自有資金利率 A (%) - 五大銀行一年期定存", min_value=0.0, max_value=20.0, value=1.70, step=0.01, key="dri")
            lr_i = _ni("借貸資金利率 B (%) - 五大銀行放款利率", min_value=0.0, max_value=20.0, value=3.26, step=0.01, key="lri")
            st.markdown("**建築投資**（自有+借貸=100%）")
            be = _sl("自有比例 C1 (%)", min_value=0.0, max_value=100.0, value=50.0, step=5.0, key="be")
            bl = 100.0 - be  # 自動計算
            st.info(f"借貸比例 D1：{bl:.0f}%（自動計算）")
            st.markdown("**土地投資**（自有+借貸=100%）")
            le = _sl("自有比例 C2 (%)", min_value=0.0, max_value=100.0, value=50.0, step=5.0, key="le")
            ll = 100.0 - le  # 自動計算
            st.info(f"借貸比例 D2：{ll:.0f}%（自動計算）")
            st.markdown("**綜合利息權重**（建築+土地=100%）")
            bw = _sl("建築利息 C3 (%)", min_value=0.0, max_value=100.0, value=48.0, step=1.0, key="bw")
            lw = 100.0 - bw  # 自動計算
            st.info(f"土地利息 D3：{lw:.0f}%（自動計算）")
            int_res = calc_composite_interest(bm, lm, dr_i, lr_i, be, bl, le, ll, bw, lw)
            # 顯示計算明細
            with st.expander("📋 計算明細", expanded=True):
                st.markdown(f"""
                **建築投資資本利息** = (C1×A + D1×B) ÷ 12 × M1 × 0.5
                = ({be:.0f}%×{dr_i}% + {bl:.0f}%×{lr_i}%) ÷ 12 × {bm} × 0.5 = **{int_res['building_int']:.3f}%**
                
                **土地投資資本利息** = (C2×A + D2×B) ÷ 12 × M2
                = ({le:.0f}%×{dr_i}% + {ll:.0f}%×{lr_i}%) ÷ 12 × {lm} = **{int_res['land_int']:.3f}%**
                
                **綜合利息資本利率** = C3×建築利息 + D3×土地利息
                = {bw:.0f}%×{int_res['building_int']:.3f}% + {lw:.0f}%×{int_res['land_int']:.3f}% = **{int_res['composite']:.3f}%**
                """)
            st.success(f"綜合利息資本利率 i = {int_res['composite']:.2f}%")
            # 儲存綜合利率（年利率）到 session_state 供預期開發分析法使用
            # 計算年利率：將總工期內的複利利率換算為年利率
            total_dev_months = (bm + lm) / 2  # 平均開發工期
            if total_dev_months > 0:
                annual_rate = int_res['composite'] / (total_dev_months / 12)  # 年利率
            else:
                annual_rate = int_res['composite']
            st.session_state['composite_annual_rate'] = annual_rate
            st.session_state['composite_interest_rate'] = int_res['composite']
            st.markdown("---")
            st.subheader("📈 利潤率 (高雄公會)")
            st.caption("花蓮縣對應高雄公會區域利潤率（第四號公報附表二）")
            dev_m = (bm + lm) / 2
            pn, pr_info = get_profit_rate_by_period(dev_m)
            st.info(f"開發工期：{dev_m:.1f}月={dev_m/12:.2f}年 → {pn} ({pr_info['min']}%~{pr_info['max']}%)")
            pr = _sl("利潤率(%)", float(pr_info["min"]), float(pr_info["max"]), float(pr_info["default"]), 0.5, key="pr")
            st.markdown("---")
            st.subheader("🔄 成熟度修正")
            st.caption("重劃後地價經成熟度修正後為特定價格（參考第十一號公報）")
            enable_maturity = _cb("啟用成熟度折現修正", value=False, key="em")
            if enable_maturity:
                st.markdown("""
                **公式：特定價格 = 正常價格 ÷ (1+K)^n**
                - K：成熟度修正率（折現率）
                - n：成熟度修正年期（滯待期間）
                """)
                maturity_K = _ni("成熟度修正率 K (%)", min_value=0.0, max_value=30.0, value=8.63, step=0.1, key="mk")
                maturity_n = _ni("成熟度修正年期 n (年)", min_value=0, max_value=20, value=6, step=1, key="mn")
                st.info(f"折現係數：1/(1+{maturity_K/100:.4f})^{maturity_n} = {1/((1+maturity_K/100)**maturity_n):.4f}")
            else:
                maturity_K, maturity_n = 0, 0
        
        with c_r:
            st.subheader("🧮 計算結果")
            bv = calc_building_volume(land_ar, bc_pct, far_pct)
            if bv:
                ld = calc_land_dev(bv, sp, base_cc, asr, mr, tr, int_res['composite'], pr, pir)
                
                # 建築規劃詳細表格
                st.markdown("#### 📐 建築規劃計算依據")
                
                # 從 bv 取得計算結果
                gf_area = bv['gf_sqm']  # 一樓面積(㎡)
                uf_area = bv['uf_sqm']  # 二樓以上主建物面積(㎡)
                bal_area = bv['bal_sqm']  # 附屬建物面積(㎡)
                bal_per_floor = bv['bal_per_floor']  # 每層陽台面積
                roof_area = bv['roof_sqm']  # 屋頂突出物面積(㎡)
                actual_bc = bv['actual_bc_pct']  # 實際建蔽率(%)
                
                # 計算屋頂突出物的說明
                roof_pct_val = gf_area * 0.125
                roof_calc_note = f"max(地面層面積×12.5%, 25㎡) = max({roof_pct_val:.2f}, 25) = {roof_area:.2f}㎡"
                
                st.markdown(f"""
                | 項目 | 計算式 | 面積(㎡) | 面積(坪) |
                |------|--------|----------|----------|
                | 基地面積 | 輸入值 | {land_ar:,.2f} | {land_ar * 0.3025:,.2f} |
                | 法定建蔽率 | 輸入值 | {bc_pct:.0f}% | - |
                | 法定容積率 | 輸入值 | {far_pct:.0f}% | - |
                | **推估樓層數** | ceil({far_pct:.0f}% ÷ {bc_pct:.0f}%) | **{bv['floors']}層** | - |
                | 實際建蔽率 | {far_pct:.0f}% ÷ {bv['floors']} | {actual_bc:.2f}% | - |
                | **一樓面積（地面層面積）** | 基地 × 實際建蔽率 | {gf_area:,.2f} | {gf_area * 0.3025:,.2f} |
                | **二樓以上主建物面積** | 基地 × 容積率 - 一樓 | {uf_area:,.2f} | {uf_area * 0.3025:,.2f} |
                | **附屬建物面積** | max(每層×10%, 5㎡) × {bv['floors']}層 | {bal_area:,.2f} | {bal_area * 0.3025:,.2f} |
                | **屋頂突出物** | max(地面層×12.5%, 25㎡) | {roof_area:,.2f} | {roof_area * 0.3025:,.2f} |
                | **總樓地板面積** | 一樓+二樓以上+附屬+屋突 | **{bv['total_sqm']:,.2f}** | **{bv['total_ping']:,.2f}** |
                """)
                
                with st.expander("📋 計算說明", expanded=False):
                    st.markdown(f"""
                    **附屬建物面積計算依據**：
                    - 每層陽台面積 = max(該層樓地板面積×10%, 5㎡)
                    - 每層樓地板面積 = {gf_area:,.2f}㎡
                    - 每層陽台 = max({gf_area:,.2f}×10%, 5) = max({gf_area*0.1:,.2f}, 5) = **{bal_per_floor:,.2f}㎡**
                    - 總附屬建物面積 = {bal_per_floor:,.2f} × {bv['floors']}層 = **{bal_area:,.2f}㎡**
                    
                    **屋頂突出物計算依據**：
                    - 屋頂突出物 = max(地面層面積×12.5%, 25㎡)
                    - {roof_calc_note}
                    """)
                
                st.info(f"**總建坪：{bv['total_ping']:,.2f} 坪** | 土地面積：{bv['land_ping']:,.2f} 坪")
                
                st.markdown("---")
                st.markdown(f"""
                #### 土地開發分析法
                **公式：V = S ÷ (1+R) ÷ (1+i) - (C+M)**
                
                | 項目 | 說明 | 金額/比率 |
                |------|------|----------|
                | **S** 總銷售金額 | {bv['total_ping']:,.2f}坪 × {sp:,.0f}元/坪 | {ld['S']:,.0f} 元 |
                | **C** 直接成本(營造費) | {bv['total_ping']:,.2f}坪 × {ld['acc']:,.0f}元/坪 | {ld['C']:,.0f} 元 |
                | **M** 間接成本合計 | | {ld['M']:,.0f} 元 |
                | 　├ 規劃設計費 | 營造費×2% | {ld['M_design']:,.0f} 元 |
                | 　├ 廣告銷售費 | 總銷售×{asr}% | {ld['M_ad']:,.0f} 元 |
                | 　├ 管理費 | 總銷售×{mr}% | {ld['M_mg']:,.0f} 元 |
                | 　└ 稅捐 | 總銷售×{tr}% | {ld['M_tx']:,.0f} 元 |
                | **R** 利潤率 | | {ld['R_pct']:.2f}% |
                | **i** 資本利息綜合利率 | | {ld['i_pct']:.2f}% |
                | C+M 總成本 | | {ld['C'] + ld['M']:,.0f} 元 |
                | S÷(1+R)÷(1+i) | 折現後銷售額 | {ld['S'] / (1 + ld['R_pct']/100) / (1 + ld['i_pct']/100):,.0f} 元 |
                | **V 土地總價（正常價格）** | 土地面積：{ld['land_sqm']:.2f}㎡ | **{ld['V']:,.0f} 元** |
                | **土地單價（正常價格）** | {ld['V']:,.0f} ÷ {ld['land_sqm']:.2f}㎡ | **{ld['V_sqm']:,.0f} 元/㎡ = {ld['V_ping']:,.0f} 元/坪** |
                """)
                
                # 決定最終使用的土地開發分析價格（正常價格，不含成熟度修正）
                dev_V_sqm = ld['V_sqm']
                dev_V_ping = ld['V_ping']
                
                # 顯示正常價格
                m1, m2, m3 = st.columns(3)
                m1.metric("土地總價", f"{ld['V']:,.0f} 元")
                m2.metric("土地單價(㎡)", f"{ld['V_sqm']:,.0f} 元/㎡")
                m3.metric("土地單價(坪)", f"{ld['V_ping']:,.0f} 元/坪")
                
                # 加權計算功能
                st.markdown("---")
                st.subheader("⚖️ 綜合評估（加權計算）")
                
                # 取得土地交易分析的結果
                land_trans_sqm = st.session_state.get('res_land_sqm', 0)
                land_trans_ping = st.session_state.get('res_land_ping', 0)
                
                if land_trans_ping > 0:
                    st.markdown(f"""
                    | 估價方法 | 土地單價（正常價格） |
                    |----------|----------|
                    | **A. 土地交易分析** | {land_trans_sqm:,.0f} 元/㎡ = {land_trans_ping:,.0f} 元/坪 |
                    | **B. 土地開發分析** | {dev_V_sqm:,.0f} 元/㎡ = {dev_V_ping:,.0f} 元/坪 |
                    """)
                    
                    st.markdown("**加權比例設定**（兩者合計須為100%）")
                    weight_A = _sl(
                        "土地交易分析權重 (%)", 
                        min_value=0, max_value=100, value=60, step=5,
                        key="weight_a",
                        help="調整此滑桿，土地開發分析權重會自動調整"
                    )
                    weight_B = 100 - weight_A  # 自動計算，確保合計100%
                    st.info(f"土地開發分析權重：{weight_B}%（自動計算）")
                    
                    # 加權計算（使用正常價格）
                    weighted_sqm = (weight_A / 100) * land_trans_sqm + (weight_B / 100) * dev_V_sqm
                    weighted_ping = (weight_A / 100) * land_trans_ping + (weight_B / 100) * dev_V_ping
                    
                    st.markdown(f"""
                    **加權計算公式**：
                    
                    綜合單價 = {weight_A}% × A + {weight_B}% × B
                    = {weight_A}% × {land_trans_sqm:,.0f} + {weight_B}% × {dev_V_sqm:,.0f}
                    = **{weighted_sqm:,.0f} 元/㎡ = {weighted_ping:,.0f} 元/坪**
                    """)
                    
                    w1, w2 = st.columns(2)
                    w1.metric("綜合評估單價(㎡)", f"{weighted_sqm:,.0f} 元/㎡")
                    w2.metric("綜合評估單價(坪)", f"{weighted_ping:,.0f} 元/坪")
                    
                    # 顯示差異比較
                    diff_from_trans = weighted_ping - land_trans_ping
                    diff_from_dev = weighted_ping - dev_V_ping
                    st.caption(f"與土地交易分析差異：{diff_from_trans:+,.0f} 元/坪 ({diff_from_trans/land_trans_ping*100:+.1f}%) | 與土地開發分析差異：{diff_from_dev:+,.0f} 元/坪 ({diff_from_dev/dev_V_ping*100:+.1f}%)")
                    
                    # 成熟度修正：套用在綜合評估結果上
                    if enable_maturity and maturity_n > 0:
                        discount_factor = 1 / ((1 + maturity_K / 100) ** maturity_n)
                        weighted_sqm_special = weighted_sqm * discount_factor
                        weighted_ping_special = weighted_ping * discount_factor
                        
                        st.markdown(f"""
                        ---
                        #### 成熟度折現修正
                        **公式：特定價格 = 綜合評估正常價格 ÷ (1+K)^n**
                        
                        | 項目 | 說明 | 金額 |
                        |------|------|------|
                        | 綜合評估正常價格 | | {weighted_sqm:,.0f} 元/㎡ = {weighted_ping:,.0f} 元/坪 |
                        | K 修正率 | | {maturity_K:.2f}% |
                        | n 修正年期 | | {maturity_n} 年 |
                        | 折現係數 | 1/(1+K)^n | {discount_factor:.4f} |
                        | **綜合評估特定價格** | {weighted_sqm:,.0f} × {discount_factor:.4f} | **{weighted_sqm_special:,.0f} 元/㎡ = {weighted_ping_special:,.0f} 元/坪** |
                        """)
                        ws1, ws2 = st.columns(2)
                        ws1.metric("綜合評估特定價格(㎡)", f"{weighted_sqm_special:,.0f} 元/㎡")
                        ws2.metric("綜合評估特定價格(坪)", f"{weighted_ping_special:,.0f} 元/坪")
                        st.caption("※ 特定價格 = 重劃後土地之特定價格（綜合評估後再考量成熟度修正）")
                        
                        # 儲存成熟度修正後的價格
                        st.session_state['weighted_price_sqm'] = weighted_sqm_special
                        st.session_state['weighted_price_ping'] = weighted_ping_special
                    else:
                        # 儲存綜合評估價格（正常價格）
                        st.session_state['weighted_price_sqm'] = weighted_sqm
                        st.session_state['weighted_price_ping'] = weighted_ping
                    
                    st.session_state['dev_land_price_sqm'] = dev_V_sqm  # 土地開發分析法正常價格
                    st.session_state['dev_land_price_ping'] = dev_V_ping
                else:
                    st.warning("⚠️ 請先在「土地交易分析」頁籤完成分析，才能進行加權計算")
                    st.caption("提示：在土地交易分析頁籤上傳土地交易資料並完成分析後，結果會自動帶入此處進行加權計算")
            else:
                st.error("請確認建蔽率與容積率")

    # ===== Tab: 預期開發分析法 =====
    elif selected_tab == TAB_EXPECTED_DEV:
        st.markdown("### 功能：依預期開發分析法計算重劃前土地價格")
        st.markdown("""
        **公式：X = { A × f － [ ( B ＋ K ) ( 1 + n × p ) + C ] } × 1/(1+n'×p) × 1/(1+r)^N × D**
        """)
        st.caption("X:重劃前素地價格 | A:重劃後土地價格 | f:有效宅地化率 | B:開發工事費 | K:拆遷補償費 | C:間接費用")
        st.caption("n:開發成本利息月數 | n':土地成本利息月數 | p:月利率 | r:成熟度修正利率 | N:成熟度修正年期 | D:個別因素修正率")
        
        c_l, c_r = st.columns([1, 2])
        
        with c_l:
            # ===== 價格日期設定 =====
            st.subheader("📅 價格日期設定")
            st.caption("設定重劃前後價格日期，用於地價指數調整")
            
            district_ed = _sb("地區", list(LAND_PRICE_INDEX_DATA.keys()), key="district_ed")
            
            st.markdown("**重劃後價格日期**")
            post_date_ed = _di("重劃後價格日期", date.today(), key="post_date_ed", label_visibility="collapsed")
            post_date_dt = datetime.combine(post_date_ed, datetime.min.time())
            post_index_ed = get_interpolated_index(post_date_dt, district_ed)
            st.info(f"重劃後價格日期指數：**{post_index_ed:.2f}**")
            
            st.markdown("**重劃前價格日期**")
            pre_date_ed = _di("重劃前價格日期", date(2024, 1, 1), key="pre_date_ed", label_visibility="collapsed")
            pre_date_dt = datetime.combine(pre_date_ed, datetime.min.time())
            pre_index_ed = get_interpolated_index(pre_date_dt, district_ed)
            st.info(f"重劃前價格日期指數：**{pre_index_ed:.2f}**")
            
            st.markdown("---")
            # ===== 取用計算結果 =====
            st.subheader("📊 取用計算結果")
            
            # 取得重劃後素地價格(A)
            weighted_sqm = st.session_state.get('weighted_price_sqm', 0)
            if weighted_sqm > 0:
                st.success(f"A 重劃後土地價格：{weighted_sqm:,.0f} 元/㎡（來自綜合評估）")
                use_weighted = _cb("使用綜合評估結果", value=True, key="use_weighted_a")
                if use_weighted:
                    A_value = weighted_sqm
                else:
                    A_value = _ni("手動輸入 A (元/㎡)", min_value=1000.0, max_value=10000000.0, value=weighted_sqm, step=1000.0, key="manual_a")
            else:
                st.warning("⚠️ 請先在「土地開發分析法」頁籤完成綜合評估")
                A_value = _ni("手動輸入 A (元/㎡)", min_value=1000.0, max_value=10000000.0, value=50000.0, step=1000.0, key="manual_a2")
            
            # 取得綜合年利率並轉換為月利率(p)
            # 從「土地開發分析法」頁籤取得利率
            composite_rate = st.session_state.get('composite_interest_rate', 0)
            annual_rate = st.session_state.get('composite_annual_rate', 0)
            
            st.markdown("---")
            st.subheader("💵 利率設定 (p)")
            if annual_rate > 0:
                st.info(f"綜合利息資本利率：{composite_rate:.3f}% (來自土地開發分析)")
                # 重新計算簡化的年利率：使用自有+借貸資金的加權年利率
                st.caption("💡 預期開發法需使用簡單月利率，建議直接輸入五大銀行利率的加權平均")
            
            # 直接讓使用者輸入年利率
            p_annual = _ni("資金年利率 (%)", min_value=0.0, max_value=20.0, value=3.0, step=0.01, key="p_annual",
                                       help="建議使用五大銀行自有+借貸資金加權年利率")
            p_monthly = p_annual / 12  # 轉換為月利率
            st.success(f"月利率 p = {p_annual:.2f}% ÷ 12 = **{p_monthly:.4f}%**")
            
            st.markdown("---")
            st.subheader("📐 土地面積")
            st.caption("輸入重劃區各類土地面積（單位：平方公尺）")

            # 自動讀取 Tab 3（街廓互動分析）推送之面積（若已推送）
            _f3_areas = st.session_state.get('f3_to_tab_areas')
            if _f3_areas:
                st.success(
                    f"✅ 已自動帶入 街廓互動分析 街廓分類結果：可建築 {_f3_areas.get('buildable_total', 0):,.2f} ㎡、"
                    f"共同負擔 {_f3_areas.get('public_common_total', 0):,.2f} ㎡、"
                    f"非共同負擔 {_f3_areas.get('public_non_common_total', 0):,.2f} ㎡、"
                    f"抵充地 {_f3_areas.get('offset_area', 0):,.2f} ㎡"
                )
                _def_offset = float(_f3_areas.get('offset_area', 0.0) or 0.0)
                _def_buildable = float(_f3_areas.get('buildable_total', 0.0) or 0.0)
                _def_pub_shared = float(_f3_areas.get('public_common_total', 0.0) or 0.0)
                _def_pub_non_shared = float(_f3_areas.get('public_non_common_total', 0.0) or 0.0)
                # 若 Tab 3 剛推送（consume flag），強制覆寫 widget 既有值避免 Streamlit widget state 殘留
                if st.session_state.pop('_f3_push_consume', False):
                    for _k, _v in (
                        ('area_offset', _def_offset),
                        ('area_buildable', _def_buildable),
                        ('area_public_shared', _def_pub_shared),
                        ('area_public_non_shared', _def_pub_non_shared),
                    ):
                        st.session_state[_k] = _v
                        st.session_state[f'_sv_{_k}'] = _v
            else:
                _def_offset, _def_buildable, _def_pub_shared, _def_pub_non_shared = 0.0, 10000.0, 4000.0, 1000.0

            area_offset = _ni("抵充地 (㎡)", min_value=0.0, max_value=10000000.0, value=_def_offset, step=100.0, key="area_offset",
                                          help="既有公共設施用地，可抵充共同負擔")
            area_buildable = _ni("可建築土地 (㎡)", min_value=0.0, max_value=10000000.0, value=_def_buildable, step=100.0, key="area_buildable",
                                             help="重劃後分回之宅地面積")
            area_public_shared = _ni("共同負擔公設用地 (㎡)", min_value=0.0, max_value=10000000.0, value=_def_pub_shared, step=100.0, key="area_public_shared",
                                                  help="道路、公園等公共設施用地")
            area_public_non_shared = _ni("非共同負擔公設用地 (㎡)", min_value=0.0, max_value=10000000.0, value=_def_pub_non_shared, step=100.0, key="area_public_non_shared",
                                                      help="學校、機關用地等")
            
            # 計算重劃區總面積和宅地化率
            # 重劃區總面積 = 可建築土地 + 共同負擔公設用地 + 非共同負擔公設用地（不含抵充地）
            total_area = area_buildable + area_public_shared + area_public_non_shared
            
            # 有效宅地化率 f = 可建築土地 / 重劃區總面積
            if total_area > 0:
                f_rate = area_buildable / total_area
            else:
                f_rate = 0
            
            # 公共設施用地平均負擔比率 = (共同負擔公設用地 - 抵充地) / (重劃區總面積 - 抵充地)
            denominator_for_burden = total_area - area_offset
            if denominator_for_burden > 0:
                public_burden_rate = (area_public_shared - area_offset) / denominator_for_burden
            else:
                public_burden_rate = 0
            
            st.info(f"""
            **面積計算結果**
            - 重劃區總面積 = {total_area:,.2f} ㎡
            - 有效宅地化率 f = {area_buildable:,.0f} / {total_area:,.0f} = **{f_rate*100:.2f}%**
            - 公共設施用地平均負擔比率 = ({area_public_shared:,.0f} - {area_offset:,.0f}) / ({total_area:,.0f} - {area_offset:,.0f}) = **{public_burden_rate*100:.2f}%**
            """)
            
            st.markdown("---")
            st.subheader("💰 開發成本")
            st.caption("輸入單位面積成本（單位：元/平方公尺）")
            
            B_cost = _ni("B 開發工事費 (元/㎡)", min_value=0.0, max_value=100000.0, value=3000.0, step=100.0, key="b_cost",
                                     help="測量、填土、道路、水溝、橋樑等直接工事費")
            K_cost = _ni("K 拆遷補償費 (元/㎡)", min_value=0.0, max_value=100000.0, value=500.0, step=100.0, key="k_cost",
                                     help="各種負擔金、拆遷補償")
            C_cost = _ni("C 間接費用 (重劃業務費) (元/㎡)", min_value=0.0, max_value=100000.0, value=200.0, step=50.0, key="c_cost",
                                     help="廣告、薪資、稅捐、管理費、登記費、設計費等")
            
            st.markdown("---")
            st.subheader("⏱️ 利息計算期間")
            
            n_months_input = _ni("工程動工至完工月數", min_value=1, max_value=120, value=24, step=1, key="n_months",
                                             help="開發成本之利息負擔以一半計算")
            n_half = n_months_input / 2  # 開發成本以工期一半計算利息
            st.info(f"n = {n_months_input} ÷ 2 = **{n_half:.1f} 月**（開發成本利息月數）")
            
            n_prime = _ni("n' 土地成本利息月數", min_value=1, max_value=120, value=36, step=1, key="n_prime",
                                       help="重劃公告至重劃後土地點交之月數")
            
            st.markdown("---")
            st.subheader("🔄 成熟度修正")
            st.caption("因A值已作過成熟度修正，此處預設為1（即N=0）")
            
            maturity_r = _ni("r 成熟度修正利率 (%)", min_value=0.0, max_value=30.0, value=0.0, step=0.1, key="maturity_r_ed",
                                         help="若A值已修正，可設為0")
            maturity_N = _ni("N 成熟度修正年期", min_value=0, max_value=20, value=0, step=1, key="maturity_n_ed",
                                         help="若A值已修正，可設為0")
            
            # 計算成熟度修正係數
            if maturity_N > 0 and maturity_r > 0:
                maturity_factor = 1 / ((1 + maturity_r / 100) ** maturity_N)
            else:
                maturity_factor = 1.0
            st.info(f"成熟度修正係數 = 1/(1+{maturity_r/100:.4f})^{maturity_N} = **{maturity_factor:.4f}**")
            
            st.markdown("---")
            st.subheader("📊 個別因素修正")
            D_factor = _ni("D 個別因素修正率 (%)", min_value=50.0, max_value=150.0, value=100.0, step=1.0, key="d_factor",
                                       help="不同所有權人土地條件差異修正")
        
        with c_r:
            st.subheader("🧮 預期開發分析法計算結果")
            
            # 檢查必要數據
            if A_value > 0 and f_rate > 0 and total_area > 0:
                # ===== 預期開發分析法公式計算 =====
                # X = { A × f － [ ( B ＋ K ) ( 1 + n × p ) + C ] } × 1/(1+n'×p) × 1/(1+r)^N × D
                
                # 將百分比轉換為小數
                p_decimal = p_monthly / 100  # 月利率小數
                D_decimal = D_factor / 100   # 個別因素修正率小數
                
                # 步驟1：計算 A × f（開發後土地價值 × 有效宅地化率）
                step1_Af = A_value * f_rate
                
                # 步驟2：計算開發成本含利息 (B + K)(1 + n × p)
                step2_BK = B_cost + K_cost
                step2_interest = 1 + n_half * p_decimal
                step2_BK_with_interest = step2_BK * step2_interest
                
                # 步驟3：加上間接費用 C
                step3_total_cost = step2_BK_with_interest + C_cost
                
                # 步驟4：計算扣除成本後的毛額
                step4_gross = step1_Af - step3_total_cost
                
                # 步驟5：土地成本利息折現 1/(1+n'×p)
                step5_land_discount = 1 / (1 + n_prime * p_decimal)
                step5_result = step4_gross * step5_land_discount
                
                # 步驟6：成熟度修正 1/(1+r)^N
                step6_result = step5_result * maturity_factor
                
                # 步驟7：個別因素修正 × D
                X_value = step6_result * D_decimal  # 重劃前土地單價（重劃後價格日期）
                
                # 顯示計算過程
                st.markdown(f"""
                #### 計算公式
                **X = {{ A × f － [ ( B ＋ K ) ( 1 + n × p ) + C ] }} × 1/(1+n'×p) × 1/(1+r)^N × D**
                
                ---
                #### 參數值
                | 參數 | 說明 | 值 |
                |------|------|-----|
                | **A** | 重劃後土地價格 | {A_value:,.0f} 元/㎡ |
                | **f** | 有效宅地化率 | {f_rate*100:.2f}% = {f_rate:.4f} |
                | **B** | 開發工事費 | {B_cost:,.0f} 元/㎡ |
                | **K** | 拆遷補償費 | {K_cost:,.0f} 元/㎡ |
                | **C** | 間接費用 | {C_cost:,.0f} 元/㎡ |
                | **n** | 開發成本利息月數 | {n_half:.1f} 月 |
                | **n'** | 土地成本利息月數 | {n_prime} 月 |
                | **p** | 月利率 | {p_monthly:.4f}% = {p_decimal:.6f} |
                | **r** | 成熟度修正利率 | {maturity_r:.2f}% |
                | **N** | 成熟度修正年期 | {maturity_N} 年 |
                | **D** | 個別因素修正率 | {D_factor:.0f}% |
                
                ---
                #### 計算過程
                
                **步驟1：A × f**（重劃後可分回土地價值）
                = {A_value:,.0f} × {f_rate:.4f} = **{step1_Af:,.2f}** 元/㎡
                
                **步驟2：(B + K)(1 + n × p)**（開發成本含利息）
                = ({B_cost:,.0f} + {K_cost:,.0f}) × (1 + {n_half:.1f} × {p_decimal:.6f})
                = {step2_BK:,.0f} × {step2_interest:.6f} = **{step2_BK_with_interest:,.2f}** 元/㎡
                
                **步驟3：步驟2 + C**（總開發成本）
                = {step2_BK_with_interest:,.2f} + {C_cost:,.0f} = **{step3_total_cost:,.2f}** 元/㎡
                
                **步驟4：步驟1 - 步驟3**（扣除成本後毛額）
                = {step1_Af:,.2f} - {step3_total_cost:,.2f} = **{step4_gross:,.2f}** 元/㎡
                
                **步驟5：× 1/(1+n'×p)**（土地成本利息折現）
                = {step4_gross:,.2f} × 1/(1 + {n_prime} × {p_decimal:.6f})
                = {step4_gross:,.2f} × {step5_land_discount:.6f} = **{step5_result:,.2f}** 元/㎡
                
                **步驟6：× 1/(1+r)^N**（成熟度修正）
                = {step5_result:,.2f} × {maturity_factor:.4f} = **{step6_result:,.2f}** 元/㎡
                
                **步驟7：× D**（個別因素修正）
                = {step6_result:,.2f} × {D_decimal:.4f} = **{X_value:,.2f}** 元/㎡
                """)
                
                st.markdown("---")
                st.subheader("📈 重劃後價格日期之重劃前地價")
                
                # 顯示結果（重劃後價格日期）
                x1, x2 = st.columns(2)
                x1.metric("重劃前土地單價(㎡)", f"{X_value:,.0f} 元/㎡")
                x2.metric("重劃前土地單價(坪)", f"{X_value / 0.3025:,.0f} 元/坪")
                
                # 計算重劃前土地總價
                X_total = X_value * total_area
                st.metric("重劃前土地總價", f"{X_total:,.0f} 元")
                st.caption(f"※ 以上為「重劃後價格日期」之重劃前平均地價")
                
                # ===== 價格日期調整 =====
                st.markdown("---")
                st.subheader("🔄 價格日期調整")
                st.caption("將預期開發法計算結果從重劃後價格日期調整至重劃前價格日期")
                
                # 使用本頁籤輸入的價格日期（而非從 session_state 讀取）
                post_date = post_date_dt
                pre_date = pre_date_dt
                post_index = post_index_ed
                pre_index = pre_index_ed
                district = district_ed
                
                if post_date and pre_date:
                    # 計算價格日期調整率
                    if post_index > 0:
                        date_adjust_factor = pre_index / post_index
                    else:
                        date_adjust_factor = 1.0
                    
                    # 調整後的重劃前地價
                    X_adjusted = X_value * date_adjust_factor
                    X_adjusted_total = X_adjusted * total_area
                    # 存入 session_state 供 Tab 5 使用
                    st.session_state['pre_land_price_sqm'] = X_adjusted
                    
                    st.markdown(f"""
                    | 項目 | 值 |
                    |------|-----|
                    | 地區 | {district} |
                    | 重劃後價格日期 | {post_date.strftime('%Y/%m/%d')} |
                    | 重劃後價格日期指數 | {post_index:.2f} |
                    | 重劃前價格日期 | {pre_date.strftime('%Y/%m/%d')} |
                    | 重劃前價格日期指數 | {pre_index:.2f} |
                    | **價格日期調整率** | {pre_index:.2f} / {post_index:.2f} = **{date_adjust_factor:.4f}** |
                    """)
                    
                    st.markdown(f"""
                    **調整計算**：
                    
                    重劃前地價（調整至重劃前價格日期）= {X_value:,.0f} × {date_adjust_factor:.4f} = **{X_adjusted:,.0f} 元/㎡**
                    """)
                    
                    st.markdown("---")
                    st.subheader("📋 最終結果：重劃前價格日期之重劃前平均地價")
                    
                    r1, r2 = st.columns(2)
                    r1.metric("重劃前土地單價(㎡)", f"{X_adjusted:,.0f} 元/㎡", 
                             delta=f"調整 {X_adjusted - X_value:+,.0f}")
                    r2.metric("重劃前土地單價(坪)", f"{X_adjusted / 0.3025:,.0f} 元/坪",
                             delta=f"調整 {(X_adjusted - X_value) / 0.3025:+,.0f}")
                    
                    st.metric("重劃前土地總價", f"{X_adjusted_total:,.0f} 元")
                    st.caption(f"※ 價格日期：{pre_date.strftime('%Y/%m/%d')} | 重劃區總面積：{total_area:,.2f} ㎡")
                    
                    # 顯示前後比較
                    st.markdown("---")
                    st.subheader("📊 重劃前後地價比較")
                    
                    # 計算漲跌幅
                    price_change = A_value - X_adjusted
                    price_change_pct = (price_change / X_adjusted * 100) if X_adjusted > 0 else 0
                    
                    col1, col2, col3 = st.columns(3)
                    col1.metric("重劃前地價", f"{X_adjusted:,.0f} 元/㎡", help=f"價格日期：{pre_date.strftime('%Y/%m/%d')}")
                    col2.metric("重劃後地價", f"{A_value:,.0f} 元/㎡", help=f"價格日期：{post_date.strftime('%Y/%m/%d')}")
                    col3.metric("增值", f"{price_change:+,.0f} 元/㎡", delta=f"{price_change_pct:+.1f}%")
                
            else:
                st.warning("⚠️ 請確認以下資料已輸入：")
                if A_value <= 0:
                    st.error("- A（重劃後土地價格）尚未設定")
                if f_rate <= 0:
                    st.error("- f（有效宅地化率）為0，請檢查土地面積輸入")
                if total_area <= 0:
                    st.error("- 重劃區總面積為0，請檢查土地面積輸入")
    
    # ===== Tab: 地價區段分析 =====
    elif selected_tab == TAB_PRICE_ZONE:
        st.markdown("### 功能：重劃前後地價區段分析")
        st.caption("依據比準區段/街廓反推各區段/街廓單價。非共同負擔公設用地（社宅、機關用地）需價購取得，納入可建築土地計算。")

        # ===== 從 Tab 3（街廓互動分析）讀取推送之街廓清單與重劃前地價區段清單 =====
        _f3_blocks = st.session_state.get('f3_to_tab_blocks')
        _f3_zones = st.session_state.get('f3_to_tab_zones')
        if _f3_blocks or _f3_zones:
            with st.expander("📦 已從 街廓互動分析 帶入之街廓／區段清單", expanded=True):
                if _f3_blocks:
                    st.markdown("**重劃後街廓清單（由 街廓互動分析 推送）**")
                    st.dataframe(pd.DataFrame(_f3_blocks), use_container_width=True, hide_index=True)
                if _f3_zones:
                    st.markdown("**重劃前地價區段清單（由 街廓互動分析 圖面點選結果推送）**")
                    st.dataframe(pd.DataFrame(_f3_zones), use_container_width=True, hide_index=True)
                st.caption("💡 如需更新，請至 街廓互動分析 重新點選區段並再次按「📤 推送至 地價區段分析」。")

        # ===== 從 Tab 4 取得各類土地面積（不再重複輸入） =====
        t4_area_buildable = st.session_state.get('area_buildable', 0)
        t4_area_public_shared = st.session_state.get('area_public_shared', 0)
        t4_area_public_non_shared = st.session_state.get('area_public_non_shared', 0)
        t4_area_offset = st.session_state.get('area_offset', 0)
        
        if t4_area_buildable <= 0:
            st.warning("⚠️ 請先至「📋 預期開發分析法」頁籤輸入各類土地面積，本頁將自動帶入該資料")
            st.info("尚未取得 預期開發分析法 面積資料，將使用預設值。完成 預期開發分析法 輸入後重新進入本頁即可帶入。")
        else:
            st.success(f"✅ 已從「預期開發分析法」帶入面積資料：可建築土地 {t4_area_buildable:,.2f} ㎡｜共同負擔公設 {t4_area_public_shared:,.2f} ㎡｜非共同負擔公設 {t4_area_public_non_shared:,.2f} ㎡｜抵充地 {t4_area_offset:,.2f} ㎡")
        
        col_left, col_right = st.columns([1, 1])
        
        # ===== 從 Tab 3 推送資料判斷是否已完成街廓分類 =====
        _has_f3_blocks = bool(_f3_blocks)
        _has_f3_zones = bool(_f3_zones)

        with col_left:
            st.subheader("📐 重劃區街廓清單（由 街廓互動分析 帶入，無需手動輸入分區數）")
            st.caption("💡 本頁不再輸入「分區數／類型／面積」。請於 街廓互動分析 完成街廓分類後，按「📤 推送至 地價區段分析」即可自動帶入。")

            # ---- 依 Tab 3 推送結果組出 general_zones / non_shared_zones ----
            general_zones = []
            non_shared_zones = []

            if _has_f3_blocks:
                for b in _f3_blocks:
                    cat = b.get('街廓分類', '') or b.get('category', '')
                    burden = F3_CATEGORY_BURDEN.get(cat, '')
                    area_m2 = float(b.get('面積_m2', 0.0) or 0.0)
                    label = b.get('街廓編號', '') or b.get('label', '')
                    if burden == '可建築土地':
                        general_zones.append({
                            'type': cat or '住宅區', 'area': area_m2,
                            'name': label, 'category': '一般'
                        })
                    elif burden == '非共同負擔':
                        non_shared_zones.append({
                            'type': cat or '其他非共同負擔之公共設施用地', 'area': area_m2,
                            'name': label, 'category': '非共同負擔'
                        })
            else:
                st.warning("⚠️ 尚未從 街廓互動分析 推送街廓清單。請先完成 街廓互動分析 的街廓分類並按「📤 推送至 地價區段分析」。")
                st.info("尚未取得 街廓互動分析 街廓清單時，本頁仍以 預期開發分析法 的總面積顯示；但「相對比率」計算需個別街廓，請優先完成 街廓互動分析。")

            # ---- 一、一般可建築土地 ----
            st.markdown("**一、一般可建築土地（住宅區、商業區等）**")
            if general_zones:
                df_gz = pd.DataFrame([
                    {'街廓編號': z['name'], '分類': z['type'], '面積(㎡)': round(z['area'], 2)}
                    for z in general_zones
                ])
                st.dataframe(df_gz, use_container_width=True, hide_index=True)
            else:
                st.caption("（街廓互動分析 尚未推送一般可建築街廓）")
            total_general = sum(z['area'] for z in general_zones)
            if t4_area_buildable > 0:
                gz_diff = total_general - t4_area_buildable
                if abs(gz_diff) > 1:
                    st.error(f"❌ 一般可建築土地（街廓互動分析）合計 {total_general:,.2f}㎡ ≠ 預期開發分析法 設定 {t4_area_buildable:,.2f}㎡（差異：{gz_diff:+,.2f}㎡）")
                else:
                    st.success(f"✅ 一般可建築土地合計 = {total_general:,.2f} ㎡（與 預期開發分析法 一致）")
            else:
                st.info(f"一般可建築土地合計：{total_general:,.2f} ㎡")

            # ---- 二、非共同負擔公共設施用地 ----
            st.markdown("---")
            st.markdown("**二、非共同負擔公共設施用地（需價購取得）**")
            if non_shared_zones:
                df_nsz = pd.DataFrame([
                    {'街廓編號': z['name'], '分類': z['type'], '面積(㎡)': round(z['area'], 2)}
                    for z in non_shared_zones
                ])
                st.dataframe(df_nsz, use_container_width=True, hide_index=True)
            else:
                st.caption("（街廓互動分析 尚未推送非共同負擔公設街廓）")
            total_non_shared = sum(z['area'] for z in non_shared_zones)
            if t4_area_public_non_shared > 0:
                nsz_diff = total_non_shared - t4_area_public_non_shared
                if abs(nsz_diff) > 1:
                    st.error(f"❌ 非共同負擔公設（街廓互動分析）合計 {total_non_shared:,.2f}㎡ ≠ 預期開發分析法 設定 {t4_area_public_non_shared:,.2f}㎡（差異：{nsz_diff:+,.2f}㎡）")
                else:
                    st.success(f"✅ 非共同負擔公設合計 = {total_non_shared:,.2f} ㎡（與 預期開發分析法 一致）")
            else:
                st.info(f"非共同負擔公設小計：{total_non_shared:,.2f} ㎡")

            # 合併所有可建築土地（用於地價計算）
            all_buildable_zones = general_zones + non_shared_zones
            total_buildable = total_general + total_non_shared
            st.success(f"**可建築土地總面積 = {total_buildable:,.2f} ㎡（共 {len(all_buildable_zones)} 個街廓）**")
            
            # ===== 共同負擔公共設施用地（從 Tab 4 帶入，不再輸入）=====
            st.markdown("---")
            st.markdown("**三、共同負擔公共設施用地**（由 預期開發分析法 帶入，無需重複輸入）")
            st.caption("道路、溝渠、公園、綠地、廣場等，由土地所有權人共同負擔")

            # 若 Tab 5（預期開發分析法）尚未填入面積，優先從 Tab 3 推送結果回讀；
            # 仍為 0 時才用 5000 當預設值避免顯示空白
            _f3_fallback = st.session_state.get('f3_to_tab_areas') or {}
            if t4_area_public_shared and t4_area_public_shared > 0:
                area_public_shared_t5 = float(t4_area_public_shared)
            elif float(_f3_fallback.get('public_common_total', 0.0) or 0.0) > 0:
                area_public_shared_t5 = float(_f3_fallback.get('public_common_total', 0.0) or 0.0)
            else:
                area_public_shared_t5 = 5000.0
            if t4_area_offset and t4_area_offset > 0:
                area_offset_t5 = float(t4_area_offset)
            else:
                area_offset_t5 = float(_f3_fallback.get('offset_area', 0.0) or 0.0)
            
            display_col1, display_col2 = st.columns(2)
            display_col1.metric("共同負擔公設用地", f"{area_public_shared_t5:,.2f} ㎡")
            display_col2.metric("其中：抵充地", f"{area_offset_t5:,.2f} ㎡")
            
            # ===== 面積彙總 =====
            st.markdown("---")
            st.subheader("📊 面積彙總")
            
            total_area_t5 = total_buildable + area_public_shared_t5
            
            st.markdown(f"""
            | 項目 | 面積(㎡) | 說明 |
            |------|----------|------|
            | 一般可建築土地 | {total_general:,.2f} | 住宅區、商業區等 |
            | 非共同負擔公設用地 | {total_non_shared:,.2f} | 社宅、機關用地（需價購） |
            | **可建築土地合計** | **{total_buildable:,.2f}** | 納入地價計算 |
            | 共同負擔公設用地 | {area_public_shared_t5:,.2f} | 道路、公園等（預期開發分析法 帶入） |
            | 　其中：抵充地 | {area_offset_t5:,.2f} | 既有公有土地（預期開發分析法 帶入） |
            | **重劃區總面積** | **{total_area_t5:,.2f}** | |
            """)
            
            # 計算負擔比率
            if total_area_t5 > 0:
                public_burden_rate = (area_public_shared_t5 - area_offset_t5) / (total_area_t5 - area_offset_t5) * 100 if (total_area_t5 - area_offset_t5) > 0 else 0
                st.info(f"公共設施用地負擔比率：{public_burden_rate:.2f}%")
            
            # 面積檢核
            st.markdown("---")
            st.subheader("🔍 面積檢核")
            errors = []
            if area_offset_t5 > area_public_shared_t5:
                errors.append("抵充地面積不能大於共同負擔公設用地面積")
            
            # 檢查街廓名稱是否重複
            all_names = [z['name'] for z in all_buildable_zones]
            if len(all_names) != len(set(all_names)):
                errors.append("街廓代號有重複，請修正")
            
            # 一般可建築土地分區數面積總和檢核
            if t4_area_buildable > 0 and abs(total_general - t4_area_buildable) > 1:
                errors.append(f"一般可建築土地分區面積總和 ({total_general:,.2f}㎡) ≠ 預期開發分析法 可建築土地總面積 ({t4_area_buildable:,.2f}㎡)")
            
            if errors:
                for err in errors:
                    st.error(f"❌ {err}")
            else:
                st.success("✅ 面積檢核通過")
        
        with col_right:
            st.subheader("💰 地價來源設定")
            
            # 取得其他頁籤的計算結果
            dev_analysis_sqm = st.session_state.get('weighted_price_sqm', 0)
            pre_price_from_t4 = st.session_state.get('pre_land_price_sqm', 0)

            st.markdown("**重劃後平均地價（可建築土地）**")
            if dev_analysis_sqm > 0:
                st.info(f"來自土地開發分析：{dev_analysis_sqm:,.0f} 元/㎡ = {dev_analysis_sqm/0.3025:,.0f} 元/坪")
                use_dev = _cb("使用土地開發分析結果", value=True, key="use_dev_t5")
                post_avg_price = dev_analysis_sqm if use_dev else _ni("手動輸入重劃後平均單價(元/㎡)", min_value=1000.0, value=50000.0, step=1000.0, key="post_avg_manual")
            else:
                post_avg_price = _ni("重劃後可建築土地平均單價 Q (元/㎡)", min_value=1000.0, value=50000.0, step=1000.0, key="post_avg")

            st.markdown("**重劃前平均地價**")
            if pre_price_from_t4 > 0:
                st.info(f"來自預期開發分析法：{pre_price_from_t4:,.0f} 元/㎡ = {pre_price_from_t4/0.3025:,.0f} 元/坪")
                use_pre_t4 = _cb("使用預期開發分析法結果", value=True, key="use_pre_t4_t5")
                pre_avg_price = pre_price_from_t4 if use_pre_t4 else _ni("手動輸入重劃前平均單價(元/㎡)", min_value=1000.0, value=30000.0, step=1000.0, key="pre_avg_manual")
            else:
                pre_avg_price = _ni("重劃前平均單價 (元/㎡)（請先至 預期開發分析法 完成計算）", min_value=1000.0, value=30000.0, step=1000.0, key="pre_avg")
        
        st.markdown("---")
        
        # ===== 重劃後分配街廓地價分析 =====
        st.subheader("🏘️ 重劃後分配街廓地價分析")
        st.markdown("""
        **計算公式**：Q × Σ(各街廓面積) = U × Σ(相對比率 × 面積)
        - Q：可建築土地平均單價
        - U：比準街廓單價
        - 各街廓單價 = U × 該街廓相對比率
        """)
        
        if len(all_buildable_zones) > 0:
            col_post1, col_post2 = st.columns([1, 1])
            
            with col_post1:
                st.markdown("**街廓比率設定**")
                
                # 選擇比準街廓
                block_names = [f"{z['name']} ({z['type']})" for z in all_buildable_zones]
                block_name_only = [z['name'] for z in all_buildable_zones]
                benchmark_idx = _sb("選擇比準街廓", range(len(block_names)), format_func=lambda x: block_names[x], key="benchmark_block")
                benchmark_block = block_name_only[benchmark_idx]
                
                st.markdown("**各街廓相對比率設定**（比準街廓=100%）")
                block_ratios = {}
                
                # 預設比率建議
                default_ratios = {
                    "住宅區": 100.0, "商業區": 120.0, "工業區": 80.0,
                    "社會住宅": 90.0, "機關用地": 80.0, "學校用地": 80.0, "其他": 100.0
                }
                
                for z in all_buildable_zones:
                    if z['name'] == benchmark_block:
                        block_ratios[z['name']] = 100.0
                        st.success(f"**{z['name']}** ({z['type']}) - 比準街廓：**100%**")
                    else:
                        default_val = default_ratios.get(z['type'], 100.0)
                        ratio = _sl(
                            f"{z['name']} ({z['type']})", 
                            min_value=50.0, max_value=150.0, 
                            value=default_val, step=5.0, 
                            key=f"ratio_{z['name']}",
                            help=f"面積：{z['area']:,.0f}㎡"
                        )
                        block_ratios[z['name']] = ratio
            
            with col_post2:
                st.markdown("**計算結果**")
                
                # 計算比準街廓單價 U
                # Q × Σ(面積) = U × Σ(比率 × 面積)
                # U = Q × Σ(面積) / Σ(比率 × 面積)
                
                sum_weighted_area = sum(z['area'] * block_ratios[z['name']] / 100 for z in all_buildable_zones)
                
                if sum_weighted_area > 0 and total_buildable > 0:
                    benchmark_unit_price = post_avg_price * total_buildable / sum_weighted_area
                    
                    st.success(f"**比準街廓單價 U = {benchmark_unit_price:,.0f} 元/㎡ = {benchmark_unit_price/0.3025:,.0f} 元/坪**")
                    
                    # 計算各街廓單價並找出最低價
                    block_results = []
                    min_price = float('inf')
                    min_price_block = ""
                    
                    for z in all_buildable_zones:
                        price = benchmark_unit_price * block_ratios[z['name']] / 100
                        block_results.append({
                            "街廓": z['name'],
                            "類型": z['type'],
                            "分類": z['category'],
                            "面積(㎡)": z['area'],
                            "相對比率(%)": block_ratios[z['name']],
                            "單價(元/㎡)": price,
                            "單價(元/坪)": price / 0.3025,
                            "總價(元)": price * z['area']
                        })
                        if price < min_price:
                            min_price = price
                            min_price_block = z['name']
                    
                    df_blocks = pd.DataFrame(block_results)
                    st.dataframe(df_blocks.style.format({
                        "面積(㎡)": "{:,.2f}",
                        "相對比率(%)": "{:.1f}",
                        "單價(元/㎡)": "{:,.0f}",
                        "單價(元/坪)": "{:,.0f}",
                        "總價(元)": "{:,.0f}"
                    }), use_container_width=True)
                    
                    # 驗算
                    total_value = sum(r['總價(元)'] for r in block_results)
                    calc_avg = total_value / total_buildable if total_buildable > 0 else 0
                    st.caption(f"驗算：總價值 {total_value:,.0f} 元 ÷ 總面積 {total_buildable:,.0f}㎡ = {calc_avg:,.0f} 元/㎡ (輸入平均單價：{post_avg_price:,.0f} 元/㎡)")
                    
                    # 共同負擔公設用地以最低地價評估
                    st.markdown("---")
                    st.markdown("**共同負擔公共設施用地評估**")
                    st.info(f"以可建築土地最低地價計算：{min_price:,.0f} 元/㎡ = {min_price/0.3025:,.0f} 元/坪（來自 {min_price_block}）")
                    
                    public_shared_value = area_public_shared_t5 * min_price
                    
                    col_pub1, col_pub2 = st.columns(2)
                    col_pub1.metric("共同負擔公設用地面積", f"{area_public_shared_t5:,.0f} ㎡")
                    col_pub2.metric("共同負擔公設用地總價", f"{public_shared_value:,.0f} 元")
                    
                    # 儲存結果供財務分析使用
                    st.session_state['post_block_results'] = block_results
                    st.session_state['post_benchmark_price'] = benchmark_unit_price
                    st.session_state['post_min_price'] = min_price
                    st.session_state['total_buildable_area'] = total_buildable
                    st.session_state['total_general_area'] = total_general
                    st.session_state['total_non_shared_area'] = total_non_shared
                    st.session_state['public_shared_area'] = area_public_shared_t5
                    st.session_state['offset_area'] = area_offset_t5
                    st.session_state['total_area_t5'] = total_area_t5
                    st.session_state['all_buildable_zones'] = all_buildable_zones
        
        st.markdown("---")
        
        # ===== 重劃前地價區段分析 =====
        st.subheader("🗺️ 重劃前地價區段分析")
        st.markdown("""
        **計算公式**：平均單價 × 總面積 = Y × Σ(相對比率 × 面積)
        - Y：比準區段單價
        - 各區段單價 = Y × 該區段相對比率
        """)
        
        col_pre1, col_pre2 = st.columns([1, 1])

        with col_pre1:
            st.markdown("**地價區段清單（由 街廓互動分析 圖面點選結果帶入，無需手動輸入區段數）**")

            # ---- 由 Tab 3 推送之重劃前地價區段清單生成 pre_zones ----
            pre_zones = []
            if _has_f3_zones:
                for z in _f3_zones:
                    pre_zones.append({
                        'name': z.get('區段', '') or f"區段{len(pre_zones)+1}",
                        'area': float(z.get('面積_m2', 0.0) or 0.0),
                        '筆數': int(z.get('筆數', 0) or 0),
                        '暫編地號': z.get('暫編地號', ''),
                    })
                df_prez = pd.DataFrame([
                    {'區段': z['name'], '筆數': z['筆數'], '面積(㎡)': round(z['area'], 2),
                     '暫編地號': z.get('暫編地號', '')}
                    for z in pre_zones
                ])
                st.dataframe(df_prez, use_container_width=True, hide_index=True)
            else:
                st.warning("⚠️ 尚未從 街廓互動分析 推送重劃前地價區段。請先於 街廓互動分析 完成圖面點選標註並按「📤 推送至 地價區段分析」。")

            total_pre_area = sum(z['area'] for z in pre_zones)
            if pre_zones:
                area_diff = abs(total_pre_area - total_area_t5)
                if total_area_t5 > 0 and area_diff > max(1.0, total_area_t5 * 0.02):
                    st.warning(f"⚠️ 區段面積合計 {total_pre_area:,.2f}㎡ ≠ 重劃區總面積 {total_area_t5:,.2f}㎡（差異：{area_diff:,.2f}㎡；可能仍有宗地未標註區段）")
                else:
                    st.success(f"✅ 區段面積合計 = {total_pre_area:,.2f}㎡")

            # 選擇比準區段 + 相對比率
            if pre_zones:
                pre_zone_names = [z['name'] for z in pre_zones]
                benchmark_pre_zone = _sb("比準區段", pre_zone_names, key="benchmark_pre_zone")

                st.markdown("**各區段相對比率設定**（比準區段=100%）")
                pre_zone_ratios = {}
                for z in pre_zones:
                    if z['name'] == benchmark_pre_zone:
                        pre_zone_ratios[z['name']] = 100.0
                        st.success(f"**{z['name']}**（比準區段）：**100%**")
                    else:
                        ratio = _sl(f"{z['name']} 相對比率(%)", min_value=50.0, max_value=150.0, value=70.0, step=5.0, key=f"pre_ratio_{z['name']}")
                        pre_zone_ratios[z['name']] = ratio
            else:
                benchmark_pre_zone = None
                pre_zone_ratios = {}
        
        with col_pre2:
            st.markdown("**計算結果**")
            
            # 計算比準區段單價
            sum_weighted_area_pre = sum(z['area'] * pre_zone_ratios[z['name']] / 100 for z in pre_zones)
            
            if sum_weighted_area_pre > 0 and total_pre_area > 0:
                benchmark_pre_price = pre_avg_price * total_pre_area / sum_weighted_area_pre
                
                st.success(f"**比準區段單價 Y = {benchmark_pre_price:,.0f} 元/㎡ = {benchmark_pre_price/0.3025:,.0f} 元/坪**")
                
                # 計算各區段單價
                pre_zone_results = []
                for z in pre_zones:
                    price = benchmark_pre_price * pre_zone_ratios[z['name']] / 100
                    pre_zone_results.append({
                        "區段": z['name'],
                        "面積(㎡)": z['area'],
                        "相對比率(%)": pre_zone_ratios[z['name']],
                        "單價(元/㎡)": price,
                        "單價(元/坪)": price / 0.3025,
                        "總價(元)": price * z['area']
                    })
                
                df_pre_zones = pd.DataFrame(pre_zone_results)
                st.dataframe(df_pre_zones.style.format({
                    "面積(㎡)": "{:,.2f}",
                    "相對比率(%)": "{:.1f}",
                    "單價(元/㎡)": "{:,.0f}",
                    "單價(元/坪)": "{:,.0f}",
                    "總價(元)": "{:,.0f}"
                }), use_container_width=True)
                
                # 儲存結果
                st.session_state['pre_zone_results'] = pre_zone_results
                st.session_state['pre_benchmark_price'] = benchmark_pre_price
                st.session_state['pre_zones'] = pre_zones
    
    # ===== Tab: 財務分析 =====
    elif selected_tab == TAB_FINANCE:
        st.markdown("### 功能：重劃財務可行性分析")
        st.caption("計算重劃總費用、現金流量表(NPV)、IRR。單位：萬元。抵費地於最後一年年初賣出。")
        
        col_fin1, col_fin2 = st.columns([1, 2])
        
        with col_fin1:
            st.subheader("📊 基本參數")
            
            # 取得前面頁籤的資料
            total_area_fin = st.session_state.get('total_area_t5', 39000)
            total_buildable_fin = st.session_state.get('total_buildable_area', 10000)
            
            # 取得土地交易分析與土地開發分析的結果（不含成熟度修正）
            land_trans_sqm = st.session_state.get('res_land_sqm', 0)  # 比較法（土地交易分析）
            dev_analysis_sqm = st.session_state.get('dev_land_price_sqm', 0)  # 土地開發分析法（正常價格）
            weighted_price = st.session_state.get('weighted_price_sqm', 0)  # 加權平均價格
            
            st.info(f"""
            **來自地價區段分析**
            - 重劃區總面積：{total_area_fin:,.0f} ㎡
            - 可建築土地面積：{total_buildable_fin:,.0f} ㎡
            """)
            
            st.markdown("---")
            st.subheader("💰 抵費地單價設定")
            st.caption("抵費地賣出價格採「土地開發分析法」與「比較法」加權平均（不做成熟度修正）")
            
            # 顯示兩種方法的價格
            col_p1, col_p2 = st.columns(2)
            col_p1.metric("比較法（土地交易分析）", f"{land_trans_sqm:,.0f} 元/㎡" if land_trans_sqm > 0 else "尚未計算")
            col_p2.metric("土地開發分析法", f"{dev_analysis_sqm:,.0f} 元/㎡" if dev_analysis_sqm > 0 else "尚未計算")
            
            if weighted_price > 0:
                st.success(f"**加權平均價格：{weighted_price:,.0f} 元/㎡ = {weighted_price/0.3025:,.0f} 元/坪**")
                use_weighted_price = _cb("使用加權平均價格", value=True, key="use_wp_fin")
                if use_weighted_price:
                    land_disposal_price = weighted_price
                else:
                    land_disposal_price = _ni("手動輸入抵費地單價（元/㎡）", min_value=1000.0, value=weighted_price, step=1000.0, key="manual_dp")
            else:
                st.warning("⚠️ 請先完成「土地開發分析法」頁籤的綜合評估")
                land_disposal_price = _ni("手動輸入抵費地單價（元/㎡）", min_value=1000.0, value=21601.0, step=1000.0, key="manual_dp2")
            
            st.markdown("---")
            st.subheader("⏱️ 開發期間設定")
            total_years = _ni("重劃總期間（年）", min_value=2, max_value=20, value=5, step=1, key="total_years")
            st.caption("※ 抵費地於最後一年年初賣出")
            
            st.markdown("---")
            st.subheader("🔧 重劃總費用單價（元/㎡）")
            st.caption("💡 本區數值直接使用「預期開發分析法」已輸入之 B／K／C 開發成本，不再重複填寫。如需修改請至 預期開發分析法。")
            # 直接讀取 Tab 5（預期開發分析法）的三個開發成本 widget：
            #   B_cost（key='b_cost'）= 開發工事費       → 工程費用
            #   K_cost（key='k_cost'）= 拆遷補償費       → 地上物拆遷補償費
            #   C_cost（key='c_cost'）= 間接費用/重劃業務費 → 行政作業費（重劃費用）
            # ※ 優先讀取「影子鍵 _sv_*」（跨 Tab rerun 不會被 Streamlit 自動清除），
            #   再 fallback 至 widget 鍵，最後 fallback 至 Tab 5 預設值（B=3000, K=500, C=200）
            def _read_persisted(widget_key, default):
                sv = st.session_state.get(f'_sv_{widget_key}')
                wv = st.session_state.get(widget_key)
                v = sv if sv is not None else (wv if wv is not None else default)
                try:
                    return float(v or 0.0)
                except (TypeError, ValueError):
                    return float(default)

            _b_in_state = ('_sv_b_cost' in st.session_state) or ('b_cost' in st.session_state)
            _k_in_state = ('_sv_k_cost' in st.session_state) or ('k_cost' in st.session_state)
            _c_in_state = ('_sv_c_cost' in st.session_state) or ('c_cost' in st.session_state)
            demolition_cost   = _read_persisted('k_cost', 500.0)
            admin_cost        = _read_persisted('c_cost', 200.0)
            construction_cost = _read_persisted('b_cost', 3000.0)

            # 顯示資料來源狀態（已連動 vs 使用 Tab 5 預設）
            _all_synced = _b_in_state and _k_in_state and _c_in_state
            if _all_synced:
                st.success("✅ 已連動 預期開發分析法 之 B／K／C 輸入值")
            else:
                _missing = []
                if not _b_in_state: _missing.append("B 工程費")
                if not _k_in_state: _missing.append("K 拆遷補償費")
                if not _c_in_state: _missing.append("C 重劃費用")
                st.warning(
                    f"⚠️ 您尚未於 預期開發分析法 輸入：{'、'.join(_missing)}；"
                    "目前顯示之數值為 **預期開發分析法 預設值**。建議先至 預期開發分析法 確認 / 修改 B／K／C，"
                    "回到本頁即會自動帶入您輸入的單價。"
                )

            _c_dc = st.columns(3)
            _c_dc[0].metric("地上物拆遷補償費 (K)", f"{demolition_cost:,.0f} 元/㎡",
                            delta="✅ 已連動 預期開發分析法" if _k_in_state else "⚠️ 尚未連動",
                            delta_color="off")
            _c_dc[1].metric("行政作業費 / 重劃費用 (C)", f"{admin_cost:,.0f} 元/㎡",
                            delta="✅ 已連動 預期開發分析法" if _c_in_state else "⚠️ 尚未連動",
                            delta_color="off")
            _c_dc[2].metric("工程費用 (B)", f"{construction_cost:,.0f} 元/㎡",
                            delta="✅ 已連動 預期開發分析法" if _b_in_state else "⚠️ 尚未連動",
                            delta_color="off")
            
            st.markdown("---")
            st.subheader("💵 貸款利率設定")
            loan_rate = _ni("貸款年利率（五大銀行放款利率）%", min_value=0.0, max_value=20.0, value=2.077, step=0.001, key="loan_rate", format="%.3f")
            
            st.markdown("---")
            st.subheader("📐 抵費地面積")
            # 計算預估所需抵費地面積（含預估貸款利息）
            # 1. B+K+C 直接費用
            total_dev_cost_estimate = (demolition_cost + admin_cost + construction_cost) * total_area_fin  # 元
            # 2. 預估貸款利息（萬元）：假設費用線性累積、收入於最後一年取得，
            #    平均待還餘額 ≈ 總費用 × (年期-1) ÷ 2，以年利率 r 計息
            _years_for_interest = max(0, int(total_years) - 1)
            estimated_loan_interest = total_dev_cost_estimate * (loan_rate / 100.0) * _years_for_interest / 2.0  # 元
            # 3. 重劃總費用（含利息）
            total_replot_cost_with_interest = total_dev_cost_estimate + estimated_loan_interest  # 元
            required_land_area = total_replot_cost_with_interest / land_disposal_price if land_disposal_price > 0 else 0

            st.info(
                f"預估所需抵費地面積：**{required_land_area:,.0f} ㎡**　"
                f"= 重劃總費用（B+K+C+貸款利息）÷ 抵費地單價\n\n"
                f"　• B+K+C 直接費用合計：{total_dev_cost_estimate/10000:,.2f} 萬元\n\n"
                f"　• 預估貸款利息：{estimated_loan_interest/10000:,.2f} 萬元　"
                f"（總費用 × {loan_rate:.3f}% × {_years_for_interest} ÷ 2，"
                f"線性累積平均餘額估算）\n\n"
                f"　• 重劃總費用（含利息）：{total_replot_cost_with_interest/10000:,.2f} 萬元"
            )

            actual_disposal_area = _ni("實際抵費地面積（㎡）", min_value=0.0, max_value=float(total_buildable_fin),
                                                    value=min(required_land_area * 1.05, float(total_buildable_fin)), step=100.0, key="disp_area")
        
        with col_fin2:
            st.subheader("💵 重劃總費用及盈餘（單位：萬元）")
            
            # 計算各項費用（轉換為萬元）
            demolition_total = demolition_cost * total_area_fin / 10000  # 萬元
            admin_total = admin_cost * total_area_fin / 10000  # 萬元
            construction_total = construction_cost * total_area_fin / 10000  # 萬元
            
            # 抵費地處分收入（萬元）
            disposal_income = actual_disposal_area * land_disposal_price / 10000  # 萬元
            
            st.markdown("**重劃總費用明細**")
            cost_data = [
                ["地上物拆遷補償費", f"{demolition_cost:,.0f}", f"{total_area_fin:,.0f}", f"{demolition_total:,.2f}"],
                ["行政作業費", f"{admin_cost:,.0f}", f"{total_area_fin:,.0f}", f"{admin_total:,.2f}"],
                ["工程費用", f"{construction_cost:,.0f}", f"{total_area_fin:,.0f}", f"{construction_total:,.2f}"],
            ]
            
            df_cost = pd.DataFrame(cost_data, columns=["項目", "單價(元/㎡)", "面積(㎡)", "小計(萬元)"])
            st.table(df_cost)
            
            st.markdown("---")
            st.subheader("📈 現金流量表（單位：萬元）")

            # ---------- 工程費 / 重劃費用 預設比例產生器（總和恰為 100%） ----------
            def _default_construction_ratio(n: int) -> list:
                """工程費：集中在中間年度（頭年與尾年低、中間年度高）。合計 = 100%。"""
                n = max(1, int(n))
                if n == 1: return [100.0]
                if n == 2: return [50.0, 50.0]
                if n == 3: return [10.0, 50.0, 40.0]
                if n == 4: return [0.0, 20.0, 50.0, 30.0]
                if n == 5: return [0.0, 5.0, 48.0, 47.0, 0.0]
                if n == 6: return [0.0, 5.0, 30.0, 35.0, 30.0, 0.0]
                # n >= 7：以「中間年度權重高、首尾權重低」的二次曲線分配
                mid = (n - 1) / 2.0
                w = [max(0.0, 1.0 - ((i - mid) / max(mid, 1)) ** 2) + 0.05 for i in range(n)]
                # 首尾壓低
                w[0] *= 0.1
                w[-1] *= 0.1
                s = sum(w) or 1.0
                raw = [round(x * 100.0 / s, 1) for x in w]
                # 修正捨入誤差，補到最大的一項
                diff = round(100.0 - sum(raw), 1)
                if diff != 0:
                    k = raw.index(max(raw))
                    raw[k] = round(raw[k] + diff, 1)
                return raw

            def _default_replot_ratio(n: int) -> list:
                """重劃費用：平均分配，合計 = 100%（修正捨入誤差）。"""
                n = max(1, int(n))
                base = round(100.0 / n, 1)
                arr = [base] * n
                diff = round(100.0 - sum(arr), 1)
                if diff != 0:
                    arr[0] = round(arr[0] + diff, 1)
                return arr

            # ---------- 偵測開發年期變動：自動重置各年度比例 ----------
            _prev_ty = st.session_state.get('_prev_total_years', None)
            _cur_ty = int(total_years)
            _years_changed = (_prev_ty != _cur_ty)
            # 首次進入（尚未設定任一 const_ratio_i）也視同需要初始化
            _needs_init = not any(f'const_ratio_{i}' in st.session_state for i in range(_cur_ty))

            if _years_changed or _needs_init:
                _new_const = _default_construction_ratio(_cur_ty)
                _new_replot = _default_replot_ratio(_cur_ty)
                # 覆寫 widget state（舊有的 i < _cur_ty 直接覆寫；超出 _cur_ty 的殘留 key 清除）
                for i in range(_cur_ty):
                    st.session_state[f'const_ratio_{i}']  = float(_new_const[i])
                    st.session_state[f'replot_ratio_{i}'] = float(_new_replot[i])
                # 清除超出年期的殘留 key（避免 Streamlit widget 狀態污染）
                for _k in list(st.session_state.keys()):
                    if _k.startswith('const_ratio_') or _k.startswith('replot_ratio_'):
                        try:
                            _idx = int(_k.split('_')[-1])
                            if _idx >= _cur_ty:
                                del st.session_state[_k]
                        except ValueError:
                            pass
                st.session_state['_prev_total_years'] = _cur_ty
                if _years_changed and _prev_ty is not None:
                    st.info(f"📐 偵測到開發年期由 {_prev_ty} 年變更為 {_cur_ty} 年，工程費與重劃費用比例已自動重新分配（合計 100%）。")

            # 年度支出分配設定
            st.markdown("**工程費各年度分配比例**（合計應為100%）")
            st.caption("💡 預設為中間年度集中型；變動開發年期時會自動重算。可手動再微調，按下方表格即會同步更新合計。")

            construction_ratios = []
            cols_const = st.columns(_cur_ty)
            _def_const = _default_construction_ratio(_cur_ty)
            for i in range(_cur_ty):
                # 若 widget 已經有 state（本次或之前設定），就不再傳 value 避免被 Streamlit 警告
                _ckey = f"const_ratio_{i}"
                if _ckey in st.session_state:
                    ratio = cols_const[i].number_input(f"第{i+1}年(%)", min_value=0.0, max_value=100.0,
                                                       step=1.0, key=_ckey)
                else:
                    ratio = cols_const[i].number_input(f"第{i+1}年(%)", min_value=0.0, max_value=100.0,
                                                       value=float(_def_const[i]), step=1.0, key=_ckey)
                construction_ratios.append(ratio / 100)

            # 檢核比例合計
            total_const_ratio = sum(construction_ratios)
            if abs(total_const_ratio - 1.0) > 0.01:
                st.warning(f"⚠️ 工程費比例合計 {total_const_ratio*100:.1f}% ≠ 100%（可按下方「自動補正」使合計精確等於 100%）")
                if st.button("🔧 自動補正工程費比例至 100%", key="btn_fix_const"):
                    _fix = _default_construction_ratio(_cur_ty)
                    for i in range(_cur_ty):
                        st.session_state[f'const_ratio_{i}'] = float(_fix[i])
                    st.rerun()
            else:
                st.success(f"✅ 工程費比例合計 {total_const_ratio*100:.1f}%")

            # 重劃費用分配（預設平均分配）
            st.markdown("**重劃費用各年度分配比例**（合計應為100%）")
            st.caption("💡 預設平均分配；變動開發年期時會自動等分重算。")

            replot_ratios = []
            cols_replot = st.columns(_cur_ty)
            _def_replot = _default_replot_ratio(_cur_ty)
            for i in range(_cur_ty):
                _rkey = f"replot_ratio_{i}"
                if _rkey in st.session_state:
                    ratio = cols_replot[i].number_input(f"第{i+1}年(%)", min_value=0.0, max_value=100.0,
                                                        step=1.0, key=_rkey)
                else:
                    ratio = cols_replot[i].number_input(f"第{i+1}年(%)", min_value=0.0, max_value=100.0,
                                                        value=float(_def_replot[i]), step=1.0, key=_rkey)
                replot_ratios.append(ratio / 100)

            # 檢核重劃費用比例合計
            total_replot_ratio = sum(replot_ratios)
            if abs(total_replot_ratio - 1.0) > 0.01:
                st.warning(f"⚠️ 重劃費用比例合計 {total_replot_ratio*100:.1f}% ≠ 100%（可按下方「自動補正」使合計精確等於 100%）")
                if st.button("🔧 自動補正重劃費用比例至 100%", key="btn_fix_replot"):
                    _fix = _default_replot_ratio(_cur_ty)
                    for i in range(_cur_ty):
                        st.session_state[f'replot_ratio_{i}'] = float(_fix[i])
                    st.rerun()
            else:
                st.success(f"✅ 重劃費用比例合計 {total_replot_ratio*100:.1f}%")
            
            # ===== 建立現金流量表 =====
            cash_flow_data = []
            cumulative_cash = 0
            yearly_cash_flows = []
            total_interest = 0
            
            # 重劃費用 = 行政作業費 + 地上物拆遷補償費
            replot_fee_total = admin_total + demolition_total
            
            for year in range(1, int(total_years) + 1):
                # 工程費（萬元）
                year_construction = construction_total * construction_ratios[year-1] if year-1 < len(construction_ratios) else 0
                
                # 重劃費用（依使用者設定比例）
                year_replot_fee = replot_fee_total * replot_ratios[year-1] if year-1 < len(replot_ratios) else 0
                
                # 支出小計
                year_expense = year_construction + year_replot_fee
                
                # 收入（最後一年年初收取抵費地款）
                year_income = disposal_income if year == total_years else 0
                
                # 當期淨現金流量
                net_cash = year_income - year_expense
                yearly_cash_flows.append(net_cash)
                
                # 期初淨額
                beginning_balance = cumulative_cash
                
                # 利息（按期初負債計算，負債才計息）
                year_interest = abs(min(0, beginning_balance)) * (loan_rate / 100) if beginning_balance < 0 else 0
                total_interest += year_interest
                
                # 期末餘額
                ending_balance = beginning_balance + net_cash - year_interest
                cumulative_cash = ending_balance
                
                cash_flow_data.append({
                    "年度": f"第{year}年",
                    "工程費": round(year_construction, 2),
                    "重劃費用": round(year_replot_fee, 2),
                    "支出小計": round(year_expense, 2),
                    "收入": round(year_income, 2),
                    "當期淨現金流量": round(net_cash, 2),
                    "期初淨額": round(beginning_balance, 2),
                    "貸款利息": round(year_interest, 2),
                    "期末餘額": round(ending_balance, 2)
                })
            
            # 加入合計列
            cash_flow_data.append({
                "年度": "合計",
                "工程費": round(construction_total, 2),
                "重劃費用": round(replot_fee_total, 2),
                "支出小計": round(construction_total + replot_fee_total, 2),
                "收入": round(disposal_income, 2),
                "當期淨現金流量": round(sum(yearly_cash_flows), 2),
                "期初淨額": "-",
                "貸款利息": round(total_interest, 2),
                "期末餘額": round(cumulative_cash, 2)
            })
            
            df_cashflow = pd.DataFrame(cash_flow_data)
            
            # 格式化顯示（直式表格）
            st.dataframe(df_cashflow.style.format({
                "工程費": "{:,.2f}",
                "重劃費用": "{:,.2f}",
                "支出小計": "{:,.2f}",
                "收入": "{:,.2f}",
                "當期淨現金流量": "{:,.2f}",
                "貸款利息": "{:,.2f}",
                "期末餘額": "{:,.2f}"
            }, na_rep="-"), use_container_width=True)
            
            # ===== 計算NPV和IRR =====
            st.markdown("---")
            st.subheader("📊 財務指標")
            
            # 重劃總費用合計（含利息）
            total_dev_cost_with_interest = construction_total + replot_fee_total + total_interest
            
            # 盈餘
            surplus = disposal_income - total_dev_cost_with_interest
            
            # 益本比
            benefit_cost_ratio = disposal_income / total_dev_cost_with_interest if total_dev_cost_with_interest > 0 else 0
            
            # NPV計算（使用貸款利率作為折現率）
            discount_rate = loan_rate / 100
            npv = sum(cf / ((1 + discount_rate) ** (i + 1)) for i, cf in enumerate(yearly_cash_flows))
            
            # IRR計算（使用二分法）
            def calc_npv_for_irr(rate, cfs):
                return sum(cf / ((1 + rate) ** (i + 1)) for i, cf in enumerate(cfs))
            
            try:
                low, high = -0.99, 1.0
                irr_found = False
                for _ in range(500):
                    mid = (low + high) / 2
                    npv_mid = calc_npv_for_irr(mid, yearly_cash_flows)
                    if abs(npv_mid) < 0.0001:
                        irr_found = True
                        break
                    if npv_mid > 0:
                        low = mid
                    else:
                        high = mid
                irr_pct = mid * 100 if irr_found else 0
            except:
                irr_pct = 0
            
            # 顯示財務指標
            col_f1, col_f2, col_f3, col_f4 = st.columns(4)
            col_f1.metric("1.重劃總費用合計", f"{total_dev_cost_with_interest:,.2f} 萬元")
            col_f2.metric("2.抵費地處分收入", f"{disposal_income:,.2f} 萬元")
            col_f3.metric("盈餘 (2)-(1)", f"{surplus:,.2f} 萬元")
            col_f4.metric("益本比", f"{benefit_cost_ratio:.3f}")
            
            col_n1, col_n2 = st.columns(2)
            col_n1.metric("NPV（折現率=貸款利率）", f"{npv:,.2f} 萬元")
            col_n2.metric("IRR", f"{irr_pct:.3f}%")
            
            # 儲存結果
            st.session_state['financial_results'] = {
                'total_dev_cost': total_dev_cost_with_interest,
                'disposal_income': disposal_income,
                'surplus': surplus,
                'npv': npv,
                'irr': irr_pct,
                'benefit_cost_ratio': benefit_cost_ratio,
                'cash_flow_df': df_cashflow,
                'total_interest': total_interest,
                'construction_total': construction_total,
                'replot_fee_total': replot_fee_total
            }
            
            # ===== 重劃前後地價檢核 =====
            st.markdown("---")
            st.subheader("🔍 重劃前後地價檢核")
            st.caption("檢核重劃後土地總價值是否大於重劃前，以確保土地所有權人有意願參與重劃")
            
            col_check1, col_check2 = st.columns([1, 1])
            
            with col_check1:
                st.markdown("**參數設定**")
                
                # 取得相關數據
                post_avg_price_check = st.session_state.get('weighted_price_sqm', land_disposal_price)
                # 重劃前平均地價應採用 Tab 5（預期開發分析法）計算之「重劃前價格日期之重劃前平均地價」
                # 不使用 Tab 1（土地交易分析）的 `res_land_sqm`（該為重劃後價格日期之比較法結果）
                pre_avg_price_check = float(st.session_state.get('pre_land_price_sqm', 0) or 0)
                public_shared_area = st.session_state.get('public_shared_area', 5000)
                offset_area = st.session_state.get('offset_area', 0)

                _pre_src_note = "來源：預期開發分析法（重劃前價格日期）"
                if pre_avg_price_check <= 0:
                    _pre_src_note = "⚠️ 尚未取得，請先完成 預期開發分析法計算"

                st.markdown(f"""
                **已知數據（來自其他頁籤）**
                - 重劃區總面積：{total_area_fin:,.0f} ㎡
                - 重劃後平均地價：{post_avg_price_check:,.0f} 元/㎡（來源：土地開發分析法 加權平均）
                - 重劃前平均地價：{pre_avg_price_check:,.0f} 元/㎡（{_pre_src_note}）
                - 共同負擔公設用地：{public_shared_area:,.0f} ㎡
                - 抵充地面積：{offset_area:,.0f} ㎡
                """)
                
                # 使用者輸入公告現值
                st.markdown("**公告現值設定**")
                avg_announced_value = _ni(
                    "本重劃區加權平均土地公告現值（元/㎡）", 
                    min_value=0.0, max_value=1000000.0, 
                    value=5000.0, step=100.0, key="avg_ann_val"
                )
            
            with col_check2:
                st.markdown("**負擔計算**")
                
                # 費用負擔 = (工程費用總額+重劃費用總額+貸款利息總額) / [重劃後平均地價 × (重劃區總面積-抵充地)]
                # 注意：現金流量表的數據是萬元，需要轉換
                cost_total_yuan = (construction_total + replot_fee_total + total_interest) * 10000  # 轉為元
                
                denominator = total_area_fin - offset_area
                if denominator > 0 and post_avg_price_check > 0:
                    # 正確公式：費用負擔 = 費用總額 / (重劃後地價 × 可分配面積)
                    fee_burden_rate = cost_total_yuan / (post_avg_price_check * denominator) * 100  # 費用負擔比率(%)
                else:
                    fee_burden_rate = 0
                
                # 公設用地負擔 = (共同負擔公設用地面積 - 抵充地) / (重劃區總面積 - 抵充地)
                public_burden_rate = (public_shared_area - offset_area) / denominator * 100 if denominator > 0 else 0
                
                # 總負擔比率
                total_burden_rate = fee_burden_rate + public_burden_rate
                
                st.markdown(f"""
                | 負擔項目 | 計算式 | 比率 |
                |----------|--------|------|
                | 費用負擔 | (工程費+重劃費+利息)÷[重劃後地價×(總面積-抵充地)] | **{fee_burden_rate:.2f}%** |
                | 公設用地負擔 | (公設用地-抵充地)÷(總面積-抵充地) | **{public_burden_rate:.2f}%** |
                | **總負擔** | 費用負擔 + 公設用地負擔 | **{total_burden_rate:.2f}%** |
                """)
                
                # 檢核45%上限
                if total_burden_rate > 45:
                    st.error(f"⚠️ 總負擔 {total_burden_rate:.2f}% 超過法定上限 45%")
                else:
                    st.success(f"✅ 總負擔 {total_burden_rate:.2f}% 未超過法定上限 45%")

            # ============================================================
            # C 值計算（回送 Tab 3 街廓互動分析）
            # 依據：《市地重劃實施辦法》第 29 條附件二
            #   C = (工程費用 + 重劃費用 + 貸款利息) ÷
            #       [重劃後平均地價 × (重劃區總面積 − 公共設施用地負擔總面積)]
            # 本頁的金額單位為「萬元」，公式內部須換算為「元」
            # ============================================================
            st.markdown("---")
            st.markdown("### 🔗 C 值計算（費用負擔係數）— 回送 街廓互動分析")

            _f3_inputs = st.session_state.get('f3_to_finance_inputs')
            if _f3_inputs:
                st.info(
                    f"✅ 已自動從 街廓互動分析帶入："
                    f"重劃區總面積 {float(_f3_inputs.get('total_area', 0) or 0):,.2f} ㎡、"
                    f"公共設施用地負擔總面積 {float(_f3_inputs.get('public_common_total', 0) or 0):,.2f} ㎡"
                )
                _c_total_area = float(_f3_inputs.get('total_area', 0) or 0) or float(total_area_fin)
                _c_pub_burden = float(_f3_inputs.get('public_common_total', 0) or 0) or float(public_shared_area)
                _c_price_after = float(_f3_inputs.get('price_after', 0) or 0) or float(post_avg_price_check)
            else:
                st.warning(
                    "⚠️ 尚未從 街廓互動分析 推送參數，將以本頁既有之「重劃區總面積」、"
                    "「共同負擔公設用地面積」及「重劃後平均地價」估算 C 值。"
                    "建議先於 街廓互動分析 完成街廓分類後按下「📤 推送 B 值與參數至 財務分析」。"
                )
                _c_total_area = float(total_area_fin)
                _c_pub_burden = float(public_shared_area)
                _c_price_after = float(post_avg_price_check)

            # 金額換算為元
            _eng_cost_yuan = float(construction_total) * 10000.0
            _redev_cost_yuan = float(replot_fee_total) * 10000.0
            _loan_int_yuan = float(total_interest) * 10000.0
            _cost_sum_yuan = _eng_cost_yuan + _redev_cost_yuan + _loan_int_yuan

            C_value = calc_C_value(
                engineering_cost=_eng_cost_yuan,
                redev_cost=_redev_cost_yuan,
                loan_interest=_loan_int_yuan,
                price_after=_c_price_after,
                total_area=_c_total_area,
                public_burden_total=_c_pub_burden,
            )

            cC1, cC2, cC3, cC4 = st.columns(4)
            cC1.metric("工程費用（元）", f"{_eng_cost_yuan:,.0f}")
            cC2.metric("重劃費用（元）", f"{_redev_cost_yuan:,.0f}")
            cC3.metric("貸款利息（元）", f"{_loan_int_yuan:,.0f}")
            cC4.metric("C 值（費用負擔係數）", f"{C_value:.6f}")

            with st.expander("📐 C 值計算明細"):
                _denom_val = _c_price_after * (_c_total_area - _c_pub_burden)
                st.markdown(f"""
                **公式**：C = (工程費用 + 重劃費用 + 貸款利息) ÷ [重劃後平均地價 × (重劃區總面積 − 公共設施用地負擔總面積)]

                | 項目 | 數值 |
                |------|------|
                | 工程費用 | {_eng_cost_yuan:,.0f} 元 |
                | 重劃費用 | {_redev_cost_yuan:,.0f} 元 |
                | 貸款利息 | {_loan_int_yuan:,.0f} 元 |
                | 費用總額 | **{_cost_sum_yuan:,.0f} 元** |
                | 重劃後平均地價 | {_c_price_after:,.0f} 元/㎡ |
                | 重劃區總面積 | {_c_total_area:,.2f} ㎡ |
                | 公共設施用地負擔總面積 | {_c_pub_burden:,.2f} ㎡ |
                | 分母 | {_denom_val:,.0f} |
                | **C 值** | **{C_value:.6f}** |
                """)

            # 自動回寫 session_state，供 Tab 3 重算 B 值
            st.session_state['f3_C_from_finance'] = float(C_value)
            st.success(
                f"✅ 已將 C = {C_value:.6f} 回送至 街廓互動分析，"
                f"街廓互動分析 的臨街地特別負擔總面積與 B 值將自動重算。"
            )

            # 計算三個地價總值
            st.markdown("---")
            st.markdown("**重劃前後地價檢核結果**")
            
            # A: 重劃後土地總價值 = 總面積 × (1-總負擔) × 重劃後平均地價
            value_A = total_area_fin * (1 - total_burden_rate/100) * post_avg_price_check
            
            # B: 重劃前土地總價值 = 重劃前平均地價 × 總面積
            value_B = pre_avg_price_check * total_area_fin if pre_avg_price_check > 0 else 0
            
            # C: 公告現值計算土地總地價 = 加權平均公告現值 × 總面積
            value_C = avg_announced_value * total_area_fin
            
            col_v1, col_v2, col_v3 = st.columns(3)
            col_v1.metric("A.重劃後土地總價值", f"{value_A/100000000:,.2f} 億元", 
                         help=f"{total_area_fin:,.0f}㎡ × (1-{total_burden_rate:.2f}%) × {post_avg_price_check:,.0f}元/㎡")
            col_v2.metric("B.重劃前土地總價值", f"{value_B/100000000:,.2f} 億元",
                         help=f"{pre_avg_price_check:,.0f}元/㎡ × {total_area_fin:,.0f}㎡")
            col_v3.metric("C.公告現值總地價", f"{value_C/100000000:,.2f} 億元",
                         help=f"{avg_announced_value:,.0f}元/㎡ × {total_area_fin:,.0f}㎡")
            
            # 檢核 A > B > C
            st.markdown("**檢核結果**")
            
            check_results = []
            if value_A > value_B:
                check_results.append(("✅", "A > B", "重劃後價值 > 重劃前價值", "土地所有權人參與重劃有利"))
            else:
                check_results.append(("❌", "A ≤ B", "重劃後價值 ≤ 重劃前價值", "土地所有權人參與重劃無利可圖"))
            
            if value_B > value_C:
                check_results.append(("✅", "B > C", "重劃前價值 > 公告現值", "評定地價合理"))
            else:
                check_results.append(("⚠️", "B ≤ C", "重劃前價值 ≤ 公告現值", "評定地價可能偏低"))
            
            if value_A > value_C:
                check_results.append(("✅", "A > C", "重劃後價值 > 公告現值", "重劃具有開發效益"))
            else:
                check_results.append(("❌", "A ≤ C", "重劃後價值 ≤ 公告現值", "重劃開發效益不佳"))
            
            for icon, condition, desc, conclusion in check_results:
                st.markdown(f"{icon} **{condition}**：{desc} → {conclusion}")
            
            # 整體評估
            if value_A > value_B > value_C:
                st.success("🎉 **整體評估通過**：A > B > C，重劃具有財務可行性，土地所有權人有意願參與重劃")
            else:
                st.warning("⚠️ **整體評估未通過**：未滿足 A > B > C 條件，建議檢討重劃計畫")
            
            # 儲存檢核結果
            st.session_state['price_check_results'] = {
                'value_A': value_A,
                'value_B': value_B,
                'value_C': value_C,
                'fee_burden_rate': fee_burden_rate,
                'public_burden_rate': public_burden_rate,
                'total_burden_rate': total_burden_rate,
                'check_passed': value_A > value_B > value_C
            }
            
            # ===== 匯出功能 =====
            st.markdown("---")
            st.subheader("📄 報表匯出（直式A4）")
            
            if st.button("匯出現金流量表Excel（直式A4）", key="export_cashflow_excel"):
                try:
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    import io
                    
                    # 創建工作簿
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "現金流量表(NPV)"
                    
                    # 定義樣式
                    title_font = Font(name='Arial', size=14, bold=True)
                    header_font = Font(name='Arial', size=11, bold=True)
                    header_font_white = Font(name='Arial', size=10, bold=True, color='FFFFFF')
                    normal_font = Font(name='Arial', size=10)
                    
                    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                    light_blue_fill = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')
                    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    total_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                    
                    thin_border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
                    
                    center_align = Alignment(horizontal='center', vertical='center')
                    right_align = Alignment(horizontal='right', vertical='center')
                    left_align = Alignment(horizontal='left', vertical='center')
                    
                    # 設定欄寬
                    col_widths = [8, 12, 12, 12, 12, 15, 12, 12, 12]
                    for i, width in enumerate(col_widths, 1):
                        ws.column_dimensions[chr(64+i)].width = width
                    
                    # 標題
                    ws.merge_cells('A1:I1')
                    ws['A1'] = '市地重劃財務可行性分析 - 年度現金流量表'
                    ws['A1'].font = title_font
                    ws['A1'].alignment = center_align
                    
                    ws.merge_cells('A2:I2')
                    ws['A2'] = '（單位：萬元）'
                    ws['A2'].font = normal_font
                    ws['A2'].alignment = Alignment(horizontal='right', vertical='center')
                    
                    # 現金流量表表頭
                    row = 4
                    headers = ['年度', '工程費', '重劃費用', '支出小計', '收入', '當期淨現金流量', '期初淨額', '貸款利息', '期末餘額']
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=row, column=col, value=header)
                        cell.font = header_font_white
                        cell.fill = header_fill
                        cell.alignment = center_align
                        cell.border = thin_border
                    
                    # 現金流量表資料
                    for data_row in cash_flow_data:
                        row += 1
                        values = [data_row['年度'], data_row['工程費'], data_row['重劃費用'], 
                                 data_row['支出小計'], data_row['收入'], data_row['當期淨現金流量'],
                                 data_row['期初淨額'], data_row['貸款利息'], data_row['期末餘額']]
                        for col, value in enumerate(values, 1):
                            cell = ws.cell(row=row, column=col, value=value if value != "-" else "")
                            cell.font = normal_font
                            cell.border = thin_border
                            if col == 1:
                                cell.alignment = center_align
                                if data_row['年度'] == '合計':
                                    cell.fill = total_fill
                                    cell.font = header_font
                            else:
                                cell.alignment = right_align
                                if isinstance(value, (int, float)):
                                    cell.number_format = '#,##0.00'
                                if data_row['年度'] == '合計':
                                    cell.fill = total_fill
                    
                    # 財務指標
                    row += 2
                    ws.merge_cells(f'A{row}:C{row}')
                    ws[f'A{row}'] = '財務指標'
                    ws[f'A{row}'].font = header_font
                    ws[f'A{row}'].fill = light_blue_fill
                    ws[f'A{row}'].alignment = center_align
                    for col in range(1, 4):
                        ws.cell(row=row, column=col).border = thin_border
                    
                    indicators = [
                        ['1.重劃總費用合計（萬元）', total_dev_cost_with_interest],
                        ['2.抵費地處分收入（萬元）', disposal_income],
                        ['盈餘 (2)-(1)（萬元）', surplus],
                        ['益本比', benefit_cost_ratio],
                        ['NPV（萬元）', npv],
                        ['IRR', f"{irr_pct:.3f}%"],
                    ]
                    
                    for ind in indicators:
                        row += 1
                        ws.cell(row=row, column=1, value=ind[0]).font = normal_font
                        ws.cell(row=row, column=1).border = thin_border
                        ws.cell(row=row, column=1).alignment = left_align
                        
                        cell_val = ws.cell(row=row, column=2, value=ind[1])
                        cell_val.font = Font(name='Arial', size=10, bold=True)
                        cell_val.border = thin_border
                        cell_val.alignment = right_align
                        if isinstance(ind[1], float):
                            cell_val.number_format = '#,##0.00' if abs(ind[1]) >= 1 else '0.000'
                        
                        ws.cell(row=row, column=3).border = thin_border
                    
                    # 重劃總費用明細
                    row += 2
                    ws.merge_cells(f'A{row}:E{row}')
                    ws[f'A{row}'] = '重劃總費用及盈餘明細'
                    ws[f'A{row}'].font = header_font
                    ws[f'A{row}'].fill = light_blue_fill
                    ws[f'A{row}'].alignment = center_align
                    
                    row += 1
                    cost_headers = ['項目', '單價(元/㎡)', '面積(㎡)', '金額(萬元)', '備註']
                    for col, header in enumerate(cost_headers, 1):
                        cell = ws.cell(row=row, column=col, value=header)
                        cell.font = header_font_white
                        cell.fill = header_fill
                        cell.alignment = center_align
                        cell.border = thin_border
                    
                    cost_items = [
                        ['地上物拆遷補償費', demolition_cost, total_area_fin, demolition_total, ''],
                        ['行政作業費', admin_cost, total_area_fin, admin_total, ''],
                        ['工程費用', construction_cost, total_area_fin, construction_total, ''],
                        ['貸款利息', f'{loan_rate:.3f}%', f'期間{total_years}年', total_interest, ''],
                        ['1.合計', '', '', total_dev_cost_with_interest, '重劃總費用總計'],
                    ]
                    
                    for item in cost_items:
                        row += 1
                        for col, value in enumerate(item, 1):
                            cell = ws.cell(row=row, column=col, value=value)
                            cell.font = normal_font
                            cell.border = thin_border
                            if col == 1:
                                cell.alignment = left_align
                                if str(item[0]).startswith('1.'):
                                    cell.font = header_font
                                    cell.fill = total_fill
                            elif col in [2, 3, 4]:
                                cell.alignment = right_align
                                if isinstance(value, (int, float)):
                                    cell.number_format = '#,##0.00' if col == 4 else '#,##0'
                                if str(item[0]).startswith('1.'):
                                    cell.fill = total_fill
                    
                    # 抵費地處分收入
                    row += 1
                    ws.cell(row=row, column=1, value='2.抵費地處分收入').font = header_font
                    ws.cell(row=row, column=1).fill = total_fill
                    ws.cell(row=row, column=1).border = thin_border
                    ws.cell(row=row, column=2, value=land_disposal_price).number_format = '#,##0'
                    ws.cell(row=row, column=2).border = thin_border
                    ws.cell(row=row, column=2).alignment = right_align
                    ws.cell(row=row, column=3, value=actual_disposal_area).number_format = '#,##0'
                    ws.cell(row=row, column=3).border = thin_border
                    ws.cell(row=row, column=3).alignment = right_align
                    ws.cell(row=row, column=4, value=disposal_income).number_format = '#,##0.00'
                    ws.cell(row=row, column=4).border = thin_border
                    ws.cell(row=row, column=4).alignment = right_align
                    ws.cell(row=row, column=4).fill = total_fill
                    ws.cell(row=row, column=5, value=f'單價×面積÷10000').border = thin_border
                    
                    # 盈餘
                    row += 1
                    ws.cell(row=row, column=1, value='盈餘 (2)-(1)').font = Font(name='Arial', size=11, bold=True)
                    ws.cell(row=row, column=1).fill = yellow_fill
                    ws.cell(row=row, column=1).border = thin_border
                    for col in range(2, 5):
                        ws.cell(row=row, column=col).border = thin_border
                    ws.cell(row=row, column=4, value=surplus)
                    ws.cell(row=row, column=4).font = Font(name='Arial', size=11, bold=True)
                    ws.cell(row=row, column=4).fill = yellow_fill
                    ws.cell(row=row, column=4).number_format = '#,##0.00'
                    ws.cell(row=row, column=4).alignment = right_align
                    
                    # 說明
                    row += 2
                    ws[f'A{row}'] = '【說明】'
                    ws[f'A{row}'].font = header_font
                    
                    notes = [
                        f'1. 重劃區總面積：{total_area_fin:,.0f} ㎡，抵費地面積：{actual_disposal_area:,.0f} ㎡',
                        f'2. 抵費地單價：{land_disposal_price:,.0f} 元/㎡（土地開發分析法與比較法加權平均）',
                        f'3. 工程費及重劃費用各年度分配比例由使用者設定',
                        f'4. 抵費地收入於第{total_years}年（最後一年）年初收取',
                        f'5. 貸款利率採五大銀行平均放款利率 {loan_rate:.3f}%',
                        f'6. 貸款利息依各年度期初負債餘額計算',
                    ]
                    
                    for note in notes:
                        row += 1
                        ws[f'A{row}'] = note
                        ws[f'A{row}'].font = Font(name='Arial', size=9)
                    
                    # 頁面設定
                    ws.print_area = f'A1:I{row}'
                    ws.page_setup.orientation = 'portrait'
                    ws.page_setup.paperSize = ws.PAPERSIZE_A4
                    ws.page_setup.fitToPage = True
                    ws.page_setup.fitToWidth = 1
                    ws.page_setup.fitToHeight = 0
                    
                    # 儲存到記憶體
                    output = io.BytesIO()
                    wb.save(output)
                    output.seek(0)
                    
                    st.download_button(
                        label="📥 下載現金流量表Excel（直式A4）",
                        data=output.getvalue(),
                        file_name="現金流量表_NPV.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    st.success("✅ Excel檔案已生成")
                    
                except Exception as e:
                    st.error(f"匯出Excel時發生錯誤：{str(e)}")
            
            # ===== 統計分析匯出功能 =====
            st.markdown("---")
            if st.button("匯出完整統計分析報告Excel", key="export_full_analysis"):
                try:
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    import io
                    
                    wb = Workbook()
                    
                    # 定義樣式
                    title_font = Font(name='Arial', size=14, bold=True)
                    header_font = Font(name='Arial', size=11, bold=True)
                    header_font_white = Font(name='Arial', size=10, bold=True, color='FFFFFF')
                    normal_font = Font(name='Arial', size=10)
                    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                    light_fill = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')
                    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                        top=Side(style='thin'), bottom=Side(style='thin'))
                    
                    # ===== 工作表1：統計分析結果 =====
                    ws1 = wb.active
                    ws1.title = "統計分析結果"
                    
                    ws1['A1'] = '市地重劃前後地價評估 - 統計分析結果'
                    ws1['A1'].font = title_font
                    ws1.merge_cells('A1:F1')
                    
                    # 重劃後地價評估
                    row = 3
                    ws1[f'A{row}'] = '一、重劃後平均地價評估'
                    ws1[f'A{row}'].font = header_font
                    ws1[f'A{row}'].fill = light_fill
                    
                    row += 1
                    headers = ['估價方法', '土地單價(元/㎡)', '土地單價(元/坪)', '權重(%)', '加權單價(元/㎡)']
                    for col, h in enumerate(headers, 1):
                        cell = ws1.cell(row=row, column=col, value=h)
                        cell.font = header_font_white
                        cell.fill = header_fill
                        cell.border = thin_border
                    
                    # 取得評估結果
                    land_trans = st.session_state.get('res_land_sqm', 0)
                    dev_price = st.session_state.get('dev_land_price_sqm', 0)
                    weighted = st.session_state.get('weighted_price_sqm', 0)
                    
                    row += 1
                    data = ['比較法（土地交易分析）', land_trans, land_trans/0.3025 if land_trans > 0 else 0, 60, land_trans * 0.6]
                    for col, val in enumerate(data, 1):
                        cell = ws1.cell(row=row, column=col, value=val)
                        cell.border = thin_border
                        if col >= 2 and isinstance(val, (int, float)):
                            cell.number_format = '#,##0'
                    
                    row += 1
                    data = ['土地開發分析法', dev_price, dev_price/0.3025 if dev_price > 0 else 0, 40, dev_price * 0.4]
                    for col, val in enumerate(data, 1):
                        cell = ws1.cell(row=row, column=col, value=val)
                        cell.border = thin_border
                        if col >= 2 and isinstance(val, (int, float)):
                            cell.number_format = '#,##0'
                    
                    row += 1
                    ws1.cell(row=row, column=1, value='加權平均').font = header_font
                    ws1.cell(row=row, column=1).fill = green_fill
                    ws1.cell(row=row, column=1).border = thin_border
                    ws1.cell(row=row, column=2, value=weighted).number_format = '#,##0'
                    ws1.cell(row=row, column=2).fill = green_fill
                    ws1.cell(row=row, column=2).border = thin_border
                    ws1.cell(row=row, column=3, value=weighted/0.3025 if weighted > 0 else 0).number_format = '#,##0'
                    ws1.cell(row=row, column=3).fill = green_fill
                    ws1.cell(row=row, column=3).border = thin_border
                    ws1.cell(row=row, column=4, value=100).fill = green_fill
                    ws1.cell(row=row, column=4).border = thin_border
                    ws1.cell(row=row, column=5, value=weighted).number_format = '#,##0'
                    ws1.cell(row=row, column=5).fill = green_fill
                    ws1.cell(row=row, column=5).border = thin_border
                    
                    # 重劃前地價評估（採用 Tab 5 預期開發分析法之「重劃前價格日期之重劃前平均地價」）
                    row += 2
                    ws1[f'A{row}'] = '二、重劃前平均地價評估（預期開發分析法）'
                    ws1[f'A{row}'].font = header_font
                    ws1[f'A{row}'].fill = light_fill

                    _pre_land_price_t5 = float(st.session_state.get('pre_land_price_sqm', 0) or 0)
                    row += 1
                    ws1.cell(row=row, column=1, value='預期開發分析法（重劃前價格日期）').border = thin_border
                    ws1.cell(row=row, column=2, value=_pre_land_price_t5).number_format = '#,##0'
                    ws1.cell(row=row, column=2).border = thin_border
                    ws1.cell(row=row, column=3, value=(_pre_land_price_t5/0.3025) if _pre_land_price_t5 > 0 else 0).number_format = '#,##0'
                    ws1.cell(row=row, column=3).border = thin_border
                    
                    # 財務分析結果
                    row += 2
                    ws1[f'A{row}'] = '三、財務可行性分析結果'
                    ws1[f'A{row}'].font = header_font
                    ws1[f'A{row}'].fill = light_fill
                    
                    fin_items = [
                        ['重劃總費用合計（萬元）', total_dev_cost_with_interest],
                        ['抵費地處分收入（萬元）', disposal_income],
                        ['盈餘（萬元）', surplus],
                        ['益本比', benefit_cost_ratio],
                        ['NPV（萬元）', npv],
                        ['IRR（%）', irr_pct],
                    ]
                    for item in fin_items:
                        row += 1
                        ws1.cell(row=row, column=1, value=item[0]).border = thin_border
                        cell = ws1.cell(row=row, column=2, value=item[1])
                        cell.border = thin_border
                        cell.number_format = '#,##0.00' if isinstance(item[1], float) else '#,##0'
                    
                    # 重劃前後地價檢核
                    row += 2
                    ws1[f'A{row}'] = '四、重劃前後地價檢核'
                    ws1[f'A{row}'].font = header_font
                    ws1[f'A{row}'].fill = light_fill
                    
                    check_data = st.session_state.get('price_check_results', {})
                    check_items = [
                        ['費用負擔比率（%）', check_data.get('fee_burden_rate', 0)],
                        ['公設用地負擔比率（%）', check_data.get('public_burden_rate', 0)],
                        ['總負擔比率（%）', check_data.get('total_burden_rate', 0)],
                        ['A.重劃後土地總價值（億元）', check_data.get('value_A', 0)/100000000],
                        ['B.重劃前土地總價值（億元）', check_data.get('value_B', 0)/100000000],
                        ['C.公告現值總地價（億元）', check_data.get('value_C', 0)/100000000],
                        ['檢核結果 A>B>C', '通過' if check_data.get('check_passed', False) else '未通過'],
                    ]
                    for item in check_items:
                        row += 1
                        ws1.cell(row=row, column=1, value=item[0]).border = thin_border
                        cell = ws1.cell(row=row, column=2, value=item[1])
                        cell.border = thin_border
                        if isinstance(item[1], float):
                            cell.number_format = '#,##0.00'
                        if item[0] == '檢核結果 A>B>C':
                            cell.fill = green_fill if item[1] == '通過' else yellow_fill
                    
                    # ===== 工作表2：現金流量表 =====
                    ws2 = wb.create_sheet("現金流量表")
                    ws2['A1'] = '年度現金流量表（單位：萬元）'
                    ws2['A1'].font = title_font
                    
                    row = 3
                    cf_headers = ['年度', '工程費', '重劃費用', '支出小計', '收入', '當期淨現金流量', '期初淨額', '貸款利息', '期末餘額']
                    for col, h in enumerate(cf_headers, 1):
                        cell = ws2.cell(row=row, column=col, value=h)
                        cell.font = header_font_white
                        cell.fill = header_fill
                        cell.border = thin_border
                    
                    for cf_row in cash_flow_data:
                        row += 1
                        values = [cf_row['年度'], cf_row['工程費'], cf_row['重劃費用'], cf_row['支出小計'],
                                 cf_row['收入'], cf_row['當期淨現金流量'], cf_row['期初淨額'], 
                                 cf_row['貸款利息'], cf_row['期末餘額']]
                        for col, val in enumerate(values, 1):
                            cell = ws2.cell(row=row, column=col, value=val if val != "-" else "")
                            cell.border = thin_border
                            if col > 1 and isinstance(val, (int, float)):
                                cell.number_format = '#,##0.00'
                    
                    # ===== 工作表3：區段分析 =====
                    ws3 = wb.create_sheet("區段分析")
                    ws3['A1'] = '重劃後分配街廓地價分析'
                    ws3['A1'].font = title_font
                    
                    block_results = st.session_state.get('post_block_results', [])
                    if block_results:
                        row = 3
                        block_headers = ['街廓', '類型', '分類', '面積(㎡)', '相對比率(%)', '單價(元/㎡)', '單價(元/坪)', '總價(元)']
                        for col, h in enumerate(block_headers, 1):
                            cell = ws3.cell(row=row, column=col, value=h)
                            cell.font = header_font_white
                            cell.fill = header_fill
                            cell.border = thin_border
                        
                        for br in block_results:
                            row += 1
                            values = [br.get('街廓', ''), br.get('類型', ''), br.get('分類', ''),
                                     br.get('面積(㎡)', 0), br.get('相對比率(%)', 0), 
                                     br.get('單價(元/㎡)', 0), br.get('單價(元/坪)', 0), br.get('總價(元)', 0)]
                            for col, val in enumerate(values, 1):
                                cell = ws3.cell(row=row, column=col, value=val)
                                cell.border = thin_border
                                if col >= 4 and isinstance(val, (int, float)):
                                    cell.number_format = '#,##0'
                    
                    # 調整欄寬
                    for ws in [ws1, ws2, ws3]:
                        for col in range(1, 10):
                            ws.column_dimensions[chr(64+col)].width = 18
                    
                    # 儲存
                    output = io.BytesIO()
                    wb.save(output)
                    output.seek(0)
                    
                    st.download_button(
                        label="📥 下載完整統計分析報告",
                        data=output.getvalue(),
                        file_name="重劃地價評估_統計分析報告.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    st.success("✅ 統計分析報告已生成")
                    
                except Exception as e:
                    st.error(f"匯出統計分析報告時發生錯誤：{str(e)}")

    # ===== Tab: 土地歸戶 =====
    elif selected_tab == TAB_LAND_AGGREGATION:
        st.markdown("### 功能：上傳地籍資料來源.xlsx → 自動對應 U_LAND 資料 → 下載重劃區地籍對照表")
        st.caption("📌 請上傳含有「U_LAND」與「重劃區地籍」兩個工作表的地籍資料來源.xlsx")

        uploaded_cadastral = st.file_uploader(
            "選擇地籍資料來源.xlsx", type=["xlsx"], key="f7_cadastral"
        )
        # 上傳時存入 session_state；切回此頁時從快取還原
        if uploaded_cadastral:
            st.session_state['f7_saved'] = (uploaded_cadastral.name, uploaded_cadastral.getvalue())
        elif st.session_state.get('f7_saved'):
            n7, d7 = st.session_state['f7_saved']
            uploaded_cadastral = _NamedBytesIO(n7, d7)
            st.caption("📂 已從快取載入上次上傳的檔案")

        if uploaded_cadastral:
            try:
                import io as _io

                with st.spinner("讀取工作表中…"):
                    xl7 = pd.ExcelFile(uploaded_cadastral, engine="openpyxl")
                    if "U_LAND" not in xl7.sheet_names or "重劃區地籍" not in xl7.sheet_names:
                        st.error("❌ 檔案內找不到「U_LAND」或「重劃區地籍」工作表，請確認檔案格式。")
                        st.stop()

                    df_uland7 = pd.read_excel(xl7, sheet_name="U_LAND", header=0, engine="openpyxl")
                    df_rezoning7 = pd.read_excel(xl7, sheet_name="重劃區地籍", header=0, engine="openpyxl")

                st.success(f"✅ U_LAND：{len(df_uland7):,} 筆｜重劃區地籍：{len(df_rezoning7)} 筆")

                # ── 前處理 U_LAND ──────────────────────────────────────────
                # 去掉段小段的數字前綴（如 '0098國股段' → '國股段'）
                df_uland7["段名"] = df_uland7["段小段"].astype(str).str.replace(r"^\d+", "", regex=True)
                df_uland7["地號int"] = pd.to_numeric(df_uland7["地號"], errors="coerce")

                def 重劃地號轉ULAND整數(地號文字):
                    """'658' → 6580000；'670-2' → 6700002"""
                    s = str(地號文字).strip()
                    if "-" in s:
                        parts = s.split("-", 1)
                        return int(parts[0]) * 10000 + int(parts[1])
                    return int(s) * 10000

                def 格式化他項權利(mort_df):
                    """整合同一地號的所有抵押資訊為字串"""
                    if mort_df.empty:
                        return ""
                    parts = []
                    # 以 (設定義務人, 姓名.2) 去重，避免因多筆所有權人重複展開
                    seen = set()
                    for _, mr in mort_df.iterrows():
                        obligor = str(mr.get("設定義務人", "")) if pd.notna(mr.get("設定義務人")) else ""
                        creditor = str(mr.get("姓名.2", "")) if pd.notna(mr.get("姓名.2")) else ""
                        deno = mr.get("債權權利範圍持分分母")
                        numer = mr.get("債權權利範圍持分分子")
                        key = (obligor, creditor)
                        if key in seen or (not obligor and not creditor):
                            continue
                        seen.add(key)
                        try:
                            ratio = f"{int(numer)}/{int(deno)}" if pd.notna(deno) and pd.notna(numer) else ""
                        except Exception:
                            ratio = ""
                        parts.append(f"義務人：{obligor}／債權人：{creditor} {ratio}".strip())
                    return "；".join(parts)

                # ── 對應主邏輯 ────────────────────────────────────────────
                rows_out = []            # 對應成功的輸出列（每位所有權人一列）
                rows_fail = []           # 對應失敗的地號
                parcel_fp = {}           # {(地段, 地號): 歸戶指紋}

                for _, rz_row in df_rezoning7.iterrows():
                    seg = str(rz_row["地段"]).strip()
                    landno_str = str(rz_row["地號"]).strip()
                    city = str(rz_row["鄉鎮市"]).strip()

                    try:
                        target_int = 重劃地號轉ULAND整數(landno_str)
                    except Exception:
                        rows_fail.append({"鄉鎮市": city, "地段": seg, "地號": landno_str, "原因": "地號格式無法解析"})
                        continue

                    matched = df_uland7[
                        (df_uland7["段名"] == seg) & (df_uland7["地號int"] == target_int)
                    ]

                    if matched.empty:
                        rows_fail.append({"鄉鎮市": city, "地段": seg, "地號": landno_str, "原因": "U_LAND 無對應資料"})
                        continue

                    # 面積、公告現值、申報地價（同地號相同，取第一非空值）
                    area7    = matched["面積"].dropna().iloc[0]    if matched["面積"].notna().any()    else None
                    gonggao7 = matched["公告現值"].dropna().iloc[0] if matched["公告現值"].notna().any() else None
                    shenbao7 = matched["申報地價"].dropna().iloc[0] if matched["申報地價"].notna().any() else None

                    # 他項權利顯示字串
                    mort_rows7 = matched[matched["設定義務人"].notna()]
                    he_xiang7 = 格式化他項權利(mort_rows7)

                    # ── 歸戶指紋計算 ──────────────────────────────────────
                    # 所有權人指紋：排序後（統編:持分）串接，確保順序不影響比對
                    own_fp_parts = []
                    own_rows7 = matched.copy()
                    own_rows7["_dedup_key"] = own_rows7.apply(
                        lambda r: str(r.get("所有權統一編號", "")).strip()
                        if pd.notna(r.get("所有權統一編號")) and str(r.get("所有權統一編號", "")).strip() not in ("", "nan")
                        else str(r.get("姓名", "")).strip(),
                        axis=1
                    )
                    own_rows7 = own_rows7.drop_duplicates(subset=["_dedup_key"])

                    for _, owner in own_rows7.iterrows():
                        uni = str(owner.get("所有權統一編號", "")).strip() if pd.notna(owner.get("所有權統一編號")) else ""
                        name7_fp = str(owner.get("姓名", "")).strip() if pd.notna(owner.get("姓名")) else ""
                        key_id = uni if uni else name7_fp
                        d7fp = owner.get("權利範圍分母")
                        n7fp = owner.get("權利範圍分子")
                        ratio_fp = f"{int(n7fp)}/{int(d7fp)}" if pd.notna(d7fp) and pd.notna(n7fp) else ""
                        own_fp_parts.append(f"{key_id}:{ratio_fp}")
                    own_fp_parts.sort()

                    # 他項權利指紋：含 權利種類（抵押/地上權等），排序後串接
                    mort_fp_parts = []
                    seen_m = set()
                    for _, mr in mort_rows7.iterrows():
                        obligor = str(mr.get("設定義務人", "")) if pd.notna(mr.get("設定義務人")) else ""
                        creditor = str(mr.get("姓名.2", ""))    if pd.notna(mr.get("姓名.2"))    else ""
                        kind    = str(mr.get("權利種類", ""))   if pd.notna(mr.get("權利種類"))   else ""
                        dm = mr.get("債權權利範圍持分分母")
                        nm = mr.get("債權權利範圍持分分子")
                        mkey = (obligor, creditor, kind)
                        if mkey in seen_m:
                            continue
                        seen_m.add(mkey)
                        mratio = f"{int(nm)}/{int(dm)}" if pd.notna(dm) and pd.notna(nm) else ""
                        mort_fp_parts.append(f"{obligor}→{creditor}[{kind}]{mratio}")
                    mort_fp_parts.sort()

                    fingerprint = "|".join(own_fp_parts) + "#" + ";".join(mort_fp_parts)
                    parcel_fp[(seg, landno_str)] = fingerprint

                    # ── 每位所有權人一列 ──────────────────────────────────
                    for _, owner in own_rows7.iterrows():
                        uni  = str(owner.get("所有權統一編號", "")) if pd.notna(owner.get("所有權統一編號")) else ""
                        name7 = str(owner.get("姓名", ""))          if pd.notna(owner.get("姓名"))          else ""
                        mgr_uni  = owner.get("管理者統一編號")
                        manager7 = str(owner.get("姓名.1", "")) if pd.notna(mgr_uni) and pd.notna(owner.get("姓名.1")) else ""
                        addr7    = str(owner.get("地址", ""))   if pd.notna(owner.get("地址")) else ""
                        deno7    = owner.get("權利範圍分母")
                        numer7   = owner.get("權利範圍分子")
                        try:
                            share_area7 = round(float(area7) * float(numer7) / float(deno7), 2) if (
                                area7 is not None and pd.notna(deno7) and pd.notna(numer7) and float(deno7) != 0
                            ) else area7
                        except Exception:
                            share_area7 = area7

                        rows_out.append({
                            "_parcel_key": (seg, landno_str),   # 暫存，後續加歸戶群組用
                            "鄉鎮市": city,
                            "地段": seg,
                            "地號": landno_str,
                            "統編": uni,
                            "所有權人": name7,
                            "管理機關": manager7,
                            "地址": addr7,
                            "土地面積㎡": area7,
                            "持分分母": int(deno7)  if pd.notna(deno7)  else "",
                            "持分分子": int(numer7) if pd.notna(numer7) else "",
                            "持分面積": share_area7,
                            "公告土地現值(元/㎡)": gonggao7,
                            "當期申報地價(元/㎡)": shenbao7,
                            "他項權利": he_xiang7,
                            "有無建物保存登記": "",
                            "有無限制登記": "",
                            "坐落街廓代碼": "",
                        })

                # ── 歸戶分組（指紋相同 → 同群組） ──────────────────────
                fp_to_group: dict = {}
                g_counter = 1
                for pk, fp in parcel_fp.items():
                    if fp not in fp_to_group:
                        fp_to_group[fp] = f"G{g_counter:03d}"
                        g_counter += 1

                for row in rows_out:
                    pk = row.pop("_parcel_key")
                    row["歸戶群組"] = fp_to_group.get(parcel_fp.get(pk, ""), "")

                # ── 寫入 session_state：供 Tab 3 步驟 K 跨街廓調配使用 ─────────
                # t8_ownership_map: {地號: 歸戶群組ID}（僅取地號，忽略地段差異）
                # t8_ownership_map_full: {(地段, 地號): 歸戶群組ID}
                # t8_ownership_groups: {歸戶群組ID: [地號列表]}
                # 🆕 Phase 5：改用模組級 _normalize_landno_module（含去前置零、子號 0 移除等強化）
                #    使 Tab 4 讀到 DXF 之「13-0628-0006」「715-0」等格式時也能查得到

                _own_map = {}
                _own_map_full = {}
                _own_groups: dict = {}
                for (seg_k, landno_k), fp in parcel_fp.items():
                    gid = fp_to_group.get(fp, "")
                    if not gid:
                        continue
                    _own_map[landno_k] = gid
                    _own_map_full[(seg_k, landno_k)] = gid
                    _own_groups.setdefault(gid, []).append(landno_k)
                    # Phase 5：改用模組級正規化函式（去前置零 + 子號 0 移除）
                    nk = _normalize_landno_module(landno_k)
                    if nk and nk != landno_k:
                        _own_map.setdefault(nk, gid)
                        _own_map_full.setdefault((seg_k, nk), gid)
                    # 同時寫入「-0 帶後綴」變體鍵，供反向查詢（Tab 1 是「715」、DXF 是「715-0」）
                    if '-' not in landno_k:
                        _own_map.setdefault(f"{landno_k}-0", gid)

                st.session_state["t8_ownership_map"] = _own_map
                st.session_state["t8_ownership_map_full"] = _own_map_full
                st.session_state["t8_ownership_groups"] = _own_groups

                # 🆕 Phase 11 Hotfix：t8_parcel_areas — {地號: 土地面積㎡}
                # 供 Tab 4 步驟 B 之地籍配對結果做面積交叉驗證
                _parcel_areas = {}
                _seen_pk = set()
                for row in rows_out:
                    seg_k = row.get('地段', '')
                    landno_k = row.get('地號', '')
                    if not landno_k:
                        continue
                    pk = (seg_k, landno_k)
                    if pk in _seen_pk:
                        continue   # 同地號之多筆持分只取第一筆 area
                    _seen_pk.add(pk)
                    try:
                        a = float(row.get('土地面積㎡', 0) or 0)
                    except Exception:
                        a = 0.0
                    if a > 0:
                        _parcel_areas[landno_k] = a
                        # 同步寫入正規化變體 key
                        nk = _normalize_landno_module(landno_k)
                        if nk and nk != landno_k:
                            _parcel_areas.setdefault(nk, a)
                        if '-' not in landno_k:
                            _parcel_areas.setdefault(f"{landno_k}-0", a)
                st.session_state["t8_parcel_areas"] = _parcel_areas

                # 欄位順序：歸戶群組放最前面
                col_order = [
                    "歸戶群組", "鄉鎮市", "地段", "地號", "統編", "所有權人", "管理機關",
                    "地址", "土地面積㎡", "持分分母", "持分分子", "持分面積",
                    "公告土地現值(元/㎡)", "當期申報地價(元/㎡)", "他項權利",
                    "有無建物保存登記", "有無限制登記", "坐落街廓代碼"
                ]

                # ── 顯示對應結果 ──────────────────────────────────────────
                st.markdown("---")
                n_ok = len(df_rezoning7) - len(rows_fail)
                n_groups = len(fp_to_group)
                st.markdown(
                    f"**對應結果：共 {len(df_rezoning7)} 筆地號，"
                    f"成功 {n_ok} 筆、失敗 {len(rows_fail)} 筆｜"
                    f"歸戶群組共 {n_groups} 組**"
                )

                if rows_fail:
                    st.warning(f"⚠️ 以下 {len(rows_fail)} 筆地號在 U_LAND 無對應資料，請手動確認：")
                    st.dataframe(pd.DataFrame(rows_fail), use_container_width=True)

                if rows_out:
                    result_df7 = pd.DataFrame(rows_out)[col_order]

                    st.markdown("#### 對應結果")
                    st.caption(
                        "歸戶群組：指紋（所有權人組合＋他項權利）相同者歸為同群組｜"
                        "O欄：有無建物保存登記｜P欄：有無限制登記　→ 選填：有／無／其他｜"
                        "坐落街廓代碼：可輸入多個代碼，以逗號分隔"
                    )

                    # 可編輯欄：O、P、坐落街廓代碼
                    disabled_cols = [
                        "歸戶群組", "鄉鎮市", "地段", "地號", "統編", "所有權人", "管理機關",
                        "地址", "土地面積㎡", "持分分母", "持分分子", "持分面積",
                        "公告土地現值(元/㎡)", "當期申報地價(元/㎡)", "他項權利"
                    ]
                    edited_df7 = st.data_editor(
                        result_df7,
                        use_container_width=True,
                        num_rows="fixed",
                        column_config={
                            "有無建物保存登記": st.column_config.SelectboxColumn(
                                "有無建物保存登記", options=["", "有", "無", "其他"], required=False),
                            "有無限制登記": st.column_config.SelectboxColumn(
                                "有無限制登記", options=["", "有", "無", "其他"], required=False),
                            "坐落街廓代碼": st.column_config.TextColumn(
                                "坐落街廓代碼", help="可輸入多個代碼，以逗號分隔，例如：R1,R2"),
                            "土地面積㎡":           st.column_config.NumberColumn(format="%.2f"),
                            "持分面積":             st.column_config.NumberColumn(format="%.2f"),
                            "公告土地現值(元/㎡)":  st.column_config.NumberColumn(format="%d"),
                            "當期申報地價(元/㎡)":  st.column_config.NumberColumn(format="%.0f"),
                        },
                        disabled=disabled_cols,
                        key="editor_tab7"
                    )

                    # ── 歸戶群組摘要表 ────────────────────────────────────
                    st.markdown("---")
                    st.subheader("📋 歸戶群組摘要")
                    grp_summary = []
                    for fp, gid in sorted(fp_to_group.items(), key=lambda x: x[1]):
                        # 找屬於此群組的所有地號
                        parcels_in_grp = [
                            f"{k[0]}{k[1]}" for k, v in parcel_fp.items() if v == fp
                        ]
                        # 從 rows_out 取出第一筆（已移除 _parcel_key）
                        sample_rows = [r for r in rows_out if fp_to_group.get(parcel_fp.get((r["地段"], r["地號"]), ""), "") == gid]
                        owners_in_grp = "、".join(
                            dict.fromkeys(r["所有權人"] for r in sample_rows if r["所有權人"])
                        )
                        he_xiang_sample = sample_rows[0]["他項權利"] if sample_rows else ""
                        total_area = sum(
                            (r["土地面積㎡"] or 0)
                            for r in sample_rows
                            if r.get("持分分子", 1) == r.get("持分分母", 1)  # 獨有地
                        )
                        grp_summary.append({
                            "歸戶群組": gid,
                            "地號清單": "、".join(parcels_in_grp),
                            "所有權人組合": owners_in_grp,
                            "他項權利": he_xiang_sample,
                            "宗地筆數": len(parcels_in_grp),
                        })
                    if grp_summary:
                        st.dataframe(pd.DataFrame(grp_summary), use_container_width=True, hide_index=True)

                    # ── 應分配面積計算 ─────────────────────────────────────
                    st.markdown("---")
                    with st.expander("📐 歸戶後應分配面積計算", expanded=False):
                        st.markdown("**全局參數**（可由其他分析頁面帶入，或手動輸入）")
                        _pre_p  = float(st.session_state.get('pre_land_price_sqm',  0) or 0)
                        _post_p = float(st.session_state.get('weighted_price_sqm', 0) or 0)
                        _auto_A = round(_post_p / _pre_p, 6) if _pre_p > 0 else 1.0
                        _ca1, _ca2, _ca3 = st.columns(3)
                        _A_val = _ca1.number_input(
                            "A：地價比（重劃後地價／重劃前地價）",
                            value=float(_auto_A), min_value=0.001, format="%.6f", key="alloc_A")
                        _B_val = _ca2.number_input(
                            "B：公共設施用地一般負擔系數",
                            value=0.0, min_value=0.0, max_value=1.0, format="%.6f", key="alloc_B")
                        _C_val = _ca3.number_input(
                            "C：費用負擔係數",
                            value=0.0, min_value=0.0, max_value=1.0, format="%.6f", key="alloc_C")
                        if _pre_p > 0 or _post_p > 0:
                            st.caption(
                                f"▸ 由 預期開發分析法 帶入重劃前地價：{_pre_p:,.0f} 元/㎡　"
                                f"由 土地開發分析法 帶入重劃後地價：{_post_p:,.0f} 元/㎡　→　自動計算 A = {_auto_A:.6f}"
                            )

                        # 收集已填入的街廓代碼
                        _all_codes: set = set()
                        for _cv in edited_df7["坐落街廓代碼"].dropna():
                            for _ck in [_x.strip() for _x in str(_cv).split(",") if _x.strip()]:
                                _all_codes.add(_ck)

                        if not _all_codes:
                            st.info("請先在上方表格「坐落街廓代碼」欄位填入各地號所坐落的街廓代碼，再展開此區塊計算。")
                        else:
                            st.markdown("---")
                            st.markdown(
                                "**各街廓道路參數**　"
                                "（僅重劃工程**新闢道路**需計算鄰街地負擔，非新闢道路請勾選取消並將 RW/S 留 0）"
                            )
                            # 建立街廓參數表初始值
                            _blk_init = []
                            for _code in sorted(_all_codes):
                                _blk_init.append({
                                    "街廓代碼":       _code,
                                    "前臨路寬RW(m)": 0.0,
                                    "側面路寬S(m)":  0.0,
                                    "深度比F":        0.0,
                                    "是否新闢道路":   False,
                                })
                            _blk_df = pd.DataFrame(_blk_init)
                            _edited_blk = st.data_editor(
                                _blk_df,
                                use_container_width=True,
                                num_rows="fixed",
                                column_config={
                                    "街廓代碼":       st.column_config.TextColumn("街廓代碼", disabled=True),
                                    "前臨路寬RW(m)": st.column_config.NumberColumn(format="%.2f"),
                                    "側面路寬S(m)":  st.column_config.NumberColumn(format="%.2f"),
                                    "深度比F":        st.column_config.NumberColumn(
                                        format="%.4f",
                                        help="深度比（鄰街地負擔深度係數），依重劃作業手冊計算後填入"),
                                    "是否新闢道路":   st.column_config.CheckboxColumn("是否新闢道路"),
                                },
                                key="editor_block_params",
                            )
                            _bp: dict = {}
                            for _, _br in _edited_blk.iterrows():
                                _c2  = _br["街廓代碼"]
                                _new = bool(_br.get("是否新闢道路", False))
                                _bp[_c2] = {
                                    "RW": float(_br["前臨路寬RW(m)"]) if _new else 0.0,
                                    "S":  float(_br["側面路寬S(m)"])  if _new else 0.0,
                                    "F":  float(_br["深度比F"]),
                                    "is_new": _new,
                                }

                            st.markdown("---")
                            st.markdown(
                                "**各歸戶群組暨街廓分配計算**　"
                                "（L1：前臨新闢路長度（m）；L2：側面新闢路長度（m）；非新闢道路填 0）"
                            )
                            # 計算各群組各街廓的重劃前面積 a
                            _grp_owner_map: dict = {}
                            _alloc_init_rows = []
                            for _gid in sorted(edited_df7["歸戶群組"].unique()):
                                _gdf = edited_df7[edited_df7["歸戶群組"] == _gid]
                                _owners = "、".join(
                                    dict.fromkeys(str(_x) for _x in _gdf["所有權人"].dropna() if str(_x).strip()))
                                _grp_owner_map[_gid] = _owners
                                _blk_area: dict = {}
                                for _, _gr in _gdf.iterrows():
                                    _cs_str = str(_gr.get("坐落街廓代碼", "") or "")
                                    _cs = [_x.strip() for _x in _cs_str.split(",") if _x.strip()]
                                    _ar = float(_gr.get("持分面積", 0) or 0)
                                    if _cs:
                                        _per = _ar / len(_cs)
                                        for _c3 in _cs:
                                            _blk_area[_c3] = _blk_area.get(_c3, 0) + _per
                                for _c3, _a3 in sorted(_blk_area.items()):
                                    _alloc_init_rows.append({
                                        "歸戶群組":       _gid,
                                        "所有權人":       _owners,
                                        "坐落街廓":       _c3,
                                        "重劃前面積a(㎡)": round(_a3, 4),
                                        "L1前臨路長(m)":  0.0,
                                        "L2側面路長(m)":  0.0,
                                    })

                            _alloc_init_df = pd.DataFrame(_alloc_init_rows) if _alloc_init_rows else pd.DataFrame(
                                columns=["歸戶群組","所有權人","坐落街廓","重劃前面積a(㎡)","L1前臨路長(m)","L2側面路長(m)"])
                            _edited_alloc = st.data_editor(
                                _alloc_init_df,
                                use_container_width=True,
                                num_rows="fixed",
                                column_config={
                                    "歸戶群組":       st.column_config.TextColumn(disabled=True),
                                    "所有權人":       st.column_config.TextColumn(disabled=True),
                                    "坐落街廓":       st.column_config.TextColumn(disabled=True),
                                    "重劃前面積a(㎡)": st.column_config.NumberColumn(format="%.4f", disabled=True),
                                    "L1前臨路長(m)":  st.column_config.NumberColumn(format="%.2f"),
                                    "L2側面路長(m)":  st.column_config.NumberColumn(format="%.2f"),
                                },
                                key="editor_alloc_table",
                            )

                            # 計算 G
                            _res_alloc_rows = []
                            for _, _ar2 in _edited_alloc.iterrows():
                                _gid2   = _ar2["歸戶群組"]
                                _code4  = _ar2["坐落街廓"]
                                _a4     = float(_ar2["重劃前面積a(㎡)"] or 0)
                                _L1     = float(_ar2["L1前臨路長(m)"]  or 0)
                                _L2     = float(_ar2["L2側面路長(m)"]  or 0)
                                _p4     = _bp.get(_code4, {"RW":0,"S":0,"F":0,"is_new":False})
                                _RW4, _S4, _F4 = _p4["RW"], _p4["S"], _p4["F"]
                                _knbr_f = _RW4 * _F4 * _L1
                                _knbr_s = _S4  * _L2
                                _G      = (_a4 * (1 - _A_val * _B_val) - _knbr_f - _knbr_s) * (1 - _C_val)
                                _ratio  = _G / _a4 if _a4 > 0 else 0.0
                                _res_alloc_rows.append({
                                    "歸戶群組":         _gid2,
                                    "所有權人":         _grp_owner_map.get(_gid2, ""),
                                    "坐落街廓":         _code4,
                                    "重劃前面積a(㎡)":  round(_a4,  4),
                                    "A（地價比）":      round(_A_val, 6),
                                    "B（公設負擔）":    round(_B_val, 6),
                                    "C（費用負擔）":    round(_C_val, 6),
                                    "RW前臨路寬(m)":    round(_RW4, 2),
                                    "深度比F":          round(_F4,  4),
                                    "L1前臨路長(m)":    round(_L1,  2),
                                    "S側面路寬(m)":     round(_S4,  2),
                                    "L2側面路長(m)":    round(_L2,  2),
                                    "鄰街地負擔(㎡)":   round(_knbr_f + _knbr_s, 4),
                                    "應分配面積G(㎡)":  round(_G,    4),
                                    "分配比G/a":        round(_ratio, 6),
                                })

                            if _res_alloc_rows:
                                st.markdown("---")
                                st.markdown("**計算結果：歸戶後應分配面積（公式：G = [a(1−A×B) − RW×F×L1 − S×L2](1−C)）**")
                                _res_alloc_df = pd.DataFrame(_res_alloc_rows)
                                # 標示 G < 0 警告
                                _neg_rows = _res_alloc_df[_res_alloc_df["應分配面積G(㎡)"] < 0]
                                if not _neg_rows.empty:
                                    st.warning(
                                        f"⚠️ 以下 {len(_neg_rows)} 筆計算結果 G < 0（面積不足分配），請確認參數是否正確：\n"
                                        + "、".join(_neg_rows["歸戶群組"].tolist()))
                                st.dataframe(_res_alloc_df, use_container_width=True, hide_index=True)
                                _tot_b = sum(r["重劃前面積a(㎡)"]  for r in _res_alloc_rows)
                                _tot_a = sum(r["應分配面積G(㎡)"] for r in _res_alloc_rows)
                                _mc1, _mc2, _mc3 = st.columns(3)
                                _mc1.metric("重劃前合計面積",   f"{_tot_b:,.4f} ㎡")
                                _mc2.metric("重劃後應分配合計", f"{_tot_a:,.4f} ㎡")
                                _mc3.metric("整體分配比 G/a",
                                    f"{_tot_a/_tot_b:.4%}" if _tot_b > 0 else "N/A")
                                # 儲存結果供 Excel 匯出
                                st.session_state["alloc_results_df"] = _res_alloc_df

                    # ── 產生 Excel 下載 ───────────────────────────────────
                    st.markdown("---")
                    if st.button("📋 產生下載檔案", key="btn_tab7_export"):
                        try:
                            from openpyxl import Workbook as _Wb
                            from openpyxl.styles import Font as _Font, PatternFill as _Fill, Alignment as _Align, Border as _Border, Side as _Side

                            wb7 = _Wb()
                            ws7 = wb7.active
                            ws7.title = "重劃區地籍對照表"

                            header_fill  = _Fill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                            grp_fill     = _Fill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                            header_font  = _Font(bold=True, color="FFFFFF", size=10)
                            thin   = _Side(style="thin")
                            border = _Border(left=thin, right=thin, top=thin, bottom=thin)
                            center = _Align(horizontal="center", vertical="center", wrap_text=True)
                            left_a = _Align(horizontal="left",   vertical="center", wrap_text=True)

                            headers7 = list(edited_df7.columns)
                            for ci, hdr in enumerate(headers7, 1):
                                c = ws7.cell(row=1, column=ci, value=hdr)
                                c.font = header_font; c.fill = header_fill
                                c.border = border;    c.alignment = center

                            for ri, data_row in enumerate(edited_df7.itertuples(index=False), 2):
                                for ci, val in enumerate(data_row, 1):
                                    c = ws7.cell(row=ri, column=ci, value=val if val != "" else None)
                                    c.border = border; c.alignment = left_a
                                    # 歸戶群組欄（第1欄）底色標示
                                    if ci == 1 and val:
                                        c.fill = grp_fill
                                        c.font = _Font(bold=True)

                            col_widths7 = {
                                "A": 10, "B": 8,  "C": 10, "D": 10, "E": 14,
                                "F": 18, "G": 14, "H": 24, "I": 10, "J": 8,
                                "K": 8,  "L": 10, "M": 14, "N": 14, "O": 36,
                                "P": 14, "Q": 12, "R": 18
                            }
                            for cl, w in col_widths7.items():
                                ws7.column_dimensions[cl].width = w
                            ws7.row_dimensions[1].height = 30

                            # 工作表2：歸戶群組摘要
                            if grp_summary:
                                ws7c = wb7.create_sheet("歸戶群組摘要")
                                sum_hdrs = list(grp_summary[0].keys())
                                for ci, h in enumerate(sum_hdrs, 1):
                                    c = ws7c.cell(row=1, column=ci, value=h)
                                    c.font = header_font; c.fill = header_fill
                                    c.border = border;    c.alignment = center
                                for ri, sr in enumerate(grp_summary, 2):
                                    for ci, h in enumerate(sum_hdrs, 1):
                                        c = ws7c.cell(row=ri, column=ci, value=sr.get(h, ""))
                                        c.border = border; c.alignment = left_a
                                for cl in ["A", "B", "C", "D", "E"]:
                                    ws7c.column_dimensions[cl].width = 20

                            # 工作表3：對應失敗地號
                            if rows_fail:
                                ws7b = wb7.create_sheet("對應失敗地號")
                                fail_hdrs = ["鄉鎮市", "地段", "地號", "原因"]
                                for ci, fh in enumerate(fail_hdrs, 1):
                                    c = ws7b.cell(row=1, column=ci, value=fh)
                                    c.font = _Font(bold=True)
                                    c.fill = _Fill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                                    c.border = border
                                for ri, fr in enumerate(rows_fail, 2):
                                    for ci, fh in enumerate(fail_hdrs, 1):
                                        ws7b.cell(row=ri, column=ci, value=fr.get(fh, "")).border = border
                                for cl in ["A", "B", "C", "D"]:
                                    ws7b.column_dimensions[cl].width = 20

                            # 工作表4：應分配面積計算表（若已計算）
                            _exp_alloc = st.session_state.get("alloc_results_df")
                            if _exp_alloc is not None and not _exp_alloc.empty:
                                ws7d = wb7.create_sheet("歸戶後應分配面積計算表")
                                alloc_fill = _Fill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                                alloc_hdrs = list(_exp_alloc.columns)
                                for ci, ah in enumerate(alloc_hdrs, 1):
                                    c = ws7d.cell(row=1, column=ci, value=ah)
                                    c.font = header_font; c.fill = header_fill
                                    c.border = border;    c.alignment = center
                                for ri, arow in enumerate(_exp_alloc.itertuples(index=False), 2):
                                    for ci, val in enumerate(arow, 1):
                                        c = ws7d.cell(row=ri, column=ci,
                                                      value=None if (isinstance(val, float) and pd.isna(val)) else val)
                                        c.border = border; c.alignment = left_a
                                        if ci == 1:  # 歸戶群組欄底色
                                            c.fill = alloc_fill; c.font = _Font(bold=True)
                                alloc_col_w = {"A":12,"B":20,"C":10,"D":14,"E":12,"F":12,"G":12,
                                               "H":12,"I":10,"J":10,"K":10,"L":10,"M":14,"N":14,"O":12}
                                for cl, w in alloc_col_w.items():
                                    ws7d.column_dimensions[cl].width = w
                                ws7d.row_dimensions[1].height = 30

                            buf7 = _io.BytesIO()
                            wb7.save(buf7)
                            buf7.seek(0)

                            st.download_button(
                                label="📥 下載重劃區地籍對照表.xlsx",
                                data=buf7.getvalue(),
                                file_name="重劃區地籍對照表.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                            st.success("✅ 檔案已產生，請點擊上方按鈕下載")
                        except Exception as e7_export:
                            st.error(f"匯出時發生錯誤：{str(e7_export)}")

            except Exception as e7:
                st.error(f"處理地籍資料時發生錯誤：{str(e7)}")
                st.exception(e7)

    # ===== Tab: 街廓互動分析 =====
    elif selected_tab == TAB_BLOCK_INTERACT:
        st.markdown("### 功能：街廓分類 + 重劃前地籍套疊 + 臨街地特別負擔 + B 值計算")
        st.caption("📌 依序完成 A→B→C→D→E→F→H→G→I，可得 B 值、臨街地特別負擔與各暫編地號 G 值分配")
        st.caption("📌 C 值須於「💰 財務分析」輸入工程費用等後自動回傳本頁")

        # 採用模組層級的常數：F3_BLOCK_CATEGORIES、F3_CATEGORY_BURDEN、F3_CATEGORY_COLORS

        # ---------- 步驟 A：上傳分配街廓 DXF ----------
        st.markdown("---")
        st.markdown("#### 📂 步驟 A：上傳分配街廓 DXF")
        up_block = st.file_uploader(
            "選擇分配街廓 DXF 檔（例：test.dxf）",
            type=["dxf"], key="f3_block_upl"
        )
        if up_block:
            st.session_state['f3_block_saved'] = (up_block.name, up_block.getvalue())
        elif st.session_state.get('f3_block_saved'):
            n8, d8 = st.session_state['f3_block_saved']
            up_block = _NamedBytesIO(n8, d8)
            st.caption(f"📂 已從快取載入：{n8}")

        if up_block is None:
            st.info("請先上傳分配街廓 DXF 檔（內含規劃完成的街廓 POLYLINE，建議每個街廓標註 R1、R2... 以方便辨識）")
            st.stop()

        # 解析 DXF
        try:
            with st.spinner("解析 DXF 中…"):
                block_data = parse_block_dxf(up_block.getvalue() if hasattr(up_block, 'getvalue') else up_block.read())
        except Exception as e8a:
            st.error(f"DXF 解析錯誤：{str(e8a)}")
            st.exception(e8a)
            st.stop()

        polys = block_data['polygons']
        if not polys:
            st.error("❌ DXF 內未找到任何閉合 POLYLINE。請確認 DXF 是否為規劃完成的分配街廓圖。")
            st.stop()

        non_outer = [p for p in polys if not p['is_outer']]
        # outer_boundary 現為座標清單（由葉節點街廓聯集而成），非單一多邊形
        outer_boundary_pts = block_data.get('outer_boundary') or []

        st.success(f"✅ 解析完成：共 {len(polys)} 個多邊形（外框 1 + 街廓 {len(non_outer)}）")
        with st.expander("🔎 DXF 圖層資訊", expanded=False):
            st.write("**DXF 圖層：**", block_data.get('layer_names', []))
            st.write(f"**重劃範圍（外框）面積：** {block_data.get('outer_area_m2', 0):,.2f} ㎡ ≈ {block_data.get('outer_area_m2', 0)*0.3025:,.2f} 坪")

        # ---------- Session state 預設值 ----------
        SS_CAT = 'f3_block_cat'
        SS_LBL = 'f3_block_label'
        SS_ROAD = 'f3_block_road'
        if SS_CAT not in st.session_state or len(st.session_state.get(SS_CAT, {})) != len(non_outer):
            st.session_state[SS_CAT] = {}
            st.session_state[SS_LBL] = {}
            st.session_state[SS_ROAD] = {}
            for p in non_outer:
                bid = p['id']
                # 🆕 Task AB（Phase 11）：優先採用 DXF 文字解析出之 category
                #     若無 → 依 R 字頭 fallback 至「住宅區」；其他 → 「未分類」
                _raw_cat = p.get('category_raw', '') or ''
                _norm_cat = _normalize_block_category(_raw_cat) if _raw_cat else ''
                if _norm_cat:
                    default_cat = _norm_cat
                elif (p['label'] or '').upper().startswith('R'):
                    default_cat = "住宅區"
                else:
                    default_cat = "未分類"
                st.session_state[SS_CAT][bid] = default_cat
                st.session_state[SS_LBL][bid] = p['label'] or f"區塊#{bid}"
                st.session_state[SS_ROAD][bid] = {
                    'front_width': 8.0, 'front_length': 0.0, 'front_new': True,
                    'side_width': 8.0, 'side_length': 0.0, 'side_new': False,
                }

        # ---------- 街廓分類 + 可編輯街廓編號表單 ----------
        st.markdown("#### 🏷️ 街廓分類（15 項使用分區；街廓編號可手動修改）")
        with st.form("f3_classify_form", clear_on_submit=False):
            col_h = st.columns([1.2, 1.5, 1.2, 1.2, 2])
            col_h[0].markdown("**DXF 標籤**")
            col_h[1].markdown("**街廓編號（可改）**")
            col_h[2].markdown("**面積(㎡)**")
            col_h[3].markdown("**面積(坪)**")
            col_h[4].markdown("**分類**")
            new_cats, new_lbls = {}, {}
            for p in non_outer:
                bid = p['id']
                orig_lbl = p['label'] or f"#{bid}"
                c1, c2, c3, c4, c5 = st.columns([1.2, 1.5, 1.2, 1.2, 2])
                c1.write(f"`{orig_lbl}`")
                new_lbls[bid] = c2.text_input(
                    f"編號_{bid}", value=st.session_state[SS_LBL].get(bid, orig_lbl),
                    key=f"f3_lbl_{bid}", label_visibility="collapsed"
                )
                c3.write(f"{p['area_m2']:,.2f}")
                c4.write(f"{p['area_m2']*0.3025:,.2f}")
                cur_cat = st.session_state[SS_CAT].get(bid, "未分類")
                idx_c = F3_BLOCK_CATEGORIES.index(cur_cat) if cur_cat in F3_BLOCK_CATEGORIES else 0
                new_cats[bid] = c5.selectbox(
                    f"分類_{bid}", F3_BLOCK_CATEGORIES, index=idx_c,
                    key=f"f3_cat_{bid}", label_visibility="collapsed"
                )
            submit_cat = st.form_submit_button("✅ 套用街廓編號與分類", use_container_width=True)
        if submit_cat:
            st.session_state[SS_CAT] = new_cats
            st.session_state[SS_LBL] = new_lbls
            st.success("街廓編號與分類已更新")
            st.rerun()

        # 整併 classified_blocks
        classified_blocks = []
        for p in non_outer:
            bid = p['id']
            lbl = st.session_state[SS_LBL].get(bid, p['label'] or f"#{bid}")
            classified_blocks.append({
                **p,
                'label': lbl,
                'category': st.session_state[SS_CAT].get(bid, "未分類"),
                'burden_type': F3_CATEGORY_BURDEN.get(st.session_state[SS_CAT].get(bid, "未分類"), "未分類"),
                # 🆕 W-A.3 §4：街角標記預設值（CAD SIDE_LINE 讀入後再覆寫，見步驟C後）
                'has_corner': False,
                'corner_sides': [],
            })
        st.session_state['f3_classified_blocks'] = classified_blocks

        # ---------- 分類面積匯總 ----------
        st.markdown("#### 📊 分類面積匯總（依負擔屬性）")
        summary = {c: {'筆數': 0, '面積(㎡)': 0.0} for c in F3_BLOCK_CATEGORIES}
        for b in classified_blocks:
            summary[b['category']]['筆數'] += 1
            summary[b['category']]['面積(㎡)'] += b['area_m2']
        buildable_total = 0.0
        public_common_total = 0.0
        public_non_common_total = 0.0
        sum_rows = []
        for c in F3_BLOCK_CATEGORIES:
            s = summary[c]
            if s['筆數'] == 0:
                continue
            bt = F3_CATEGORY_BURDEN.get(c, "未分類")
            sum_rows.append({
                '分類': c, '負擔屬性': bt, '筆數': s['筆數'],
                '面積(㎡)': round(s['面積(㎡)'], 2),
                '面積(坪)': round(s['面積(㎡)'] * 0.3025, 2),
            })
            if bt == "可建築土地":
                buildable_total += s['面積(㎡)']
            elif bt == "共同負擔":
                public_common_total += s['面積(㎡)']
            elif bt == "非共同負擔":
                public_non_common_total += s['面積(㎡)']
        if sum_rows:
            st.dataframe(pd.DataFrame(sum_rows), use_container_width=True, hide_index=True)
        mc1, mc2, mc3 = st.columns(3)
        mc1.metric("可建築土地合計", f"{buildable_total:,.2f} ㎡", f"{buildable_total*0.3025:,.2f} 坪")
        mc2.metric("公設共同負擔合計", f"{public_common_total:,.2f} ㎡", f"{public_common_total*0.3025:,.2f} 坪")
        mc3.metric("公設非共同負擔合計", f"{public_non_common_total:,.2f} ㎡", f"{public_non_common_total*0.3025:,.2f} 坪")

        # ---------- 街廓配色圖 ----------
        st.markdown("#### 🗺️ 街廓配色圖")
        st.caption("🟥 紅色線段 = 街廓分配線（街廓邊界），與後續重劃前地籍 DXF 套疊時仍可清楚區分。")
        import plotly.graph_objects as go
        fig_a = go.Figure()
        if outer_boundary_pts:
            fig_a.add_trace(go.Scatter(
                x=[v[0] for v in outer_boundary_pts], y=[v[1] for v in outer_boundary_pts],
                mode='lines', line=dict(color='#B30000', width=3, dash='dash'),
                name='重劃範圍外框', hoverinfo='skip'
            ))
        # 填色區塊（單一合併 trace 節省渲染時間）
        for b in classified_blocks:
            color = F3_CATEGORY_COLORS.get(b['category'], "#BDBDBD")
            xs = [v[0] for v in b['vertices']] + [b['vertices'][0][0]]
            ys = [v[1] for v in b['vertices']] + [b['vertices'][0][1]]
            fig_a.add_trace(go.Scatter(
                x=xs, y=ys, mode='lines', fill='toself', fillcolor=color,
                line=dict(color=color, width=0.5), opacity=0.45,
                name=f"{b['label']} ({b['category']})",
                hovertemplate=f"<b>{b['label']}</b><br>{b['category']}<br>{b['area_m2']:,.2f} ㎡<extra></extra>",
            ))
            fig_a.add_annotation(
                x=b['centroid'][0], y=b['centroid'][1], text=f"<b>{b['label']}</b>",
                showarrow=False, font=dict(size=11, color="white"),
                bgcolor="rgba(0,0,0,0.45)", borderpad=2,
            )
        # 🟥 街廓分配線（紅色邊界）— 合併為一條 trace 提升效能
        red_xs, red_ys = [], []
        for b in classified_blocks:
            for i, (x, y) in enumerate(b['vertices']):
                red_xs.append(x); red_ys.append(y)
            # 封閉每個街廓
            if b['vertices']:
                red_xs.append(b['vertices'][0][0]); red_ys.append(b['vertices'][0][1])
            red_xs.append(None); red_ys.append(None)  # 段間斷開
        fig_a.add_trace(go.Scatter(
            x=red_xs, y=red_ys, mode='lines',
            line=dict(color='#D62728', width=2),
            name='街廓分配線', hoverinfo='skip',
        ))
        fig_a.update_layout(
            height=560, showlegend=True,
            paper_bgcolor='white', plot_bgcolor='white',
            xaxis=dict(title="TWD97 X (m)", gridcolor='#E5E5E5', zerolinecolor='#CCCCCC',
                       showline=True, linecolor='#888888'),
            yaxis=dict(title="TWD97 Y (m)", gridcolor='#E5E5E5', zerolinecolor='#CCCCCC',
                       showline=True, linecolor='#888888', scaleanchor="x", scaleratio=1),
            legend=dict(orientation="v", yanchor="top", y=1, xanchor="left", x=1.02,
                        bgcolor='rgba(255,255,255,0.9)', bordercolor='#CCCCCC', borderwidth=1),
            margin=dict(l=40, r=200, t=20, b=40),
        )
        st.plotly_chart(fig_a, use_container_width=True)

        # ---------- 步驟 B：地籍多邊形（🆕 Phase 11 Task AA：重用 Step A 之 DXF）----------
        # 使用者已將 BLOCK / CADASTRAL_BOUND / CADASTRAL_TEXT / BASELINE 等所有圖層整合於同一個 DXF。
        # 不再要求第二個檔案上傳；直接以 Step A 上傳之 DXF bytes 餵給 parse_cadastral_geofile。
        st.markdown("---")
        st.markdown("#### 🧬 步驟 B：自動解析重劃前地籍（沿用 Step A 之 DXF）")
        st.caption(
            "🚀 **One-File Workflow**：系統自動讀取 Step A 已上傳之 DXF 中的地籍圖層"
            "（ROAD 界址線 + OPAR 地號），無需再上傳第二份檔案。"
        )
        # 重用 Step A 之 up_block bytes 與檔名
        _block_bytes = up_block.getvalue() if hasattr(up_block, 'getvalue') else None
        _block_name = getattr(up_block, 'name', 'block.dxf')
        if _block_bytes is None:
            up_cad = None
        else:
            up_cad = _NamedBytesIO(_block_name, _block_bytes)

        parcel_polys = []
        if up_cad:
            try:
                with st.spinner("解析地籍圖層中…（沿用 Step A DXF）"):
                    _bytes = _block_bytes
                    cad_data = parse_cadastral_geofile(_bytes, _block_name)
            except Exception as eC:
                st.error(f"地籍解析錯誤：{str(eC)}")
                st.exception(eC)
                cad_data = None
            if cad_data:
                parcel_polys = cad_data.get('parcel_polygons', []) or []

                # 🆕 Phase 11 Hotfix：以 Tab 1 登記面積交叉驗證 + 自動修正換位錯配
                _t1_areas = st.session_state.get('t8_parcel_areas', {}) or {}
                _swap_log = []
                _suspects = []
                if _t1_areas:
                    try:
                        _validate_res = validate_parcel_assignments_by_area(
                            parcel_polys, _t1_areas,
                            area_diff_threshold=0.30,
                            max_swap_distance_m=30.0,
                        )
                        parcel_polys = _validate_res['parcel_polys']
                        _swap_log = _validate_res['swaps']
                        _suspects = _validate_res['suspects_remaining']
                        # 把修正後 parcel_polys 寫回 cad_data，供下游使用
                        cad_data['parcel_polygons'] = parcel_polys
                    except Exception as _ev:
                        st.warning(f"⚠️ 面積交叉驗證失敗（不影響原配對）：{str(_ev)}")

                _msg = (f"✅ 地籍讀取完成：{len(cad_data.get('parcels', [])):,} 筆地號；"
                        f"重建 {len(parcel_polys):,} 筆重劃前宗地多邊形")
                if _swap_log:
                    _msg += f"｜🔄 土地歸戶 面積交叉驗證自動修正 **{len(_swap_log)} 對**錯配"
                st.success(_msg)
                # 列出修正詳情
                if _swap_log:
                    with st.expander(f"🔄 土地歸戶 面積交叉驗證：自動換位修正 {len(_swap_log)} 對", expanded=True):
                        for sw in _swap_log:
                            st.write(
                                f"・**{sw['parcel_a_orig']}** ↔ **{sw['parcel_b_orig']}** "
                                f"（中心距離 {sw['centroid_dist_m']:.1f} m；"
                                f"偏差和 {sw['err_before']:.2%} → {sw['err_after']:.2%}）"
                            )
                if _suspects:
                    with st.expander(f"⚠️ 仍有 {len(_suspects)} 筆面積偏差 > 30%（無相鄰可換位）", expanded=False):
                        for i in _suspects:
                            p = parcel_polys[i]
                            pn = p.get('parcel', '')
                            actual = float(p.get('area_m2', 0) or 0)
                            reg = float(_t1_areas.get(pn, 0) or 0)
                            st.write(
                                f"・{pn}：實測 {actual:.2f} ㎡ vs 登記 {reg:.2f} ㎡"
                                + (f"（偏差 {abs(actual-reg)/reg:.1%}）" if reg > 0 else "")
                            )
                # DXF 解析診斷資訊（深度重構 Task 1）
                _filtered = st.session_state.get('f3_dxf_filtered_opar', [])
                _tiers = st.session_state.get('f3_dxf_match_tiers', {})
                if _filtered or _tiers:
                    with st.expander("🔍 DXF 解析診斷", expanded=False):
                        if _filtered:
                            st.info(
                                f"⚠️ 已過濾 {len(_filtered)} 筆非地號 OPAR 文字：{', '.join(_filtered[:20])}"
                                + (f"…等共 {len(_filtered)} 筆" if len(_filtered) > 20 else "")
                            )
                        if _tiers:
                            _t1 = sum(1 for v in _tiers.values() if v == 1)
                            _t2 = sum(1 for v in _tiers.values() if v == 2)
                            _t3 = sum(1 for v in _tiers.values() if v == 3)
                            st.markdown(
                                f"地號配對統計：**Tier 1**（嚴格包含）{_t1} 筆　"
                                f"**Tier 2**（最近鄰 5m）{_t2} 筆　"
                                f"**Tier 3**（1m 緩衝）{_t3} 筆"
                            )

        # ---------- 步驟 C：暫編地號（多邊形 intersection） ----------
        temp_parcels = []
        if parcel_polys:
            st.markdown("---")
            st.markdown("#### 🧮 步驟 C：自動計算暫編地號（街廓 × 重劃前宗地 intersection）")
            try:
                with st.spinner("套疊計算中…"):
                    temp_parcels = overlay_polygons_to_blocks(parcel_polys, classified_blocks)
                    # 🚨 W-A 規格一：四欄面積後處理（補登記/分攤登記、重設 面積_m2=0）
                    temp_parcels = _assign_four_column_areas(
                        temp_parcels,
                        st.session_state.get('t8_parcel_areas', {}) or {},
                    )
                    # 🆕 W-A.3 §2：標記暫編切割性質 cut_type（逕割/分配/非分割/整筆）
                    temp_parcels = _annotate_temp_parcel_cut_type(temp_parcels)
                st.session_state['f3_temp_parcels'] = temp_parcels
                # W-A.2 §1：不再對「有原地號歸屬」之交集塊做形狀剔除（_sliver_n 應為 0）
                _attrib_n = int(st.session_state.get('f3_wa2_attrib_repaired_count', 0) or 0)
                _unc_n = int(st.session_state.get('f3_unassigned_added_count', 0) or 0)
                _ghost_n = int(st.session_state.get('f3_unassigned_ghost_count', 0) or 0)
                _msg_extra = []
                if _attrib_n > 0:
                    _msg_extra.append(f"🔧 §2 歸屬修復 **{_attrib_n}** 塊（救回真實土地）")
                if _unc_n > 0:
                    _msg_extra.append(f"🟧 _UNC 真．無主殘料 **{_unc_n}** 筆")
                if _ghost_n > 0:
                    _msg_extra.append(f"👻 ghost 邊界/小碎塊 **{_ghost_n}** 筆")
                st.success(
                    f"✅ 產生 {len(temp_parcels)} 筆暫編地號"
                    + ("（" + "、".join(_msg_extra) + "）" if _msg_extra else "")
                )
                # 🆕 W-D.3（純加性診斷 warn，零行為變更）：ghost 零面積不變量。
                #   [H-ghost] 已決 H-a 之等價全靠 ghost 零面積（Step G 不排 ghost）；
                #   若任一 ghost 面積≠0 → 具名警示（不改分配控制流；harness 側為 RuntimeError）。
                _ghost_nonzero = [tp for tp in temp_parcels
                                  if tp.get('_is_ghost_sliver')
                                  and (float(tp.get('幾何面積_m2', 0) or 0) != 0.0
                                       or float(tp.get('面積_m2', 0) or 0) != 0.0)]
                if _ghost_nonzero:
                    st.warning(
                        f"⚠️ ghost 零面積不變量破：**{len(_ghost_nonzero)}** 筆 ghost 面積≠0"
                        "（[H-ghost] H-a 等價根基失守，真面積應僅存 `_ghost_area_m2`）。"
                        "請查 overlay 之 ghost 建構（幾何面積_m2/面積_m2 應恆 0）。"
                    )

                # 🚨 W-A.2 §3：完整性警示清單（缺口率 > 3% 之地號）
                _integrity_warnings = (
                    st.session_state.get('f3_integrity_warnings', []) or []
                )
                if _integrity_warnings:
                    with st.expander(
                        f"⚠️ 切割不完整警示清單（{len(_integrity_warnings)} 筆地號缺口率 > 3%）",
                        expanded=False,
                    ):
                        st.caption(
                            "缺口率 = |Σ各暫編幾何面積 − 登記面積| / 登記面積。"
                            "本清單中之地號**分攤登記面積已退化 = 幾何面積**"
                            "（不縮放登記面積，避免將登記面積硬攤在不完整切割塊上）。"
                            "可能原因：跨重劃區邊界、CAD 拓樸缺角、圖簿差。"
                        )
                        st.dataframe(
                            pd.DataFrame(_integrity_warnings),
                            use_container_width=True, hide_index=True,
                        )

                # ═══════════════════════════════════════════════════════════
                # 🔬 W-A.2 §0 診斷：_UNC 匿名塊根因定位（給 KL 判定是否需 §4）
                # ───────────────────────────────────────────────────────────
                # 印出：
                # (1) 628-12 parcel_polys polygon 面積 vs Tab 1 登記面積
                #     → ≈200（完整）走 §1；≈55（不完整）需 §4 polygonize 邊界閉合。
                # (2) 所有 _UNC 各自面積、所屬街廓、與原地號 polygon 交集 top 3。
                # ═══════════════════════════════════════════════════════════
                with st.expander("🔬 W-A.2 §0 診斷：_UNC 匿名塊根因定位", expanded=True):
                    from shapely.geometry import Polygon as _SP_diag_wa2
                    # ── 1. 628-12 polygon 完整性 ──
                    st.markdown("**【1】628-12 parcel_polys polygon 面積（判 polygonize 完整性）**")
                    _t1_areas_diag = st.session_state.get('t8_parcel_areas', {}) or {}
                    _p628_list = [p for p in parcel_polys
                                  if str(p.get('parcel', '')).strip() == '628-12']
                    if not _p628_list:
                        st.warning("⚠️ parcel_polys 中找不到 628-12（檢查 DXF 解析）")
                    else:
                        _p628 = _p628_list[0]
                        _coords_628 = _p628.get('polygon_coords', []) or []
                        _reg_628 = float(_t1_areas_diag.get('628-12', 0) or 0)
                        _poly_area_628 = float(_p628.get('area_m2', 0) or 0)
                        if len(_coords_628) >= 3:
                            try:
                                _poly_628 = _SP_diag_wa2(_coords_628)
                                if not _poly_628.is_valid:
                                    _poly_628 = _poly_628.buffer(0)
                                _poly_area_628 = float(_poly_628.area)
                            except Exception as _e628:
                                st.write(f"  ⚠️ polygon 重算失敗：{_e628}")
                        st.write(f"- 多邊形面積（DXF 重建）≈ **{_poly_area_628:.2f} ㎡**")
                        st.write(f"- 土地歸戶 登記面積 = **{_reg_628:.2f} ㎡**")
                        if _reg_628 > 0:
                            _diff_628 = abs(_poly_area_628 - _reg_628)
                            _ratio_628 = _diff_628 / _reg_628 * 100
                            if _ratio_628 < 5.0:
                                st.success(
                                    f"✅ polygon 完整（差 {_diff_628:.2f} ㎡ / "
                                    f"{_ratio_628:.1f}%）→ G1 交集塊被 sliver / "
                                    f"0.1m 垂距誤刪，走 §1 即可救回。"
                                )
                            else:
                                st.error(
                                    f"⚠️ polygon 不完整（差 {_diff_628:.2f} ㎡ / "
                                    f"{_ratio_628:.1f}%）→ polygonize 在重劃區邊界未閉合，"
                                    f"除 §1 外另需 §4。**請 KL 確認 §4 方案再動工。**"
                                )

                    # ── 2. 所有 _UNC 細節 ──
                    st.markdown("---")
                    _uncs_diag = [tp for tp in temp_parcels
                                  if str(tp.get('原地號', '')).startswith('_UNC')]
                    st.markdown(f"**【2】_UNC 匿名塊明細（共 {len(_uncs_diag)} 筆）**")
                    if not _uncs_diag:
                        st.info("（本批暫編無 _UNC）")
                    for _u in _uncs_diag:
                        _u_no = _u.get('暫編地號', '')
                        _u_area = float(_u.get('幾何面積_m2',
                                                _u.get('面積_m2', 0)) or 0)
                        _u_blk = _u.get('所屬街廓', '')
                        _u_coords = _u.get('polygon_coords', []) or []
                        st.markdown(
                            f"---\n**{_u_no}**　面積 **{_u_area:.2f} ㎡**　"
                            f"街廓 `{_u_blk}`"
                        )
                        if len(_u_coords) < 3:
                            st.write("  （無 polygon 座標、無法比對）")
                            continue
                        try:
                            _u_poly = _SP_diag_wa2(_u_coords)
                            if not _u_poly.is_valid:
                                _u_poly = _u_poly.buffer(0)
                            if _u_poly.is_empty:
                                st.write("  （polygon 空）")
                                continue
                        except Exception as _eup:
                            st.write(f"  （polygon 解析失敗：{_eup}）")
                            continue

                        _hits = []
                        for _pp in parcel_polys:
                            _pno = str(_pp.get('parcel', '') or '').strip()
                            _pcoords = _pp.get('polygon_coords', []) or []
                            if not _pno or len(_pcoords) < 3:
                                continue
                            try:
                                _pp_poly = _SP_diag_wa2(_pcoords)
                                if not _pp_poly.is_valid:
                                    _pp_poly = _pp_poly.buffer(0)
                                _inter = _u_poly.intersection(_pp_poly)
                                if _inter.is_empty:
                                    continue
                                _ia = float(_inter.area)
                                if _ia >= 0.05:
                                    _hits.append(
                                        (_pno, _ia,
                                         (_ia / _u_area * 100) if _u_area > 0 else 0)
                                    )
                            except Exception:
                                continue
                        _hits.sort(key=lambda x: -x[1])
                        if not _hits:
                            st.write("  - 與所有原地號交集 < 0.05 ㎡ → 真．無主殘料")
                        else:
                            for _h in _hits[:3]:
                                st.write(
                                    f"  - **{_h[0]}**　交集 **{_h[1]:.2f} ㎡**　"
                                    f"占 _UNC {_h[2]:.1f}%"
                                )

                # 🆕 Phase 7 Module 1：CAD 精準圖層直讀（BASELINE / CENTERLINE / FRONT_LINE / SIDE_LINE）
                try:
                    if hasattr(up_cad, 'getvalue'):
                        _doc_cad_p7 = _read_dxf_any_encoding(up_cad.getvalue())
                        _cad_layers = parse_cad_precision_layers(_doc_cad_p7, classified_blocks)
                        _exist_mbl = dict(st.session_state.get('f3_manual_baseline', {}))
                        for _lbl, _bv in _cad_layers['baselines'].items():
                            _exist_mbl[_lbl] = _bv
                        st.session_state['f3_manual_baseline'] = _exist_mbl
                        _exist_mc = dict(st.session_state.get('f3_manual_road_centerlines', {}))
                        for _lbl, _pts in _cad_layers['centerlines'].items():
                            _exist_mc[_lbl] = _pts
                        st.session_state['f3_manual_road_centerlines'] = _exist_mc
                        st.session_state['f3_cad_front_lengths'] = _cad_layers['front_lengths']
                        st.session_state['f3_cad_side_lengths'] = _cad_layers['side_lengths']
                        # 🆕 Phase 11 v2：保存 FRONT_LINE 完整端點（供 G 迭代取得 d_hat）
                        st.session_state['f3_cad_front_lines'] = (
                            _cad_layers.get('front_lines', {}) or {}
                        )
                        # 🆕 Hotfix：左/右側長度分開儲存
                        st.session_state['f3_cad_side_lengths_by_side'] = (
                            _cad_layers.get('side_lengths_by_side', {}) or {}
                        )
                        # 🚨 W-B §1-2：SIDE_LINE 座標（含 p1/p2/mid）
                        st.session_state['f3_cad_side_lines_by_side'] = (
                            _cad_layers.get('side_lines_by_side', {}) or {}
                        )
                        # 🚨 W-B §1-3：宗地分配線方向
                        st.session_state['f3_cad_alloc_dir'] = (
                            _cad_layers.get('alloc_dir_by_block', {}) or {}
                        )
                        # 🚨 W-B §1-1b：未吻合 SIDE_LINE 警示
                        _sl_unmatched = (
                            _cad_layers.get('side_unmatched_warnings', []) or []
                        )
                        for _uw in _sl_unmatched:
                            _mid_uw = _uw.get('mid', (0, 0))
                            st.warning(
                                f"⚠️ SIDE_LINE（中點 ({_mid_uw[0]:.1f}, {_mid_uw[1]:.1f})）"
                                f"兩端均未接到任何 FRONT_LINE 端點"
                                f"（最近 {_uw.get('dmin', 0):.2f}m"
                                f" > {_uw.get('tol', 0.5):.1f}m）"
                                "——請檢查 DXF：SIDE_LINE 端點須接 FRONT_LINE 端點"
                            )
                        # 🆕 W-A.3 §4：依 SIDE_LINE 左右側標記街廓有無街角（has_corner/corner_sides）
                        classified_blocks = _annotate_block_corner_flags(
                            classified_blocks,
                            _cad_layers.get('side_lengths_by_side', {}) or {},
                        )
                        st.session_state['f3_classified_blocks'] = classified_blocks
                        # 🆕 W-D.1.3-a 切換：CAD-import 後單一 post-pass，以拓樸定錨重建各塊 theoretical_corners
                        #   （取代長度法、單一真相源）。就地 mutate geom_restore → 重存 session_state →
                        #   下游（corner-range 迴圈、PK cutoff）使用當下才從 session 讀，無 early-consumer 破口。
                        try:
                            _fl_map_pp = _cad_layers.get('front_lines', {}) or {}
                            _sl_map_pp = _cad_layers.get('side_lines_by_side', {}) or {}
                            for _blk_pp in classified_blocks:
                                _gr_pp = _blk_pp.get('geom_restore')
                                if not isinstance(_gr_pp, dict):
                                    continue
                                _lbl_pp = _blk_pp.get('label')
                                _edges_pp = (_gr_pp.get('classification') or {}).get('edges') or []
                                _new_tcs_pp = _rebuild_corners_topology(
                                    _edges_pp, _fl_map_pp.get(_lbl_pp), _sl_map_pp.get(_lbl_pp))
                                if _new_tcs_pp is None:
                                    # 無法定錨（缺 edges/FRONT）：街角街廓→停機警告（不靜默退長度法）；非街角→留空
                                    if bool(_blk_pp.get('corner_sides')):
                                        st.warning(
                                            f"⚠️ 街廓 {_lbl_pp}：缺必要圖層（FRONT/SIDE/ALLOC/BASELINE），"
                                            f"無法拓樸定錨截角 → 停機（請補圖層、不靜默退長度法）。"
                                        )
                                    _new_tcs_pp = []
                                elif bool(_blk_pp.get('corner_sides')) and not _new_tcs_pp:
                                    # reviewer W1：有街角(corner_sides)但截角拓樸配對失敗（SIDE 共線配對空）
                                    #   → _rebuild 回 [] 而非 None、會繞過上面停機。此處補警示、不靜默歸零。
                                    st.warning(
                                        f"⚠️ 街廓 {_lbl_pp}：有街角(corner_sides={_blk_pp.get('corner_sides')})"
                                        f"但截角拓樸配對失敗（FRONT/SIDE 共線配對空）→ 截角歸 0；"
                                        f"請檢查 FRONT/SIDE 圖層或容差（不靜默編造，CLAUDE.md）。"
                                    )
                                # 就地覆寫 theoretical_corners + 同步 cutoff_total_area（KL 點2：903 副本勿陳舊）
                                _gr_pp['theoretical_corners'] = _new_tcs_pp
                                _gr_pp['cutoff_total_area'] = round(
                                    sum(float(_t.get('cutoff_area_m2', 0) or 0) for _t in _new_tcs_pp), 4)
                                _blk_pp['cutoff_total_area_m2'] = _gr_pp['cutoff_total_area']
                            st.session_state['f3_classified_blocks'] = classified_blocks
                        except Exception as _e_pp:
                            st.warning(f"⚠️ W-D.1.3-a 截角拓樸定錨 post-pass 失敗：{_e_pp}")
                        st.session_state['f3_cad_diagnostics'] = _cad_layers['diagnostics']
                        # 🆕 Hotfix：條數改用 raw matched count（DXF 實際畫了幾條）
                        _n_b = _cad_layers.get('baselines_matched_count',
                                                 len(_cad_layers['baselines']))
                        _n_c = _cad_layers.get('centerlines_matched_count',
                                                 len(_cad_layers['centerlines']))
                        _n_f = _cad_layers.get('front_lines_matched_count',
                                                 len(_cad_layers['front_lengths']))
                        _n_s = _cad_layers.get('side_lines_matched_count',
                                                 len(_cad_layers['side_lengths']))
                        if (_n_b + _n_c + _n_f + _n_s) > 0:
                            st.success(
                                f"✅ CAD 精準圖層直讀：{_n_b} 條基準線、{_n_c} 條中心線、"
                                f"{_n_f} 條正面長度、{_n_s} 條側面長度"
                            )
                        elif _cad_layers['diagnostics']['layers_found']:
                            st.info(
                                f"ℹ️ 偵測到 CAD 精準圖層 {_cad_layers['diagnostics']['layers_found']} "
                                "但未綁定任何街廓（請檢查線段位置是否靠近街廓邊界 < 0.1m）"
                            )
                except Exception as _eC7:
                    pass


                # 🆕 Phase 6 Task O：DXF 解析後即時問題診斷
                try:
                    _own_map_diag = st.session_state.get('t8_ownership_map', {}) or {}
                    _problems_init = _detect_problem_parcels(temp_parcels, _own_map_diag)
                    _n_no_landno = len(_problems_init['no_landno'])
                    _n_no_owner = len(_problems_init['no_owner'])
                    if _n_no_landno > 0 or _n_no_owner > 0:
                        st.warning(
                            f"⚠️ 偵測到 **{_n_no_landno}** 筆無地號地塊 + "
                            f"**{_n_no_owner}** 筆未歸戶地塊。"
                            "請於下方「🗺️ 統一 GIS 互動分析地圖」勾選「⚠️ 問題地塊」圖層檢視，"
                            "並用「✏️ 補正地號」操作模式手動補正。"
                        )
                except Exception:
                    pass
            except Exception as eD:
                st.error(f"暫編地號計算失敗：{str(eD)}")

            # ─── Tab 1 歸戶交叉檢核（W-A.2 §5-1：純資訊性，不剔除有原地號歸屬之暫編）───
            _own_map_check = st.session_state.get('t8_ownership_map', {}) or {}

            # 地號正規化：Tab 1（「715」「628-37」）vs Tab 4（DXF OPAR：「715-0」「13-628-37」）
            #   1) DXF 段碼前綴 → 取後 2 段，例如「13-628-37」→「628-37」
            #   2) 子號為 0 後綴 → 去除，例如「715-0」→「715」
            def _normalize_landno(s):
                if not s:
                    return ''
                t = str(s).strip()
                if not t:
                    return ''
                parts = t.split('-')
                if len(parts) >= 3:
                    parts = parts[-2:]
                if len(parts) == 2 and parts[1] in ('0', '00', '000'):
                    parts = parts[:1]
                return '-'.join(parts)

            if temp_parcels and _own_map_check:
                _own_norm_set = set()
                _own_raw_set = set()
                for k in _own_map_check.keys():
                    _own_raw_set.add(str(k).strip())
                    _own_norm_set.add(_normalize_landno(k))

                _orphans = []
                _matched_via_norm = 0
                for tp in temp_parcels:
                    ono_raw = str(tp.get('原地號', '') or '').strip()
                    if not ono_raw:
                        continue
                    # W-A.2 §5：真．無主殘料（_UNC / _GHOST）不視為 orphan
                    if ono_raw.startswith('_UNC') or ono_raw.startswith('_GHOST'):
                        continue
                    ono_norm = _normalize_landno(ono_raw)
                    if ono_raw in _own_raw_set or ono_norm in _own_norm_set:
                        if ono_raw not in _own_raw_set and ono_norm in _own_norm_set:
                            _matched_via_norm += 1
                        continue
                    _orphans.append(tp)
                if _orphans:
                    # W-A.2 §5-1：移除「建議剔除」按鈕——
                    # KL 2026-06-15 §0 確認：所謂「未在 Tab 1 歸戶清單」的暫編
                    # 多為 _UNC 匿名標籤誤讀；§1 修好後實際 orphan 應極少（純粹
                    # Tab 1 漏填或地號編碼差異）。系統不再自動剔除有原地號歸屬之暫編。
                    st.info(
                        f"ℹ️ 偵測到 **{len(_orphans)} 筆**暫編地號之原地號未在 土地歸戶 歸戶清單中"
                        f"（已嘗試正規化比對：去除段碼前綴與 -0 子號後綴）。"
                        f"請檢核 土地歸戶 是否漏填，或為地號編碼差異。"
                    )
                    _orphan_rows = [{
                        '暫編地號': tp.get('暫編地號', ''),
                        '原地號': tp.get('原地號', '') or '(無)',
                        '正規化後地號': _normalize_landno(tp.get('原地號', '')),
                        '所屬街廓': tp.get('所屬街廓', ''),
                        '幾何面積(㎡)': round(float(tp.get('幾何面積_m2', tp.get('面積_m2', 0)) or 0), 2),
                        '分攤登記面積(㎡)': round(float(tp.get('分攤登記面積_m2', 0) or 0), 2),
                    } for tp in _orphans]
                    st.dataframe(pd.DataFrame(_orphan_rows),
                                 use_container_width=True, hide_index=True)
                    st.caption(
                        "※ W-A.2 §5-1：系統不再對有原地號歸屬之暫編做自動剔除。"
                        "舊版「建議剔除非重劃範圍內地號」邏輯已停用——"
                        "歸屬到真實地號的暫編一律保留。"
                    )
                else:
                    _msg_norm = f"（其中 {_matched_via_norm} 筆透過格式正規化對應上）" if _matched_via_norm > 0 else ""
                    _own_groups_check = st.session_state.get('t8_ownership_groups', {}) or {}
                    _real_landno_count = sum(len(v) for v in _own_groups_check.values())
                    st.success(
                        f"✅ 土地歸戶 歸戶交叉檢核通過：{len(temp_parcels)} 筆暫編地號全部對應到 土地歸戶 歸戶清單"
                        f"（共 {_real_landno_count} 筆地號）{_msg_norm}"
                    )
            elif temp_parcels and not _own_map_check:
                st.info(
                    "ℹ️ 尚未匯入 土地歸戶 資料。建議先至 土地歸戶 上傳地籍 Excel，"
                    "系統會自動比對暫編地號與歸戶清單。"
                )

        # ---------- 步驟 C-2：暫編地號圖面定位與高亮（任務二） ----------
        if temp_parcels:
            st.markdown("---")
            st.markdown("#### 🔍 步驟 C-2：暫編地號定位（圖面高亮）")
            st.caption(
                "DXF 轉入後若有地號錯誤或無地號，可在此選擇要定位之暫編地號 → "
                "圖面以亮紅色高亮 + 質心標記 + 自動縮放，便於對照修正。"
            )

            # 收集所有暫編地號 + 「(無地號)」特殊選項
            _no_parent_count = sum(1 for tp in temp_parcels if not tp.get('原地號'))
            _temp_no_options = ['（不高亮）']
            if _no_parent_count > 0:
                _temp_no_options.append(f'(無地號) — 共 {_no_parent_count} 筆')
            _temp_no_options.extend(sorted(
                {str(tp.get('暫編地號', '') or '') for tp in temp_parcels
                 if tp.get('暫編地號')}
            ))

            _hl_col1, _hl_col2 = st.columns([2, 1])
            _selected_hl = _hl_col1.selectbox(
                "選擇欲定位之暫編地號",
                options=_temp_no_options,
                key='f3_highlight_temp_no',
                help="選擇後地圖會自動縮放至該地號並以亮紅色高亮"
            )
            _zoom_on = _hl_col2.checkbox(
                "自動縮放至該地號", value=True,
                key='f3_highlight_auto_zoom'
            )

            # 解析使用者選擇 → 真正的高亮鍵
            _hl_key = None
            if _selected_hl == '（不高亮）':
                _hl_key = None
            elif _selected_hl.startswith('(無地號)'):
                _hl_key = '(無地號)'
            else:
                _hl_key = _selected_hl

            # 將解析後高亮鍵寫至獨立 session_state key（不可覆寫 widget 自身綁定的 key）
            st.session_state['f3_highlight_temp_no_resolved'] = _hl_key
            # f3_highlight_auto_zoom 由 checkbox 自動維護於 session_state，下游直接讀取即可

            if _hl_key:
                _zmap_preview = dict(st.session_state.get('f3_parcel_zones', {}))
                _fig_locator = _render_f3_overlay_figure(
                    classified_blocks, temp_parcels, _zmap_preview,
                    interactive=False, height=520,
                    highlight_temp_no=_hl_key,
                    auto_zoom_to_highlight=_zoom_on,
                )
                st.plotly_chart(_fig_locator, use_container_width=True)

                # 找出對應地號之詳細資訊
                _matched = []
                for tp in temp_parcels:
                    if _hl_key == '(無地號)' and not tp.get('原地號'):
                        _matched.append(tp)
                    elif (str(tp.get('暫編地號', '') or '') == _hl_key
                          or str(tp.get('原地號', '') or '') == _hl_key):
                        _matched.append(tp)
                if _matched:
                    _detail = []
                    for tp in _matched:
                        _detail.append({
                            '暫編地號': tp.get('暫編地號', ''),
                            '原地號': tp.get('原地號', '') or '(無)',
                            '所屬街廓': tp.get('所屬街廓', ''),
                            '面積(㎡)': round(float(tp.get('分攤登記面積_m2', tp.get('面積_m2', 0)) or 0), 2),
                            '質心 X': round(float(tp.get('centroid_x', 0) or 0), 2),
                            '質心 Y': round(float(tp.get('centroid_y', 0) or 0), 2),
                        })
                    st.dataframe(pd.DataFrame(_detail),
                                 use_container_width=True, hide_index=True)
                else:
                    st.info(f"找不到符合「{_hl_key}」之暫編地號")
            else:
                st.info("選擇暫編地號後，地圖會自動高亮並放大至該位置。")

        # ---------- 🆕 Phase 4 Task G：統一 GIS 互動分析地圖 ----------
        if temp_parcels:
            st.markdown("---")
            st.markdown("### 🗺️ 統一 GIS 互動分析地圖")
            st.caption(
                "**1️⃣** 在右側勾選需顯示的圖層　**2️⃣** 在「點擊將執行」選擇要做的動作　"
                "**3️⃣** 在地圖上點擊地號 / 點位　**4️⃣** 完成街角地與基準線設定後，於下方「步驟 G」按按鈕產生試分配地籍圖。"
            )

            # 預設圖層
            _layers = dict(st.session_state.get('f3_visible_layers', {
                'block_boundaries': True, 'pre_cadastral': True, 'price_zones': True,
                'ownership_dist': False, 'corner_status': True,
                'manual_baseline': True, 'trial_allocation': False,
                'road_centerlines': True, 'problem_parcels': False,
                'ownership_after_alloc': False,   # 🆕 V12 模組 4-B
            }))
            # 圖層 checkbox（兩列共 10 個，含 V12 重劃後歸戶分佈）
            _lc1, _lc2, _lc3, _lc4, _lc5 = st.columns(5)
            _layers['block_boundaries'] = _lc1.checkbox("街廓邊界", value=_layers.get('block_boundaries', True), key='f3_um_lyr_blocks')
            _layers['pre_cadastral']    = _lc2.checkbox("重劃前地籍線", value=_layers.get('pre_cadastral', True), key='f3_um_lyr_cad')
            _layers['price_zones']      = _lc3.checkbox("地價區段著色", value=_layers.get('price_zones', True), key='f3_um_lyr_zones')
            _layers['ownership_dist']   = _lc4.checkbox("歸戶分佈（前）", value=_layers.get('ownership_dist', False), key='f3_um_lyr_owners')
            _layers['corner_status']    = _lc5.checkbox("街角狀態", value=_layers.get('corner_status', True), key='f3_um_lyr_corners')
            _lc6, _lc7, _lc8, _lc9, _lc10 = st.columns(5)
            _layers['manual_baseline']  = _lc6.checkbox("分配基準線", value=_layers.get('manual_baseline', True), key='f3_um_lyr_baseline')
            _layers['trial_allocation'] = _lc7.checkbox("試分配結果", value=_layers.get('trial_allocation', False), key='f3_um_lyr_trial')
            _layers['road_centerlines'] = _lc8.checkbox("道路中心線", value=_layers.get('road_centerlines', True), key='f3_um_lyr_road')
            _layers['problem_parcels']  = _lc9.checkbox(
                "⚠️ 問題地塊", value=_layers.get('problem_parcels', False),
                key='f3_um_lyr_problems',
                help="標示 (1) 無地號之地塊（紅色）(2) 有地號但 土地歸戶 找不到歸戶者（黃色）",
            )
            # 🆕 V12 模組 4-B：重劃後土地歸戶分佈
            _layers['ownership_after_alloc'] = _lc10.checkbox(
                "🎨 重劃後歸戶分佈", value=_layers.get('ownership_after_alloc', False),
                key='f3_um_lyr_alloc_owners',
                help="依 G 值迭代結果之 cut_coords，按歸戶 ID 著色。若顏色出現於非可建築街廓 → 視為幾何錯誤警示。"
            )
            st.session_state['f3_visible_layers'] = _layers

            # 操作模式 radio
            # 🆕 Phase 6：新增 N（精準中心線）+ O（補正地號），移除舊「✂️ 手動分割道路」
            _action_options = [
                '🔍 瀏覽',
                '🏷️ 指定地價區段',
                '⭐ 切換街角地',
                # 🆕 Phase 11：使用者直接於 DXF BASELINE / CENTERLINE 圖層繪製，
                # 系統自動讀入；以下兩項已停用：
                #   '📐 設定分配基準線（點擊邊界）',
                #   '🛣️ 生成精準中心線（點擊道路邊界平移）',
                '✏️ 補正地號（點選問題地塊）',
            ]
            _curr_action = st.session_state.get('f3_unified_action', '🔍 瀏覽')
            if _curr_action not in _action_options:
                _curr_action = '🔍 瀏覽'
            _action = st.radio(
                "🎯 點擊地圖將執行：",
                options=_action_options,
                index=_action_options.index(_curr_action),
                horizontal=True, key='f3_um_action',
            )
            st.session_state['f3_unified_action'] = _action

            # 模式特定參數列
            _mode_param_col = st.container()
            with _mode_param_col:
                if _action == '🏷️ 指定地價區段':
                    _zc1, _zc2 = st.columns([3, 1])
                    _curr_zone = _zc1.text_input(
                        "區段代碼（下次點擊將寫入此值）",
                        value=st.session_state.get('f3_current_zone', '甲段'),
                        key='f3_um_zone_input',
                    )
                    st.session_state['f3_current_zone'] = _curr_zone
                    _bulk_clicked = _zc2.button(
                        "🔄 一鍵填入剩餘", use_container_width=True, key='f3_um_bulk_fill',
                        help="將所有目前未標註的地號一次填入此區段代碼"
                    )
                    if _bulk_clicked:
                        _zmap_bulk = dict(st.session_state.get('f3_parcel_zones', {}))
                        _all_orig = {tp.get('原地號', '') for tp in (temp_parcels or [])}
                        _added = 0
                        for orig in _all_orig:
                            if orig and not _zmap_bulk.get(orig):
                                _zmap_bulk[orig] = _curr_zone
                                _added += 1
                        st.session_state['f3_parcel_zones'] = _zmap_bulk
                        st.toast(f"✅ 已將 {_added} 筆未標註地號一次填入「{_curr_zone}」", icon='🔄')
                        st.rerun()
                elif _action == '⭐ 切換街角地':
                    _default_side = st.radio(
                        "預設街角側別（自動偵測失敗時 fallback 用此值）",
                        options=['左側', '右側'], horizontal=True, key='f3_um_default_side',
                        help="點選街角地時系統會以幾何投影自動判定左/右側；僅當投影失敗時才使用此預設值。"
                    )
                elif _action == '📐 設定分配基準線（點擊邊界）':
                    # 🆕 Phase 6 Task M：街廓邊 hover 高亮 + 一鍵鎖定
                    st.info(
                        "💡 **滑鼠移至街廓任一條邊（橙色高亮）→ tooltip 顯示邊資訊 → 點擊即鎖定**\n\n"
                        "📐 系統將該邊作為該街廓之分配基準線。"
                        "如方向相反，可使用下方「🔄 反轉 180°」按鈕即時校正。"
                    )
                    _mbl_show = st.session_state.get('f3_manual_baseline', {}) or {}
                    _enabled_lbl = [lbl for lbl, v in _mbl_show.items()
                                    if v.get('enabled') and v.get('point')]
                    if _enabled_lbl:
                        # 顯示來源（CAD 直讀 / 手動）
                        _src_marks = []
                        for lbl in _enabled_lbl:
                            src = _mbl_show[lbl].get('source', 'manual')
                            _src_marks.append(f"{lbl}{'(📏CAD)' if src=='cad' else ''}")
                        st.caption(
                            f"🟢 目前已設定基準線之街廓：**{', '.join(_src_marks)}**"
                            f"（共 {len(_enabled_lbl)} 個；CAD = 由 DXF BASELINE 圖層直讀）"
                        )
                        # 🆕 Phase 7 Module 3 (M3-4)：180° 反轉按鈕
                        _rc1, _rc2 = st.columns([3, 2])
                        _rev_sel = _rc1.selectbox(
                            "選擇要反轉方向之街廓基準線",
                            options=_enabled_lbl,
                            key='f3_um_reverse_select',
                        )
                        if _rc2.button("🔄 反轉 180°", use_container_width=True,
                                        key='f3_um_reverse_baseline',
                                        help="將該街廓基準線之方向角加 180°（即時反轉左右推進視角）"):
                            _mbl_w = dict(_mbl_show)
                            _curr = dict(_mbl_w[_rev_sel])
                            _new_ang = (float(_curr.get('angle_deg', 0.0)) + 180.0) % 360.0
                            # 修正回 [-180, 180] 範圍
                            if _new_ang > 180:
                                _new_ang -= 360
                            _curr['angle_deg'] = _new_ang
                            _mbl_w[_rev_sel] = _curr
                            st.session_state['f3_manual_baseline'] = _mbl_w
                            st.toast(
                                f"🔄 {_rev_sel} 基準線已反轉 → 方向 {_new_ang:.1f}°",
                                icon='🔄'
                            )
                            st.rerun()
                elif _action == '🛣️ 生成精準中心線（點擊道路邊界平移）':
                    # 🆕 Phase 6 Task N：CAD 級平移生成
                    _re_now = st.session_state.get('f3_road_edge_pick', None)
                    if _re_now:
                        st.info(
                            f"🟡 已鎖定道路 **{_re_now['road_label']}** 之第 "
                            f"{_re_now['edge_idx']+1} 條邊（長度 {_re_now['length']:.2f}m）"
                        )
                        _w1, _w2, _w3 = st.columns([2, 1, 1])
                        _w_val = _w1.number_input(
                            "請輸入該道路之計畫寬度 W (m)",
                            min_value=1.0, max_value=50.0,
                            value=float(st.session_state.get('f3_road_w_default', 8.0)),
                            step=0.5, format="%.2f",
                            key='f3_um_road_w_input',
                            help="計畫道路全寬。系統將以 W/2 平移該邊向道路內側生成中心線。",
                        )
                        st.session_state['f3_road_w_default'] = _w_val
                        if _w2.button("🔧 生成精準中心線", use_container_width=True,
                                       type='primary', key='f3_um_offset_centerline'):
                            _result = _offset_edge_to_centerline(
                                road_label=_re_now['road_label'],
                                edge_p1=_re_now['p1'],
                                edge_p2=_re_now['p2'],
                                width_m=float(_w_val),
                                classified_blocks=classified_blocks,
                            )
                            if _result and _result.get('centerline_pts'):
                                _mc_w = dict(st.session_state.get('f3_manual_road_centerlines', {}))
                                _mc_w[_re_now['road_label']] = _result['centerline_pts']
                                st.session_state['f3_manual_road_centerlines'] = _mc_w
                                st.session_state['f3_road_edge_pick'] = None
                                st.toast(
                                    f"✅ 道路 {_re_now['road_label']} 中心線已生成"
                                    f"（W/2={_w_val/2:.2f}m 平移）",
                                    icon='🔧'
                                )
                                st.rerun()
                            else:
                                st.warning("⚠️ 平移失敗，請檢查道路 polygon 是否有效")
                        if _w3.button("✖️ 取消鎖定", key='f3_um_cancel_redge',
                                       use_container_width=True):
                            st.session_state['f3_road_edge_pick'] = None
                            st.rerun()
                    else:
                        st.info(
                            "💡 **CAD 級精準中心線生成（Offset Curve 演算法）**\n\n"
                            "1️⃣ 點擊地圖上任一條**亮黃色道路邊界**（道路類街廓邊已自動高亮）\n"
                            "2️⃣ 上方輸入該道路之計畫寬度 W\n"
                            "3️⃣ 按「🔧 生成精準中心線」→ 系統將該邊向道路內側平移 W/2\n\n"
                            "🚫 **不採手動連點繪製** — 避免人為誤差導致地主分配面積錯誤"
                        )
                        _mc_show = st.session_state.get('f3_manual_road_centerlines', {}) or {}
                        if _mc_show:
                            with st.expander(
                                f"🟢 目前已生成 {len(_mc_show)} 條精準中心線", expanded=False
                            ):
                                for _r, _pts in _mc_show.items():
                                    _c1, _c2 = st.columns([3, 1])
                                    _c1.write(f"- **{_r}**：{len(_pts)} 點")
                                    if _c2.button("🗑️", key=f'f3_del_mc_{_r}'):
                                        _mc_show2 = dict(_mc_show)
                                        _mc_show2.pop(_r, None)
                                        st.session_state['f3_manual_road_centerlines'] = _mc_show2
                                        st.rerun()
                elif _action == '✏️ 補正地號（點選問題地塊）':
                    # 🆕 Phase 6 Task O：補正地號
                    _correcting = st.session_state.get('f3_correcting_parcel', None)
                    if _correcting:
                        st.warning(
                            f"📝 正在補正暫編地號 **{_correcting}** 的原地號："
                        )
                        _new_no = st.text_input(
                            "請輸入正確地號（例如 628-6）",
                            value="",
                            key='f3_um_correct_landno_input',
                        )
                        _cc1, _cc2 = st.columns(2)
                        if _cc1.button("✅ 確認補正", use_container_width=True,
                                        key='f3_um_correct_confirm', type='primary'):
                            if _new_no.strip():
                                _ovr = dict(st.session_state.get('f3_parcel_no_overrides', {}))
                                _ovr[_correcting] = _new_no.strip()
                                st.session_state['f3_parcel_no_overrides'] = _ovr
                                # 🆕 Phase 7 Module 3 (M3-5)：即時連動 temp_parcels
                                # 直接覆寫 '原地號' 欄位，使下游歸戶查找立即生效
                                _tp_list_now = st.session_state.get('f3_temp_parcels', [])
                                for _tp_iter in _tp_list_now:
                                    if _tp_iter.get('暫編地號') == _correcting:
                                        _tp_iter['原地號_override'] = _new_no.strip()
                                        _tp_iter['原地號'] = _new_no.strip()
                                        break
                                st.session_state['f3_temp_parcels'] = _tp_list_now
                                st.session_state['f3_correcting_parcel'] = None
                                st.toast(
                                    f"✅ 已將暫編 {_correcting} 補正為「{_new_no.strip()}」"
                                    "（即時連動歸戶查找）",
                                    icon='✏️'
                                )
                                st.rerun()
                            else:
                                st.warning("⚠️ 地號不可為空")
                        if _cc2.button("✖️ 取消", use_container_width=True,
                                        key='f3_um_correct_cancel'):
                            st.session_state['f3_correcting_parcel'] = None
                            st.rerun()
                    else:
                        st.info(
                            "💡 請於地圖上點擊**問題地塊**（紅色斜線=無地號 / 黃色斜線=未歸戶）\n\n"
                            "✨ 補正後的地號將寫入 `f3_parcel_no_overrides`，後續歸戶查找時優先使用。\n\n"
                            "🔔 請先勾選上方「⚠️ 問題地塊」圖層方可看到問題地塊。"
                        )
                        _ovr_show = st.session_state.get('f3_parcel_no_overrides', {}) or {}
                        if _ovr_show:
                            with st.expander(
                                f"🟢 目前已補正 {len(_ovr_show)} 筆地號", expanded=False
                            ):
                                for _k, _v in _ovr_show.items():
                                    _o1, _o2 = st.columns([3, 1])
                                    _o1.write(f"- 暫編 **{_k}** → 地號 `{_v}`")
                                    if _o2.button("🗑️", key=f'f3_del_ovr_{_k}'):
                                        _ovr_show2 = dict(_ovr_show)
                                        _ovr_show2.pop(_k, None)
                                        st.session_state['f3_parcel_no_overrides'] = _ovr_show2
                                        st.rerun()
                else:  # 🔍 瀏覽
                    st.caption("ℹ️ 目前為瀏覽模式：點擊不會觸發任何動作。")

            # ─── 統一地圖（fragment）──────────────────────────────
            try:
                from streamlit_plotly_events import plotly_events as _pevents_um
                _has_pe_um = True
            except Exception:
                _pevents_um = None
                _has_pe_um = False

            @st.fragment
            def _f3_unified_map_fragment():
                _zmap_now = dict(st.session_state.get('f3_parcel_zones', {}))
                _own_map_now = st.session_state.get('t8_ownership_map', {}) or {}
                _g_vals_now = st.session_state.get('f3_G_values', []) or []
                _mbl_now = st.session_state.get('f3_manual_baseline', {}) or {}
                _splits_now = st.session_state.get('f3_manual_road_splits', {}) or {}
                _params_now = st.session_state.get('f3_g_iter_params', {}) or {}
                _vl_now = st.session_state.get('f3_visible_layers', {}) or {}
                _pm_now = st.session_state.get('f3_baseline_pick_mode', None)
                _rs_now = st.session_state.get('f3_road_split_mode', None)
                _action_now = st.session_state.get('f3_unified_action', '🔍 瀏覽')

                # 🆕 Phase 6 新增：手動精準中心線、道路邊鎖定狀態、action_mode
                _mc_now = st.session_state.get('f3_manual_road_centerlines', {}) or {}
                _redge_pick_now = st.session_state.get('f3_road_edge_pick', None)

                (_fig, _clk_offset, _clk_parcels,
                 _edge_offset, _edge_meta,
                 _redge_offset, _redge_meta) = _render_f3_unified_map(
                    classified_blocks=classified_blocks,
                    temp_parcels=temp_parcels,
                    zones_map=_zmap_now,
                    ownership_map=_own_map_now,
                    g_values=_g_vals_now,
                    manual_baselines=_mbl_now,
                    manual_road_splits=_splits_now,
                    manual_road_centerlines=_mc_now,
                    params_map=_params_now,
                    visible_layers=_vl_now,
                    pick_mode_state=_pm_now,
                    road_split_state=_rs_now,
                    road_edge_pick=_redge_pick_now,
                    action_mode=_action_now,
                    height=620,
                )

                if not _has_pe_um:
                    st.warning("⚠️ 缺少 streamlit-plotly-events，僅顯示靜態地圖（無法點選）。")
                    st.plotly_chart(_fig, use_container_width=True)
                    return

                _ck_um = st.session_state.get('f3_unified_click_counter', 0)
                # 🆕 Phase 7 Module 3 (M3-2)：將 action name 雜湊入 key，
                #    強制清除 Plotly 殘留快取（切換模式時）
                import hashlib as _hl_um
                _action_hash = _hl_um.md5(str(_action_now).encode()).hexdigest()[:6]
                _picked = _pevents_um(_fig, click_event=True, hover_event=False,
                                       key=f'f3_um_clicker_{_ck_um}_{_action_hash}',
                                       override_height=620)

                if not _picked:
                    return

                ev = _picked[0]
                cn = ev.get('curveNumber')
                cx, cy = ev.get('x'), ev.get('y')

                # ── 點擊解析：優先用 curveNumber 對到 click target trace；
                #               否則用 x/y + shapely 後援 ──
                _matched = None
                if cn is not None and _clk_offset <= cn < _clk_offset + len(_clk_parcels):
                    _matched = _clk_parcels[cn - _clk_offset]
                elif cx is not None and cy is not None:
                    # 後援：可能點到非 click-target trace（如歸戶填色），改 shapely 解析
                    _matched = _resolve_clicked_parcel(cx, cy, temp_parcels)

                # ── 動作分流 ──
                if _action_now == '🔍 瀏覽':
                    return

                elif _action_now == '🏷️ 指定地價區段':
                    if _matched:
                        _zone = st.session_state.get('f3_current_zone', '甲段')
                        _zmap_now[_matched.get('原地號', '')] = _zone
                        st.session_state['f3_parcel_zones'] = _zmap_now
                        st.toast(f"✅ 已標 {_matched.get('暫編地號','')} → 區段 {_zone}", icon='🏷️')
                    else:
                        st.toast("⚠️ 點擊位置未對到任何地號", icon='⚠️')

                elif _action_now == '⭐ 切換街角地':
                    matched = _matched
                    if matched:
                        k = matched['暫編地號']
                        blk_label = matched.get('所屬街廓', '')
                        blk_meta = next((b for b in classified_blocks
                                         if b.get('label') == blk_label), None)
                        # 🆕 Phase 11 v2：以使用者「預設街角側別」下拉為主導
                        # 邏輯：
                        #   - 未標 → 標記為使用者下拉之側別（_default）
                        #   - 已標且 side == 使用者下拉 → 取消街角地（清除）
                        #   - 已標且 side != 使用者下拉 → 切換為使用者下拉之側別
                        # 自動偵測（_auto）僅做為 toast 提示，不會覆寫使用者下拉之選擇
                        _auto = _auto_detect_corner_side(matched, blk_meta) if blk_meta else ''
                        _default = st.session_state.get('f3_um_default_side', '左側')
                        if _default not in ('左側', '右側'):
                            _default = '左側'
                        curr = dict(_params_now.get(k, {}))
                        _was_corner = bool(curr.get('is_corner', False))
                        _was_side = curr.get('side', '無')

                        if not _was_corner:
                            # 第 1 次點擊：以使用者下拉為準
                            curr['is_corner'] = True
                            curr['side'] = _default
                            _hint = (f"（系統幾何判定為 {_auto}）"
                                     if _auto in ('左側', '右側') and _auto != _default
                                     else "")
                            _msg = f"⭐ 已標街角地：**{_default}**{_hint}"
                        elif _was_side == _default:
                            # 已標且側別與下拉相同 → 取消街角地
                            curr['is_corner'] = False
                            curr['side'] = '無'
                            _msg = f"❎ 已取消街角地（原為 {_default}）"
                        else:
                            # 已標但側別與下拉不同 → 切換為下拉之側別
                            curr['is_corner'] = True
                            curr['side'] = _default
                            _msg = f"🔁 切換街角側別：{_was_side} → **{_default}**"

                        if 'W' not in curr:
                            curr['W'] = 0.0
                        _params_now[k] = curr
                        st.session_state['f3_g_iter_params'] = _params_now
                        st.session_state['f3_g_needs_recalc'] = True
                        st.toast(f"{_msg}：{k}", icon='🎯')

                elif _action_now == '📐 設定分配基準線（點擊邊界）':
                    # 🆕 Phase 6 Task M：優先用 edge_interactive trace 解析（curveNumber-based）
                    _edge_picked = None
                    if (cn is not None and _edge_meta
                        and _edge_offset <= cn < _edge_offset + len(_edge_meta)):
                        _edge_picked = _edge_meta[cn - _edge_offset]
                    if _edge_picked:
                        # ── 直接從 edge_meta 取得邊資訊 ──
                        blk_lbl, edge_idx, p1, p2, length = _edge_picked
                        import math as _m_um
                        import numpy as _np_um
                        dx = p2[0] - p1[0]
                        dy = p2[1] - p1[1]
                        angle = _m_um.degrees(_m_um.atan2(dy, dx))
                        # 反向防呆
                        try:
                            from shapely.geometry import Polygon as _SP_chk
                            _bmeta = next((b for b in classified_blocks
                                           if b.get('label') == blk_lbl), None)
                            if _bmeta and _bmeta.get('vertices'):
                                _bp = _SP_chk(_bmeta['vertices'])
                                if _bp.is_valid:
                                    _mbr_dir, _ = _detect_block_allocation_dir(
                                        _bp, _np_um.array([1.0, 0.0])
                                    )
                                    _ang_rad = _m_um.radians(angle)
                                    _uvec = _np_um.array([_m_um.cos(_ang_rad),
                                                           _m_um.sin(_ang_rad)])
                                    if float(_np_um.dot(_uvec, _mbr_dir)) < 0:
                                        st.toast(
                                            "ℹ️ 您鎖定的邊與街廓長軸夾角 > 90°；"
                                            "如方向相反可至步驟 G 微調 ±180°",
                                            icon='ℹ️'
                                        )
                        except Exception:
                            pass
                        _mbl_w = dict(st.session_state.get('f3_manual_baseline', {}))
                        _mbl_w[blk_lbl] = {
                            'point': (float(p1[0]), float(p1[1])),
                            'angle_deg': float(angle),
                            'enabled': True,
                        }
                        st.session_state['f3_manual_baseline'] = _mbl_w
                        st.session_state['f3_baseline_pick_mode'] = None
                        st.toast(
                            f"✅ 街廓 {blk_lbl} 第 {edge_idx+1} 條邊（長 {length:.1f}m）"
                            f"已設為分配基準線，方向 {angle:.1f}°",
                            icon='📐'
                        )
                    else:
                        # ── Fallback：Phase 5 _snap_baseline_to_block_edge ──
                        _blk_for_snap = None
                        if _matched and _matched.get('所屬街廓'):
                            _blk_for_snap = next(
                                (b for b in classified_blocks
                                 if b.get('label') == _matched['所屬街廓']),
                                None
                            )
                        if _blk_for_snap is None and cx is not None and cy is not None:
                            _blk_for_snap = _resolve_clicked_block(
                                cx, cy, classified_blocks, tolerance=5.0
                            )
                        if _blk_for_snap is not None:
                            _snap = _snap_baseline_to_block_edge(
                                cx, cy, _blk_for_snap, max_dist=15.0
                            )
                            if _snap is not None:
                                _blk_lbl = _blk_for_snap.get('label', '')
                                _mbl_w = dict(st.session_state.get('f3_manual_baseline', {}))
                                _mbl_w[_blk_lbl] = {
                                    'point': _snap['point'],
                                    'angle_deg': _snap['angle_deg'],
                                    'enabled': True,
                                }
                                st.session_state['f3_manual_baseline'] = _mbl_w
                                st.toast(
                                    f"✅ 已將 {_blk_lbl} 第 {_snap['edge_idx']+1} 條邊"
                                    f"設為分配基準線（fallback 模式）",
                                    icon='📐'
                                )
                            else:
                                st.toast("⚠️ 點擊位置未對到任何街廓邊", icon='⚠️')
                        else:
                            st.toast("⚠️ 點擊位置未對到任何街廓邊", icon='⚠️')

                elif _action_now == '🛣️ 生成精準中心線（點擊道路邊界平移）':
                    # 🆕 Phase 6 Task N：點擊道路邊界 → 鎖定（後續輸入 W → 平移生成）
                    _redge_picked = None
                    if (cn is not None and _redge_meta
                        and _redge_offset <= cn < _redge_offset + len(_redge_meta)):
                        _redge_picked = _redge_meta[cn - _redge_offset]
                    if _redge_picked:
                        road_lbl, edge_idx, p1, p2, length, road_blk = _redge_picked
                        st.session_state['f3_road_edge_pick'] = {
                            'road_label': road_lbl,
                            'edge_idx': edge_idx,
                            'p1': (float(p1[0]), float(p1[1])),
                            'p2': (float(p2[0]), float(p2[1])),
                            'length': float(length),
                        }
                        st.toast(
                            f"🟡 已鎖定道路 {road_lbl} 第 {edge_idx+1} 條邊（{length:.1f}m），"
                            "請於上方輸入路寬 → 按「生成精準中心線」",
                            icon='🛣️'
                        )
                    else:
                        st.toast(
                            "⚠️ 請點擊**亮黃色**道路邊界（非道路類街廓邊不會高亮）",
                            icon='⚠️'
                        )

                elif _action_now == '✏️ 補正地號（點選問題地塊）':
                    # 🆕 Phase 6 Task O：點選問題地塊 → 設定 correcting parcel
                    if _matched:
                        st.session_state['f3_correcting_parcel'] = _matched.get('暫編地號', '')
                        st.toast(
                            f"📝 已選定 {_matched.get('暫編地號','')}，請於上方輸入正確地號",
                            icon='✏️'
                        )
                    else:
                        st.toast("⚠️ 點擊位置未對到任何地塊", icon='⚠️')

                elif _action_now == '__OBSOLETE_baseline_two_click__':
                    # 已被 Phase 5 Task J 取代，保留 dead code 區段以防舊 session 觸發
                    clicks = list((_pm_now or {}).get('clicks', []))
                    if not clicks:
                        _target_lbl = ''
                        if _matched and _matched.get('所屬街廓'):
                            _target_lbl = _matched['所屬街廓']
                        if not _target_lbl and cx is not None and cy is not None:
                            blk = _resolve_clicked_block(cx, cy, classified_blocks)
                            if blk:
                                _target_lbl = blk.get('label', '')
                        if _target_lbl and cx is not None and cy is not None:
                            clicks.append((float(cx), float(cy)))
                            st.session_state['f3_baseline_pick_mode'] = {
                                'block': _target_lbl, 'clicks': clicks
                            }
                            st.toast(f"📍 已鎖定 {_target_lbl}，請點擊第 2 點", icon='1️⃣')
                        else:
                            st.toast("⚠️ 第 1 點未落入任何街廓內", icon='⚠️')
                    else:
                        clicks.append((float(cx), float(cy)))
                        import math as _m_um
                        import numpy as _np_um
                        _x1, _y1 = clicks[0]
                        _x2, _y2 = clicks[1]
                        _dx = _x2 - _x1; _dy = _y2 - _y1
                        _seg = _m_um.hypot(_dx, _dy)
                        _blk_lbl = (_pm_now or {}).get('block', '')
                        if _seg < 1e-6:
                            st.warning("⚠️ 兩點過近無法計算方向")
                            st.session_state['f3_baseline_pick_mode'] = None
                        else:
                            _angle = _m_um.degrees(_m_um.atan2(_dy, _dx))
                            try:
                                _bmeta = next((b for b in classified_blocks
                                               if b.get('label') == _blk_lbl), None)
                                if _bmeta and _bmeta.get('vertices'):
                                    from shapely.geometry import Polygon as _SP_chk
                                    _bp = _SP_chk(_bmeta['vertices'])
                                    if _bp.is_valid:
                                        _mbr_dir, _ = _detect_block_allocation_dir(
                                            _bp, _np_um.array([1.0, 0.0])
                                        )
                                        _uvec = _np_um.array([_dx, _dy]) / _seg
                                        if float(_np_um.dot(_uvec, _mbr_dir)) < 0:
                                            st.warning(
                                                "⚠️ 您指定的方向與街廓長軸夾角 > 90°（即與街廓主走向相反）。"
                                                "如需反向，可手動於下方步驟 G「手動基準線設定」調整方向角 ±180°。"
                                            )
                            except Exception:
                                pass
                            _mbl_w = dict(st.session_state.get('f3_manual_baseline', {}))
                            _mbl_w[_blk_lbl] = {
                                'point': (_x1, _y1),
                                'angle_deg': _angle,
                                'enabled': True,
                            }
                            st.session_state['f3_manual_baseline'] = _mbl_w
                            st.session_state['f3_baseline_pick_mode'] = None
                            st.toast(
                                f"📐 街廓 {_blk_lbl} 基準線已設定：起點 ({_x1:.2f}, {_y1:.2f})、"
                                f"方向 {_angle:.2f}°", icon='✅'
                            )

                elif _action_now == '✂️ 手動分割道路（連點 2 點）':
                    clicks = list((_rs_now or {}).get('clicks', []))
                    if not clicks:
                        # 第 1 點 → 自動鎖定道路街廓
                        # 🆕 Phase 5 Task K：寬鬆分類比對（含「市區道路」「縣道」等變體）
                        #                   + 容差 2.0 → 3.0m（道路邊界點選更友善）
                        _r_lbl = ''
                        if cx is not None and cy is not None:
                            road_blk = _resolve_clicked_block(
                                cx, cy, classified_blocks,
                                category_predicate=_is_road_like_category,
                                tolerance=3.0,
                            )
                            if road_blk:
                                _r_lbl = road_blk.get('label', '')
                        if _r_lbl and cx is not None and cy is not None:
                            clicks.append((float(cx), float(cy)))
                            st.session_state['f3_road_split_mode'] = {
                                'road': _r_lbl, 'clicks': clicks
                            }
                            st.toast(f"📍 已鎖定道路 {_r_lbl}，請點擊第 2 點", icon='1️⃣')
                        else:
                            st.toast("⚠️ 第 1 點需落在道路用地內", icon='⚠️')
                    else:
                        # 第 2 點 → 切割
                        clicks.append((float(cx), float(cy)))
                        from shapely.geometry import LineString as _SL_s, Polygon as _SP_s
                        from shapely.ops import split as _ssplit
                        _r_lbl = (_rs_now or {}).get('road', '')
                        _r_blk = next((b for b in classified_blocks
                                       if b.get('label') == _r_lbl), None)
                        if _r_blk and _r_blk.get('vertices'):
                            try:
                                _rp = _SP_s(_r_blk['vertices'])
                                if not _rp.is_valid:
                                    _rp = _rp.buffer(0)
                                _splitter = _SL_s([clicks[0], clicks[1]])
                                _result = _ssplit(_rp, _splitter)
                                _pieces = [g for g in _result.geoms if g.area >= 1.0]
                                _splits_w = dict(st.session_state.get('f3_manual_road_splits', {}))
                                _splits_w[_r_lbl] = _pieces
                                st.session_state['f3_manual_road_splits'] = _splits_w
                                st.session_state['f3_road_split_mode'] = None
                                st.toast(f"✂️ 道路 {_r_lbl} 已切為 {len(_pieces)} 塊", icon='✅')
                            except Exception as _eS:
                                st.toast(f"⚠️ 切割失敗：{_eS}", icon='⚠️')
                                st.session_state['f3_road_split_mode'] = None

                # 旋轉 click counter
                st.session_state['f3_unified_click_counter'] = _ck_um + 1
                st.rerun(scope='app')

            _f3_unified_map_fragment()

        # ---------- 步驟 D：重劃前地價區段標註（Phase 4 簡化版：移除舊圖選器）----------
        # 🆕 Phase 4：地圖點選功能已整合至上方「統一 GIS 互動分析地圖」（操作模式 = 🏷️ 指定地價區段）。
        # 此步驟保留：(1) 進度顯示 + 一鍵清空；(2) 暫編地號清單表格。
        if temp_parcels:
            st.markdown("---")
            st.markdown("#### 🏷️ 步驟 D：重劃前地價區段標註（進度與清單）")
            st.caption(
                "💡 圖面點選功能已遷移至上方「🗺️ 統一 GIS 互動分析地圖」。"
                "請於該地圖選擇「🏷️ 指定地價區段」操作模式後在地圖上點選地號。"
            )
            if 'f3_parcel_zones' not in st.session_state:
                st.session_state['f3_parcel_zones'] = {}

            all_original_parcels = sorted({tp['原地號'] for tp in temp_parcels})
            _zmap_d = dict(st.session_state.get('f3_parcel_zones', {}))
            _labeled = sum(1 for op in all_original_parcels if _zmap_d.get(op))
            _total = len(all_original_parcels)
            _dc1, _dc2 = st.columns([3, 1])
            _dc1.metric("📊 已標註進度", f"{_labeled} / {_total} 筆")
            if _dc2.button("🧹 清空全部已標註區段", key='btn_f3_d_clear_zones',
                            use_container_width=True):
                st.session_state['f3_parcel_zones'] = {}
                st.rerun()

            # 暫編地號清單表格（含區段欄）
            _latest_zones = st.session_state.get('f3_parcel_zones', {})
            # 🆕 W-A.3：has_corner 掛在街廓，建快查 dict 供暫編表顯示
            _blk_corner_map = {
                b.get('label', ''): b.get('has_corner', False)
                for b in (classified_blocks or [])
            }
            _df_tp_simple = pd.DataFrame([
                {'原地號': tp['原地號'], '暫編地號': tp['暫編地號'],
                 '所屬街廓': tp['所屬街廓'], '街廓分類': tp['街廓分類'],
                 # 🚨 W-A：分攤登記面積（顯示用）+ 幾何面積（DXF 真值，供驗收對拍）
                 '分攤登記面積(㎡)': round(float(tp.get('分攤登記面積_m2', tp.get('面積_m2', 0)) or 0), 2),
                 '幾何面積(㎡)': round(float(tp.get('幾何面積_m2', tp.get('面積_m2', 0)) or 0), 2),
                 '重劃前地價區段': _latest_zones.get(tp['原地號'], ''),
                 # 🆕 W-A.3 §2/§4 驗收欄
                 'cut_type': tp.get('cut_type', ''),
                 '街廓有街角': _blk_corner_map.get(tp.get('所屬街廓', ''), False)}
                for tp in temp_parcels
            ])
            st.dataframe(_df_tp_simple, use_container_width=True, hide_index=True, height=320)
            # 🆕 W-A.3 驗收用下載按鈕
            st.download_button(
                label="⬇️ 下載暫編地號表（CSV，含 cut_type / 街廓有街角）",
                data=_df_tp_simple.to_csv(index=False, encoding='utf-8-sig'),
                file_name='temp_parcels_wa3.csv',
                mime='text/csv',
                key='btn_download_tp_wa3',
            )

            # 同步更新 temp_parcels 之區段欄位（供下游 Step E/F/G/Tab 5/6/7/8 讀取）
            zones_map = st.session_state.get('f3_parcel_zones', {})
            for tp in temp_parcels:
                tp['重劃前地價區段'] = zones_map.get(tp['原地號'], '')

        # ---------- (已移除舊 Step D fragment / @st.fragment 包裝)
        if False:
            try:
                from streamlit_plotly_events import plotly_events as _pevents_d
                _has_pe_d = True
            except Exception:
                _pevents_d = None
                _has_pe_d = False

            @st.fragment
            def _f3_step_d_fragment():
                _zmap = dict(st.session_state.get('f3_parcel_zones', {}))
                labeled = sum(1 for op in all_original_parcels if _zmap.get(op))
                total = len(all_original_parcels)

                # 工具列：區段代碼、進度、清空
                tc1, tc2, tc3 = st.columns([2, 1, 1])
                current_zone = tc1.text_input(
                    "區段代碼（下一次點選將寫入此值）",
                    value=st.session_state.get('f3_current_zone', '甲段'),
                    key="f3_cur_zone_in"
                )
                st.session_state['f3_current_zone'] = current_zone
                tc2.metric("已標註", f"{labeled}/{total}")
                if tc3.button("🧹 清空全部", key="f3_clear_zones", use_container_width=True):
                    st.session_state['f3_parcel_zones'] = {}
                    st.rerun(scope='app')

                # 取得任務二之高亮地號（若已選）→ 步驟 D 同步顯示
                _hl_carry = st.session_state.get('f3_highlight_temp_no_resolved')
                _hl_zoom_carry = st.session_state.get('f3_highlight_auto_zoom', True)

                # 地圖（一律可點選）
                # 🆕 Task E：傳入 visible_layers（GIS 控制台之圖層選擇）
                _vl_now = st.session_state.get('f3_visible_layers', None)
                if not _has_pe_d:
                    st.warning("⚠️ 缺少 streamlit-plotly-events，僅顯示靜態地圖（無法點選）。")
                    _fig_static = _render_f3_overlay_figure(
                        classified_blocks, temp_parcels, _zmap,
                        interactive=False, height=520,
                        highlight_temp_no=_hl_carry,
                        auto_zoom_to_highlight=_hl_zoom_carry,
                        visible_layers=_vl_now,
                    )
                    st.plotly_chart(_fig_static, use_container_width=True)
                else:
                    st.caption(f"💡 點擊任一封閉多邊形 → 立即標記為「{current_zone}」區段（fragment 優化，只重繪此區塊）。")
                    _fig = _render_f3_overlay_figure(
                        classified_blocks, temp_parcels, _zmap,
                        interactive=True, height=520,
                        highlight_temp_no=_hl_carry,
                        auto_zoom_to_highlight=_hl_zoom_carry,
                        visible_layers=_vl_now,
                    )
                    _zk_cntr = st.session_state.get('f3_zone_click_counter', 0)
                    picked = _pevents_d(_fig, click_event=True, hover_event=False,
                                        key=f"f3_click_zone_{_zk_cntr}", override_height=520)
                    # 🆕 Phase 4 註：上方統一 GIS 互動地圖為主要操作介面，
                    #               此處保留為輔助；移除舊 Phase 3 的 f3_map_mode 守門
                    if picked:
                        ev = picked[0]
                        cn = ev.get('curveNumber')
                        cx_click = ev.get('x')
                        cy_click = ev.get('y')
                        matched_parcel = None
                        valid_parcels = [tp for tp in temp_parcels
                                         if len((tp.get('polygon_coords') or [])) >= 3]
                        # 圖層順序：街廓填色(N)→ 紅線(1)→ 區段分群(1+uniq_zones)→ 點擊靶(N_parcels)
                        # 🆕 Task E：依 visible_layers 動態調整 offset
                        _vl_off = st.session_state.get('f3_visible_layers', {}) or {}
                        _show_blk_off = bool(_vl_off.get('block_boundaries', True))
                        _show_cad_off = bool(_vl_off.get('pre_cadastral', True))
                        n_blocks = len(classified_blocks) if _show_blk_off else 0
                        n_red_line = 1 if _show_blk_off else 0
                        n_uniq_zones = len({v for v in _zmap.values() if v})
                        n_zone_traces = (1 + n_uniq_zones) if _show_cad_off else 0
                        click_offset = n_blocks + n_red_line + n_zone_traces
                        if cn is not None and click_offset <= cn < click_offset + len(valid_parcels):
                            matched_parcel = valid_parcels[cn - click_offset]
                        else:
                            # 備援：x/y + shapely
                            try:
                                from shapely.geometry import Point as _SPoint, Polygon as _SPoly
                                if cx_click is not None and cy_click is not None:
                                    pt = _SPoint(float(cx_click), float(cy_click))
                                    best_area = None
                                    for tp in temp_parcels:
                                        coords = tp.get('polygon_coords') or []
                                        if len(coords) < 3:
                                            continue
                                        try:
                                            poly = _SPoly(coords)
                                            if poly.is_valid and (poly.contains(pt) or poly.distance(pt) < 0.5):
                                                a = float(poly.area)
                                                if best_area is None or a < best_area:
                                                    best_area = a
                                                    matched_parcel = tp
                                        except Exception:
                                            continue
                            except Exception:
                                pass
                        if matched_parcel is not None:
                            _zmap[matched_parcel['原地號']] = current_zone
                            st.session_state['f3_parcel_zones'] = _zmap
                            st.session_state['f3_zone_click_counter'] = _zk_cntr + 1
                            st.toast(f"✅ {matched_parcel['原地號']} → 「{current_zone}」", icon="🎯")
                            st.rerun(scope='fragment')  # 快速：只重繪此 fragment

                # 一鍵填入未標註空白 + 儲存按鈕
                st.markdown("---")
                blank_count = total - labeled
                bf1, bf2, bf3 = st.columns([2, 1, 1])
                bulk_zone = bf1.text_input(
                    f"🪄 將剩餘 {blank_count} 筆未標註地號一鍵填入此區段代碼",
                    value=current_zone, key="f3_bulk_fill_zone"
                )
                if bf2.button(f"🪄 全部填入「{(bulk_zone or '?').strip()}」",
                              use_container_width=True, key="f3_bulk_fill_btn",
                              disabled=(blank_count == 0)):
                    zn = (bulk_zone or '').strip()
                    if not zn:
                        st.warning("請先輸入區段代碼。")
                    else:
                        new_map = dict(_zmap)
                        filled = 0
                        for op in all_original_parcels:
                            if not new_map.get(op):
                                new_map[op] = zn
                                filled += 1
                        st.session_state['f3_parcel_zones'] = new_map
                        st.session_state['f3_current_zone'] = zn
                        st.toast(f"✅ 已一鍵填入 {filled} 筆未標註地號為「{zn}」", icon="🪄")
                        st.rerun(scope='app')
                if bf3.button("💾 儲存並更新全頁",
                              use_container_width=True, key="f3_save_btn",
                              type='primary'):
                    _zone_count = len({v for v in _zmap.values() if v})
                    st.success(f"✅ 已儲存 {len(_zmap)} 筆地號標註，共 {_zone_count} 個重劃前地價區段。"
                               f"後續步驟（Step E/F/G 與下游分析頁）將以此為基礎。")
                    st.rerun(scope='app')

                # 暫編地號清單（直接讀 session_state 取得最新狀態）
                st.markdown("##### 📋 暫編地號清單")
                _latest = st.session_state.get('f3_parcel_zones', {})
                df_tp = pd.DataFrame([
                    {'原地號': tp['原地號'], '暫編地號': tp['暫編地號'],
                     '所屬街廓': tp['所屬街廓'], '街廓分類': tp['街廓分類'],
                     # 🚨 W-A：分攤登記面積（顯示用）+ 幾何面積（DXF 真值，供驗收對拍）
                     '分攤登記面積(㎡)': round(float(tp.get('分攤登記面積_m2', tp.get('面積_m2', 0)) or 0), 2),
                     '幾何面積(㎡)': round(float(tp.get('幾何面積_m2', tp.get('面積_m2', 0)) or 0), 2),
                     '重劃前地價區段': _latest.get(tp['原地號'], '')}
                    for tp in temp_parcels
                ])
                st.dataframe(df_tp, use_container_width=True, hide_index=True, height=320)

            _f3_step_d_fragment()

            # 更新 temp_parcels 對應區段（main 流程使用，下游 Step E/F/G/Tab 5/6/7/8 會讀取）
            zones_map = st.session_state.get('f3_parcel_zones', {})
            for tp in temp_parcels:
                tp['重劃前地價區段'] = zones_map.get(tp['原地號'], '')


        # ---------- 步驟 E：可建築街廓路寬 ----------
        build_blocks = [b for b in classified_blocks if b['burden_type'] == "可建築土地"]
        if build_blocks:
            st.markdown("---")
            st.markdown("#### 🛣️ 步驟 E：可建築街廓路寬、長度與雙側面道路")
            st.caption(
                "路寬長度計算至「臨正街延伸與臨側街延伸之交點」，不計道路截角。僅新闢道路計入臨街地特別負擔。"
                "正面／側面長度由街廓 DXF 最小外接矩形自動帶入；若街廓有兩條側面道路（左+右），可分別設定。"
            )
            # 先按一次以讓使用者可以重設全部
            if st.button("🔄 以 DXF 重新帶入所有街廓長度（覆寫現有值）", key="f3_reset_from_dxf"):
                for b in build_blocks:
                    ext = calc_block_extension_lengths(b.get('vertices', []))
                    cur = st.session_state[SS_ROAD].get(b['id'], {})
                    cur.update({
                        'front_length':       ext['front_length'],
                        'left_side_length':   ext['left_side_length'],
                        'right_side_length':  ext['right_side_length'],
                        '_auto_angle':        ext['angle_deg'],
                    })
                    st.session_state[SS_ROAD][b['id']] = cur
                st.success("✅ 已依 DXF 最小外接矩形自動帶入各街廓長度")
                st.rerun()

            with st.form("f3_road_form", clear_on_submit=False):
                new_roads = {}
                for b in build_blocks:
                    bid = b['id']
                    # DXF 自動計算結果（供預填與顯示）
                    ext = calc_block_extension_lengths(b.get('vertices', []))
                    r = st.session_state[SS_ROAD].get(bid, {})
                    with st.expander(
                        f"🏗️ {b['label']}　面積 {b['area_m2']:,.2f} ㎡",
                        expanded=False,
                    ):
                        # 🆕 Phase 7 Module 2：CAD 精準長度覆寫
                        _cad_fl_map_e = st.session_state.get('f3_cad_front_lengths', {}) or {}
                        _cad_sl_map_e = st.session_state.get('f3_cad_side_lengths', {}) or {}
                        # 🆕 Hotfix：左/右側獨立 map
                        _cad_sl_by_side_e = st.session_state.get('f3_cad_side_lengths_by_side', {}) or {}
                        _b_lbl_e = b['label']
                        _has_cad_fl = _b_lbl_e in _cad_fl_map_e
                        _b_sides = _cad_sl_by_side_e.get(_b_lbl_e, {})
                        _has_cad_left = _b_sides.get('left', 0) > 0
                        _has_cad_right = _b_sides.get('right', 0) > 0
                        _has_cad_sl = _has_cad_left or _has_cad_right or (_b_lbl_e in _cad_sl_map_e)

                        st.markdown("**🔹 正面道路（沿街廓長邊方向）**")
                        ca1, ca2, ca3 = st.columns([1, 1, 1.2])
                        fw = ca1.number_input(
                            "正面路寬 (m)", min_value=0.0,
                            value=float(r.get('front_width', 8.0)), step=0.5,
                            key=f"f3_fw_{bid}"
                        )
                        # 🆕 Phase 7：若有 CAD 直讀正面長度 → 預設值改用 CAD 值 + 加註標記
                        _fl_default = (_cad_fl_map_e[_b_lbl_e] if _has_cad_fl else
                                        float(r.get('front_length', ext['front_length'])))
                        fl = ca2.number_input(
                            f"正面長度 (m){' 📏 CAD' if _has_cad_fl else ''}",
                            min_value=0.0, value=_fl_default,
                            step=0.5, key=f"f3_fl_{bid}",
                            help=("📏 此值由 DXF FRONT_LINE 圖層精準直讀，建議勿手動修改。"
                                  if _has_cad_fl else
                                  f"DXF 自動：{ext['front_length']:.2f} m（延伸交點長度，不計截角）")
                        )
                        fn = ca3.checkbox(
                            "正面為新闢道路", value=bool(r.get('front_new', True)), key=f"f3_fn_{bid}"
                        )

                        st.markdown("**🔹 左側面道路（若該街廓左側無臨路請填 0）**")
                        cl1, cl2, cl3 = st.columns([1, 1, 1.2])
                        # 🆕 Phase 11：預設 8m，使用者可改 0 表示無臨路
                        lw = cl1.number_input(
                            "左側路寬 (m)", min_value=0.0,
                            value=float(r.get('left_side_width', r.get('side_width', 8.0))),
                            step=0.5, key=f"f3_lw_{bid}",
                            help="該街廓左側道路寬度；填 0 表示左側無臨路（如該街廓緊鄰其他街廓而非道路）",
                        )
                        # 🆕 Phase 7 + Hotfix：若有 CAD 直讀左側長度 → 預設值改用 CAD 值
                        if _has_cad_left:
                            _ll_default = float(_b_sides.get('left', 0.0))
                        elif _has_cad_sl and not _has_cad_right:
                            # 向下相容：若舊版 side_lengths（單值）有資料但未分左右 → 視為左側
                            _ll_default = float(_cad_sl_map_e.get(_b_lbl_e, 0.0))
                        else:
                            _ll_default = float(r.get('left_side_length',
                                                       r.get('side_length',
                                                             ext['left_side_length'])))
                        ll = cl2.number_input(
                            f"左側長度 (m){' 📏 CAD' if (_has_cad_left or (_has_cad_sl and not _has_cad_right)) else ''}",
                            min_value=0.0, value=_ll_default,
                            step=0.5, key=f"f3_ll_{bid}",
                            help=("📏 此值由 DXF SIDE_LINE 圖層精準直讀。"
                                  if (_has_cad_left or _has_cad_sl) else
                                  f"DXF 自動：{ext['left_side_length']:.2f} m")
                        )
                        ln = cl3.checkbox(
                            "左側為新闢道路",
                            value=bool(r.get('left_side_new', r.get('side_new', False))),
                            key=f"f3_ln_{bid}"
                        )

                        st.markdown("**🔹 右側面道路（若該街廓右側無臨路請填 0）**")
                        cr1, cr2, cr3 = st.columns([1, 1, 1.2])
                        # 🆕 Phase 11：預設 8m，使用者可改 0 表示無臨路
                        rw = cr1.number_input(
                            "右側路寬 (m)", min_value=0.0,
                            value=float(r.get('right_side_width', 8.0)), step=0.5,
                            key=f"f3_rw_{bid}",
                            help="該街廓右側道路寬度；填 0 表示右側無臨路（如該街廓緊鄰其他街廓而非道路）",
                        )
                        # 🆕 Hotfix：右側長度也支援 CAD 直讀
                        _rl_default = (float(_b_sides.get('right', 0.0)) if _has_cad_right
                                        else float(r.get('right_side_length',
                                                           ext['right_side_length'])))
                        rl = cr2.number_input(
                            f"右側長度 (m){' 📏 CAD' if _has_cad_right else ''}",
                            min_value=0.0, value=_rl_default,
                            step=0.5, key=f"f3_rl_{bid}",
                            help=("📏 此值由 DXF SIDE_LINE 圖層精準直讀。"
                                  if _has_cad_right else
                                  f"DXF 自動：{ext['right_side_length']:.2f} m")
                        )
                        rn = cr3.checkbox(
                            "右側為新闢道路", value=bool(r.get('right_side_new', False)),
                            key=f"f3_rn_{bid}"
                        )

                        new_roads[bid] = {
                            'front_width': fw, 'front_length': fl, 'front_new': fn,
                            'left_side_width': lw, 'left_side_length': ll, 'left_side_new': ln,
                            'right_side_width': rw, 'right_side_length': rl, 'right_side_new': rn,
                            # 相容舊欄位（供舊程式碼讀取）
                            'side_width': lw, 'side_length': ll, 'side_new': ln,
                            '_auto_angle': ext['angle_deg'],
                        }
                submit_road = st.form_submit_button("✅ 儲存路寬資料", use_container_width=True)
            if submit_road:
                st.session_state[SS_ROAD].update(new_roads)
                st.success("路寬資料已儲存（正面 + 左側 + 右側三組）")
                st.rerun()

        # ---------- 步驟 F：全區參數 ----------
        st.markdown("---")
        st.markdown("#### 💰 步驟 F：全區參數")

        # ---- 一鍵匯入 Tab 4（土地開發分析法）與 Tab 5（預期開發分析法）之平均地價 ----
        _post_from_t4 = float(st.session_state.get('weighted_price_sqm', 0.0) or 0.0)
        _pre_from_t5  = float(st.session_state.get('pre_land_price_sqm', 0.0) or 0.0)
        qi_c1, qi_c2, qi_c3 = st.columns([1, 1, 2])
        qi_c1.metric("土地開發分析法 重劃後平均地價", f"{_post_from_t4:,.0f} 元/㎡" if _post_from_t4 > 0 else "（尚未計算）")
        qi_c2.metric("預期開發分析法 重劃前平均地價", f"{_pre_from_t5:,.0f} 元/㎡"  if _pre_from_t5  > 0 else "（尚未計算）")
        if qi_c3.button(
            "🔄 一鍵匯入 土地開發分析法/預期開發分析法 計算之重劃前後平均地價",
            use_container_width=True,
            disabled=(_post_from_t4 <= 0 and _pre_from_t5 <= 0),
            help="自動將 土地開發分析法 的加權重劃後平均地價與 預期開發分析法 的重劃前平均地價帶入下方欄位",
        ):
            st.session_state['f3_price_before'] = _pre_from_t5
            st.session_state['f3_price_after']  = _post_from_t4
            st.success(f"✅ 已匯入：重劃前 {_pre_from_t5:,.0f} 元/㎡、重劃後 {_post_from_t4:,.0f} 元/㎡")
            st.rerun()

        with st.form("f3_global_form", clear_on_submit=False):
            gf1, gf2, gf3 = st.columns(3)
            total_area_in = gf1.number_input(
                "重劃區總面積（㎡）", min_value=0.0,
                value=float(st.session_state.get('f3_total_area', block_data.get('outer_area_m2', 0.0) or 0.0)),
                step=100.0, help="預設為 DXF 外框多邊形面積"
            )
            preexist_public_in = gf2.number_input(
                "重劃前已徵收取得公設面積（㎡）", min_value=0.0,
                value=float(st.session_state.get('f3_preexisting_public', 0.0)), step=10.0,
                help="B 值分母中 (總面積 − 此值)"
            )
            offset_in = gf3.number_input(
                "抵充地面積（㎡）", min_value=0.0,
                value=float(st.session_state.get('f3_offset_area', 0.0)), step=10.0
            )
            gf4, gf5 = st.columns(2)
            price_before_in = gf4.number_input(
                "重劃前平均地價（元/㎡）", min_value=0.0,
                value=float(st.session_state.get('f3_price_before', 0.0)), step=100.0
            )
            price_after_in = gf5.number_input(
                "重劃後平均地價（元/㎡）", min_value=0.0,
                value=float(st.session_state.get('f3_price_after', 0.0)), step=100.0
            )
            submit_gl = st.form_submit_button("✅ 儲存全區參數", use_container_width=True)
        if submit_gl:
            st.session_state['f3_total_area'] = total_area_in
            st.session_state['f3_preexisting_public'] = preexist_public_in
            st.session_state['f3_offset_area'] = offset_in
            st.session_state['f3_price_before'] = price_before_in
            st.session_state['f3_price_after'] = price_after_in
            st.success("全區參數已儲存")
            st.rerun()

        # ---------- 步驟 H：負擔計算（B 值 / 臨街地特別負擔） ----------
        st.markdown("---")
        st.markdown("#### 📐 步驟 H：臨街地特別負擔 & B 值計算")

        C_value = st.session_state.get('f3_C_from_finance', None)
        if C_value is None:
            st.warning("⚠️ 尚未取得 C 值。請於「💰 財務分析」輸入工程費用、重劃費用、貸款利息後自動回傳；目前以 C=0 試算。")
            C_for_calc = 0.0
        else:
            st.info(f"✅ 已讀取 財務分析 之 C 值 = {C_value:.6f}")
            C_for_calc = float(C_value)

        road_data = []
        for b in build_blocks:
            r = st.session_state[SS_ROAD].get(b['id'], {})
            road_data.append({
                'label': b['label'],
                'front_width':       r.get('front_width', 0.0),
                'front_length':      r.get('front_length', 0.0),
                'front_new':         r.get('front_new', False),
                # 新欄位：左側 + 右側
                'left_side_width':   r.get('left_side_width',  r.get('side_width',  0.0)),
                'left_side_length':  r.get('left_side_length', r.get('side_length', 0.0)),
                'left_side_new':     r.get('left_side_new',    r.get('side_new',    False)),
                'right_side_width':  r.get('right_side_width', 0.0),
                'right_side_length': r.get('right_side_length', 0.0),
                'right_side_new':    r.get('right_side_new', False),
            })
        sb = calc_special_burden_total(road_data, C_for_calc)
        # W-G G.1：純加性曝值（供「七級調配」區 live 組 ctx；harvest 保留 main() 但不呼叫→零 diff）
        st.session_state['f3_sb_rows'] = sb['rows']

        if sb['rows']:
            st.markdown("##### 📋 各街廓臨街地負擔計算")
            st.dataframe(pd.DataFrame(sb['rows']), use_container_width=True, hide_index=True)

        offset_area = float(st.session_state.get('f3_offset_area', 0.0))
        preexisting_public = float(st.session_state.get('f3_preexisting_public', 0.0))
        total_area = float(st.session_state.get('f3_total_area', 0.0))
        price_before = float(st.session_state.get('f3_price_before', 0.0))
        price_after = float(st.session_state.get('f3_price_after', 0.0))

        general_burden = max(public_common_total - offset_area - sb['special_total'], 0.0)
        B_value = calc_B_value(
            general_burden_area=general_burden,
            price_before=price_before, price_after=price_after,
            total_area=total_area, preexisting_public_area=preexisting_public,
        )

        rh1, rh2, rh3 = st.columns(3)
        rh1.metric("正面道路負擔面積 (㎡)", f"{sb['front_total']:,.2f}")
        rh2.metric("側面道路負擔面積 (㎡)", f"{sb['side_total']:,.2f}")
        rh3.metric("臨街地特別負擔總面積 (㎡)", f"{sb['special_total']:,.2f}")
        rj1, rj2, rj3 = st.columns(3)
        rj1.metric("公設共同負擔總面積 (㎡)", f"{public_common_total:,.2f}")
        rj2.metric("一般負擔總面積 (㎡)", f"{general_burden:,.2f}")
        rj3.metric("B 值（一般負擔係數）", f"{B_value:.6f}")

        with st.expander("🔎 公式說明", expanded=False):
            st.markdown("""
**臨街地特別負擔總面積** = (Σ正面道路負擔面積 + Σ側面道路負擔面積) × (1 − C)
- 正面道路負擔面積 = 正街負擔尺度 × 正面長度（僅新闢道路）
- 側面道路負擔面積 = 側街負擔尺度 × 側面長度（僅新闢道路）
- 正街負擔尺度：≤4M → 0；4<路寬<8 → (w−4)/2；8≤路寬<20 → w/4；≥20 → 5
- 側街負擔尺度 = 正街尺度 × 1/2

**一般負擔總面積** = 公設共同負擔總面積 − 抵充地 − 臨街地特別負擔總面積

**B 值** = (一般負擔總面積 × 重劃前平均地價) ÷ (重劃後平均地價 × (重劃區總面積 − 重劃前已徵收取得公設面積))
""")

        # ---------- 步驟 G：宗地分配計算（完整 G 值迭代版） ----------
        if temp_parcels:
            st.markdown("---")
            st.markdown("#### 🧭 步驟 G：宗地分配計算（完整迭代版）")
            st.markdown("""
**公式**：G = [a·(1 − A·B) − Rw·F·l₁ − S·l₂] × (1 − C)

- **a**：1 筆土地（未歸戶）坐落於該街廓面積
- **A**：地價比 = 重劃後單價 / 重劃前單價（自動由 預期開發分析法 同街廓／同重劃前區段單價查得）
- **B**：一般負擔係數（街廓互動分析）　**C**：費用負擔係數（財務分析）
- **l₁ = 側面負擔尺度**（配 Rw·F·l₁）；**l₂ = 正面負擔尺度**（配 S·l₂）
- **F**：街角第一筆土地面臨側面道路之長度＝勾選「街角側別」後，自動帶入步驟 E 該街廓之
  「左側長度」或「右側長度」（左右定義：站在街廓內面向正面道路，左手邊為左側、右手邊為右側）
- **W**：宗地側街臨街線實際長度之中點，向宗地分配線作垂直線所量之距離。
  **按「🔍 自動計算 W」**可由「街廓多邊形 + 宗地多邊形 + 街廓分配線（若未指定，以街廓 MBR 長邊方向替代）」自動幾何推導；亦可直接在表格手動輸入。
- **Rw（%）**：由 104 版作業手冊 P.103 累積表（W=1→17.4、W=2→24.4、…、W=18 以上→100）線性內插。
- **街廓平均深度 = 街廓面積 ÷ 正面臨接道路長度**
- **迭代初值 S₀ = a·(1 − 財務分析 總負擔比率) / 街廓平均深度**
- 收斂條件：|G_k − G_(k-1)| < 0.005 且 |S_(k+1) − S_k| < 0.005（皆取小數點後 2 位）

⚠️ **關於圖面點選街角地**：下方「🖱️ 圖面點選街角地」expander 提供 fragment 優化之地圖點選器，
點擊地號可立即切換街角地狀態（紅色填色＝已標街角；藍色框＝非街角）。標記後再回到下方表格選擇「街角側別」。
若不想用圖選，仍可直接在下表勾選「街角地」欄位。
""")

            # ---- 參數來源自動收集 ----
            # 1. Tab 6 總負擔比率（Tab 7 回送）
            _tab6_burden = st.session_state.get('f3_total_burden_rate_from_finance', None)
            if _tab6_burden is None:
                _tab6_burden = float(B_value) + float(C_for_calc)
                # ⚠️ B+C 僅為迭代初值的粗略估計；非「地主實際平均負擔比率」之正確公式
                # 正確公式請參考 Step G 完成後說明欄與 Excel 報表 Sheet 3
                _burden_note = (f"（以 B+C = {_tab6_burden:.4f} **粗估初值**，"
                                "財務分析 更新後將自動覆寫；正確平均負擔比率請見 Step G 結果說明）")
            else:
                _burden_note = f"（來自 財務分析：{float(_tab6_burden):.4f}）"
            st.caption(f"📌 初值用 財務分析 總負擔比率 = **{float(_tab6_burden):.4f}** {_burden_note}")

            # 2. 各街廓 l_front/l_side_L/l_side_R/正面長度/面積/左右側長度
            sb_rows_by_label = {r['街廓']: r for r in (sb.get('rows') or [])}
            # 🐛 Phase 11 Hotfix：補建 shapely polygon — solve_G_binary 需要這個
            # （否則會 fallback 至代數迭代，導致 cut_coords 永遠為空、重劃後地籍圖無法繪製）
            from shapely.geometry import Polygon as _SP_block
            block_meta_by_label = {}
            for b in classified_blocks:
                _meta = dict(b)
                if _meta.get('shapely') is None:
                    _verts = _meta.get('vertices', []) or []
                    if len(_verts) >= 3:
                        try:
                            _sp = _SP_block(_verts)
                            if not _sp.is_valid:
                                _sp = _sp.buffer(0)
                            _meta['shapely'] = _sp
                        except Exception:
                            _meta['shapely'] = None
                block_meta_by_label[b['label']] = _meta

            # 3. Tab 5 每街廓重劃後單價、每重劃前區段重劃前單價
            post_block_results = st.session_state.get('post_block_results') or []
            pre_zone_results = st.session_state.get('pre_zone_results') or []
            post_price_by_block = {r['街廓']: float(r.get('單價(元/㎡)', 0) or 0) for r in post_block_results}
            pre_price_by_zone = {r['區段']: float(r.get('單價(元/㎡)', 0) or 0) for r in pre_zone_results}

            # 🆕 Phase 11：以下兩項手動 GIS 操作已停用（使用者直接於 DXF BASELINE
            # 與 CENTERLINE 圖層繪製，由 parse_cad_precision_layers 自動讀入）：
            #   - 「⚙️ 手動分配基準線設定」expander
            #   - 「📐 設定分配基準線（點擊邊界）」、「🛣️ 生成精準中心線」操作模式

            build_parcels = [tp for tp in temp_parcels
                             if F3_CATEGORY_BURDEN.get(tp['街廓分類'], '') == '可建築土地'
                             # 🆕 V13 修正 #1：排除已整併之公設地（仍保留供視覺化）
                             and not tp.get('_merged_into_g', False)]
            # W-G G.1：純加性曝值（供「七級調配」區 live 組 ctx）
            st.session_state['f3_build_parcels'] = build_parcels

            if not build_parcels:
                st.info("目前沒有可建築街廓之暫編地號，無法進行 G 值迭代。")
            else:
                import pandas as _pd
                _param_key = 'f3_g_iter_params'
                _prev_params = st.session_state.get(_param_key, {})

                # ---------- 步驟 L：街角地參數 + 第 1 宗 PK（Step G 前置 / Phase A 已前移） ----------
                st.markdown("---")
                st.markdown("#### 📍 步驟 L：街角地參數設定 + 第 1 宗街角地優先權 PK（Step G 前置）")
                st.markdown("""
🆕 **本步驟已前移至 Step G 之前**（Phase A 重構），Step G 迭代會自動採用本步驟 PK 結果。

依《市地重劃作業手冊》P.182-184 + 《花蓮縣畸零地使用規則》：
- **街角地最小分配寬度** = 法定最小建築寬度 + 退縮距離
- **街角地最小分配面積** = (街角地最小分配寬度 × 街廓分配深度) − 截角面積
  （截角面積由步驟 A 之街廓 DXF 自動偵測；可手動微調）
- **資格條件**：暫編地號 polygon 與「街角地最小面積範圍」幾何相交即可（任意大小）
- **第 1 宗優先權指數** = (臨正街長度/基準) × 0.4 + (臨側街長度/基準) × 0.2 + (跨占街角面積/基準) × 0.4
  - 全部候選 G 值都未達門檻 → 街角地第 1 宗強制留設為**抵費地**（forced_offset）
""")
                with st.expander("⚙️ 街角地參數（共用設定 + 各街廓微調）", expanded=False):
                    # 🚨 Phase 9.12 Issue 4：UI 雙向綁定 + on_change callback 清除 G 值 cache
                    def _f3L_invalidate_g_cache():
                        """退縮距離 / 分配深度變更時，清空 G 值結果並標記需重算"""
                        try:
                            _st_inv = st
                            _st_inv.session_state.pop('f3_G_values', None)
                            _st_inv.session_state.pop('f3_G_trace', None)
                            _st_inv.session_state.pop('f3_corner_winners', None)
                            _st_inv.session_state.pop('f3L_corner_winners', None)
                            _st_inv.session_state['f3_g_needs_rerun'] = True
                        except Exception:
                            pass
                    _setback_init = float(st.session_state.get('f3L_setback_default', 3.5))
                    _setback_default = st.number_input(
                        "街角地退縮距離（m，全區共用預設）",
                        min_value=0.0, max_value=10.0,
                        value=_setback_init, step=0.1,
                        key='f3L_setback_default',
                        on_change=_f3L_invalidate_g_cache,
                        help='💡 修改後 Step G 迭代將自動採用最新 PK 結果')
                    if st.session_state.get('f3_g_needs_rerun'):
                        st.info(
                            "ℹ️ **參數已修改 → 點擊下方按鈕重新執行街角地 PK，"
                            "再至 Step G 點擊「執行 G 值迭代計算」即可採用最新 PK 結果**"
                        )
                    # 🆕 W-C §5a：街廓分配深度沿 ALLOC_LINE 量（方法 A 剖面差值），廢 area/front_len
                    #   與「全域單值套全街廓」；改逐街廓 D_avg_i + 逐街廓選填覆寫（留空0=auto）。
                    _build_blocks = [b for b in classified_blocks
                                     if F3_CATEGORY_BURDEN.get(b.get('category', ''), '') == '可建築土地']
                    import math as _math_dep
                    _depth_info_by_blk = {}; _depth_use_by_blk = {}; _min_alloc_area_by_blk = {}
                    _min_width_by_blk = {}
                    _adir_dep = st.session_state.get('f3_cad_alloc_dir', {}) or {}
                    _fl_dep = st.session_state.get('f3_cad_front_lines', {}) or {}
                    _mbl_dep = st.session_state.get('f3_manual_baseline', {}) or {}
                    for b in _build_blocks:
                        _lbl = b['label']
                        _sb_d = sb_rows_by_label.get(_lbl, {})
                        _fl_len_d = float(_sb_d.get('正面長度(m)', 0.0) or 0.0)
                        _area_d = float(b.get('area_m2', 0.0) or 0.0)
                        _ad_d = _adir_dep.get(_lbl)
                        _fpts = None
                        _flb = _fl_dep.get(_lbl) or {}
                        if _flb.get('p1') and _flb.get('p2'):
                            _fpts = [_flb['p1'], _flb['p2']]
                        _bpts = None
                        _mb = _mbl_dep.get(_lbl) or {}
                        if _mb.get('point') and _mb.get('angle_deg') is not None:
                            _bx, _by = _mb['point']; _ang = _math_dep.radians(float(_mb['angle_deg']))
                            # 🆕 §5a-1：合成 BASELINE 裁到街廓 bbox 對角線（非 ±1000m），
                            #   否則近 90° 的 BASELINE 會讓方法 A depth(w) 掃到 1000（R2 D_avg=500 之因）。
                            _vb = b.get('vertices') or []
                            if _vb:
                                _xs = [v[0] for v in _vb]; _ys = [v[1] for v in _vb]
                                _Lb = (((max(_xs) - min(_xs)) ** 2 + (max(_ys) - min(_ys)) ** 2) ** 0.5) or 200.0
                            else:
                                _Lb = 200.0
                            _bpts = [(_bx - _Lb * _math_dep.cos(_ang), _by - _Lb * _math_dep.sin(_ang)),
                                     (_bx + _Lb * _math_dep.cos(_ang), _by + _Lb * _math_dep.sin(_ang))]
                        _dinfo = _compute_block_depth_alloc(b.get('vertices') or [], _area_d, _ad_d,
                                                         front_pts=_fpts, baseline_pts=_bpts)
                        if _dinfo is None:
                            # 缺 ALLOC_LINE → 退回 area/front_len（並於 §7/診斷標示）
                            _dval = (_area_d / _fl_len_d) if _fl_len_d > 0 else 0.0
                            _dinfo = {'D_avg': round(_dval, 2), 'D_min': round(_dval, 2),
                                   'D_max': round(_dval, 2), 'D_avg_B': round(_dval, 2),
                                   'method': '⚠️ 缺 ALLOC → area/front_len', 'note': '缺 f3_cad_alloc_dir'}
                        _depth_info_by_blk[_lbl] = _dinfo
                    # 逐街廓選填覆寫（留空0=用 D_avg_i；移除舊「單值套全街廓」）
                    with st.expander("📏 街廓分配深度（沿 ALLOC_LINE 自動量 D_avg；可逐街廓選填覆寫）", expanded=False):
                        st.caption("留空(0)=自動沿 ALLOC_LINE 量之 D_avg；填值=覆寫該街廓。"
                                   "D_min/D_max 僅診斷（差很大→街廓畫歪）。")
                        for b in _build_blocks:
                            _lbl = b['label']; _dinfo = _depth_info_by_blk[_lbl]
                            _ov = st.number_input(
                                f"{_lbl} 深度覆寫（m）", min_value=0.0, max_value=300.0,
                                value=float(st.session_state.get(f'f3L_depth_ov_{_lbl}', 0.0)),
                                step=0.5, key=f'f3L_depth_ov_{_lbl}',
                                on_change=_f3L_invalidate_g_cache,
                                help=(f"自動 D_avg={_dinfo['D_avg']}m（{_dinfo['method']}）；"
                                      f"D_min={_dinfo['D_min']}/D_max={_dinfo['D_max']}；{_dinfo['note']}"))
                            _depth_use_by_blk[_lbl] = float(_ov) if _ov > 0 else float(_dinfo['D_avg'])
                            # min_alloc_area_i = D_avg_i × min_width_i（逐 Ri 查正面路寬）
                            _fw_d = float(sb_rows_by_label.get(_lbl, {}).get('正面路寬(m)', 0.0) or 0.0)
                            _mw_d = float(get_min_lot_size(b.get('category', ''), _fw_d).get('min_width', 0.0) or 0.0)
                            _min_width_by_blk[_lbl] = _mw_d
                            _min_alloc_area_by_blk[_lbl] = (round(_depth_use_by_blk[_lbl] * _mw_d, 2)
                                                            if _mw_d > 0 else None)
                    st.session_state['f3_block_depth_by_label'] = _depth_info_by_blk
                    st.session_state['f3_alloc_depth_by_label'] = _depth_use_by_blk
                    st.session_state['f3_min_alloc_area_by_label'] = _min_alloc_area_by_blk
                    st.session_state['f3_min_width_by_label'] = _min_width_by_blk
                    # 🆕 W-C §5 / 7-0a 前置地基檢查：region_min vs 重劃區總面積×C%（v3.1 §7-0a）
                    #   region_min = min(各可建築街廓 D_avg_i×min_width_i)（非街角範圍面積）。
                    _valid_mins = [v for v in _min_alloc_area_by_blk.values() if v is not None and v > 0]
                    _region_min = min(_valid_mins) if _valid_mins else None
                    _total_area_70a = sum(float(_b.get('area_m2', 0.0) or 0.0)
                                          for _b in classified_blocks)
                    _C_70a = float(C_for_calc or 0.0)
                    _pool_70a = _total_area_70a * _C_70a
                    st.session_state['f3_70a'] = {
                        'region_min': _region_min, 'total_area': round(_total_area_70a, 2),
                        'C': _C_70a, 'pool': round(_pool_70a, 2),
                        'argmin_blk': (min(((k, v) for k, v in _min_alloc_area_by_blk.items()
                                            if v is not None and v > 0),
                                           key=lambda kv: kv[1])[0] if _valid_mins else None),
                        'pass': (_region_min is None) or (_C_70a <= 0) or (_region_min <= _pool_70a),
                    }
                    if _region_min is not None and _C_70a > 0 and _region_min > _pool_70a:
                        st.error(
                            f"🔴 §7-0a 前置地基檢查未過：全區最小分配面積 {_region_min:.2f}㎡ "
                            f"> 重劃區總面積×C% = {_total_area_70a:,.0f}×{_C_70a:.4f} "
                            f"= {_pool_70a:.2f}㎡。**分配街廓需重新劃設**，請勿續跑配地。"
                        )
                    elif _region_min is not None and _C_70a > 0:
                        st.caption(
                            f"✅ §7-0a 地基檢查 PASS：全區最小分配面積 {_region_min:.2f}㎡（最淺乘積街廓 "
                            f"{st.session_state['f3_70a']['argmin_blk']}）≤ 重劃區總面積×C% "
                            f"{_pool_70a:,.2f}㎡（{_total_area_70a:,.0f}×{_C_70a:.4f}）"
                        )
                    # 各街廓建立 corner_min_area 表
                    _corner_rows_init = []
                    _anchor_diag = []   # 🆕 W-D.1.3-a 診斷：截角拓樸定錨（已切換 production；閘只認截角、front_idx 資訊）
                    _burden_18m_by_blk = {}   # 🆕 W-C §2：負擔範圍 W=18m 面積（每側）
                    for b in _build_blocks:
                        _lbl = b['label']
                        _sb_row = sb_rows_by_label.get(_lbl, {})
                        _fw = float(_sb_row.get('正面路寬(m)', 0.0) or 0.0)
                        _fl = float(_sb_row.get('正面長度(m)', 0.0) or 0.0)
                        _area = float(b.get('area_m2', 0.0) or 0.0)
                        _depth_use = _depth_use_by_blk.get(_lbl, (_area / _fl) if _fl > 0 else 0.0)
                        _cutoff_total = float(b.get('cutoff_total_area_m2', 0.0) or 0.0)
                        # 法定最小寬/深（僅供顯示）
                        _size_info = get_min_lot_size(b.get('category', ''), _fw)
                        _legal_w_disp = float(_size_info.get('min_width', 0.0) or 0.0)
                        _legal_d_disp = float(_size_info.get('min_depth', 0.0) or 0.0)
                        _lw_side = float(_sb_row.get('左側路寬(m)', 0.0) or 0.0)
                        _rw_side = float(_sb_row.get('右側路寬(m)', 0.0) or 0.0)
                        # 🚨 W-B 修正：街角有無改用 SIDE_LINE 判定（W-A.3 corner_sides），
                        #   不再用路寬>0.5（路寬兩側皆有路時會誤判幽靈側別、誤跑 PK）。
                        _corner_sides_for_blk = list(b.get('corner_sides', []) or [])
                        if not _corner_sides_for_blk:
                            # 防呆：corner_sides 於 rerun 重建為空時，回退至 W-A.3 同源 session
                            _sbs_sess = st.session_state.get(
                                'f3_cad_side_lengths_by_side', {}) or {}
                            _sides_sess = _sbs_sess.get(_lbl, {}) or {}
                            try:
                                if float(_sides_sess.get('left', 0) or 0) > 0:
                                    _corner_sides_for_blk.append('left')
                                if float(_sides_sess.get('right', 0) or 0) > 0:
                                    _corner_sides_for_blk.append('right')
                            except Exception:
                                pass
                        has_left  = 'left'  in _corner_sides_for_blk
                        has_right = 'right' in _corner_sides_for_blk
                        # 🆕 W-D.1.3-a 診斷（已切換 production=拓樸）：舊截角idx＝現行 geom_restore(已拓樸)、
                        #   新截角idx＝即時 _anchor → 應全等(✅ 確認切換生效)；閘只認截角，front_idx 為資訊性(不 gate)。
                        try:
                            _gr_a = b.get('geom_restore') or {}
                            _cls_a = _gr_a.get('classification') or {}
                            _edges_a = _cls_a.get('edges') or []
                            _old_fi = _cls_a.get('front_idx')
                            _old_ch = sorted({int(tc['cutoff_idx'])
                                              for tc in (_gr_a.get('theoretical_corners') or [])
                                              if tc.get('cutoff_idx') is not None})
                            _fl_a = (st.session_state.get('f3_cad_front_lines', {}) or {}).get(_lbl)
                            _sls_a = (st.session_state.get(
                                'f3_cad_side_lines_by_side', {}) or {}).get(_lbl, {})
                            _topo_a = _anchor_chamfers_topology(_edges_a, _fl_a, _sls_a)
                            _new_fi = _topo_a.get('front_idx')
                            _new_ch = sorted({ci for _s in ('left', 'right')
                                              for ci in ((_topo_a['sides'].get(_s) or {}).get('chamfer_idxs') or [])})
                            _fi_ok = (_old_fi == _new_fi)
                            _ch_ok = (_old_ch == _new_ch)
                            _anchor_diag.append({
                                '街廓': _lbl,
                                '舊front_idx': _old_fi, '新front_idx': _new_fi,
                                'front(資訊)': ('同' if _fi_ok else '異·改FRONT共線邊(不 gate)'),
                                '舊截角idx': str(_old_ch), '新截角idx': str(_new_ch),
                                '截角': ('✅' if _ch_ok else '🔴'),
                                '左': ((_topo_a['sides'].get('left') or {}).get('status', '—')
                                       if has_left else '無此側'),
                                '右': ((_topo_a['sides'].get('right') or {}).get('status', '—')
                                       if has_right else '無此側'),
                                '閘(截角)': ('✅ 綠（截角idx同→gate鏈+live幾何量不變）' if _ch_ok
                                            else '🔴 截角idx異→停查（切換引入 bug 或碰 903 副本陷阱）'),
                            })
                        except Exception as _e_anc:
                            _anchor_diag.append({'街廓': _lbl, '閘(截角)': f'⚠️ 診斷例外：{_e_anc}'})
                        _cad_fl_for_split = (st.session_state.get(
                            'f3_cad_front_lines', {}) or {}).get(_lbl, {})
                        _fl_p1_split = _cad_fl_for_split.get('p1') if _cad_fl_for_split else None
                        _fl_p2_split = _cad_fl_for_split.get('p2') if _cad_fl_for_split else None
                        if _fl_p1_split and _fl_p2_split:
                            _per_end = _compute_per_end_cutoff_areas(b, _fl_p1_split, _fl_p2_split)
                            _cutoff_left = float(_per_end['p1_end']['cutoff_area'])
                            _cutoff_right = float(_per_end['p2_end']['cutoff_area'])
                        else:
                            _cutoff_left, _cutoff_right = _split_cutoffs_by_side(b)
                        # 幾何街角規定範圍面積（W-B _build_corner_range_v2，取代矩形近似）
                        _adir_cr = st.session_state.get('f3_cad_alloc_dir', {}) or {}
                        _slbs_cr = st.session_state.get('f3_cad_side_lines_by_side', {}) or {}
                        _alloc_cr = _adir_cr.get(_lbl)
                        _side_cr = _slbs_cr.get(_lbl, {})
                        _blk_verts_cr = b.get('vertices') or []
                        _blk_cen_cr = b.get('centroid') or (0.0, 0.0)
                        _shift_cr = _setback_default + _legal_w_disp
                        _left_min = None; _right_min = None
                        if _alloc_cr and len(_blk_verts_cr) >= 3:
                            if has_left and 'left' in _side_cr:
                                _chi_l = _make_chamfer_tri_wb(b, 'left')
                                _rng_l = _build_corner_range_v2(
                                    _side_cr['left']['mid'], _blk_verts_cr, _blk_cen_cr,
                                    _alloc_cr, _shift_cr, _chi_l)
                                _left_min = (round(float(_rng_l.area), 2)
                                             if _rng_l is not None else None)
                            if has_right and 'right' in _side_cr:
                                _chi_r = _make_chamfer_tri_wb(b, 'right')
                                _rng_r = _build_corner_range_v2(
                                    _side_cr['right']['mid'], _blk_verts_cr, _blk_cen_cr,
                                    _alloc_cr, _shift_cr, _chi_r)
                                _right_min = (round(float(_rng_r.area), 2)
                                              if _rng_r is not None else None)
                        # 🆕 W-C §2：負擔範圍 W=18m（shift=18、不扣截角；v3.1 §5）
                        _burden_l = _burden_r = None
                        if _alloc_cr and len(_blk_verts_cr) >= 3:
                            if has_left and 'left' in _side_cr:
                                _br_l = _build_corner_range_v2(
                                    _side_cr['left']['mid'], _blk_verts_cr, _blk_cen_cr,
                                    _alloc_cr, 18.0, None)
                                _burden_l = round(float(_br_l.area), 2) if _br_l is not None else None
                            if has_right and 'right' in _side_cr:
                                _br_r = _build_corner_range_v2(
                                    _side_cr['right']['mid'], _blk_verts_cr, _blk_cen_cr,
                                    _alloc_cr, 18.0, None)
                                _burden_r = round(float(_br_r.area), 2) if _br_r is not None else None
                        _burden_18m_by_blk[_lbl] = {'left': _burden_l, 'right': _burden_r}
                        _corner_rows_init.append({
                            '街廓': _lbl,
                            '分類': b.get('category', ''),
                            '正面路寬(m)': round(_fw, 2),
                            '【左】路寬(m)': (round(_lw_side, 2) if has_left else '—'),
                            '【右】路寬(m)': (round(_rw_side, 2) if has_right else '—'),
                            '街廓分配深度(m)': round(_depth_use, 2),
                            '法定最小寬(m)': _legal_w_disp,
                            '法定最小深(m)': _legal_d_disp,
                            '【左】截角(㎡)': (round(_cutoff_left, 2) if has_left else '—'),
                            '【右】截角(㎡)': (round(_cutoff_right, 2) if has_right else '—'),
                            '【左】街角最小面積(㎡)': _left_min,
                            '【右】街角最小面積(㎡)': _right_min,
                        })
                    st.session_state['f3_burden_range_18m'] = _burden_18m_by_blk
                    if _corner_rows_init:
                        _df_corner = _pd.DataFrame([
                            {k: v for k, v in r.items() if not k.startswith('_')}
                            for r in _corner_rows_init
                        ])
                        st.dataframe(_df_corner, use_container_width=True, hide_index=True)
                        st.session_state['f3L_corner_min_table'] = _corner_rows_init
                    else:
                        st.info("無可建築土地街廓 → 不需計算街角地最小分配面積")

                    # 🆕 W-D.1.3-a 診斷：截角拓樸定錨（已切換 production）；閘只認截角、front_idx 資訊性
                    if _anchor_diag:
                        with st.expander(
                            "🔬 W-D.1.3-a 診斷：截角拓樸定錨（已切換 production；閘只認截角、front_idx 資訊性）",
                            expanded=False
                        ):
                            st.dataframe(_pd.DataFrame(_anchor_diag),
                                         use_container_width=True, hide_index=True)
                            st.caption(
                                "💡 **已切換**：production `theoretical_corners` 已由 CAD-import post-pass 以拓樸定錨重建。"
                                "`舊截角idx`＝現行 geom_restore（已拓樸）、`新截角idx`＝即時 `_anchor_chamfers_topology` → **應全等（✅，確認切換生效）**。\n\n"
                                "**閘只認截角 + live 幾何量**（截角idx／深度／街角最小面積／winner）；`front(資訊)` 欄 `舊`＝classify 最長邊、`新`＝FRONT 共線邊，"
                                "**5 塊異為預期（狹長街廓長度法標到屁股邊）、vestigial 不驅動 production、不 gate**。\n\n"
                                "🔴 **截角欄任一異** ＝ 切換引入 bug 或碰 903 副本陷阱 → **停、交 KL**（D-2）。KL 另對照切換前後：街角最小面積／winner／優先權指數應逐項相同（僅 front_idx 欄改）。"
                            )

                # 優先權指數選位（按下執行才跑）
                if st.button("🏁 執行第 1 宗街角地優先權選位（左右側獨立）",
                              use_container_width=True, key='btn_f3L_corner_priority'):
                    _own_map = st.session_state.get('t8_ownership_map', {}) or {}
                    if not _own_map:
                        st.warning("⚠️ 尚未偵測到 土地歸戶 資料，請先至 土地歸戶 匯入歸戶 Excel。")
                    else:
                        from shapely.geometry import LineString as _SLine
                        _corner_select_results = []
                        _corner_cand_diag = []   # 🆕 W-D.1.2 診斷：逐候選三分項（揭露 §1 指數退化，供 KL 核 D-3）
                        # 🆕 Phase A：使用 build_parcels（temp_parcels 子集）取代 f3_G_values
                        # build_parcels 為「可建築土地」之 temp_parcels；此時 G 值尚未計算
                        # 但 PK 所需欄位（暫編地號、原地號、所屬街廓、面積_m2）皆已具備
                        _g_rows = build_parcels
                        by_blk = {}
                        for r in _g_rows:
                            by_blk.setdefault(r['所屬街廓'], []).append(r)
                        _front_line_by_blk = {}
                        for b in _build_blocks:
                            _verts = b.get('vertices') or []
                            if len(_verts) < 3:
                                _front_line_by_blk[b['label']] = None
                                continue
                            try:
                                _longest = None; _ll = 0.0
                                for i in range(len(_verts)):
                                    _p1 = _verts[i]; _p2 = _verts[(i + 1) % len(_verts)]
                                    _L = ((_p2[0] - _p1[0]) ** 2 + (_p2[1] - _p1[1]) ** 2) ** 0.5
                                    if _L > _ll:
                                        _ll = _L
                                        _longest = _SLine([tuple(_p1), tuple(_p2)])
                                _front_line_by_blk[b['label']] = _longest
                            except Exception:
                                _front_line_by_blk[b['label']] = None
                        _side_warnings = []
                        _f3_corner_winners_state = {}
                        for _row in (_corner_rows_init or []):
                            _lbl = _row['街廓']
                            st.session_state['f3_current_pk_block'] = _lbl  # 🚨 W-B §2 bugfix
                            _blk_meta_for_side = next(
                                (b for b in _build_blocks if b.get('label') == _lbl), None
                            )
                            _cad_fl_lstep = (st.session_state.get(
                                'f3_cad_front_lines', {}) or {}).get(_lbl, {})
                            _fl_p1_lstep = _cad_fl_lstep.get('p1') if _cad_fl_lstep else None
                            _fl_p2_lstep = _cad_fl_lstep.get('p2') if _cad_fl_lstep else None
                            _param_dict_for_pk = (
                                st.session_state.get('f3_g_iter_params', {}) or {}
                            )
                            _all_in_blk = by_blk.get(_lbl, [])
                            # 🚨 Patch D-1（Hotfix Fix A 補套用）：候選池一律全自動 PK
                            # 廢除「使用者人工標記覆寫」之分支，避免候選池被使用者誤勾選縮減而漏挑
                            _user_marked_in_blk = [
                                r for r in _all_in_blk
                                if (_param_dict_for_pk.get(r.get('暫編地號', ''), {})
                                    .get('is_corner', False))
                            ]
                            _candidates_pool = _all_in_blk
                            _has_user_marked = bool(_user_marked_in_blk)
                            _candidate_source = 'auto_pk'
                            # 🆕 W-D.2 v2 轉正（§2 tiebreaker 換源）：正典原位次＝
                            #   _projection_order 投影序 rank（單一真相源；廢距角序暫行近似）
                            _rank_by_tpid = {
                                tp.get('暫編地號'): _i_rk + 1
                                for _i_rk, tp in enumerate(
                                    _projection_order(_all_in_blk, _fl_p1_lstep, _fl_p2_lstep))
                            }
                            _candidates = []
                            for r in _candidates_pool:
                                _parent = r.get('原地號', '')
                                _gid = _own_map.get(_parent, '')
                                if not _gid:
                                    continue
                                _tp = next((tp for tp in (temp_parcels or [])
                                            if tp.get('暫編地號') == r.get('暫編地號')), None)
                                _cen_x = float(_tp.get('centroid_x', 0)) if _tp else 0.0
                                _cen_y = float(_tp.get('centroid_y', 0)) if _tp else 0.0
                                _G_est = _estimate_G_for_qualification(
                                    float(r.get('幾何面積_m2', r.get('面積_m2', 0.0)) or 0.0)
                                )
                                _cad_fl_priority = (st.session_state.get(
                                    'f3_cad_front_lengths', {}) or {}).get(_lbl, 0.0)
                                _cad_sl_priority = (st.session_state.get(
                                    'f3_cad_side_lengths', {}) or {}).get(_lbl, 0.0)
                                _front_len_priority = (
                                    _cad_fl_priority if _cad_fl_priority > 0
                                    else float(_row.get('正面長度(m)',
                                                         _row.get('正面路寬(m)', 0.0)) or 0.0)
                                )
                                _side_len_priority = (
                                    _cad_sl_priority if _cad_sl_priority > 0
                                    else 0.0  # Phase A：l₁ 為 G 後欄位，前置 PK 階段 fallback 為 0
                                )
                                _candidates.append({
                                    '歸戶群組': _gid,
                                    '歸戶': _gid,
                                    '暫編地號': r.get('暫編地號', ''),
                                    '原地號': _parent,
                                    'centroid': (_cen_x, _cen_y),
                                    'polygon_coords': (_tp.get('polygon_coords') if _tp else None),
                                    'G_estimated': _G_est,
                                    'G_value': _G_est,
                                    'front_length': _front_len_priority,
                                    'side_length': _side_len_priority,
                                    'physical_overlap_area': float(r.get('幾何面積_m2', r.get('面積_m2', 0.0)) or 0.0),
                                    '臨正街長度_m': _front_len_priority,
                                    '臨側街長度_m': _side_len_priority,
                                    '跨占街角面積_m2': float(r.get('幾何面積_m2', r.get('面積_m2', 0.0)) or 0.0),
                                    # 🆕 W-D.2 v2：§2 正典原位次（tiebreaker 單一真相源）
                                    '_pre_position_rank': _rank_by_tpid.get(
                                        r.get('暫編地號', ''), float('inf')),
                                })
                            _bf = max((c['臨正街長度_m'] for c in _candidates), default=1.0) or 1.0
                            _bs = max((c['臨側街長度_m'] for c in _candidates), default=1.0) or 1.0
                            _ba = max((c['跨占街角面積_m2'] for c in _candidates), default=1.0) or 1.0
                            _l_min_val = _row.get('【左】街角最小面積(㎡)')
                            _r_min_val = _row.get('【右】街角最小面積(㎡)')
                            _use_v13 = (_fl_p1_lstep is not None and _fl_p2_lstep is not None)
                            if _use_v13:
                                _min_p1 = (float(_l_min_val) if _l_min_val is not None
                                            else float('inf'))
                                _min_p2 = (float(_r_min_val) if _r_min_val is not None
                                            else float('inf'))
                                # 🐛 Fix：截角欄位無臨路時存字串 '—'，不可 float()
                                def _safe_cutoff(v):
                                    try:
                                        if v is None or v == '' or v == '—':
                                            return 0.0
                                        return float(v)
                                    except (TypeError, ValueError):
                                        return 0.0
                                _cutoff_p1_for_pk = _safe_cutoff(_row.get('【左】截角(㎡)'))
                                _cutoff_p2_for_pk = _safe_cutoff(_row.get('【右】截角(㎡)'))
                                _g_map = {c['暫編地號']: c['G_estimated'] for c in _candidates}
                                _v13 = select_corner_lots_both_sides_v12(
                                    candidates=_candidates,
                                    front_line_p1=_fl_p1_lstep,
                                    front_line_p2=_fl_p2_lstep,
                                    cutoff_p1_end=_cutoff_p1_for_pk,
                                    cutoff_p2_end=_cutoff_p2_for_pk,
                                    base_front_len_m=_bf,
                                    base_side_len_m_p1=_bs,
                                    base_side_len_m_p2=_bs,
                                    min_corner_area_p1=_min_p1,
                                    min_corner_area_p2=_min_p2,
                                    g_values_map=_g_map,
                                )
                                _l_v13 = _v13['p1_end']; _r_v13 = _v13['p2_end']
                                # 🆕 W-D.1.2 診斷：逐候選三分項攤現況（揭露 §1 指數退化，供 KL 核 D-3）
                                #   端 p1_end→左、p2_end→右（沿用本區 _l/_r 顯示對應）；
                                #   達標 = 通過第一關門檻；分數僅 qualified 有（eliminated 顯示 —）。
                                for _dg_side, _dg_res in (('左', _l_v13), ('右', _r_v13)):
                                    _dg_win = ((_dg_res.get('winner') or {}).get('暫編地號'))
                                    for _dg_pass, _dg_list in (('達標', _dg_res.get('qualified', [])),
                                                               ('未達標', _dg_res.get('eliminated', []))):
                                        for _dc in (_dg_list or []):
                                            _corner_cand_diag.append({
                                                '街廓': _lbl,
                                                '端': _dg_side,
                                                '候選地號': _dc.get('暫編地號', ''),
                                                '原地號': _dc.get('原地號', ''),
                                                '真交集(㎡)': round(float(_dc.get('_corner_intersection_area', 0) or 0), 2),
                                                '整筆幾何(㎡)': round(float(_dc.get('_full_parcel_area',
                                                                                  _dc.get('physical_overlap_area', 0)) or 0), 2),
                                                '範圍面積(㎡)': round(float(_dc.get('_corner_range_area', 0) or 0), 2),
                                                'G估(㎡)': round(float(_dc.get('G_for_threshold', 0) or 0), 2),
                                                '門檻(㎡)': round(float(_dc.get('min_area_to_apply', 0) or 0), 2),
                                                # 🆕 W-D.1.3-b 交叉檢查（KL）：範圍面積 應逐塊 == 門檻（項三分母＝G-gate 最小面積＝同顆法定 range 多邊形）
                                                '範圍=門檻?': ('✅' if abs(float(_dc.get('_corner_range_area', 0) or 0)
                                                                          - float(_dc.get('min_area_to_apply', 0) or 0)) < 0.5
                                                              else '🔴異源·停查'),
                                                # 🆕 項三原始比＝真交集/範圍（分子分母同源應 ≤1；若 🔴>1 表 clamp 觸發、多邊形不一致紅旗）
                                                '項三比(≤1)': ('🔴>1' if (float(_dc.get('_corner_intersection_area', 0) or 0)
                                                                          > float(_dc.get('_corner_range_area', 0) or 0) + 0.01
                                                                          and float(_dc.get('_corner_range_area', 0) or 0) > 0)
                                                               else round(float(_dc.get('_corner_intersection_area', 0) or 0)
                                                                          / max(float(_dc.get('_corner_range_area', 0) or 0), 1e-9), 4)),
                                                '達標': _dg_pass,
                                                # 🆕 W-D.1.3-c 逐筆臨長（供項一/項二核；分子⊆分母、Σ臨≤邊）
                                                '截角邊(range)': round(float(_dc.get('_corner_cut_den', 0) or 0), 3),
                                                '臨截角': round(float(_dc.get('_corner_cut_len', 0) or 0), 3),
                                                '側街邊(range)': round(float(_dc.get('_side_line_den', 0) or 0), 3),
                                                '臨側街': round(float(_dc.get('_side_line_len', 0) or 0), 3),
                                                '正街角分(0.4)': (round(float(_dc.get('_score_corner_cut', 0) or 0), 4)
                                                                  if '_score_corner_cut' in _dc else '—'),
                                                '側街分(0.2)': (round(float(_dc.get('_score_side', 0) or 0), 4)
                                                                if '_score_side' in _dc else '—'),
                                                '跨占分(0.4)': (round(float(_dc.get('_score_overlap', 0) or 0), 4)
                                                                if '_score_overlap' in _dc else '—'),
                                                '總分': (round(float(_dc.get('priority_index', 0) or 0), 4)
                                                         if 'priority_index' in _dc else '—'),
                                                # 🆕 W-D.2 v2 轉正：原位次＝§2 正典投影序 rank（單一真相源；
                                                #   廢距角序暫行欄——v1 診斷 baseline 該欄豁免記帳見 verify/README）
                                                '原位次(投影序)': int(_dc.get('_pre_position_rank', 0) or 0),
                                                '選中': ('✅' if (_dg_win and _dc.get('暫編地號') == _dg_win) else ''),
                                            })
                                _l_disp_min = ('無此側' if _min_p1 == float('inf')
                                                else f"{round(_min_p1, 2)}")
                                _r_disp_min = ('無此側' if _min_p2 == float('inf')
                                                else f"{round(_min_p2, 2)}")
                                _l_winner = _l_v13.get('winner')
                                _r_winner = _r_v13.get('winner')
                                _l_disp_winner = (
                                    '無此側' if _l_disp_min == '無此側'
                                    else (f"{_l_winner['歸戶群組']}（{_l_winner.get('原地號','')}）"
                                          f"[{_l_winner.get('暫編地號','')}]"
                                          if _l_winner else '⚠️ 強制抵費地')
                                )
                                _r_disp_winner = (
                                    '無此側' if _r_disp_min == '無此側'
                                    else (f"{_r_winner['歸戶群組']}（{_r_winner.get('原地號','')}）"
                                          f"[{_r_winner.get('暫編地號','')}]"
                                          if _r_winner else '⚠️ 強制抵費地')
                                )
                                _l_disp_score = (
                                    round(float(_l_winner.get('priority_index', 0)), 4)
                                    if (_l_winner and _l_disp_min != '無此側') else '—'
                                )
                                _r_disp_score = (
                                    round(float(_r_winner.get('priority_index', 0)), 4)
                                    if (_r_winner and _r_disp_min != '無此側') else '—'
                                )
                                _l_qcount = (len(_l_v13.get('qualified', []))
                                              if _l_disp_min != '無此側' else '—')
                                _r_qcount = (len(_r_v13.get('qualified', []))
                                              if _r_disp_min != '無此側' else '—')
                                _f3_corner_winners_state[_lbl] = {
                                    'p1_end': (_l_winner.get('暫編地號', '') if _l_winner else None),
                                    'p2_end': (_r_winner.get('暫編地號', '') if _r_winner else None),
                                    'method': 'V13_spatial_binding',
                                }
                            else:
                                # 🆕 W-D.1.3-d（O1 裁定）：缺 FRONT_LINE → 停機警告，不靜默退回 V12 生面積評分。
                                #   _use_v13 由 FRONT_LINE p1/p2 是否存在判定；缺 FRONT 則街角三指數
                                #   （臨截角/臨側街/真交集）皆無幾何依據 → 具名中文警示 + 跳過該塊（不 append 假結果）。
                                st.error(
                                    f"🛑 街廓 {_lbl}：缺 FRONT_LINE 圖層（起點 p1＝左、迄點 p2＝右）→ "
                                    f"無法執行第 1 宗街角地優先權選位（街角三指數需 FRONT/SIDE/ALLOC/BASELINE 圖層齊備）。\n\n"
                                    f"請至 CAD 於 FRONT_LINE 圖層補畫該街廓正面臨路線後，重新匯入 DXF 並重跑此步驟。"
                                )
                                continue
                            _corner_select_results.append({
                                '街廓': _lbl,
                                '演算法': ('V13' if _use_v13 else 'V12'),
                                '候選來源': ('🎯 使用者標記' if _candidate_source == 'user_marked'
                                             else '🤖 自動 PK'),
                                '候選數': len(_candidates),
                                '【左】最小面積(㎡)': _l_disp_min,
                                '【右】最小面積(㎡)': _r_disp_min,
                                '【左】達資格候選': _l_qcount,
                                '【左】第1宗指配': _l_disp_winner,
                                '【左】優先權指數': _l_disp_score,
                                '【右】達資格候選': _r_qcount,
                                '【右】第1宗指配': _r_disp_winner,
                                '【右】優先權指數': _r_disp_score,
                            })
                        if _f3_corner_winners_state:
                            st.session_state['f3_corner_winners'] = _f3_corner_winners_state
                        # 🆕 Phase B-1：寫入 forced_offset 供 Step G 雙端鎖定 + Phase C 抵費地 buffer 邏輯使用
                        _forced_offset_map = {}

                        def _fo_min_area(_v):
                            """'無此側'/None → 0.0；數字字串（如 '300.52'）→ float。"""
                            try:
                                return float(_v)
                            except (TypeError, ValueError):
                                return 0.0
                        for _r_pk in (_corner_select_results or []):
                            _lbl_pk = _r_pk.get('街廓', '')
                            _l_forced = ('強制抵費地' in str(_r_pk.get('【左】第1宗指配', '')))
                            _r_forced = ('強制抵費地' in str(_r_pk.get('【右】第1宗指配', '')))
                            _l_has_side = (_r_pk.get('【左】最小面積(㎡)') != '無此側')
                            _r_has_side = (_r_pk.get('【右】最小面積(㎡)') != '無此側')
                            _forced_offset_map[_lbl_pk] = {
                                'left_forced_offset': bool(_l_forced and _l_has_side),
                                'right_forced_offset': bool(_r_forced and _r_has_side),
                                'left_has_side': bool(_l_has_side),
                                'right_has_side': bool(_r_has_side),
                                # 🆕 W-D.2 §3（M3 接線・餵入端）：角落抵費地面積＝該側 range 面積
                                #   （＝【左/右】最小面積同源）。消費端＝_spatial_order_parcels_v2
                                #   回傳 corner_offset_area → Step G 守恆 ledger 拆帳。
                                'left_corner_min_area': (
                                    _fo_min_area(_r_pk.get('【左】最小面積(㎡)'))
                                    if (_l_forced and _l_has_side) else 0.0),
                                'right_corner_min_area': (
                                    _fo_min_area(_r_pk.get('【右】最小面積(㎡)'))
                                    if (_r_forced and _r_has_side) else 0.0),
                            }
                        st.session_state['f3L_forced_offset'] = _forced_offset_map
                        if _corner_select_results:
                            st.markdown("##### 🥇 第 1 宗街角地指配結果（左右側獨立）")
                            if _side_warnings:
                                with st.expander(
                                    f"⚠️ {len(_side_warnings)} 筆「使用者標註側別」與「幾何投影建議」不一致",
                                    expanded=False
                                ):
                                    for _w in _side_warnings:
                                        st.warning(_w)
                                    st.info(
                                        "ℹ️ 說明：上述地號之側別由使用者於步驟 G 圖選器手動標註，"
                                        "但本系統幾何投影演算法（`LineString.project()`）建議的側別不同。"
                                        "已採用使用者標註值；如需修正，請至步驟 G 圖選器重新點選。"
                                    )
                            st.dataframe(_pd.DataFrame(_corner_select_results),
                                         use_container_width=True, hide_index=True)
                            st.session_state['f3L_corner_winners'] = _corner_select_results
                            st.session_state['f3L_corner_side_warnings'] = _side_warnings
                            st.caption(
                                "💡 **左右側分類規則**：(1) 使用者標註優先；(2) 未標註者用 `LineString.project()` "
                                "沿正面道路投影距離，前半段=左側、後半段=右側；(3) Phase 4 H-1 — 圖選時即時自動偵測。"
                                "🆕 **最小面積**：左右側使用各自路寬計算 + 各自截角獨立扣除（H-3b）；"
                                "**Tiebreaker**：同分時依跨占面積 → 正街長度 → 歸戶字母序決勝（F-3）。"
                                "🆕 **無此側**：表示該街廓對應側無臨路，無街角地候選資格。"
                                "🆕 **Phase B**：強制抵費地之街角將於 Phase C 自動預留為抵費地，不分配給任何宗地。"
                            )
                            # 🆕 W-D.1.2 診斷：第 1 宗街角地逐候選評點明細（揭露 §1 指數退化，供 KL 核 D-3）
                            if _corner_cand_diag:
                                with st.expander(
                                    "🔬 W-D.1.2 診斷：第 1 宗街角地逐候選評點明細"
                                    f"（{len(_corner_cand_diag)} 筆候選 · 三分項 0.4/0.2/0.4 · 達標/選中）",
                                    expanded=False
                                ):
                                    st.dataframe(_pd.DataFrame(_corner_cand_diag),
                                                 use_container_width=True, hide_index=True)
                                    st.caption(
                                        "💡 **三分項**（手冊街角地優先權指數）：正街角線 0.4 + 側街 0.2 + 跨占面積 0.4 = 總分。"
                                        "**達標**＝過第一關門檻（G估 ≥ 街角地最小分配面積）；未達標不計分（顯示 —）。"
                                        "**選中**＝該端達標候選中總分最高者（位次 1 街角地）。\n\n"
                                        "🔎 **W-D.1.3-b（項三已改真交集）**："
                                        "『跨占分』＝ 0.4×(真交集 / 範圍面積)（值域 0–0.4；已刪 15m 框）；"
                                        "『真交集』＝ parcel ∩ `_build_corner_range_v2`（**已用於評分**）、『範圍面積』＝同 range 多邊形面積（項三分母）；"
                                        "winner＝達標候選中真交集最大者（項一/項二仍常數 0.4/0.2，per-parcel 為 -c）；"
                                        "**交叉檢查（KL）：『範圍=門檻?』應逐塊全 ✅（項三分母＝G-gate 最小面積＝同顆法定 range 多邊形）、『項三比』應全 ≤1（無 clamp）；任一 🔴 → 停查**；"
                                        "『側街分』為 0 或 — 表示 `side_length` 未供（-c 逐筆化）。\n\n"
                                        "🆕 **原位次(投影序)**：tiebreaker 可稽核欄＝§2 正典原位次"
                                        "（`_projection_order` 投影序 rank，單一真相源；W-D.2 v2 轉正時換源、廢距角序暫行）。"
                                    )
                            # 🚨 W-B §6：街角規定範圍面積驗收
                            _cr_areas_ui = st.session_state.get(
                                'f3_corner_range_areas', {}
                            ) or {}
                            if _cr_areas_ui:
                                with st.expander(
                                    "🔍 W-B 驗收：街角規定範圍面積（ALLOC_LINE 五邊形法）",
                                    expanded=False
                                ):
                                    _cr_rows_ui = []
                                    for _b_ui in sorted(_cr_areas_ui.keys()):
                                        _v = _cr_areas_ui[_b_ui]
                                        _cr_rows_ui.append({
                                            '街廓': _b_ui,
                                            '【左】規定範圍(㎡)': (
                                                f"{_v['left']:.2f}"
                                                if _v.get('left') is not None else '—'
                                            ),
                                            '【右】規定範圍(㎡)': (
                                                f"{_v['right']:.2f}"
                                                if _v.get('right') is not None else '—'
                                            ),
                                        })
                                    import pandas as _pd_cr
                                    st.dataframe(
                                        _pd_cr.DataFrame(_cr_rows_ui),
                                        use_container_width=True, hide_index=True
                                    )
                                    st.caption(
                                        "驗收基準：R5 左側 ≈ **146.50㎡**（退縮 0m，"
                                        "V6.dxf ALLOC_LINE + SIDE_LINE 五邊形∩BLOCK−截角）"
                                    )
                            # 🆕 W-D.1.3-d（§7.4 圖9）：抵費地情形驗收（additive 診斷、純讀既有指配結果）
                            _offset_diag_rows = []
                            for _r_off in (_corner_select_results or []):
                                for _end_lbl_off, _min_key_off, _win_key_off in (
                                    ('左', '【左】最小面積(㎡)', '【左】第1宗指配'),
                                    ('右', '【右】最小面積(㎡)', '【右】第1宗指配')):
                                    if '強制抵費地' in str(_r_off.get(_win_key_off, '')):
                                        _offset_diag_rows.append({
                                            '街廓': _r_off.get('街廓', ''),
                                            '端': _end_lbl_off,
                                            '抵費地面積＝range(㎡)': _r_off.get(_min_key_off),
                                            '指配': '強制抵費地',
                                        })
                            if _offset_diag_rows:
                                with st.expander(
                                    f"🔍 W-D.1.3-d 驗收：抵費地情形（§7.4 圖9 · {len(_offset_diag_rows)} 端強制抵費地）",
                                    expanded=False
                                ):
                                    st.dataframe(_pd.DataFrame(_offset_diag_rows),
                                                 use_container_width=True, hide_index=True)
                                    st.caption(
                                        "💡 全候選未過 G 門檻 → 第 1 宗街角地留設為**抵費地**（合法輸出、非失敗）；"
                                        "其**面積鎖定＝街角最小分配面積範圍面積**（§1.1 之 T-多邊形，即本表『抵費地面積＝range』）。\n\n"
                                        "🎯 **3.5m 靶**：R5左／R2左／R3右（range 面積 ≈ 300.52／309.05／308.93㎡）。\n\n"
                                        "⚠️ 守恆『ΣG＋角落抵費地＋中央池＝街廓』之『中央池／角落抵費地＝總池重定位』接線屬 **W-D.2 §3 滑池槽**；"
                                        "本表僅驗『抵費地面積＝range 面積』，不接 ΣG＋池（接線＝行為變更、非本波範圍）。"
                                    )

                # ---- (已停用舊 Step G 圖選器 fragment，僅保留程式碼供參考) ----
                if False:
                    @st.fragment
                    def _f3_corner_picker_fragment():
                        st.caption("💡 點擊地圖上某筆地號 → 切換街角地狀態（🔴 紅色填色＝已標街角；🔵 藍色框＝非街角）。"
                                   "點選前先選好「預設街角側別」，新標記的地號會套用此側別，可後續至下表調整。")
                        import plotly.graph_objects as go

                        # fragment 重跑時從 session_state 重讀最新參數
                        _params = dict(st.session_state.get(_param_key, {}))

                        # 預設側別下拉
                        # 🆕 H-1：將預設改為 '左側'（市地重劃實務之主分配方向）+ 加註「優先採自動偵測」
                        default_side = st.radio(
                            "預設街角側別（自動偵測失敗時 fallback 用此值）",
                            options=['左側', '右側'], horizontal=True, key='f3_corner_default_side',
                            help="點選街角地時，系統會以幾何投影自動判定左/右側（依宗地質心於街廓正面道路上的投影距離）。"
                                  "僅當自動偵測失敗時，才使用此預設值。",
                        )

                        fig = go.Figure()
                        # 1️⃣ 街廓填色 + 邊框
                        valid_blocks = [b for b in build_blocks if len(b.get('vertices', [])) >= 3]
                        for b in valid_blocks:
                            verts = b['vertices']
                            xs = [v[0] for v in verts] + [verts[0][0]]
                            ys = [v[1] for v in verts] + [verts[0][1]]
                            fig.add_trace(go.Scatter(
                                x=xs, y=ys, mode='lines',
                                fill='toself', fillcolor='rgba(200,200,200,0.15)',
                                line=dict(color='#888888', width=1.2),
                                hoverinfo='skip', showlegend=False,
                            ))
                        block_trace_count = len(valid_blocks)

                        # 2️⃣ 街廓 label 文字（合成一條 trace）
                        if valid_blocks:
                            fig.add_trace(go.Scatter(
                                x=[b['centroid'][0] for b in valid_blocks],
                                y=[b['centroid'][1] for b in valid_blocks],
                                mode='text',
                                text=[f"<b>{b.get('label', '')}</b>" for b in valid_blocks],
                                textfont=dict(size=11, color='#444'),
                                hoverinfo='skip', showlegend=False,
                            ))
                            label_trace_count = 1
                        else:
                            label_trace_count = 0

                        # 3️⃣ 各筆地號可點擊 trace（每筆 1 條）
                        valid_parcels = [tp for tp in build_parcels
                                         if len(tp.get('polygon_coords') or []) >= 3]
                        for tp in valid_parcels:
                            coords = tp['polygon_coords']
                            xs_c = [c[0] for c in coords] + [coords[0][0]]
                            ys_c = [c[1] for c in coords] + [coords[0][1]]
                            k = tp['暫編地號']
                            is_c = bool(_params.get(k, {}).get('is_corner', False))
                            side_str = _params.get(k, {}).get('side', '無')
                            if is_c:
                                _fillc = 'rgba(214,39,40,0.45)'
                                _linec = '#D62728'; _lw = 2.0
                                _state = f"★ 街角地（{side_str}）"
                            else:
                                _fillc = 'rgba(31,119,180,0.10)'
                                _linec = '#1F77B4'; _lw = 0.6
                                _state = "一般地號"
                            fig.add_trace(go.Scatter(
                                x=xs_c, y=ys_c, mode='lines',
                                fill='toself', fillcolor=_fillc,
                                line=dict(color=_linec, width=_lw),
                                hoverinfo='text',
                                text=f"暫編地號 {k}<br>所屬街廓 {tp['所屬街廓']}<br>狀態：{_state}",
                                showlegend=False, name=k,
                            ))

                        # 4️⃣ Task C：手動基準線拾取的紅色標記（已點擊的點 + 已設定的基準線方向）
                        _pmode_now = st.session_state.get('f3_baseline_pick_mode', None)
                        if _pmode_now and _pmode_now.get('clicks'):
                            _picked_pts = _pmode_now['clicks']
                            fig.add_trace(go.Scatter(
                                x=[p[0] for p in _picked_pts],
                                y=[p[1] for p in _picked_pts],
                                mode='markers+text',
                                marker=dict(symbol='cross', size=14, color='#D62728',
                                            line=dict(color='#FFFFFF', width=2)),
                                text=[f"{i+1}" for i in range(len(_picked_pts))],
                                textposition='top center',
                                textfont=dict(size=14, color='#D62728'),
                                hoverinfo='skip', showlegend=False,
                            ))
                        # 已設定的基準線：紅色虛線箭頭
                        _mbl_render = st.session_state.get('f3_manual_baseline', {}) or {}
                        for _bl, _bv in _mbl_render.items():
                            if not _bv.get('enabled') or not _bv.get('point'):
                                continue
                            try:
                                import math as _math_r
                                _bx, _by = _bv['point']
                                _bang = _math_r.radians(float(_bv.get('angle_deg', 0.0)))
                                _alen = 30.0   # 顯示 30m 長度
                                _ex = _bx + _alen * _math_r.cos(_bang)
                                _ey = _by + _alen * _math_r.sin(_bang)
                                fig.add_trace(go.Scatter(
                                    x=[_bx, _ex], y=[_by, _ey],
                                    mode='lines+markers',
                                    line=dict(color='#D62728', width=2, dash='dash'),
                                    marker=dict(symbol='arrow-up', size=12, color='#D62728',
                                                angleref='previous'),
                                    name=f"{_bl} 基準線",
                                    hoverinfo='text',
                                    text=[f"{_bl} 起點", f"{_bl} 方向 ({_bv.get('angle_deg', 0):.1f}°)"],
                                    showlegend=False,
                                ))
                            except Exception:
                                pass

                        fig.update_layout(
                            height=520,
                            paper_bgcolor='white', plot_bgcolor='white',
                            xaxis=dict(title="TWD97 X", gridcolor='#E5E5E5'),
                            yaxis=dict(title="TWD97 Y", gridcolor='#E5E5E5',
                                       scaleanchor='x', scaleratio=1),
                            margin=dict(l=40, r=40, t=10, b=40),
                            showlegend=False,
                        )

                        try:
                            from streamlit_plotly_events import plotly_events as _pe
                            _has_pe = True
                        except Exception:
                            _has_pe = False

                        if not _has_pe:
                            st.warning("⚠️ 缺少 streamlit-plotly-events 套件，僅顯示預覽圖。")
                            st.plotly_chart(fig, use_container_width=True)
                            return

                        # 用 click_counter 旋轉 key，強制 plotly_events 元件重置
                        # 否則 component 會把上次點擊「記住」每次重跑都重複回傳同一筆 → toast 無限跳
                        _ck_cntr = st.session_state.get('f3_corner_click_counter', 0)
                        picked = _pe(fig, click_event=True, hover_event=False,
                                     key=f'f3_corner_clicker_{_ck_cntr}', override_height=520)

                        # 圖層 trace 順序：街廓填色(N) → label(1) → 地號(N_parcels)
                        click_offset = block_trace_count + label_trace_count
                        # 🆕 Phase 4 註：上方統一 GIS 互動地圖為主要操作介面；
                        #               此處保留為輔助，移除 Phase 3 的 f3_map_mode 守門
                        if picked:
                            ev = picked[0]
                            # ── Task C：兩點定線拾取模式（優先攔截） ──
                            _pmode = st.session_state.get('f3_baseline_pick_mode', None)
                            if _pmode and _pmode.get('block'):
                                _cx = ev.get('x'); _cy = ev.get('y')
                                if _cx is not None and _cy is not None:
                                    _clicks = list(_pmode.get('clicks', []))
                                    _clicks.append((float(_cx), float(_cy)))
                                    _blk = _pmode['block']
                                    if len(_clicks) >= 2:
                                        # 第 2 點到位 → 計算角度並回填
                                        import math as _math
                                        _x1, _y1 = _clicks[0]
                                        _x2, _y2 = _clicks[1]
                                        _dx = _x2 - _x1; _dy = _y2 - _y1
                                        _seg_len = _math.hypot(_dx, _dy)
                                        if _seg_len < 1e-6:
                                            st.warning("⚠️ 兩點過近無法計算方向，已重置拾取模式")
                                            st.session_state['f3_baseline_pick_mode'] = None
                                        else:
                                            _angle_deg = _math.degrees(_math.atan2(_dy, _dx))
                                            # 反向防呆：與街廓 MBR 長軸夾角 > 90° 警告
                                            try:
                                                _blk_meta = block_meta_by_label.get(_blk)
                                                if _blk_meta and len(_blk_meta.get('vertices', [])) >= 3:
                                                    from shapely.geometry import Polygon as _SP_chk
                                                    import numpy as _np_chk
                                                    _bp_chk = _SP_chk(_blk_meta['vertices'])
                                                    if _bp_chk.is_valid:
                                                        _mbr_dir, _ = _detect_block_allocation_dir(
                                                            _bp_chk, _np_chk.array([1.0, 0.0])
                                                        )
                                                        _user_vec = _np_chk.array([_dx, _dy]) / _seg_len
                                                        _dot = float(_np_chk.dot(_user_vec, _mbr_dir))
                                                        if _dot < 0:
                                                            st.warning(
                                                                "⚠️ 您指定的方向與街廓長軸夾角 > 90°（即與街廓主走向相反）。"
                                                                "請確認左右側分配是否符合預期；如需反向可手動調整方向角 ±180°。"
                                                            )
                                            except Exception:
                                                pass
                                            # 寫入 manual baseline
                                            _mbl_now = dict(st.session_state.get('f3_manual_baseline', {}))
                                            _mbl_now[_blk] = {
                                                'point': (_x1, _y1),
                                                'angle_deg': _angle_deg,
                                                'enabled': True,
                                            }
                                            st.session_state['f3_manual_baseline'] = _mbl_now
                                            st.session_state['f3_baseline_pick_mode'] = None
                                            st.toast(
                                                f"✅ 街廓 {_blk} 基準線已設定：起點 ({_x1:.2f}, {_y1:.2f})、"
                                                f"方向 {_angle_deg:.2f}°", icon='📐'
                                            )
                                    else:
                                        # 僅取得第 1 點 → 更新狀態，等第 2 點
                                        st.session_state['f3_baseline_pick_mode'] = {
                                            'block': _blk, 'clicks': _clicks
                                        }
                                        st.toast(
                                            f"📍 街廓 {_blk} 第 1 點已記錄 ({_x1:.2f}, {_y1:.2f})，請繼續點第 2 點",
                                            icon='1️⃣'
                                        )
                                    # 旋轉 counter 防止重複觸發
                                    st.session_state['f3_corner_click_counter'] = _ck_cntr + 1
                                    st.rerun(scope='app')

                            # ── 既有：街角地切換邏輯 + 🆕 H-1 自動側別偵測 ──
                            cn = ev.get('curveNumber')
                            if cn is not None and click_offset <= cn < click_offset + len(valid_parcels):
                                tp = valid_parcels[cn - click_offset]
                                k = tp['暫編地號']
                                curr = dict(_params.get(k, {}))
                                new_is = not bool(curr.get('is_corner', False))
                                curr['is_corner'] = new_is
                                if new_is:
                                    # 🆕 H-1：優先採自動偵測（依幾何投影），失敗才 fallback 至 default_side
                                    _blk_label_for_auto = tp.get('所屬街廓', '')
                                    _blk_meta_for_auto = next(
                                        (b for b in valid_blocks
                                         if b.get('label') == _blk_label_for_auto),
                                        None
                                    )
                                    _auto = (_auto_detect_corner_side(tp, _blk_meta_for_auto)
                                             if _blk_meta_for_auto else '')
                                    if _auto in ('左側', '右側'):
                                        curr['side'] = _auto
                                        _side_note = f"自動判定為 {_auto}"
                                    else:
                                        # 自動偵測失敗 → 採 default_side
                                        if curr.get('side', '無') == '無':
                                            curr['side'] = default_side
                                        _side_note = f"幾何投影失敗，採預設 {curr['side']}"
                                else:
                                    curr['side'] = '無'
                                    _side_note = ''
                                if 'W' not in curr:
                                    curr['W'] = 0.0
                                _params[k] = curr
                                st.session_state[_param_key] = _params
                                # 遞增 counter 旋轉 key（下次重跑後 plotly_events 不再回傳此筆）
                                st.session_state['f3_corner_click_counter'] = _ck_cntr + 1
                                # 🆕 Task F-1：觸發式重算旗標（下次 Step G 區塊渲染時自動執行 G 計算）
                                st.session_state['f3_g_needs_recalc'] = True
                                _msg = (f"✅ 已標記街角地（{_side_note}）" if new_is
                                        else "❎ 已取消街角地")
                                st.toast(f"{_msg}：{k}", icon='🎯')
                                # 用 app scope 重跑，讓下方 data_editor 表格也立即反映新狀態
                                st.rerun(scope='app')

                        # 列出目前已標記之街角地
                        marked = [(k, v.get('side', '無'))
                                  for k, v in _params.items() if v.get('is_corner')]
                        if marked:
                            st.markdown("**📋 目前已標記街角地：**")
                            st.write("、".join(f"{k}（{s}）" for k, s in marked))
                        else:
                            st.info("尚未標記任何街角地。請於上方地圖點擊地號多邊形。")

                    _f3_corner_picker_fragment()

                # 每筆組出 seed row；F 依 is_corner + 側別 自動帶入
                _seed_rows = []
                for tp in build_parcels:
                    k = tp['暫編地號']
                    prev = _prev_params.get(k, {})
                    blk_label = tp['所屬街廓']
                    sb_row = sb_rows_by_label.get(blk_label, {})
                    is_c = bool(prev.get('is_corner', False))
                    side = prev.get('side', '無')  # '無' / '左側' / '右側'
                    if is_c and side == '左側':
                        F_auto = float(sb_row.get('左側長度(m)', 0.0) or 0.0)
                    elif is_c and side == '右側':
                        F_auto = float(sb_row.get('右側長度(m)', 0.0) or 0.0)
                    else:
                        F_auto = 0.0
                    _seed_rows.append({
                        '暫編地號':   k,
                        '所屬街廓':   blk_label,
                        '重劃前區段': tp.get('重劃前地價區段', ''),
                        '街角地':     is_c,
                        '街角側別':   side,
                        'F（自動）(m)': round(F_auto, 2),
                        'W (m)':       float(prev.get('W', 0.0)),
                    })
                _df_seed = _pd.DataFrame(_seed_rows)
                # data_editor 純為「顯示鏡像」：街角地 / 街角側別 / F 為唯讀（避免 widget cache 誤判覆寫 session_state）
                # 街角地的修改一律透過上方 fragment（圖面點選）完成；W 仍可在表格手動輸入或按 Auto W 計算。
                _ed_key_suffix = st.session_state.get('f3_corner_click_counter', 0)
                _edited = st.data_editor(
                    _df_seed,
                    key=f'f3_g_iter_editor_{_ed_key_suffix}',
                    hide_index=True,
                    use_container_width=True,
                    column_config={
                        '暫編地號':    st.column_config.TextColumn(disabled=True),
                        '所屬街廓':    st.column_config.TextColumn(disabled=True),
                        '重劃前區段':  st.column_config.TextColumn(disabled=True),
                        '街角地':      st.column_config.CheckboxColumn(disabled=True,
                            help="此欄為唯讀；請於上方「🖱️ 圖面點選街角地」expander 點選圖形以切換"),
                        '街角側別':    st.column_config.TextColumn(disabled=True,
                            help="此欄為唯讀；街角地的側別由上方圖選器設定（預設套用點選前選的右側/左側）"),
                        'F（自動）(m)': st.column_config.NumberColumn(disabled=True,
                            help="依街角側別，自動取步驟 E 該街廓之左/右側長度"),
                        'W (m)':       st.column_config.NumberColumn(min_value=0.0, step=0.1,
                            help="分配土地寬度 — 可手動輸入或按「🔍 自動計算 W」推導"),
                    },
                )

                # 寫回邏輯：只回寫「W (m)」欄位的使用者編輯，街角地 / 側別 完全由 fragment 決定，避免互相覆寫
                _existing = dict(st.session_state.get(_param_key, {}))
                _widget_state = st.session_state.get(f'f3_g_iter_editor_{_ed_key_suffix}', {})
                _user_edits = (_widget_state.get('edited_rows', {})
                               if isinstance(_widget_state, dict) else {})
                if _user_edits:
                    for _row_key, _col_edits in _user_edits.items():
                        try:
                            _row_idx = int(_row_key)
                        except Exception:
                            continue
                        if _row_idx < 0 or _row_idx >= len(_seed_rows):
                            continue
                        _k = str(_seed_rows[_row_idx].get('暫編地號', ''))
                        if not _k or 'W (m)' not in _col_edits:
                            continue
                        _curr = dict(_existing.get(_k, {}))
                        try:
                            _curr['W'] = float(_col_edits['W (m)'] or 0.0)
                            _existing[_k] = _curr
                        except Exception:
                            pass
                    st.session_state[_param_key] = _existing
                _new_params = dict(_existing)

                # ---- 🧮 執行 G 值迭代計算（幾何二分法 + 邊界繼承） ----
                st.caption("📐 改用「幾何二分法」演算：不再事前計算 W，每筆宗地由 baseline 沿正面街切出 block strip，"
                           "比對實際幾何面積與法規 G_target；同街廓多筆採「邊界繼承」連續配地。"
                           "第 1 筆若為街角地走完整側街運算，第 2 筆起 F=0、Rw=0、l₁=0。")
                # 🆕 Task F-1：觸發式重算 — 街角地點選後自動跑 G 值計算
                _auto_recalc = bool(st.session_state.pop('f3_g_needs_recalc', False))
                _btn_clicked = st.button("🧮 執行 G 值迭代計算（雙向夾擠 + 抵費地）",
                                          use_container_width=True,
                                          key='btn_run_g_iter', type='primary')
                if _btn_clicked or _auto_recalc:
                    if _auto_recalc and not _btn_clicked:
                        st.info("🔄 偵測到街角地變動，自動重算 G 值…")
                    g_rows = []
                    detail_trace = {}
                    _params_for_g = dict(st.session_state.get(_param_key, _new_params))

                    # 🆕 V13 修正 #2：若 Step L 已執行 → 用其 winner 覆寫 _params_for_g
                    # （Step L 為法定流程，PK 結果優先於使用者個別標記）
                    _step_l_winners = (st.session_state.get('f3_corner_winners', {}) or {})
                    if _step_l_winners:
                        # 收集所有受 Step L 影響之街廓
                        _affected_blocks = set(_step_l_winners.keys())
                        # 先把這些街廓所有暫編地號之 is_corner 設為 False
                        for tp in build_parcels:
                            if tp.get('所屬街廓', '') in _affected_blocks:
                                k_tp = tp.get('暫編地號', '')
                                _existing = dict(_params_for_g.get(k_tp, {}))
                                _existing['is_corner'] = False
                                _existing['side'] = '無'
                                _params_for_g[k_tp] = _existing
                        # 再標記 Step L 之 winner（p1_end → 左側、p2_end → 右側）
                        _override_count = 0
                        for blk_lbl_w, winners in _step_l_winners.items():
                            p1_winner = winners.get('p1_end')
                            p2_winner = winners.get('p2_end')
                            if p1_winner:
                                _existing = dict(_params_for_g.get(p1_winner, {}))
                                _existing['is_corner'] = True
                                _existing['side'] = '左側'
                                _params_for_g[p1_winner] = _existing
                                _override_count += 1
                            if p2_winner:
                                _existing = dict(_params_for_g.get(p2_winner, {}))
                                _existing['is_corner'] = True
                                _existing['side'] = '右側'
                                _params_for_g[p2_winner] = _existing
                                _override_count += 1
                        if _override_count > 0:
                            st.info(f"📍 已採用步驟 L 之街角地 PK 結果（覆寫 {_override_count} 筆暫編地號之街角標記）")

                    # 🆕 Phase B：讀取 forced_offset 標記並通知使用者
                    # forced_offset = True 表示該側街角範圍將強制留設為抵費地（Phase C 自動處理面積保留）
                    _f3L_forced_offset = (st.session_state.get('f3L_forced_offset', {}) or {})
                    if _f3L_forced_offset:
                        _forced_blocks = []
                        for _blk_lbl_fo, _fo in _f3L_forced_offset.items():
                            _sides_fo = []
                            if _fo.get('left_forced_offset'):
                                _sides_fo.append('左側')
                            if _fo.get('right_forced_offset'):
                                _sides_fo.append('右側')
                            if _sides_fo:
                                _forced_blocks.append(f"{_blk_lbl_fo}（{', '.join(_sides_fo)}）")
                        if _forced_blocks:
                            st.warning(
                                f"⚠️ **{len(_forced_blocks)} 個街廓有強制抵費地街角**："
                                + "、".join(_forced_blocks)
                                + "\n\n💡 強制抵費地之街角範圍將於 **Phase C（公設地調配 / 抵費地 buffer）"
                                "** 自動預留為抵費地（不分配給任何宗地）。"
                                "目前 G 值迭代仍按原邏輯計算，Phase C 完成前抵費地面積可能偏大。"
                            )

                    # 1. 將 build_parcels 依「所屬街廓」分組；每街廓內以「街角地優先 + 輸入順序」排序
                    parcels_by_block = {}
                    for tp in build_parcels:
                        parcels_by_block.setdefault(tp['所屬街廓'], []).append(tp)
                    # （Task D：移除舊 corner-priority 排序；🚨 Patch E-2 後改於各街廓內以 _spatial_order_parcels_v2 處理）

                    # 2. 逐街廓處理（雙向夾擠 + 抵費地）— Task D 核心重構
                    import numpy as _np_d
                    from shapely.geometry import Polygon as _SP_d
                    from shapely.ops import unary_union as _uunion_d
                    import math as _math_d

                    def _build_g_row(_k, _tp, _blk_label, _blk_area, _front_len, _avg_depth,
                                     _zone, _A_ratio, _l_front, _l_side, _F, _is_corner_mark,
                                     _is_first_corner, _side, _res, _solver_label, _alloc_side):
                        """組成單筆 g_rows 項目（DRY helper）"""
                        # 🚨 W-A 規格一：a_for_G = 分攤登記面積_m2 + 面積_m2(a' 累加器)
                        if '分攤登記面積_m2' in _tp:
                            _a_m2 = round(float(_tp.get('分攤登記面積_m2', 0) or 0)
                                          + float(_tp.get('面積_m2', 0) or 0), 2)
                        else:
                            _a_m2 = round(float(_tp.get('面積_m2', 0) or 0), 2)
                        return {
                            '暫編地號': _k, '原地號': _tp.get('原地號', ''),
                            '所屬街廓': _blk_label,
                            '重劃前區段': _zone,
                            'a 面積(㎡)': _a_m2,
                            '街廓面積(㎡)': round(_blk_area, 2),
                            '正面長度(m)': round(_front_len, 2),
                            '平均深度(m)': round(_avg_depth, 2),
                            'A 地價比': round(_A_ratio, 4),
                            'l₂ 正面尺度': round(_l_front, 2),
                            'l₁ 側面尺度': round(_l_side, 2),
                            '街角地': '是' if _is_corner_mark else '否',
                            '第1筆街角': '是' if _is_first_corner else '否',
                            '街角側別': _side if _is_corner_mark else '—',
                            'F(m)': round(_F, 2),
                            'W(m)': round(_res.get('W', 0.0), 2),
                            'Rw(%)': round(_res.get('Rw_pct', 0.0), 2),
                            'S(m)': round(_res.get('S', 0.0), 2),
                            '幾何面積(㎡)': round(_res.get('area_geom', 0.0), 2),
                            'G(㎡)': round(_res.get('G', 0.0), 2),
                            '累積S(m)': round(float(_res.get('_alloc_cum_S', 0.0)), 2),
                            '推進側別': _alloc_side,    # 🆕 Task D：'left' / 'right' / '單向' / '抵費地'
                            '解法': _solver_label,
                            '迭代次數': _res.get('iterations', 0),
                            '是否收斂': _res.get('是否收斂_override',
                                                 '✅' if _res.get('converged') else '⚠️'),
                            '負擔比率': round(1 - _res['G']/_a_m2, 4)
                                       if _a_m2 > 0 and _res.get('G') else 0,
                            '宗地寬度(m)': round(_res.get('_宗地寬度', 0.0), 2),   # 🆕 §1 判去留用（單筆）
                            '畸零地旗標': _res.get('_畸零旗標', ''),               # 🆕 §1-4 第1調配順位
                            'cut_coords': _res.get('cut_coords', []) or [],
                        }

                    def _solve_one(_a_m2, _A, _l_front, _l_side, _F, _blk_poly, _d_hat,
                                   _baseline_pt, _S_max, _is_corner, _side, _avg_depth,
                                   _allocation_dir=None, _side_mid=None, _W_prev=0.0):
                        """求解單筆宗地 — 幾何二分法優先，失敗 fallback 至代數迭代

                        🆕 W-C §0.5-B/§4：_allocation_dir = rot90(f3_cad_alloc_dir)（臨街向）；
                        _side_mid = SIDE_LINE 中點；_W_prev = 前一筆累積 W。三者驅動
                        solve_G_binary 內 W（沿 ALLOC 法向累積）與 Rw 差額（§3）。
                        """
                        if _blk_poly is not None and _d_hat is not None and _baseline_pt is not None:
                            try:
                                _r = solve_G_binary(
                                    a=_a_m2, A=_A, B=B_value, C=C_for_calc,
                                    l_front=_l_front, l_side=_l_side, F=_F,
                                    block_poly=_blk_poly, d_hat=_d_hat,
                                    baseline_pt=_baseline_pt,
                                    S_max_limit=_S_max,
                                    is_corner=_is_corner,
                                    side_label=_side if _side in ('左側', '右側') else '左側',
                                    tol=0.01, max_iter=80,
                                    allocation_dir=_allocation_dir,
                                    side_mid=_side_mid, W_prev=_W_prev,
                                )
                                return _r, '幾何二分法'
                            except Exception:
                                pass
                        # fallback：代數迭代（同樣攜帶 W_prev 累積差額，§4）
                        _r = iterate_G_S(
                            a=_a_m2, A=_A, B=B_value, C=C_for_calc,
                            l_front=_l_front, l_side=_l_side, F=_F, W=0.0,
                            avg_depth=_avg_depth,
                            is_corner=_is_corner,
                            tab6_total_burden=_tab6_burden,
                            W_prev=_W_prev,
                        )
                        _r['area_geom'] = round(_r.get('S', 0) * _avg_depth, 2)
                        _r['cut_coords'] = []
                        return _r, '代數迭代(fallback)'

                    st.session_state['f3_wd2_pool_diag'] = {}   # 🆕 W-D.2 §3：每輪重建（防殘留舊塊）
                    for blk_label, parcels_in_blk in parcels_by_block.items():
                        blk_meta = block_meta_by_label.get(blk_label, {})
                        blk_poly = blk_meta.get('shapely', None)
                        blk_area = float(blk_meta.get('area_m2', 0.0) or 0.0)
                        sb_row = sb_rows_by_label.get(blk_label, {})
                        front_len = float(sb_row.get('正面長度(m)', 0.0) or 0.0)
                        l_front = float(sb_row.get('正街尺度', 0.0) or 0.0)
                        # 🆕 W-C §5a：avg_depth 用沿 ALLOC 量之 D_avg（含逐街廓覆寫）；缺則退 area/front_len
                        avg_depth_default = float(
                            (st.session_state.get('f3_alloc_depth_by_label', {}) or {}).get(blk_label)
                            or ((blk_area / front_len) if front_len > 0 else 0.0))

                        # 街廓 d_hat / 角點：依首筆街角地的「街角側別」決定
                        first_corner_side = None
                        for tp in parcels_in_blk:
                            p = _params_for_g.get(tp['暫編地號'], {})
                            if p.get('is_corner') and p.get('side', '無') in ('左側', '右側'):
                                first_corner_side = p['side']
                                break
                        side_for_block = first_corner_side or '左側'

                        d_hat = None; corner_pt = None; S_block_max = front_len or 100.0
                        allocation_dir_block = None   # 🆕 Phase 11 v2

                        # ── 🆕 Phase 11 v2：FRONT_LINE 為 d_hat + corner_pt 之主來源 ──
                        # 重劃後分配土地必須臨正面道路 → 推進方向 d_hat 必須沿 FRONT_LINE
                        _cad_fl_blk = (st.session_state.get('f3_cad_front_lines', {}) or {}).get(blk_label, {})
                        if _cad_fl_blk and _cad_fl_blk.get('p1') and _cad_fl_blk.get('p2'):
                            try:
                                _p1_fl = _cad_fl_blk['p1']; _p2_fl = _cad_fl_blk['p2']
                                _dx_fl = _p2_fl[0] - _p1_fl[0]; _dy_fl = _p2_fl[1] - _p1_fl[1]
                                _L_fl = (_dx_fl ** 2 + _dy_fl ** 2) ** 0.5
                                if _L_fl > 0.1:
                                    d_hat = _np_d.array([_dx_fl / _L_fl, _dy_fl / _L_fl])
                                    corner_pt = _np_d.array(_p1_fl, dtype=float)
                                    S_block_max = float(_L_fl)
                            except Exception:
                                d_hat = None; corner_pt = None

                        # ── 🆕 W-C §0.5-B/§0.5-C：allocation_dir = rot90(f3_cad_alloc_dir) ──
                        # f3_cad_alloc_dir[blk] = 地界線/宗地分配線方向 (ux,uy)；rot90 → (-uy,ux)
                        # = 臨街向（= W 量測軸 n_alloc），使 _block_strip 之 n_hat ∥ 地界線。
                        # 廢除舊「BASELINE/MBR 當 allocation_dir」與 V13 強制校正 hack（§0.5-C）。
                        _alloc_dir_cad = (st.session_state.get('f3_cad_alloc_dir', {})
                                          or {}).get(blk_label)
                        allocation_dir_block = alloc_normal_axis(_alloc_dir_cad)
                        if allocation_dir_block is None:
                            st.warning(
                                f"⚠️ 街廓 {blk_label} 缺宗地分配線方向（f3_cad_alloc_dir），"
                                "W／Rw 無法量測（側街負擔以 0 計），請補畫 ALLOC_LINE 後重出 DXF"
                            )
                        # BASELINE 僅保留為 d_hat fallback（FRONT_LINE 缺時的推進方向），不再當 allocation_dir
                        _mbl_block = (st.session_state.get('f3_manual_baseline', {}) or {}).get(blk_label, {})
                        if (d_hat is None and _mbl_block.get('enabled')
                                and _mbl_block.get('point') and blk_poly is not None):
                            try:
                                _ang_bl = _math_d.radians(float(_mbl_block.get('angle_deg', 0.0)))
                                d_hat = _np_d.array([_math_d.cos(_ang_bl), _math_d.sin(_ang_bl)])
                                corner_pt = _np_d.array(_mbl_block['point'], dtype=float)
                                S_block_max = _block_max_S(blk_poly, corner_pt, d_hat)
                            except Exception:
                                pass

                        # ── 🆕 §0.5-C/F.2：缺 FRONT_LINE/BASELINE → 不退 MBR（fail-loud）──
                        # d_hat（推進向）唯一來源 = FRONT_LINE（次選 BASELINE）；廢 MBR 自動偵測
                        # （_get_block_d_hat MBR 長邊會讓斜街廓帶鋪歪、端點楔形空白）。
                        if d_hat is None or corner_pt is None:
                            st.warning(
                                f"⚠️ 街廓 {blk_label} 缺 FRONT_LINE（且無 BASELINE）→ 無推進方向 d_hat，"
                                "退化為輸入順序、不做雙向夾擠（**不退 MBR**）。請補畫 FRONT_LINE 後重出 DXF。"
                            )

                        # 🆕 W-C §0.5-B 軸向「真檢查」（KL 指正：原檢查是套套邏輯，恆真、抓不到缺 ALLOC）
                        # 真正要驗的：f3_cad_alloc_dir[blk]（ALLOC=地界線/深度向）⊥ 該塊 FRONT_LINE
                        #   → 夾角應 ≈90°（§0.5-E V6.dxf 真值：六塊皆 ⊥）。<80° 即 ALLOC 配錯/缺。
                        try:
                            _diag_dict = st.session_state.setdefault(
                                'f3_g_iter_diagnostics', {})
                            # 取該塊 FRONT_LINE 真方向（非 d_hat，d_hat 可能退 BASELINE）
                            _front_dir = None
                            if _cad_fl_blk and _cad_fl_blk.get('p1') and _cad_fl_blk.get('p2'):
                                _fp1 = _cad_fl_blk['p1']; _fp2 = _cad_fl_blk['p2']
                                _fdx = _fp2[0] - _fp1[0]; _fdy = _fp2[1] - _fp1[1]
                                _fL = (_fdx ** 2 + _fdy ** 2) ** 0.5
                                if _fL > 1e-6:
                                    _front_dir = (_fdx / _fL, _fdy / _fL)
                            _perp_deg = None; _perp_ok = None
                            if _alloc_dir_cad is not None and _front_dir is not None:
                                _ax, _ay = float(_alloc_dir_cad[0]), float(_alloc_dir_cad[1])
                                _an = (_ax * _ax + _ay * _ay) ** 0.5
                                if _an > 1e-9:
                                    _cosaf = abs((_ax / _an) * _front_dir[0] + (_ay / _an) * _front_dir[1])
                                    _cosaf = max(-1.0, min(1.0, _cosaf))
                                    _perp_deg = round(_math_d.degrees(_math_d.acos(_cosaf)), 1)  # ≈90 為佳
                                    _perp_ok = bool(_perp_deg >= 80.0)
                            if _perp_ok is False:
                                st.error(
                                    f"🔴 街廓 {blk_label}：f3_cad_alloc_dir 與 FRONT_LINE 夾角 {_perp_deg}°"
                                    "（應≈90°）→ ALLOC_LINE 疑配錯/歸錯塊（非 ⊥FRONT），W/Rw/界線方向不可信，"
                                    "請檢查 DXF 該塊 ALLOC_LINE。"
                                )
                            _diag_dict[blk_label] = {
                                'd_hat': (d_hat.tolist() if d_hat is not None else None),
                                'allocation_dir': (allocation_dir_block.tolist()
                                                   if allocation_dir_block is not None else None),
                                'ALLOC⊥FRONT_deg': _perp_deg,
                                'ALLOC⊥FRONT_ok': _perp_ok,
                                'note': ('✅ ALLOC ⊥ FRONT（軸向正確）' if _perp_ok
                                         else ('⚠️ 缺 ALLOC 或 FRONT，無法驗（W/Rw=0）' if _perp_ok is None
                                               else '🔴 ALLOC 非 ⊥FRONT，疑配錯塊')),
                            }
                        except Exception:
                            pass

                        # ═══════════════════════════════════════════════════════════
                        # 🚨 Patch E-2 → 🆕 W-D.2 §3：v2 排序 + 滑池槽 k* 切分推進
                        # 廢除：v1 兼容層、winner insert(0)、舊 fallback；
                        #       「KL 規範 N/2 中點規則」（過渡作法，KL 2026-07-05 裁可刪）
                        #       ——池槽唯一來源＝_select_pool_slot（J 最大化，D-1 bootstrap）。
                        # 設計：
                        #   v2.ordered = [位次1=p1端winner, ..., 位次N=p2端winner]
                        #   left_group  = ordered[0:k]                     ← 從 p1 端推
                        #   right_group = reversed(ordered[k:])            ← 從 p2 端推
                        #   池＝雙向中間剩餘；k 由基準趟真寬度餵 _select_pool_slot 決定
                        # ═══════════════════════════════════════════════════════════
                        _v2_res = None
                        _degenerate_order = (d_hat is None or corner_pt is None)
                        if _degenerate_order:
                            # 沒有幾何資訊 → 退化為輸入順序、不做雙向夾擠（單趟、全左群）
                            ordered_v2 = []
                            for tp in parcels_in_blk:
                                _pm = _params_for_g.get(tp['暫編地號'], {})
                                ordered_v2.append({
                                    'tp': tp,
                                    'pre_position': 0,
                                    'is_corner_winner': bool(_pm.get('is_corner', False)),
                                    'is_first_corner_marker': bool(_pm.get('is_corner', False)),
                                    'side': '中段',
                                })
                            ordered_v2.sort(key=lambda e: (not e['is_corner_winner'],))
                        else:
                            # 從 session 取 PK winner / forced_offset / FRONT_LINE 端點
                            _v2_pk_winners = (
                                st.session_state.get('f3_corner_winners', {}) or {}
                            ).get(blk_label, {}) or {}
                            _v2_forced = (
                                st.session_state.get('f3L_forced_offset', {}) or {}
                            ).get(blk_label, {}) or {}
                            _v2_cad_fl = (
                                st.session_state.get('f3_cad_front_lines', {}) or {}
                            ).get(blk_label, {}) or {}
                            _v2_fl_p1 = _v2_cad_fl.get('p1')
                            _v2_fl_p2 = _v2_cad_fl.get('p2')
                            if not _v2_fl_p1 or not _v2_fl_p2:
                                _v2_fl_p1 = (float(corner_pt[0]), float(corner_pt[1]))
                                _v2_fl_p2 = (
                                    float(corner_pt[0]) + float(d_hat[0]) * float(S_block_max),
                                    float(corner_pt[1]) + float(d_hat[1]) * float(S_block_max),
                                )

                            # 呼叫 v2（不再 except fallback；失敗就讓它噴錯）
                            _v2_res = _spatial_order_parcels_v2(
                                parcels_in_block=parcels_in_blk,
                                d_hat=d_hat,
                                front_line_p1=_v2_fl_p1,
                                front_line_p2=_v2_fl_p2,
                                pk_winners=_v2_pk_winners,
                                forced_offset=_v2_forced,
                            )
                            ordered_v2 = list(_v2_res.get('ordered', []) or [])

                        # 🆕 W-D.2 §3：註記原位次 index（基準趟寬度→_select_pool_slot 映射用）。
                        #   k 切分與 side 標籤依 k 而變 → 移入 _advance_block_with_split（趟內建）。
                        for _i_ov2, _e_ov2 in enumerate(ordered_v2):
                            _e_ov2['_ov2_idx'] = _i_ov2

                        # ── 🆕 Phase C：forced_offset 預留（街角強制抵費地 buffer）──
                        # 若 PK 結果為「強制抵費地」→ 該側街角範圍預留為抵費地，
                        #   不參與分配；雙向推進之 cum_S 起始值即等於 buffer 寬度
                        # buffer 寬度 = 街角最小分配面積 ÷ 街廓平均深度
                        _fo_block = (st.session_state.get(
                            'f3L_forced_offset', {}) or {}).get(blk_label, {})
                        _fo_left = bool(_fo_block.get('left_forced_offset', False))
                        _fo_right = bool(_fo_block.get('right_forced_offset', False))
                        # 從 f3L_corner_min_table 讀取最小面積
                        _corner_min_table = st.session_state.get('f3L_corner_min_table', []) or []
                        _row_for_buffer = next(
                            (r for r in _corner_min_table if r.get('街廓') == blk_label), None
                        )
                        _left_buffer_S = 0.0
                        _right_buffer_S = 0.0
                        if _row_for_buffer and avg_depth_default > 0:
                            if _fo_left:
                                _l_min = _row_for_buffer.get('【左】街角最小面積(㎡)')
                                try:
                                    if _l_min is not None and _l_min != float('inf'):
                                        _left_buffer_S = float(_l_min) / avg_depth_default
                                except (TypeError, ValueError):
                                    _left_buffer_S = 0.0
                            if _fo_right:
                                _r_min = _row_for_buffer.get('【右】街角最小面積(㎡)')
                                try:
                                    if _r_min is not None and _r_min != float('inf'):
                                        _right_buffer_S = float(_r_min) / avg_depth_default
                                except (TypeError, ValueError):
                                    _right_buffer_S = 0.0
                        # 顯示通知
                        if _left_buffer_S > 0 or _right_buffer_S > 0:
                            _msg_buf = []
                            if _left_buffer_S > 0:
                                _msg_buf.append(f"左側 {_left_buffer_S:.2f}m")
                            if _right_buffer_S > 0:
                                _msg_buf.append(f"右側 {_right_buffer_S:.2f}m")
                            st.info(
                                f"🟡 街廓 {blk_label}：強制抵費地 buffer 預留 "
                                f"{', '.join(_msg_buf)}（不參與分配，將自動成為抵費地）"
                            )
                        # 🆕 W-C §0.5-F.3：記錄強制街角抵費地狀態（供診斷；該 buffer 區應 render 成橘色抵費地）
                        st.session_state.setdefault('f3_forced_offset_diag', {})[blk_label] = {
                            'left_forced': _fo_left, 'right_forced': _fo_right,
                            'left_buffer_S': round(_left_buffer_S, 2),
                            'right_buffer_S': round(_right_buffer_S, 2),
                        }

                        # ── 🆕 W-C §3：角側區塊常數（F、l_side、SIDE_LINE 中點、n_alloc）──
                        # 側街負擔 Rw·F·l_side 攤到「角側全筆」（非僅第1筆）；F、l_side 為街廓常數。
                        # n_alloc = allocation_dir_block = rot90(f3_cad_alloc_dir) = W 量測軸（§0.5-B）。
                        _side_lines_blk = (st.session_state.get(
                            'f3_cad_side_lines_by_side', {}) or {}).get(blk_label, {}) or {}
                        _sl_left = _side_lines_blk.get('left') or {}
                        _sl_right = _side_lines_blk.get('right') or {}
                        _side_mid_left = _sl_left.get('mid')
                        _side_mid_right = _sl_right.get('mid')
                        _has_left_corner = _side_mid_left is not None
                        _has_right_corner = _side_mid_right is not None
                        _F_left = float(sb_row.get('左側長度(m)', 0.0) or 0.0) if _has_left_corner else 0.0
                        _lside_left = float(sb_row.get('左側尺度', 0.0) or 0.0) if _has_left_corner else 0.0
                        _F_right = float(sb_row.get('右側長度(m)', 0.0) or 0.0) if _has_right_corner else 0.0
                        _lside_right = float(sb_row.get('右側尺度', 0.0) or 0.0) if _has_right_corner else 0.0
                        _n_alloc_blk = allocation_dir_block   # rot90(f3_cad_alloc_dir)
                        # 臨街投影係數 _cos_dn：S 增量 → W 增量（W += S·|d_hat·n_alloc|）。
                        # 用於 thread W_前 初值（forced_offset buffer 之臨街寬）與一致性。
                        _cos_dn = 1.0
                        if _n_alloc_blk is not None and d_hat is not None:
                            try:
                                _dh0 = _np_d.asarray(d_hat, dtype=float)
                                _dl0 = float(_np_d.linalg.norm(_dh0))
                                if _dl0 > 1e-9:
                                    _cos_dn = abs(float(_np_d.dot(
                                        _dh0 / _dl0, _np_d.asarray(_n_alloc_blk, dtype=float))))
                            except Exception:
                                _cos_dn = 1.0
                        # 🆕 W-C §1-4：本街廓畸零地最小面寬（判去留用 W=單筆宗地寬度，非 S、非累積）
                        _mw_blk = float((st.session_state.get('f3_min_width_by_label', {})
                                         or {}).get(blk_label, 0.0) or 0.0)

                        def _mark_zaling(_res):
                            """單筆宗地寬度 = S·|d_hat·n_alloc|（相鄰兩地界線垂距）；< 最小面寬→第1調配順位旗標。"""
                            _pw = float(_res.get('S', 0.0)) * _cos_dn
                            _res['_宗地寬度'] = round(_pw, 2)
                            _res['_畸零旗標'] = ('⚠️移出/第1調配順位'
                                                 if (_mw_blk > 0 and _pw < _mw_blk) else '')
                            return _res

                        # ── 🆕 W-D.2 §3（D-1 bootstrap 載體）：單塊可重入推進 ──
                        def _advance_block_with_split(_k_split, _commit):
                            """依池槽 _k_split 切分後執行雙向推進（Task D-2 左右迴圈整段內移）。
                            _k_split：左群=ordered_v2[:k]、右群=reversed(ordered_v2[k:])、池插中間；
                                      _degenerate_order 時＝全左群（沿既有退化語意）。
                            _commit：False＝基準趟（僅取真寬度 w_i）——不寫 g_rows/detail_trace、
                                     所有 st.* 訊息靜默（防雙趟重複）；True＝正式趟。
                            回傳 {'rows','trace','widths'(原位次 左→右),'left_cum_S','right_cum_S',
                                  'left_results','right_results'}。不改 a、不寫 session_state。"""
                            _N_f = len(ordered_v2)
                            if _N_f == 0:
                                left_group = []; right_group = []
                            elif _degenerate_order or _k_split >= _N_f:
                                left_group = list(ordered_v2); right_group = []
                            elif _k_split <= 0:
                                left_group = []; right_group = list(reversed(ordered_v2))
                            else:
                                left_group = list(ordered_v2[:_k_split])
                                right_group = list(reversed(ordered_v2[_k_split:]))
                            # v1 風格 'side' 標籤（規則同舊；群組依 k 而變故在趟內標）
                            for _idx_l, _e in enumerate(left_group):
                                if _idx_l == 0 and _e.get('is_first_corner_marker', False):
                                    _e['side'] = '左側'
                                else:
                                    _e['side'] = '無'
                            for _idx_r, _e in enumerate(right_group):
                                if _idx_r == 0 and _e.get('is_first_corner_marker', False):
                                    _e['side'] = '右側'
                                else:
                                    _e['side'] = '無'

                            _rows_local = []
                            _trace_local = {}
                            _widths_local = [0.0] * _N_f

                            # ── Task D-2：左側推進（d_hat 正向）──
                            # 🆕 Phase C：若左側 forced_offset → 從 buffer 寬度起算（跳過街角）
                            left_cum_S = float(_left_buffer_S)
                            right_cum_S = float(_right_buffer_S)   # 同理右側
                            # 🆕 W-C §4：thread 累積 W_前（首筆=0；forced_offset 時=buffer 臨街寬）
                            _W_prev_left = (_left_buffer_S * _cos_dn) if _has_left_corner else 0.0
                            _W_prev_right = (_right_buffer_S * _cos_dn) if _has_right_corner else 0.0
                            first_corner_used_left = False
                            left_results = []
                            for entry in left_group:
                                tp = entry['tp']
                                k = tp['暫編地號']
                                # 🚨 W-A 規格一：a_for_G = 分攤登記面積_m2 + 面積_m2(a' 累加器)
                                if '分攤登記面積_m2' in tp:
                                    a_m2 = round(float(tp.get('分攤登記面積_m2', 0) or 0)
                                                 + float(tp.get('面積_m2', 0) or 0), 2)
                                else:
                                    a_m2 = round(float(tp.get('面積_m2', 0) or 0), 2)
                                side = entry.get('side', '無')
                                # 🚨 Patch E-2.3：用 v2 entry 的 is_first_corner_marker 旗標
                                # 不再依賴 side 字串比對，因為 left_group 第 1 筆若是 winner 就是 first_corner
                                is_corner_marked = bool(entry.get('is_corner_winner', False))
                                is_first_corner_l = (
                                    bool(entry.get('is_first_corner_marker', False))
                                    and not first_corner_used_left
                                )
                                # 🆕 W-C §3：左側角側全筆皆套街廓常數 F/l_side（非僅第1筆）
                                if _has_left_corner:
                                    l_side_use = _lside_left; F_use = _F_left
                                else:
                                    l_side_use = 0.0; F_use = 0.0
                                zone = tp.get('重劃前地價區段', '')
                                post_p = post_price_by_block.get(blk_label, 0.0)
                                pre_p = pre_price_by_zone.get(zone, 0.0)
                                A_ratio = (post_p / pre_p) if (pre_p > 0 and post_p > 0) else 1.0
                                S_remain = max(0.1, S_block_max - left_cum_S - right_cum_S)
                                baseline_pt = (corner_pt + left_cum_S * d_hat
                                               if (d_hat is not None and corner_pt is not None) else None)
                                res, solver_label = _solve_one(
                                    a_m2, A_ratio, l_front, l_side_use, F_use,
                                    blk_poly, d_hat, baseline_pt, S_remain,
                                    is_first_corner_l, side, avg_depth_default,
                                    _allocation_dir=allocation_dir_block,
                                    _side_mid=(_side_mid_left if _has_left_corner else None),
                                    _W_prev=_W_prev_left,
                                )
                                if _has_left_corner:   # thread 累積 W_前 給下一筆
                                    _W_prev_left = float(res.get('W_far', _W_prev_left))
                                _S_actual = float(res.get('S', 0.0))
                                # 極端防呆 3：S_remain 觸頂
                                _G_target = float(res.get('G', 0.0))
                                _area_actual = float(res.get('area_geom', 0.0))
                                if (abs(_S_actual - S_remain) < 0.05 and _G_target > 0
                                    and _area_actual < _G_target * 0.95):
                                    res['是否收斂_override'] = '⚠️ 空間不足(夾擠限制)'
                                left_cum_S += _S_actual
                                res['_alloc_cum_S'] = left_cum_S
                                _mark_zaling(res)   # 🆕 §1-4 判去留旗標
                                _widths_local[entry['_ov2_idx']] = float(
                                    res.get('_宗地寬度', 0.0) or 0.0)   # 🆕 W-D.2 真寬度（D-1）
                                if is_first_corner_l:
                                    first_corner_used_left = True
                                _rows_local.append(_build_g_row(
                                    k, tp, blk_label, blk_area, front_len, avg_depth_default,
                                    zone, A_ratio, l_front, l_side_use, F_use, is_corner_marked,
                                    is_first_corner_l, side, res, solver_label, 'left',
                                ))
                                _trace_local[k] = res.get('trace', [])
                                left_results.append((entry, res))

                            # ── Task D-2：右側推進（d_hat 反向）──
                            # 補充 1（最後一哩路）：以街廓所有頂點在 d_hat 方向最大投影值精準定 end_pt，
                            # 杜絕菱形/三角形/狹長傾斜街廓的懸空問題（無需試探）
                            if d_hat is not None and corner_pt is not None and blk_meta.get('vertices'):
                                try:
                                    _proj_pts = [
                                        float(_np_d.dot(
                                            _np_d.array([v[0] - corner_pt[0], v[1] - corner_pt[1]]),
                                            d_hat
                                        )) for v in blk_meta['vertices']
                                    ]
                                    actual_max_proj = max(_proj_pts) if _proj_pts else S_block_max
                                except Exception:
                                    actual_max_proj = S_block_max
                                end_pt = corner_pt + actual_max_proj * d_hat
                                d_hat_rev = -d_hat
                            else:
                                actual_max_proj = S_block_max
                                end_pt = None
                                d_hat_rev = None

                            first_corner_used_right = False
                            right_results = []
                            for entry in right_group:
                                tp = entry['tp']
                                k = tp['暫編地號']
                                # 🚨 W-A 規格一：a_for_G = 分攤登記面積_m2 + 面積_m2(a' 累加器)
                                if '分攤登記面積_m2' in tp:
                                    a_m2 = round(float(tp.get('分攤登記面積_m2', 0) or 0)
                                                 + float(tp.get('面積_m2', 0) or 0), 2)
                                else:
                                    a_m2 = round(float(tp.get('面積_m2', 0) or 0), 2)
                                side = entry.get('side', '無')
                                # 🚨 Patch E-2.4：與 E-2.3 對稱
                                is_corner_marked = bool(entry.get('is_corner_winner', False))
                                is_first_corner_r = (
                                    bool(entry.get('is_first_corner_marker', False))
                                    and not first_corner_used_right
                                )
                                # 🆕 W-C §3：右側角側全筆皆套街廓常數 F/l_side（非僅第1筆）
                                if _has_right_corner:
                                    l_side_use = _lside_right; F_use = _F_right
                                else:
                                    l_side_use = 0.0; F_use = 0.0
                                zone = tp.get('重劃前地價區段', '')
                                post_p = post_price_by_block.get(blk_label, 0.0)
                                pre_p = pre_price_by_zone.get(zone, 0.0)
                                A_ratio = (post_p / pre_p) if (pre_p > 0 and post_p > 0) else 1.0
                                S_remain = max(0.1, actual_max_proj - left_cum_S - right_cum_S)
                                baseline_pt = (end_pt + right_cum_S * d_hat_rev
                                               if (d_hat_rev is not None and end_pt is not None) else None)
                                res, solver_label = _solve_one(
                                    a_m2, A_ratio, l_front, l_side_use, F_use,
                                    blk_poly, d_hat_rev, baseline_pt, S_remain,
                                    is_first_corner_r, side, avg_depth_default,
                                    _allocation_dir=allocation_dir_block,
                                    _side_mid=(_side_mid_right if _has_right_corner else None),
                                    _W_prev=_W_prev_right,
                                )
                                # 極端防呆 2 後援：右側起點數值微修
                                if (float(res.get('area_geom', 0)) < 0.5
                                    and d_hat_rev is not None and baseline_pt is not None):
                                    for _adj in (0.1, 0.3, 0.5):
                                        _try_pt = baseline_pt + _adj * d_hat_rev
                                        _try_S = max(0.1, S_remain - _adj)
                                        _r2, _sl2 = _solve_one(
                                            a_m2, A_ratio, l_front, l_side_use, F_use,
                                            blk_poly, d_hat_rev, _try_pt, _try_S,
                                            is_first_corner_r, side, avg_depth_default,
                                            _allocation_dir=allocation_dir_block,
                                            _side_mid=(_side_mid_right if _has_right_corner else None),
                                            _W_prev=_W_prev_right,
                                        )
                                        if float(_r2.get('area_geom', 0)) >= 0.5:
                                            res, solver_label = _r2, _sl2
                                            if _commit:   # 🆕 W-D.2 reviewer WARNING：深巢 st.info 顯式 gate
                                                st.info(
                                                    f"ℹ️ 街廓 {blk_label} 右側起點數值微修 {_adj}m 後成功切出土地"
                                                )
                                            break
                                if _has_right_corner:   # thread 累積 W_前 給下一筆
                                    _W_prev_right = float(res.get('W_far', _W_prev_right))
                                _S_actual = float(res.get('S', 0.0))
                                # 極端防呆 3：S_remain 觸頂
                                _G_target = float(res.get('G', 0.0))
                                _area_actual = float(res.get('area_geom', 0.0))
                                if (abs(_S_actual - S_remain) < 0.05 and _G_target > 0
                                    and _area_actual < _G_target * 0.95):
                                    res['是否收斂_override'] = '⚠️ 空間不足(夾擠限制)'
                                right_cum_S += _S_actual
                                res['_alloc_cum_S'] = right_cum_S
                                _mark_zaling(res)   # 🆕 §1-4 判去留旗標
                                _widths_local[entry['_ov2_idx']] = float(
                                    res.get('_宗地寬度', 0.0) or 0.0)   # 🆕 W-D.2 真寬度（D-1）
                                if is_first_corner_r:
                                    first_corner_used_right = True
                                _rows_local.append(_build_g_row(
                                    k, tp, blk_label, blk_area, front_len, avg_depth_default,
                                    zone, A_ratio, l_front, l_side_use, F_use, is_corner_marked,
                                    is_first_corner_r, side, res, solver_label, 'right',
                                ))
                                _trace_local[k] = res.get('trace', [])
                                right_results.append((entry, res))

                            return {
                                'rows': _rows_local, 'trace': _trace_local,
                                'widths': _widths_local,
                                'left_cum_S': left_cum_S, 'right_cum_S': right_cum_S,
                                'left_results': left_results, 'right_results': right_results,
                            }

                        # ── 🆕 W-D.2 J-fix：warn-on-zero-weight（no-silent-fallback）──
                        #   有 SIDE_LINE 之側其 J 權重 F×l₁ 必 >0；零＝上游參數層靜默缺
                        #   （Step E 該街廓側街「新闢」未勾/路寬未設 → 尺度 0 → J 死值、
                        #    G 公式側街負擔項**同源**歸零）。具名警示、不得靜默吞 0。
                        for _side_nm_w2, _has_w2, _F_w2, _l1_w2 in (
                                ('左', _has_left_corner, _F_left, _lside_left),
                                ('右', _has_right_corner, _F_right, _lside_right)):
                            if _has_w2 and (_F_w2 * _l1_w2) <= 0.0:
                                st.warning(
                                    f"⚠️ 街廓 {blk_label} {_side_nm_w2}側：有 SIDE_LINE 但滑池槽權重 "
                                    f"F×l₁ = {_F_w2:.2f}×{_l1_w2:.2f} = 0 → J 退化為 0"
                                    "（選槽將僅剩 min-dev 居中），且 G 公式側街負擔項同源歸零。"
                                    "請檢查**步驟 E** 該街廓側街「路寬／新闢道路」設定"
                                    "（尺度=0 多因未勾新闢；參數正典見 verify/case_params 快照）後重跑。"
                                )

                        # ── 🆕 W-D.2 §3 選槽 orchestration（D-1：基準趟→k*→正式趟）──
                        _N = len(ordered_v2)
                        _k_naive = (_N + 1) // 2 if _N % 2 == 1 else _N // 2   # 僅作基準趟切點
                        _slot_res = None
                        if _degenerate_order or _N <= 1:
                            # 無選槽自由度 → 單趟正式（退化語意不變）
                            _k_star = _N
                            _adv_final = _advance_block_with_split(_k_star, True)
                        else:
                            # ① 基準趟（k=naive、不落 rows）取真寬度 w_i（⊥ALLOC；D-1 Option A）
                            _adv_base = _advance_block_with_split(_k_naive, False)
                            # ② 真寬度餵 _select_pool_slot（Q2：b＝buffer_S×cos_dn 入 W 軸；
                            #    Q3：F/l1 與推進迴圈同源＝_F_left/_lside_left/_F_right/_lside_right）
                            _slot_res = _select_pool_slot(
                                _adv_base['widths'],
                                {'has': _has_left_corner, 'F': _F_left,
                                 'l1': _lside_left, 'b': _left_buffer_S * _cos_dn},
                                {'has': _has_right_corner, 'F': _F_right,
                                 'l1': _lside_right, 'b': _right_buffer_S * _cos_dn},
                            )
                            _k_star = int(_slot_res['k'])
                            # 停機②（J 下降）看守：argmax 保證 J(k*)≥J(naive)；破＝實作 bug
                            _J_by_k = {t['k']: t['J'] for t in _slot_res['table']}
                            if (_k_naive in _J_by_k
                                    and _J_by_k.get(_k_star, 0.0) < _J_by_k[_k_naive] - 1e-9):
                                st.error(
                                    f"🔴 停機②（J 下降）街廓 {blk_label}：J(k*={_k_star})="
                                    f"{_J_by_k.get(_k_star, 0.0):.4f} < J(naive={_k_naive})="
                                    f"{_J_by_k[_k_naive]:.4f} — 滑池槽最佳化反變糟，停、上呈 KL＋claude.ai。"
                                )
                            # ③ 正式趟（k*）才落 rows
                            _adv_final = _advance_block_with_split(_k_star, True)
                        g_rows.extend(_adv_final['rows'])
                        detail_trace.update(_adv_final['trace'])
                        left_cum_S = _adv_final['left_cum_S']
                        right_cum_S = _adv_final['right_cum_S']
                        left_results = _adv_final['left_results']
                        right_results = _adv_final['right_results']

                        # ── 雙向重疊警告 ──
                        if left_cum_S + right_cum_S > S_block_max + 0.5:
                            st.warning(
                                f"⚠️ 街廓 {blk_label}：左 ({left_cum_S:.2f}m) + "
                                f"右 ({right_cum_S:.2f}m) = {left_cum_S + right_cum_S:.2f}m "
                                f"已超過街廓深度 {S_block_max:.2f}m，可能有重疊區"
                            )

                        # ── Task D-3：抵費地（Offset Land）幾何自動生成 ──
                        _pool_total_blk = None   # 🆕 W-D.2 ledger：幾何剩餘總量（None＝無幾何/失敗）
                        if blk_poly is not None:
                            try:
                                allocated_polys = []
                                for _entry, _res in (left_results + right_results):
                                    _coords = _res.get('cut_coords') or []
                                    if len(_coords) >= 3:
                                        try:
                                            _p = _SP_d(_coords)
                                            if not _p.is_valid:
                                                _p = _p.buffer(0)
                                            if not _p.is_empty and _p.area >= 0.5:
                                                allocated_polys.append(_p)
                                        except Exception:
                                            continue
                                if allocated_polys:
                                    # 補充 2：buffer(0.001) 消除浮點碎邊
                                    allocated_union = _uunion_d(allocated_polys).buffer(0.001)
                                    offset_land = blk_poly.difference(allocated_union)
                                    if hasattr(offset_land, 'is_valid') and not offset_land.is_valid:
                                        offset_land = offset_land.buffer(0)
                                else:
                                    offset_land = blk_poly

                                # 處理 MultiPolygon 與面積過濾
                                if offset_land.geom_type == 'MultiPolygon':
                                    parts = sorted(offset_land.geoms,
                                                   key=lambda g: g.area, reverse=True)
                                    offset_geoms = [g for g in parts if g.area >= 1.0]
                                elif offset_land.geom_type == 'Polygon':
                                    offset_geoms = ([offset_land]
                                                    if offset_land.area >= 1.0 else [])
                                else:
                                    offset_geoms = []

                                _pool_total_blk = float(sum(_g.area for _g in offset_geoms))  # 🆕 W-D.2 ledger

                                # 極端防呆 4：抵費地碎裂孤島警告
                                if len(offset_geoms) > 1:
                                    _total_a = sum(g.area for g in offset_geoms)
                                    st.warning(
                                        f"⚠️ 街廓 {blk_label} 的抵費地因分配夾擠被分割成 "
                                        f"{len(offset_geoms)} 塊不連續區域（總面積 {_total_a:.2f} ㎡），"
                                        f"請檢視圖形是否合理，或考量手動調整分配基準線。"
                                    )

                                # 寫入 g_rows 作為「抵費地」項目
                                _min_block = min(50.0, blk_area * 0.05)   # 抵費地最小門檻：街廓 5% 或 50㎡
                                for _i, _g in enumerate(offset_geoms):
                                    _suffix = '' if len(offset_geoms) == 1 else f'-{_i+1}'
                                    _conv_flag = ('🟡' if _g.area >= _min_block
                                                  else '⚠️ < 最小分配')
                                    try:
                                        _coords_list = [[float(c[0]), float(c[1])]
                                                        for c in list(_g.exterior.coords)]
                                    except Exception:
                                        _coords_list = []
                                    g_rows.append({
                                        '暫編地號': f'{blk_label}-抵費地{_suffix}',
                                        '原地號': '—',
                                        '所屬街廓': blk_label,
                                        '重劃前區段': '—',
                                        'a 面積(㎡)': 0.0,
                                        '街廓面積(㎡)': round(blk_area, 2),
                                        '正面長度(m)': round(front_len, 2),
                                        '平均深度(m)': round(avg_depth_default, 2),
                                        'A 地價比': 0.0,
                                        'l₂ 正面尺度': 0.0,
                                        'l₁ 側面尺度': 0.0,
                                        '街角地': '—',
                                        '第1筆街角': '—',
                                        '街角側別': '—',
                                        'F(m)': 0.0,
                                        'W(m)': 0.0,
                                        'Rw(%)': 0.0,
                                        'S(m)': 0.0,
                                        '幾何面積(㎡)': round(_g.area, 2),
                                        'G(㎡)': 0.0,
                                        '累積S(m)': 0.0,
                                        '推進側別': '抵費地',
                                        '解法': '幾何剩餘',
                                        '迭代次數': 0,
                                        '是否收斂': _conv_flag,
                                        '負擔比率': 0.0,
                                        'cut_coords': _coords_list,
                                    })
                            except Exception as _eOff:
                                st.warning(f"⚠️ 街廓 {blk_label} 抵費地計算失敗：{_eOff}")

                        # ── 🆕 W-D.2 §3：守恆 ledger（M3 接線・消費端）──
                        # 角落抵費地／中央池＝幾何剩餘之「拆帳呈示」（池重定位、非新增面積）。
                        # 守恆：ΣG（配地）＋池總（幾何剩餘）＝街廓 DXF 面積，殘差 <1㎡。
                        _sum_G_blk = sum(float(r.get('G(㎡)', 0) or 0) for r in _adv_final['rows'])
                        _sum_geom_blk = sum(float(r.get('幾何面積(㎡)', 0) or 0)
                                            for r in _adv_final['rows'])
                        _corner_off_L = (float(_v2_res.get('left_corner_offset_area', 0.0) or 0.0)
                                         if _v2_res else 0.0)
                        _corner_off_R = (float(_v2_res.get('right_corner_offset_area', 0.0) or 0.0)
                                         if _v2_res else 0.0)

                        def _rw_real_wd2(_side_tag):
                            return round(sum(float(r.get('Rw(%)', 0) or 0)
                                             for r in _adv_final['rows']
                                             if r.get('推進側別') == _side_tag
                                             and float(r.get('F(m)', 0) or 0) > 0), 2)
                        _tbl_wd2 = (_slot_res or {}).get('table') or []
                        _row_at = {t['k']: t for t in _tbl_wd2}
                        _t_star = _row_at.get(_k_star, {})
                        _t_naive = _row_at.get(_k_naive, {})
                        if _pool_total_blk is not None:
                            _resid_wd2 = round(_sum_G_blk + _pool_total_blk - blk_area, 2)
                            _verdict_wd2 = ('✅' if abs(_resid_wd2) < 1.0 else '🔴 守恆破')
                        else:
                            _resid_wd2 = None
                            _verdict_wd2 = '—（無街廓幾何）'
                        st.session_state['f3_wd2_pool_diag'][blk_label] = {
                            'n': _N, 'k_naive': _k_naive, 'k*': _k_star,
                            'J(naive)': round(float(_t_naive.get('J', 0.0)), 4),
                            'J(k*)': round(float(_t_star.get('J', 0.0)), 4),
                            'ΣRw_L理論@k*(%)': round(float(_t_star.get('ΣRw_L', 0.0)), 2),
                            'ΣRw_R理論@k*(%)': round(float(_t_star.get('ΣRw_R', 0.0)), 2),
                            'ΣRw_L實跑(%)': _rw_real_wd2('left'),
                            'ΣRw_R實跑(%)': _rw_real_wd2('right'),
                            'ΣG(㎡)': round(_sum_G_blk, 2),
                            'Σ配地幾何(㎡)': round(_sum_geom_blk, 2),
                            '池總=幾何剩餘(㎡)': (round(_pool_total_blk, 2)
                                                  if _pool_total_blk is not None else None),
                            '角落抵費地L(㎡)': round(_corner_off_L, 2),
                            '角落抵費地R(㎡)': round(_corner_off_R, 2),
                            '中央池(㎡)': (round(_pool_total_blk - _corner_off_L - _corner_off_R, 2)
                                           if _pool_total_blk is not None else None),
                            '守恆殘差(㎡)': _resid_wd2,
                            '判定': _verdict_wd2,
                            # 🆕 KL 條件④（2026-07-05）：ledger vs 幾何片對照——ledger 角落＝
                            #   range 規劃值拆帳、幾何片＝實際切片；兩算總量逐塊相等（池總），
                            #   位移非漏帳。此欄防未來誤判。
                            '幾何片明細(㎡)': (str([round(float(_g.area), 2) for _g in offset_geoms])
                                               if _pool_total_blk is not None else ''),
                            '片數': (len(offset_geoms) if _pool_total_blk is not None else 0),
                            'note': ((_slot_res or {}).get('note', '') or
                                     ('degenerate/N≤1 單趟' if (_degenerate_order or _N <= 1) else ''))
                                    + ('；naive 切點不在合法域(pin)，停機②看守略過比較'
                                       if (_slot_res and _k_naive not in _row_at) else ''),
                            'slot_table': [dict(t) for t in _tbl_wd2],
                        }
                        if _verdict_wd2 == '🔴 守恆破':
                            st.error(
                                f"🔴 停機③（守恆破）街廓 {blk_label}：ΣG {_sum_G_blk:.2f}＋池 "
                                f"{_pool_total_blk:.2f} vs 街廓 {blk_area:.2f}"
                                f"（殘差 {_resid_wd2:+.2f}㎡ ≥1㎡）— 停、上呈 KL＋claude.ai。"
                            )

                    # 🆕 V12 模組 1 補強 B：孤立公設地虛擬 G 值結算
                    # 微調防護 1：B/C 優先，total_burden_ratio fallback
                    _orphans_for_g = st.session_state.get('f3_orphan_parcels', []) or []
                    _n_orphan_added = 0
                    if _orphans_for_g:
                        if (B_value is not None and C_for_calc is not None
                            and B_value > 0 and C_for_calc > 0):
                            _burden_for_orphan = float(B_value) + float(C_for_calc)
                        else:
                            _burden_for_orphan = float(st.session_state.get(
                                'f3_total_burden_rate_from_finance', 0.40) or 0.40)
                        _burden_for_orphan = max(0.0, min(0.95, _burden_for_orphan))
                        for _orph in _orphans_for_g:
                            # 🚨 W-A：孤立公設地虛擬 G 用登記軌（分攤登記，幾何 fallback）
                            _a_o = float(_orph.get('分攤登記面積_m2',
                                                   _orph.get('幾何面積_m2',
                                                             _orph.get('面積_m2', 0))) or 0)
                            if _a_o <= 0:
                                continue
                            _G_virtual = _a_o * (1.0 - _burden_for_orphan)
                            g_rows.append({
                                '暫編地號': _orph.get('暫編地號', ''),
                                '原地號': _orph.get('原地號', ''),
                                '所屬街廓': _orph.get('所屬街廓', ''),
                                '重劃前區段': _orph.get('重劃前地價區段', ''),
                                'a 面積(㎡)': round(_a_o, 2),
                                '街廓面積(㎡)': 0.0,
                                '正面長度(m)': 0.0,
                                '平均深度(m)': 0.0,
                                'A 地價比': 1.0,
                                'l₂ 正面尺度': 0.0,
                                'l₁ 側面尺度': 0.0,
                                '街角地': '—',
                                '第1筆街角': '—',
                                '街角側別': '—',
                                'F(m)': 0.0,
                                'W(m)': 0.0,
                                'Rw(%)': 0.0,
                                'S(m)': 0.0,
                                '幾何面積(㎡)': 0.0,
                                'G(㎡)': round(_G_virtual, 2),
                                '累積S(m)': 0.0,
                                '推進側別': '🟠 孤立公設地',
                                # 🚨 Phase 9.10 法定敘明文字（精確不可變動）
                                '解法': '無法調配至其他街廓，建議領錢及申請合併分配。',
                                '迭代次數': 0,
                                '是否收斂': '—',
                                '負擔比率': round(_burden_for_orphan, 4),
                                'cut_coords': None,
                            })
                            _n_orphan_added += 1

                    # 🚨 Phase 9.9：cash compensation list 同樣寫入 g_rows
                    # （這些是因擠壓上限被退回的公設地，與 orphan 同等待遇）
                    _cash_comp = st.session_state.get('f3_cash_compensation_list', []) or []
                    _n_cash_added = 0
                    if _cash_comp:
                        for _cc in _cash_comp:
                            _a_cc = float(_cc.get('面積(㎡)', 0) or 0)
                            if _a_cc <= 0:
                                continue
                            # 同樣依 burden 算虛擬 G
                            _G_v_cc = _a_cc * (1.0 - _burden_for_orphan)
                            g_rows.append({
                                '暫編地號': _cc.get('暫編地號', ''),
                                '原地號': _cc.get('原地號', ''),
                                '所屬街廓': _cc.get('公設街廓', ''),
                                '重劃前區段': '',
                                'a 面積(㎡)': round(_a_cc, 2),
                                '街廓面積(㎡)': 0.0,
                                '正面長度(m)': 0.0, '平均深度(m)': 0.0,
                                'A 地價比': 1.0,
                                'l₂ 正面尺度': 0.0, 'l₁ 側面尺度': 0.0,
                                '街角地': '—', '第1筆街角': '—', '街角側別': '—',
                                'F(m)': 0.0, 'W(m)': 0.0, 'Rw(%)': 0.0,
                                'S(m)': 0.0, '幾何面積(㎡)': 0.0,
                                'G(㎡)': round(_G_v_cc, 2),
                                '累積S(m)': 0.0,
                                '推進側別': '💰 現金補償',
                                # 🚨 Phase 9.10 法定敘明文字（精確不可變動）
                                '解法': '無法調配至其他街廓，建議領錢及申請合併分配。',
                                '迭代次數': 0, '是否收斂': '—',
                                '負擔比率': round(_burden_for_orphan, 4),
                                'cut_coords': None,
                            })
                            _n_cash_added += 1

                    # 🆕 Phase 8 Issue 5：抵費地 < 5㎡ 碎片自動合併
                    # （含防護二：buffer(0.001).buffer(-0.001) 消縫隙）
                    try:
                        from shapely.geometry import Polygon as _SP_off_m
                        from shapely.ops import unary_union as _uu_off
                        # 依街廓分組
                        _offsets_by_blk = {}
                        for _r in g_rows:
                            if _r.get('推進側別') == '抵費地':
                                _bl = _r.get('所屬街廓', '')
                                _offsets_by_blk.setdefault(_bl, []).append(_r)
                        _merged_fragments_count = 0
                        for _bl, _offs in _offsets_by_blk.items():
                            if len(_offs) < 2:
                                continue
                            # 找最大 + 找小碎片（< 5㎡）
                            _largest = max(_offs,
                                            key=lambda r: float(r.get('幾何面積(㎡)', 0) or 0))
                            _smalls = [r for r in _offs
                                        if r is not _largest
                                        and float(r.get('幾何面積(㎡)', 0) or 0) < 5.0]
                            if not _smalls:
                                continue
                            for _s in _smalls:
                                _largest['幾何面積(㎡)'] = round(
                                    float(_largest.get('幾何面積(㎡)', 0) or 0)
                                    + float(_s.get('幾何面積(㎡)', 0) or 0), 2
                                )
                                # 合併 cut_coords（防護二：buffer 雙向消縫隙）
                                try:
                                    _l_coords = _largest.get('cut_coords') or []
                                    _s_coords = _s.get('cut_coords') or []
                                    if (len(_l_coords) >= 3 and len(_s_coords) >= 3):
                                        _poly_l = _SP_off_m(_l_coords)
                                        _poly_s = _SP_off_m(_s_coords)
                                        if not _poly_l.is_valid:
                                            _poly_l = _poly_l.buffer(0)
                                        if not _poly_s.is_valid:
                                            _poly_s = _poly_s.buffer(0)
                                        # 雙向 buffer 消除浮點縫隙
                                        _merged = _uu_off([_poly_l, _poly_s]).buffer(0.001).buffer(-0.001)
                                        if (_merged.geom_type == 'Polygon'
                                            and not _merged.is_empty):
                                            _largest['cut_coords'] = list(
                                                _merged.exterior.coords
                                            )
                                        elif _merged.geom_type == 'MultiPolygon':
                                            _largest_geom = max(
                                                _merged.geoms, key=lambda g: g.area
                                            )
                                            _largest['cut_coords'] = list(
                                                _largest_geom.exterior.coords
                                            )
                                except Exception:
                                    pass
                                # 從 g_rows 移除小碎片
                                try:
                                    g_rows.remove(_s)
                                    _merged_fragments_count += 1
                                except ValueError:
                                    pass
                        if _merged_fragments_count > 0:
                            st.session_state['f3_offset_fragments_merged'] = _merged_fragments_count
                    except Exception:
                        pass


                    st.session_state['f3_G_values'] = g_rows
                    st.session_state['f3_G_trace'] = detail_trace
                    # 🚨 Phase 9.12 Issue 4：清除 rerun flag（G 值已重新計算完成）
                    st.session_state.pop('f3_g_needs_rerun', None)
                    _n_offset = sum(1 for r in g_rows if r.get('推進側別') == '抵費地')
                    _n_orphan_in = sum(1 for r in g_rows if r.get('推進側別') == '🟠 孤立公設地')
                    # 🚨 Phase 9.10：計算現金補償筆數
                    _n_cash_in = sum(1 for r in g_rows if r.get('推進側別') == '💰 現金補償')
                    _n_alloc = len(g_rows) - _n_offset - _n_orphan_in - _n_cash_in
                    _msg = f"✅ 已完成 {_n_alloc} 筆宗地之 G 值計算（雙向夾擠）"
                    if _n_offset > 0:
                        _msg += f"，並產生 {_n_offset} 筆抵費地"
                    if _n_orphan_in > 0:
                        _msg += f"；🟠 並結算 {_n_orphan_in} 筆孤立公設地之虛擬 G 值"
                    if _n_cash_in > 0:
                        _msg += f"；💰 並標記 {_n_cash_in} 筆建議現金補償地"
                    st.success(_msg)


                    # 🚨 Phase 9.10：💰 建議領取現金補償 / 申請合併分配清單
                    _cash_list_show = st.session_state.get('f3_cash_compensation_list', []) or []
                    with st.expander(
                        "💰 建議領取現金補償 / 申請合併分配清單（未調配公設地）"
                        + (f"（{len(_cash_list_show)} 筆）" if _cash_list_show else ""),
                        expanded=bool(_cash_list_show)
                    ):
                        if _cash_list_show:
                            st.error(
                                f"⚠️ 偵測到 **{len(_cash_list_show)} 筆**公設地"
                                "無法在『不擠壓原位次地主』前提下調配（街廓上限 85%）"
                            )
                            try:
                                import pandas as _pd_cc
                                # 將「嘗試街廓清單」list 轉為字串供 dataframe 顯示
                                _cc_display = []
                                for _c in _cash_list_show:
                                    _row_d = dict(_c)
                                    _tried = _c.get('嘗試街廓清單', [])
                                    if isinstance(_tried, list):
                                        _row_d['嘗試街廓清單'] = '；'.join(str(t) for t in _tried)
                                    _cc_display.append(_row_d)
                                st.dataframe(_pd_cc.DataFrame(_cc_display),
                                              use_container_width=True, hide_index=True)
                            except Exception as _eCC:
                                st.json(_cash_list_show)
                            st.info(
                                "📜 **法定處理**：依市地重劃實施辦法 §31，"
                                "上列公設地之地主應**領取現金補償**或**申請合併分配**。\n\n"
                                "📊 g_rows 報表中此類地塊標記為「💰 現金補償」"
                                "；解法欄位記載「無法調配至其他街廓，建議領錢及申請合併分配。」"
                            )
                        else:
                            st.success("✅ 所有公設地皆成功併入鄰近建地或結算為孤立公設地，無需現金補償")

                    # 🆕 Phase 9 Task 2：抵費地比例的法定觀念說明（依重劃實務修正）
                    # ⚠️ 法規重點（依市地重劃實施辦法 §29 附件二 + 重劃實務）：
                    #
                    # B 值（一般負擔係數）公式為：
                    #   一般負擔總面積 × 重劃前平均地價 / [重劃後平均地價 × (重劃區總面積 - 抵充地)]
                    #   其中：一般負擔總面積 = 公設用地負擔總面積 - 抵充地 - 臨街地特別負擔總面積
                    #
                    # 因此 B 值 **已扣除「臨街地特別負擔」**，
                    # 性質上不等於「公共設施用地平均負擔比率」。
                    #
                    # 正確之分項：
                    #   * 公共設施用地平均負擔比率
                    #     = (共同負擔公設面積 - 政府已取得公設 - 抵充地)
                    #       / (重劃區總面積 - 政府已取得公設 - 抵充地)
                    #   * 費用負擔比率
                    #     = (工程 + 重劃 + 利息)
                    #       / [重劃後平均地價 × (重劃區總面積 - 政府已取得公設 - 抵充地)]
                    #     ※ 注意：此分母與 C 值公式分母（用「總面積 - 公設總面積」）不同
                    #
                    # 全區地主實際負擔比率 ≈ 公設用地平均負擔比率 + 費用負擔比率
                    # 抵費地（R 街廓內）大致對應「費用負擔比率」一項（實務通常 10-15%）
                    st.info(
                        "💡 **抵費地比例之法定觀念（重要）**：\n\n"
                        "**抵費地** 為 R 可建築街廓內，「抵付工程 / 重劃 / 利息費用」之保留地，"
                        "其比例**約等於「費用負擔比率」**（市地重劃實務通常 10-15%）。\n\n"
                        "* **B 值（一般負擔係數）**：依附件二公式已扣除臨街地特別負擔，"
                        "**不等於「公共設施用地平均負擔比率」**\n"
                        "* **C 值（費用負擔係數）** 之分母用「總面積 − 公設總面積」，"
                        "**亦不等於「費用負擔比率」**（後者分母為「總面積 − 政府公設 − 抵充地」）\n"
                        "* 故 **抵費地比 ≠ B + C，亦 ≠ C 值本身**；正確需另計\n\n"
                        "若實際抵費比顯著高於 10-15%（典型值），通常代表：\n"
                        "1. 重劃前地籍圖被「街廓線」切碎 → 4-Tier 公設整併失效（請檢查圖層）\n"
                        "2. 部分公設地仍為**真孤立**狀態（無同地號可合併）→ 待跨街廓調配（§31 機制）"
                    )
                    if _n_orphan_in > 0:
                        st.info(
                            "💡 **孤立公設地註**：目前抵費地面積尚包含「尚未跨街廓調配之孤立公設地保留額度」。"
                            f"目前有 **{_n_orphan_in}** 筆真孤立公設地以「虛擬 G 值」結算"
                            "（市地重劃實施辦法 §31 之跨街廓調配機制）。"
                            "實際抵費地面積將於**跨街廓指配 / 現金補償**（§31 機制）完成後縮減至最終法定值。"
                        )
                    st.rerun()

                # 🆕 W-C §0.5-B：顯示軸向診斷（含地界線 ∥ f3_cad_alloc_dir 之 90° 誤接檢查）
                _diag = st.session_state.get('f3_g_iter_diagnostics', {}) or {}
                if _diag:
                    with st.expander(f"🔬 G 迭代幾何診斷 / §0.5-B 軸向檢查（{len(_diag)} 街廓）", expanded=False):
                        st.caption(
                            "📐 **§0.5-B 真檢查**：f3_cad_alloc_dir（ALLOC=地界線/深度向）必須 ⊥ 該塊 FRONT_LINE，"
                            "夾角應 ≈90°（§0.5-E V6.dxf 真值：六塊皆 ⊥）。<80° = ALLOC 配錯/歸錯塊 🔴，"
                            "請檢查 DXF 勿續跑。"
                        )
                        for blk_lbl_d, info in sorted(_diag.items()):
                            _perp = info.get('ALLOC⊥FRONT_deg')
                            _perp_str = (f"{_perp:.1f}°" if _perp is not None else 'N/A')
                            st.markdown(
                                f"・**{blk_lbl_d}**：ALLOC⊥FRONT = {_perp_str}（應≈90）"
                                f"　→ {info.get('note', '')}"
                            )

                # 顯示結果
                if st.session_state.get('f3_G_values'):
                    st.markdown("##### 📊 G 值計算結果（幾何二分法；面積均以小數點後 2 位顯示）")
                    _df_g = _pd.DataFrame(st.session_state['f3_G_values'])
                    _fmt = {
                        'a 面積(㎡)': '{:.2f}', '街廓面積(㎡)': '{:.2f}',
                        '正面長度(m)': '{:.2f}', '平均深度(m)': '{:.2f}',
                        'A 地價比': '{:.4f}',
                        'l₂ 正面尺度': '{:.2f}', 'l₁ 側面尺度': '{:.2f}',
                        'F(m)': '{:.2f}', 'W(m)': '{:.2f}', 'Rw(%)': '{:.2f}',
                        'S(m)': '{:.2f}', 'G(㎡)': '{:.2f}',
                        '幾何面積(㎡)': '{:.2f}', '累積S(m)': '{:.2f}',
                        '負擔比率': '{:.2%}',
                    }
                    # 只套用實際存在的欄位，避免 KeyError
                    _fmt = {k: v for k, v in _fmt.items() if k in _df_g.columns}
                    # cut_coords 欄不適合在表格中顯示（list of list），先過濾掉
                    _display_cols = [c for c in _df_g.columns if c != 'cut_coords']
                    st.dataframe(
                        _df_g[_display_cols].style.format(_fmt),
                        use_container_width=True, hide_index=True,
                    )

                    # 🆕 W-D.2 §3：滑池槽診斷（k 選位／J／ΣRw 前後／守恆拆帳 ledger）
                    _wd2_diag = st.session_state.get('f3_wd2_pool_diag', {}) or {}
                    if _wd2_diag:
                        with st.expander(
                            f"🔬 W-D.2 §3 滑池槽診斷（{len(_wd2_diag)} 街廓 · "
                            "k 選位/J/ΣRw/守恆拆帳）",
                            expanded=False):
                            _rows_wd2 = []
                            for _lbl_w2 in sorted(_wd2_diag):
                                _d2 = _wd2_diag[_lbl_w2]
                                _rows_wd2.append({'街廓': _lbl_w2,
                                                  **{k2: v2 for k2, v2 in _d2.items()
                                                     if k2 != 'slot_table'}})
                            st.dataframe(_pd.DataFrame(_rows_wd2),
                                         use_container_width=True, hide_index=True)
                            _rows_jt = []
                            for _lbl_w2 in sorted(_wd2_diag):
                                for _t2 in (_wd2_diag[_lbl_w2].get('slot_table') or []):
                                    _rows_jt.append({
                                        '街廓': _lbl_w2, 'k': _t2['k'],
                                        'J': round(float(_t2['J']), 4),
                                        'dev(m)': round(float(_t2['dev']), 2),
                                        'ΣRw_L(%)': round(float(_t2['ΣRw_L']), 2),
                                        'ΣRw_R(%)': round(float(_t2['ΣRw_R']), 2),
                                    })
                            if _rows_jt:
                                st.markdown("**逐槽 J 表（各候選 k：J／dev／ΣRw 兩側）**")
                                st.dataframe(_pd.DataFrame(_rows_jt),
                                             use_container_width=True, hide_index=True)
                            st.caption(
                                "💡 **k\\***＝『兩側私有實扛側街負擔加權總和 J 最大』之池槽"
                                "（J=ΣRw·F·l₁ 兩側加總；ε 平手→最靠中央 dev 最小）；"
                                "naive＝舊 N/2 中點（僅存為 D-1 基準趟切點）。"
                                "**角落抵費地／中央池＝幾何剩餘之拆帳呈示（池重定位、非新增面積）**；"
                                "守恆 ΣG＋池＝街廓 DXF（<1㎡）。G/ΣRw 變動＝W-D.2 預期產出（D-2）；"
                                "停機僅三種：不變量破／J 下降／守恆破。\n\n"
                                "⚠️ **W1（reviewer，交 KL＋claude.ai 判）**：『中央池』欄＝池總（幾何差集）"
                                "−角落（PK min_area **規劃值**）——forced 塊若幾何剩餘＜規劃值，此欄可能"
                                "**顯負**；不影響守恆殘差（殘差不吃角落拆帳）。3.5m 之 R5左/R2左/R3右 "
                                "為觀察靶，請匯出本表交判讀。"
                            )

                    # 🆕 W-C §7：數學引擎對拍（負擔範圍 / Rw 累積 / 三量 / 深度 / 7-0a）
                    with st.expander(
                        "📐 W-C 數學引擎對拍（負擔範圍 W=18 / Rw 累積 ΣRw=100% / 三量 / 深度 / 7-0a）",
                        expanded=False):
                        _gv = st.session_state.get('f3_G_values', []) or []
                        _dep = st.session_state.get('f3_block_depth_by_label', {}) or {}
                        # ① 負擔範圍 W=18m
                        _br = st.session_state.get('f3_burden_range_18m', {}) or {}
                        if _br:
                            st.markdown("**① 負擔範圍 W=18m 面積（未扣截角，軟驗證 ≈ 18×D_avg）**")
                            _rows_br = []
                            for _lbl in sorted(_br):
                                _da = (_dep.get(_lbl, {}) or {}).get('D_avg')
                                _rows_br.append({
                                    '街廓': _lbl, '左 W18(㎡)': _br[_lbl].get('left'),
                                    '右 W18(㎡)': _br[_lbl].get('right'), 'D_avg(m)': _da,
                                    '18×D_avg(㎡)': round(18.0 * _da, 1) if _da else None})
                            st.dataframe(_pd.DataFrame(_rows_br), use_container_width=True, hide_index=True)
                        # ② ΣRw per 角側（B-2 修假警報：驗收式改 telescoping）
                        #   ΣRw_側 = R(末筆W) − R(起始W)；起始W = 該側 forced buffer 寬，無 forced 則 0。
                        #   forced 側（buffer 段不扛側街負擔）ΣRw 本就 <100%，舊式拿 100% 比會誤判。
                        st.markdown("**② 各街廓角側 Rw 累積差額 ΣRw（應 = R(末筆W) − R(起始W)；"
                                    "起始W = forced buffer 寬、無 forced 則 0）**")
                        _fo_diag2 = st.session_state.get('f3_forced_offset_diag', {}) or {}
                        _sides = {}
                        for r in _gv:
                            if float(r.get('F(m)', 0) or 0) > 0:
                                _sides.setdefault((r.get('所屬街廓', ''), r.get('推進側別', '')), []).append(r)
                        _rows_rw = []
                        for key in sorted(_sides):
                            _blk_k, _side_k = key
                            lst = _sides[key]
                            s = sum(float(x.get('Rw(%)', 0) or 0) for x in lst)
                            wlast = max(float(x.get('W(m)', 0) or 0) for x in lst)
                            # 起始W = 該側 forced buffer 寬（'left'/'right'_buffer_S）；無 forced → 0
                            _w_start = float((_fo_diag2.get(_blk_k, {}) or {}).get(
                                f'{_side_k}_buffer_S', 0.0) or 0.0)
                            _expect = rw_from_width(wlast) - rw_from_width(_w_start)
                            _rows_rw.append({
                                '街廓': _blk_k, '側': _side_k, '筆數': len(lst),
                                '起始W(m)': round(_w_start, 2),
                                'ΣRw(%)': round(s, 2), '末筆W(m)': round(wlast, 2),
                                '預期 R(末)−R(起)(%)': round(_expect, 2),
                                '判定': ('✅' if abs(s - _expect) < 1.0 else '⚠️ 檢查')})
                        if _rows_rw:
                            st.dataframe(_pd.DataFrame(_rows_rw), use_container_width=True, hide_index=True)
                        # ③ 三量對照
                        st.markdown("**③ 三量對照：W（ALLOC 法向累積）/ S（FRONT_LINE）/ 宗地寬度（單筆）**")
                        _rows_3 = []
                        for r in _gv:
                            if float(r.get('F(m)', 0) or 0) > 0:
                                _W = float(r.get('W(m)', 0) or 0); _S = float(r.get('S(m)', 0) or 0)
                                _rows_3.append({
                                    '暫編地號': r.get('暫編地號'), '街廓': r.get('所屬街廓'),
                                    'W累積(m)': _W, 'S(m)': _S, '宗地寬度(m)': r.get('宗地寬度(m)'),
                                    'Rw(%)': r.get('Rw(%)'), 'G(㎡)': r.get('G(㎡)'),
                                    '斜街廓註': 'S>W' if _S > _W + 0.3 else ''})
                        if _rows_3:
                            st.dataframe(_pd.DataFrame(_rows_3), use_container_width=True, hide_index=True)
                        # ④ 判去留旗標
                        _zl = [r for r in _gv if r.get('畸零地旗標')]
                        st.markdown(f"**④ 判去留：W<法定最小寬 → 第1調配順位旗標（實際調配 W-E）— {len(_zl)} 筆**")
                        if _zl:
                            st.dataframe(_pd.DataFrame([{
                                '暫編地號': r.get('暫編地號'), '街廓': r.get('所屬街廓'),
                                '宗地寬度(m)': r.get('宗地寬度(m)'), '旗標': r.get('畸零地旗標')}
                                for r in _zl]), use_container_width=True, hide_index=True)
                        else:
                            st.caption("（無畸零地）")
                        # ⑤ 街廓深度 + 最小分配面積
                        st.markdown("**⑤ 街廓深度（沿 ALLOC，方法A主用/B驗證）+ 最小分配面積（D_avg×min_width）**")
                        _ma = st.session_state.get('f3_min_alloc_area_by_label', {}) or {}
                        _rows_d = []
                        for _lbl in sorted(_dep):
                            _d = _dep[_lbl]
                            _rows_d.append({
                                '街廓': _lbl, 'D_avg(m)': _d.get('D_avg'), 'D_min': _d.get('D_min'),
                                'D_max': _d.get('D_max'), 'D_avg_B(驗證)': _d.get('D_avg_B'),
                                '法': _d.get('method'), '最小分配面積(㎡)': _ma.get(_lbl),
                                '註': _d.get('note')})
                        if _rows_d:
                            st.dataframe(_pd.DataFrame(_rows_d), use_container_width=True, hide_index=True)
                        # ⑥ 7-0a
                        _70 = st.session_state.get('f3_70a', {}) or {}
                        if _70:
                            _pf = '✅ PASS' if _70.get('pass') else '🔴 FAIL'
                            st.markdown(
                                f"**⑥ 7-0a 前置地基檢查：{_pf}**　region_min = "
                                f"{_70.get('region_min')}㎡（最淺乘積街廓 {_70.get('argmin_blk')}）"
                                f" vs 重劃區總面積×C% = {_70.get('total_area')}×{_70.get('C')} "
                                f"= {_70.get('pool')}㎡")
                        # ⑦ 強制街角抵費地（§0.5-F.3：該 buffer 區應 render 成橘色抵費地；白色=漏上色）
                        _fo = st.session_state.get('f3_forced_offset_diag', {}) or {}
                        _fo_on = {k: v for k, v in _fo.items()
                                  if v.get('left_forced') or v.get('right_forced')}
                        st.markdown(f"**⑦ 強制街角抵費地（forced_offset）— {len(_fo_on)} 街廓觸發**")
                        if _fo_on:
                            st.caption("該 buffer 區（街角最小面積÷D_avg）不參與分配、應顯示為橘色抵費地；"
                                       "若圖上是白色 = 漏上色（render bug）。")
                            st.dataframe(_pd.DataFrame([{
                                '街廓': k, '左forced': v.get('left_forced'),
                                '左buffer(m)': v.get('left_buffer_S'),
                                '右forced': v.get('right_forced'),
                                '右buffer(m)': v.get('right_buffer_S')} for k, v in sorted(_fo_on.items())]),
                                use_container_width=True, hide_index=True)
                        else:
                            st.caption("（無街廓觸發強制街角抵費地）")

                    # 🔬 W-C 一次性唯讀診斷（白縫量 / Tier3 對照 / 抵費地重疊或縫 / forced buffer 落點）
                    #    純讀取 f3_G_values + classified_blocks + 既有診斷，不改 g_rows、不動分配、不碰守恆。
                    with st.expander("🔬 W-C 白縫診斷（唯讀：blk−lot−抵費地 gap / Tier3 / forced buffer 落點）",
                                     expanded=False):
                        from shapely.geometry import Polygon as _SPdg
                        from shapely.ops import unary_union as _uudg
                        _gvd = st.session_state.get('f3_G_values', []) or []
                        _t3log = st.session_state.get('f3_tier3_log', []) or []
                        _crp = st.session_state.get('f3_corner_range_polys', {}) or {}
                        _fo_dg = st.session_state.get('f3_forced_offset_diag', {}) or {}
                        _OFF_TAGS = {'抵費地'}
                        _SKIP_TAGS = {'抵費地', '🟠 孤立公設地', '💰 現金補償',
                                      '🟢 Tier2 跨街廓配地'}
                        _t3_by_ri = {}
                        for _t in _t3log:
                            _t3_by_ri[_t.get('ri', '')] = (_t3_by_ri.get(_t.get('ri', ''), 0.0)
                                                           + float(_t.get('distributed_area', 0) or 0))

                        def _poly_dg(coords):
                            try:
                                if not coords or len(coords) < 3:
                                    return None
                                _p = _SPdg([(float(c[0]), float(c[1])) for c in coords])
                                if not _p.is_valid:
                                    _p = _p.buffer(0)
                                return _p if (not _p.is_empty and _p.area > 0) else None
                            except Exception:
                                return None

                        _blk_poly_dg = {}
                        for _b in classified_blocks:
                            _pp = _poly_dg(_b.get('vertices') or [])
                            if _pp is not None:
                                _blk_poly_dg[_b.get('label', '')] = _pp

                        def _lots_offs(_lbl):
                            _lots = []; _offs = []
                            for r in _gvd:
                                if r.get('所屬街廓', '') != _lbl:
                                    continue
                                _pp = _poly_dg(r.get('cut_coords'))
                                if _pp is None:
                                    continue
                                if r.get('推進側別') in _OFF_TAGS:
                                    _offs.append(_pp)
                                elif r.get('推進側別') not in _SKIP_TAGS:
                                    _lots.append(_pp)
                            return _lots, _offs

                        _rows_gap = []
                        for _lbl, _bp in sorted(_blk_poly_dg.items()):
                            _lots, _offs = _lots_offs(_lbl)
                            _ba = float(_bp.area)
                            _slot = sum(p.area for p in _lots)
                            _soff = sum(p.area for p in _offs)
                            _lot_u = _uudg(_lots) if _lots else None
                            _off_u = _uudg(_offs) if _offs else None
                            _offland_a = None; _uncov = None; _ovl = None
                            try:
                                _offland = _bp.difference(_lot_u) if _lot_u is not None else _bp
                                _offland_a = float(_offland.area)
                                if _off_u is not None:
                                    _uncov = float(_offland.difference(_off_u).area)
                                    _ovl = float(_off_u.difference(_offland).area)
                                else:
                                    _uncov = _offland_a
                            except Exception:
                                pass
                            _rows_gap.append({
                                '街廓': _lbl, 'blk_area': round(_ba, 2),
                                'Σlot(㎡)': round(_slot, 2), 'Σ抵費地(㎡)': round(_soff, 2),
                                'gap=blk−lot−抵(㎡)': round(_ba - _slot - _soff, 2),
                                'Tier3灌G(㎡)': round(_t3_by_ri.get(_lbl, 0.0), 2),
                                'offset_land重算(㎡)': (round(_offland_a, 2)
                                                       if _offland_a is not None else None),
                                '池區未成抵費地(白縫㎡)': (round(_uncov, 2)
                                                          if _uncov is not None else None),
                                '抵費地越界入lot(㎡)': (round(_ovl, 2) if _ovl is not None else None),
                            })
                        if _rows_gap:
                            st.markdown("**① 白縫量 + Tier3 對照 + 抵費地 vs offset_land（重算）**")
                            st.dataframe(_pd.DataFrame(_rows_gap),
                                         use_container_width=True, hide_index=True)
                            st.caption(
                                "gap≈Tier3灌G → 白縫源自 Tier3 把面積灌 G 沒灌幾何（A-1）；"
                                "『池區未成抵費地』>0 → offset_land 有片沒畫成抵費地（B-1 (b) 類）；"
                                "『抵費地越界入lot』>0 → 抵費地與 lot 重疊雙畫。"
                                "（offset_land 重算未套 15166 的 buffer(0.001)，故為真實幾何）")
                        else:
                            st.caption("（尚無 f3_G_values，請先跑步驟 C）")

                        _rows_fo = []
                        for _lbl, _fo in sorted(_fo_dg.items()):
                            if _lbl not in _blk_poly_dg:
                                continue
                            _lots, _offs = _lots_offs(_lbl)
                            _lot_u = _uudg(_lots) if _lots else None
                            _off_u = _uudg(_offs) if _offs else None
                            for _side in ('left', 'right'):
                                if not _fo.get(f'{_side}_forced'):
                                    continue
                                _crpoly = _poly_dg((_crp.get(_lbl) or {}).get(_side))
                                _pt = None
                                if _crpoly is not None:
                                    try:
                                        _pt = _crpoly.representative_point()
                                    except Exception:
                                        _pt = None
                                _in_lot = (bool(_lot_u.contains(_pt))
                                           if (_pt is not None and _lot_u is not None) else None)
                                _in_off = (bool(_off_u.contains(_pt))
                                           if (_pt is not None and _off_u is not None) else None)
                                if _pt is None:
                                    _verd = '代表點無（缺 corner_range_polys，需重跑步驟C）'
                                elif _in_lot:
                                    _verd = '在 lot 內 → 真因 (a)：buffer 被首筆 lot 吃'
                                elif _in_off:
                                    _verd = '在抵費地內 → 已正確上色（白色另有他因）'
                                else:
                                    _verd = '皆不在 → 真因 (b)：白洞（difference/<1㎡ 過濾丟）'
                                _rows_fo.append({
                                    '街廓': _lbl, '側': _side,
                                    'buffer_S(m)': _fo.get(f'{_side}_buffer_S'),
                                    'corner_range面積(㎡)': (round(_crpoly.area, 2)
                                                            if _crpoly is not None else None),
                                    '代表點在lot': _in_lot, '代表點在抵費地': _in_off,
                                    '判定': _verd})
                        if _rows_fo:
                            st.markdown("**② forced 側 corner buffer 代表點落點（判 B-1 真因 (a)/(b)）**")
                            st.dataframe(_pd.DataFrame(_rows_fo),
                                         use_container_width=True, hide_index=True)

                    # 迭代軌跡檢視
                    with st.expander("🔎 檢視單筆迭代軌跡"):
                        _keys = list(st.session_state.get('f3_G_trace', {}).keys())
                        if _keys:
                            _sel = st.selectbox("選擇暫編地號", _keys, key='f3_g_trace_sel')
                            _trace = st.session_state['f3_G_trace'].get(_sel, [])
                            if _trace:
                                _df_tr = _pd.DataFrame(_trace)
                                st.dataframe(_df_tr, use_container_width=True, hide_index=True)

                    # ─── 任務七 B + Task D-4：重劃後試分配地籍圖（雙向夾擠 + 抵費地）───
                    st.markdown("##### 🗺️ 重劃後試分配地籍圖")
                    st.caption(
                        "🟢 **綠色** = 一般宗地分配後幾何　|　"
                        "🟠 **橘色** = 街角地（左右錨定端點）　|　"
                        "🟡 **黃色** = 抵費地（雙向夾擠後剩餘空地，自動標註面積）　|　"
                        "灰色 = 街廓底圖。Hover 可查看詳細數據；雙向推進方向見表格『推進側別』欄。"
                    )
                    import plotly.graph_objects as _go_post
                    _fig_post = _go_post.Figure()

                    # 1) 街廓灰底
                    for _b in classified_blocks:
                        _verts = _b.get('vertices') or []
                        if not _verts:
                            continue
                        _xs = [v[0] for v in _verts] + [_verts[0][0]]
                        _ys = [v[1] for v in _verts] + [_verts[0][1]]
                        _fig_post.add_trace(_go_post.Scatter(
                            x=_xs, y=_ys, mode='lines',
                            fill='toself', fillcolor='rgba(200,200,200,0.18)',
                            line=dict(color='#777', width=0.6),
                            hoverinfo='text',
                            text=f"{_b.get('label','')} ({_b.get('category','')})",
                            showlegend=False,
                        ))

                    # 2) 各暫編地號之 cut polygon — Task D-4：三色（綠/橘/黃）
                    _palette = ['#2E8B33', '#3CB371', '#1F77B4', '#9467BD',
                                '#8C564B', '#E377C2', '#17BECF', '#BCBD22']
                    _drawn = 0; _missed = 0; _offset_drawn = 0
                    for _i, _row in enumerate(st.session_state['f3_G_values']):
                        _coords = _row.get('cut_coords') or []
                        if len(_coords) < 3:
                            _missed += 1
                            continue
                        _xs = [c[0] for c in _coords] + [_coords[0][0]]
                        _ys = [c[1] for c in _coords] + [_coords[0][1]]
                        # 三色判斷：抵費地（黃） > 街角地（橘） > 一般（綠）
                        _is_offset = (_row.get('推進側別') == '抵費地')
                        _is_corner = (_row.get('街角地') == '是')
                        if _is_offset:
                            _fill_rgba = 'rgba(255, 200, 0, 0.55)'   # 黃色
                            _line_color = '#E8A317'
                            _line_width = 2.0
                            _hover = (
                                f"<b>抵費地</b><br>"
                                f"編號 {_row.get('暫編地號','')}<br>"
                                f"街廓 {_row.get('所屬街廓','')}<br>"
                                f"面積 {_row.get('幾何面積(㎡)',0):.2f} ㎡<br>"
                                f"狀態：{_row.get('是否收斂','—')}"
                            )
                            _trace_name = f"抵費地 {_row.get('暫編地號','')}"
                        elif _is_corner:
                            _fill_rgba = 'rgba(255,165,0,0.55)'      # 橘色
                            _line_color = '#FF8C00'
                            _line_width = 1.5
                            _hover = (
                                f"<b>街角地</b> {_row.get('暫編地號','')}<br>"
                                f"原地號 {_row.get('原地號','')}<br>"
                                f"街廓 {_row.get('所屬街廓','')}<br>"
                                f"a={_row.get('a 面積(㎡)',0):.2f} ㎡<br>"
                                f"S={_row.get('S(m)',0):.2f} m<br>"
                                f"G={_row.get('G(㎡)',0):.2f} ㎡<br>"
                                f"推進側別：{_row.get('推進側別','—')}"
                            )
                            _trace_name = f"{_row.get('暫編地號','')} ({_row.get('G(㎡)',0):.0f}㎡)"
                        else:
                            _fill_rgba = 'rgba(50,180,80,0.50)'       # 綠色
                            _line_color = _palette[_i % len(_palette)]
                            _line_width = 1.5
                            _hover = (
                                f"暫編 {_row.get('暫編地號','')}<br>"
                                f"原地號 {_row.get('原地號','')}<br>"
                                f"街廓 {_row.get('所屬街廓','')}<br>"
                                f"a={_row.get('a 面積(㎡)',0):.2f} ㎡<br>"
                                f"S={_row.get('S(m)',0):.2f} m<br>"
                                f"G={_row.get('G(㎡)',0):.2f} ㎡<br>"
                                f"推進側別：{_row.get('推進側別','—')}"
                            )
                            _trace_name = f"{_row.get('暫編地號','')} ({_row.get('G(㎡)',0):.0f}㎡)"
                        _fig_post.add_trace(_go_post.Scatter(
                            x=_xs, y=_ys, mode='lines',
                            fill='toself', fillcolor=_fill_rgba,
                            line=dict(color=_line_color, width=_line_width),
                            name=_trace_name,
                            hoverinfo='text', text=_hover,
                            showlegend=False,
                        ))
                        _drawn += 1
                        if _is_offset:
                            _offset_drawn += 1

                    # 抵費地中心文字標註 — Task D-4
                    for _row in st.session_state['f3_G_values']:
                        if _row.get('推進側別') != '抵費地':
                            continue
                        _coords = _row.get('cut_coords') or []
                        if len(_coords) < 3:
                            continue
                        try:
                            from shapely.geometry import Polygon as _SP_ann
                            _cen = _SP_ann(_coords).centroid
                            _fig_post.add_annotation(
                                x=_cen.x, y=_cen.y,
                                text=f"<b>抵費地</b><br>{_row.get('幾何面積(㎡)',0):.1f} ㎡",
                                showarrow=False,
                                font=dict(size=11, color='#5D4037'),
                                bgcolor='rgba(255,255,255,0.75)',
                                bordercolor='#E8A317', borderwidth=1, borderpad=3,
                            )
                        except Exception:
                            pass

                    if _drawn == 0:
                        st.info(
                            "ℹ️ 尚無可繪製之 cut polygon。可能原因：(1) 尚未執行 G 值迭代計算 "
                            "(2) 該街廓無正面道路角點/方向資訊 → 退回代數迭代解法（無幾何 cut）"
                        )
                    else:
                        if _missed > 0:
                            st.caption(
                                f"💡 已繪製 {_drawn} 筆"
                                + (f"（含 {_offset_drawn} 筆抵費地）" if _offset_drawn > 0 else "")
                                + f"；另有 {_missed} 筆無 cut 幾何（多為代數迭代解法）"
                            )
                        _fig_post.update_layout(
                            height=560, paper_bgcolor='white', plot_bgcolor='white',
                            xaxis=dict(title="TWD97 X", gridcolor='#E5E5E5'),
                            yaxis=dict(title="TWD97 Y", gridcolor='#E5E5E5',
                                       scaleanchor="x", scaleratio=1),
                            margin=dict(l=40, r=40, t=10, b=40), showlegend=False,
                        )
                        st.plotly_chart(_fig_post, use_container_width=True)
                        st.caption(
                            "🟢 綠色 = 一般宗地　🟠 橘色 = 街角地（第 1 筆走完整側街運算）　"
                            "🩶 灰色 = 街廓底圖"
                        )

                    # ---------- 🆕 Phase 7 Module 4：法定成果匯出 ----------
                    st.markdown("---")
                    st.markdown("##### 📥 法定成果匯出（Phase 7）")
                    st.caption(
                        "完成 G 值計算後，可匯出符合 AutoCAD 與印表機列印之法定文件。"
                    )
                    _ec1, _ec2 = st.columns(2)
                    # DXF 匯出
                    try:
                        _dxf_bytes_p7 = export_allocated_dxf(
                            g_rows=st.session_state.get('f3_G_values', []),
                            classified_blocks=classified_blocks,
                            manual_road_centerlines=st.session_state.get(
                                'f3_manual_road_centerlines', {}
                            ),
                        )
                        _ec1.download_button(
                            label="📥 下載重劃後土地分配 DXF",
                            data=_dxf_bytes_p7,
                            file_name="重劃後土地分配圖.dxf",
                            mime="application/octet-stream",
                            key='dl_p7_dxf',
                            use_container_width=True,
                            type='primary',
                            help="包含 NEW_PARCEL（綠）+ OFFSET_LAND（黃）+ NEW_CENTERLINE（紅）+ PARCEL_TEXT 圖層",
                        )
                    except Exception as _eDXF:
                        _ec1.error(f"DXF 匯出失敗：{_eDXF}")
                    # Excel 匯出
                    try:
                        _xls_bytes_p7 = export_legal_excel(
                            g_rows=st.session_state.get('f3_G_values', []),
                            ownership_map=st.session_state.get('t8_ownership_map', {}),
                            ownership_groups=st.session_state.get('t8_ownership_groups', {}),
                            B_value=float(B_value or 0),
                            C_value=float(C_for_calc or 0),
                            total_area=float(sum(
                                b.get('area_m2', 0) or 0 for b in classified_blocks
                            )),
                            pre_avg_price=float(st.session_state.get('pre_land_price_sqm', 0) or 0),
                            post_avg_price=float(st.session_state.get('weighted_price_sqm', 0) or 0),
                            engineering_cost=float(st.session_state.get(
                                'f3_engineering_cost_from_finance', 0) or 0),
                            redev_cost=float(st.session_state.get(
                                'f3_redev_cost_from_finance', 0) or 0),
                            loan_interest=float(st.session_state.get(
                                'f3_loan_interest_from_finance', 0) or 0),
                        )
                        _ec2.download_button(
                            label="📥 下載法定報表 Excel（3 表）",
                            data=_xls_bytes_p7,
                            file_name="重劃法定報表.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key='dl_p7_xls',
                            use_container_width=True,
                            type='primary',
                            help="Sheet 1: 對照清冊 + Sheet 2: 歸戶負擔 + Sheet 3: 公設統計（橫向 A4）",
                        )
                    except Exception as _eXLS:
                        _ec2.error(f"Excel 匯出失敗：{_eXLS}")

            # ---------- 步驟 J：最小分配面積判斷 + 同原地號合併 ----------
            if st.session_state.get('f3_G_values'):
                st.markdown("---")
                st.markdown("#### 🧮 步驟 J：最小分配面積判斷 + 同原地號合併")
                st.markdown("""
依《花蓮縣畸零地使用規則》§3 第 1 款（一般建築基地）附表，以**該暫編地號所屬街廓的分類 + 正面路寬**
查得最小分配寬度 × 深度 → 最小分配面積。

**判斷規則（市地重劃實施辦法 §30 + 平均地權條例 §60-1）：**
1. 暫編地號 G 值 **< 最小分配面積 1/2** → 建議**現金補償**
2. 暫編地號 G 值 **≥ 1/2 但 < 最小分配面積** → 合併至**同原地號面積最大之暫編地號**
3. 若同原地號所有暫編地號 G 值**加總後仍未達最小分配面積** → 需**跨街廓調配**（§31 機制）
""")

                # 建立每街廓之最小分配面積
                _min_area_by_block = {}
                _min_info_rows = []
                for b in classified_blocks:
                    _lbl = b['label']
                    _cat = b.get('category', '')
                    _sb_row = sb_rows_by_label.get(_lbl, {})
                    _fw = float(_sb_row.get('正面路寬(m)', 0.0) or 0.0)
                    _info = get_min_lot_size(_cat, _fw)
                    # 🚨 Patch E-1 §6：min_area 改用「Ri 實際分配深度」（非法定固定 14m）
                    _alloc_depth_w1 = float(_sb_row.get('街廓分配深度(m)', 0.0) or 0.0)
                    if _alloc_depth_w1 > 0 and _info['min_width'] > 0:
                        _cutoff_w1 = (_safe_num_e1(_sb_row.get('【左】截角(㎡)'))
                                      + _safe_num_e1(_sb_row.get('【右】截角(㎡)')))
                        _min_area_by_block[_lbl] = round(
                            max(_info['min_width'] * _alloc_depth_w1 - _cutoff_w1, 0.0), 2)
                    else:
                        _min_area_by_block[_lbl] = _info['min_area']  # fallback 法定值
                    if _info['min_area'] > 0:
                        _min_info_rows.append({
                            '街廓': _lbl, '分類': _cat, '正面路寬(m)': round(_fw, 2),
                            '查表分類': _info['table_key'],
                            '最小寬度(m)': _info['min_width'],
                            '最小深度(m)': _info['min_depth'],
                            '最小分配面積(㎡)': _info['min_area'],
                        })

                if _min_info_rows:
                    with st.expander("📑 各街廓最小分配面積對照（依畸零地附表）", expanded=False):
                        st.dataframe(_pd.DataFrame(_min_info_rows),
                                     use_container_width=True, hide_index=True)

                # 🚨 Patch B-2：寬度驗證（並列於既有面積驗證）
                # 從各 g_row 之 cut_coords 反推 strip width，若 < 法定最小寬度（住宅 3.5m）→
                # 強制標記 _below_min_width=True；後續 merge_subparcels_by_parent 會以
                # 「擴充合併條件」處理（面積 < min OR 寬度 < min 任一即觸發合併）
                _min_width_by_block = {}
                for _b_w in classified_blocks:
                    _lbl_w = _b_w['label']
                    _cat_w = _b_w.get('category', '')
                    _sb_w = sb_rows_by_label.get(_lbl_w, {})
                    _fw_w = float(_sb_w.get('正面路寬(m)', 0.0) or 0.0)
                    _mw_info = get_min_lot_size(_cat_w, _fw_w)
                    _min_width_by_block[_lbl_w] = float(_mw_info.get('min_width', 0.0) or 0.0)

                _width_violation_count = 0
                for _r_jw in st.session_state.get('f3_G_values', []) or []:
                    if _r_jw.get('推進側別') in ('抵費地', '🟠 孤立公設地', '💰 現金補償'):
                        continue
                    # 🚨 Patch E-1.6：寬度判定優先用 G 迭代的 S(m) 推進值
                    _s_from_iter = _r_jw.get('S(m)')
                    if _s_from_iter is not None:
                        try:
                            _w_jw = float(_s_from_iter)
                        except Exception:
                            _w_jw = 0.0
                    else:
                        _cut_jw = _r_jw.get('cut_coords') or []
                        if len(_cut_jw) < 3:
                            continue
                        _w_jw = _compute_strip_width(_cut_jw, d_hat=None)
                    _r_jw['實際寬度(m)'] = round(_w_jw, 2)
                    _legal_w = float(_min_width_by_block.get(
                        _r_jw.get('所屬街廓', ''), 0.0) or 0.0)
                    if _legal_w > 0 and _w_jw < _legal_w:
                        _r_jw['_below_min_width'] = True
                        _r_jw['_width_violation_note'] = (
                            f"寬度 {_w_jw:.2f}m < 法定最小 {_legal_w:.2f}m"
                        )
                        # 將寬度未達者之 G 強制壓為 < min_area，觸發合併條件
                        # （具體做法：將 G 設為 1，使其落入 < min_area 區段；
                        #  保留原 G 至 _G_before_width_violation 供回溯）
                        _orig_G = float(_r_jw.get('G(㎡)', 0.0) or 0.0)
                        _min_a_jw = float(_min_area_by_block.get(
                            _r_jw.get('所屬街廓', ''), 0.0) or 0.0)
                        if _orig_G >= _min_a_jw and _min_a_jw > 0:
                            _r_jw['_G_before_width_violation'] = _orig_G
                            _r_jw['G(㎡)'] = round(min(_orig_G, _min_a_jw - 0.01), 2)
                        _width_violation_count += 1
                if _width_violation_count > 0:
                    st.warning(
                        f"📏 **Patch B-2 寬度驗證**：共偵測到 **{_width_violation_count} 筆**"
                        f"暫編地號之實際寬度 < 法定最小寬度 → 已標記為「待合併」"
                        f"（觸發同原地號合併或跨街廓調配）"
                    )

                # 執行合併
                _merge_res = merge_subparcels_by_parent(
                    st.session_state['f3_G_values'], _min_area_by_block)

                # 顯示合併結果
                st.markdown("##### 🔀 合併結果（同原地號之暫編地號）")
                _df_merge = _pd.DataFrame(_merge_res['merged_rows'])
                if not _df_merge.empty:
                    _show_cols = [c for c in [
                        '暫編地號', '原地號', '所屬街廓', 'a 面積(㎡)',
                        'G(㎡)', '合併後G(㎡)', '最小分配面積(㎡)', '是否達最小',
                        '合併來源', '合併至',
                    ] if c in _df_merge.columns]
                    _fmt = {c: '{:.2f}' for c in ('a 面積(㎡)', 'G(㎡)', '合併後G(㎡)',
                                                   '最小分配面積(㎡)') if c in _df_merge.columns}
                    st.dataframe(_df_merge[_show_cols].style.format(_fmt),
                                 use_container_width=True, hide_index=True)

                # 現金補償建議
                if _merge_res['cash_compensation']:
                    st.warning(f"⚠️ 共 {len(_merge_res['cash_compensation'])} 筆暫編地號未達最小分配面積 1/2，建議現金補償：")
                    st.dataframe(_pd.DataFrame(_merge_res['cash_compensation']),
                                 use_container_width=True, hide_index=True)

                # 儲存結果
                st.session_state['f3_J_merged'] = _merge_res['merged_rows']
                st.session_state['f3_J_cash_compensation'] = _merge_res['cash_compensation']
                st.session_state['f3_J_cross_block_needed'] = _merge_res['cross_block_needed']
                st.session_state['f3_min_area_by_block'] = _min_area_by_block

                # 提示跨街廓調配
                if _merge_res['cross_block_needed']:
                    st.info(f"ℹ️ 共 {len(_merge_res['cross_block_needed'])} 個原地號於所屬街廓合併後仍未達最小分配面積 → 需進行**跨街廓調配**（§31 機制）")


            # ---------- 步驟 L 已前移至 Step G 內部（Phase A 重構） ----------
            # 原 Step L 區塊已移除；新版位於 Step G 之 _prev_params 之後（PK 前置）
            # 此處保留結構性空行，後接 Step M

            # ---------- 步驟 M：公共設施用地分配（任務四） ----------
            st.markdown("---")
            st.markdown("#### 🌳 步驟 M：公共設施用地分配（與 土地歸戶 連動）")
            st.markdown("""
依《平均地權條例》§60 + 《市地重劃實施辦法》§21~28：
- **道路（共同負擔）**：以道路中心線為界，將道路面積分向兩側合併分配
- **廣場 / 鄰里公園 / 零售市場（共同負擔）**：合併於相鄰住宅區
- **非共同負擔（機關、社會住宅）**：公地優先指配 → 不足時按私有歸戶面積比例發還（不受最小分配限制）
""")
            with st.expander("📋 自動計算公設分配（依目前 土地歸戶 + 街廓互動分析 街廓分類）",
                              expanded=False):
                _own_map = st.session_state.get('t8_ownership_map', {}) or {}
                _own_groups = st.session_state.get('t8_ownership_groups', {}) or {}
                if not _own_map:
                    st.warning("⚠️ 尚未偵測到 土地歸戶 資料，請先至 土地歸戶 匯入歸戶 Excel。")
                else:
                    # 把街廓依用途分群
                    _road_blocks = [b for b in classified_blocks
                                    if (b.get('category', '') in ('道路', '溝渠'))]
                    _plaza_blocks = [b for b in classified_blocks
                                     if (b.get('category', '') in ('廣場', '鄰里公園',
                                                                    '兒童遊樂場', '零售市場'))]
                    _non_common_blocks = [b for b in classified_blocks
                                          if (b.get('category', '') in ('機關用地', '社會住宅',
                                                                         '其他非共同負擔之公共設施用地'))]

                    # ─── M-1 道路（任務六：幾何驅動切割） ───
                    if _road_blocks:
                        st.markdown("##### 🛣️ M-1 道路用地（共同負擔，以道路中心線幾何切割分配）")
                        from shapely.geometry import Polygon as _SPolyM
                        _road_alloc_summary = []
                        # 預先把暫編地號依「街廓」分組，方便依左右側街廓清單聚合持分
                        _parcel_by_block = {}
                        for tp in (temp_parcels or []):
                            _bk = tp.get('所屬街廓', '')
                            _parcel_by_block.setdefault(_bk, []).append(tp)
                        # 取得非道路街廓清單（道路兩側可能的街廓）
                        _non_road_blocks = [b for b in classified_blocks
                                            if (b.get('category', '') not in ('道路', '溝渠'))]

                        for rb in _road_blocks:
                            # 1) 重建道路 polygon
                            try:
                                _verts = rb.get('vertices') or []
                                _road_poly = _SPolyM(_verts) if len(_verts) >= 3 else None
                                if _road_poly is not None and not _road_poly.is_valid:
                                    _road_poly = _road_poly.buffer(0)
                            except Exception:
                                _road_poly = None
                            # 2) 取得中心線
                            #    🆕 Phase 6 Task N：優先使用使用者「精準平移生成」之手動中心線
                            #    🆕 Task B：否則 fallback 至自動演算法（含相鄰街廓邊界延伸）
                            _cls = []
                            _r_lbl = rb.get('label', '')
                            _manual_mc_dict = (st.session_state.get(
                                'f3_manual_road_centerlines', {}) or {})
                            if _r_lbl in _manual_mc_dict and len(_manual_mc_dict[_r_lbl]) >= 2:
                                # 使用手動精準中心線
                                from shapely.geometry import LineString as _SL_mc
                                try:
                                    _pts_mc = _manual_mc_dict[_r_lbl]
                                    _cls = [_SL_mc([(p[0], p[1]) for p in _pts_mc])]
                                except Exception:
                                    _cls = []
                            if not _cls and _road_poly is not None:
                                # fallback：自動演算法
                                try:
                                    _adj_blocks = []
                                    for _b in _non_road_blocks:
                                        _bv = _b.get('vertices') or []
                                        if len(_bv) < 3:
                                            continue
                                        try:
                                            _bp = _SPolyM(_bv)
                                            if not _bp.is_valid:
                                                _bp = _bp.buffer(0)
                                            _adj_blocks.append({
                                                'label': _b.get('label', ''),
                                                'shapely': _bp,
                                            })
                                        except Exception:
                                            continue
                                    _cls = _extract_road_centerlines(
                                        _road_poly, adjacent_blocks=_adj_blocks
                                    )
                                except Exception:
                                    _cls = []
                            # 3) 把非道路街廓依中心線分為左右
                            _side_split = (classify_blocks_by_centerline(_non_road_blocks, _cls)
                                           if _cls else {'left': [], 'right': [], 'on_line': []})
                            _left_blk_ids = set(_side_split.get('left', []))
                            _right_blk_ids = set(_side_split.get('right', []))
                            # 把街廓 id 對到 label（或反之）以匹配 _parcel_by_block 的 key
                            _id_to_label = {b.get('id', b.get('label', '')): b.get('label', '')
                                            for b in _non_road_blocks}
                            _left_blk_labels = {_id_to_label.get(i, '') for i in _left_blk_ids}
                            _right_blk_labels = {_id_to_label.get(i, '') for i in _right_blk_ids}
                            # 4) 為每個歸戶聚合左右側真實持分面積
                            _holdings = {}
                            for gid in _own_groups.keys():
                                _l_a = 0.0; _r_a = 0.0
                                for tp in (temp_parcels or []):
                                    _own = _own_map.get(tp.get('原地號', ''), '')
                                    if _own != gid:
                                        continue
                                    _bk = tp.get('所屬街廓', '')
                                    if _bk in _left_blk_labels:
                                        _l_a += float(tp.get('幾何面積_m2', tp.get('面積_m2', 0.0)) or 0.0)
                                    elif _bk in _right_blk_labels:
                                        _r_a += float(tp.get('幾何面積_m2', tp.get('面積_m2', 0.0)) or 0.0)
                                _holdings[gid] = {'left': _l_a, 'right': _r_a}
                            # 5) 呼叫幾何驅動分配
                            _r = allocate_road_to_adjacent_owners(
                                road_area_m2=float(rb.get('area_m2', 0.0)),
                                ownership_residential_holdings=_holdings,
                                min_area_m2=0.0,
                                road_poly=_road_poly,
                                centerlines=_cls,
                            )
                            _road_alloc_summary.append({
                                '道路街廓': rb.get('label', ''),
                                '面積(㎡)': round(float(rb.get('area_m2', 0.0)), 2),
                                '中心線分支數': len(_cls),
                                '左側街廓數': len(_left_blk_ids),
                                '右側街廓數': len(_right_blk_ids),
                                '左側面積(㎡)': _r.get('left_area', 0.0),
                                '右側面積(㎡)': _r.get('right_area', 0.0),
                                '切割方式': _r.get('method', ''),
                                '受配歸戶數': len(_r['allocations']),
                                '需調配歸戶數': len(_r['cross_block_needed']),
                            })
                        st.dataframe(_pd.DataFrame(_road_alloc_summary),
                                     use_container_width=True, hide_index=True)

                    # ─── M-2 廣場/公園/市場 ───
                    if _plaza_blocks:
                        st.markdown("##### 🌲 M-2 廣場/鄰里公園/零售市場（合併於相鄰住宅區）")
                        _plaza_alloc_summary = []
                        # 簡化：「相鄰住宅區面積」以該歸戶於可建築街廓之 G 值總和代理
                        _g_rows = st.session_state.get('f3_G_values', []) or []
                        _resi_holdings = {}
                        for r in _g_rows:
                            _gid = _own_map.get(r.get('原地號', ''), '')
                            if not _gid:
                                continue
                            _resi_holdings[_gid] = _resi_holdings.get(_gid, 0.0) \
                                + float(r.get('G(㎡)', 0.0) or 0.0)
                        for pb in _plaza_blocks:
                            _r = allocate_plaza_to_adjacent_residential(
                                plaza_area_m2=float(pb.get('area_m2', 0.0)),
                                ownership_residential_holdings=_resi_holdings,
                                min_area_m2=0.0,
                            )
                            _plaza_alloc_summary.append({
                                '公設街廓': pb.get('label', ''),
                                '分類': pb.get('category', ''),
                                '面積(㎡)': round(float(pb.get('area_m2', 0.0)), 2),
                                '受配歸戶數': len(_r['allocations']),
                                '需調配歸戶數': len(_r['cross_block_needed']),
                            })
                        st.dataframe(_pd.DataFrame(_plaza_alloc_summary),
                                     use_container_width=True, hide_index=True)

                    # ─── M-3 非共同負擔 ───
                    if _non_common_blocks:
                        st.markdown("##### 🏛️ M-3 非共同負擔公設（公地優先 + 按比例發還）")
                        _public_supply_input = st.number_input(
                            "公有地可供量(㎡)（縣市 / 鄉鎮 / 國有 合計）",
                            min_value=0.0, value=0.0, step=100.0,
                            key='f3M_public_supply')
                        _offset_supply_input = st.number_input(
                            "抵費地可供量(㎡)",
                            min_value=0.0, value=0.0, step=100.0,
                            key='f3M_offset_supply')
                        _g_rows = st.session_state.get('f3_G_values', []) or []
                        _private_holdings = {}
                        for r in _g_rows:
                            _gid = _own_map.get(r.get('原地號', ''), '')
                            if not _gid:
                                continue
                            _private_holdings[_gid] = _private_holdings.get(_gid, 0.0) \
                                + float(r.get('a 面積(㎡)', 0.0) or 0.0)
                        _nc_alloc_summary = []
                        for nb in _non_common_blocks:
                            _r = allocate_non_common_public_land(
                                public_land_area_m2=float(nb.get('area_m2', 0.0)),
                                public_owned_supply_m2=_public_supply_input,
                                offset_land_supply_m2=_offset_supply_input,
                                private_holdings=_private_holdings,
                                min_area_m2=0.0,
                            )
                            _nc_alloc_summary.append({
                                '公設街廓': nb.get('label', ''),
                                '分類': nb.get('category', ''),
                                '需指配(㎡)': round(float(nb.get('area_m2', 0.0)), 2),
                                '公有地指配(㎡)': _r['public_used'],
                                '抵費地指配(㎡)': _r['offset_used'],
                                '按比例發還合計(㎡)': round(sum(p['按比例發還']
                                                              for p in _r['pro_rata_returns']), 2),
                                '受配歸戶數': len(_r['pro_rata_returns']),
                            })
                        st.dataframe(_pd.DataFrame(_nc_alloc_summary),
                                     use_container_width=True, hide_index=True)

                    if not _road_blocks and not _plaza_blocks and not _non_common_blocks:
                        st.info("無公共設施用地街廓需分配（請至步驟 A 確認街廓分類）")

            # ─── M-4 圖形預覽（任務三）：直觀顯示公設與相鄰住宅區之相對位置 ───
            _all_public_blocks = [b for b in classified_blocks
                                  if b.get('category', '') in (
                                      '道路', '溝渠', '廣場', '鄰里公園',
                                      '兒童遊樂場', '零售市場', '機關用地',
                                      '停車場', '社會住宅',
                                      '其他非共同負擔之公共設施用地')]
            if _all_public_blocks:
                with st.expander("🗺️ M-4 圖形預覽：公設用地 ↔ 相鄰歸戶土地（任務三）",
                                  expanded=False):
                    st.caption(
                        "選擇要預覽之公設用地與歸戶 → 同時畫出該公設街廓（藍色半透明）+ "
                        "該歸戶於相鄰可建築街廓內的暫編地號（綠色填充）"
                        "+（道路類）輔助道路中心線。"
                    )

                    _own_groups = st.session_state.get('t8_ownership_groups', {}) or {}
                    _own_map = st.session_state.get('t8_ownership_map', {}) or {}

                    _pcol1, _pcol2 = st.columns(2)
                    _public_lbls = [
                        f"{b.get('label','')}（{b.get('category','')}, "
                        f"{float(b.get('area_m2',0)):,.1f} ㎡）"
                        for b in _all_public_blocks
                    ]
                    _public_idx = _pcol1.selectbox(
                        "選擇公設街廓", options=list(range(len(_public_lbls))),
                        format_func=lambda i: _public_lbls[i],
                        key='f3M_preview_public_idx',
                    )
                    _gid_options = ['（不指定，顯示全部歸戶）']
                    _gid_options.extend(sorted(_own_groups.keys()))
                    _gid_pick = _pcol2.selectbox(
                        "選擇歸戶（綠色顯示其土地）", options=_gid_options,
                        key='f3M_preview_gid',
                    )

                    _selected_block = _all_public_blocks[_public_idx]
                    _is_road_like = _selected_block.get('category', '') in ('道路', '溝渠')

                    # 找該歸戶所有暫編地號（只取可建築土地者）
                    _gid_parcel_set = set()
                    if _gid_pick != '（不指定，顯示全部歸戶）':
                        _gid_parcel_set = {
                            ono for ono, gid in _own_map.items()
                            if gid == _gid_pick
                        }

                    # 建立 Plotly 圖
                    import plotly.graph_objects as go
                    _fig_m = go.Figure()

                    # 1) 全部街廓：淡灰底
                    for b in classified_blocks:
                        verts = b.get('vertices', []) or []
                        if not verts:
                            continue
                        xs = [v[0] for v in verts] + [verts[0][0]]
                        ys = [v[1] for v in verts] + [verts[0][1]]
                        _fig_m.add_trace(go.Scatter(
                            x=xs, y=ys, mode='lines',
                            fill='toself', fillcolor='rgba(200,200,200,0.15)',
                            line=dict(color='#888888', width=0.6),
                            hoverinfo='text',
                            text=f"{b.get('label','')} - {b.get('category','')}",
                            showlegend=False,
                        ))

                    # 2) 目標公設用地：藍色半透明粗邊
                    _sel_verts = _selected_block.get('vertices', []) or []
                    if _sel_verts:
                        _xs = [v[0] for v in _sel_verts] + [_sel_verts[0][0]]
                        _ys = [v[1] for v in _sel_verts] + [_sel_verts[0][1]]
                        _fig_m.add_trace(go.Scatter(
                            x=_xs, y=_ys, mode='lines',
                            fill='toself', fillcolor='rgba(30, 100, 255, 0.35)',
                            line=dict(color='#1F4FE0', width=3.5),
                            name=f"🟦 公設：{_selected_block.get('label','')}"
                                 f"（{_selected_block.get('category','')}）",
                            hoverinfo='text',
                            text=f"{_selected_block.get('label','')} "
                                 f"{_selected_block.get('category','')} "
                                 f"{float(_selected_block.get('area_m2',0)):,.1f} ㎡",
                        ))

                    # 3) 🆕 Phase 11 重寫：使用 CAD CENTERLINE 圖層之中心線（不再自動偵測）
                    #    讀取 f3_manual_road_centerlines 中對應該道路街廓之中心線
                    _cad_centerlines = st.session_state.get('f3_manual_road_centerlines', {}) or {}
                    _ln_pts_for_split = None    # 供下游切割使用之 LineString
                    if _is_road_like and _sel_verts:
                        _cl_pts = _cad_centerlines.get(_selected_block.get('label', ''))
                        if _cl_pts and len(_cl_pts) >= 2:
                            _cl_xs = [p[0] for p in _cl_pts]
                            _cl_ys = [p[1] for p in _cl_pts]
                            _fig_m.add_trace(go.Scatter(
                                x=_cl_xs, y=_cl_ys, mode='lines',
                                line=dict(color='#FFA000', width=3, dash='dash'),
                                name='— · — 道路中心線（CAD CENTERLINE 圖層）',
                                hoverinfo='skip',
                            ))
                            try:
                                from shapely.geometry import LineString as _LS_split
                                _ln_pts_for_split = _LS_split(_cl_pts)
                            except Exception:
                                _ln_pts_for_split = None

                    # 4) 🆕 公設用地內之暫編地號 + 配合中心線切割 + 顯示分配示意
                    #    步驟：
                    #    a. 找出位於選定公設街廓內之暫編地號（temp_parcels.所屬街廓 == selected）
                    #    b. 對每筆暫編地號，依其原地號之歸戶 → 找該歸戶於相鄰可建築街廓之土地
                    #    c. 若道路類有 CAD 中心線 → 將公設暫編地號 polygon 沿中心線切成 2 半
                    #    d. 將每半依距離分配至最近的「同歸戶相鄰可建築街廓土地」
                    #    e. 用箭頭視覺化「公設暫編地號 → 目的可建築街廓」對映關係
                    from shapely.geometry import Polygon as _SP_M, Point as _SPt_M
                    from shapely.ops import split as _ops_split, unary_union as _ops_union

                    _own_map_for_m = st.session_state.get('t8_ownership_map', {}) or {}
                    _own_groups_for_m = st.session_state.get('t8_ownership_groups', {}) or {}

                    # 收集選定公設街廓內之暫編地號
                    _public_parcels_in_block = [
                        tp for tp in (temp_parcels or [])
                        if tp.get('所屬街廓', '') == _selected_block.get('label', '')
                    ]

                    _split_overlays = []   # [(coords, label, fill_color)]
                    _arrows = []           # [(x_from, y_from, x_to, y_to, label)]
                    _alloc_summary = []    # [(原地號, 歸戶, 公設面積, 配至街廓, 配至面積)]

                    for _pp in _public_parcels_in_block:
                        _pcoords = _pp.get('polygon_coords') or []
                        if len(_pcoords) < 3:
                            continue
                        try:
                            _pp_poly = _SP_M(_pcoords)
                            if not _pp_poly.is_valid:
                                _pp_poly = _pp_poly.buffer(0)
                        except Exception:
                            continue
                        _pp_orig = _pp.get('原地號', '')
                        _pp_gid = _own_map_for_m.get(_pp_orig, '')
                        _pp_area = float(_pp.get('幾何面積_m2', _pp.get('面積_m2', 0)) or 0)

                        # 找該歸戶於可建築街廓內之相鄰土地
                        _gid_lands = []   # [(tp_in_buildable, distance)]
                        if _pp_gid:
                            _grp_parcels = _own_groups_for_m.get(_pp_gid, []) or []
                            for tp2 in (temp_parcels or []):
                                if tp2.get('街廓分類', '') not in ('住宅區', '商業區'):
                                    continue
                                if tp2.get('原地號', '') not in _grp_parcels:
                                    continue
                                _t2coords = tp2.get('polygon_coords') or []
                                if len(_t2coords) < 3:
                                    continue
                                try:
                                    _t2poly = _SP_M(_t2coords)
                                    _d = _t2poly.distance(_pp_poly)
                                    _gid_lands.append((tp2, _d))
                                except Exception:
                                    continue
                            _gid_lands.sort(key=lambda x: x[1])

                        # 道路類 + 有中心線 + 同歸戶於兩側皆有土地 → 切成 2 半
                        _halves = []   # [(half_poly, target_tp)]
                        if (_is_road_like and _ln_pts_for_split is not None
                                and len(_gid_lands) >= 2):
                            try:
                                _splitted = _ops_split(_pp_poly, _ln_pts_for_split.buffer(0.01).boundary)
                                _pieces = list(_splitted.geoms) if hasattr(_splitted, 'geoms') else [_splitted]
                                _pieces = [pc for pc in _pieces
                                            if hasattr(pc, 'area') and pc.area >= 1.0]
                                if len(_pieces) >= 2:
                                    # 對每片找最近的 _gid_lands
                                    for pc in _pieces[:2]:
                                        try:
                                            _pc_cen = pc.centroid
                                            _best_tp = None; _best_d = float('inf')
                                            for tp2, _ in _gid_lands:
                                                _t2poly = _SP_M(tp2['polygon_coords'])
                                                _d = _t2poly.distance(_pc_cen)
                                                if _d < _best_d:
                                                    _best_d = _d
                                                    _best_tp = tp2
                                            if _best_tp is not None:
                                                _halves.append((pc, _best_tp))
                                        except Exception:
                                            continue
                            except Exception:
                                pass

                        if _halves:
                            # 道路切成 2 半，分別分配至兩側
                            for pc, tp_target in _halves:
                                try:
                                    _coords_half = list(pc.exterior.coords)
                                    _split_overlays.append((
                                        _coords_half,
                                        f"{_pp_orig} → {tp_target.get('所屬街廓','')}",
                                        'rgba(255,165,0,0.45)',   # 橘色
                                    ))
                                    # 箭頭：half centroid → target centroid
                                    _h_cen = pc.centroid
                                    _t_cen_x = float(tp_target.get('centroid_x',
                                                                    pc.centroid.x))
                                    _t_cen_y = float(tp_target.get('centroid_y',
                                                                    pc.centroid.y))
                                    _arrows.append((_h_cen.x, _h_cen.y,
                                                    _t_cen_x, _t_cen_y,
                                                    f"{_pp_orig}({pc.area:.1f}㎡)"))
                                    _alloc_summary.append({
                                        '公設原地號': _pp_orig,
                                        '歸戶': _pp_gid,
                                        '切割面積(㎡)': round(float(pc.area), 2),
                                        '配至街廓': tp_target.get('所屬街廓', ''),
                                        '配至原地號': tp_target.get('原地號', ''),
                                    })
                                except Exception:
                                    continue
                        elif _gid_lands:
                            # 不切割 → 整筆配至最近之同歸戶可建築土地
                            tp_target = _gid_lands[0][0]
                            try:
                                _t_cen_x = float(tp_target.get('centroid_x',
                                                                _pp_poly.centroid.x))
                                _t_cen_y = float(tp_target.get('centroid_y',
                                                                _pp_poly.centroid.y))
                                _arrows.append((_pp_poly.centroid.x, _pp_poly.centroid.y,
                                                _t_cen_x, _t_cen_y,
                                                f"{_pp_orig}({_pp_area:.1f}㎡)"))
                                _alloc_summary.append({
                                    '公設原地號': _pp_orig,
                                    '歸戶': _pp_gid,
                                    '切割面積(㎡)': round(_pp_area, 2),
                                    '配至街廓': tp_target.get('所屬街廓', ''),
                                    '配至原地號': tp_target.get('原地號', ''),
                                })
                            except Exception:
                                pass

                    # 繪製切割後 polygon（橘色填充）
                    for _coords_half, _lbl_h, _color_h in _split_overlays:
                        _xs_h = [c[0] for c in _coords_half] + [_coords_half[0][0]]
                        _ys_h = [c[1] for c in _coords_half] + [_coords_half[0][1]]
                        _fig_m.add_trace(go.Scatter(
                            x=_xs_h, y=_ys_h, mode='lines',
                            fill='toself', fillcolor=_color_h,
                            line=dict(color='#E07A00', width=1.5),
                            name=f"🟠 切割：{_lbl_h}",
                            hoverinfo='text', text=_lbl_h,
                            showlegend=True,
                        ))

                    # 繪製箭頭（公設地 → 目的地）
                    for _ax, _ay, _bx, _by, _lbl_a in _arrows:
                        _fig_m.add_annotation(
                            ax=_ax, ay=_ay, x=_bx, y=_by,
                            xref='x', yref='y', axref='x', ayref='y',
                            showarrow=True, arrowhead=3, arrowsize=1.5,
                            arrowwidth=2, arrowcolor='#D62728',
                            text='', opacity=0.75,
                        )

                    # 5) 該歸戶之暫編地號（在可建築街廓內之自有土地）— 綠色填充
                    if _gid_parcel_set:
                        _matched_count = 0
                        for tp in (temp_parcels or []):
                            if tp.get('原地號', '') in _gid_parcel_set:
                                coords = tp.get('polygon_coords') or []
                                if len(coords) < 3:
                                    continue
                                xs_p = [c[0] for c in coords] + [coords[0][0]]
                                ys_p = [c[1] for c in coords] + [coords[0][1]]
                                _fig_m.add_trace(go.Scatter(
                                    x=xs_p, y=ys_p, mode='lines',
                                    fill='toself', fillcolor='rgba(50,180,80,0.45)',
                                    line=dict(color='#2E8B33', width=2),
                                    name=f"🟢 {_gid_pick}：{tp.get('原地號','')}"
                                         + (f"（{tp.get('暫編地號','')}）"
                                            if tp.get('暫編地號') else ''),
                                    hoverinfo='text',
                                    text=f"歸戶 {_gid_pick}\n"
                                         f"原地號 {tp.get('原地號','')}\n"
                                         f"街廓 {tp.get('所屬街廓','')}\n"
                                         f"面積 {float(tp.get('分攤登記面積_m2', tp.get('面積_m2',0))):,.2f} ㎡",
                                    showlegend=(_matched_count < 5),
                                ))
                                _matched_count += 1

                        if _matched_count == 0:
                            st.warning(
                                f"⚠️ 歸戶 **{_gid_pick}** 在重劃前地籍中找不到對應暫編地號"
                                f"（地號清單：{', '.join(sorted(_gid_parcel_set)[:5])}"
                                f"{'…' if len(_gid_parcel_set) > 5 else ''}）"
                            )

                    _fig_m.update_layout(
                        height=560,
                        paper_bgcolor='white', plot_bgcolor='white',
                        xaxis=dict(title="TWD97 X", gridcolor='#E5E5E5'),
                        yaxis=dict(title="TWD97 Y", gridcolor='#E5E5E5',
                                   scaleanchor="x", scaleratio=1),
                        margin=dict(l=40, r=40, t=10, b=40),
                        showlegend=True,
                        legend=dict(orientation="h", yanchor="bottom", y=1.02,
                                    xanchor="left", x=0,
                                    bgcolor='rgba(255,255,255,0.9)',
                                    bordercolor='#CCCCCC', borderwidth=1),
                    )
                    st.plotly_chart(_fig_m, use_container_width=True)

                    # 操作提示
                    st.caption(
                        "🔵 藍色 = 目前選定之公設用地　"
                        "🟠 橘色 = 公設地經中心線切割後配至兩側之示意　"
                        "🟢 綠色 = 該歸戶在相鄰街廓之自有暫編地號　"
                        "🟡 黃色虛線 = CAD CENTERLINE 圖層之道路中心線　"
                        "🔴 紅色箭頭 = 公設地 → 目的可建築街廓對映"
                    )

                    # 分配對映表
                    if _alloc_summary:
                        with st.expander(f"📋 公設用地分配明細（共 {len(_alloc_summary)} 筆對映）",
                                          expanded=False):
                            import pandas as _pd_m
                            st.dataframe(_pd_m.DataFrame(_alloc_summary),
                                         use_container_width=True, hide_index=True)

        # ---------- 步驟 I：推送資料至下游 Tab ----------
        st.markdown("---")
        st.markdown("#### 🔗 步驟 I：推送資料至下游分頁（預期開發分析法 / 地價區段分析 / 財務分析）")
        pcol1, pcol2, pcol3 = st.columns(3)

        # ------ 共用：把 Tab 3 計算出的四項面積同步寫入 Tab 5/6 widget session_state ------
        def _sync_areas_to_tabs():
            _push_map = {
                'area_offset': float(offset_area or 0.0),
                'area_buildable': float(buildable_total or 0.0),
                'area_public_shared': float(public_common_total or 0.0),
                'area_public_non_shared': float(public_non_common_total or 0.0),
            }
            for _k, _v in _push_map.items():
                st.session_state[_k] = _v
                st.session_state[f'_sv_{_k}'] = _v
            st.session_state['f3_to_tab_areas'] = {
                'buildable_total': buildable_total,
                'public_common_total': public_common_total,
                'public_non_common_total': public_non_common_total,
                'offset_area': offset_area,
                'by_category': {c: summary[c]['面積(㎡)'] for c in F3_BLOCK_CATEGORIES},
            }
            st.session_state['_f3_push_consume'] = True  # Tab 5 下次進入時強制覆寫殘留 widget 狀態

        if pcol1.button("📤 推送至 預期開發分析法", use_container_width=True):
            _sync_areas_to_tabs()
            st.success(f"✅ 已推送；可建築 {buildable_total:,.2f} ㎡、共同負擔 {public_common_total:,.2f} ㎡、非共同負擔 {public_non_common_total:,.2f} ㎡、抵充地 {offset_area:,.2f} ㎡")
            st.rerun()

        if pcol2.button("📤 推送至 地價區段分析", use_container_width=True):
            blocks_out = [
                {'街廓編號': b['label'], '街廓分類': b['category'], '面積_m2': round(b['area_m2'], 2)}
                for b in classified_blocks
            ]
            zones_agg = {}
            for tp in (temp_parcels or []):
                z = tp.get('重劃前地價區段', '')
                if not z:
                    continue
                zones_agg.setdefault(z, {'筆數': 0, '面積_m2': 0.0, '暫編地號': []})
                zones_agg[z]['筆數'] += 1
                zones_agg[z]['面積_m2'] += float(tp.get('幾何面積_m2', tp.get('面積_m2', 0)) or 0)
                zones_agg[z]['暫編地號'].append(tp['暫編地號'])
            zones_out = [
                {'區段': k, '筆數': v['筆數'], '面積_m2': round(v['面積_m2'], 2),
                 '暫編地號': '、'.join(v['暫編地號'][:6]) + ('…' if len(v['暫編地號']) > 6 else '')}
                for k, v in zones_agg.items()
            ]
            st.session_state['f3_to_tab_blocks'] = blocks_out
            st.session_state['f3_to_tab_zones'] = zones_out
            _sync_areas_to_tabs()  # 同步面積到 Tab 5，Tab 6 從 Tab 5 讀取面積才會正確
            st.success(f"✅ 已推送 {len(blocks_out)} 個街廓、{len(zones_out)} 個重劃前地價區段（含四項面積同步至 預期開發分析法/地價區段分析）")
            st.rerun()

        if pcol3.button("📤 推送 B 值與參數至 財務分析", use_container_width=True):
            st.session_state['f3_to_finance_inputs'] = {
                'total_area': total_area,
                'preexisting_public': preexisting_public,
                'offset_area': offset_area,
                'public_common_total': public_common_total,
                'general_burden_area': general_burden,
                'price_after': price_after,
                'special_burden_total': sb['special_total'],
            }
            st.session_state['f3_to_tab_B'] = B_value
            _sync_areas_to_tabs()  # 連同四項面積一起同步
            st.success(f"✅ 已推送參數與 B={B_value:.6f} 至 財務分析（含四項面積同步至 預期開發分析法/地價區段分析）")
            st.rerun()

        # ---------- W-G G.1：七級調配（§7 引擎接線；單一真相源＝verify/wf_f0~f4）----------
        st.markdown("---")
        st.markdown("#### 🔷 七級調配（§7 引擎）")
        st.caption("單一真相源＝verify/wf_f0~f4；本區以 live session 注入呼叫**同一引擎模組**（禁 fork、禁重寫調配邏輯）。")
        _wg_ss = st.session_state
        if not _wg_ss.get('f3_G_values'):
            st.info("ℹ️ 請先完成 **Step G（G 值計算）**，再執行七級調配。")
        elif not _is_uc9898(_wg_ss):
            st.warning(
                "⚠️ §7 七級調配引擎為 **UC9898 本案凍結法定真相**（內含 67 項案例錨定斷言，如六格 G(Σa)、"
                "SNAP_WAVG、楔形面積錨）。目前載入資料非 UC9898，引擎**不適用**、暫不執行（避免觸錨崩潰）。"
                "通用化（去錨）列 **未來『泛化波』backlog**——引擎不得為 UI 妥協。")
        else:
            with st.expander("📌 誠實圍欄：財務 ctx 來源＝β 混源（live 幾何 ＋ 快照凍結財務）", expanded=True):
                st.caption("裁定②（KL 2026-07-12）：live 幾何/宗地/Step-G 輸出 ＋ **快照 財務接線_v3**"
                           "（verify/case_params_UC9898.json）。app live 財務值無法逐位複現 KL 授權精度"
                           "（如總面積精確 34949.888 反推），故財務採快照凍結值；下表明示差異、非阻斷、禁靜默混源。")
                try:
                    _fin = _wf_load_snapshot(__file__)["財務接線_v3"]
                    _live = {
                        "重劃前均價(元/㎡)": _wg_ss.get('pre_land_price_sqm'),
                        "重劃後均價(元/㎡)": _wg_ss.get('weighted_price_sqm'),
                        "單位工程費(元/㎡)": _wg_ss.get('_sv_b_cost') or _wg_ss.get('b_cost'),
                        "拆遷費(元/㎡)":     _wg_ss.get('_sv_k_cost') or _wg_ss.get('k_cost'),
                        "行政費(元/㎡)":     _wg_ss.get('_sv_c_cost') or _wg_ss.get('c_cost'),
                    }
                    _snap_v = {
                        "重劃前均價(元/㎡)": _fin["重劃前平均地價_元每m2"],
                        "重劃後均價(元/㎡)": _fin["重劃後平均地價_元每m2"],
                        "單位工程費(元/㎡)": _fin["單位工程費_元每m2"],
                        "拆遷費(元/㎡)":     _fin["拆遷費_元每m2"],
                        "行政費(元/㎡)":     _fin["行政費_元每m2"],
                    }
                    _fence_rows = []
                    for _k, _sv in _snap_v.items():
                        _lv = _live.get(_k)
                        if _lv is None:
                            _d = "— (app 未提供)"
                        elif abs(float(_lv) - float(_sv)) < 1e-6:
                            _d = "✅ 一致"
                        else:
                            _d = f"⚠️ 差異 (live {float(_lv):,.4f})"
                        _fence_rows.append({"項目": _k, "快照凍結值(採用)": f"{float(_sv):,.4f}",
                                            "app live": ("—" if _lv is None else f"{float(_lv):,.4f}"), "狀態": _d})
                    _fence_rows.append({"項目": "重劃區總面積(㎡)",
                                        "快照凍結值(採用)": f"{float(_fin['重劃區總面積_m2_精確']):,.3f}（反推·B/C 分母）",
                                        "app live": "≈34950 (DXF/輸入)", "狀態": "⚠️ 反推精確值 app 無 live 源"})
                    _fence_rows.append({"項目": "貸款利息(元)", "快照凍結值(採用)": f"{int(_fin['貸款利息_元']):,}",
                                        "app live": "NPV×1e4 (精度較粗)", "狀態": "⚠️ 模型/精度不同源"})
                    st.dataframe(pd.DataFrame(_fence_rows), use_container_width=True, hide_index=True)
                    st.caption("🔒 全節凍結採用：重劃區總面積_m2_精確、後街廓/前區段面積單價、原地號_區段 等 財務接線_v3。")
                except Exception as _e_fence:
                    st.warning(f"誠實圍欄表生成失敗（不影響引擎）：{_e_fence}")

            if st.button("🔷 執行七級調配（wf_f0 → f4 全鏈）", key="wg_run_seven", type="primary"):
                try:
                    import sys as _sys
                    import os as _os
                    _vdir = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "verify")
                    if _vdir not in _sys.path:
                        _sys.path.insert(0, _vdir)
                    import wf_f0 as _wf0
                    import wf_f1 as _wf1
                    import wf_f2 as _wf2
                    import wf_f3 as _wf3
                    import wf_f4 as _wf4
                    with st.spinner("執行七級調配（f0→f4 全鏈；含 7-5 最佳化與終態整形）…"):
                        _tag = _wf_tag_of(_wg_ss['f3L_setback_default'])
                        _cbt = {_tag: _build_wf_ctx(_wg_ss, _tag)}
                        _f0r = _wf0.compute(_cbt)
                        _wf1.compute(_cbt, _f0r)
                        _f2r = _wf2.compute(_cbt, _f0r)
                        _f3r = _wf3.compute(_cbt, _f2r)
                        _f4r = _wf4.compute(_cbt, _f0r, _f2r, _f3r)
                    _wg_ss['f3_wg_f4'] = _f4r[_tag]
                    _wg_ss['f3_wg_tag'] = _tag
                    st.success(f"✅ 七級調配完成（f0→f4 全鏈；情境 {_tag}）")
                except Exception as _e_wg:
                    import traceback as _tb
                    st.error(f"🔴 七級調配停機（引擎斷言/合約）：{_e_wg}")
                    st.code(_tb.format_exc()[-1500:])

            _f4o = _wg_ss.get('f3_wg_f4')
            if _f4o:
                st.markdown(f"##### 終態摘要（情境 {_wg_ss.get('f3_wg_tag', '')}）")
                _wtabs = st.tabs(["7-4 三級調配", "7-5 雙出口", "池終態", "終態整形", "33 群總決算"])
                with _wtabs[0]:
                    st.dataframe(pd.DataFrame(_f4o.get("conv_rows", [])), use_container_width=True, hide_index=True)
                with _wtabs[1]:
                    st.dataframe(pd.DataFrame(_f4o.get("exit_rows", [])), use_container_width=True, hide_index=True)
                    st.caption("意思決定項（§7-5 雙出口／增配>0／拆單候選）**照旗標呈現、不自動裁**（合議制作業介面覆寫）。")
                with _wtabs[2]:
                    st.dataframe(pd.DataFrame(_f4o.get("pool_rows", [])), use_container_width=True, hide_index=True)
                    st.caption("池三則：各塊 0 或 ≥MinA。")
                with _wtabs[3]:
                    st.dataframe(pd.DataFrame(_f4o.get("reshape_rows", [])), use_container_width=True, hide_index=True)
                with _wtabs[4]:
                    st.dataframe(pd.DataFrame(_f4o.get("ledger_rows", [])), use_container_width=True, hide_index=True)
                st.caption("🖼️ 逐代圖層切換／池·碎片·雙出口 GIS 視覺化為 **G.2**；本區＝引擎接線＋表格終態。"
                           "雙路同源終驗（app live vs wf/f4 baseline 逐格）為 **G.3**。")

if __name__ == "__main__":
    main()