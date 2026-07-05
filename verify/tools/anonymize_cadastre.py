# -*- coding: utf-8 -*-
"""
anonymize_cadastre.py — 地籍資料等值保持匿名化（U_LAND 格式）
================================================================
用途：將含個資之地籍資料來源.xlsx（工作表 U_LAND＋重劃區地籍）產出匿名版，
      並以 app 歸戶指紋演算法驗證「歸戶分群前後完全同構」。
作者：claude.ai 蒸餾（原始三段 pass 合併）；KL 三輪 review 定案。
安全：本腳本不含任何 PII；對照表輸出檔「限 KL 本機保存、永不進 repo」。

用法：
    python anonymize_cadastre.py 來源.xlsx 匿名版輸出.xlsx 對照表輸出.xlsx [--style fake|mask] [--mask-all]

    --style fake  假名版（預設）：人名→同長度假中文名（單射、結構保證同構）
    --style mask  遮蔽版：施怡韋→施*韋（第二類謄本式；非單射，須靠驗證步實證同構）
    --mask-all    連 8 位法人/機關統編與名稱也全匿（預設保留）

等值保持原則（歸戶語意不變的根據）：
  1. 全域字典、跨 16 個 PII 欄一致：同一真值→同一假值；fake 模式加保證單射。
  2. 統編：自然人(1英9數)→A9########（第2碼9=明顯合成、格式合規）；
     異常碼→Z########；8位法人/機關→保留（--mask-all 時→90######）。
  3. 姓名：法人名（與8位統編同列共現）→保留；團體名（≥5字或含關鍵字）→測試團體NNN；
     餘依 --style 產假名或遮蔽名。假名不與任何真名相同。
  4. 地址一律匿；出生日期非空→4NNNNN；'null' 字面值原樣；空值原樣（na 樣式不變）。
  5. 非 PII 欄、列序、工作表名、欄頭、其他工作表完全原樣。

驗證（內建，跑完自動執行）：
  複刻 app.py 歸戶指紋（owner 去重鍵=統編否則姓名；parts=key:持分排序；
  他項=（設定義務人,權利人,權利種類）去重＋債權持分）於匿名前後各算一次，
  斷言：宗地鍵集一致、分群劃分同構（逐群成員集合相同）、指紋雙射。
  任何斷言失敗 → 不輸出匿名檔、直接報錯。

已知界限：
  - mask 模式同構為「該資料集實證」而非結構保證——統編空值列或他項姓名相撞成併時可能破，
    每次換資料必重跑本驗證（哲學：見 skills/no-silent-fallback、validation-runbook）。
  - PII 欄位表對應 U_LAND 76 欄版式（AG/AH/AI/AK、AN-AR、BH、BL-BP、BS-BW）；
    版式變更時先核欄序再跑。
"""
import argparse, re, sys
from collections import defaultdict
import openpyxl

ID_COLS   = [33, 40, 64, 71]        # AG所有權統編 AN管理者統編 BL權利人統編 BS他項管理者統編
NAME_COLS = [34, 41, 60, 65, 72]    # AH姓名 AO管理者姓名 BH設定義務人 BM權利人姓名 BT他項管理者姓名
ADDR_COLS = [35, 42, 66, 73]        # AI AP BN BU
BIRTH_COLS= [37, 44, 68, 75]        # AK AR BP BW
PAIRS     = [(33,34),(40,41),(64,65),(71,72)]   # (統編,姓名) 同列配對 → 法人名偵測
ORG_KW = ('公司','公業','祭祀','寺','廟','堂','會','行號','政府','公所','農會','國有','縣','市','鄉',
          '銀行','合作社','學校','企業','工業','商行','工廠','社','局','署','部','院','基金','管理')
SURN = list('陳林黃張李王吳劉蔡楊許鄭謝郭洪曾邱廖賴徐周葉蘇莊呂江何蕭羅高潘簡朱鍾游彭詹胡施沈余趙盧梁顏柯翁魏孫戴范方宋鄧杜傅侯曹薛丁卓阮馬董溫唐藍蔣石古紀姚連馮歐程湯田康姜白汪鄒尤巫鐘黎涂龔嚴韓袁金')
GIVEN= list('雅怡君宜芳婷淑惠美玲珍秀慧文欣佩琪思穎靜宛庭家豪志明建宏俊傑冠廷承翰宗哲瑋柏韋辰宇軒睿彥倫振豐鴻成信良德旭昌盛凱勝弘毅安仁義禮智賢')


def norm(v):
    """儲存格值 → 正規字串鍵；空/None → None。整數浮點去 .0（Excel 型別噪音）。"""
    if v is None: return None
    if isinstance(v, float) and v.is_integer(): v = int(v)
    s = str(v)
    return s if s.strip() != '' else None


def scan(src):
    """Pass 1：全表掃描 → PII 相異值集合、法人名集合、逐宗地指紋原料。"""
    ids, names, addrs, births, entity_names = set(), set(), set(), set(), set()
    own = defaultdict(list); mort = defaultdict(list)
    wb = openpyxl.load_workbook(src, read_only=True)
    ws = wb['U_LAND']
    for row in ws.iter_rows(min_row=2, values_only=True):
        for c in ID_COLS:
            v = norm(row[c-1])
            if v: ids.add(v)
        for c in NAME_COLS:
            v = norm(row[c-1])
            if v: names.add(v)
        for c in ADDR_COLS:
            v = norm(row[c-1])
            if v: addrs.add(v)
        for c in BIRTH_COLS:
            v = norm(row[c-1])
            if v and v.strip().lower() != 'null': births.add(v)
        for cid, cnm in PAIRS:
            vi, vn = norm(row[cid-1]), norm(row[cnm-1])
            if vi and vn and re.fullmatch(r'[0-9]{8}', vi.strip()):
                entity_names.add(vn)
        seg, land = norm(row[0]), norm(row[1])
        if seg and land:
            key = (seg, land)
            uni, nm = norm(row[32]), norm(row[33])
            uni_s = uni.strip() if uni else ''
            nm_s  = nm.strip() if nm else ''
            d, nu = row[28], row[29]                     # AC分母 AD分子
            ratio = f"{int(nu)}/{int(d)}" if (d is not None and nu is not None) else ""
            own[key].append((uni_s, nm_s, ratio))
            ob = row[59]                                  # BH
            if ob is not None and str(ob).strip() != '':
                cred, kind = norm(row[64]), norm(row[60])  # BM BI
                dm, nm2 = row[57], row[58]                 # BF BG
                mr = f"{int(nm2)}/{int(dm)}" if (dm is not None and nm2 is not None) else ""
                mort[key].append((str(ob), str(cred) if cred else '', str(kind) if kind else '', mr))
    wb.close()
    return ids, names, addrs, births, entity_names, dict(own), dict(mort)


def build_maps(ids, names, entity_names, addrs, births, style, mask_all):
    id_map = {}; pc = ec = zc = 0
    for v in sorted(ids):
        s = v.strip()
        if re.fullmatch(r'[0-9]{8}', s) and not mask_all:
            id_map[v] = v; continue
        if re.fullmatch(r'[A-Z][0-9]{9}', s):
            pc += 1; id_map[v] = f"A9{pc:08d}"
        elif re.fullmatch(r'[0-9]{8}', s):
            ec += 1; id_map[v] = f"90{ec:06d}"
        else:
            zc += 1; id_map[v] = f"Z{zc:08d}"
    name_map = {}; used = set(); g = 0; gc = 0
    for v in sorted(names):
        if v in entity_names and not mask_all:
            name_map[v] = v; continue
        if len(v) >= 5 or any(k in v for k in ORG_KW):
            gc += 1; name_map[v] = f"測試團體{gc:03d}"; continue
        if style == 'mask':
            if len(v) == 2:   name_map[v] = v[0] + '*'
            elif len(v) == 3: name_map[v] = v[0] + '*' + v[2]
            else:             name_map[v] = v[0] + '*'*(len(v)-2) + v[-1]
            continue
        L = min(max(len(v), 2), 4)
        while True:
            g += 1
            sn = SURN[g % len(SURN)]
            giv = ''.join(GIVEN[(g // (len(GIVEN)**i) + i*17) % len(GIVEN)] for i in range(L-1))
            cand = (sn + giv)[:L]
            if cand not in used and cand not in names:
                used.add(cand); name_map[v] = cand; break
    addr_map  = {v: f"花蓮縣測試市匿名路{i}號" for i, v in enumerate(sorted(addrs), 1)}
    birth_map = {v: f"4{i:05d}"               for i, v in enumerate(sorted(births), 1)}
    return id_map, name_map, addr_map, birth_map


def fingerprint(own, mort, id_map, name_map):
    """複刻 app 歸戶指紋；傳 identity map 即為原始版。"""
    out = {}
    for key, olist in own.items():
        seen = set(); parts = []
        for uni, nm, ratio in olist:
            u2 = id_map.get(uni, uni) if uni else ''
            n2 = name_map.get(nm, nm) if nm else ''
            dk = u2 if u2 not in ('', 'nan') else n2
            if dk in seen: continue
            seen.add(dk)
            kid = u2 if u2 else n2
            parts.append(f"{kid}:{ratio}")
        parts.sort()
        sm = set(); mparts = []
        for ob, cr, kd, mr in mort.get(key, []):
            ob2 = name_map.get(ob, ob); cr2 = name_map.get(cr, cr)
            t = (ob2, cr2, kd)
            if t in sm: continue
            sm.add(t); mparts.append(f"{ob2}→{cr2}[{kd}]{mr}")
        mparts.sort()
        out[key] = "|".join(parts) + "#" + ";".join(mparts)
    return out


def verify_isomorphic(own, mort, id_map, name_map):
    ident_i = {}; ident_n = {}
    of = fingerprint(own, mort, ident_i, ident_n)
    af = fingerprint(own, mort, id_map, name_map)
    assert set(of) == set(af), "宗地鍵集不一致"
    go = defaultdict(set); ga = defaultdict(set)
    for k, f in of.items(): go[f].add(k)
    for k, f in af.items(): ga[f].add(k)
    po = {frozenset(v) for v in go.values()}
    pa = {frozenset(v) for v in ga.values()}
    assert po == pa, f"歸戶分群不同構（原 {len(po)} 群 vs 匿 {len(pa)} 群）——停，勿輸出"
    o2a = {}; a2o = {}
    for k in of:
        o, a = of[k], af[k]
        assert o2a.setdefault(o, a) == a and a2o.setdefault(a, o) == o, "指紋非雙射"
    return len(po)


def rewrite(src, dst, id_map, name_map, addr_map, birth_map):
    def sub(v, mp):
        k = norm(v)
        return mp.get(k, v) if k is not None else v
    wb = openpyxl.load_workbook(src, read_only=True)
    out = openpyxl.Workbook(write_only=True)
    for sn in wb.sheetnames:
        wsi = wb[sn]; wso = out.create_sheet(sn)
        if sn != 'U_LAND':
            for row in wsi.iter_rows(values_only=True): wso.append(list(row))
            continue
        for i, row in enumerate(wsi.iter_rows(values_only=True), 1):
            r = list(row)
            if i > 1:
                for c in ID_COLS:    r[c-1] = sub(r[c-1], id_map)
                for c in NAME_COLS:  r[c-1] = sub(r[c-1], name_map)
                for c in ADDR_COLS:  r[c-1] = sub(r[c-1], addr_map)
                for c in BIRTH_COLS: r[c-1] = sub(r[c-1], birth_map)
            wso.append(r)
    out.save(dst); wb.close()


def write_mapping(path, id_map, name_map, addr_map, birth_map):
    mp = openpyxl.Workbook(write_only=True)
    for title, m in (('統編', id_map), ('姓名', name_map), ('地址', addr_map), ('出生', birth_map)):
        s = mp.create_sheet(title); s.append(['真值', '假值'])
        for k, v in sorted(m.items()):
            if k != v: s.append([k, v])
    mp.save(path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('src'); ap.add_argument('dst'); ap.add_argument('mapping')
    ap.add_argument('--style', choices=['fake', 'mask'], default='fake')
    ap.add_argument('--mask-all', action='store_true')
    a = ap.parse_args()
    print("Pass1 掃描…")
    ids, names, addrs, births, entity_names, own, mort = scan(a.src)
    print(f"  ids {len(ids)} / names {len(names)} / addrs {len(addrs)} / births {len(births)} / 法人名 {len(entity_names)} / 宗地 {len(own)}")
    id_map, name_map, addr_map, birth_map = build_maps(ids, names, entity_names, addrs, births, a.style, a.mask_all)
    print("驗證歸戶同構…")
    n = verify_isomorphic(own, mort, id_map, name_map)
    print(f"  ✅ 同構（{n} 群）")
    print("Pass2 改寫…")
    rewrite(a.src, a.dst, id_map, name_map, addr_map, birth_map)
    write_mapping(a.mapping, id_map, name_map, addr_map, birth_map)
    print(f"完成：{a.dst}\n對照表（限本機保存、勿進 repo）：{a.mapping}")


if __name__ == '__main__':
    main()
